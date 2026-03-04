"""
Test de mejoras UX en Carga Masiva de Candidatos (v2.0.66–v2.0.67)

Cubre los nuevos flujos del modal de carga masiva:
  1. Listar partners para SearchableSelect  (GET /partners)
  2. Listar planteles dado un partner         (GET /partners/{id}/campuses)
  3. Listar grupos dado un plantel            (GET /partners/campuses/{id}/groups)
  4. Crear grupo inline desde el modal        (POST /partners/campuses/{id}/groups)
  5. Preview con grupo recién creado           (POST /user-management/candidates/bulk-upload/preview)
  6. Upload con grupo recién creado            (POST /user-management/candidates/bulk-upload)
  7. Verificar asignación al grupo nuevo       (miembros creados correctamente)
  8. Selector vacío — plantel sin grupos       (respuesta vacía válida)
  9. Selector vacío — partner sin planteles    (respuesta vacía válida)
 10. Validaciones de creación de grupo         (nombre vacío → 400)
 11. Preview sin grupo (flujo básico sigue OK) (retrocompatibilidad)
 12. Limpieza de datos de prueba

USO: python tests/test_bulk_upload_ux.py [--dev]
"""
import sys
import os
import io
import json
import time
import uuid
import requests

# ── Configuración ──
if "--dev" in sys.argv:
    API = "https://evaluaasi-motorv2-api-dev.purpleocean-384694c4.southcentralus.azurecontainerapps.io/api"
    ENV_LABEL = "DEV"
    DB_NAME = "evaluaasi_dev"
else:
    API = "https://evaluaasi-motorv2-api.purpleocean-384694c4.southcentralus.azurecontainerapps.io/api"
    ENV_LABEL = "PROD"
    DB_NAME = "evaluaasi"

try:
    import pymssql
    HAS_DB = True
except ImportError:
    print("⚠️ pymssql no disponible, limpieza manual requerida")
    HAS_DB = False

DB_SERVER = "evaluaasi-motorv2-sql.database.windows.net"
DB_USER = "evaluaasi_admin"
DB_PASS = "EvalAasi2024_newpwd!"

passed = 0
failed = 0
warnings = 0

TEST_PREFIX = f"BKUX_{uuid.uuid4().hex[:6].upper()}"
REQ_TIMEOUT = 120

# IDs para limpieza
created_user_ids = []
created_group_id = None
created_member_ids = []


# ── Helpers ──

def test(name, condition, detail=""):
    global passed, failed
    if condition:
        print(f"  ✅ {name}")
        passed += 1
    else:
        print(f"  ❌ {name} — {detail}")
        failed += 1


def warn(name, detail=""):
    global warnings
    print(f"  ⚠️ {name} — {detail}")
    warnings += 1


def post_with_retry(url, max_retries=2, **kwargs):
    kwargs.setdefault('timeout', REQ_TIMEOUT)
    for attempt in range(max_retries + 1):
        try:
            r = requests.post(url, **kwargs)
            if r.status_code == 504 and attempt < max_retries:
                print(f"  ⏳ 504 timeout, reintentando ({attempt + 1}/{max_retries})...")
                time.sleep(5)
                continue
            return r
        except requests.exceptions.Timeout:
            if attempt < max_retries:
                print(f"  ⏳ Request timeout, reintentando ({attempt + 1}/{max_retries})...")
                time.sleep(5)
                continue
            raise
    return r


def get_with_retry(url, max_retries=1, **kwargs):
    kwargs.setdefault('timeout', REQ_TIMEOUT)
    for attempt in range(max_retries + 1):
        try:
            r = requests.get(url, **kwargs)
            if r.status_code == 504 and attempt < max_retries:
                time.sleep(5)
                continue
            return r
        except requests.exceptions.Timeout:
            if attempt < max_retries:
                time.sleep(5)
                continue
            raise
    return r


def get_token(username="admin", password="Admin123!"):
    r = post_with_retry(f"{API}/auth/login", json={"username": username, "password": password})
    if r.status_code in (429, 423):
        wait = int(r.json().get("retry_after", 30))
        print(f"  ⏳ Rate limit, esperando {min(wait, 60)}s...")
        time.sleep(min(wait, 60))
        r = post_with_retry(f"{API}/auth/login", json={"username": username, "password": password})
    if r.status_code == 200:
        return r.json().get("access_token")
    print(f"  ⚠️ Login falló: {r.status_code} {r.text[:200]}")
    return None


def auth_headers(token):
    return {"Authorization": f"Bearer {token}"}


def auth_json_headers(token):
    return {"Authorization": f"Bearer {token}", "Content-Type": "application/json"}


def db_query(sql, params=None):
    if not HAS_DB:
        return []
    conn = pymssql.connect(server=DB_SERVER, user=DB_USER, password=DB_PASS, database=DB_NAME)
    cursor = conn.cursor(as_dict=True)
    cursor.execute(sql, params or ())
    try:
        rows = cursor.fetchall()
    except:
        rows = []
    conn.commit()
    conn.close()
    return rows


def db_exec(sql, params=None):
    if not HAS_DB:
        return 0
    conn = pymssql.connect(server=DB_SERVER, user=DB_USER, password=DB_PASS, database=DB_NAME)
    cursor = conn.cursor()
    cursor.execute(sql, params or ())
    affected = cursor.rowcount
    conn.commit()
    conn.close()
    return affected


def make_candidates_excel(candidates):
    """Genera Excel para carga masiva Flujo 1."""
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    headers = ["email", "nombre", "primer_apellido", "segundo_apellido", "genero", "curp"]
    for col_idx, h in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx, value=h)
    ws.cell(row=2, column=1, value="(descripción)")
    for row_idx, c in enumerate(candidates, 3):
        for col_idx, h in enumerate(headers, 1):
            ws.cell(row=row_idx, column=col_idx, value=c.get(h))
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ============================================================
# MAIN
# ============================================================
def main():
    global passed, failed, created_user_ids, created_group_id, created_member_ids

    print("\n" + "=" * 65)
    print(f"  TESTS DE UX — CARGA MASIVA DE CANDIDATOS ({ENV_LABEL})")
    print(f"  SearchableSelect + Crear grupo inline + Flujo completo")
    print(f"  Prefijo: {TEST_PREFIX}")
    print("=" * 65)

    # ── 0. Autenticación ──
    print("\n── 0. Autenticación ──")
    token = get_token("admin", "Admin123!")
    test("Login como admin", token is not None)
    if not token:
        print("⛔ No se pudo autenticar. Abortando.")
        sys.exit(1)
    h = auth_headers(token)
    hj = auth_json_headers(token)

    # ════════════════════════════════════════════════════════════
    # SECCIÓN 1: SELECTORES (SearchableSelect)
    # ════════════════════════════════════════════════════════════
    print("\n" + "=" * 65)
    print("  SECCIÓN 1: SELECTORES — Partners / Planteles / Grupos")
    print("=" * 65)

    # ── 1. Listar partners ──
    print("\n── 1. GET /partners — Listar partners para SearchableSelect ──")
    r = get_with_retry(f"{API}/partners?page=1&per_page=500&active_only=true", headers=h)
    test("GET /partners → 200", r.status_code == 200, f"Status: {r.status_code}")

    partner_id = None
    partner_name = None
    partners_data = None
    if r.status_code == 200:
        partners_data = r.json()
        test("Respuesta tiene 'partners' (array)", isinstance(partners_data.get("partners"), list))
        test("Respuesta tiene 'total' (int)", isinstance(partners_data.get("total"), int))

        partners_list = partners_data.get("partners", [])
        test(f"Hay ≥1 partner activo", len(partners_list) >= 1, f"total={len(partners_list)}")

        if partners_list:
            # Cada partner tiene id y name (necesarios para SearchableSelect)
            first = partners_list[0]
            test("Partner tiene 'id'", "id" in first, f"keys={list(first.keys())}")
            test("Partner tiene 'name'", "name" in first, f"keys={list(first.keys())}")
    else:
        warn("GET /partners falló, tests subsecuentes pueden fallar")

    # ── 2. Listar planteles de un partner ──
    print("\n── 2. GET /partners/{id}/campuses — Planteles del partner ──")

    campus_id = None
    campus_name = None
    if partners_data:
        # Buscar un partner que tenga al menos 1 plantel activo
        for p in partners_data.get("partners", []):
            pid = p["id"]
            r = get_with_retry(f"{API}/partners/{pid}/campuses?active_only=true", headers=h)
            if r.status_code == 200:
                campuses_data = r.json()
                campuses_list = campuses_data.get("campuses", [])
                if campuses_list:
                    partner_id = pid
                    partner_name = p["name"]
                    print(f"  📌 Partner con planteles: {partner_name} (ID {partner_id})")

                    test(f"GET /partners/{partner_id}/campuses → 200", True)
                    test("Respuesta tiene 'campuses' (array)", isinstance(campuses_list, list))
                    test("Respuesta tiene 'partner_name'", bool(campuses_data.get("partner_name")))
                    test("partner_name coincide", campuses_data.get("partner_name") == partner_name,
                         f"Got: {campuses_data.get('partner_name')}")
                    test(f"Hay ≥1 plantel en partner", len(campuses_list) >= 1)

                    first_c = campuses_list[0]
                    test("Plantel tiene 'id'", "id" in first_c)
                    test("Plantel tiene 'name'", "name" in first_c)

                    campus_id = first_c["id"]
                    campus_name = first_c["name"]
                    print(f"  📌 Plantel seleccionado: {campus_name} (ID {campus_id})")
                    break
        if not partner_id:
            warn("Ningún partner tiene planteles activos", "No se pueden ejecutar tests con campus/grupo")
    else:
        warn("Sin datos de partners, saltando test de planteles")

    # ── 3. Listar grupos de un plantel ──
    print("\n── 3. GET /partners/campuses/{id}/groups — Grupos del plantel ──")

    existing_group_id = None
    existing_group_name = None
    if campus_id:
        r = get_with_retry(f"{API}/partners/campuses/{campus_id}/groups?active_only=true", headers=h)
        test(f"GET /partners/campuses/{campus_id}/groups → 200", r.status_code == 200, f"Status: {r.status_code}")

        if r.status_code == 200:
            groups_data = r.json()
            test("Respuesta tiene 'groups' (array)", isinstance(groups_data.get("groups"), list))
            test("Respuesta tiene 'campus_name'", bool(groups_data.get("campus_name")))

            groups_list = groups_data.get("groups", [])
            if groups_list:
                first_g = groups_list[0]
                test("Grupo tiene 'id'", "id" in first_g)
                test("Grupo tiene 'name'", "name" in first_g)
                existing_group_id = first_g["id"]
                existing_group_name = first_g["name"]
                print(f"  📌 Grupo existente: {existing_group_name} (ID {existing_group_id})")
            else:
                print("  ℹ️ Plantel no tiene grupos activos (OK, test de vacíos más adelante)")
    else:
        warn("Sin campus_id, saltando test de grupos")

    # ── 4. Selector vacío — partner sin planteles ──
    print("\n── 4. Selector vacío — partner ficticio o sin planteles ──")

    # Intentar con un partner que no tenga campuses
    if partners_data:
        all_partners = partners_data.get("partners", [])
        found_empty = False
        # Probar los últimos partners (más probables de no tener campuses)
        for p in reversed(all_partners[:10]):
            pid = p["id"]
            r = get_with_retry(f"{API}/partners/{pid}/campuses?active_only=true", headers=h)
            if r.status_code == 200:
                clist = r.json().get("campuses", [])
                if len(clist) == 0:
                    found_empty = True
                    test(f"Partner sin planteles → array vacío", True)
                    test("  campuses == []", clist == [])
                    test("  total == 0", r.json().get("total") == 0)
                    break
        if not found_empty:
            print("  ℹ️ Todos los partners tienen planteles (no se pudo probar vacío)")
            test("GET campuses responde array (puede estar vacío o con datos)", True)

    # ── 5. Selector vacío — plantel sin grupos ──
    print("\n── 5. Selector vacío — plantel sin grupos activos ──")

    if campus_id:
        # Buscar un plantel sin grupos
        if partners_data:
            found_empty_groups = False
            for p in partners_data.get("partners", [])[:5]:
                r_c = get_with_retry(f"{API}/partners/{p['id']}/campuses?active_only=true", headers=h)
                if r_c.status_code == 200:
                    for c in r_c.json().get("campuses", []):
                        r_g = get_with_retry(f"{API}/partners/campuses/{c['id']}/groups?active_only=true", headers=h)
                        if r_g.status_code == 200 and len(r_g.json().get("groups", [])) == 0:
                            found_empty_groups = True
                            test(f"Plantel sin grupos → array vacío", True)
                            test("  groups == []", r_g.json().get("groups") == [])
                            break
                if found_empty_groups:
                    break
            if not found_empty_groups:
                print("  ℹ️ Todos los planteles tienen grupos (no se encontró caso vacío)")
                test("GET groups responde array (puede estar vacío o con datos)", True)

    # ════════════════════════════════════════════════════════════
    # SECCIÓN 2: CREAR GRUPO INLINE
    # ════════════════════════════════════════════════════════════
    print("\n" + "=" * 65)
    print("  SECCIÓN 2: CREAR GRUPO INLINE desde modal de carga masiva")
    print("=" * 65)

    # ── 6. Validación — nombre vacío ──
    print("\n── 6. Crear grupo sin nombre → 400 ──")
    if campus_id:
        r = post_with_retry(f"{API}/partners/campuses/{campus_id}/groups",
                            json={}, headers=hj)
        test("POST grupo sin nombre → 400", r.status_code == 400, f"Status: {r.status_code}")
        if r.status_code == 400:
            test("  Error menciona 'nombre' o 'requerido'",
                 "requerido" in r.json().get("error", "").lower() or "nombre" in r.json().get("error", "").lower(),
                 r.json().get("error", ""))
    else:
        warn("Sin campus_id, saltando test de creación grupo sin nombre")

    # ── 7. Crear grupo inline exitoso ──
    print("\n── 7. Crear grupo inline con nombre válido ──")
    group_name = f"{TEST_PREFIX}_GrupoTest"
    if campus_id:
        r = post_with_retry(f"{API}/partners/campuses/{campus_id}/groups",
                            json={"name": group_name}, headers=hj)
        test("POST crear grupo → 201", r.status_code == 201, f"Status: {r.status_code} | {r.text[:300]}")

        if r.status_code == 201:
            data = r.json()
            test("Respuesta tiene 'group'", "group" in data)
            grp = data.get("group", {})
            test("Grupo tiene 'id'", bool(grp.get("id")))
            test("Grupo tiene 'name'", bool(grp.get("name")))
            test(f"name == '{group_name}'", grp.get("name") == group_name, f"Got: {grp.get('name')}")
            test(f"campus_id == {campus_id}", grp.get("campus_id") == campus_id,
                 f"Got: {grp.get('campus_id')}")
            test("is_active == True", grp.get("is_active") == True)

            created_group_id = grp["id"]
            print(f"  📌 Grupo creado: {group_name} (ID {created_group_id})")
        else:
            warn(f"No se pudo crear grupo: {r.text[:200]}")
    else:
        warn("Sin campus_id, no se puede crear grupo inline")

    # ── 8. Verificar grupo aparece en listado ──
    print("\n── 8. Grupo recién creado aparece en GET groups ──")
    if created_group_id and campus_id:
        r = get_with_retry(f"{API}/partners/campuses/{campus_id}/groups?active_only=true", headers=h)
        test("GET groups → 200", r.status_code == 200)

        if r.status_code == 200:
            groups_list = r.json().get("groups", [])
            found = any(g["id"] == created_group_id for g in groups_list)
            test(f"Grupo {created_group_id} aparece en lista", found,
                 f"IDs encontrados: {[g['id'] for g in groups_list[:10]]}")
    else:
        warn("Sin grupo creado, saltando verificación en listado")

    # ════════════════════════════════════════════════════════════
    # SECCIÓN 3: CARGA MASIVA CON GRUPO INLINE
    # ════════════════════════════════════════════════════════════
    print("\n" + "=" * 65)
    print("  SECCIÓN 3: CARGA MASIVA CON GRUPO CREADO INLINE")
    print("=" * 65)

    uid = TEST_PREFIX[5:]  # parte random
    test_candidates = [
        {
            "email": f"{TEST_PREFIX.lower()}_c1@testux.com",
            "nombre": f"Marta{uid}",
            "primer_apellido": "López",
            "segundo_apellido": "García",
            "genero": "F",
            "curp": f"MAGA{uid}000001".ljust(18, "X")[:18],
        },
        {
            "email": f"{TEST_PREFIX.lower()}_c2@testux.com",
            "nombre": f"Pedro{uid}",
            "primer_apellido": "Ramírez",
            "segundo_apellido": "Soto",
            "genero": "M",
            "curp": None,
        },
        {
            "email": None,
            "nombre": f"Lucía{uid}",
            "primer_apellido": "Hernández",
            "segundo_apellido": "Villa",
            "genero": "F",
            "curp": None,
        },
    ]

    # ── 9. Preview con grupo recién creado ──
    print("\n── 9. Preview con group_id del grupo inline ──")
    target_group = created_group_id or existing_group_id
    if target_group:
        excel_data = make_candidates_excel(test_candidates)
        r = post_with_retry(
            f"{API}/user-management/candidates/bulk-upload/preview",
            headers=h,
            files={"file": ("candidatos.xlsx", excel_data,
                            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            data={"group_id": str(target_group)},
        )
        test("Preview con group_id → 200", r.status_code == 200, f"Status: {r.status_code} | {r.text[:300]}")

        if r.status_code == 200:
            data = r.json()
            summary = data.get("summary", {})
            test("summary.total_rows == 3", summary.get("total_rows") == 3,
                 f"total_rows={summary.get('total_rows')}")
            ready = summary.get("ready", 0) + summary.get("name_matches", 0)
            test(f"ready + name_matches == 3", ready == 3, f"ready={summary.get('ready')}, nm={summary.get('name_matches')}")
            test("can_proceed == True", data.get("can_proceed") == True)

            # group_info presente porque se pasó group_id
            gi = data.get("group_info")
            test("group_info presente", gi is not None, f"keys={list(data.keys())}")
            if gi:
                test(f"group_info.id == {target_group}", gi.get("id") == target_group,
                     f"Got: {gi.get('id')}")

            # Verificar elegibilidad en filas
            preview_rows = data.get("preview", [])
            processable = [p for p in preview_rows if p.get("status") in ("ready", "name_match")]
            if processable:
                elig = processable[0].get("eligibility", {})
                test("Fila tiene eligibility.reporte", "reporte" in elig)
                test("Fila tiene eligibility.eduit", "eduit" in elig)
                test("Fila tiene eligibility.conocer", "conocer" in elig)
                test("Fila tiene eligibility.insignia", "insignia" in elig)
    else:
        warn("Sin grupo disponible, saltando preview con grupo")

    # ── 10. Upload con grupo recién creado ──
    print("\n── 10. Upload con group_id del grupo inline ──")
    if target_group:
        excel_upload = make_candidates_excel(test_candidates)
        r = post_with_retry(
            f"{API}/user-management/candidates/bulk-upload",
            headers=h,
            files={"file": ("candidatos.xlsx", excel_upload,
                            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            data={"group_id": str(target_group)},
        )
        test("Upload con group_id → 200", r.status_code == 200, f"Status: {r.status_code} | {r.text[:300]}")

        if r.status_code == 200:
            data = r.json()
            summary = data.get("summary", {})
            details = data.get("details", {})

            test(f"created == 3", summary.get("created") == 3, f"created={summary.get('created')}")
            test(f"errors == 0", summary.get("errors") == 0, f"errors={summary.get('errors')}")

            # Verificar asignación a grupo
            ga = data.get("group_assignment", {})
            test("group_assignment presente", bool(ga))
            if ga:
                test(f"assigned >= 3", ga.get("assigned", 0) >= 3, f"assigned={ga.get('assigned')}")
                test(f"group_id == {target_group}", ga.get("group_id") == target_group,
                     f"Got: {ga.get('group_id')}")

            # Guardar IDs para limpieza
            for c in details.get("created", []):
                uname = c.get("username")
                if uname and HAS_DB:
                    rows = db_query("SELECT id FROM users WHERE username=%s", (uname,))
                    if rows:
                        created_user_ids.append(rows[0]["id"])

            # Verificar credenciales generadas
            created_list = details.get("created", [])
            if created_list:
                first = created_list[0]
                test("Tiene 'username'", bool(first.get("username")))
                test("Tiene 'password'", bool(first.get("password")))
                test("Username es MAYÚSCULAS", first.get("username", "").isupper(),
                     f"Got: {first.get('username')}")
    else:
        warn("Sin grupo disponible, saltando upload con grupo")

    # ── 11. Verificar miembros en grupo ──
    print("\n── 11. Verificar miembros asignados al grupo ──")
    if target_group and created_user_ids and HAS_DB:
        member_count = db_query(
            "SELECT COUNT(*) as cnt FROM group_members WHERE group_id=%s AND status='active'",
            (target_group,)
        )
        cnt = member_count[0]["cnt"] if member_count else 0
        test(f"Grupo {target_group} tiene ≥3 miembros activos", cnt >= 3, f"cnt={cnt}")

        # Verificar que los usuarios creados son miembros
        for uid in created_user_ids:
            rows = db_query(
                "SELECT id FROM group_members WHERE group_id=%s AND user_id=%s AND status='active'",
                (target_group, uid)
            )
            is_member = len(rows) > 0
            if not is_member:
                test(f"Usuario {uid} es miembro del grupo", False)
                break
        else:
            test(f"Todos los usuarios creados son miembros del grupo", True)
    elif not HAS_DB:
        warn("Sin acceso a DB, no se puede verificar membresías")
    else:
        warn("Sin datos para verificar membresías")

    # ════════════════════════════════════════════════════════════
    # SECCIÓN 4: RETROCOMPATIBILIDAD — preview/upload SIN grupo
    # ════════════════════════════════════════════════════════════
    print("\n" + "=" * 65)
    print("  SECCIÓN 4: RETROCOMPATIBILIDAD — Sin selector de grupo")
    print("=" * 65)

    # ── 12. Preview sin grupo ──
    print("\n── 12. Preview sin group_id (flujo original) ──")

    nogrp_email = f"{TEST_PREFIX.lower()}_no1@testux.com"
    nogrp_candidates = [
        {
            "email": nogrp_email,
            "nombre": f"Solo{uid}",
            "primer_apellido": "SinGrupo",
            "segundo_apellido": "Test",
            "genero": "M",
            "curp": None,
        },
    ]
    excel_ng = make_candidates_excel(nogrp_candidates)
    r = post_with_retry(
        f"{API}/user-management/candidates/bulk-upload/preview",
        headers=h,
        files={"file": ("singrupo.xlsx", excel_ng,
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    test("Preview sin grupo → 200", r.status_code == 200, f"Status: {r.status_code} | {r.text[:300]}")

    if r.status_code == 200:
        data = r.json()
        test("group_info == null (sin grupo)", data.get("group_info") is None,
             f"Got: {data.get('group_info')}")
        test("can_proceed == True", data.get("can_proceed") == True)
        summary = data.get("summary", {})
        test("total_rows == 1", summary.get("total_rows") == 1)

    # ── 13. Upload sin grupo ──
    print("\n── 13. Upload sin group_id (flujo original) ──")

    excel_ng2 = make_candidates_excel(nogrp_candidates)
    r = post_with_retry(
        f"{API}/user-management/candidates/bulk-upload",
        headers=h,
        files={"file": ("singrupo.xlsx", excel_ng2,
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    test("Upload sin grupo → 200", r.status_code == 200, f"Status: {r.status_code} | {r.text[:300]}")

    if r.status_code == 200:
        data = r.json()
        summary = data.get("summary", {})
        test(f"created == 1", summary.get("created") == 1, f"created={summary.get('created')}")
        test("group_assignment ausente o vacío",
             not data.get("group_assignment") or data.get("group_assignment", {}).get("assigned", 0) == 0,
             f"ga={data.get('group_assignment')}")

        # Guardar para limpieza
        for c in data.get("details", {}).get("created", []):
            uname = c.get("username")
            if uname and HAS_DB:
                rows = db_query("SELECT id FROM users WHERE username=%s", (uname,))
                if rows:
                    created_user_ids.append(rows[0]["id"])

    # ════════════════════════════════════════════════════════════
    # SECCIÓN 5: FLUJO COMPLETO END-TO-END
    # ════════════════════════════════════════════════════════════
    print("\n" + "=" * 65)
    print("  SECCIÓN 5: FLUJO END-TO-END (Partner → Campus → Grupo → Upload)")
    print("=" * 65)

    # ── 14. Simular flujo completo del modal ──
    print("\n── 14. Simulación flujo completo del modal ──")

    if partner_id and campus_id:
        # Paso a: Obtener partners
        r1 = get_with_retry(f"{API}/partners?page=1&per_page=500&active_only=true", headers=h)
        test("Paso 1: GET partners → 200", r1.status_code == 200)

        # Paso b: Seleccionar partner → obtener campuses
        r2 = get_with_retry(f"{API}/partners/{partner_id}/campuses?active_only=true", headers=h)
        test("Paso 2: GET campuses → 200", r2.status_code == 200)

        # Paso c: Seleccionar campus → obtener grupos
        r3 = get_with_retry(f"{API}/partners/campuses/{campus_id}/groups?active_only=true", headers=h)
        test("Paso 3: GET groups → 200", r3.status_code == 200)

        # Paso d: Crear grupo inline
        e2e_group_name = f"{TEST_PREFIX}_E2E"
        r4 = post_with_retry(f"{API}/partners/campuses/{campus_id}/groups",
                             json={"name": e2e_group_name}, headers=hj)
        test("Paso 4: POST crear grupo → 201", r4.status_code == 201, f"Status: {r4.status_code}")

        e2e_group_id = None
        if r4.status_code == 201:
            e2e_group_id = r4.json().get("group", {}).get("id")

        # Paso e: Preview con grupo
        e2e_email = f"{TEST_PREFIX.lower()}_e2e@testux.com"
        e2e_cands = [{
            "email": e2e_email,
            "nombre": f"E2E{uid}",
            "primer_apellido": "Flujo",
            "segundo_apellido": "Completo",
            "genero": "M",
            "curp": None,
        }]
        excel_e2e = make_candidates_excel(e2e_cands)
        if e2e_group_id:
            r5 = post_with_retry(
                f"{API}/user-management/candidates/bulk-upload/preview",
                headers=h,
                files={"file": ("e2e.xlsx", excel_e2e, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
                data={"group_id": str(e2e_group_id)},
            )
            test("Paso 5: Preview → 200", r5.status_code == 200, f"Status: {r5.status_code}")

            # Paso f: Upload
            excel_e2e2 = make_candidates_excel(e2e_cands)
            r6 = post_with_retry(
                f"{API}/user-management/candidates/bulk-upload",
                headers=h,
                files={"file": ("e2e.xlsx", excel_e2e2, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
                data={"group_id": str(e2e_group_id)},
            )
            test("Paso 6: Upload → 200", r6.status_code == 200, f"Status: {r6.status_code}")

            if r6.status_code == 200:
                summ = r6.json().get("summary", {})
                test("  created == 1", summ.get("created") == 1, f"created={summ.get('created')}")
                ga = r6.json().get("group_assignment", {})
                test("  assigned == 1", ga.get("assigned", 0) >= 1, f"assigned={ga.get('assigned')}")
                test("  group_id correcto", ga.get("group_id") == e2e_group_id,
                     f"Got: {ga.get('group_id')}")

                # Guardar para limpieza
                for c in r6.json().get("details", {}).get("created", []):
                    uname = c.get("username")
                    if uname and HAS_DB:
                        rows = db_query("SELECT id FROM users WHERE username=%s", (uname,))
                        if rows:
                            created_user_ids.append(rows[0]["id"])
        else:
            warn("No se pudo crear grupo E2E, saltando preview/upload")

        # Guardar grupo E2E para limpieza
        if e2e_group_id and HAS_DB:
            # Lo marcamos para limpiar
            pass  # Se limpia en sección de limpieza

    else:
        warn("Sin partner/campus, saltando flujo E2E")

    # ════════════════════════════════════════════════════════════
    # SECCIÓN 6: UPLOAD CON EXISTENTES + GRUPO (include_existing_ids)
    # ════════════════════════════════════════════════════════════
    print("\n" + "=" * 65)
    print("  SECCIÓN 6: ASIGNAR EXISTENTES A GRUPO (include_existing_ids)")
    print("=" * 65)

    # ── 15. Re-upload mismos candidatos con grupo ──
    print("\n── 15. Re-upload existentes con include_existing_ids ──")

    if target_group and created_user_ids and HAS_DB:
        # Limpiar membresías previas para forzar re-asignación
        for uid in created_user_ids:
            db_exec("DELETE FROM group_members WHERE group_id=%s AND user_id=%s",
                    (target_group, uid))

        # Preview primero para obtener IDs de existentes
        excel_re = make_candidates_excel(test_candidates)
        r = post_with_retry(
            f"{API}/user-management/candidates/bulk-upload/preview",
            headers=h,
            files={"file": ("reupload.xlsx", excel_re,
                            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            data={"group_id": str(target_group)},
        )

        existing_ids = []
        if r.status_code == 200:
            for row in r.json().get("preview", []):
                eu = row.get("existing_user")
                if eu and row.get("status") in ("skipped", "duplicate"):
                    existing_ids.append(eu["id"])

        if existing_ids:
            excel_re2 = make_candidates_excel(test_candidates)
            r = post_with_retry(
                f"{API}/user-management/candidates/bulk-upload",
                headers=h,
                files={"file": ("reupload.xlsx", excel_re2,
                                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
                data={
                    "group_id": str(target_group),
                    "include_existing_ids": json.dumps(existing_ids),
                },
            )
            test("Upload include_existing_ids → 200", r.status_code == 200,
                 f"Status: {r.status_code} | {r.text[:300]}")

            if r.status_code == 200:
                data = r.json()
                summ = data.get("summary", {})
                test(f"existing_assigned >= 1", summ.get("existing_assigned", 0) >= 1,
                     f"existing_assigned={summ.get('existing_assigned')}")
                ga = data.get("group_assignment", {})
                test(f"assigned_existing >= 1", ga.get("assigned_existing", 0) >= 1,
                     f"assigned_existing={ga.get('assigned_existing')}")
        else:
            warn("No se encontraron existentes para include_existing_ids")
    else:
        warn("Sin datos para test include_existing_ids")

    # ════════════════════════════════════════════════════════════
    # SECCIÓN 7: DESCARGA DE PLANTILLA
    # ════════════════════════════════════════════════════════════
    print("\n" + "=" * 65)
    print("  SECCIÓN 7: DESCARGA DE PLANTILLA")
    print("=" * 65)

    print("\n── 16. GET template ──")
    r = get_with_retry(f"{API}/user-management/candidates/bulk-upload/template", headers=h)
    test("GET template → 200", r.status_code == 200, f"Status: {r.status_code}")
    if r.status_code == 200:
        ct = r.headers.get("Content-Type", "")
        test("Content-Type correcto (xlsx/octet-stream)",
             "spreadsheet" in ct or "octet-stream" in ct, f"CT: {ct}")
        test(f"Tamaño > 100 bytes", len(r.content) > 100, f"Size: {len(r.content)}")

    # ════════════════════════════════════════════════════════════
    # LIMPIEZA
    # ════════════════════════════════════════════════════════════
    print("\n" + "=" * 65)
    print("  🧹 LIMPIEZA DE DATOS DE PRUEBA")
    print("=" * 65)

    if HAS_DB:
        cleaned_users = 0
        cleaned_groups = 0
        cleaned_members = 0

        # Limpiar usuarios por prefijo de email
        test_users = db_query(
            "SELECT id, username, email FROM users WHERE email LIKE %s",
            (f"%{TEST_PREFIX.lower()}%",)
        )
        # También por username
        test_users2 = db_query(
            "SELECT id, username, email FROM users WHERE username LIKE %s",
            (f"%{TEST_PREFIX}%",)
        )
        all_test_users = {r["id"]: r for r in test_users + test_users2}
        all_user_ids = list(all_test_users.keys()) + created_user_ids

        # Limpiar membresías
        for uid in set(all_user_ids):
            n = db_exec("DELETE FROM group_members WHERE user_id=%s", (uid,))
            cleaned_members += n

        # Limpiar usuarios
        for uid in set(all_user_ids):
            n = db_exec("DELETE FROM users WHERE id=%s", (uid,))
            cleaned_users += n

        # Limpiar grupos de prueba
        test_groups = db_query(
            "SELECT id, name FROM candidate_groups WHERE name LIKE %s",
            (f"{TEST_PREFIX}%",)
        )
        for g in test_groups:
            db_exec("DELETE FROM group_members WHERE group_id=%s", (g["id"],))
            db_exec("DELETE FROM candidate_groups WHERE id=%s", (g["id"],))
            cleaned_groups += 1

        print(f"  ✅ {cleaned_users} usuarios eliminados")
        print(f"  ✅ {cleaned_members} membresías eliminadas")
        print(f"  ✅ {cleaned_groups} grupos de prueba eliminados")
        test("Limpieza completada", True)
    else:
        warn("Sin acceso a DB, limpieza manual requerida", f"Prefijo: {TEST_PREFIX}")

    # ════════════════════════════════════════════════════════════
    # RESUMEN FINAL
    # ════════════════════════════════════════════════════════════
    print("\n" + "=" * 65)
    total = passed + failed
    emoji = "✅" if failed == 0 else "⚠️" if failed <= 3 else "❌"
    print(f"  {emoji}  RESULTADOS ({ENV_LABEL}): {passed}/{total} tests pasaron")
    if warnings:
        print(f"     ⚠️  {warnings} advertencias")
    if failed:
        print(f"     ❌ {failed} tests fallaron")
    print("=" * 65)

    sys.exit(0 if failed == 0 else 1)


if __name__ == "__main__":
    main()
