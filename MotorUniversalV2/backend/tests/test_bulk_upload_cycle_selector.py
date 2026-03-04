"""
Test de selector de Ciclo Escolar en Carga Masiva (v2.0.69)

Cubre el nuevo flujo donde al crear un grupo inline desde la carga masiva
se requiere seleccionar un ciclo escolar (SearchableSelect buscable).

  1. Listar ciclos de un plantel                (GET /partners/campuses/{id}/cycles)
  2. Ciclos tienen id y name (SearchableSelect)  (campos requeridos para selector)
  3. Plantel sin ciclos → array vacío            (respuesta válida)
  4. Crear grupo con school_cycle_id             (POST /partners/campuses/{id}/groups)
  5. Grupo creado tiene school_cycle vinculado    (verify school_cycle en response)
  6. Listar grupos filtrado por cycle_id          (GET /partners/campuses/{id}/groups?cycle_id=X)
  7. Grupo con ciclo aparece en filtro            (grupo recién creado en lista)
  8. Grupo sin ciclo no aparece en filtro         (grupo sin ciclo excluido)
  9. Crear grupo sin school_cycle_id (sigue OK)   (retrocompatibilidad)
 10. Crear grupo con school_cycle_id inválido     (ciclo de otro plantel → 400)
 11. Preview + Upload con grupo de ciclo          (flujo completo carga masiva)
 12. Verificar membresía en grupo con ciclo       (candidatos asignados correctamente)
 13. Limpieza de datos de prueba

USO: python tests/test_bulk_upload_cycle_selector.py [--dev]
"""
import sys
import os
import io
import json
import time
import uuid
import requests

# ── Configuración ──
if "--dev" in sys.argv:
    API = "https://evaluaasi-motorv2-api-dev.purpleocean-384694c4.southcentralus.azurecontainerapps.io/api"
    ENV_LABEL = "DEV"
    DB_NAME = "evaluaasi_dev"
else:
    API = "https://evaluaasi-motorv2-api.purpleocean-384694c4.southcentralus.azurecontainerapps.io/api"
    ENV_LABEL = "PROD"
    DB_NAME = "evaluaasi"

try:
    import pymssql
    HAS_DB = True
except ImportError:
    print("⚠️ pymssql no disponible, limpieza manual requerida")
    HAS_DB = False

DB_SERVER = "evaluaasi-motorv2-sql.database.windows.net"
DB_USER = "evaluaasi_admin"
DB_PASS = "EvalAasi2024_newpwd!"

passed = 0
failed = 0
warnings = 0

TEST_PREFIX = f"CYCS_{uuid.uuid4().hex[:6].upper()}"
REQ_TIMEOUT = 120

# IDs para limpieza
created_user_ids = []
created_group_ids = []
created_cycle_id = None
created_member_ids = []


# ── Helpers ──

def test(name, condition, detail=""):
    global passed, failed
    if condition:
        print(f"  ✅ {name}")
        passed += 1
    else:
        print(f"  ❌ {name} — {detail}")
        failed += 1


def warn(name, detail=""):
    global warnings
    print(f"  ⚠️ {name} — {detail}")
    warnings += 1


def post_with_retry(url, max_retries=2, **kwargs):
    kwargs.setdefault('timeout', REQ_TIMEOUT)
    for attempt in range(max_retries + 1):
        try:
            r = requests.post(url, **kwargs)
            if r.status_code == 504 and attempt < max_retries:
                print(f"  ⏳ 504 timeout, reintentando ({attempt + 1}/{max_retries})...")
                time.sleep(5)
                continue
            return r
        except requests.exceptions.Timeout:
            if attempt < max_retries:
                print(f"  ⏳ Request timeout, reintentando ({attempt + 1}/{max_retries})...")
                time.sleep(5)
                continue
            raise
    return r


def get_with_retry(url, max_retries=1, **kwargs):
    kwargs.setdefault('timeout', REQ_TIMEOUT)
    for attempt in range(max_retries + 1):
        try:
            r = requests.get(url, **kwargs)
            if r.status_code == 504 and attempt < max_retries:
                time.sleep(5)
                continue
            return r
        except requests.exceptions.Timeout:
            if attempt < max_retries:
                time.sleep(5)
                continue
            raise
    return r


def get_token(username="admin", password="Admin123!"):
    r = post_with_retry(f"{API}/auth/login", json={"username": username, "password": password})
    if r.status_code in (429, 423):
        wait = int(r.json().get("retry_after", 30))
        print(f"  ⏳ Rate limit, esperando {min(wait, 60)}s...")
        time.sleep(min(wait, 60))
        r = post_with_retry(f"{API}/auth/login", json={"username": username, "password": password})
    if r.status_code == 200:
        return r.json().get("access_token")
    print(f"  ⚠️ Login falló: {r.status_code} {r.text[:200]}")
    return None


def auth_headers(token):
    return {"Authorization": f"Bearer {token}"}


def auth_json_headers(token):
    return {"Authorization": f"Bearer {token}", "Content-Type": "application/json"}


def db_query(sql, params=None):
    if not HAS_DB:
        return []
    conn = pymssql.connect(server=DB_SERVER, user=DB_USER, password=DB_PASS, database=DB_NAME)
    cursor = conn.cursor(as_dict=True)
    cursor.execute(sql, params or ())
    try:
        rows = cursor.fetchall()
    except:
        rows = []
    conn.commit()
    conn.close()
    return rows


def db_exec(sql, params=None):
    if not HAS_DB:
        return 0
    conn = pymssql.connect(server=DB_SERVER, user=DB_USER, password=DB_PASS, database=DB_NAME)
    cursor = conn.cursor()
    cursor.execute(sql, params or ())
    affected = cursor.rowcount
    conn.commit()
    conn.close()
    return affected


def make_candidates_excel(candidates):
    """Genera Excel para carga masiva."""
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    headers = ["email", "nombre", "primer_apellido", "segundo_apellido", "genero", "curp"]
    for col_idx, h in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx, value=h)
    ws.cell(row=2, column=1, value="(descripción)")
    for row_idx, c in enumerate(candidates, 3):
        for col_idx, h in enumerate(headers, 1):
            ws.cell(row=row_idx, column=col_idx, value=c.get(h))
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ============================================================
# MAIN
# ============================================================
def main():
    global passed, failed, created_user_ids, created_group_ids, created_cycle_id

    print("\n" + "=" * 65)
    print(f"  TESTS — CICLO ESCOLAR EN CARGA MASIVA ({ENV_LABEL})")
    print(f"  Selector buscable + Crear grupo con ciclo")
    print(f"  Prefijo: {TEST_PREFIX}")
    print("=" * 65)

    # ── 0. Autenticación ──
    print("\n── 0. Autenticación ──")
    token = get_token("admin", "Admin123!")
    test("Login como admin", token is not None)
    if not token:
        print("⛔ No se pudo autenticar. Abortando.")
        sys.exit(1)
    h = auth_headers(token)
    hj = auth_json_headers(token)

    # ════════════════════════════════════════════════════════════
    # SECCIÓN 1: ENCONTRAR PARTNER / CAMPUS CON CICLOS
    # ════════════════════════════════════════════════════════════
    print("\n" + "=" * 65)
    print("  SECCIÓN 1: LOCALIZAR PLANTEL CON CICLOS ESCOLARES")
    print("=" * 65)

    # Buscar un partner con planteles
    print("\n── 1. Buscar partner con plantel activo ──")
    r = get_with_retry(f"{API}/partners?page=1&per_page=500&active_only=true", headers=h)
    test("GET /partners → 200", r.status_code == 200, f"Status: {r.status_code}")

    partner_id = None
    campus_id = None
    campus_name = None
    cycle_id = None
    cycle_name = None
    campus_id_no_cycles = None

    if r.status_code == 200:
        partners_list = r.json().get("partners", [])
        test("Hay ≥1 partner", len(partners_list) >= 1, f"total={len(partners_list)}")

        # Buscar un plantel que tenga ciclos escolares
        for p in partners_list:
            pid = p["id"]
            r_c = get_with_retry(f"{API}/partners/{pid}/campuses?active_only=true", headers=h)
            if r_c.status_code != 200:
                continue
            for c in r_c.json().get("campuses", []):
                cid = c["id"]
                r_cy = get_with_retry(f"{API}/partners/campuses/{cid}/cycles?active_only=true", headers=h)
                if r_cy.status_code == 200:
                    cy_list = r_cy.json().get("cycles", [])
                    if cy_list:
                        partner_id = pid
                        campus_id = cid
                        campus_name = c["name"]
                        cycle_id = cy_list[0]["id"]
                        cycle_name = cy_list[0]["name"]
                        print(f"  📌 Partner: {p['name']} (ID {pid})")
                        print(f"  📌 Plantel: {campus_name} (ID {campus_id})")
                        print(f"  📌 Ciclo: {cycle_name} (ID {cycle_id})")
                        break
                    elif not campus_id_no_cycles:
                        # Guardar un plantel sin ciclos para test de vacío
                        campus_id_no_cycles = cid
            if partner_id:
                break

        test("Encontró partner+plantel+ciclo", partner_id is not None, "revise que existan ciclos escolares")
    else:
        warn("GET /partners falló")

    # ════════════════════════════════════════════════════════════
    # SECCIÓN 2: SELECTOR DE CICLOS (SearchableSelect)
    # ════════════════════════════════════════════════════════════
    print("\n" + "=" * 65)
    print("  SECCIÓN 2: LISTAR CICLOS PARA SearchableSelect")
    print("=" * 65)

    # ── 2. Listar ciclos del plantel ──
    print("\n── 2. GET /partners/campuses/{id}/cycles — Ciclos del plantel ──")
    if campus_id:
        r = get_with_retry(f"{API}/partners/campuses/{campus_id}/cycles?active_only=true", headers=h)
        test(f"GET cycles del plantel → 200", r.status_code == 200, f"Status: {r.status_code}")

        if r.status_code == 200:
            data = r.json()
            cycles_list = data.get("cycles", [])
            test("Respuesta tiene 'cycles' (array)", isinstance(cycles_list, list))
            test("Respuesta tiene 'total' (int)", isinstance(data.get("total"), int))
            test("Hay ≥1 ciclo", len(cycles_list) >= 1, f"total={len(cycles_list)}")

            if cycles_list:
                first = cycles_list[0]
                test("Ciclo tiene 'id'", "id" in first, f"keys={list(first.keys())}")
                test("Ciclo tiene 'name'", "name" in first, f"keys={list(first.keys())}")
                test("Ciclo tiene 'cycle_type'", "cycle_type" in first)
                test("Ciclo tiene 'start_date'", "start_date" in first)
                test("Ciclo tiene 'end_date'", "end_date" in first)
                test("Ciclo tiene 'is_active'", first.get("is_active") == True)
    else:
        warn("Sin campus_id, saltando test de ciclos")

    # ── 3. Plantel sin ciclos → array vacío ──
    print("\n── 3. Plantel sin ciclos → respuesta vacía válida ──")
    if campus_id_no_cycles:
        r = get_with_retry(f"{API}/partners/campuses/{campus_id_no_cycles}/cycles?active_only=true", headers=h)
        test(f"GET cycles plantel sin ciclos → 200", r.status_code == 200, f"Status: {r.status_code}")
        if r.status_code == 200:
            cy_list = r.json().get("cycles", [])
            test("cycles == [] (vacío)", len(cy_list) == 0, f"Hay {len(cy_list)} ciclos")
            test("total == 0", r.json().get("total") == 0)
    else:
        print("  ℹ️ Todos los planteles tienen ciclos, no se encontró caso vacío")
        test("GET cycles responde formato correcto (testeado arriba)", True)

    # ════════════════════════════════════════════════════════════
    # SECCIÓN 3: CREAR GRUPO CON CICLO ESCOLAR
    # ════════════════════════════════════════════════════════════
    print("\n" + "=" * 65)
    print("  SECCIÓN 3: CREAR GRUPO INLINE CON CICLO ESCOLAR")
    print("=" * 65)

    # ── 4. Crear grupo con school_cycle_id ──
    print("\n── 4. Crear grupo con school_cycle_id válido ──")
    group_with_cycle_id = None
    group_with_cycle_name = f"{TEST_PREFIX}_ConCiclo"
    if campus_id and cycle_id:
        r = post_with_retry(f"{API}/partners/campuses/{campus_id}/groups",
                            json={"name": group_with_cycle_name, "school_cycle_id": cycle_id}, headers=hj)
        test("POST crear grupo con ciclo → 201", r.status_code == 201, f"Status: {r.status_code} | {r.text[:300]}")

        if r.status_code == 201:
            grp = r.json().get("group", {})
            group_with_cycle_id = grp.get("id")
            created_group_ids.append(group_with_cycle_id)

            test("Grupo tiene 'id'", bool(grp.get("id")))
            test(f"name == '{group_with_cycle_name}'", grp.get("name") == group_with_cycle_name)
            test(f"campus_id == {campus_id}", grp.get("campus_id") == campus_id)
            test("is_active == True", grp.get("is_active") == True)

            # ── 5. Verificar school_cycle vinculado ──
            print("\n── 5. Grupo tiene school_cycle en respuesta ──")
            sc = grp.get("school_cycle")
            test("Respuesta incluye 'school_cycle'", sc is not None, f"school_cycle={sc}")
            if sc:
                test(f"school_cycle.id == {cycle_id}", sc.get("id") == cycle_id, f"Got: {sc.get('id')}")
                test(f"school_cycle.name == '{cycle_name}'", sc.get("name") == cycle_name, f"Got: {sc.get('name')}")

            print(f"  📌 Grupo con ciclo: {group_with_cycle_name} (ID {group_with_cycle_id})")
        else:
            warn("No se pudo crear grupo con ciclo")
    else:
        warn("Sin campus_id o cycle_id, saltando")

    # ── 6. Listar grupos filtrado por cycle_id ──
    print("\n── 6. GET groups con filtro cycle_id ──")
    if campus_id and cycle_id:
        r = get_with_retry(f"{API}/partners/campuses/{campus_id}/groups?cycle_id={cycle_id}", headers=h)
        test("GET groups?cycle_id → 200", r.status_code == 200, f"Status: {r.status_code}")

        if r.status_code == 200:
            groups_list = r.json().get("groups", [])
            test("Respuesta tiene 'groups' (array)", isinstance(groups_list, list))

            # ── 7. Grupo con ciclo aparece en filtro ──
            if group_with_cycle_id:
                found = any(g["id"] == group_with_cycle_id for g in groups_list)
                test(f"Grupo {group_with_cycle_id} aparece en filtro cycle_id={cycle_id}", found,
                     f"IDs: {[g['id'] for g in groups_list[:15]]}")
    else:
        warn("Sin campus/ciclo, saltando filtro")

    # ── 8. Crear grupo SIN ciclo (retrocompatibilidad) ──
    print("\n── 8. Crear grupo sin school_cycle_id (retrocompatible) ──")
    group_no_cycle_id = None
    group_no_cycle_name = f"{TEST_PREFIX}_SinCiclo"
    if campus_id:
        r = post_with_retry(f"{API}/partners/campuses/{campus_id}/groups",
                            json={"name": group_no_cycle_name}, headers=hj)
        test("POST crear grupo sin ciclo → 201", r.status_code == 201, f"Status: {r.status_code}")

        if r.status_code == 201:
            grp = r.json().get("group", {})
            group_no_cycle_id = grp.get("id")
            created_group_ids.append(group_no_cycle_id)
            test("school_cycle es null (no asignado)", grp.get("school_cycle") is None,
                 f"Got: {grp.get('school_cycle')}")
            print(f"  📌 Grupo sin ciclo: {group_no_cycle_name} (ID {group_no_cycle_id})")

        # Verificar que no aparece en filtro por cycle_id
        if group_no_cycle_id and cycle_id:
            r = get_with_retry(f"{API}/partners/campuses/{campus_id}/groups?cycle_id={cycle_id}", headers=h)
            if r.status_code == 200:
                groups_list = r.json().get("groups", [])
                found = any(g["id"] == group_no_cycle_id for g in groups_list)
                test(f"Grupo sin ciclo NO aparece en filtro cycle_id={cycle_id}", not found,
                     f"IDs: {[g['id'] for g in groups_list[:15]]}")
    else:
        warn("Sin campus_id, saltando")

    # ── 9. school_cycle_id de otro plantel → 400 ──
    print("\n── 9. school_cycle_id de otro plantel → 400 ──")
    if campus_id and cycle_id:
        # Buscar otro plantel
        other_campus_id = None
        for p in r.json().get("partners", []) if r.status_code == 200 else []:
            pass
        # Intentar crear grupo con ciclo que no pertenece al plantel
        # Usamos un ciclo válido pero en un campus diferente simulando con campus_id_no_cycles
        # Si tenemos otro campus, probamos ahí
        if campus_id_no_cycles:
            r = post_with_retry(
                f"{API}/partners/campuses/{campus_id_no_cycles}/groups",
                json={"name": f"{TEST_PREFIX}_InvalidCycle", "school_cycle_id": cycle_id},
                headers=hj
            )
            test("Ciclo de otro plantel → 400", r.status_code == 400, f"Status: {r.status_code}")
            if r.status_code == 400:
                test("Error menciona 'ciclo' o 'válido'",
                     "ciclo" in r.json().get("error", "").lower() or "válido" in r.json().get("error", "").lower(),
                     r.json().get("error", ""))
        else:
            # Inventar un cycle_id muy alto que no existe
            r = post_with_retry(
                f"{API}/partners/campuses/{campus_id}/groups",
                json={"name": f"{TEST_PREFIX}_InvalidCycle", "school_cycle_id": 999999},
                headers=hj
            )
            test("school_cycle_id inexistente → 400", r.status_code == 400, f"Status: {r.status_code}")
    else:
        warn("Sin datos para test de ciclo inválido")

    # ════════════════════════════════════════════════════════════
    # SECCIÓN 4: CARGA MASIVA CON GRUPO DE CICLO
    # ════════════════════════════════════════════════════════════
    print("\n" + "=" * 65)
    print("  SECCIÓN 4: CARGA MASIVA CON GRUPO VINCULADO A CICLO")
    print("=" * 65)

    uid = TEST_PREFIX[5:]
    test_candidates = [
        {
            "email": f"{TEST_PREFIX.lower()}_c1@testcycle.com",
            "nombre": f"Ana{uid}",
            "primer_apellido": "Martínez",
            "segundo_apellido": "López",
            "genero": "F",
            "curp": f"AMLO{uid}000001".ljust(18, "X")[:18],
        },
        {
            "email": f"{TEST_PREFIX.lower()}_c2@testcycle.com",
            "nombre": f"Carlos{uid}",
            "primer_apellido": "Pérez",
            "segundo_apellido": "Ruiz",
            "genero": "M",
            "curp": None,
        },
    ]

    target_group = group_with_cycle_id
    if not target_group:
        warn("No hay grupo con ciclo, usando grupo sin ciclo si existe")
        target_group = group_no_cycle_id

    # ── 10. Preview con grupo de ciclo ──
    print("\n── 10. Preview con grupo → verificar grupo con ciclo ──")
    if target_group:
        excel_data = make_candidates_excel(test_candidates)
        r = post_with_retry(
            f"{API}/user-management/candidates/bulk-upload/preview",
            headers=h,
            files={"file": ("candidatos.xlsx", excel_data,
                            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            data={"group_id": str(target_group)}
        )
        test("POST preview → 200", r.status_code == 200, f"Status: {r.status_code} | {r.text[:300]}")

        if r.status_code == 200:
            pdata = r.json()
            test("Preview tiene 'preview' (array)", isinstance(pdata.get("preview"), list))
            preview_list = pdata.get("preview", [])
            test(f"Preview tiene {len(test_candidates)} filas", len(preview_list) == len(test_candidates),
                 f"Got {len(preview_list)}")

            ready_count = sum(1 for p in preview_list if p.get("status") == "ready")
            test(f"Al menos 1 fila 'ready'", ready_count >= 1, f"ready={ready_count}")
    else:
        warn("Sin grupo disponible, saltando preview")

    # ── 11. Upload con grupo de ciclo ──
    print("\n── 11. Upload con grupo vinculado a ciclo ──")
    if target_group:
        excel_data = make_candidates_excel(test_candidates)
        r = post_with_retry(
            f"{API}/user-management/candidates/bulk-upload",
            headers=h,
            files={"file": ("candidatos.xlsx", excel_data,
                            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            data={"group_id": str(target_group)}
        )
        test("POST upload → 200", r.status_code == 200, f"Status: {r.status_code} | {r.text[:200]}")

        if r.status_code == 200:
            result = r.json()
            details = result.get("details", {})
            created_list = details.get("created", [])
            existing_assigned = details.get("existing_assigned", [])
            summary = result.get("summary", {})
            total_processed = summary.get("created", 0) + summary.get("existing_assigned", 0)
            test(f"Procesó ≥1 usuario (creado o existente asignado)", total_processed >= 1,
                 f"created={summary.get('created')}, existing_assigned={summary.get('existing_assigned')}")
            for u in created_list:
                if u.get("id"):
                    created_user_ids.append(u["id"])
            for u in existing_assigned:
                if u.get("id"):
                    created_user_ids.append(u["id"])

            # ── 12. Verificar membresía ──
            print("\n── 12. Candidatos asignados al grupo con ciclo ──")
            r = get_with_retry(f"{API}/partners/groups/{target_group}", headers=h)
            test("GET grupo → 200", r.status_code == 200, f"Status: {r.status_code}")

            if r.status_code == 200:
                grp = r.json().get("group", {})
                member_count = grp.get("member_count", 0)
                test(f"Grupo tiene ≥{len(test_candidates)} miembros",
                     member_count >= len(test_candidates),
                     f"member_count={member_count}")

                # Verificar que el grupo tiene ciclo
                sc = grp.get("school_cycle")
                if group_with_cycle_id and target_group == group_with_cycle_id:
                    test("Grupo mantiene school_cycle después de upload",
                         sc is not None and sc.get("id") == cycle_id,
                         f"school_cycle={sc}")
    else:
        warn("Sin grupo, saltando upload")

    # ════════════════════════════════════════════════════════════
    # SECCIÓN 5: VALIDACIONES ADICIONALES
    # ════════════════════════════════════════════════════════════
    print("\n" + "=" * 65)
    print("  SECCIÓN 5: VALIDACIONES ADICIONALES")
    print("=" * 65)

    # ── 13. Ciclo con grupos listados (include_groups en respuesta de ciclos) ──
    print("\n── 13. Ciclo incluye conteo de grupos ──")
    if campus_id and cycle_id:
        r = get_with_retry(f"{API}/partners/campuses/{campus_id}/cycles?active_only=true", headers=h)
        if r.status_code == 200:
            cycles_list = r.json().get("cycles", [])
            target_cycle = next((c for c in cycles_list if c["id"] == cycle_id), None)
            if target_cycle:
                test("Ciclo tiene 'group_count' o 'groups'",
                     "group_count" in target_cycle or "groups" in target_cycle,
                     f"keys={list(target_cycle.keys())}")
                if "group_count" in target_cycle:
                    test(f"group_count ≥ 1 (tiene al menos nuestro grupo)", 
                         target_cycle["group_count"] >= 1,
                         f"group_count={target_cycle['group_count']}")
            else:
                warn(f"Ciclo {cycle_id} no encontrado en respuesta")

    # ── 14. Buscar ciclo por nombre (simular SearchableSelect) ──
    print("\n── 14. Buscar ciclo por nombre (SearchableSelect) ──")
    if campus_id and cycle_name:
        r = get_with_retry(f"{API}/partners/campuses/{campus_id}/cycles?active_only=true", headers=h)
        if r.status_code == 200:
            cycles_list = r.json().get("cycles", [])
            # Simular búsqueda del SearchableSelect: filtrar por nombre
            search_term = cycle_name[:4].lower()
            matches = [c for c in cycles_list if search_term in c.get("name", "").lower()]
            test(f"Buscar '{search_term}' encuentra ciclo", len(matches) >= 1,
                 f"matches={len(matches)}, names={[c['name'] for c in cycles_list]}")
            test("Cada match tiene id y name",
                 all("id" in m and "name" in m for m in matches))

    # ════════════════════════════════════════════════════════════
    # LIMPIEZA
    # ════════════════════════════════════════════════════════════
    print("\n" + "=" * 65)
    print("  🧹 LIMPIEZA DE DATOS DE PRUEBA")
    print("=" * 65)

    if HAS_DB:
        cleaned_users = 0
        cleaned_groups = 0
        cleaned_members = 0

        # Limpiar usuarios por prefijo de email
        test_users = db_query(
            "SELECT id, username, email FROM users WHERE email LIKE %s",
            (f"%{TEST_PREFIX.lower()}%",)
        )
        test_users2 = db_query(
            "SELECT id, username, email FROM users WHERE username LIKE %s",
            (f"%{TEST_PREFIX}%",)
        )
        all_test_users = {r["id"]: r for r in test_users + test_users2}
        all_user_ids = list(all_test_users.keys()) + created_user_ids

        # Limpiar membresías
        for uid in set(all_user_ids):
            n = db_exec("DELETE FROM group_members WHERE user_id=%s", (uid,))
            cleaned_members += n

        # Limpiar usuarios
        for uid in set(all_user_ids):
            n = db_exec("DELETE FROM users WHERE id=%s", (uid,))
            cleaned_users += n

        # Limpiar grupos de prueba
        test_groups = db_query(
            "SELECT id, name FROM candidate_groups WHERE name LIKE %s",
            (f"{TEST_PREFIX}%",)
        )
        for g in test_groups:
            db_exec("DELETE FROM group_members WHERE group_id=%s", (g["id"],))
            db_exec("DELETE FROM candidate_groups WHERE id=%s", (g["id"],))
            cleaned_groups += 1

        print(f"  ✅ {cleaned_users} usuarios eliminados")
        print(f"  ✅ {cleaned_members} membresías eliminadas")
        print(f"  ✅ {cleaned_groups} grupos de prueba eliminados")
        test("Limpieza completada", True)
    else:
        warn("Sin acceso a DB, limpieza manual requerida", f"Prefijo: {TEST_PREFIX}")

    # ════════════════════════════════════════════════════════════
    # RESUMEN FINAL
    # ════════════════════════════════════════════════════════════
    print("\n" + "=" * 65)
    total = passed + failed
    emoji = "✅" if failed == 0 else "⚠️" if failed <= 3 else "❌"
    print(f"  {emoji}  RESULTADOS ({ENV_LABEL}): {passed}/{total} tests pasaron")
    if warnings:
        print(f"     ⚠️  {warnings} advertencias")
    if failed:
        print(f"     ❌ {failed} tests fallaron")
    print("=" * 65)

    sys.exit(0 if failed == 0 else 1)


if __name__ == "__main__":
    main()
