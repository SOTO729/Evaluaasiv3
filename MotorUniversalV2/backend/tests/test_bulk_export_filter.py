"""
Tests del filtro de status en el export de carga masiva.
Verifica que GET /user-management/bulk-history/:id/export?status=X
filtra correctamente los miembros exportados en el Excel.

Escenarios:
  1. Export sin filtro devuelve Excel con Content-Type correcto
  2. Export sin filtro incluye todos los miembros
  3. Export con status válido filtra solo los miembros de ese status
  4. Export con status sin matches devuelve Excel solo con header (0 filas data)
  5. Export con status inválido/inyección ignora el filtro y devuelve todos
  6. Export requiere autenticación (401 sin token)
  7. Cada status válido funciona como filtro individualmente
  8. El tamaño del Excel filtrado <= tamaño del Excel sin filtro
"""

import time
import io
import requests
import pytest

DEV_API = "https://evaluaasi-motorv2-api-dev.purpleocean-384694c4.southcentralus.azurecontainerapps.io/api"
VALID_STATUSES = ['created', 'existing_assigned', 'error', 'skipped', 'curp_invalid', 'curp_verified']


@pytest.fixture(scope="session")
def api():
    return DEV_API


@pytest.fixture(scope="session")
def admin_token(api):
    for attempt in range(3):
        r = requests.post(
            f"{api}/auth/login",
            json={"username": "admin", "password": "admin123"},
            timeout=30,
        )
        if r.status_code == 200:
            return r.json()["access_token"]
        if r.status_code in [502, 503]:
            time.sleep(10)
    pytest.fail("Admin login failed after 3 attempts")


@pytest.fixture(scope="session")
def headers(admin_token):
    return {"Authorization": f"Bearer {admin_token}"}


@pytest.fixture(scope="session")
def test_batch(api, headers):
    """Find a batch with >0 members for testing."""
    r = requests.get(f"{api}/user-management/bulk-history?per_page=50", headers=headers, timeout=15)
    assert r.status_code == 200
    batches = r.json()["batches"]
    for b in batches:
        if b["total_processed"] > 0:
            detail = requests.get(f"{api}/user-management/bulk-history/{b['id']}", headers=headers, timeout=15)
            if detail.status_code == 200:
                data = detail.json()
                if data.get("members") and len(data["members"]) > 0:
                    return data
    pytest.skip("No batch with members found in DEV")


@pytest.fixture(scope="session")
def batch_statuses(test_batch):
    """Map of status -> count for the test batch."""
    counts = {}
    for m in test_batch["members"]:
        s = m["status"]
        counts[s] = counts.get(s, 0) + 1
    return counts


def _parse_xlsx_rows(content_bytes):
    """Parse Excel bytes and return number of data rows (excluding header)."""
    try:
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(content_bytes))
        ws = wb.active
        return ws.max_row - 1  # subtract header
    except ImportError:
        pytest.skip("openpyxl not installed locally")


# ── Tests ──

def test_export_returns_xlsx(api, headers, test_batch):
    """T1: Export sin filtro devuelve Excel válido."""
    r = requests.get(
        f"{api}/user-management/bulk-history/{test_batch['id']}/export",
        headers=headers, timeout=20,
    )
    assert r.status_code == 200
    assert "spreadsheetml" in r.headers.get("Content-Type", "")


def test_export_no_filter_all_members(api, headers, test_batch):
    """T2: Export sin filtro incluye todos los miembros."""
    r = requests.get(
        f"{api}/user-management/bulk-history/{test_batch['id']}/export",
        headers=headers, timeout=20,
    )
    assert r.status_code == 200
    rows = _parse_xlsx_rows(r.content)
    assert rows == len(test_batch["members"]), f"Expected {len(test_batch['members'])} rows, got {rows}"


def test_export_with_matching_status(api, headers, test_batch, batch_statuses):
    """T3: Export con status válido que existe filtra correctamente."""
    # Pick a status that exists in this batch
    status = list(batch_statuses.keys())[0]
    expected = batch_statuses[status]

    r = requests.get(
        f"{api}/user-management/bulk-history/{test_batch['id']}/export?status={status}",
        headers=headers, timeout=20,
    )
    assert r.status_code == 200
    rows = _parse_xlsx_rows(r.content)
    assert rows == expected, f"Filtered by {status}: expected {expected} rows, got {rows}"


def test_export_with_no_match_status(api, headers, test_batch, batch_statuses):
    """T4: Export con status sin matches devuelve solo header."""
    # Find a valid status NOT present in this batch
    missing = [s for s in VALID_STATUSES if s not in batch_statuses]
    if not missing:
        pytest.skip("Batch has all possible statuses")
    status = missing[0]

    r = requests.get(
        f"{api}/user-management/bulk-history/{test_batch['id']}/export?status={status}",
        headers=headers, timeout=20,
    )
    assert r.status_code == 200
    rows = _parse_xlsx_rows(r.content)
    assert rows == 0, f"Expected 0 rows for status={status}, got {rows}"


def test_export_ignores_invalid_status(api, headers, test_batch):
    """T5: Export con status inválido ignora filtro, devuelve todos."""
    total = len(test_batch["members"])

    for bad_status in ["invalid", "DROP TABLE", "<script>", "' OR 1=1 --"]:
        r = requests.get(
            f"{api}/user-management/bulk-history/{test_batch['id']}/export",
            params={"status": bad_status},
            headers=headers, timeout=20,
        )
        assert r.status_code == 200
        rows = _parse_xlsx_rows(r.content)
        assert rows == total, f"Invalid status '{bad_status}': expected all {total} rows, got {rows}"


def test_export_requires_auth(api, test_batch):
    """T6: Export sin token devuelve 401."""
    r = requests.get(
        f"{api}/user-management/bulk-history/{test_batch['id']}/export",
        timeout=10,
    )
    assert r.status_code in [401, 422], f"Expected 401/422, got {r.status_code}"


def test_each_valid_status_filter(api, headers, test_batch, batch_statuses):
    """T7: Cada status válido funciona como filtro (devuelve 200)."""
    for status in VALID_STATUSES:
        r = requests.get(
            f"{api}/user-management/bulk-history/{test_batch['id']}/export?status={status}",
            headers=headers, timeout=20,
        )
        assert r.status_code == 200, f"status={status} returned {r.status_code}"
        rows = _parse_xlsx_rows(r.content)
        expected = batch_statuses.get(status, 0)
        assert rows == expected, f"status={status}: expected {expected}, got {rows}"


def test_filtered_size_leq_unfiltered(api, headers, test_batch, batch_statuses):
    """T8: El tamaño del Excel filtrado <= sin filtro."""
    r_all = requests.get(
        f"{api}/user-management/bulk-history/{test_batch['id']}/export",
        headers=headers, timeout=20,
    )
    size_all = len(r_all.content)

    status = list(batch_statuses.keys())[0]
    r_filtered = requests.get(
        f"{api}/user-management/bulk-history/{test_batch['id']}/export?status={status}",
        headers=headers, timeout=20,
    )
    size_filtered = len(r_filtered.content)
    assert size_filtered <= size_all, f"Filtered ({size_filtered}) > unfiltered ({size_all})"


if __name__ == "__main__":
    pytest.main([__file__, "-v"])
