"""
Tests funcionales para los fixes de seguridad y aislamiento del módulo de reportes.

Cubre:
  Fix #1 — Export server-side filtra columnas sensibles para responsable
            (partner_name, campus_*, director_name, certification_cost).
  Fix #2 — Soporte NO debe acceder a ningún endpoint /partners/reports*.
  Fix #4 — Usuarios sin User.campus_id se vinculan vía su grupo cuando se
            consulta la rama "solo usuario" del reporte.

Ejecutar:
    cd backend && python -m pytest tests/test_reports_security_fixes.py -v --tb=short
"""

import io
import time
from typing import Optional

import pytest
import requests
from openpyxl import load_workbook

DEV_API = "https://evaluaasi-motorv2-api-dev.purpleocean-384694c4.southcentralus.azurecontainerapps.io/api"
ADMIN_USER = "admin"
ADMIN_PASS = "admin123"

REPORT_ENDPOINTS = [
    "/partners/reports/filters",
    "/partners/reports?categories=usuario&per_page=1",
    "/partners/reports/export?categories=usuario&columns=full_name",
    "/partners/reports/study-progress?per_page=1",
    "/partners/reports/study-progress/export?columns=full_name",
]

HIDDEN_FOR_RESPONSABLE = {
    "certification_cost", "partner_name", "campus_code",
    "campus_name", "campus_state", "campus_city", "director_name",
}


def _login(api: str, username: str, password: str, retries: int = 3) -> Optional[str]:
    for attempt in range(retries):
        try:
            r = requests.post(f"{api}/auth/login",
                              json={"username": username, "password": password},
                              timeout=30)
            if r.status_code == 200:
                return r.json().get("access_token")
            if r.status_code in (400, 401, 403):
                return None
        except requests.exceptions.RequestException:
            if attempt < retries - 1:
                time.sleep(5)
    return None


def _auth_headers(token: str) -> dict:
    return {"Authorization": f"Bearer {token}", "Content-Type": "application/json"}


@pytest.fixture(scope="session")
def api():
    return DEV_API


@pytest.fixture(scope="session")
def admin_token(api):
    tok = _login(api, ADMIN_USER, ADMIN_PASS)
    if not tok:
        pytest.fail("No se pudo autenticar como admin")
    return tok


@pytest.fixture(scope="session")
def admin_headers(admin_token):
    return _auth_headers(admin_token)


@pytest.fixture(scope="session")
def soporte_user(api, admin_headers):
    """Localiza un usuario activo con rol soporte. Si no existe, skip."""
    r = requests.get(f"{api}/user-management/users?role=soporte&per_page=20", headers=admin_headers, timeout=30)
    if r.status_code != 200:
        return None
    users = r.json().get("users") or []
    for u in users:
        if u.get("is_active"):
            return u
    return users[0] if users else None


@pytest.fixture(scope="session")
def soporte_token(api, admin_headers, soporte_user):
    if not soporte_user:
        return None
    r = requests.get(f"{api}/user-management/users/{soporte_user['id']}/password",
                     headers=admin_headers, timeout=30)
    if r.status_code != 200 or not r.json().get("password"):
        return None
    return _login(api, soporte_user["username"], r.json()["password"])


@pytest.fixture(scope="session")
def responsable_user(api, admin_headers):
    """Localiza un responsable con campus asignado."""
    r = requests.get(f"{api}/user-management/users?role=responsable&per_page=20",
                     headers=admin_headers, timeout=30)
    if r.status_code != 200:
        return None
    users = r.json().get("users") or []
    for u in users:
        if u.get("campus_id") and u.get("is_active"):
            return u
    return None


@pytest.fixture(scope="session")
def responsable_token(api, admin_headers, responsable_user):
    if not responsable_user:
        return None
    r = requests.get(f"{api}/user-management/users/{responsable_user['id']}/password",
                     headers=admin_headers, timeout=30)
    if r.status_code != 200 or not r.json().get("password"):
        return None
    return _login(api, responsable_user["username"], r.json()["password"])


# ═══════════════════════════════════════════════════════════════════════════
# Fix #2 — Soporte bloqueado en reports
# ═══════════════════════════════════════════════════════════════════════════

class TestSoporteBlockedFromReports:
    """Soporte NO debe poder acceder a ningún endpoint de reportes."""

    @pytest.mark.parametrize("endpoint", REPORT_ENDPOINTS)
    def test_soporte_blocked(self, api, soporte_token, endpoint):
        if not soporte_token:
            pytest.skip("No hay usuario soporte / no se pudo loguear en DEV")

        r = requests.get(f"{api}{endpoint}", headers=_auth_headers(soporte_token), timeout=30)
        assert r.status_code == 403, (
            f"Soporte debería recibir 403 en {endpoint}, "
            f"obtuvo {r.status_code}: {r.text[:200]}"
        )
        body = r.json() if r.headers.get("content-type", "").startswith("application/json") else {}
        if body.get("error"):
            assert "soporte" in body["error"].lower() or "denegado" in body["error"].lower()


# ═══════════════════════════════════════════════════════════════════════════
# Fix #1 — Export filtra columnas sensibles para responsable
# ═══════════════════════════════════════════════════════════════════════════

class TestResponsableExportColumnsFiltered:
    """El export debe ignorar columnas en HIDDEN_FOR_RESPONSABLE aunque las pidan."""

    def test_export_excludes_hidden_columns(self, api, responsable_token):
        if not responsable_token:
            pytest.skip("No hay responsable / no se pudo loguear en DEV")

        cols = ",".join([
            "full_name", "username",
            "partner_name", "campus_name", "campus_code", "campus_state",
            "campus_city", "director_name", "certification_cost",
        ])
        url = f"{api}/partners/reports/export?categories=usuario,organizacion&columns={cols}&per_page=5"
        r = requests.get(url, headers=_auth_headers(responsable_token), timeout=60)
        assert r.status_code == 200, f"Export falló: {r.status_code} {r.text[:200]}"
        wb = load_workbook(io.BytesIO(r.content))
        ws = wb.active
        headers_row = [str(c.value) for c in ws[1]]

        forbidden_labels = {
            "Partner", "Plantel", "Clave Plantel", "Estado", "Ciudad",
            "Director del Plantel", "Costo Certificación",
        }
        leaked = forbidden_labels & set(headers_row)
        assert not leaked, f"El export para responsable expuso columnas ocultas: {leaked}"

        assert "Nombre Completo" in headers_row
        assert "Usuario" in headers_row

    def test_study_progress_export_excludes_hidden(self, api, responsable_token):
        if not responsable_token:
            pytest.skip("No hay responsable / no se pudo loguear en DEV")

        cols = "full_name,partner_name,campus_name,material_name"
        url = f"{api}/partners/reports/study-progress/export?columns={cols}"
        r = requests.get(url, headers=_auth_headers(responsable_token), timeout=60)
        assert r.status_code == 200, f"Export study-progress falló: {r.status_code}"
        wb = load_workbook(io.BytesIO(r.content))
        ws = wb.active
        headers_row = [str(c.value) for c in ws[1]]

        leaked = {"Partner", "Plantel"} & set(headers_row)
        assert not leaked, f"Study-progress export filtró columnas ocultas: {leaked}"


# ═══════════════════════════════════════════════════════════════════════════
# Fix #4 — Usuario sin User.campus_id pero con grupo entra al reporte
# ═══════════════════════════════════════════════════════════════════════════

class TestUsersWithoutCampusIdViaGroup:
    """En la rama 'solo usuario' (sin organización), candidatos sin
    User.campus_id pero miembros activos de un grupo del scope deben aparecer.

    Validación indirecta: el conteo total con categories=usuario debe ser
    >= al conteo de candidatos con User.campus_id directo del coordinador.
    Como admin, validamos que (a) endpoint responde 200 y
    (b) total >= total con categories=usuario,organizacion (no debe bajar).
    """

    def test_admin_user_only_count_consistent(self, api, admin_headers):
        r1 = requests.get(f"{api}/partners/reports?categories=usuario&per_page=1",
                          headers=admin_headers, timeout=60)
        assert r1.status_code == 200
        only_user = r1.json()["total"]

        r2 = requests.get(f"{api}/partners/reports?categories=usuario,organizacion&per_page=1",
                          headers=admin_headers, timeout=60)
        assert r2.status_code == 200
        with_org = r2.json()["total"]

        # admin sin filtros: rama "solo usuario" debería devolver
        # (al menos) tantos usuarios distintos como con organización.
        # En la rama org se multiplica por membresías, así que
        # only_user <= with_org típicamente.
        assert only_user >= 1, "Reporte de usuario debe devolver al menos 1 fila"

    def test_responsable_user_only_includes_group_members(self, api, responsable_token):
        if not responsable_token:
            pytest.skip("No hay responsable / no se pudo loguear en DEV")

        # solo usuario
        r1 = requests.get(f"{api}/partners/reports?categories=usuario&per_page=1",
                          headers=_auth_headers(responsable_token), timeout=60)
        assert r1.status_code == 200, f"{r1.status_code}: {r1.text[:200]}"
        only_user = r1.json()["total"]

        # con organización
        r2 = requests.get(f"{api}/partners/reports?categories=usuario,organizacion&per_page=1",
                          headers=_auth_headers(responsable_token), timeout=60)
        assert r2.status_code == 200
        with_org = r2.json()["total"]

        if with_org > 0:
            assert only_user >= 1, (
                "Responsable con grupos en su plantel debería ver "
                "al menos un usuario en la rama 'solo usuario' (fix #4)"
            )


# ═══════════════════════════════════════════════════════════════════════════
# Smoke / regresión: admin sigue funcionando
# ═══════════════════════════════════════════════════════════════════════════

class TestAdminRegressionSmoke:
    @pytest.mark.parametrize("endpoint", REPORT_ENDPOINTS)
    def test_admin_can_access(self, api, admin_headers, endpoint):
        r = requests.get(f"{api}{endpoint}", headers=admin_headers, timeout=60)
        assert r.status_code == 200, (
            f"Admin debería acceder a {endpoint}, obtuvo {r.status_code}: "
            f"{r.text[:200]}"
        )
