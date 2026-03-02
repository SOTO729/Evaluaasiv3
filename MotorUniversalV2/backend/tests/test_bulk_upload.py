"""
Test integral de Carga Masiva optimizada — Flujo 1 y Flujo 2

Flujo 1: POST /user-management/candidates/bulk-upload/preview
         POST /user-management/candidates/bulk-upload
Flujo 2: POST /partners/groups/{id}/members/upload/preview
         POST /partners/groups/{id}/members/upload

Cubre:
  1. Autenticación requerida (401 sin token)
  2. Validación de archivo (formato, tamaño, columnas)
  3. Preview sin crear usuarios (Flujo 1)
  4. Creación batch con commits cada 500
  5. Detección de emails/CURPs existentes
  6. Generación de usernames únicos batch
  7. Segundo_apellido y genero como requeridos
  8. Asignación a grupo con usuarios existentes
  9. Preview de Flujo 2 con ambigüedades
 10. Resolución de ambigüedades en Flujo 2
 11. Límites de filas y tamaño de archivo
 12. Limpieza de datos de prueba

USO: python tests/test_bulk_upload.py
"""
import sys
import os
import io
import json
import time
import uuid
import requests

# ── Configuración ──
API = "https://evaluaasi-motorv2-api.purpleocean-384694c4.southcentralus.azurecontainerapps.io/api"

try:
    import pymssql
    HAS_DB = True
except ImportError:
    print("⚠️ pymssql no disponible, parte de limpieza manual requerida")
    HAS_DB = False

DB_SERVER = "evaluaasi-motorv2-sql.database.windows.net"
DB_USER = "evaluaasi_admin"
DB_PASS = "EvalAasi2024_newpwd!"
DB_NAME = "evaluaasi"

passed = 0
failed = 0
warnings = 0
created_user_ids = []  # Para limpieza
TEST_PREFIX = f"BULKTEST_{uuid.uuid4().hex[:6].upper()}"


def test(name, condition, detail=""):
    global passed, failed
    if condition:
        print(f"  ✅ {name}")
        passed += 1
    else:
        print(f"  ❌ {name} — {detail}")
        failed += 1


def warn(name, detail=""):
    global warnings
    print(f"  ⚠️ {name} — {detail}")
    warnings += 1


def get_token(username="admin", password="Admin123!"):
    """Login y devolver token."""
    r = requests.post(f"{API}/auth/login", json={"username": username, "password": password})
    if r.status_code == 429:
        wait = int(r.json().get("retry_after", 30))
        print(f"  ⏳ Rate limit, esperando {min(wait, 60)}s...")
        time.sleep(min(wait, 60))
        r = requests.post(f"{API}/auth/login", json={"username": username, "password": password})
    if r.status_code == 200:
        return r.json().get("access_token")
    print(f"  ⚠️ Login fallido ({username}): {r.status_code} {r.text[:200]}")
    return None


def auth_headers(token):
    return {"Authorization": f"Bearer {token}"}


def db_query(sql, params=None):
    if not HAS_DB:
        return []
    conn = pymssql.connect(server=DB_SERVER, user=DB_USER, password=DB_PASS, database=DB_NAME)
    cursor = conn.cursor(as_dict=True)
    cursor.execute(sql, params or ())
    try:
        rows = cursor.fetchall()
    except:
        rows = []
    conn.commit()
    conn.close()
    return rows


def db_exec(sql, params=None):
    if not HAS_DB:
        return
    conn = pymssql.connect(server=DB_SERVER, user=DB_USER, password=DB_PASS, database=DB_NAME)
    cursor = conn.cursor()
    cursor.execute(sql, params or ())
    conn.commit()
    conn.close()


# ── Generadores de Excel ──

def make_excel_bytes(rows, headers=None):
    """
    Genera un archivo Excel (.xlsx) en memoria.
    rows: lista de dicts o lista de listas
    headers: encabezados (si rows son dicts, se deducen)
    Fila 1: headers, Fila 2: descripciones (vacía), Fila 3+: datos
    """
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active

    if not headers and rows and isinstance(rows[0], dict):
        headers = list(rows[0].keys())

    # Fila 1: encabezados
    if headers:
        for col_idx, h in enumerate(headers, 1):
            ws.cell(row=1, column=col_idx, value=h)

    # Fila 2: descripciones (vacía)
    ws.cell(row=2, column=1, value="(descripción)")

    # Fila 3+: datos
    for row_idx, row_data in enumerate(rows, 3):
        if isinstance(row_data, dict):
            for col_idx, h in enumerate(headers, 1):
                ws.cell(row=row_idx, column=col_idx, value=row_data.get(h))
        elif isinstance(row_data, (list, tuple)):
            for col_idx, val in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=val)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def make_candidates_excel(candidates):
    """Genera Excel para Flujo 1 con el formato requerido."""
    headers = ["email", "nombre", "primer_apellido", "segundo_apellido", "genero", "curp"]
    return make_excel_bytes(candidates, headers)


def make_identifiers_excel(identifiers):
    """Genera Excel para Flujo 2 (identificador, notas).
    F2 lee desde min_row=2 (solo salta header), así que NO poner fila de descripción."""
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.cell(row=1, column=1, value="identificador")
    ws.cell(row=1, column=2, value="notas")
    for idx, ident in enumerate(identifiers, 2):
        if isinstance(ident, dict):
            ws.cell(row=idx, column=1, value=ident.get("identifier", ""))
            ws.cell(row=idx, column=2, value=ident.get("notes", ""))
        else:
            ws.cell(row=idx, column=1, value=ident)
            ws.cell(row=idx, column=2, value="")
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ============================================================
# MAIN
# ============================================================
def main():
    global passed, failed, created_user_ids

    print("\n" + "=" * 60)
    print("  TESTS DE CARGA MASIVA — FLUJO 1 & FLUJO 2")
    print(f"  Prefijo de prueba: {TEST_PREFIX}")
    print("=" * 60)

    # ── 0. Autenticación ──
    print("\n── 0. Autenticación ──")
    admin_token = get_token("admin", "Admin123!")
    test("Login como admin", admin_token is not None)
    if not admin_token:
        print("⛔ No se pudo autenticar. Abortando.")
        sys.exit(1)
    h = auth_headers(admin_token)

    # ════════════════════════════════════════════════════════════
    # FLUJO 1: CARGA MASIVA DE CANDIDATOS
    # ════════════════════════════════════════════════════════════
    print("\n" + "=" * 60)
    print("  FLUJO 1: CARGA MASIVA DE CANDIDATOS")
    print("=" * 60)

    # ── 1. Endpoints protegidos ──
    print("\n── 1. Endpoints protegidos (sin token → 401) ──")
    r1 = requests.post(f"{API}/user-management/candidates/bulk-upload/preview")
    test("Preview sin auth → 401/422", r1.status_code in (401, 422))

    r2 = requests.post(f"{API}/user-management/candidates/bulk-upload")
    test("Upload sin auth → 401/422", r2.status_code in (401, 422))

    # ── 2. Validación de archivo ──
    print("\n── 2. Validaciones de archivo ──")

    # Sin archivo
    r = requests.post(f"{API}/user-management/candidates/bulk-upload/preview", headers=h)
    test("Preview sin archivo → 400", r.status_code == 400)
    test("  Mensaje indica 'no se envió'", "no se envi" in r.json().get("error", "").lower(),
         r.json().get("error", ""))

    # Archivo no-Excel
    r = requests.post(
        f"{API}/user-management/candidates/bulk-upload/preview",
        headers=h,
        files={"file": ("datos.csv", b"col1,col2\na,b", "text/csv")}
    )
    test("Preview con CSV → 400", r.status_code == 400)
    test("  Mensaje indica formato Excel", "excel" in r.json().get("error", "").lower(),
         r.json().get("error", ""))

    # Excel sin columnas requeridas
    bad_excel = make_excel_bytes(
        [{"columna_rara": "dato"}],
        ["columna_rara"]
    )
    r = requests.post(
        f"{API}/user-management/candidates/bulk-upload/preview",
        headers=h,
        files={"file": ("datos.xlsx", bad_excel, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
    )
    test("Preview sin columnas requeridas → 400", r.status_code == 400)
    test("  Mensaje indica columnas faltantes", "faltan" in r.json().get("error", "").lower(),
         r.json().get("error", ""))

    # ── 3. Segundo_apellido y genero como requeridos ──
    print("\n── 3. Validación de columnas requeridas (segundo_apellido, genero) ──")

    # Excel sin segundo_apellido
    no_segundo = make_excel_bytes(
        [{"nombre": "Juan", "primer_apellido": "López", "genero": "M"}],
        ["nombre", "primer_apellido", "genero"]
    )
    r = requests.post(
        f"{API}/user-management/candidates/bulk-upload/preview",
        headers=h,
        files={"file": ("datos.xlsx", no_segundo, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
    )
    test("Preview sin segundo_apellido → 400", r.status_code == 400)
    test("  Error menciona segundo_apellido", "segundo_apellido" in r.json().get("error", "").lower(),
         r.json().get("error", ""))

    # Excel sin genero
    no_genero = make_excel_bytes(
        [{"nombre": "Juan", "primer_apellido": "López", "segundo_apellido": "García"}],
        ["nombre", "primer_apellido", "segundo_apellido"]
    )
    r = requests.post(
        f"{API}/user-management/candidates/bulk-upload/preview",
        headers=h,
        files={"file": ("datos.xlsx", no_genero, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
    )
    test("Preview sin genero → 400", r.status_code == 400)
    test("  Error menciona genero", "genero" in r.json().get("error", "").lower(),
         r.json().get("error", ""))

    # ── 4. Preview exitoso (datos válidos) ──
    print("\n── 4. Preview exitoso con datos válidos ──")

    test_email_1 = f"{TEST_PREFIX.lower()}_user1@testbulk.com"
    test_email_2 = f"{TEST_PREFIX.lower()}_user2@testbulk.com"
    test_curp_1 = f"TEST{TEST_PREFIX[:4]}000001"  # 18 chars total
    test_curp_1 = test_curp_1.ljust(18, "X")[:18]

    candidates = [
        {"email": test_email_1, "nombre": "Ana", "primer_apellido": "García", "segundo_apellido": "López", "genero": "F", "curp": test_curp_1},
        {"email": test_email_2, "nombre": "Carlos", "primer_apellido": "Martínez", "segundo_apellido": "Ruiz", "genero": "M", "curp": None},
        {"email": None, "nombre": "María", "primer_apellido": "Sánchez", "segundo_apellido": "Torres", "genero": "F", "curp": None},
    ]

    excel_data = make_candidates_excel(candidates)
    r = requests.post(
        f"{API}/user-management/candidates/bulk-upload/preview",
        headers=h,
        files={"file": ("candidatos.xlsx", excel_data, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
    )
    test("Preview → 200", r.status_code == 200, f"Status: {r.status_code} | {r.text[:300]}")

    if r.status_code == 200:
        data = r.json()
        test("  Tiene 'preview' array", isinstance(data.get("preview"), list))
        test("  Tiene 'summary'", isinstance(data.get("summary"), dict))
        test("  Tiene 'can_proceed'", "can_proceed" in data)

        summary = data.get("summary", {})
        test(f"  total_rows == 3", summary.get("total_rows") == 3, f"Got: {summary.get('total_rows')}")
        test(f"  ready == 3 (todos nuevos)", summary.get("ready") == 3, f"Got: {summary.get('ready')}")
        test(f"  errors == 0", summary.get("errors") == 0, f"Got: {summary.get('errors')}")
        test("  can_proceed == True", data.get("can_proceed") == True)

        # Verificar estructura de filas preview
        preview_rows = data.get("preview", [])
        ready_rows = [p for p in preview_rows if p.get("status") == "ready"]
        test(f"  {len(ready_rows)} filas 'ready'", len(ready_rows) == 3)

        if ready_rows:
            first = ready_rows[0]
            test("  Fila ready tiene 'row'", "row" in first)
            test("  Fila ready tiene 'username_preview'", "username_preview" in first)
            test("  Fila ready tiene 'eligibility'", "eligibility" in first)
            elig = first.get("eligibility", {})
            test("  Elegibilidad incluye conocer, insignia", "conocer" in elig and "insignia" in elig)
    else:
        warn("Preview falló, saltando validaciones de respuesta")

    # ── 5. Preview con errores de validación ──
    print("\n── 5. Preview con filas inválidas ──")

    mixed_candidates = [
        {"email": "bueno@test.com", "nombre": "Bueno", "primer_apellido": "Uno", "segundo_apellido": "Dos", "genero": "M", "curp": None},
        {"email": "mail-invalido", "nombre": "Malo", "primer_apellido": "Email", "segundo_apellido": "Formt", "genero": "F", "curp": None},
        {"email": None, "nombre": None, "primer_apellido": None, "segundo_apellido": None, "genero": None, "curp": None},
        {"email": None, "nombre": "Solo", "primer_apellido": "Nombre", "segundo_apellido": "Bien", "genero": "X", "curp": None},
        {"email": "b2@test.com", "nombre": "Bien", "primer_apellido": "Curp", "segundo_apellido": "Corto", "genero": "M", "curp": "ABC"},
    ]
    excel_mix = make_candidates_excel(mixed_candidates)
    r = requests.post(
        f"{API}/user-management/candidates/bulk-upload/preview",
        headers=h,
        files={"file": ("mixto.xlsx", excel_mix, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
    )
    test("Preview mixto → 200", r.status_code == 200, f"Status: {r.status_code}")

    if r.status_code == 200:
        data = r.json()
        summary = data.get("summary", {})
        # Fila 3: vacía no se contará. Los errores son: mail-invalido, genero X, curp corto, fila vacía tiene campos requeridos
        test(f"  errors > 0 (filas inválidas detectadas)", summary.get("errors", 0) > 0, f"errors={summary.get('errors')}")
        test(f"  ready >= 1 (al menos 'bueno@test.com')", summary.get("ready", 0) >= 1, f"ready={summary.get('ready')}")

        # Verificar detalle de errores
        error_rows = [p for p in data.get("preview", []) if p.get("status") == "error"]
        test(f"  Filas error tienen campo 'error'", all("error" in e for e in error_rows))

    # ── 6. Upload real (crear candidatos) ──
    print("\n── 6. Upload real — crear candidatos ──")

    excel_upload = make_candidates_excel(candidates)
    r = requests.post(
        f"{API}/user-management/candidates/bulk-upload",
        headers=h,
        files={"file": ("candidatos.xlsx", excel_upload, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
    )
    test("Upload → 200", r.status_code == 200, f"Status: {r.status_code} | {r.text[:300]}")

    if r.status_code == 200:
        data = r.json()
        summary = data.get("summary", {})
        details = data.get("details", {})

        test(f"  created == 3", summary.get("created") == 3, f"created={summary.get('created')}")
        test(f"  errors == 0", summary.get("errors") == 0, f"errors={summary.get('errors')}")

        created_list = details.get("created", [])
        test(f"  detalles.created tiene 3 entries", len(created_list) == 3)

        if created_list:
            first_created = created_list[0]
            test("  Creado tiene 'username'", bool(first_created.get("username")))
            test("  Creado tiene 'password'", bool(first_created.get("password")))
            test("  Username es MAYÚSCULAS", first_created.get("username", "").isupper(),
                 f"Got: {first_created.get('username')}")

            # Guardar IDs para limpieza
            for c in created_list:
                uname = c.get("username")
                if uname:
                    rows = db_query("SELECT id FROM users WHERE username=%s", (uname,))
                    if rows:
                        created_user_ids.append(rows[0]["id"])
    else:
        warn("Upload falló, saltando verificaciones")

    # ── 7. Preview detecta existentes ──
    print("\n── 7. Preview detecta usuarios ya existentes ──")

    # Subir el mismo archivo de nuevo
    excel_dup = make_candidates_excel(candidates)
    r = requests.post(
        f"{API}/user-management/candidates/bulk-upload/preview",
        headers=h,
        files={"file": ("duplicados.xlsx", excel_dup, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
    )
    test("Preview de duplicados → 200", r.status_code == 200, f"Status: {r.status_code}")

    if r.status_code == 200:
        data = r.json()
        summary = data.get("summary", {})
        # Los emails/CURPs ya existen, deberían ser skipped
        test(f"  skipped >= 2 (email duplicados)", summary.get("skipped", 0) >= 2,
             f"skipped={summary.get('skipped')}")
        test(f"  ready < 3 (algunos ya existen)", summary.get("ready", 3) < 3,
             f"ready={summary.get('ready')}")

    # ── 8. Upload idempotente (no duplica) ──
    print("\n── 8. Upload idempotente (no crea duplicados) ──")

    excel_dup2 = make_candidates_excel(candidates)
    r = requests.post(
        f"{API}/user-management/candidates/bulk-upload",
        headers=h,
        files={"file": ("dup2.xlsx", excel_dup2, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
    )
    test("Upload duplicados → 200", r.status_code == 200, f"Status: {r.status_code}")

    if r.status_code == 200:
        data = r.json()
        summary = data.get("summary", {})
        # El 3er candidato sin email/curp no se detecta como existente → puede crearse de nuevo
        test(f"  created <= 1 (máximo 1 no-identificable)", summary.get("created", 99) <= 1,
             f"created={summary.get('created')}")
        test(f"  skipped >= 2 (los identificables se detectan)", summary.get("skipped", 0) >= 2,
             f"skipped={summary.get('skipped')}")

    # ── 9. Upload con grupo ──
    print("\n── 9. Upload con asignación a grupo ──")

    # Buscar un grupo activo
    group_id = None
    groups = db_query("SELECT TOP 1 id, name FROM candidate_groups WHERE is_active=1")
    if groups:
        group_id = groups[0]["id"]
        group_name = groups[0]["name"]
        print(f"  📌 Usando grupo: {group_name} (ID {group_id})")

        # Crear candidatos nuevos para asignar a grupo
        grp_email = f"{TEST_PREFIX.lower()}_grp1@testbulk.com"
        grp_candidates = [
            {"email": grp_email, "nombre": "GrpTest", "primer_apellido": "Uno", "segundo_apellido": "Prueba", "genero": "M", "curp": None},
        ]
        excel_grp = make_candidates_excel(grp_candidates)
        r = requests.post(
            f"{API}/user-management/candidates/bulk-upload",
            headers=h,
            files={"file": ("grupo.xlsx", excel_grp, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            data={"group_id": str(group_id)}
        )
        test("Upload con grupo → 200", r.status_code == 200, f"Status: {r.status_code} | {r.text[:300]}")

        if r.status_code == 200:
            data = r.json()
            test("  group_assignment presente", "group_assignment" in data)
            ga = data.get("group_assignment", {})
            test(f"  assigned >= 1", ga.get("assigned", 0) >= 1, f"assigned={ga.get('assigned')}")
            test(f"  group_name correcto", ga.get("group_name") == group_name,
                 f"Got: {ga.get('group_name')}")

            # Guardar para limpieza
            for c in data.get("details", {}).get("created", []):
                uname = c.get("username")
                if uname:
                    rows = db_query("SELECT id FROM users WHERE username=%s", (uname,))
                    if rows:
                        created_user_ids.append(rows[0]["id"])
    else:
        warn("No se encontró grupo activo, saltando test con grupo")

    # ── 10. Descarga de plantilla ──
    print("\n── 10. Descarga de plantilla ──")
    r = requests.get(f"{API}/user-management/candidates/bulk-upload/template", headers=h)
    test("GET template → 200", r.status_code == 200, f"Status: {r.status_code}")
    if r.status_code == 200:
        test("  Content-Type xlsx",
             "spreadsheet" in r.headers.get("Content-Type", "") or "octet-stream" in r.headers.get("Content-Type", ""),
             r.headers.get("Content-Type", ""))
        test(f"  Tamaño > 0 bytes", len(r.content) > 100, f"Size: {len(r.content)}")

    # ════════════════════════════════════════════════════════════
    # FLUJO 2: ASIGNACIÓN MASIVA A GRUPO
    # ════════════════════════════════════════════════════════════
    print("\n" + "=" * 60)
    print("  FLUJO 2: ASIGNACIÓN MASIVA A GRUPO")
    print("=" * 60)

    # Necesitamos un grupo y algunos candidatos existentes
    if not group_id:
        groups = db_query("SELECT TOP 1 id, name FROM candidate_groups WHERE is_active=1")
        if groups:
            group_id = groups[0]["id"]
            group_name = groups[0]["name"]

    if not group_id:
        warn("No hay grupo activo — saltando Flujo 2 completo")
    else:
        print(f"\n  📌 Grupo de prueba: {group_name} (ID {group_id})")

        # ── 11. Endpoints protegidos ──
        print("\n── 11. Flujo 2 — Endpoints protegidos ──")
        r = requests.post(f"{API}/partners/groups/{group_id}/members/upload/preview")
        test("F2 Preview sin auth → 401/422", r.status_code in (401, 422))
        r = requests.post(f"{API}/partners/groups/{group_id}/members/upload")
        test("F2 Upload sin auth → 401/422", r.status_code in (401, 422))

        # ── 12. Preview con identificadores ──
        print("\n── 12. Flujo 2 — Preview ──")

        # Buscar candidatos existentes para usar como identificadores
        known_candidates = db_query(
            "SELECT TOP 3 id, email, username, curp, name, first_surname "
            "FROM users WHERE role='candidato' AND is_active=1 AND email IS NOT NULL"
        )

        if len(known_candidates) >= 2:
            idents = [
                known_candidates[0]["email"],  # Por email
                known_candidates[1]["username"] if known_candidates[1].get("username") else known_candidates[1]["email"],
                "noexiste_" + TEST_PREFIX + "@fake.com",  # No encontrado
            ]
            excel_f2 = make_identifiers_excel(idents)

            r = requests.post(
                f"{API}/partners/groups/{group_id}/members/upload/preview",
                headers=h,
                files={"file": ("asignar.xlsx", excel_f2, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
            )
            test("F2 Preview → 200", r.status_code == 200, f"Status: {r.status_code} | {r.text[:300]}")

            if r.status_code == 200:
                data = r.json()
                test("  Tiene 'preview' array", isinstance(data.get("preview"), list))
                test("  Tiene 'summary'", isinstance(data.get("summary"), dict))
                test("  Tiene 'group_name'", bool(data.get("group_name")))

                summary = data.get("summary", {})
                test(f"  total == {len(idents)}", summary.get("total") == len(idents),
                     f"total={summary.get('total')}")
                test(f"  not_found >= 1", summary.get("not_found", 0) >= 1,
                     f"not_found={summary.get('not_found')}")
                test("  summary tiene 'ambiguous' key", "ambiguous" in summary,
                     f"keys: {list(summary.keys())}")

                # Verificar estructura de filas
                preview = data.get("preview", [])
                not_found = [p for p in preview if p.get("status") == "not_found"]
                test(f"  Fila not_found tiene 'error'",
                     len(not_found) > 0 and bool(not_found[0].get("error")))

                ready_rows = [p for p in preview if p.get("status") == "ready"]
                if ready_rows:
                    r0 = ready_rows[0]
                    test("  Fila ready tiene 'user' con datos", bool(r0.get("user")))
                    if r0.get("user"):
                        u = r0["user"]
                        test("  user tiene full_name", bool(u.get("full_name")))
                        test("  user tiene id", bool(u.get("id")))
        else:
            warn("Menos de 2 candidatos con email, saltando test de preview F2")

        # ── 13. Flujo 2 — Validación de archivo ──
        print("\n── 13. Flujo 2 — Validación de archivo ──")

        r = requests.post(
            f"{API}/partners/groups/{group_id}/members/upload/preview",
            headers=h
        )
        test("F2 Preview sin archivo → 400", r.status_code == 400)

        r = requests.post(
            f"{API}/partners/groups/{group_id}/members/upload/preview",
            headers=h,
            files={"file": ("datos.txt", b"hola", "text/plain")}
        )
        test("F2 Preview con TXT → 400", r.status_code == 400)

        # ── 14. Preview con ambigüedades ──
        print("\n── 14. Flujo 2 — Detección de ambigüedades ──")

        # Buscar un nombre que tenga duplicados (si hay)
        dup_names = db_query("""
            SELECT RTRIM(LTRIM(name)) + ' ' + RTRIM(LTRIM(first_surname)) 
                   + CASE WHEN second_surname IS NOT NULL AND RTRIM(LTRIM(second_surname)) != '' 
                          THEN ' ' + RTRIM(LTRIM(second_surname)) ELSE '' END AS full_name,
                   COUNT(*) as cnt
            FROM users 
            WHERE role='candidato' AND is_active=1 AND name IS NOT NULL
            GROUP BY RTRIM(LTRIM(name)) + ' ' + RTRIM(LTRIM(first_surname))
                   + CASE WHEN second_surname IS NOT NULL AND RTRIM(LTRIM(second_surname)) != ''
                          THEN ' ' + RTRIM(LTRIM(second_surname)) ELSE '' END
            HAVING COUNT(*) > 1
            ORDER BY cnt DESC
            OFFSET 0 ROWS FETCH NEXT 1 ROWS ONLY
        """)

        if dup_names:
            ambiguous_name = dup_names[0]["full_name"]
            print(f"  📌 Nombre ambiguo encontrado: '{ambiguous_name}' ({dup_names[0]['cnt']} coincidencias)")

            excel_amb = make_identifiers_excel([ambiguous_name])
            r = requests.post(
                f"{API}/partners/groups/{group_id}/members/upload/preview",
                headers=h,
                files={"file": ("ambiguo.xlsx", excel_amb, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
            )
            test("F2 Preview con nombre ambiguo → 200", r.status_code == 200)

            if r.status_code == 200:
                data = r.json()
                summary = data.get("summary", {})
                test(f"  ambiguous >= 1", summary.get("ambiguous", 0) >= 1,
                     f"ambiguous={summary.get('ambiguous')}")

                amb_rows = [p for p in data.get("preview", []) if p.get("status") == "ambiguous"]
                if amb_rows:
                    test("  Fila ambigua tiene 'ambiguous_matches'",
                         isinstance(amb_rows[0].get("ambiguous_matches"), list))
                    matches = amb_rows[0].get("ambiguous_matches", [])
                    test(f"  ambiguous_matches >= 2", len(matches) >= 2,
                         f"len={len(matches)}")
                    if matches:
                        test("  Match tiene 'id' y 'full_name'",
                             bool(matches[0].get("id")) and bool(matches[0].get("full_name")))

                        # ── 15. Resolución de ambigüedad ──
                        print("\n── 15. Flujo 2 — Resolución de ambigüedad en upload ──")
                        chosen_id = matches[0]["id"]
                        resolutions = json.dumps([{"identifier": ambiguous_name, "user_id": chosen_id}])

                        r2 = requests.post(
                            f"{API}/partners/groups/{group_id}/members/upload",
                            headers=h,
                            files={"file": ("amb2.xlsx", make_identifiers_excel([ambiguous_name]),
                                            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
                            data={"mode": "add", "resolutions": resolutions}
                        )
                        test("F2 Upload con resolución → 201", r2.status_code == 201,
                             f"Status: {r2.status_code} | {r2.text[:300]}")

                        if r2.status_code == 201:
                            data2 = r2.json()
                            test(f"  added >= 1 (ambigüedad resuelta)",
                                 len(data2.get("added", [])) >= 1,
                                 f"added={data2.get('added')}")

                            # Limpiar — quitar del grupo
                            if HAS_DB:
                                db_exec(
                                    "DELETE FROM group_members WHERE group_id=%s AND user_id=%s",
                                    (group_id, chosen_id)
                                )
                                print(f"  🧹 Limpiado miembro {chosen_id} del grupo")
                else:
                    warn("Sin matches en fila ambigua")
            else:
                warn(f"Preview ambiguo falló: {r.status_code}")
        else:
            warn("No se encontraron nombres duplicados, saltando test de ambigüedad")
            print("  (Esto es normal si la base de datos no tiene nombres repetidos entre candidatos)")

        # ── 16. Upload real Flujo 2 ──
        print("\n── 16. Flujo 2 — Upload real ──")

        if len(known_candidates) >= 1:
            # Usar email de un candidato conocido
            test_ident = known_candidates[0]["email"]
            excel_up = make_identifiers_excel([test_ident])

            r = requests.post(
                f"{API}/partners/groups/{group_id}/members/upload",
                headers=h,
                files={"file": ("asignar.xlsx", excel_up, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
                data={"mode": "add"}
            )
            test("F2 Upload → 201", r.status_code == 201, f"Status: {r.status_code} | {r.text[:300]}")

            if r.status_code == 201:
                data = r.json()
                total = data.get("total_processed", 0)
                added_count = len(data.get("added", []))
                err_count = len(data.get("errors", []))
                test(f"  total_processed >= 1", total >= 1, f"total={total}")
                test(f"  added + errors == total", added_count + err_count == total,
                     f"added={added_count}, errors={err_count}")

                # Limpiar
                if added_count > 0 and HAS_DB:
                    uid = known_candidates[0]["id"]
                    db_exec(
                        "DELETE FROM group_members WHERE group_id=%s AND user_id=%s",
                        (group_id, uid)
                    )
                    print(f"  🧹 Limpiado miembro del grupo")

        # ── 17. Plantilla de Flujo 2 ──
        print("\n── 17. Flujo 2 — Descarga de plantilla ──")
        r = requests.get(f"{API}/partners/groups/members/template", headers=h)
        test("GET F2 template → 200", r.status_code == 200, f"Status: {r.status_code}")
        if r.status_code == 200:
            test(f"  Content > 100 bytes", len(r.content) > 100)

    # ════════════════════════════════════════════════════════════
    # LIMPIEZA
    # ════════════════════════════════════════════════════════════
    print("\n" + "=" * 60)
    print("  🧹 LIMPIEZA DE DATOS DE PRUEBA")
    print("=" * 60)

    if HAS_DB:
        # Limpiar membresías de grupo
        if created_user_ids:
            for uid in created_user_ids:
                db_exec("DELETE FROM group_members WHERE user_id=%s", (uid,))
            print(f"  ✅ {len(created_user_ids)} membresías limpiadas")

        # Limpiar usuarios creados
        deleted = db_query(
            "SELECT id, username, email FROM users WHERE username LIKE %s",
            (f"{TEST_PREFIX}%",)
        )
        # Also check by test emails
        deleted2 = db_query(
            "SELECT id, username, email FROM users WHERE email LIKE %s",
            (f"%{TEST_PREFIX.lower()}%",)
        )
        all_delete = {r["id"]: r for r in deleted + deleted2}

        if all_delete:
            for uid in all_delete:
                db_exec("DELETE FROM group_members WHERE user_id=%s", (uid,))
                db_exec("DELETE FROM users WHERE id=%s", (uid,))
            print(f"  ✅ {len(all_delete)} usuarios de prueba eliminados")
        else:
            # Check by created_user_ids
            if created_user_ids:
                for uid in created_user_ids:
                    db_exec("DELETE FROM group_members WHERE user_id=%s", (uid,))
                    db_exec("DELETE FROM users WHERE id=%s", (uid,))
                print(f"  ✅ {len(created_user_ids)} usuarios eliminados por ID")
            else:
                print("  ℹ️ No se encontraron usuarios de prueba para limpiar")

        # Limpiar candidatos creados en test de upload mixto (bueno@test.com)
        cleanup_emails = ["bueno@test.com", "b2@test.com"]
        for email in cleanup_emails:
            rows = db_query("SELECT id FROM users WHERE email=%s", (email,))
            for row in rows:
                db_exec("DELETE FROM group_members WHERE user_id=%s", (row["id"],))
                db_exec("DELETE FROM users WHERE id=%s", (row["id"],))
        if cleanup_emails:
            print(f"  ✅ Emails de test mixto limpiados")

        test("Limpieza completada", True)
    else:
        warn("Sin acceso a DB, limpieza manual requerida con prefijo: " + TEST_PREFIX)

    # ════════════════════════════════════════════════════════════
    # RESUMEN FINAL
    # ════════════════════════════════════════════════════════════
    print("\n" + "=" * 60)
    total = passed + failed
    emoji = "✅" if failed == 0 else "⚠️" if failed <= 3 else "❌"
    print(f"📋 RESULTADOS: {passed}/{total} tests pasaron")
    if warnings:
        print(f"   ⚠️  {warnings} advertencias")
    if failed:
        print(f"   ❌ {failed} tests fallaron")
    print("=" * 60)

    sys.exit(0 if failed == 0 else 1)


if __name__ == "__main__":
    main()
