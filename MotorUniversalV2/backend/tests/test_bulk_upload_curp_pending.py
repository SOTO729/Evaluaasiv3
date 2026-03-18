"""
Tests para dos fixes de carga masiva:

1. Name match con info de grupo/plantel:
   - Preview con coincidencia de nombre muestra grupo y campus del usuario existente.

2. CURP pending state con grupo:
   - Usuarios con CURP + grupo se crean con is_active=False.
   - GroupMember se crea con status='curp_pending'.
   - El endpoint de miembros de grupo NO muestra miembros curp_pending.
   - Si CURP es inválida → usuario y membresía se eliminan.
   - Si CURP es válida → usuario se activa y membresía pasa a 'active'.

USO: python tests/test_bulk_upload_curp_pending.py
"""
import sys
import os
import io
import json
import time
import uuid
import requests

# ── Configuración ──
API = os.environ.get(
    "TEST_API",
    "https://evaluaasi-motorv2-api-dev.purpleocean-384694c4.southcentralus.azurecontainerapps.io/api"
)

try:
    import pymssql
    HAS_DB = True
except ImportError:
    print("⚠️  pymssql no disponible, tests que requieren DB se omitirán")
    HAS_DB = False

DB_SERVER = "evaluaasi-motorv2-sql.database.windows.net"
DB_USER = "evaluaasi_admin"
DB_PASS = "EvalAasi2024_newpwd!"
DB_NAME = os.environ.get("TEST_DB", "evaluaasi_dev")

passed = 0
failed = 0
warnings = 0
TEST_PREFIX = f"CPTEST_{uuid.uuid4().hex[:6].upper()}"
created_user_ids = []

REQ_TIMEOUT = 120


def test(name, condition, detail=""):
    global passed, failed
    if condition:
        print(f"  ✅ {name}")
        passed += 1
    else:
        print(f"  ❌ {name} — {detail}")
        failed += 1


def warn(name, detail=""):
    global warnings
    print(f"  ⚠️  {name} — {detail}")
    warnings += 1


def post_with_retry(url, max_retries=2, **kwargs):
    kwargs.setdefault('timeout', REQ_TIMEOUT)
    for attempt in range(max_retries + 1):
        try:
            r = requests.post(url, **kwargs)
            if r.status_code == 504 and attempt < max_retries:
                print(f"  ⏳ 504 timeout, reintentando ({attempt + 1}/{max_retries})...")
                time.sleep(5)
                continue
            return r
        except requests.exceptions.Timeout:
            if attempt < max_retries:
                print(f"  ⏳ Request timeout, reintentando ({attempt + 1}/{max_retries})...")
                time.sleep(5)
                continue
            raise
    return r


def get_with_retry(url, max_retries=1, **kwargs):
    kwargs.setdefault('timeout', REQ_TIMEOUT)
    for attempt in range(max_retries + 1):
        try:
            r = requests.get(url, **kwargs)
            if r.status_code == 504 and attempt < max_retries:
                time.sleep(5)
                continue
            return r
        except requests.exceptions.Timeout:
            if attempt < max_retries:
                time.sleep(5)
                continue
            raise
    return r


def db_query(sql, params=None):
    if not HAS_DB:
        return []
    conn = pymssql.connect(server=DB_SERVER, user=DB_USER, password=DB_PASS, database=DB_NAME)
    cursor = conn.cursor(as_dict=True)
    cursor.execute(sql, params or ())
    try:
        rows = cursor.fetchall()
    except Exception:
        rows = []
    conn.commit()
    conn.close()
    return rows


def db_exec(sql, params=None):
    if not HAS_DB:
        return
    conn = pymssql.connect(server=DB_SERVER, user=DB_USER, password=DB_PASS, database=DB_NAME)
    cursor = conn.cursor()
    cursor.execute(sql, params or ())
    conn.commit()
    conn.close()


def get_token(username="admin", password="admin123"):
    r = post_with_retry(f"{API}/auth/login", json={"username": username, "password": password})
    if r.status_code == 429:
        wait = int(r.json().get("retry_after", 30))
        print(f"  ⏳ Rate limit, esperando {min(wait, 60)}s...")
        time.sleep(min(wait, 60))
        r = post_with_retry(f"{API}/auth/login", json={"username": username, "password": password})
    if r.status_code == 200:
        return r.json().get("access_token")
    print(f"  ⚠️  Login fallido ({username}): {r.status_code} {r.text[:200]}")
    return None


def auth_headers(token):
    return {"Authorization": f"Bearer {token}"}


def make_candidates_excel(candidates):
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    headers = ["email", "nombre", "primer_apellido", "segundo_apellido", "genero", "curp"]
    for col_idx, h in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx, value=h)
    ws.cell(row=2, column=1, value="(descripción)")
    for row_idx, c in enumerate(candidates, 3):
        for col_idx, h in enumerate(headers, 1):
            ws.cell(row=row_idx, column=col_idx, value=c.get(h))
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ============================================================
# MAIN
# ============================================================
def main():
    global passed, failed, created_user_ids

    print("\n" + "=" * 60)
    print("  TESTS: NAME MATCH GROUPS + CURP PENDING STATE")
    print(f"  Prefijo: {TEST_PREFIX}")
    print(f"  API: {API}")
    print(f"  DB: {DB_NAME}")
    print("=" * 60)

    # ── 0. Auth ──
    print("\n── 0. Autenticación ──")
    token = get_token()
    test("Login como admin", token is not None)
    if not token:
        print("⛔ No se pudo autenticar. Abortando.")
        sys.exit(1)
    h = auth_headers(token)

    # ── Buscar grupo activo con campus ──
    group_id = None
    group_name = None
    campus_name = None
    if HAS_DB:
        rows = db_query(
            "SELECT TOP 1 cg.id as gid, cg.name as gname, c.name as cname "
            "FROM candidate_groups cg "
            "JOIN campuses c ON cg.campus_id = c.id "
            "WHERE cg.is_active = 1"
        )
        if rows:
            group_id = rows[0]["gid"]
            group_name = rows[0]["gname"]
            campus_name = rows[0]["cname"]
            print(f"  📌 Grupo: {group_name} (ID {group_id}), Campus: {campus_name}")
    if not group_id:
        warn("No se encontró grupo activo con campus, algunos tests se omitirán")

    # ════════════════════════════════════════════════════════════
    # TEST 1: NAME MATCH PREVIEW CON INFO DE GRUPO/PLANTEL
    # ════════════════════════════════════════════════════════════
    print("\n" + "=" * 60)
    print("  TEST 1: NAME MATCH PREVIEW CON INFO DE GRUPO/PLANTEL")
    print("=" * 60)

    seed_user_id = None
    seed_username = None
    if HAS_DB and group_id:
        # Crear un usuario seed en la DB y asignarlo al grupo
        seed_user_id = str(uuid.uuid4())
        seed_username = f"{TEST_PREFIX}_SEED"
        seed_name = TEST_PREFIX
        seed_first = "NAMEMATCH"
        seed_gender = "M"

        db_exec(
            "INSERT INTO users (id, username, name, first_surname, second_surname, gender, role, is_active, is_verified, password_hash, created_at, updated_at) "
            "VALUES (%s, %s, %s, %s, %s, %s, %s, 1, 0, 'dummy', GETUTCDATE(), GETUTCDATE())",
            (seed_user_id, seed_username, seed_name, seed_first, "TEST", seed_gender, "candidato")
        )
        db_exec(
            "INSERT INTO group_members (group_id, user_id, status) VALUES (%s, %s, %s)",
            (group_id, seed_user_id, "active")
        )
        created_user_ids.append(seed_user_id)
        print(f"  📌 Seed user creado: {seed_username} (ID {seed_user_id})")

        # Preview con un candidato que tiene mismo nombre/primer_apellido/genero
        candidates = [{
            "email": f"{TEST_PREFIX.lower()}_nm@test.com",
            "nombre": seed_name,
            "primer_apellido": seed_first,
            "segundo_apellido": "OTRO",
            "genero": seed_gender,
            "curp": None,
        }]
        excel = make_candidates_excel(candidates)
        r = post_with_retry(
            f"{API}/user-management/candidates/bulk-upload/preview",
            headers=h,
            files={"file": ("test.xlsx", excel, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        test("Preview → 200", r.status_code == 200, f"Status: {r.status_code} | {r.text[:300]}")

        if r.status_code == 200:
            data = r.json()
            preview = data.get("preview", [])
            test("Hay al menos 1 fila en preview", len(preview) >= 1, f"preview len: {len(preview)}")

            # Buscar la fila con name_matches
            nm_row = None
            for p in preview:
                if p.get("name_matches"):
                    nm_row = p
                    break

            test("Fila con name_matches encontrada", nm_row is not None,
                 f"Statuses: {[p.get('status') for p in preview]}")

            if nm_row:
                matches = nm_row["name_matches"]
                test("name_matches tiene al menos 1 coincidencia", len(matches) >= 1)

                # Buscar nuestro seed user en las coincidencias
                seed_match = None
                for m in matches:
                    if m.get("username") == seed_username:
                        seed_match = m
                        break

                test("Seed user aparece en coincidencias", seed_match is not None,
                     f"Matches: {[m.get('username') for m in matches]}")

                if seed_match:
                    test("Match tiene campo 'groups'", "groups" in seed_match,
                         f"Keys: {list(seed_match.keys())}")

                    groups = seed_match.get("groups", [])
                    test("groups tiene al menos 1 entrada", len(groups) >= 1,
                         f"groups: {groups}")

                    if groups:
                        g = groups[0]
                        test("groups[0] tiene 'group_name'", "group_name" in g, f"Keys: {list(g.keys())}")
                        test("groups[0] tiene 'campus_name'", "campus_name" in g, f"Keys: {list(g.keys())}")
                        test(f"group_name = '{group_name}'", g.get("group_name") == group_name,
                             f"Got: {g.get('group_name')}")
                        test(f"campus_name = '{campus_name}'", g.get("campus_name") == campus_name,
                             f"Got: {g.get('campus_name')}")
    else:
        warn("Omitiendo test name match (requiere DB + grupo)")

    # ════════════════════════════════════════════════════════════
    # TEST 2: CURP PENDING STATE
    # ════════════════════════════════════════════════════════════
    print("\n" + "=" * 60)
    print("  TEST 2: CURP PENDING STATE CON GRUPO")
    print("=" * 60)

    if group_id:
        # Usar una CURP con formato válido pero que RENAPO rechazará (persona inexistente)
        invalid_curp = "ZZZZ990101HDFZZZQ9"
        curp_candidate = [{
            "email": f"{TEST_PREFIX.lower()}_curp@test.com",
            "nombre": f"{TEST_PREFIX}CURP",
            "primer_apellido": "PENDING",
            "segundo_apellido": "TEST",
            "genero": "M",
            "curp": invalid_curp,
        }]
        excel = make_candidates_excel(curp_candidate)

        print("\n── 2a. Upload con CURP + grupo ──")
        r = post_with_retry(
            f"{API}/user-management/candidates/bulk-upload",
            headers=h,
            files={"file": ("curp_test.xlsx", excel, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            data={"group_id": str(group_id)},
        )
        test("Upload con CURP + grupo → 200", r.status_code == 200, f"Status: {r.status_code} | {r.text[:300]}")

        uploaded_username = None
        uploaded_user_id = None
        if r.status_code == 200:
            data = r.json()
            created_list = data.get("details", {}).get("created", [])
            errors_list = data.get("details", {}).get("errors", [])
            skipped_list = data.get("details", {}).get("skipped", [])
            if errors_list:
                print(f"  ⚠️  Errores: {errors_list[:3]}")
            if skipped_list:
                print(f"  ⚠️  Skipped: {skipped_list[:3]}")
            test("1 usuario creado", len(created_list) == 1, f"created: {len(created_list)}, errors: {len(errors_list)}, skipped: {len(skipped_list)}")

            if created_list:
                uploaded_username = created_list[0].get("username")
                print(f"  📌 Username creado: {uploaded_username}")

        # Verificar estado en BD inmediatamente (antes de que el thread background termine)
        if HAS_DB and uploaded_username:
            print("\n── 2b. Verificar estado inicial en BD ──")
            # Pequeña pausa para que el commit se propague
            time.sleep(1)

            user_rows = db_query(
                "SELECT id, is_active, curp_verified FROM users WHERE username = %s",
                (uploaded_username,)
            )
            test("Usuario encontrado en BD", len(user_rows) == 1, f"Rows: {len(user_rows)}")

            if user_rows:
                uploaded_user_id = user_rows[0]["id"]
                created_user_ids.append(uploaded_user_id)

                is_active = user_rows[0]["is_active"]
                test("is_active = False (pendiente CURP)", is_active == False,
                     f"is_active = {is_active}")

                # Verificar GroupMember con status='curp_pending'
                gm_rows = db_query(
                    "SELECT status FROM group_members WHERE group_id = %s AND user_id = %s",
                    (group_id, uploaded_user_id)
                )
                test("GroupMember existe", len(gm_rows) == 1, f"Rows: {len(gm_rows)}")
                if gm_rows:
                    test("GroupMember status = 'curp_pending'",
                         gm_rows[0]["status"] == "curp_pending",
                         f"status = '{gm_rows[0]['status']}'")

            print("\n── 2c. Verificar que no aparece en endpoint de miembros ──")
            r2 = get_with_retry(
                f"{API}/partners/groups/{group_id}/members?per_page=1000",
                headers=h,
            )
            test("GET group members → 200", r2.status_code == 200, f"Status: {r2.status_code}")

            if r2.status_code == 200:
                members = r2.json().get("members", [])
                member_ids = [m.get("user_id") or m.get("id") for m in members]
                test("Usuario curp_pending NO aparece en lista de miembros",
                     uploaded_user_id not in member_ids,
                     f"User ID {uploaded_user_id} found in {len(members)} members")

            # Esperar a que el background thread procese (CURP inválida → debe eliminar)
            print("\n── 2d. Esperar verificación CURP background ──")
            max_wait = 45  # RENAPO puede tardar (2s rate limit + scraping time)
            elapsed = 0
            user_deleted = False
            while elapsed < max_wait:
                time.sleep(2)
                elapsed += 2
                rows = db_query("SELECT id FROM users WHERE id = %s", (uploaded_user_id,))
                if not rows:
                    user_deleted = True
                    print(f"  ✅ Usuario eliminado después de {elapsed}s")
                    break
                print(f"  ⏳ {elapsed}s — usuario aún existe")

            test("CURP inválida → usuario eliminado por background thread",
                 user_deleted,
                 f"Todavía existe después de {max_wait}s (RENAPO puede tardar más)")
            if not user_deleted:
                warn("Background thread no completó en tiempo esperado (RENAPO lento)", "El usuario será eliminado eventualmente")

            if user_deleted:
                # Verificar que GroupMember también se eliminó
                gm_rows = db_query(
                    "SELECT id FROM group_members WHERE group_id = %s AND user_id = %s",
                    (group_id, uploaded_user_id)
                )
                test("GroupMember también eliminado", len(gm_rows) == 0,
                     f"Aún hay {len(gm_rows)} membresía(s)")
                # Ya no necesitamos limpiar este usuario
                if uploaded_user_id in created_user_ids:
                    created_user_ids.remove(uploaded_user_id)

        elif not HAS_DB:
            warn("Omitiendo verificación BD (pymssql no disponible)")

        # ── 2e. Test de usuario SIN CURP con grupo → debe ser is_active=True, status=active ──
        print("\n── 2e. Upload SIN CURP + grupo → debe ser activo inmediato ──")
        no_curp_candidate = [{
            "email": f"{TEST_PREFIX.lower()}_nocurp@test.com",
            "nombre": f"{TEST_PREFIX}NOCURP",
            "primer_apellido": "ACTIVO",
            "segundo_apellido": "TEST",
            "genero": "M",
            "curp": None,
        }]
        excel2 = make_candidates_excel(no_curp_candidate)
        r3 = post_with_retry(
            f"{API}/user-management/candidates/bulk-upload",
            headers=h,
            files={"file": ("nocurp.xlsx", excel2, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            data={"group_id": str(group_id)},
        )
        test("Upload sin CURP + grupo → 200", r3.status_code == 200, f"Status: {r3.status_code}")

        nocurp_username = None
        nocurp_user_id = None
        if r3.status_code == 200:
            created3 = r3.json().get("details", {}).get("created", [])
            if created3:
                nocurp_username = created3[0].get("username")

        if HAS_DB and nocurp_username:
            time.sleep(1)
            rows = db_query(
                "SELECT id, is_active FROM users WHERE username = %s",
                (nocurp_username,)
            )
            if rows:
                nocurp_user_id = rows[0]["id"]
                created_user_ids.append(nocurp_user_id)
                test("Sin CURP → is_active = True", rows[0]["is_active"] == True,
                     f"is_active = {rows[0]['is_active']}")

                gm = db_query(
                    "SELECT status FROM group_members WHERE group_id = %s AND user_id = %s",
                    (group_id, nocurp_user_id)
                )
                if gm:
                    test("Sin CURP → GroupMember status = 'active'",
                         gm[0]["status"] == "active",
                         f"status = '{gm[0]['status']}'")

        # ── 2f. Test CURP genérica extranjera → debe ser activo inmediato ──
        print("\n── 2f. Upload con CURP genérica extranjera + grupo → activo inmediato ──")
        foreign_curp = "XEXX010101HNEXXXA4"  # CURP genérica para extranjeros hombres
        foreign_candidate = [{
            "email": f"{TEST_PREFIX.lower()}_foreign@test.com",
            "nombre": f"{TEST_PREFIX}FOREIGN",
            "primer_apellido": "EXTRANJERO",
            "segundo_apellido": "TEST",
            "genero": "H",
            "curp": foreign_curp,
        }]
        excel3 = make_candidates_excel(foreign_candidate)
        r4 = post_with_retry(
            f"{API}/user-management/candidates/bulk-upload",
            headers=h,
            files={"file": ("foreign.xlsx", excel3, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            data={"group_id": str(group_id)},
        )
        test("Upload CURP extranjera + grupo → 200", r4.status_code == 200,
             f"Status: {r4.status_code} | {r4.text[:300]}")

        foreign_username = None
        foreign_user_id = None
        if r4.status_code == 200:
            created4 = r4.json().get("details", {}).get("created", [])
            if created4:
                foreign_username = created4[0].get("username")

        if HAS_DB and foreign_username:
            time.sleep(1)
            rows = db_query(
                "SELECT id, is_active FROM users WHERE username = %s",
                (foreign_username,)
            )
            if rows:
                foreign_user_id = rows[0]["id"]
                created_user_ids.append(foreign_user_id)
                test("CURP extranjera → is_active = True (sin verificación pendiente)",
                     rows[0]["is_active"] == True,
                     f"is_active = {rows[0]['is_active']}")

                gm = db_query(
                    "SELECT status FROM group_members WHERE group_id = %s AND user_id = %s",
                    (group_id, foreign_user_id)
                )
                if gm:
                    test("CURP extranjera → GroupMember status = 'active'",
                         gm[0]["status"] == "active",
                         f"status = '{gm[0]['status']}'")
    else:
        warn("Omitiendo tests CURP pending (requiere grupo)")

    # ════════════════════════════════════════════════════════════
    # LIMPIEZA
    # ════════════════════════════════════════════════════════════
    print("\n" + "=" * 60)
    print("  LIMPIEZA")
    print("=" * 60)

    if HAS_DB:
        # Limpiar por prefijo
        test_users = db_query(
            "SELECT id, username FROM users WHERE username LIKE %s",
            (f"{TEST_PREFIX}%",)
        )
        all_ids = set(created_user_ids) | {r["id"] for r in test_users}

        if all_ids:
            for uid in all_ids:
                db_exec("DELETE FROM bulk_upload_members WHERE username IN (SELECT username FROM users WHERE id=%s)", (uid,))
                db_exec("DELETE FROM group_members WHERE user_id=%s", (uid,))
                db_exec("DELETE FROM users WHERE id=%s", (uid,))
            print(f"  ✅ {len(all_ids)} usuario(s) de prueba eliminados")
        else:
            print("  ℹ️  No hay usuarios de prueba para limpiar")

        # También limpiar por email pattern
        email_users = db_query(
            "SELECT id FROM users WHERE email LIKE %s",
            (f"%{TEST_PREFIX.lower()}%",)
        )
        for eu in email_users:
            if eu["id"] not in all_ids:
                db_exec("DELETE FROM group_members WHERE user_id=%s", (eu["id"],))
                db_exec("DELETE FROM users WHERE id=%s", (eu["id"],))
                print(f"  ✅ Limpiado usuario extra por email: {eu['id']}")

        test("Limpieza completada", True)
    else:
        warn("Sin DB, limpieza manual requerida con prefijo: " + TEST_PREFIX)

    # ════════════════════════════════════════════════════════════
    # RESUMEN
    # ════════════════════════════════════════════════════════════
    print("\n" + "=" * 60)
    total = passed + failed
    emoji = "✅" if failed == 0 else "⚠️" if failed <= 2 else "❌"
    print(f"{emoji} RESULTADOS: {passed}/{total} tests pasaron")
    if warnings:
        print(f"   ⚠️  {warnings} advertencias")
    if failed:
        print(f"   ❌ {failed} tests fallaron")
    print("=" * 60)

    sys.exit(0 if failed == 0 else 1)


if __name__ == "__main__":
    main()
