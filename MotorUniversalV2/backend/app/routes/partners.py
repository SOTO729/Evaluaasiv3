"""
Rutas para gestión de Partners, Planteles y Grupos
Solo accesibles por coordinadores y admins
"""
import secrets
import string
from datetime import datetime
from io import BytesIO
from flask import Blueprint, request, jsonify, g, send_file
from functools import wraps
from flask_jwt_extended import jwt_required, get_jwt_identity
from sqlalchemy.orm import joinedload
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from app import db
from app.models import (
    Partner, PartnerStatePresence, Campus, CandidateGroup, GroupMember,
    User, MEXICAN_STATES, SchoolCycle
)
from app.models.partner import GroupExam, GroupExamMember, GroupExamMaterial
from app.models.user import decrypt_password, encrypt_password
from app.models.balance import CoordinatorBalance, BalanceTransaction, create_balance_transaction
from app.models.competency_standard import CompetencyStandard
from app.models.result import Result
from app.models.student_progress import StudentTopicProgress

bp = Blueprint('partners', __name__)


def coordinator_required(f):
    """Decorador que requiere rol de coordinador, developer o admin"""
    @wraps(f)
    def decorated(*args, **kwargs):
        user_id = get_jwt_identity()
        user = User.query.get(user_id)
        if not user:
            return jsonify({'error': 'No autorizado'}), 401
        if user.role not in ['admin', 'developer', 'coordinator']:
            return jsonify({'error': 'Acceso denegado. Se requiere rol de coordinador'}), 403
        g.current_user = user
        return f(*args, **kwargs)
    return decorated


# ============== CURP EXTRANJEROS ==============
# Personas extranjeras reciben un CURP genérico basado en su género.
# Estos CURPs pueden repetirse en la app.
FOREIGN_CURP_MALE = 'XEXX010101HNEXXXA4'
FOREIGN_CURP_FEMALE = 'XEXX010101MNEXXXA8'
GENERIC_FOREIGN_CURPS = {FOREIGN_CURP_MALE, FOREIGN_CURP_FEMALE}


def _get_foreign_curp(gender: str) -> str:
    """Asigna CURP genérico de extranjero según género (M=masculino→H, F=femenino→M)."""
    if gender == 'M':  # Masculino
        return FOREIGN_CURP_MALE
    else:  # Femenino u Otro
        return FOREIGN_CURP_FEMALE


def _is_generic_foreign_curp(curp: str) -> bool:
    """Verifica si un CURP es el genérico de extranjero (puede repetirse)."""
    return curp and curp.upper().strip() in GENERIC_FOREIGN_CURPS


# ============== ESTADOS MEXICANOS ==============

@bp.route('/mexican-states', methods=['GET'])
@jwt_required()
def get_mexican_states():
    """Obtener lista de estados mexicanos"""
    return jsonify({'states': MEXICAN_STATES})


@bp.route('/countries', methods=['GET'])
@jwt_required()
def get_countries():
    """Obtener lista de países disponibles"""
    from app.models.partner import AVAILABLE_COUNTRIES
    return jsonify({'countries': AVAILABLE_COUNTRIES})


# ============== HELPERS MULTI-TENANT ==============

def _get_coordinator_filter(user):
    """
    Retorna el coordinator_id para filtrar datos, o None si el usuario es admin/developer (ve todo).
    """
    if user.role == 'coordinator':
        return user.id
    return None


def _verify_partner_access(partner_id, user):
    """Verifica que el coordinador tenga acceso al partner. 
    Retorna (partner, None) si ok, o (None, error_response) si no tiene acceso."""
    partner = Partner.query.get_or_404(partner_id)
    coord_id = _get_coordinator_filter(user)
    if coord_id and partner.coordinator_id != coord_id:
        return None, (jsonify({'error': 'No tienes acceso a este partner'}), 403)
    return partner, None


def _verify_campus_access(campus_id, user):
    """Verifica que el coordinador tenga acceso al campus (a través de su partner)."""
    from app.models.partner import Campus
    campus = Campus.query.get_or_404(campus_id)
    coord_id = _get_coordinator_filter(user)
    if coord_id:
        partner = Partner.query.get(campus.partner_id)
        if not partner or partner.coordinator_id != coord_id:
            return None, (jsonify({'error': 'No tienes acceso a este plantel'}), 403)
    return campus, None


def _verify_group_access(group_id, user):
    """Verifica que el coordinador tenga acceso al grupo (a través de campus → partner)."""
    from app.models.partner import Campus
    group = CandidateGroup.query.get_or_404(group_id)
    coord_id = _get_coordinator_filter(user)
    if coord_id:
        campus = Campus.query.get(group.campus_id)
        if campus:
            partner = Partner.query.get(campus.partner_id)
            if not partner or partner.coordinator_id != coord_id:
                return None, (jsonify({'error': 'No tienes acceso a este grupo'}), 403)
        else:
            return None, (jsonify({'error': 'No tienes acceso a este grupo'}), 403)
    return group, None


# ============== PARTNERS ==============

@bp.route('', methods=['GET'])
@jwt_required()
@coordinator_required
def get_partners():
    """Listar todos los partners"""
    try:
        page = request.args.get('page', 1, type=int)
        per_page = request.args.get('per_page', 20, type=int)
        search = request.args.get('search', '')
        active_only = request.args.get('active_only', 'true').lower() == 'true'
        
        query = Partner.query
        
        # Multi-tenant: coordinadores solo ven sus partners
        coord_id = _get_coordinator_filter(g.current_user)
        if coord_id:
            query = query.filter(Partner.coordinator_id == coord_id)
        
        if active_only:
            query = query.filter(Partner.is_active == True)
            
        if search:
            search_term = f'%{search}%'
            query = query.filter(
                db.or_(
                    Partner.name.ilike(search_term),
                    Partner.legal_name.ilike(search_term),
                    Partner.rfc.ilike(search_term)
                )
            )
        
        query = query.order_by(Partner.name)
        
        pagination = query.paginate(page=page, per_page=per_page, max_per_page=1000, error_out=False)
        
        return jsonify({
            'partners': [p.to_dict(include_states=True) for p in pagination.items],
            'total': pagination.total,
            'pages': pagination.pages,
            'current_page': page
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@bp.route('/<int:partner_id>', methods=['GET'])
@jwt_required()
@coordinator_required
def get_partner(partner_id):
    """Obtener detalle de un partner"""
    try:
        partner, error = _verify_partner_access(partner_id, g.current_user)
        if error:
            return error
        return jsonify({
            'partner': partner.to_dict(include_states=True, include_campuses=True)
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@bp.route('', methods=['POST'])
@jwt_required()
@coordinator_required
def create_partner():
    """Crear un nuevo partner"""
    try:
        data = request.get_json()
        
        if not data.get('name'):
            return jsonify({'error': 'El nombre es requerido'}), 400
        
        # Verificar RFC único si se proporciona
        if data.get('rfc'):
            existing = Partner.query.filter_by(rfc=data['rfc']).first()
            if existing:
                return jsonify({'error': 'Ya existe un partner con ese RFC'}), 400
        
        partner = Partner(
            name=data['name'],
            legal_name=data.get('legal_name'),
            rfc=data.get('rfc'),
            country=data.get('country', 'México'),
            coordinator_id=_get_coordinator_filter(g.current_user) or data.get('coordinator_id'),
            email=data.get('email'),
            phone=data.get('phone'),
            website=data.get('website'),
            logo_url=data.get('logo_url'),
            notes=data.get('notes'),
            is_active=data.get('is_active', True)
        )
        
        db.session.add(partner)
        db.session.commit()
        
        return jsonify({
            'message': 'Partner creado exitosamente',
            'partner': partner.to_dict()
        }), 201
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500


@bp.route('/<int:partner_id>', methods=['PUT'])
@jwt_required()
@coordinator_required
def update_partner(partner_id):
    """Actualizar un partner"""
    try:
        partner, error = _verify_partner_access(partner_id, g.current_user)
        if error:
            return error
        data = request.get_json()
        
        # Verificar RFC único si cambia
        if data.get('rfc') and data['rfc'] != partner.rfc:
            existing = Partner.query.filter_by(rfc=data['rfc']).first()
            if existing:
                return jsonify({'error': 'Ya existe un partner con ese RFC'}), 400
        
        # Actualizar campos
        for field in ['name', 'legal_name', 'rfc', 'country', 'email', 'phone', 'website', 'logo_url', 'notes', 'is_active']:
            if field in data:
                setattr(partner, field, data[field])
        
        db.session.commit()
        
        return jsonify({
            'message': 'Partner actualizado exitosamente',
            'partner': partner.to_dict(include_states=True)
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500


@bp.route('/<int:partner_id>', methods=['DELETE'])
@jwt_required()
@coordinator_required
def delete_partner(partner_id):
    """Eliminar un partner (soft delete)"""
    try:
        partner, error = _verify_partner_access(partner_id, g.current_user)
        if error:
            return error
        partner.is_active = False
        db.session.commit()
        
        return jsonify({'message': 'Partner desactivado exitosamente'})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500


# ============== PRESENCIA EN ESTADOS ==============

@bp.route('/<int:partner_id>/states', methods=['GET'])
@jwt_required()
@coordinator_required
def get_partner_states(partner_id):
    """Obtener estados donde tiene presencia un partner"""
    try:
        partner, error = _verify_partner_access(partner_id, g.current_user)
        if error:
            return error
        presences = PartnerStatePresence.query.filter_by(partner_id=partner_id).all()
        
        return jsonify({
            'partner_id': partner_id,
            'partner_name': partner.name,
            'states': [p.to_dict() for p in presences]
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@bp.route('/<int:partner_id>/states', methods=['POST'])
@jwt_required()
@coordinator_required
def add_partner_state(partner_id):
    """Agregar presencia en un estado"""
    try:
        partner, error = _verify_partner_access(partner_id, g.current_user)
        if error:
            return error
        data = request.get_json()
        
        state_name = data.get('state_name')
        if not state_name:
            return jsonify({'error': 'El nombre del estado es requerido'}), 400
            
        if state_name not in MEXICAN_STATES:
            return jsonify({'error': 'Estado no válido'}), 400
        
        # Verificar si ya existe
        existing = PartnerStatePresence.query.filter_by(
            partner_id=partner_id,
            state_name=state_name
        ).first()
        
        if existing:
            return jsonify({'error': 'El partner ya tiene presencia en ese estado'}), 400
        
        presence = PartnerStatePresence(
            partner_id=partner_id,
            state_name=state_name,
            regional_contact_name=data.get('regional_contact_name'),
            regional_contact_email=data.get('regional_contact_email'),
            regional_contact_phone=data.get('regional_contact_phone')
        )
        
        db.session.add(presence)
        db.session.commit()
        
        return jsonify({
            'message': f'Presencia en {state_name} agregada',
            'presence': presence.to_dict()
        }), 201
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500


@bp.route('/<int:partner_id>/states/<int:presence_id>', methods=['DELETE'])
@jwt_required()
@coordinator_required
def remove_partner_state(partner_id, presence_id):
    """Eliminar presencia en un estado"""
    try:
        partner, error = _verify_partner_access(partner_id, g.current_user)
        if error:
            return error
        presence = PartnerStatePresence.query.filter_by(
            id=presence_id,
            partner_id=partner_id
        ).first_or_404()
        
        state_name = presence.state_name
        db.session.delete(presence)
        db.session.commit()
        
        return jsonify({'message': f'Presencia en {state_name} eliminada'})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500


# ============== PLANTELES (CAMPUSES) ==============

@bp.route('/<int:partner_id>/campuses', methods=['GET'])
@jwt_required()
@coordinator_required
def get_campuses(partner_id):
    """Listar planteles de un partner"""
    try:
        partner, error = _verify_partner_access(partner_id, g.current_user)
        if error:
            return error
        
        state_filter = request.args.get('state')
        active_only = request.args.get('active_only', 'true').lower() == 'true'
        
        query = Campus.query.filter_by(partner_id=partner_id)
        
        if active_only:
            query = query.filter(Campus.is_active == True)
            
        if state_filter:
            query = query.filter(Campus.state_name == state_filter)
        
        campuses = query.order_by(Campus.state_name, Campus.name).all()
        
        return jsonify({
            'partner_id': partner_id,
            'partner_name': partner.name,
            'campuses': [c.to_dict(include_groups=True) for c in campuses],
            'total': len(campuses)
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@bp.route('/<int:partner_id>/campuses', methods=['POST'])
@jwt_required()
@coordinator_required
def create_campus(partner_id):
    """Crear un nuevo plantel"""
    import secrets
    import string
    
    def generate_campus_code():
        """Genera un código único de 15 caracteres alfanuméricos"""
        alphabet = string.ascii_uppercase + string.digits
        while True:
            code = ''.join(secrets.choice(alphabet) for _ in range(15))
            # Verificar que no exista
            if not Campus.query.filter_by(code=code).first():
                return code
    
    try:
        partner, error = _verify_partner_access(partner_id, g.current_user)
        if error:
            return error
        data = request.get_json()
        
        if not data.get('name'):
            return jsonify({'error': 'El nombre es requerido'}), 400
        
        # El país por defecto es México
        country = data.get('country', 'México')
        
        # Si el país es México, el estado es requerido
        if country == 'México':
            if not data.get('state_name'):
                return jsonify({'error': 'El estado es requerido para México'}), 400
            if data['state_name'] not in MEXICAN_STATES:
                return jsonify({'error': 'Estado no válido'}), 400
        
        # El state_name puede ser None para países que no son México
        state_name = data.get('state_name') if country == 'México' else data.get('state_name', '')
            
        # Código postal es opcional
            
        if not data.get('director_name'):
            return jsonify({'error': 'El nombre del director es requerido'}), 400
        
        if not data.get('director_first_surname'):
            return jsonify({'error': 'El primer apellido del director es requerido'}), 400
        
        if not data.get('director_second_surname'):
            return jsonify({'error': 'El segundo apellido del director es requerido'}), 400
            
        if not data.get('director_email'):
            return jsonify({'error': 'El correo del director es requerido'}), 400
            
        if not data.get('director_phone'):
            return jsonify({'error': 'El teléfono del director es requerido'}), 400
        
        if not data.get('director_gender'):
            return jsonify({'error': 'El género del director es requerido'}), 400
        
        if data['director_gender'] not in ['M', 'F', 'O']:
            return jsonify({'error': 'Género del director inválido. Use M, F u O'}), 400
        
        # Determinar si el plantel es extranjero
        is_foreign = country != 'México'
        
        if is_foreign:
            # Para planteles extranjeros: asignar CURP genérico según género
            director_curp = _get_foreign_curp(data['director_gender'])
        else:
            if not data.get('director_curp'):
                return jsonify({'error': 'El CURP del director es requerido'}), 400
            
            # Validar CURP del director (18 caracteres)
            director_curp = data['director_curp'].upper().strip()
            if len(director_curp) != 18:
                return jsonify({'error': 'El CURP del director debe tener 18 caracteres'}), 400
            
            # Verificar que el CURP del director no esté ya registrado como usuario
            existing_curp_user = User.query.filter_by(curp=director_curp).first()
            if existing_curp_user:
                return jsonify({'error': f'Ya existe un usuario registrado con ese CURP ({director_curp}). No se puede usar la misma persona como director en otro plantel.'}), 400
        
        if not data.get('director_date_of_birth'):
            return jsonify({'error': 'La fecha de nacimiento del director es requerida'}), 400
        
        # Validar formato de fecha de nacimiento
        from datetime import datetime as dt
        try:
            director_dob = dt.strptime(data['director_date_of_birth'], '%Y-%m-%d').date()
        except ValueError:
            return jsonify({'error': 'Formato de fecha de nacimiento inválido. Use YYYY-MM-DD'}), 400
        
        state_auto_created = False
        
        # Verificar si el partner ya tiene presencia en ese estado
        # Si no, crearla automáticamente
        existing_presence = PartnerStatePresence.query.filter_by(
            partner_id=partner_id,
            state_name=state_name
        ).first()
        
        if not existing_presence:
            # Crear automáticamente la presencia en el estado
            new_presence = PartnerStatePresence(
                partner_id=partner_id,
                state_name=state_name,
                is_active=True
            )
            db.session.add(new_presence)
            state_auto_created = True
        
        # Generar código único para el plantel
        campus_code = generate_campus_code()
        
        # Los planteles nuevos se crean inactivos por defecto
        # Se activarán al completar el proceso de activación (asignar responsable, etc.)
        campus = Campus(
            partner_id=partner_id,
            name=data['name'],
            code=campus_code,
            country=country,
            state_name=state_name if state_name else None,
            city=data.get('city'),
            address=data.get('address'),
            postal_code=data.get('postal_code'),
            email=data.get('director_email'),
            phone=data.get('director_phone'),
            website=data.get('website'),
            # Director del plantel (datos completos)
            director_name=data.get('director_name'),
            director_first_surname=data.get('director_first_surname'),
            director_second_surname=data.get('director_second_surname'),
            director_email=data.get('director_email'),
            director_phone=data.get('director_phone'),
            director_curp=director_curp,
            director_gender=data.get('director_gender'),
            director_date_of_birth=director_dob,
            is_active=False,  # Inactivo hasta completar activación
            activation_status='pending'  # Estado inicial de activación
        )
        
        db.session.add(campus)
        db.session.flush()  # Para obtener el ID del campus
        
        # Crear usuario del director del plantel como responsable disponible
        import random
        import uuid
        
        def generate_unique_username():
            chars = string.ascii_uppercase + string.digits
            while True:
                username = ''.join(random.choices(chars, k=10))
                if not User.query.filter_by(username=username).first():
                    return username
        
        director_username = generate_unique_username()
        director_password = secrets.token_urlsafe(12)
        
        # Crear el usuario del director con rol responsable
        # El usuario del director se identifica porque su CURP coincide con director_curp del campus
        director_user = User(
            id=str(uuid.uuid4()),
            email=data.get('director_email'),
            username=director_username,
            name=data.get('director_name').strip(),
            first_surname=data.get('director_first_surname').strip(),
            second_surname=data.get('director_second_surname').strip(),
            gender=data.get('director_gender'),
            curp=director_curp,
            date_of_birth=director_dob,
            role='responsable',
            campus_id=campus.id,  # Asociado al plantel
            is_active=True,
            is_verified=True,
            can_bulk_create_candidates=False,
            can_manage_groups=False,
            can_view_reports=True
        )
        director_user.set_password(director_password)
        director_user.encrypted_password = encrypt_password(director_password)
        
        db.session.add(director_user)
        
        db.session.commit()
        
        # Obtener los estados actualizados del partner
        updated_states = [p.to_dict() for p in partner.state_presences.all()]
        
        message = 'Plantel creado exitosamente'
        if state_auto_created:
            message += f'. Se registró automáticamente la presencia en {state_name}'
        
        return jsonify({
            'message': message,
            'campus': campus.to_dict(),
            'state_auto_created': state_auto_created,
            'partner_states': updated_states,
            'director_user': {
                'id': director_user.id,
                'username': director_username,
                'full_name': director_user.full_name,
                'email': director_user.email,
                'temporary_password': director_password
            }
        }), 201
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500


@bp.route('/campuses/<int:campus_id>', methods=['GET'])
@jwt_required()
@coordinator_required
def get_campus(campus_id):
    """Obtener detalle de un plantel con ciclos escolares"""
    try:
        campus, error = _verify_campus_access(campus_id, g.current_user)
        if error:
            return error
        return jsonify({
            'campus': campus.to_dict(include_groups=True, include_partner=True, include_cycles=True, include_responsable=True)
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@bp.route('/campuses/<int:campus_id>', methods=['PUT'])
@jwt_required()
@coordinator_required
def update_campus(campus_id):
    """Actualizar un plantel"""
    try:
        campus, error = _verify_campus_access(campus_id, g.current_user)
        if error:
            return error
        data = request.get_json()
        
        # Validar estado solo si el país es México
        country = data.get('country', campus.country or 'México')
        if country == 'México' and data.get('state_name') and data['state_name'] not in MEXICAN_STATES:
            return jsonify({'error': 'Estado no válido'}), 400
        
        # Si cambia a un país que no es México, el state_name puede quedar vacío
        if 'country' in data and data['country'] != 'México' and 'state_name' not in data:
            data['state_name'] = None
        
        # Determinar si el plantel es extranjero
        is_foreign = country != 'México'
        
        # Si el plantel es extranjero y se envía género del director, auto-asignar CURP genérico
        if is_foreign:
            director_gender = data.get('director_gender', campus.director_gender)
            if director_gender:
                data['director_curp'] = _get_foreign_curp(director_gender)
        
        # Validar y procesar campos del director si se envían (solo para México)
        if 'director_curp' in data and data['director_curp'] and not is_foreign:
            data['director_curp'] = data['director_curp'].upper().strip()
            if len(data['director_curp']) != 18:
                return jsonify({'error': 'El CURP del director debe tener 18 caracteres'}), 400
            # Verificar que el CURP no esté ya registrado como otro usuario (solo si no es genérico de extranjero)
            if not _is_generic_foreign_curp(data['director_curp']):
                if data['director_curp'] != (campus.director_curp or '').upper().strip():
                    existing_curp_user = User.query.filter_by(curp=data['director_curp']).first()
                    if existing_curp_user:
                        return jsonify({'error': f'Ya existe un usuario registrado con ese CURP ({data["director_curp"]}). No se puede usar la misma persona como director en otro plantel.'}), 400
        
        if 'director_gender' in data and data['director_gender']:
            if data['director_gender'] not in ['M', 'F', 'O']:
                return jsonify({'error': 'Género del director inválido. Use M, F u O'}), 400
        
        if 'director_date_of_birth' in data and data['director_date_of_birth']:
            from datetime import datetime as dt
            try:
                data['director_date_of_birth'] = dt.strptime(data['director_date_of_birth'], '%Y-%m-%d').date()
            except ValueError:
                return jsonify({'error': 'Formato de fecha de nacimiento inválido. Use YYYY-MM-DD'}), 400
        
        # Actualizar campos (incluyendo los nuevos del director)
        for field in ['name', 'country', 'state_name', 'city', 'address', 'postal_code',
                      'website', 
                      'director_name', 'director_first_surname', 'director_second_surname',
                      'director_email', 'director_phone', 'director_curp', 'director_gender', 
                      'director_date_of_birth', 'is_active']:
            if field in data:
                setattr(campus, field, data[field])
        
        # Auto-sync email/phone from director fields
        if campus.director_email:
            campus.email = campus.director_email
        if campus.director_phone:
            campus.phone = campus.director_phone
        
        db.session.commit()
        
        return jsonify({
            'message': 'Plantel actualizado exitosamente',
            'campus': campus.to_dict()
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500


@bp.route('/campuses/<int:campus_id>', methods=['DELETE'])
@jwt_required()
@coordinator_required
def delete_campus(campus_id):
    """Eliminar un plantel (soft delete)"""
    try:
        campus, error = _verify_campus_access(campus_id, g.current_user)
        if error:
            return error
        campus.is_active = False
        db.session.commit()
        
        return jsonify({'message': 'Plantel desactivado exitosamente'})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500


@bp.route('/campuses/<int:campus_id>/permanent-delete', methods=['DELETE'])
@jwt_required()
def permanent_delete_campus(campus_id):
    """
    Eliminar permanentemente un plantel (SOLO ADMINISTRADOR)
    Elimina:
    - El plantel
    - Los ciclos escolares del plantel
    - Los grupos del plantel
    - Los miembros de grupos (GroupMember)
    - Los exámenes asignados a grupos (GroupExam) y sus miembros específicos
    - Los materiales de estudio asignados a grupos (GroupStudyMaterial) y sus miembros específicos
    - Los estándares de competencia del plantel (CampusCompetencyStandard)
    
    NO elimina y permanece intacto:
    - Los usuarios (candidatos, responsables) - solo se desvinculan del plantel
    - Los vouchers ya asignados - las certificaciones permanecen
    - Los resultados de exámenes
    """
    from app.models.partner import GroupStudyMaterial, GroupStudyMaterialMember, GroupExamMember
    
    try:
        # Verificar que el usuario sea admin
        user_id = get_jwt_identity()
        user = User.query.get(user_id)
        if not user or user.role not in ['admin', 'developer']:
            return jsonify({'error': 'Acceso denegado. Solo administradores pueden eliminar planteles permanentemente'}), 403
        
        campus = Campus.query.get_or_404(campus_id)
        partner_id = campus.partner_id
        campus_name = campus.name
        
        # Contadores para el resumen
        stats = {
            'groups_deleted': 0,
            'cycles_deleted': 0,
            'members_removed': 0,
            'exams_unassigned': 0,
            'materials_unassigned': 0,
            'users_unlinked': 0,
            'competency_standards_deleted': 0
        }
        
        # 1. Obtener todos los grupos del plantel
        groups = CandidateGroup.query.filter_by(campus_id=campus_id).all()
        
        for group in groups:
            # Contar miembros que se eliminarán
            members_count = GroupMember.query.filter_by(group_id=group.id).count()
            stats['members_removed'] += members_count
            
            # Obtener exámenes asignados y eliminar sus miembros específicos primero
            group_exams = GroupExam.query.filter_by(group_id=group.id).all()
            for ge in group_exams:
                GroupExamMember.query.filter_by(group_exam_id=ge.id).delete()
            stats['exams_unassigned'] += len(group_exams)
            
            # Obtener materiales asignados y eliminar sus miembros específicos primero
            group_materials = GroupStudyMaterial.query.filter_by(group_id=group.id).all()
            for gm in group_materials:
                GroupStudyMaterialMember.query.filter_by(group_study_material_id=gm.id).delete()
            stats['materials_unassigned'] += len(group_materials)
            
            # Eliminar miembros de grupo (los usuarios NO se eliminan, solo la membresía)
            GroupMember.query.filter_by(group_id=group.id).delete()
            
            # Eliminar asignaciones de exámenes a grupos (GroupExam)
            # Nota: Esto NO elimina los resultados del usuario, solo la asignación del grupo
            GroupExam.query.filter_by(group_id=group.id).delete()
            
            # Eliminar asignaciones de materiales a grupos
            GroupStudyMaterial.query.filter_by(group_id=group.id).delete()
            
            stats['groups_deleted'] += 1
        
        # 2. Eliminar grupos
        CandidateGroup.query.filter_by(campus_id=campus_id).delete()
        
        # 3. Eliminar ciclos escolares
        cycles_count = SchoolCycle.query.filter_by(campus_id=campus_id).count()
        SchoolCycle.query.filter_by(campus_id=campus_id).delete()
        stats['cycles_deleted'] = cycles_count
        
        # 4. Eliminar estándares de competencia del plantel
        from app.models.partner import CampusCompetencyStandard
        competency_count = CampusCompetencyStandard.query.filter_by(campus_id=campus_id).count()
        CampusCompetencyStandard.query.filter_by(campus_id=campus_id).delete()
        stats['competency_standards_deleted'] = competency_count
        
        # 5. Desvincular usuarios del plantel (NO eliminarlos)
        # Los usuarios con campus_id = este plantel se desvinculan
        users_in_campus = User.query.filter_by(campus_id=campus_id).all()
        for u in users_in_campus:
            u.campus_id = None
            stats['users_unlinked'] += 1
        
        # 6. Desvincular responsable del plantel
        campus.responsable_id = None
        
        # 7. Finalmente, eliminar el plantel
        db.session.delete(campus)
        db.session.commit()
        
        return jsonify({
            'message': f'Plantel "{campus_name}" eliminado permanentemente',
            'campus_id': campus_id,
            'partner_id': partner_id,
            'stats': stats
        }), 200
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'Error al eliminar el plantel: {str(e)}'}), 500


# ============== ACTIVACIÓN DE PLANTEL ==============

@bp.route('/campuses/<int:campus_id>/responsable', methods=['POST'])
@jwt_required()
@coordinator_required
def create_campus_responsable(campus_id):
    """
    Crear usuario responsable del plantel (Paso 1 de activación)
    Este endpoint crea un nuevo usuario con rol 'responsable' y lo asocia al plantel
    Si ya existe un responsable y se pasa replace_existing=true, permite actualizar o reemplazar
    """
    import random
    import string
    import uuid
    import secrets
    from datetime import datetime
    
    try:
        campus = Campus.query.get_or_404(campus_id)
        data = request.get_json()
        
        # Verificar si se quiere reemplazar/editar el responsable existente
        replace_existing = data.get('replace_existing', False)
        
        # Guardar el ID del responsable actual para verificaciones posteriores
        current_responsable_id = campus.responsable_id
        current_responsable = None
        
        # Verificar que el plantel no tenga ya un responsable activo
        if current_responsable_id:
            current_responsable = User.query.get(current_responsable_id)
            if current_responsable and current_responsable.is_active:
                if not replace_existing:
                    return jsonify({
                        'error': 'El plantel ya tiene un responsable asignado',
                        'current_responsable': {
                            'id': current_responsable.id,
                            'full_name': current_responsable.full_name,
                            'email': current_responsable.email
                        }
                    }), 400
                # Si replace_existing=true, continuamos sin desvincular aún
        
        # Validar campos requeridos
        # Determinar si el plantel es extranjero (CURP no requerido para extranjeros)
        is_foreign = (campus.country or 'México') != 'México'
        
        required_fields = {
            'name': 'Nombre(s)',
            'first_surname': 'Apellido paterno',
            'second_surname': 'Apellido materno',
            'email': 'Correo electrónico',
            'gender': 'Género',
            'date_of_birth': 'Fecha de nacimiento'
        }
        if not is_foreign:
            required_fields['curp'] = 'CURP'
        
        for field, label in required_fields.items():
            if not data.get(field):
                return jsonify({'error': f'El campo {label} es requerido'}), 400
        
        email = data['email'].strip().lower()
        
        # Para extranjeros: auto-asignar CURP genérico según género
        if is_foreign:
            curp = _get_foreign_curp(data['gender'])
        else:
            curp = data['curp'].upper().strip()
        
        # Validar formato de email
        import re
        email_pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
        if not re.match(email_pattern, email):
            return jsonify({'error': 'Formato de correo electrónico inválido'}), 400
        
        # Validar CURP solo para nacionales (18 caracteres alfanuméricos)
        if not is_foreign:
            curp_pattern = r'^[A-Z]{4}[0-9]{6}[HM][A-Z]{5}[A-Z0-9][0-9]$'
            if not re.match(curp_pattern, curp):
                return jsonify({'error': 'Formato de CURP inválido. Debe tener 18 caracteres'}), 400
        
        # Validar género
        if data['gender'] not in ['M', 'F', 'O']:
            return jsonify({'error': 'Género inválido. Use M (masculino), F (femenino) u O (otro)'}), 400
        
        # Validar fecha de nacimiento
        try:
            date_of_birth = datetime.strptime(data['date_of_birth'], '%Y-%m-%d').date()
        except ValueError:
            return jsonify({'error': 'Formato de fecha inválido. Use YYYY-MM-DD'}), 400
        
        # Verificar email único (excepto si es el responsable actual que estamos actualizando)
        existing_email_user = User.query.filter_by(email=email).first()
        if existing_email_user:
            # Si replace_existing=true y el usuario con ese email es el responsable actual, actualizarlo
            if replace_existing and current_responsable_id and existing_email_user.id == current_responsable_id:
                # Actualizar el responsable existente en lugar de crear uno nuevo
                existing_email_user.name = data['name'].strip()
                existing_email_user.first_surname = data['first_surname'].strip()
                existing_email_user.second_surname = data['second_surname'].strip()
                existing_email_user.gender = data['gender']
                existing_email_user.curp = curp
                existing_email_user.date_of_birth = date_of_birth
                existing_email_user.can_bulk_create_candidates = data.get('can_bulk_create_candidates', False)
                existing_email_user.can_manage_groups = data.get('can_manage_groups', False)
                existing_email_user.can_view_reports = data.get('can_view_reports', True)
                existing_email_user.campus_id = campus.id
                
                db.session.commit()
                
                return jsonify({
                    'message': 'Responsable del plantel actualizado exitosamente',
                    'responsable': {
                        'id': existing_email_user.id,
                        'username': existing_email_user.username,
                        'full_name': existing_email_user.full_name,
                        'email': existing_email_user.email,
                        'curp': existing_email_user.curp,
                        'gender': existing_email_user.gender,
                        'date_of_birth': existing_email_user.date_of_birth.isoformat(),
                        'can_bulk_create_candidates': existing_email_user.can_bulk_create_candidates,
                        'can_manage_groups': existing_email_user.can_manage_groups,
                        'can_view_reports': existing_email_user.can_view_reports
                        # No incluimos temporary_password porque no cambia
                    },
                    'campus': {
                        'id': campus.id,
                        'activation_status': campus.activation_status
                    }
                }), 200
            else:
                return jsonify({'error': 'Ya existe un usuario con ese correo electrónico'}), 400
        
        # Verificar CURP único (excepto si es el responsable actual o es CURP genérico de extranjero)
        if not _is_generic_foreign_curp(curp):
            existing_curp_user = User.query.filter_by(curp=curp).first()
            if existing_curp_user:
                if replace_existing and current_responsable_id and existing_curp_user.id == current_responsable_id:
                    # Ya manejado arriba, no debería llegar aquí si el email es el mismo
                    pass
                else:
                    return jsonify({'error': 'Ya existe un usuario con ese CURP'}), 400
        
        # Generar username único de 10 caracteres (letras y números en mayúsculas)
        def generate_unique_username():
            chars = string.ascii_uppercase + string.digits
            while True:
                username = ''.join(random.choices(chars, k=10))
                if not User.query.filter_by(username=username).first():
                    return username
        
        username = generate_unique_username()
        
        # Generar contraseña segura
        password = secrets.token_urlsafe(12)
        
        # Obtener permisos configurables (defaults a False, excepto can_view_reports)
        can_bulk_create = data.get('can_bulk_create_candidates', False)
        can_manage_groups = data.get('can_manage_groups', False)
        can_view_reports = data.get('can_view_reports', True)
        
        # Si hay un responsable anterior y estamos creando uno nuevo, desvincularlo
        if replace_existing and current_responsable:
            current_responsable.campus_id = None
            db.session.add(current_responsable)
        
        # Crear el usuario responsable
        new_user = User(
            id=str(uuid.uuid4()),
            email=email,
            username=username,
            name=data['name'].strip(),
            first_surname=data['first_surname'].strip(),
            second_surname=data['second_surname'].strip(),
            gender=data['gender'],
            curp=curp,
            date_of_birth=date_of_birth,
            role='responsable',
            campus_id=campus.id,
            is_active=True,
            is_verified=True,  # Pre-verificado ya que lo crea un admin/coordinator
            can_bulk_create_candidates=can_bulk_create,
            can_manage_groups=can_manage_groups,
            can_view_reports=can_view_reports
        )
        new_user.set_password(password)
        new_user.encrypted_password = encrypt_password(password)
        
        db.session.add(new_user)
        
        # Asociar el responsable al plantel y actualizar estado de activación
        campus.responsable_id = new_user.id
        campus.activation_status = 'configuring'  # Avanza al siguiente estado
        
        db.session.commit()
        
        return jsonify({
            'message': 'Responsable del plantel creado exitosamente',
            'responsable': {
                'id': new_user.id,
                'username': new_user.username,
                'full_name': new_user.full_name,
                'email': new_user.email,
                'curp': new_user.curp,
                'gender': new_user.gender,
                'date_of_birth': new_user.date_of_birth.isoformat(),
                'can_bulk_create_candidates': new_user.can_bulk_create_candidates,
                'can_manage_groups': new_user.can_manage_groups,
                'can_view_reports': new_user.can_view_reports,
                'temporary_password': password  # Solo se muestra una vez
            },
            'campus': {
                'id': campus.id,
                'name': campus.name,
                'code': campus.code,
                'activation_status': campus.activation_status
            }
        }), 201
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500


@bp.route('/campuses/<int:campus_id>/responsable', methods=['GET'])
@jwt_required()
@coordinator_required
def get_campus_responsable(campus_id):
    """Obtener información del responsable del plantel"""
    try:
        campus = Campus.query.get_or_404(campus_id)
        
        if not campus.responsable_id:
            return jsonify({
                'message': 'El plantel no tiene responsable asignado',
                'responsable': None,
                'activation_status': campus.activation_status
            })
        
        responsable = User.query.get(campus.responsable_id)
        if not responsable:
            return jsonify({
                'message': 'El responsable asignado no existe',
                'responsable': None,
                'activation_status': campus.activation_status
            })
        
        return jsonify({
            'responsable': {
                'id': responsable.id,
                'username': responsable.username,
                'full_name': responsable.full_name,
                'email': responsable.email,
                'curp': responsable.curp,
                'gender': responsable.gender,
                'date_of_birth': responsable.date_of_birth.isoformat() if responsable.date_of_birth else None,
                'can_bulk_create_candidates': responsable.can_bulk_create_candidates,
                'can_manage_groups': responsable.can_manage_groups,
                'is_active': responsable.is_active,
                'created_at': responsable.created_at.isoformat() if responsable.created_at else None
            },
            'activation_status': campus.activation_status
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@bp.route('/campuses/<int:campus_id>/responsable', methods=['PUT'])
@jwt_required()
@coordinator_required
def update_campus_responsable(campus_id):
    """Actualizar datos o permisos del responsable del plantel"""
    try:
        campus = Campus.query.get_or_404(campus_id)
        
        if not campus.responsable_id:
            return jsonify({'error': 'El plantel no tiene responsable asignado'}), 404
        
        responsable = User.query.get(campus.responsable_id)
        if not responsable:
            return jsonify({'error': 'El responsable asignado no existe'}), 404
        
        data = request.get_json()
        
        # Campos actualizables
        updatable_fields = ['name', 'first_surname', 'second_surname', 'gender', 'phone']
        for field in updatable_fields:
            if field in data:
                setattr(responsable, field, data[field].strip() if isinstance(data[field], str) else data[field])
        
        # Actualizar permisos
        if 'can_bulk_create_candidates' in data:
            responsable.can_bulk_create_candidates = bool(data['can_bulk_create_candidates'])
        if 'can_manage_groups' in data:
            responsable.can_manage_groups = bool(data['can_manage_groups'])
        
        # Actualizar estado activo
        if 'is_active' in data:
            responsable.is_active = bool(data['is_active'])
        
        db.session.commit()
        
        return jsonify({
            'message': 'Responsable actualizado exitosamente',
            'responsable': {
                'id': responsable.id,
                'username': responsable.username,
                'full_name': responsable.full_name,
                'email': responsable.email,
                'curp': responsable.curp,
                'gender': responsable.gender,
                'date_of_birth': responsable.date_of_birth.isoformat() if responsable.date_of_birth else None,
                'can_bulk_create_candidates': responsable.can_bulk_create_candidates,
                'can_manage_groups': responsable.can_manage_groups,
                'is_active': responsable.is_active
            }
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500


@bp.route('/campuses/<int:campus_id>/available-responsables', methods=['GET'])
@jwt_required()
@coordinator_required
def get_available_responsables(campus_id):
    """
    Obtener lista de responsables disponibles que pueden ser asignados a este plantel.
    Solo muestra responsables del mismo partner que no estén asignados a otro plantel activo.
    El director del plantel (identificado por su CURP) aparece primero en la lista.
    """
    try:
        campus = Campus.query.get_or_404(campus_id)
        
        # Obtener todos los planteles del mismo partner
        partner_campus_ids = [c.id for c in Campus.query.filter_by(partner_id=campus.partner_id).all()]
        
        # Buscar usuarios con rol 'responsable' que pertenezcan a algún plantel del partner
        # y que no estén asignados actualmente a ningún plantel activo
        responsables = User.query.filter(
            User.role == 'responsable',
            User.is_active == True,
            User.campus_id.in_(partner_campus_ids)
        ).all()
        
        available = []
        for resp in responsables:
            # Verificar si está asignado a un plantel activo (diferente al actual)
            assigned_campus = Campus.query.filter(
                Campus.responsable_id == resp.id,
                Campus.id != campus_id,
                Campus.is_active == True
            ).first()
            
            # También verificar si está asignado al plantel actual
            is_current = campus.responsable_id == resp.id
            
            # Verificar si es el director del plantel (su CURP coincide con director_curp)
            is_director = resp.curp and campus.director_curp and resp.curp.upper() == campus.director_curp.upper()
            
            if not assigned_campus:
                available.append({
                    'id': resp.id,
                    'full_name': resp.full_name,
                    'email': resp.email,
                    'curp': resp.curp,
                    'gender': resp.gender,
                    'date_of_birth': resp.date_of_birth.isoformat() if resp.date_of_birth else None,
                    'username': resp.username,
                    'can_bulk_create_candidates': resp.can_bulk_create_candidates,
                    'can_manage_groups': resp.can_manage_groups,
                    'can_view_reports': resp.can_view_reports,
                    'is_current': is_current,
                    'is_director': is_director,
                    'campus_id': resp.campus_id
                })
        
        # Ordenar: director primero, luego por nombre
        available.sort(key=lambda x: (not x['is_director'], x['full_name'].lower()))
        
        return jsonify({
            'available_responsables': available,
            'total': len(available),
            'campus_id': campus_id,
            'partner_id': campus.partner_id
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@bp.route('/campuses/<int:campus_id>/assign-responsable', methods=['POST'])
@jwt_required()
@coordinator_required
def assign_existing_responsable(campus_id):
    """
    Asignar un responsable existente a un plantel.
    El responsable debe pertenecer al mismo partner y no estar asignado a otro plantel activo.
    """
    try:
        campus = Campus.query.get_or_404(campus_id)
        data = request.get_json()
        
        responsable_id = data.get('responsable_id')
        if not responsable_id:
            return jsonify({'error': 'El ID del responsable es requerido'}), 400
        
        responsable = User.query.get(responsable_id)
        if not responsable:
            return jsonify({'error': 'Responsable no encontrado'}), 404
        
        if responsable.role != 'responsable':
            return jsonify({'error': 'El usuario seleccionado no es un responsable'}), 400
        
        if not responsable.is_active:
            return jsonify({'error': 'El responsable no está activo'}), 400
        
        # Verificar que el responsable pertenezca a un plantel del mismo partner
        resp_campus = Campus.query.get(responsable.campus_id) if responsable.campus_id else None
        if not resp_campus or resp_campus.partner_id != campus.partner_id:
            return jsonify({'error': 'El responsable debe pertenecer al mismo partner'}), 400
        
        # Verificar que no esté asignado a otro plantel activo
        assigned_campus = Campus.query.filter(
            Campus.responsable_id == responsable_id,
            Campus.id != campus_id,
            Campus.is_active == True
        ).first()
        
        if assigned_campus:
            return jsonify({
                'error': f'El responsable ya está asignado al plantel activo: {assigned_campus.name}'
            }), 400
        
        # Si el plantel ya tiene otro responsable, liberarlo
        if campus.responsable_id and campus.responsable_id != responsable_id:
            # El responsable anterior mantiene su rol pero queda sin plantel asignado
            pass  # No hacemos nada con el anterior, solo lo reemplazamos
        
        # Asignar el responsable al plantel
        campus.responsable_id = responsable_id
        responsable.campus_id = campus.id  # Actualizar el campus_id del responsable
        
        # Actualizar permisos si se envían
        if 'can_bulk_create_candidates' in data:
            responsable.can_bulk_create_candidates = bool(data['can_bulk_create_candidates'])
        if 'can_manage_groups' in data:
            responsable.can_manage_groups = bool(data['can_manage_groups'])
        if 'can_view_reports' in data:
            responsable.can_view_reports = bool(data['can_view_reports'])
        
        # Avanzar el estado de activación
        if campus.activation_status == 'pending':
            campus.activation_status = 'configuring'
        
        db.session.commit()
        
        return jsonify({
            'message': 'Responsable asignado exitosamente',
            'responsable': {
                'id': responsable.id,
                'username': responsable.username,
                'full_name': responsable.full_name,
                'email': responsable.email,
                'curp': responsable.curp,
                'gender': responsable.gender,
                'date_of_birth': responsable.date_of_birth.isoformat() if responsable.date_of_birth else None,
                'can_bulk_create_candidates': responsable.can_bulk_create_candidates,
                'can_manage_groups': responsable.can_manage_groups,
                'can_view_reports': responsable.can_view_reports,
                'is_active': responsable.is_active
            },
            'campus': {
                'id': campus.id,
                'name': campus.name,
                'code': campus.code,
                'activation_status': campus.activation_status
            }
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500


# ============== MULTI-RESPONSABLES ==============

@bp.route('/campuses/<int:campus_id>/responsables', methods=['GET'])
@jwt_required()
@coordinator_required
def list_campus_responsables(campus_id):
    """Listar TODOS los responsables del plantel (no solo el principal)"""
    try:
        campus = Campus.query.get_or_404(campus_id)

        # Buscar todos los usuarios con role='responsable' y campus_id=este campus
        responsables = User.query.filter(
            User.role == 'responsable',
            User.campus_id == campus.id
        ).order_by(User.created_at.asc()).all()

        result = []
        for r in responsables:
            result.append({
                'id': r.id,
                'username': r.username,
                'full_name': r.full_name,
                'email': r.email,
                'curp': r.curp,
                'gender': r.gender,
                'date_of_birth': r.date_of_birth.isoformat() if r.date_of_birth else None,
                'can_bulk_create_candidates': r.can_bulk_create_candidates,
                'can_manage_groups': r.can_manage_groups,
                'can_view_reports': r.can_view_reports,
                'is_active': r.is_active,
                'is_primary': r.id == campus.responsable_id,
                'created_at': r.created_at.isoformat() if r.created_at else None
            })

        return jsonify({
            'responsables': result,
            'total': len(result),
            'primary_responsable_id': campus.responsable_id
        })

    except Exception as e:
        return jsonify({'error': str(e)}), 500


@bp.route('/campuses/<int:campus_id>/responsables', methods=['POST'])
@jwt_required()
@coordinator_required
def add_campus_responsable(campus_id):
    """
    Crear un nuevo responsable adicional para el plantel.
    Crea un usuario real con rol 'responsable' y lo asocia al plantel.
    """
    import random
    import string
    import uuid
    import secrets
    from datetime import datetime as dt

    try:
        campus = Campus.query.get_or_404(campus_id)
        data = request.get_json()
        
        # Determinar si el plantel es extranjero
        is_foreign = (campus.country or 'México') != 'México'

        # Validar campos requeridos
        required_fields = {
            'name': 'Nombre(s)',
            'first_surname': 'Apellido paterno',
            'second_surname': 'Apellido materno',
            'email': 'Correo electrónico',
            'gender': 'Género',
            'date_of_birth': 'Fecha de nacimiento'
        }
        if not is_foreign:
            required_fields['curp'] = 'CURP'

        for field, label in required_fields.items():
            if not data.get(field):
                return jsonify({'error': f'El campo {label} es requerido'}), 400

        email = data['email'].strip().lower()
        
        # Para extranjeros: auto-asignar CURP genérico según género
        if is_foreign:
            curp = _get_foreign_curp(data['gender'])
        else:
            curp = data['curp'].upper().strip()

        # Validar formato de email
        import re
        email_pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
        if not re.match(email_pattern, email):
            return jsonify({'error': 'Formato de correo electrónico inválido'}), 400

        # Validar CURP solo para nacionales
        if not is_foreign:
            curp_pattern = r'^[A-Z]{4}[0-9]{6}[HM][A-Z]{5}[A-Z0-9][0-9]$'
            if not re.match(curp_pattern, curp):
                return jsonify({'error': 'Formato de CURP inválido. Debe tener 18 caracteres'}), 400

        if data['gender'] not in ['M', 'F', 'O']:
            return jsonify({'error': 'Género inválido'}), 400

        try:
            date_of_birth = dt.strptime(data['date_of_birth'], '%Y-%m-%d').date()
        except ValueError:
            return jsonify({'error': 'Formato de fecha inválido. Use YYYY-MM-DD'}), 400

        # Verificar unicidad de email y CURP
        if User.query.filter_by(email=email).first():
            return jsonify({'error': 'Ya existe un usuario registrado con ese correo electrónico. No se puede usar el mismo correo para otro responsable.'}), 400
        # CURP genérico de extranjero puede repetirse
        if not _is_generic_foreign_curp(curp):
            existing_curp_user = User.query.filter_by(curp=curp).first()
            if existing_curp_user:
                return jsonify({'error': f'Ya existe un usuario registrado con ese CURP ({curp}). No se puede usar la misma persona como responsable en otro plantel.'}), 400

        # Generar username y password
        chars = string.ascii_uppercase + string.digits
        while True:
            username = ''.join(random.choices(chars, k=10))
            if not User.query.filter_by(username=username).first():
                break
        password = secrets.token_urlsafe(12)

        new_user = User(
            id=str(uuid.uuid4()),
            email=email,
            username=username,
            name=data['name'].strip(),
            first_surname=data['first_surname'].strip(),
            second_surname=data['second_surname'].strip(),
            gender=data['gender'],
            curp=curp,
            date_of_birth=date_of_birth,
            role='responsable',
            campus_id=campus.id,
            is_active=True,
            is_verified=True,
            can_bulk_create_candidates=bool(data.get('can_bulk_create_candidates', False)),
            can_manage_groups=bool(data.get('can_manage_groups', False)),
            can_view_reports=bool(data.get('can_view_reports', True)),
        )
        new_user.set_password(password)

        db.session.add(new_user)

        # Si el plantel no tiene responsable principal, asignar este
        if not campus.responsable_id:
            campus.responsable_id = new_user.id
            if campus.activation_status == 'pending':
                campus.activation_status = 'configuring'

        db.session.commit()

        return jsonify({
            'message': 'Responsable creado exitosamente',
            'responsable': {
                'id': new_user.id,
                'username': new_user.username,
                'full_name': new_user.full_name,
                'email': new_user.email,
                'curp': new_user.curp,
                'gender': new_user.gender,
                'date_of_birth': new_user.date_of_birth.isoformat(),
                'can_bulk_create_candidates': new_user.can_bulk_create_candidates,
                'can_manage_groups': new_user.can_manage_groups,
                'can_view_reports': new_user.can_view_reports,
                'is_active': True,
                'is_primary': new_user.id == campus.responsable_id,
                'temporary_password': password,
            },
            'campus': {
                'id': campus.id,
                'activation_status': campus.activation_status
            }
        }), 201

    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500


@bp.route('/campuses/<int:campus_id>/responsables/<string:user_id>', methods=['PUT'])
@jwt_required()
@coordinator_required
def update_responsable_permissions(campus_id, user_id):
    """Actualizar permisos de un responsable específico del plantel"""
    try:
        campus = Campus.query.get_or_404(campus_id)
        responsable = User.query.get_or_404(user_id)

        if responsable.role != 'responsable' or responsable.campus_id != campus.id:
            return jsonify({'error': 'El usuario no es responsable de este plantel'}), 400

        data = request.get_json()

        if 'can_bulk_create_candidates' in data:
            responsable.can_bulk_create_candidates = bool(data['can_bulk_create_candidates'])
        if 'can_manage_groups' in data:
            responsable.can_manage_groups = bool(data['can_manage_groups'])
        if 'can_view_reports' in data:
            responsable.can_view_reports = bool(data['can_view_reports'])
        if 'is_active' in data:
            responsable.is_active = bool(data['is_active'])

        db.session.commit()

        return jsonify({
            'message': 'Permisos actualizados exitosamente',
            'responsable': {
                'id': responsable.id,
                'username': responsable.username,
                'full_name': responsable.full_name,
                'email': responsable.email,
                'can_bulk_create_candidates': responsable.can_bulk_create_candidates,
                'can_manage_groups': responsable.can_manage_groups,
                'can_view_reports': responsable.can_view_reports,
                'is_active': responsable.is_active,
                'is_primary': responsable.id == campus.responsable_id,
            }
        })

    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500


@bp.route('/campuses/<int:campus_id>/responsables/<string:user_id>', methods=['DELETE'])
@jwt_required()
@coordinator_required
def remove_campus_responsable(campus_id, user_id):
    """Desactivar un responsable del plantel"""
    try:
        campus = Campus.query.get_or_404(campus_id)
        responsable = User.query.get_or_404(user_id)

        if responsable.role != 'responsable' or responsable.campus_id != campus.id:
            return jsonify({'error': 'El usuario no es responsable de este plantel'}), 400

        # No permitir eliminar al principal si es el único activo
        active_count = User.query.filter(
            User.role == 'responsable',
            User.campus_id == campus.id,
            User.is_active == True,
            User.id != user_id
        ).count()

        if responsable.id == campus.responsable_id and active_count == 0:
            return jsonify({'error': 'No se puede eliminar al único responsable activo del plantel'}), 400

        responsable.is_active = False

        # Si era el principal, reasignar a otro activo
        if responsable.id == campus.responsable_id and active_count > 0:
            next_resp = User.query.filter(
                User.role == 'responsable',
                User.campus_id == campus.id,
                User.is_active == True,
                User.id != user_id
            ).first()
            if next_resp:
                campus.responsable_id = next_resp.id

        db.session.commit()

        return jsonify({
            'message': 'Responsable desactivado exitosamente',
            'responsable_id': user_id
        })

    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500


@bp.route('/campuses/<int:campus_id>/activate', methods=['POST'])
@jwt_required()
@coordinator_required
def activate_campus(campus_id):
    """Activar un plantel después de completar el proceso de configuración, o reactivar uno previamente configurado"""
    try:
        campus = Campus.query.get_or_404(campus_id)
        
        # Verificar que el plantel no esté ya activo
        if campus.is_active:
            return jsonify({
                'error': 'El plantel ya está activo'
            }), 400
        
        # Si ya fue configurado previamente, permitir reactivación directa
        if campus.configuration_completed:
            campus.is_active = True
            campus.activation_status = 'active'
            if not campus.activated_at:
                campus.activated_at = datetime.utcnow()
            
            db.session.commit()
            
            return jsonify({
                'message': 'Plantel reactivado exitosamente',
                'campus': campus.to_dict(include_config=True)
            })
        
        # Primera activación: verificar requisitos
        if not campus.responsable_id:
            return jsonify({
                'error': 'El plantel debe tener un responsable asignado antes de activarse'
            }), 400
        
        if not campus.configuration_completed:
            return jsonify({
                'error': 'Debe completar la configuración del plantel antes de activarlo'
            }), 400
        
        # Activar el plantel
        campus.is_active = True
        campus.activation_status = 'active'
        campus.activated_at = datetime.utcnow()
        
        db.session.commit()
        
        return jsonify({
            'message': 'Plantel activado exitosamente',
            'campus': campus.to_dict(include_config=True)
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500


@bp.route('/campuses/<int:campus_id>/configure', methods=['POST'])
@jwt_required()
@coordinator_required
def configure_campus(campus_id):
    """Configurar un plantel (paso 2 del proceso de activación)"""
    try:
        campus = Campus.query.get_or_404(campus_id)
        
        # Verificar que el plantel tenga un responsable asignado
        if not campus.responsable_id:
            return jsonify({
                'error': 'El plantel debe tener un responsable asignado antes de configurarse'
            }), 400
        
        data = request.get_json()
        
        # Versión de Office
        if 'office_version' in data:
            valid_versions = ['office_2016', 'office_2019', 'office_365']
            if data['office_version'] not in valid_versions:
                return jsonify({'error': f'Versión de Office inválida. Opciones: {", ".join(valid_versions)}'}), 400
            campus.office_version = data['office_version']
        
        # Tiers de certificación
        if 'enable_tier_basic' in data:
            campus.enable_tier_basic = bool(data['enable_tier_basic'])
        if 'enable_tier_standard' in data:
            campus.enable_tier_standard = bool(data['enable_tier_standard'])
        if 'enable_tier_advanced' in data:
            campus.enable_tier_advanced = bool(data['enable_tier_advanced'])
        if 'enable_digital_badge' in data:
            campus.enable_digital_badge = bool(data['enable_digital_badge'])
        
        # Evaluaciones parciales
        if 'enable_partial_evaluations' in data:
            campus.enable_partial_evaluations = bool(data['enable_partial_evaluations'])
        if 'enable_unscheduled_partials' in data:
            campus.enable_unscheduled_partials = bool(data['enable_unscheduled_partials'])
        
        # Máquinas virtuales
        if 'enable_virtual_machines' in data:
            campus.enable_virtual_machines = bool(data['enable_virtual_machines'])
        
        # Pagos en línea
        if 'enable_online_payments' in data:
            campus.enable_online_payments = bool(data['enable_online_payments'])
        
        # Visibilidad de certificados para candidatos
        if 'enable_candidate_certificates' in data:
            campus.enable_candidate_certificates = bool(data['enable_candidate_certificates'])
        
        # PIN de seguridad para exámenes
        if 'require_exam_pin' in data:
            campus.require_exam_pin = bool(data['require_exam_pin'])
        
        # Calendario de sesiones
        if 'enable_session_calendar' in data:
            campus.enable_session_calendar = bool(data['enable_session_calendar'])
        if 'session_scheduling_mode' in data:
            mode = data['session_scheduling_mode']
            if mode in ('leader_only', 'candidate_self'):
                campus.session_scheduling_mode = mode
        
        # Vigencia de asignaciones (meses)
        if 'assignment_validity_months' in data:
            val = data['assignment_validity_months']
            campus.assignment_validity_months = int(val) if val else 12
        
        # Costos
        if 'certification_cost' in data:
            campus.certification_cost = float(data['certification_cost']) if data['certification_cost'] else 0
        if 'retake_cost' in data:
            campus.retake_cost = float(data['retake_cost']) if data['retake_cost'] else 0
        if 'max_retakes' in data:
            val = data['max_retakes']
            campus.max_retakes = int(val) if val else 0
        
        # ECM (Estándares de Competencia) asignados al plantel
        if 'competency_standard_ids' in data:
            from app.models.partner import CampusCompetencyStandard
            from app.models.competency_standard import CompetencyStandard
            
            standard_ids = data['competency_standard_ids']
            if not isinstance(standard_ids, list):
                return jsonify({'error': 'competency_standard_ids debe ser una lista'}), 400
            
            # Validar que todos los IDs existan y estén activos
            for std_id in standard_ids:
                standard = CompetencyStandard.query.get(std_id)
                if not standard:
                    return jsonify({'error': f'Estándar de competencia con ID {std_id} no encontrado'}), 404
                if not standard.is_active:
                    return jsonify({'error': f'El estándar {standard.code} no está activo'}), 400
            
            # Eliminar asignaciones existentes y crear nuevas
            CampusCompetencyStandard.query.filter_by(campus_id=campus_id).delete()
            for std_id in standard_ids:
                new_assignment = CampusCompetencyStandard(
                    campus_id=campus_id,
                    competency_standard_id=std_id
                )
                db.session.add(new_assignment)
        
        # Marcar configuración como completada si se proporciona el flag
        if data.get('complete_configuration'):
            from app.models.partner import CampusCompetencyStandard
            
            # Validar que se hayan configurado los campos requeridos
            if not campus.assignment_validity_months or campus.assignment_validity_months <= 0:
                return jsonify({'error': 'Debe establecer los meses de vigencia de las asignaciones'}), 400
            
            # Validar costos > 0
            if not campus.certification_cost or campus.certification_cost <= 0:
                return jsonify({'error': 'El costo de certificación debe ser mayor a $0'}), 400
            if not campus.retake_cost or campus.retake_cost <= 0:
                return jsonify({'error': 'El costo de retoma debe ser mayor a $0'}), 400
            
            # Al menos un tier debe estar habilitado
            if not (campus.enable_tier_basic or campus.enable_tier_standard or campus.enable_tier_advanced or campus.enable_digital_badge):
                return jsonify({'error': 'Debe habilitar al menos un tipo de certificación'}), 400
            
            # Al menos un ECM debe estar asignado
            ecm_count = CampusCompetencyStandard.query.filter_by(campus_id=campus_id).count()
            if ecm_count == 0:
                return jsonify({'error': 'Debe asignar al menos un Estándar de Competencia (ECM) al plantel'}), 400
            
            campus.configuration_completed = True
            campus.configuration_completed_at = datetime.utcnow()
        
        db.session.commit()
        
        return jsonify({
            'message': 'Configuración guardada exitosamente',
            'campus': campus.to_dict(include_config=True, include_responsable=True)
        })
        
    except ValueError as e:
        db.session.rollback()
        return jsonify({'error': f'Error en formato de datos: {str(e)}'}), 400
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500


@bp.route('/campuses/<int:campus_id>/config', methods=['GET'])
@jwt_required()
@coordinator_required
def get_campus_config(campus_id):
    """Obtener la configuración de un plantel"""
    try:
        campus = Campus.query.get_or_404(campus_id)
        
        return jsonify({
            'campus_id': campus.id,
            'campus_name': campus.name,
            'configuration': {
                'office_version': campus.office_version or 'office_365',
                'enable_tier_basic': campus.enable_tier_basic if campus.enable_tier_basic is not None else True,
                'enable_tier_standard': campus.enable_tier_standard or False,
                'enable_tier_advanced': campus.enable_tier_advanced or False,
                'enable_digital_badge': campus.enable_digital_badge or False,
                'enable_partial_evaluations': campus.enable_partial_evaluations or False,
                'enable_unscheduled_partials': campus.enable_unscheduled_partials or False,
                'enable_virtual_machines': campus.enable_virtual_machines or False,
                'enable_online_payments': campus.enable_online_payments or False,
                'enable_candidate_certificates': campus.enable_candidate_certificates or False,
            'require_exam_pin': campus.require_exam_pin or False,
            'daily_exam_pin': campus.get_daily_pin() if campus.require_exam_pin else None,
                'enable_session_calendar': campus.enable_session_calendar or False,
                'session_scheduling_mode': campus.session_scheduling_mode or 'leader_only',
                'assignment_validity_months': campus.assignment_validity_months or 12,
                'certification_cost': float(campus.certification_cost) if campus.certification_cost else 0,
                'retake_cost': float(campus.retake_cost) if campus.retake_cost else 0,
                'max_retakes': campus.max_retakes if campus.max_retakes is not None else 0,
                'configuration_completed': campus.configuration_completed or False,
                'configuration_completed_at': campus.configuration_completed_at.isoformat() if campus.configuration_completed_at else None,
            }
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@bp.route('/campuses/<int:campus_id>/deactivate', methods=['POST'])
@jwt_required()
@coordinator_required
def deactivate_campus(campus_id):
    """Desactivar un plantel (preserva configuración para reactivación rápida)"""
    try:
        campus = Campus.query.get_or_404(campus_id)
        
        # Solo desactivar, preservar toda la configuración
        campus.is_active = False
        campus.activation_status = 'inactive'
        
        db.session.commit()
        
        return jsonify({
            'message': 'Plantel desactivado. Puede reactivarlo en cualquier momento.',
            'campus': campus.to_dict(include_config=True)
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500


# ============== ECM DEL PLANTEL (Estándares de Competencia) ==============

@bp.route('/campuses/<int:campus_id>/competency-standards', methods=['GET'])
@jwt_required()
@coordinator_required
def get_campus_competency_standards(campus_id):
    """Obtener los ECM asignados a un plantel"""
    try:
        from app.models.partner import CampusCompetencyStandard
        from app.models.competency_standard import CompetencyStandard
        from app.models.brand import Brand
        
        campus = Campus.query.get_or_404(campus_id)
        
        # Obtener los ECM asignados al plantel
        campus_standards = CampusCompetencyStandard.query.filter_by(campus_id=campus_id).all()
        
        standards_data = []
        for cs in campus_standards:
            standard = CompetencyStandard.query.get(cs.competency_standard_id)
            if standard:
                brand_name = None
                brand_logo = None
                if standard.brand_id:
                    brand_obj = Brand.query.get(standard.brand_id)
                    if brand_obj:
                        brand_name = brand_obj.name
                        brand_logo = brand_obj.logo_url
                standards_data.append({
                    'id': cs.id,
                    'competency_standard_id': standard.id,
                    'code': standard.code,
                    'name': standard.name,
                    'sector': standard.sector,
                    'level': standard.level,
                    'is_active': standard.is_active,
                    'logo_url': standard.logo_url,
                    'brand': brand_name,
                    'brand_logo_url': brand_logo,
                    'assigned_at': cs.created_at.isoformat() if cs.created_at else None
                })
        
        return jsonify({
            'campus_id': campus_id,
            'campus_name': campus.name,
            'competency_standards': standards_data,
            'total': len(standards_data)
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@bp.route('/campuses/<int:campus_id>/competency-standards', methods=['PUT'])
@jwt_required()
@coordinator_required
def update_campus_competency_standards(campus_id):
    """Actualizar los ECM asignados a un plantel (reemplaza todos)"""
    try:
        from app.models.partner import CampusCompetencyStandard
        from app.models.competency_standard import CompetencyStandard
        
        campus = Campus.query.get_or_404(campus_id)
        data = request.get_json()
        
        # Lista de IDs de competency_standards a asignar
        standard_ids = data.get('competency_standard_ids', [])
        
        if not isinstance(standard_ids, list):
            return jsonify({'error': 'competency_standard_ids debe ser una lista'}), 400
        
        # Validar que todos los IDs existan
        for std_id in standard_ids:
            standard = CompetencyStandard.query.get(std_id)
            if not standard:
                return jsonify({'error': f'Estándar de competencia con ID {std_id} no encontrado'}), 404
            if not standard.is_active:
                return jsonify({'error': f'El estándar {standard.code} no está activo'}), 400
        
        # Eliminar asignaciones existentes
        CampusCompetencyStandard.query.filter_by(campus_id=campus_id).delete()
        
        # Crear nuevas asignaciones
        for std_id in standard_ids:
            new_assignment = CampusCompetencyStandard(
                campus_id=campus_id,
                competency_standard_id=std_id
            )
            db.session.add(new_assignment)
        
        db.session.commit()
        
        # Obtener datos actualizados
        campus_standards = CampusCompetencyStandard.query.filter_by(campus_id=campus_id).all()
        standards_data = []
        for cs in campus_standards:
            standard = CompetencyStandard.query.get(cs.competency_standard_id)
            if standard:
                standards_data.append({
                    'id': cs.id,
                    'competency_standard_id': standard.id,
                    'code': standard.code,
                    'name': standard.name,
                    'sector': standard.sector,
                    'is_active': standard.is_active,
                    'assigned_at': cs.created_at.isoformat() if cs.created_at else None
                })
        
        return jsonify({
            'message': f'{len(standard_ids)} estándar(es) de competencia asignado(s) al plantel',
            'campus_id': campus_id,
            'competency_standards': standards_data,
            'total': len(standards_data)
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500


@bp.route('/competency-standards/available', methods=['GET'])
@jwt_required()
@coordinator_required
def get_available_competency_standards():
    """Obtener lista de todos los ECM activos disponibles para asignar"""
    try:
        from app.models.competency_standard import CompetencyStandard
        from sqlalchemy.orm import joinedload
        
        standards = CompetencyStandard.query.options(
            joinedload(CompetencyStandard.brand)
        ).filter_by(is_active=True).order_by(CompetencyStandard.code).all()
        
        return jsonify({
            'competency_standards': [
                {
                    'id': s.id,
                    'code': s.code,
                    'name': s.name,
                    'description': s.description,
                    'sector': s.sector,
                    'level': s.level,
                    'validity_years': s.validity_years,
                    'certifying_body': s.certifying_body,
                    'logo_url': s.logo_url,
                    'brand': s.brand.name if s.brand else None,
                    'brand_logo_url': s.brand.logo_url if s.brand else None,
                }
                for s in standards
            ],
            'total': len(standards)
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ============== CICLOS ESCOLARES ==============

@bp.route('/campuses/<int:campus_id>/cycles', methods=['GET'])
@jwt_required()
@coordinator_required
def get_school_cycles(campus_id):
    """Listar ciclos escolares de un plantel"""
    try:
        campus = Campus.query.get_or_404(campus_id)
        
        active_only = request.args.get('active_only', 'true').lower() == 'true'
        
        query = campus.school_cycles
        if active_only:
            query = query.filter_by(is_active=True)
        
        cycles = query.order_by(SchoolCycle.start_date.desc()).all()
        
        return jsonify({
            'cycles': [c.to_dict(include_groups=True) for c in cycles],
            'total': len(cycles)
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@bp.route('/campuses/<int:campus_id>/cycles', methods=['POST'])
@jwt_required()
@coordinator_required
def create_school_cycle(campus_id):
    """Crear un nuevo ciclo escolar"""
    try:
        campus = Campus.query.get_or_404(campus_id)
        
        # Verificar que el plantel esté activo
        if not campus.is_active:
            return jsonify({
                'error': 'No se pueden crear ciclos escolares en un plantel inactivo',
                'message': 'El plantel debe completar el proceso de activación antes de crear ciclos escolares',
                'activation_status': campus.activation_status
            }), 400
        
        data = request.get_json()
        
        # Validaciones
        if not data.get('name'):
            return jsonify({'error': 'El nombre del ciclo es requerido'}), 400
        if not data.get('cycle_type') or data['cycle_type'] not in ['annual', 'semester']:
            return jsonify({'error': 'El tipo de ciclo debe ser "annual" o "semester"'}), 400
        if not data.get('start_date') or not data.get('end_date'):
            return jsonify({'error': 'Las fechas de inicio y fin son requeridas'}), 400
        
        from datetime import datetime
        start_date = datetime.strptime(data['start_date'], '%Y-%m-%d').date()
        end_date = datetime.strptime(data['end_date'], '%Y-%m-%d').date()
        
        if end_date <= start_date:
            return jsonify({'error': 'La fecha de fin debe ser posterior a la fecha de inicio'}), 400
        
        # Si se marca como ciclo actual, desmarcar los demás
        if data.get('is_current', False):
            campus.school_cycles.filter_by(is_current=True).update({'is_current': False})
        
        cycle = SchoolCycle(
            campus_id=campus_id,
            name=data['name'],
            cycle_type=data['cycle_type'],
            start_date=start_date,
            end_date=end_date,
            is_current=data.get('is_current', False)
        )
        
        db.session.add(cycle)
        db.session.commit()
        
        return jsonify({
            'message': 'Ciclo escolar creado exitosamente',
            'cycle': cycle.to_dict()
        }), 201
        
    except ValueError as ve:
        return jsonify({'error': f'Formato de fecha inválido: {str(ve)}'}), 400
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500


@bp.route('/cycles/<int:cycle_id>', methods=['GET'])
@jwt_required()
@coordinator_required
def get_school_cycle(cycle_id):
    """Obtener detalle de un ciclo escolar"""
    try:
        cycle = SchoolCycle.query.get_or_404(cycle_id)
        return jsonify({
            'cycle': cycle.to_dict(include_groups=True, include_campus=True)
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@bp.route('/cycles/<int:cycle_id>', methods=['PUT'])
@jwt_required()
@coordinator_required
def update_school_cycle(cycle_id):
    """Actualizar un ciclo escolar"""
    try:
        cycle = SchoolCycle.query.get_or_404(cycle_id)
        data = request.get_json()
        
        if 'name' in data:
            cycle.name = data['name']
        if 'cycle_type' in data:
            if data['cycle_type'] not in ['annual', 'semester']:
                return jsonify({'error': 'El tipo de ciclo debe ser "annual" o "semester"'}), 400
            cycle.cycle_type = data['cycle_type']
        
        if 'start_date' in data:
            from datetime import datetime
            cycle.start_date = datetime.strptime(data['start_date'], '%Y-%m-%d').date()
        if 'end_date' in data:
            from datetime import datetime
            cycle.end_date = datetime.strptime(data['end_date'], '%Y-%m-%d').date()
        
        if 'is_active' in data:
            cycle.is_active = data['is_active']
        
        # Si se marca como ciclo actual, desmarcar los demás del mismo campus
        if data.get('is_current', False) and not cycle.is_current:
            SchoolCycle.query.filter_by(campus_id=cycle.campus_id, is_current=True).update({'is_current': False})
            cycle.is_current = True
        elif 'is_current' in data:
            cycle.is_current = data['is_current']
        
        db.session.commit()
        
        return jsonify({
            'message': 'Ciclo escolar actualizado exitosamente',
            'cycle': cycle.to_dict()
        })
        
    except ValueError as ve:
        return jsonify({'error': f'Formato de fecha inválido: {str(ve)}'}), 400
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500


@bp.route('/cycles/<int:cycle_id>', methods=['DELETE'])
@jwt_required()
@coordinator_required
def delete_school_cycle(cycle_id):
    """Eliminar un ciclo escolar (soft delete)"""
    try:
        cycle = SchoolCycle.query.get_or_404(cycle_id)
        cycle.is_active = False
        cycle.is_current = False
        db.session.commit()
        
        return jsonify({'message': 'Ciclo escolar desactivado exitosamente'})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500


@bp.route('/cycles/<int:cycle_id>/permanent-delete', methods=['DELETE'])
@jwt_required()
def permanent_delete_cycle(cycle_id):
    """
    Eliminar permanentemente un ciclo escolar (SOLO ADMINISTRADOR)
    Elimina:
    - El ciclo escolar
    - Los grupos del ciclo
    - Los miembros de grupos (GroupMember)
    - Los exámenes asignados a grupos (GroupExam) y sus miembros específicos
    - Los materiales de estudio asignados a grupos (GroupStudyMaterial) y sus miembros específicos
    
    NO elimina y permanece intacto:
    - Los usuarios (candidatos) - solo se elimina su membresía en grupos
    - Los vouchers ya asignados - las certificaciones permanecen
    - Los resultados de exámenes
    """
    from app.models.partner import GroupStudyMaterial, GroupStudyMaterialMember, GroupExamMember
    
    try:
        # Verificar que el usuario sea admin
        user_id = get_jwt_identity()
        user = User.query.get(user_id)
        if not user or user.role not in ['admin', 'developer']:
            return jsonify({'error': 'Acceso denegado. Solo administradores pueden eliminar ciclos permanentemente'}), 403
        
        cycle = SchoolCycle.query.get_or_404(cycle_id)
        cycle_name = cycle.name
        campus_id = cycle.campus_id
        
        # Contadores para el resumen
        stats = {
            'groups_deleted': 0,
            'members_removed': 0,
            'exams_unassigned': 0,
            'materials_unassigned': 0,
        }
        
        # 1. Obtener todos los grupos del ciclo
        groups = CandidateGroup.query.filter_by(school_cycle_id=cycle_id).all()
        
        for group in groups:
            # Contar miembros que se eliminarán
            members_count = GroupMember.query.filter_by(group_id=group.id).count()
            stats['members_removed'] += members_count
            
            # Obtener exámenes asignados y eliminar sus miembros específicos primero
            group_exams = GroupExam.query.filter_by(group_id=group.id).all()
            for ge in group_exams:
                GroupExamMember.query.filter_by(group_exam_id=ge.id).delete()
            stats['exams_unassigned'] += len(group_exams)
            
            # Obtener materiales asignados y eliminar sus miembros específicos primero
            group_materials = GroupStudyMaterial.query.filter_by(group_id=group.id).all()
            for gm in group_materials:
                GroupStudyMaterialMember.query.filter_by(group_study_material_id=gm.id).delete()
            stats['materials_unassigned'] += len(group_materials)
            
            # Eliminar miembros de grupo (los usuarios NO se eliminan, solo la membresía)
            GroupMember.query.filter_by(group_id=group.id).delete()
            
            # Eliminar asignaciones de exámenes a grupos
            GroupExam.query.filter_by(group_id=group.id).delete()
            
            # Eliminar asignaciones de materiales a grupos
            GroupStudyMaterial.query.filter_by(group_id=group.id).delete()
            
            stats['groups_deleted'] += 1
        
        # 2. Eliminar grupos del ciclo
        CandidateGroup.query.filter_by(school_cycle_id=cycle_id).delete()
        
        # 3. Eliminar el ciclo
        db.session.delete(cycle)
        db.session.commit()
        
        return jsonify({
            'message': f'Ciclo escolar "{cycle_name}" eliminado permanentemente',
            'cycle_name': cycle_name,
            'stats': stats
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500


# ============== GRUPOS ==============

@bp.route('/campuses/<int:campus_id>/groups', methods=['GET'])
@jwt_required()
@coordinator_required
def get_groups(campus_id):
    """Listar grupos de un plantel"""
    try:
        campus, error = _verify_campus_access(campus_id, g.current_user)
        if error:
            return error
        
        active_only = request.args.get('active_only', 'true').lower() == 'true'
        cycle_id = request.args.get('cycle_id', type=int)
        include_config = request.args.get('include_config', 'false').lower() == 'true'
        
        # Usar joinedload para cargar la relación campus (necesaria para effective_config)
        query = CandidateGroup.query.options(joinedload(CandidateGroup.campus)).filter_by(campus_id=campus_id)
        
        if active_only:
            query = query.filter(CandidateGroup.is_active == True)
        
        # Filtrar por ciclo escolar si se proporciona
        if cycle_id:
            query = query.filter(CandidateGroup.school_cycle_id == cycle_id)
        
        groups = query.order_by(CandidateGroup.name).all()
        
        # Precio base del campus para comparación
        campus_certification_cost = float(campus.certification_cost) if campus.certification_cost else 0
        
        return jsonify({
            'campus_id': campus_id,
            'campus_name': campus.name,
            'campus_certification_cost': campus_certification_cost,
            'groups': [g.to_dict(include_members=True, include_cycle=True, include_config=include_config) for g in groups],
            'total': len(groups)
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@bp.route('/campuses/<int:campus_id>/groups', methods=['POST'])
@jwt_required()
@coordinator_required
def create_group(campus_id):
    """Crear un nuevo grupo"""
    try:
        campus, error = _verify_campus_access(campus_id, g.current_user)
        if error:
            return error
        data = request.get_json()
        
        if not data.get('name'):
            return jsonify({'error': 'El nombre es requerido'}), 400
        
        # Validar el ciclo escolar si se proporciona
        school_cycle_id = data.get('school_cycle_id')
        if school_cycle_id:
            cycle = SchoolCycle.query.get(school_cycle_id)
            if not cycle or cycle.campus_id != campus_id:
                return jsonify({'error': 'Ciclo escolar no válido para este plantel'}), 400
        
        group = CandidateGroup(
            campus_id=campus_id,
            school_cycle_id=school_cycle_id,
            name=data['name'],
            code=data.get('code'),
            description=data.get('description'),
            start_date=data.get('start_date'),
            end_date=data.get('end_date'),
            is_active=data.get('is_active', True)
        )
        
        db.session.add(group)
        db.session.commit()
        
        return jsonify({
            'message': 'Grupo creado exitosamente',
            'group': group.to_dict(include_cycle=True)
        }), 201
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500


# ======================================================================
# IMPORTANTE: Esta ruta debe ir ANTES de /groups/<int:group_id>
# para evitar que Flask intente interpretar 'list-all' como un int
# ======================================================================
@bp.route('/groups/list-all', methods=['GET'])
@jwt_required()
@coordinator_required
def list_all_groups():
    """Listar todos los grupos para selectores (sin paginación, info mínima)"""
    try:
        from sqlalchemy import func, text
        
        # Una sola query con LEFT JOIN para contar miembros (elimina N+1)
        member_count_sub = db.session.query(
            GroupMember.group_id,
            func.count(GroupMember.id).label('member_count')
        ).filter(
            GroupMember.status == 'active'
        ).group_by(GroupMember.group_id).subquery()
        
        groups = db.session.query(
            CandidateGroup.id,
            CandidateGroup.name,
            Campus.name.label('campus_name'),
            Partner.name.label('partner_name'),
            func.coalesce(member_count_sub.c.member_count, 0).label('current_members')
        ).join(
            Campus, CandidateGroup.campus_id == Campus.id, isouter=True
        ).join(
            Partner, Campus.partner_id == Partner.id, isouter=True
        ).outerjoin(
            member_count_sub, CandidateGroup.id == member_count_sub.c.group_id
        ).filter(
            CandidateGroup.is_active == True
        )
        
        # Multi-tenant: coordinadores solo ven grupos de sus partners
        coord_id = _get_coordinator_filter(g.current_user)
        if coord_id:
            groups = groups.filter(Partner.coordinator_id == coord_id)
        
        groups = groups.order_by(CandidateGroup.name).all()
        
        result = [{
            'id': g.id,
            'name': g.name,
            'campus_name': g.campus_name,
            'partner_name': g.partner_name,
            'current_members': g.current_members
        } for g in groups]
        
        return jsonify({
            'groups': result,
            'total': len(result)
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@bp.route('/groups/search', methods=['GET'])
@jwt_required()
@coordinator_required
def search_groups_paginated():
    """Buscar grupos con paginación server-side, búsqueda, filtros y ordenamiento.
    Optimizado para cientos de miles de registros.
    
    Query params:
    - page (int, default 1)
    - per_page (int, default 150, max 1000)
    - search (str) — busca en nombre de grupo, campus, partner, ciclo escolar
    - campus_id (int) — filtrar por campus
    - active_only (bool, default true)
    - cycle_name (str) — filtrar por nombre de ciclo escolar
    - sort_by (str) — name, member_count, campus_name, partner_name, school_cycle, created_at, is_active
    - sort_dir (str) — asc, desc
    """
    try:
        from sqlalchemy import func, text as sa_text

        page = request.args.get('page', 1, type=int)
        per_page = min(request.args.get('per_page', 150, type=int), 1000)
        search = request.args.get('search', '').strip()
        campus_id = request.args.get('campus_id', type=int)
        partner_id = request.args.get('partner_id', type=int)
        active_only = request.args.get('active_only', 'true').lower() == 'true'
        cycle_name = request.args.get('cycle_name', '').strip()
        status_filter = request.args.get('status', '')  # active, inactive, ''
        sort_by = request.args.get('sort_by', 'name')
        sort_dir = request.args.get('sort_dir', 'asc')

        # Subquery para contar miembros
        member_count_sub = db.session.query(
            GroupMember.group_id,
            func.count(GroupMember.id).label('member_count')
        ).filter(
            GroupMember.status == 'active'
        ).group_by(GroupMember.group_id).subquery()

        query = db.session.query(
            CandidateGroup.id,
            CandidateGroup.name,
            CandidateGroup.description,
            CandidateGroup.is_active,
            CandidateGroup.created_at,
            CandidateGroup.campus_id,
            Campus.name.label('campus_name'),
            Campus.country.label('campus_country'),
            Campus.state_name.label('campus_state'),
            Partner.name.label('partner_name'),
            Partner.id.label('partner_id'),
            SchoolCycle.name.label('school_cycle_name'),
            SchoolCycle.id.label('school_cycle_id'),
            func.coalesce(member_count_sub.c.member_count, 0).label('member_count'),
        ).join(
            Campus, CandidateGroup.campus_id == Campus.id, isouter=True
        ).join(
            Partner, Campus.partner_id == Partner.id, isouter=True
        ).outerjoin(
            SchoolCycle, CandidateGroup.school_cycle_id == SchoolCycle.id
        ).outerjoin(
            member_count_sub, CandidateGroup.id == member_count_sub.c.group_id
        )

        # Multi-tenant: coordinadores solo ven sus propios grupos
        coord_id = _get_coordinator_filter(g.current_user)
        if coord_id:
            query = query.filter(Partner.coordinator_id == coord_id)

        # Filtro por partner
        if partner_id:
            query = query.filter(Partner.id == partner_id)

        # Filtro por campus
        if campus_id:
            query = query.filter(CandidateGroup.campus_id == campus_id)

        # Filtro de estado activo/inactivo
        if status_filter == 'active':
            query = query.filter(CandidateGroup.is_active == True)
        elif status_filter == 'inactive':
            query = query.filter(CandidateGroup.is_active == False)
        elif active_only:
            query = query.filter(CandidateGroup.is_active == True)

        # Filtro por ciclo escolar
        if cycle_name:
            query = query.filter(SchoolCycle.name == cycle_name)

        # Búsqueda textual
        if search:
            search_term = f'%{search}%'
            query = query.filter(
                db.or_(
                    CandidateGroup.name.ilike(search_term),
                    CandidateGroup.description.ilike(search_term),
                    Campus.name.ilike(search_term),
                    Partner.name.ilike(search_term),
                    SchoolCycle.name.ilike(search_term),
                )
            )

        # Ordenamiento
        sort_map = {
            'name': CandidateGroup.name,
            'campus_name': Campus.name,
            'campus_state': Campus.state_name,
            'partner_name': Partner.name,
            'school_cycle': SchoolCycle.name,
            'created_at': CandidateGroup.created_at,
            'is_active': CandidateGroup.is_active,
            'member_count': func.coalesce(member_count_sub.c.member_count, 0),
        }
        sort_col = sort_map.get(sort_by, CandidateGroup.name)
        if sort_dir == 'desc':
            query = query.order_by(sort_col.desc())
        else:
            query = query.order_by(sort_col.asc())

        # Paginación
        total = query.count()
        total_pages = max(1, (total + per_page - 1) // per_page)
        page = min(page, total_pages)
        results = query.offset((page - 1) * per_page).limit(per_page).all()

        # También obtener la lista de ciclos escolares únicos y campuses para filtros
        cycle_names_q = db.session.query(SchoolCycle.name).join(
            CandidateGroup, CandidateGroup.school_cycle_id == SchoolCycle.id
        ).join(
            Campus, CandidateGroup.campus_id == Campus.id
        ).join(
            Partner, Campus.partner_id == Partner.id
        )
        if coord_id:
            cycle_names_q = cycle_names_q.filter(Partner.coordinator_id == coord_id)
        cycle_names = sorted(set(r[0] for r in cycle_names_q.distinct().all() if r[0]))

        groups = [{
            'id': r.id,
            'name': r.name,
            'description': r.description,
            'is_active': r.is_active,
            'created_at': r.created_at.isoformat() if r.created_at else None,
            'campus_id': r.campus_id,
            'campus_name': r.campus_name,
            'campus_country': r.campus_country,
            'campus_state': r.campus_state,
            'partner_name': r.partner_name,
            'partner_id': r.partner_id,
            'school_cycle': {'id': r.school_cycle_id, 'name': r.school_cycle_name} if r.school_cycle_name else None,
            'member_count': r.member_count,
        } for r in results]

        # Obtener partners disponibles para filtro
        partners_q = db.session.query(Partner.id, Partner.name).join(
            Campus, Campus.partner_id == Partner.id
        ).join(
            CandidateGroup, CandidateGroup.campus_id == Campus.id
        )
        if coord_id:
            partners_q = partners_q.filter(Partner.coordinator_id == coord_id)
        available_partners = sorted(
            [{'id': r[0], 'name': r[1]} for r in partners_q.distinct().all()],
            key=lambda x: x['name']
        )

        return jsonify({
            'groups': groups,
            'total': total,
            'page': page,
            'pages': total_pages,
            'per_page': per_page,
            'available_cycles': cycle_names,
            'available_partners': available_partners,
        })

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@bp.route('/groups/<int:group_id>/members/count', methods=['GET'])
@jwt_required()
@coordinator_required
def get_group_members_count(group_id):
    """Obtener solo el conteo y lista de IDs de miembros (endpoint ligero)"""
    try:
        from sqlalchemy import text
        group = CandidateGroup.query.get_or_404(group_id)
        
        status_filter = request.args.get('status', 'active')
        
        result = db.session.execute(text("""
            SELECT gm.user_id
            FROM group_members gm
            WHERE gm.group_id = :group_id AND gm.status = :status
        """), {'group_id': group_id, 'status': status_filter})
        
        member_ids = [str(row[0]) for row in result]
        
        return jsonify({
            'group_id': group_id,
            'count': len(member_ids),
            'member_ids': member_ids
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@bp.route('/groups/<int:group_id>', methods=['GET'])
@jwt_required()
@coordinator_required
def get_group(group_id):
    """Obtener detalle de un grupo"""
    try:
        group, error = _verify_group_access(group_id, g.current_user)
        if error:
            return error
        return jsonify({
            'group': group.to_dict(include_members=True, include_campus=True, include_cycle=True, include_config=True)
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@bp.route('/groups/<int:group_id>', methods=['PUT'])
@jwt_required()
@coordinator_required
def update_group(group_id):
    """Actualizar un grupo"""
    try:
        group, error = _verify_group_access(group_id, g.current_user)
        if error:
            return error
        data = request.get_json()
        
        # Validar el ciclo escolar si se proporciona
        if 'school_cycle_id' in data:
            school_cycle_id = data['school_cycle_id']
            if school_cycle_id:
                cycle = SchoolCycle.query.get(school_cycle_id)
                if not cycle or cycle.campus_id != group.campus_id:
                    return jsonify({'error': 'Ciclo escolar no válido para este plantel'}), 400
            group.school_cycle_id = school_cycle_id
        
        # Actualizar campos
        for field in ['name', 'code', 'description', 'start_date', 'end_date', 'is_active']:
            if field in data:
                setattr(group, field, data[field])
        
        db.session.commit()
        
        return jsonify({
            'message': 'Grupo actualizado exitosamente',
            'group': group.to_dict(include_cycle=True)
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500


@bp.route('/groups/<int:group_id>', methods=['DELETE'])
@jwt_required()
@coordinator_required
def delete_group(group_id):
    """Eliminar un grupo (soft delete)"""
    try:
        group, error = _verify_group_access(group_id, g.current_user)
        if error:
            return error
        group.is_active = False
        db.session.commit()
        
        return jsonify({'message': 'Grupo desactivado exitosamente'})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500


# ============== CONFIGURACIÓN DE GRUPO ==============

@bp.route('/groups/<int:group_id>/config', methods=['GET'])
@jwt_required()
@coordinator_required
def get_group_config(group_id):
    """Obtener configuración del grupo con valores heredados del campus"""
    try:
        group = CandidateGroup.query.get_or_404(group_id)
        campus = group.campus
        
        if not campus:
            return jsonify({'error': 'Campus no encontrado'}), 404
        
        # Configuración del campus (base)
        campus_config = {
            'office_version': campus.office_version,
            'enable_tier_basic': campus.enable_tier_basic,
            'enable_tier_standard': campus.enable_tier_standard,
            'enable_tier_advanced': campus.enable_tier_advanced,
            'enable_digital_badge': campus.enable_digital_badge,
            'enable_partial_evaluations': campus.enable_partial_evaluations,
            'enable_unscheduled_partials': campus.enable_unscheduled_partials,
            'enable_virtual_machines': campus.enable_virtual_machines,
            'enable_online_payments': campus.enable_online_payments,
            'enable_candidate_certificates': campus.enable_candidate_certificates if campus.enable_candidate_certificates is not None else False,
            'require_exam_pin': campus.require_exam_pin or False,
            'enable_session_calendar': campus.enable_session_calendar or False,
            'session_scheduling_mode': campus.session_scheduling_mode or 'leader_only',
            'certification_cost': float(campus.certification_cost) if campus.certification_cost else 0,
            'retake_cost': float(campus.retake_cost) if campus.retake_cost else 0,
            'max_retakes': campus.max_retakes if campus.max_retakes is not None else 0,
            'assignment_validity_months': campus.assignment_validity_months or 12,
        }
        
        # Overrides del grupo
        group_overrides = {
            'office_version_override': group.office_version_override,
            'enable_tier_basic_override': group.enable_tier_basic_override,
            'enable_tier_standard_override': group.enable_tier_standard_override,
            'enable_tier_advanced_override': group.enable_tier_advanced_override,
            'enable_digital_badge_override': group.enable_digital_badge_override,
            'enable_partial_evaluations_override': group.enable_partial_evaluations_override,
            'enable_unscheduled_partials_override': group.enable_unscheduled_partials_override,
            'enable_virtual_machines_override': group.enable_virtual_machines_override,
            'enable_online_payments_override': group.enable_online_payments_override,
            'enable_candidate_certificates_override': group.enable_candidate_certificates_override,
            'require_exam_pin_override': group.require_exam_pin_override,
            'enable_session_calendar_override': group.enable_session_calendar_override,
            'session_scheduling_mode_override': group.session_scheduling_mode_override,
            'certification_cost_override': float(group.certification_cost_override) if group.certification_cost_override is not None else None,
            'retake_cost_override': float(group.retake_cost_override) if group.retake_cost_override is not None else None,
            'max_retakes_override': group.max_retakes_override,
            'assignment_validity_months_override': group.assignment_validity_months_override,
        }
        
        # Configuración efectiva (combinando campus y grupo)
        effective_config = {
            'office_version': group.office_version_override or campus.office_version,
            'enable_tier_basic': group.enable_tier_basic_override if group.enable_tier_basic_override is not None else campus.enable_tier_basic,
            'enable_tier_standard': group.enable_tier_standard_override if group.enable_tier_standard_override is not None else campus.enable_tier_standard,
            'enable_tier_advanced': group.enable_tier_advanced_override if group.enable_tier_advanced_override is not None else campus.enable_tier_advanced,
            'enable_digital_badge': group.enable_digital_badge_override if group.enable_digital_badge_override is not None else campus.enable_digital_badge,
            'enable_partial_evaluations': group.enable_partial_evaluations_override if group.enable_partial_evaluations_override is not None else campus.enable_partial_evaluations,
            'enable_unscheduled_partials': group.enable_unscheduled_partials_override if group.enable_unscheduled_partials_override is not None else campus.enable_unscheduled_partials,
            'enable_virtual_machines': group.enable_virtual_machines_override if group.enable_virtual_machines_override is not None else campus.enable_virtual_machines,
            'enable_online_payments': group.enable_online_payments_override if group.enable_online_payments_override is not None else campus.enable_online_payments,
            'enable_candidate_certificates': group.enable_candidate_certificates_override if group.enable_candidate_certificates_override is not None else (campus.enable_candidate_certificates or False),
            'require_exam_pin': group.require_exam_pin_override if group.require_exam_pin_override is not None else (campus.require_exam_pin or False),
            'enable_session_calendar': group.enable_session_calendar_override if group.enable_session_calendar_override is not None else (campus.enable_session_calendar or False),
            'session_scheduling_mode': group.session_scheduling_mode_override if group.session_scheduling_mode_override else (campus.session_scheduling_mode or 'leader_only'),
            'certification_cost': float(group.certification_cost_override) if group.certification_cost_override is not None else (float(campus.certification_cost) if campus.certification_cost else 0),
            'retake_cost': float(group.retake_cost_override) if group.retake_cost_override is not None else (float(campus.retake_cost) if campus.retake_cost else 0),
            'max_retakes': group.max_retakes_override if group.max_retakes_override is not None else (campus.max_retakes if campus.max_retakes is not None else 0),
            'assignment_validity_months': group.assignment_validity_months_override if group.assignment_validity_months_override is not None else (campus.assignment_validity_months or 12),
        }
        
        # Obtener conteo de candidatos sin CURP/email para advertencias
        # Solo si el grupo tiene habilitados conocer_certificate o digital_badge
        members_without_curp = []
        members_without_email = []
        
        conocer_enabled = effective_config.get('enable_tier_advanced', False)
        badge_enabled = effective_config.get('enable_digital_badge', False)
        
        if conocer_enabled or badge_enabled:
            # Obtener miembros del grupo con sus usuarios
            members = GroupMember.query.filter_by(group_id=group_id, status='active').all()
            for member in members:
                if member.user:
                    if conocer_enabled and not member.user.curp:
                        members_without_curp.append({
                            'id': member.user_id,
                            'name': member.user.full_name if hasattr(member.user, 'full_name') else f"{member.user.name} {member.user.first_surname}"
                        })
                    if badge_enabled and not member.user.email:
                        members_without_email.append({
                            'id': member.user_id,
                            'name': member.user.full_name if hasattr(member.user, 'full_name') else f"{member.user.name} {member.user.first_surname}"
                        })
        
        # Advertencias para el responsable
        warnings = []
        if conocer_enabled and members_without_curp:
            warnings.append({
                'type': 'curp_required_for_conocer',
                'message': f'{len(members_without_curp)} candidato(s) sin CURP no podrán recibir certificado CONOCER',
                'affected_members': members_without_curp
            })
        if badge_enabled and members_without_email:
            warnings.append({
                'type': 'email_required_for_badge',
                'message': f'{len(members_without_email)} candidato(s) sin email no podrán recibir insignia digital',
                'affected_members': members_without_email
            })
        
        # Verificar si el grupo tiene asignaciones (certificaciones asignadas)
        from app.models import GroupExam
        assignment_count = GroupExam.query.filter_by(group_id=group_id, is_active=True).count()
        
        return jsonify({
            'group_id': group.id,
            'group_name': group.name,
            'campus_id': campus.id,
            'campus_name': campus.name,
            'use_custom_config': group.use_custom_config or False,
            'campus_config': campus_config,
            'group_overrides': group_overrides,
            'effective_config': effective_config,
            'warnings': warnings,
            'has_assignments': assignment_count > 0,
            'assignment_count': assignment_count,
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@bp.route('/groups/<int:group_id>/config', methods=['PUT'])
@jwt_required()
@coordinator_required
def update_group_config(group_id):
    """Actualizar configuración del grupo (overrides)"""
    try:
        group = CandidateGroup.query.get_or_404(group_id)
        
        # Bloquear si el grupo ya tiene asignaciones
        from app.models import GroupExam
        if GroupExam.query.filter_by(group_id=group_id, is_active=True).first():
            return jsonify({
                'error': 'La configuración del grupo está bloqueada porque ya tiene certificaciones asignadas. No se puede modificar.'
            }), 403
        
        data = request.get_json()
        
        # Campos booleanos de override
        bool_fields = [
            'enable_tier_basic_override',
            'enable_tier_standard_override',
            'enable_tier_advanced_override',
            'enable_digital_badge_override',
            'enable_partial_evaluations_override',
            'enable_unscheduled_partials_override',
            'enable_virtual_machines_override',
            'enable_online_payments_override',
            'enable_candidate_certificates_override',
            'require_exam_pin_override',
            'enable_session_calendar_override',
        ]
        
        for field in bool_fields:
            if field in data:
                setattr(group, field, data[field])
        
        # Modo de calendario de sesiones
        if 'session_scheduling_mode_override' in data:
            mode = data['session_scheduling_mode_override']
            if mode in ('leader_only', 'candidate_self', None):
                group.session_scheduling_mode_override = mode
        
        # Campo de versión de office
        if 'office_version_override' in data:
            group.office_version_override = data['office_version_override']
        
        # Campos de costos
        if 'certification_cost_override' in data:
            group.certification_cost_override = data['certification_cost_override']
        if 'retake_cost_override' in data:
            group.retake_cost_override = data['retake_cost_override']
        if 'max_retakes_override' in data:
            val = data['max_retakes_override']
            group.max_retakes_override = int(val) if val is not None else None
        
        # Vigencia de asignaciones
        if 'assignment_validity_months_override' in data:
            val = data['assignment_validity_months_override']
            group.assignment_validity_months_override = int(val) if val else None
        
        # Flag de configuración personalizada
        if 'use_custom_config' in data:
            group.use_custom_config = data['use_custom_config']
        
        db.session.commit()
        
        return jsonify({
            'message': 'Configuración del grupo actualizada',
            'group': group.to_dict(include_config=True)
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500


@bp.route('/groups/<int:group_id>/config/reset', methods=['POST'])
@jwt_required()
@coordinator_required
def reset_group_config(group_id):
    """Resetear configuración del grupo a valores del campus"""
    try:
        group = CandidateGroup.query.get_or_404(group_id)
        
        # Bloquear si el grupo ya tiene asignaciones
        from app.models import GroupExam
        if GroupExam.query.filter_by(group_id=group_id, is_active=True).first():
            return jsonify({
                'error': 'La configuración del grupo está bloqueada porque ya tiene certificaciones asignadas. No se puede restablecer.'
            }), 403
        
        # Limpiar todos los overrides
        group.use_custom_config = False
        group.office_version_override = None
        group.enable_tier_basic_override = None
        group.enable_tier_standard_override = None
        group.enable_tier_advanced_override = None
        group.enable_digital_badge_override = None
        group.enable_partial_evaluations_override = None
        group.enable_unscheduled_partials_override = None
        group.enable_virtual_machines_override = None
        group.enable_online_payments_override = None
        group.enable_candidate_certificates_override = None
        group.require_exam_pin_override = None
        group.enable_session_calendar_override = None
        group.session_scheduling_mode_override = None
        group.certification_cost_override = None
        group.retake_cost_override = None
        group.max_retakes_override = None
        group.group_start_date = None
        group.group_end_date = None
        group.assignment_validity_months_override = None
        
        db.session.commit()
        
        return jsonify({
            'message': 'Configuración del grupo restablecida a valores del campus',
            'group': group.to_dict(include_config=True)
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500


# ============== MIEMBROS DE GRUPO ==============

@bp.route('/groups/<int:group_id>/members', methods=['GET'])
@jwt_required()
@coordinator_required
def get_group_members(group_id):
    """Listar miembros de un grupo con paginación server-side, búsqueda y filtros — optimizado para 100K+ registros"""
    try:
        from sqlalchemy import text, func

        group = CandidateGroup.query.get_or_404(group_id)

        # ── Parámetros de paginación ──
        page = request.args.get('page', 1, type=int)
        per_page = min(request.args.get('per_page', 150, type=int), 1000)
        status_filter = request.args.get('status')
        search = request.args.get('search', '').strip()
        search_field = request.args.get('search_field', '')
        sort_by = request.args.get('sort_by', 'name')
        sort_dir = request.args.get('sort_dir', 'asc')

        # Filtros de elegibilidad
        has_email = request.args.get('has_email')
        has_curp = request.args.get('has_curp')

        # ── Construir query base (JOIN con User para filtrar/ordenar en SQL) ──
        query = GroupMember.query.join(User, GroupMember.user_id == User.id).filter(
            GroupMember.group_id == group_id
        )

        if status_filter:
            query = query.filter(GroupMember.status == status_filter)

        # ── Búsqueda textual ──
        if search:
            search_term = f'%{search}%'
            valid_fields = {'name', 'first_surname', 'second_surname', 'email', 'curp', 'username'}
            if search_field and search_field in valid_fields:
                field_map = {
                    'name': User.name,
                    'first_surname': User.first_surname,
                    'second_surname': User.second_surname,
                    'email': User.email,
                    'curp': User.curp,
                    'username': User.username,
                }
                query = query.filter(field_map[search_field].ilike(search_term))
            else:
                query = query.filter(
                    db.or_(
                        User.name.ilike(search_term),
                        User.first_surname.ilike(search_term),
                        User.second_surname.ilike(search_term),
                        User.email.ilike(search_term),
                        User.curp.ilike(search_term),
                        User.username.ilike(search_term),
                    )
                )

        # ── Filtros has_email / has_curp ──
        if has_email == 'yes':
            query = query.filter(User.email.isnot(None), User.email != '')
        elif has_email == 'no':
            query = query.filter(db.or_(User.email.is_(None), User.email == ''))

        if has_curp == 'yes':
            query = query.filter(User.curp.isnot(None), User.curp != '')
        elif has_curp == 'no':
            query = query.filter(db.or_(User.curp.is_(None), User.curp == ''))

        # ── Ordenamiento ──
        sort_options = {
            'name': [User.first_surname, User.name],
            'email': [User.email],
            'curp': [User.curp],
            'username': [User.username],
            'joined': [GroupMember.joined_at],
            'role': [User.role],
        }
        sort_cols = sort_options.get(sort_by, sort_options['name'])
        if sort_dir == 'desc':
            query = query.order_by(*[c.desc() for c in sort_cols])
        else:
            query = query.order_by(*sort_cols)

        # ── Paginación SQL ──
        pagination = query.paginate(page=page, per_page=per_page, max_per_page=1000, error_out=False)
        page_members = pagination.items
        total_all = pagination.total

        # ── IDs de usuarios de la página actual ──
        page_user_ids = [m.user_id for m in page_members]

        # ── Datos de asignaciones (solo para la página) ──
        group_exams = GroupExam.query.filter_by(group_id=group_id, is_active=True).all()
        group_exam_ids = [ge.id for ge in group_exams]

        exam_user_assignments = {}
        if group_exam_ids and page_user_ids:
            result = db.session.execute(text("""
                SELECT gem.user_id, gem.group_exam_id
                FROM group_exam_members gem
                WHERE gem.group_exam_id IN :exam_ids AND gem.user_id IN :user_ids
            """), {
                'exam_ids': tuple(group_exam_ids),
                'user_ids': tuple(page_user_ids),
            })
            for row in result:
                exam_user_assignments.setdefault(row[0], set()).add(row[1])

        exams_for_all = [ge.id for ge in group_exams if ge.assignment_type == 'all']

        materials_result = db.session.execute(text("""
            SELECT gsm.id, gsm.assignment_type
            FROM group_study_materials gsm
            WHERE gsm.group_id = :group_id AND gsm.is_active = 1
        """), {'group_id': group_id})

        materials_for_all_ids = []
        group_material_ids = []
        for row in materials_result:
            group_material_ids.append(row[0])
            if row[1] == 'all':
                materials_for_all_ids.append(row[0])

        material_user_assignments = {}
        if group_material_ids and page_user_ids:
            mat_result = db.session.execute(text("""
                SELECT gsmm.user_id, gsmm.group_study_material_id
                FROM group_study_material_members gsmm
                WHERE gsmm.group_study_material_id IN :mat_ids AND gsmm.user_id IN :user_ids
            """), {
                'mat_ids': tuple(group_material_ids),
                'user_ids': tuple(page_user_ids),
            })
            for row in mat_result:
                material_user_assignments.setdefault(row[0], set()).add(row[1])

        exams_with_materials = set()
        if group_exam_ids:
            exam_mat_result = db.session.execute(text("""
                SELECT DISTINCT gem.group_exam_id
                FROM group_exam_materials gem
                WHERE gem.group_exam_id IN :exam_ids
            """), {'exam_ids': tuple(group_exam_ids)})
            exams_with_materials = {row[0] for row in exam_mat_result}

        # ── Resultados de certificación (solo página) ──
        real_exam_ids = [ge.exam_id for ge in group_exams]
        certification_status_map = {}
        if page_user_ids and real_exam_ids:
            from app.models.result import Result
            results = Result.query.filter(
                Result.user_id.in_(page_user_ids),
                Result.exam_id.in_(real_exam_ids),
            ).all()
            for r in results:
                if r.user_id not in certification_status_map:
                    certification_status_map[r.user_id] = {'status': 0, 'result': 0}
                current = certification_status_map[r.user_id]
                if r.result == 1 and r.status == 1:
                    current['status'] = 1
                    current['result'] = 1
                elif r.status == 0 and current['result'] != 1:
                    current['status'] = 0
                    current['result'] = 0
                elif r.status == 1 and r.result == 0 and current['status'] != 0 and current['result'] != 1:
                    current['status'] = 1
                    current['result'] = 0

        # ── Construir respuesta para la página ──
        members_data = []
        for m in page_members:
            member_dict = m.to_dict(include_user=True)

            user_exam_ids = set()
            if exams_for_all:
                user_exam_ids.update(exams_for_all)
            if m.user_id in exam_user_assignments:
                user_exam_ids.update(exam_user_assignments[m.user_id])

            has_exam = len(user_exam_ids) > 0
            has_material_from_exam = bool(user_exam_ids & exams_with_materials)
            has_direct_material = len(materials_for_all_ids) > 0 or m.user_id in material_user_assignments
            has_material = has_direct_material or has_material_from_exam

            if has_exam and has_material:
                assignment_status = 'exam_and_material'
            elif has_exam:
                assignment_status = 'exam_only'
            elif has_material:
                assignment_status = 'material_only'
            else:
                assignment_status = 'none'

            member_dict['assignment_status'] = assignment_status
            member_dict['has_exam'] = has_exam
            member_dict['has_material'] = has_material

            cert_info = certification_status_map.get(m.user_id)
            if cert_info:
                if cert_info['result'] == 1 and cert_info['status'] == 1:
                    member_dict['certification_status'] = 'certified'
                elif cert_info['status'] == 0:
                    member_dict['certification_status'] = 'in_progress'
                else:
                    member_dict['certification_status'] = 'failed'
            else:
                member_dict['certification_status'] = 'pending'

            user = m.user
            has_curp_flag = bool(user.curp) if user else False
            has_email_flag = bool(user.email) if user else False
            member_dict['eligibility'] = {
                'has_curp': has_curp_flag,
                'has_email': has_email_flag,
                'can_receive_eduit': True,
                'can_receive_certificate': True,
                'can_receive_conocer': has_curp_flag,
                'can_receive_badge': has_email_flag,
            }

            members_data.append(member_dict)

        # ── Resumen de elegibilidad (COUNT queries ligeras sobre TODOS los miembros) ──
        base_count_q = (
            db.session.query(func.count(GroupMember.id))
            .join(User, GroupMember.user_id == User.id)
            .filter(GroupMember.group_id == group_id)
        )
        if status_filter:
            base_count_q = base_count_q.filter(GroupMember.status == status_filter)

        total_members = base_count_q.scalar()
        members_with_curp = base_count_q.filter(User.curp.isnot(None), User.curp != '').scalar()
        members_with_email_count = base_count_q.filter(User.email.isnot(None), User.email != '').scalar()
        members_fully_eligible = base_count_q.filter(
            User.curp.isnot(None), User.curp != '',
            User.email.isnot(None), User.email != '',
        ).scalar()
        members_without_curp = total_members - members_with_curp
        members_without_email = total_members - members_with_email_count

        campus = group.campus
        conocer_enabled = False
        badge_enabled = False
        if campus:
            if group.use_custom_config and group.enable_tier_advanced_override is not None:
                conocer_enabled = group.enable_tier_advanced_override
            else:
                conocer_enabled = campus.enable_tier_advanced or False
            if group.use_custom_config and group.enable_digital_badge_override is not None:
                badge_enabled = group.enable_digital_badge_override
            else:
                badge_enabled = campus.enable_digital_badge or False

        eligibility_summary = {
            'total_members': total_members,
            'fully_eligible': members_fully_eligible,
            'members_with_curp': members_with_curp,
            'members_with_email': members_with_email_count,
            'members_without_curp': members_without_curp,
            'members_without_email': members_without_email,
            'conocer_enabled': conocer_enabled,
            'badge_enabled': badge_enabled,
            'warnings': [],
        }
        if conocer_enabled and members_without_curp > 0:
            eligibility_summary['warnings'].append({
                'type': 'missing_curp',
                'message': f'{members_without_curp} candidato(s) sin CURP no podrán recibir certificado CONOCER',
                'count': members_without_curp,
            })
        if badge_enabled and members_without_email > 0:
            eligibility_summary['warnings'].append({
                'type': 'missing_email',
                'message': f'{members_without_email} candidato(s) sin email no podrán recibir insignia digital',
                'count': members_without_email,
            })

        return jsonify({
            'group_id': group_id,
            'group_name': group.name,
            'members': members_data,
            'total': total_all,
            'pages': pagination.pages,
            'current_page': page,
            'per_page': per_page,
            'eligibility_summary': eligibility_summary,
        })

    except Exception as e:
        import traceback
        return jsonify({'error': str(e), 'traceback': traceback.format_exc()}), 500


@bp.route('/groups/<int:group_id>/campus-responsables', methods=['GET'])
@jwt_required()
@coordinator_required
def get_group_campus_responsables(group_id):
    """Obtener los responsables del plantel al que pertenece el grupo,
    indicando cuáles ya son miembros del grupo."""
    try:
        group = CandidateGroup.query.get_or_404(group_id)
        if not group.campus_id:
            return jsonify({'responsables': [], 'total': 0})

        # Responsables del plantel
        responsables = User.query.filter(
            User.role == 'responsable',
            User.campus_id == group.campus_id,
            User.is_active == True
        ).order_by(User.created_at.asc()).all()

        # IDs ya miembros del grupo
        existing_ids = {m.user_id for m in GroupMember.query.filter_by(group_id=group_id).all()}

        result = []
        for r in responsables:
            result.append({
                'id': r.id,
                'username': r.username,
                'full_name': r.full_name,
                'email': r.email,
                'curp': r.curp,
                'gender': r.gender,
                'date_of_birth': r.date_of_birth.isoformat() if r.date_of_birth else None,
                'is_primary': r.id == Campus.query.get(group.campus_id).responsable_id if group.campus_id else False,
                'is_member': r.id in existing_ids,
                'role': 'responsable',
            })

        return jsonify({
            'responsables': result,
            'total': len(result),
            'campus_id': group.campus_id,
            'campus_name': group.campus.name if group.campus else None,
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@bp.route('/groups/<int:group_id>/members', methods=['POST'])
@jwt_required()
@coordinator_required
def add_group_member(group_id):
    """Agregar un candidato o responsable del plantel al grupo"""
    try:
        group = CandidateGroup.query.get_or_404(group_id)
        data = request.get_json()
        
        user_id = data.get('user_id')
        if not user_id:
            return jsonify({'error': 'El ID del usuario es requerido'}), 400
        
        # Verificar que el usuario existe
        user = User.query.get(user_id)
        if not user:
            return jsonify({'error': 'Usuario no encontrado'}), 404
            
        # Permitir candidatos y responsables del plantel del grupo
        if user.role == 'responsable':
            # Validar que el responsable pertenece al plantel del grupo
            if not group.campus_id or user.campus_id != group.campus_id:
                return jsonify({'error': 'Solo se pueden agregar responsables del plantel al que pertenece el grupo'}), 400
        elif user.role != 'candidato':
            return jsonify({'error': 'Solo se pueden agregar usuarios con rol candidato o responsable del plantel'}), 400
        
        # Verificar que no esté ya en el grupo
        existing = GroupMember.query.filter_by(group_id=group_id, user_id=user_id).first()
        if existing:
            return jsonify({'error': 'El usuario ya es miembro de este grupo'}), 400
        
        member = GroupMember(
            group_id=group_id,
            user_id=user_id,
            status=data.get('status', 'active'),
            notes=data.get('notes')
        )
        
        db.session.add(member)
        db.session.commit()
        
        return jsonify({
            'message': 'Miembro agregado exitosamente',
            'member': member.to_dict(include_user=True)
        }), 201
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500


@bp.route('/groups/<int:group_id>/members/bulk', methods=['POST'])
@jwt_required()
@coordinator_required
def add_group_members_bulk(group_id):
    """Agregar múltiples candidatos al grupo - optimizado para miles de candidatos.
    
    Parámetros opcionales:
    - auto_assign_exam_ids: lista de IDs de GroupExam. Si se envía, los nuevos
      miembros se agregan como GroupExamMember a esas asignaciones 'selected'.
      Para asignaciones 'all', no hace falta (ya están cubiertos).
    """
    try:
        group = CandidateGroup.query.get_or_404(group_id)
        data = request.get_json()
        
        user_ids = data.get('user_ids', [])
        auto_assign_exam_ids = data.get('auto_assign_exam_ids', [])
        
        if not user_ids:
            return jsonify({'error': 'Se requiere al menos un ID de usuario'}), 400
        
        added = []
        errors = []
        
        # Procesar en chunks para evitar queries demasiado grandes
        CHUNK_SIZE = 500
        
        for i in range(0, len(user_ids), CHUNK_SIZE):
            chunk_ids = user_ids[i:i + CHUNK_SIZE]
            
            # Batch: obtener todos los usuarios válidos (candidatos + responsables del plantel)
            valid_users = User.query.filter(
                User.id.in_(chunk_ids),
                db.or_(
                    User.role == 'candidato',
                    db.and_(User.role == 'responsable', User.campus_id == group.campus_id)
                )
            ).all()
            valid_user_map = {u.id: u for u in valid_users}
            
            # Batch: obtener membresías existentes en una sola query
            existing_members = GroupMember.query.filter(
                GroupMember.group_id == group_id,
                GroupMember.user_id.in_(chunk_ids)
            ).all()
            existing_member_ids = {m.user_id for m in existing_members}
            
            for user_id in chunk_ids:
                if user_id not in valid_user_map:
                    errors.append({'user_id': user_id, 'error': 'Usuario no encontrado o no es candidato/responsable del plantel'})
                    continue
                    
                if user_id in existing_member_ids:
                    errors.append({'user_id': user_id, 'error': 'Ya es miembro'})
                    continue
                
                member = GroupMember(
                    group_id=group_id,
                    user_id=user_id,
                    status='active'
                )
                db.session.add(member)
                added.append(user_id)
            
            # Flush cada chunk para liberar memoria
            if added:
                db.session.flush()
        
        # Auto-asignar a exámenes si se solicitó
        auto_assigned_exams = 0
        if added and auto_assign_exam_ids:
            from app.models import GroupExam
            from app.models.partner import GroupExamMember
            
            for ge_id in auto_assign_exam_ids:
                group_exam = GroupExam.query.filter_by(id=ge_id, group_id=group_id, is_active=True).first()
                if not group_exam:
                    continue
                
                # Para asignaciones 'all' no necesitamos agregar GroupExamMember
                if group_exam.assignment_type == 'all':
                    continue
                
                # Obtener miembros ya asignados a este examen
                existing_exam_members = set(
                    m.user_id for m in GroupExamMember.query.filter(
                        GroupExamMember.group_exam_id == ge_id,
                        GroupExamMember.user_id.in_(added)
                    ).all()
                )
                
                count = 0
                for uid in added:
                    if uid not in existing_exam_members:
                        db.session.add(GroupExamMember(group_exam_id=ge_id, user_id=uid))
                        count += 1
                
                if count > 0:
                    auto_assigned_exams += 1
                    db.session.flush()
        
        db.session.commit()
        
        return jsonify({
            'message': f'{len(added)} miembros agregados',
            'added': added,
            'errors': errors,
            'auto_assigned_exams': auto_assigned_exams
        }), 201
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500


@bp.route('/groups/<int:group_id>/members/bulk-assign-by-criteria', methods=['POST'])
@jwt_required()
@coordinator_required
def bulk_assign_by_criteria(group_id):
    """Asignar masivamente candidatos por criterios de búsqueda - optimizado para cientos de miles"""
    try:
        group = CandidateGroup.query.get_or_404(group_id)
        data = request.get_json() or {}
        
        # Mismos filtros que search_candidates_advanced
        search = data.get('search', '')
        search_field = data.get('search_field', '')
        has_group = data.get('has_group')
        gender = data.get('gender')
        state = data.get('state')
        
        query = User.query.filter(
            User.role == 'candidato',
            User.is_active == True
        )
        
        # Multi-tenant
        coord_id = _get_coordinator_filter(g.current_user)
        if coord_id:
            query = query.filter(User.coordinator_id == coord_id)
        
        # Filtro de búsqueda textual
        if search:
            search_term = f'%{search}%'
            if search_field and search_field in ['name', 'first_surname', 'second_surname', 'email', 'curp']:
                field_map = {
                    'name': User.name,
                    'first_surname': User.first_surname,
                    'second_surname': User.second_surname,
                    'email': User.email,
                    'curp': User.curp,
                }
                query = query.filter(field_map[search_field].ilike(search_term))
            else:
                query = query.filter(
                    db.or_(
                        User.name.ilike(search_term),
                        User.first_surname.ilike(search_term),
                        User.second_surname.ilike(search_term),
                        User.email.ilike(search_term),
                        User.curp.ilike(search_term)
                    )
                )
        
        # Filtro por género
        if gender and gender in ['M', 'F', 'O']:
            query = query.filter(User.gender == gender)
        
        # Filtro: tiene grupo / sin grupo
        if has_group == 'yes':
            candidates_with_group = db.session.query(GroupMember.user_id).filter(
                GroupMember.status == 'active'
            ).subquery()
            query = query.filter(User.id.in_(candidates_with_group))
        elif has_group == 'no':
            candidates_with_group = db.session.query(GroupMember.user_id).filter(
                GroupMember.status == 'active'
            ).subquery()
            query = query.filter(~User.id.in_(candidates_with_group))
        
        # Filtro por estado
        if state:
            groups_of_state = db.session.query(CandidateGroup.id).join(Campus).filter(
                Campus.state_name == state
            ).subquery()
            members_of_state = db.session.query(GroupMember.user_id).filter(
                GroupMember.group_id.in_(groups_of_state),
                GroupMember.status == 'active'
            ).subquery()
            query = query.filter(User.id.in_(members_of_state))
        
        # SIEMPRE excluir candidatos que ya están en este grupo
        members_of_group = db.session.query(GroupMember.user_id).filter(
            GroupMember.group_id == group_id
        ).subquery()
        query = query.filter(~User.id.in_(members_of_group))
        
        # Obtener TODOS los IDs que coinciden
        all_user_ids = [row[0] for row in query.with_entities(User.id).all()]
        
        if not all_user_ids:
            return jsonify({
                'message': 'No hay candidatos que coincidan con los criterios',
                'added': 0,
                'skipped': 0,
                'total_matched': 0,
            }), 200
        
        # Asignar en chunks
        added = []
        CHUNK_SIZE = 500
        
        for i in range(0, len(all_user_ids), CHUNK_SIZE):
            chunk_ids = all_user_ids[i:i + CHUNK_SIZE]
            
            existing_members = GroupMember.query.filter(
                GroupMember.group_id == group_id,
                GroupMember.user_id.in_(chunk_ids)
            ).all()
            existing_member_ids = {m.user_id for m in existing_members}
            
            for user_id in chunk_ids:
                if user_id in existing_member_ids:
                    continue
                
                member = GroupMember(
                    group_id=group_id,
                    user_id=user_id,
                    status='active'
                )
                db.session.add(member)
                added.append(user_id)
            
            if added:
                db.session.flush()
        
        db.session.commit()
        
        return jsonify({
            'message': f'{len(added)} candidatos asignados al grupo',
            'added': len(added),
            'skipped': len(all_user_ids) - len(added),
            'total_matched': len(all_user_ids),
        }), 201
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500


@bp.route('/groups/<int:group_id>/members/<int:member_id>', methods=['PUT'])
@jwt_required()
@coordinator_required
def update_group_member(group_id, member_id):
    """Actualizar estado de un miembro"""
    try:
        member = GroupMember.query.filter_by(id=member_id, group_id=group_id).first_or_404()
        data = request.get_json()
        
        if 'status' in data:
            if data['status'] not in ['active', 'suspended']:
                return jsonify({'error': 'Estado no válido. Use: active, suspended'}), 400
            member.status = data['status']
            
        if 'notes' in data:
            member.notes = data['notes']
        
        db.session.commit()
        
        return jsonify({
            'message': 'Miembro actualizado',
            'member': member.to_dict(include_user=True)
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500


@bp.route('/groups/<int:group_id>/members/<int:member_id>', methods=['DELETE'])
@jwt_required()
@coordinator_required
def remove_group_member(group_id, member_id):
    """Eliminar un miembro del grupo (las asignaciones existentes se conservan)"""
    try:
        member = GroupMember.query.filter_by(id=member_id, group_id=group_id).first_or_404()
        
        db.session.delete(member)
        db.session.commit()
        
        return jsonify({'message': 'Miembro eliminado del grupo'})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500


@bp.route('/groups/<int:group_id>/members/<int:member_id>/check-assignments', methods=['GET'])
@jwt_required()
@coordinator_required
def check_member_assignments(group_id, member_id):
    """Verificar si un miembro tiene asignaciones activas en este grupo (exámenes o materiales)"""
    try:
        member = GroupMember.query.filter_by(id=member_id, group_id=group_id).first_or_404()
        user_id = member.user_id
        
        from app.models.partner import GroupStudyMaterial, GroupStudyMaterialMember
        
        # Buscar asignaciones de exámenes en este grupo
        group_exams = GroupExam.query.filter_by(group_id=group_id).all()
        exam_assignments = []
        for ge in group_exams:
            # Si es 'all' todos están asignados, si es 'selected' verificar en GroupExamMember
            if ge.assignment_type == 'all':
                exam_assignments.append({
                    'group_exam_id': ge.id,
                    'exam_name': ge.exam.name if ge.exam else f'Examen #{ge.exam_id}',
                    'assignment_type': 'all'
                })
            else:
                gem = GroupExamMember.query.filter_by(group_exam_id=ge.id, user_id=user_id).first()
                if gem:
                    exam_assignments.append({
                        'group_exam_id': ge.id,
                        'exam_name': ge.exam.name if ge.exam else f'Examen #{ge.exam_id}',
                        'assignment_type': 'selected'
                    })
        
        # Buscar asignaciones de materiales de estudio en este grupo
        group_materials = GroupStudyMaterial.query.filter_by(group_id=group_id).all()
        material_assignments = []
        for gm in group_materials:
            if gm.assignment_type == 'all':
                material_assignments.append({
                    'group_material_id': gm.id,
                    'material_name': gm.study_material.title if gm.study_material else f'Material #{gm.study_material_id}',
                    'assignment_type': 'all'
                })
            else:
                gmm = GroupStudyMaterialMember.query.filter_by(group_study_material_id=gm.id, user_id=user_id).first()
                if gmm:
                    material_assignments.append({
                        'group_material_id': gm.id,
                        'material_name': gm.study_material.title if gm.study_material else f'Material #{gm.study_material_id}',
                        'assignment_type': 'selected'
                    })
        
        has_assignments = len(exam_assignments) > 0 or len(material_assignments) > 0
        
        return jsonify({
            'member_id': member_id,
            'user_id': user_id,
            'has_assignments': has_assignments,
            'exam_assignments': exam_assignments,
            'material_assignments': material_assignments,
            'total_assignments': len(exam_assignments) + len(material_assignments)
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ============== BÚSQUEDA DE CANDIDATOS ==============

@bp.route('/candidates/search', methods=['GET'])
@jwt_required()
@coordinator_required
def search_candidates():
    """Buscar candidatos para agregar a grupos"""
    try:
        search = request.args.get('search', '')
        search_field = request.args.get('search_field', '')  # Campo específico de búsqueda
        exclude_group_id = request.args.get('exclude_group_id', type=int)
        page = request.args.get('page', 1, type=int)
        per_page = request.args.get('per_page', 20, type=int)
        
        query = User.query.filter(
            User.role == 'candidato',
            User.is_active == True
        )
        
        # Multi-tenant: coordinadores solo ven candidatos que les pertenecen
        coord_id = _get_coordinator_filter(g.current_user)
        if coord_id:
            query = query.filter(User.coordinator_id == coord_id)
        
        if search:
            search_term = f'%{search}%'
            # Si hay un campo específico, buscar solo en ese campo
            if search_field and search_field in ['name', 'first_surname', 'second_surname', 'email', 'curp']:
                field_map = {
                    'name': User.name,
                    'first_surname': User.first_surname,
                    'second_surname': User.second_surname,
                    'email': User.email,
                    'curp': User.curp,
                }
                query = query.filter(field_map[search_field].ilike(search_term))
            else:
                # Búsqueda en todos los campos relevantes
                query = query.filter(
                    db.or_(
                        User.name.ilike(search_term),
                        User.first_surname.ilike(search_term),
                        User.second_surname.ilike(search_term),
                        User.email.ilike(search_term),
                        User.curp.ilike(search_term)
                    )
                )
        
        # Excluir candidatos que ya están en un grupo específico
        if exclude_group_id:
            existing_members = db.session.query(GroupMember.user_id).filter(
                GroupMember.group_id == exclude_group_id
            ).subquery()
            query = query.filter(~User.id.in_(existing_members))
        
        query = query.order_by(User.first_surname, User.name)
        
        pagination = query.paginate(page=page, per_page=per_page, max_per_page=1000, error_out=False)
        
        candidates = []
        for user in pagination.items:
            # Obtener información del grupo actual del candidato (si tiene)
            current_membership = GroupMember.query.filter_by(
                user_id=user.id,
                status='active'
            ).first()
            
            group_info = None
            if current_membership and current_membership.group:
                group = current_membership.group
                campus = group.campus
                cycle = group.school_cycle
                partner = campus.partner if campus else None
                
                group_info = {
                    'group_id': group.id,
                    'group_name': group.name,
                    'campus_id': campus.id if campus else None,
                    'campus_name': campus.name if campus else None,
                    'state_name': campus.state_name if campus else None,
                    'city': campus.city if campus else None,
                    'school_cycle_id': cycle.id if cycle else None,
                    'school_cycle_name': cycle.name if cycle else None,
                    'partner_id': partner.id if partner else None,
                    'partner_name': partner.name if partner else None,
                }
            
            candidates.append({
                'id': user.id,
                'email': user.email,
                'name': user.name,
                'first_surname': user.first_surname,
                'second_surname': user.second_surname,
                'full_name': f"{user.name} {user.first_surname} {user.second_surname or ''}".strip(),
                'curp': user.curp,
                'gender': user.gender,
                'created_at': user.created_at.isoformat() if user.created_at else None,
                'current_group': group_info,
            })
        
        return jsonify({
            'candidates': candidates,
            'total': pagination.total,
            'pages': pagination.pages,
            'current_page': page
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ============== PLANTILLA Y CARGA MASIVA DE ASIGNACIÓN ==============

def _resolve_identifiers_to_users(all_identifiers):
    """Resolver una lista de identificadores a objetos User.
    
    Intenta matchear por (en orden de prioridad):
    1. Email (case-insensitive)
    2. CURP (case-insensitive)
    3. Username (case-insensitive)
    4. Nombre completo (nombre + primer_apellido + segundo_apellido, case-insensitive)
    
    Returns:
        dict: {identifier_original: User} para los encontrados
        dict: {identifier_original: [User, ...]} para los ambiguos (múltiples matches por nombre)
    """
    from sqlalchemy import func as sa_func
    
    user_by_identifier = {}
    ambiguous = {}
    resolved = set()
    CHUNK_SIZE = 500
    
    # --- 1. Match por Email (case-insensitive) ---
    for i in range(0, len(all_identifiers), CHUNK_SIZE):
        chunk = all_identifiers[i:i + CHUNK_SIZE]
        chunk_lower = [c.strip().lower() for c in chunk]
        users = User.query.filter(
            sa_func.lower(User.email).in_(chunk_lower),
            User.role == 'candidato',
            User.is_active == True
        ).all()
        email_map = {}
        for u in users:
            if u.email:
                email_map.setdefault(u.email.lower(), u)
        for c in chunk:
            key = c.strip().lower()
            if key in email_map and c not in resolved:
                user_by_identifier[c] = email_map[key]
                resolved.add(c)
    
    # --- 2. Match por CURP (case-insensitive) ---
    remaining = [c for c in all_identifiers if c not in resolved]
    if remaining:
        for i in range(0, len(remaining), CHUNK_SIZE):
            chunk = remaining[i:i + CHUNK_SIZE]
            chunk_upper = [c.strip().upper() for c in chunk]
            users = User.query.filter(
                sa_func.upper(User.curp).in_(chunk_upper),
                User.role == 'candidato',
                User.is_active == True
            ).all()
            curp_map = {}
            for u in users:
                if u.curp:
                    curp_map.setdefault(u.curp.upper(), u)
            for c in chunk:
                key = c.strip().upper()
                if key in curp_map and c not in resolved:
                    user_by_identifier[c] = curp_map[key]
                    resolved.add(c)
    
    # --- 3. Match por Username (case-insensitive) ---
    remaining = [c for c in all_identifiers if c not in resolved]
    if remaining:
        for i in range(0, len(remaining), CHUNK_SIZE):
            chunk = remaining[i:i + CHUNK_SIZE]
            chunk_lower = [c.strip().lower() for c in chunk]
            users = User.query.filter(
                sa_func.lower(User.username).in_(chunk_lower),
                User.role == 'candidato',
                User.is_active == True
            ).all()
            uname_map = {}
            for u in users:
                if u.username:
                    uname_map.setdefault(u.username.lower(), u)
            for c in chunk:
                key = c.strip().lower()
                if key in uname_map and c not in resolved:
                    user_by_identifier[c] = uname_map[key]
                    resolved.add(c)
    
    # --- 4. Match por Nombre Completo (case-insensitive) — BATCH ---
    remaining = [c for c in all_identifiers if c not in resolved]
    if remaining:
        full_name_expr = (
            sa_func.rtrim(sa_func.ltrim(User.name)) + ' ' +
            sa_func.rtrim(sa_func.ltrim(User.first_surname)) +
            sa_func.coalesce(
                ' ' + sa_func.nullif(sa_func.rtrim(sa_func.ltrim(User.second_surname)), ''),
                ''
            )
        )
        # Chunks más pequeños para expresiones complejas
        NAME_CHUNK = 200
        for i in range(0, len(remaining), NAME_CHUNK):
            chunk = remaining[i:i + NAME_CHUNK]
            clean_to_originals = {}
            for c in chunk:
                clean = c.strip()
                if clean and len(clean) >= 3:
                    key = clean.lower()
                    clean_to_originals.setdefault(key, []).append(c)

            clean_values = list(clean_to_originals.keys())
            if not clean_values:
                continue

            users = User.query.filter(
                User.role == 'candidato',
                User.is_active == True,
                sa_func.lower(full_name_expr).in_(clean_values)
            ).all()

            name_to_users = {}
            for u in users:
                fname = f"{u.name} {u.first_surname} {u.second_surname or ''}".strip().lower()
                name_to_users.setdefault(fname, []).append(u)

            for key, originals in clean_to_originals.items():
                matches = name_to_users.get(key, [])
                for orig in originals:
                    if orig in resolved:
                        continue
                    if len(matches) == 1:
                        user_by_identifier[orig] = matches[0]
                        resolved.add(orig)
                    elif len(matches) > 1:
                        ambiguous[orig] = matches
    
    return user_by_identifier, ambiguous


@bp.route('/groups/members/template', methods=['GET'])
@jwt_required()
@coordinator_required
def download_group_members_template():
    """Descargar plantilla Excel para asignación masiva de candidatos a grupo"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from io import BytesIO
    
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Asignación de Candidatos"
        
        # Estilos
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="5B21B6", end_color="5B21B6", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Encabezados
        headers = [
            ("Identificador (Email, CURP, Usuario o Nombre Completo)", 55),
            ("Notas (Opcional)", 30),
        ]
        
        for col, (header, width) in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width
        
        # Ejemplos
        examples = [
            ("candidato@ejemplo.com", "Búsqueda por email"),
            ("ABCD123456HDFRRR00", "Búsqueda por CURP"),
            ("juan.perez", "Búsqueda por nombre de usuario"),
            ("Juan Pérez López", "Búsqueda por nombre completo"),
        ]
        
        example_fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
        for row, (identifier, notes) in enumerate(examples, 2):
            ws.cell(row=row, column=1, value=identifier).fill = example_fill
            ws.cell(row=row, column=2, value=notes).fill = example_fill
        
        # Instrucciones en la hoja 2
        ws_help = wb.create_sheet(title="Instrucciones")
        ws_help.column_dimensions['A'].width = 80
        
        instructions = [
            "INSTRUCCIONES PARA ASIGNACIÓN MASIVA DE CANDIDATOS",
            "",
            "1. En la columna 'Identificador' ingrese UNO de los siguientes datos del candidato:",
            "   - Email (ej: candidato@ejemplo.com)",
            "   - CURP (ej: ABCD123456HDFRRR00)",
            "   - Nombre de usuario (ej: juan.perez)",
            "   - Nombre completo (ej: Juan Pérez López)",
            "",
            "2. El sistema buscará en este orden: email → CURP → usuario → nombre completo.",
            "3. El candidato debe existir previamente en el sistema como usuario activo.",
            "4. El candidato debe tener el rol 'candidato'.",
            "5. Las notas son opcionales y se guardan como referencia.",
            "6. Si el candidato ya está en el grupo, se omitirá.",
            "7. Elimine las filas de ejemplo antes de subir el archivo.",
            "",
            "IMPORTANTE:",
            "- Si un nombre completo coincide con múltiples candidatos, se marcará como ambiguo.",
            "- Para evitar ambigüedades, prefiera usar email, CURP o nombre de usuario.",
            "- Los candidatos deben estar registrados previamente en el sistema.",
        ]
        
        for row, text in enumerate(instructions, 1):
            cell = ws_help.cell(row=row, column=1, value=text)
            if row == 1:
                cell.font = Font(bold=True, size=14)
            elif text.startswith("IMPORTANTE"):
                cell.font = Font(bold=True, color="DC2626")
        
        # Guardar en memoria
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='plantilla_asignacion_candidatos.xlsx'
        )
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@bp.route('/groups/<int:group_id>/members/upload', methods=['POST'])
@jwt_required()
@coordinator_required
def upload_group_members(group_id):
    """Procesar archivo Excel para asignar candidatos al grupo
    
    Parámetros de form-data:
    - file: Archivo Excel con candidatos
    - mode: 'move' (mover de grupo anterior) o 'add' (agregar, puede estar en múltiples grupos)
    - resolutions: JSON string con resoluciones de ambigüedades
      Formato: [{"identifier": "...", "user_id": "..."}, ...]
    """
    from openpyxl import load_workbook
    import io
    import json

    MAX_ROWS = 50000
    MAX_FILE_MB = 20
    
    try:
        group = CandidateGroup.query.get_or_404(group_id)
        
        if 'file' not in request.files:
            return jsonify({'error': 'No se envió ningún archivo'}), 400
        
        file = request.files['file']
        if not file.filename.endswith(('.xlsx', '.xls')):
            return jsonify({'error': 'El archivo debe ser Excel (.xlsx o .xls)'}), 400

        # File size limit
        file.seek(0, 2)
        file_size = file.tell()
        file.seek(0)
        if file_size > MAX_FILE_MB * 1024 * 1024:
            return jsonify({'error': f'Archivo excede {MAX_FILE_MB}MB'}), 400
        
        # Modo de asignación: 'move' o 'add'
        mode = request.form.get('mode', 'add')

        # Parse ambiguity resolutions
        resolutions_map = {}  # identifier -> user_id
        resolutions_raw = request.form.get('resolutions', '')
        if resolutions_raw:
            try:
                resolutions_list = json.loads(resolutions_raw)
                for r in resolutions_list:
                    if r.get('identifier') and r.get('user_id'):
                        resolutions_map[r['identifier']] = r['user_id']
            except (json.JSONDecodeError, TypeError):
                pass
        
        # Cargar el archivo Excel
        wb = load_workbook(io.BytesIO(file.read()))
        ws = wb.active
        
        added = []
        moved = []
        errors = []
        current_count = GroupMember.query.filter_by(group_id=group_id, status='active').count()
        
        # Recolectar todos los identificadores primero
        all_identifiers = []
        row_data = []
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if len(row_data) >= MAX_ROWS:
                return jsonify({'error': f'Límite excedido: máximo {MAX_ROWS:,} filas'}), 400
            identifier = row[0] if len(row) > 0 else None
            notes = row[1] if len(row) > 1 else None
            if not identifier:
                continue
            identifier = str(identifier).strip()
            all_identifiers.append(identifier)
            row_data.append((row_num, identifier, notes))
        
        # Resolver identificadores: email → CURP → username → nombre completo
        user_by_identifier, ambiguous = _resolve_identifiers_to_users(all_identifiers)

        # Apply ambiguity resolutions from frontend
        for identifier, user_id in resolutions_map.items():
            if identifier in ambiguous:
                # Find the user in the ambiguous matches
                resolved_user = None
                for u in ambiguous[identifier]:
                    if u.id == user_id:
                        resolved_user = u
                        break
                if not resolved_user:
                    resolved_user = User.query.get(user_id)
                if resolved_user:
                    user_by_identifier[identifier] = resolved_user
                    del ambiguous[identifier]
        
        # Batch: obtener membresías existentes en ESTE grupo
        CHUNK_SIZE = 500
        found_user_ids = list(set(u.id for u in user_by_identifier.values()))
        existing_in_target_map = {}
        if found_user_ids:
            for i in range(0, len(found_user_ids), CHUNK_SIZE):
                chunk = found_user_ids[i:i + CHUNK_SIZE]
                existing = GroupMember.query.filter(
                    GroupMember.group_id == group_id,
                    GroupMember.user_id.in_(chunk)
                ).all()
                for m in existing:
                    existing_in_target_map[m.user_id] = m
        
        # Batch: obtener membresías en OTROS grupos (solo si modo move)
        existing_other_map = {}
        if mode == 'move' and found_user_ids:
            for i in range(0, len(found_user_ids), CHUNK_SIZE):
                chunk = found_user_ids[i:i + CHUNK_SIZE]
                others = GroupMember.query.filter(
                    GroupMember.user_id.in_(chunk),
                    GroupMember.group_id != group_id,
                    GroupMember.status == 'active'
                ).all()
                for m in others:
                    if m.user_id not in existing_other_map:
                        existing_other_map[m.user_id] = m
        
        # Procesar cada fila con los datos pre-cargados
        batch_count = 0
        for row_num, identifier, notes in row_data:
            # Verificar si es ambiguo (múltiples matches por nombre)
            if identifier in ambiguous:
                matches = ambiguous[identifier]
                names = [f"{u.name} {u.first_surname} ({u.email or u.curp or u.username})" for u in matches[:3]]
                errors.append({
                    'identifier': identifier,
                    'error': f'Nombre ambiguo: {len(matches)} candidatos coinciden. Ej: {", ".join(names)}'
                })
                continue
            
            user = user_by_identifier.get(identifier)
            
            if not user:
                errors.append({
                    'identifier': identifier,
                    'error': 'Candidato no encontrado. Verifique que el email, CURP, usuario o nombre completo sea correcto'
                })
                continue
            
            # Verificar si ya es miembro de ESTE grupo
            existing_in_target = existing_in_target_map.get(user.id)
            if existing_in_target:
                if existing_in_target.status == 'active':
                    errors.append({
                        'identifier': identifier,
                        'error': 'Ya es miembro del grupo destino'
                    })
                    continue
                else:
                    # Reactivar membresía existente
                    existing_in_target.status = 'active'
                    existing_in_target.notes = str(notes) if notes else existing_in_target.notes
                    added.append(identifier)
                    current_count += 1
                    batch_count += 1
                    if batch_count >= CHUNK_SIZE:
                        db.session.commit()
                        batch_count = 0
                    continue
            
            # Verificar si está en otro grupo
            existing_other = existing_other_map.get(user.id)
            
            if existing_other and mode == 'move':
                # Mover: desactivar membresía anterior
                previous_group_name = existing_other.group.name if existing_other.group else 'Grupo desconocido'
                existing_other.status = 'removed'
                moved.append({
                    'identifier': identifier,
                    'from_group': previous_group_name
                })
            
            # Crear nueva membresía
            member = GroupMember(
                group_id=group_id,
                user_id=user.id,
                status='active',
                notes=str(notes) if notes else None
            )
            db.session.add(member)
            if not (existing_other and mode == 'move'):
                added.append(identifier)
            current_count += 1
            batch_count += 1
            if batch_count >= CHUNK_SIZE:
                db.session.commit()
                batch_count = 0
        
        if batch_count > 0:
            db.session.commit()
        
        result_message = f'{len(added)} candidato(s) agregado(s)'
        if moved:
            result_message += f', {len(moved)} movido(s) de otros grupos'
        
        return jsonify({
            'message': result_message,
            'added': added,
            'moved': moved,
            'errors': errors,
            'total_processed': len(added) + len(moved) + len(errors),
            'mode': mode
        }), 201
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500


# ============== DASHBOARD / ESTADÍSTICAS ==============

@bp.route('/dashboard', methods=['GET'])
@jwt_required()
@coordinator_required
def get_dashboard():
    """Obtener estadísticas generales para coordinador"""
    try:
        coord_id = _get_coordinator_filter(g.current_user)
        
        # Base queries con scoping por coordinador
        partner_query = Partner.query.filter_by(is_active=True)
        if coord_id:
            partner_query = partner_query.filter(Partner.coordinator_id == coord_id)
        
        total_partners = partner_query.count()
        
        # Obtener IDs de partners del coordinador para filtrar campuses/grupos
        if coord_id:
            partner_ids = [p.id for p in partner_query.with_entities(Partner.id).all()]
            campus_query = Campus.query.filter(Campus.partner_id.in_(partner_ids), Campus.is_active == True) if partner_ids else Campus.query.filter(db.false())
            campus_ids = [c.id for c in campus_query.with_entities(Campus.id).all()]
            group_query = CandidateGroup.query.filter(CandidateGroup.campus_id.in_(campus_ids), CandidateGroup.is_active == True) if campus_ids else CandidateGroup.query.filter(db.false())
            group_ids = [gr.id for gr in group_query.with_entities(CandidateGroup.id).all()]
        else:
            campus_query = Campus.query.filter_by(is_active=True)
            group_query = CandidateGroup.query.filter_by(is_active=True)
            campus_ids = None
            group_ids = None
        
        total_campuses = campus_query.count() if coord_id else Campus.query.filter_by(is_active=True).count()
        total_groups = group_query.count() if coord_id else CandidateGroup.query.filter_by(is_active=True).count()
        
        if coord_id and group_ids:
            total_members = GroupMember.query.filter(GroupMember.group_id.in_(group_ids), GroupMember.status == 'active').count()
        elif coord_id:
            total_members = 0
        else:
            total_members = GroupMember.query.filter_by(status='active').count()
        
        # Partners por estado (via campus states)
        if coord_id and partner_ids:
            partners_by_state = db.session.query(
                Campus.state_name,
                db.func.count(db.func.distinct(Campus.partner_id))
            ).filter(
                Campus.partner_id.in_(partner_ids),
                Campus.is_active == True,
                Campus.state_name.isnot(None)
            ).group_by(Campus.state_name).all()
        elif coord_id:
            partners_by_state = []
        else:
            partners_by_state = db.session.query(
                PartnerStatePresence.state_name,
                db.func.count(PartnerStatePresence.id)
            ).filter(
                PartnerStatePresence.is_active == True
            ).group_by(PartnerStatePresence.state_name).all()
        
        # Últimos grupos creados
        if coord_id and campus_ids:
            recent_groups = CandidateGroup.query.filter(
                CandidateGroup.campus_id.in_(campus_ids),
                CandidateGroup.is_active == True
            ).order_by(CandidateGroup.created_at.desc()).limit(5).all()
        elif coord_id:
            recent_groups = []
        else:
            recent_groups = CandidateGroup.query.filter_by(is_active=True).order_by(
                CandidateGroup.created_at.desc()
            ).limit(5).all()
        
        return jsonify({
            'stats': {
                'total_partners': total_partners,
                'total_campuses': total_campuses,
                'total_groups': total_groups,
                'total_members': total_members
            },
            'partners_by_state': [
                {'state': state, 'count': count} 
                for state, count in partners_by_state
            ],
            'recent_groups': [g.to_dict(include_campus=True) for g in recent_groups]
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ============== ASOCIACIÓN USUARIO-PARTNER ==============

@bp.route('/<int:partner_id>/users', methods=['GET'])
@jwt_required()
@coordinator_required
def get_partner_users(partner_id):
    """Obtener usuarios (candidatos) asociados a un partner"""
    try:
        partner, error = _verify_partner_access(partner_id, g.current_user)
        if error:
            return error
        
        page = request.args.get('page', 1, type=int)
        per_page = request.args.get('per_page', 20, type=int)
        search = request.args.get('search', '')
        
        query = partner.users
        
        if search:
            search_term = f'%{search}%'
            query = query.filter(
                db.or_(
                    User.name.ilike(search_term),
                    User.email.ilike(search_term),
                    User.curp.ilike(search_term)
                )
            )
        
        pagination = query.paginate(page=page, per_page=per_page, max_per_page=1000, error_out=False)
        
        return jsonify({
            'partner_id': partner_id,
            'partner_name': partner.name,
            'users': [u.to_dict(include_private=True) for u in pagination.items],
            'total': pagination.total,
            'pages': pagination.pages,
            'current_page': page
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@bp.route('/<int:partner_id>/users/<string:user_id>', methods=['POST'])
@jwt_required()
@coordinator_required
def add_user_to_partner(partner_id, user_id):
    """Asociar un usuario a un partner"""
    try:
        partner = Partner.query.get_or_404(partner_id)
        user = User.query.get_or_404(user_id)
        
        # Verificar si ya está asociado
        if partner.users.filter_by(id=user_id).first():
            return jsonify({'error': 'El usuario ya está asociado a este partner'}), 400
        
        partner.users.append(user)
        db.session.commit()
        
        return jsonify({
            'message': f'Usuario {user.full_name} asociado a {partner.name}',
            'user': user.to_dict(include_partners=True)
        }), 201
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500


@bp.route('/<int:partner_id>/users/<string:user_id>', methods=['DELETE'])
@jwt_required()
@coordinator_required
def remove_user_from_partner(partner_id, user_id):
    """Desasociar un usuario de un partner"""
    try:
        partner = Partner.query.get_or_404(partner_id)
        user = User.query.get_or_404(user_id)
        
        # Verificar si está asociado
        if not partner.users.filter_by(id=user_id).first():
            return jsonify({'error': 'El usuario no está asociado a este partner'}), 400
        
        partner.users.remove(user)
        db.session.commit()
        
        return jsonify({
            'message': f'Usuario {user.full_name} desasociado de {partner.name}'
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500


@bp.route('/users/<string:user_id>/partners', methods=['GET'])
@jwt_required()
@coordinator_required
def get_user_partners(user_id):
    """Obtener los partners a los que está asociado un usuario (solo para coordinadores)
    NOTA: Los coordinadores solo pueden ver la información del usuario en sus propios partners,
    no pueden ver a qué otros partners está asociado el usuario.
    """
    try:
        user = User.query.get_or_404(user_id)
        current_user = g.current_user
        
        partners_data = []
        for partner in user.partners.all():
            # Si el coordinador no es admin, solo puede ver sus propios partners
            # (esto requeriría una relación coordinator-partner, por ahora todos los coordinadores ven todos)
            partner_dict = partner.to_dict(include_states=True, include_campuses=True)
            # Obtener grupos del usuario en este partner
            user_groups = []
            for campus in partner.campuses.all():
                for group in campus.groups.all():
                    membership = GroupMember.query.filter_by(
                        group_id=group.id,
                        user_id=user_id
                    ).first()
                    if membership:
                        user_groups.append({
                            'group': group.to_dict(),
                            'campus': campus.to_dict(),
                            'membership_status': membership.status,
                            'joined_at': membership.joined_at.isoformat() if membership.joined_at else None
                        })
            partner_dict['user_groups'] = user_groups
            partners_data.append(partner_dict)
        
        return jsonify({
            'user_id': user_id,
            'user_name': user.full_name,
            'partners': partners_data,
            'total': len(partners_data)
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@bp.route('/users/<string:user_id>/partners', methods=['POST'])
@jwt_required()
@coordinator_required
def set_user_partners(user_id):
    """Establecer los partners de un usuario (reemplaza los existentes)"""
    try:
        user = User.query.get_or_404(user_id)
        data = request.get_json()
        
        partner_ids = data.get('partner_ids', [])
        
        # Limpiar partners existentes
        user.partners = []
        
        # Agregar nuevos partners
        for pid in partner_ids:
            partner = Partner.query.get(pid)
            if partner:
                user.partners.append(partner)
        
        db.session.commit()
        
        return jsonify({
            'message': 'Partners actualizados exitosamente',
            'user': user.to_dict(include_partners=True)
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500


# ============== ENDPOINTS PARA CANDIDATOS (SUS PROPIOS PARTNERS) ==============

@bp.route('/my-partners', methods=['GET'])
@jwt_required()
def get_my_partners():
    """Obtener los partners a los que está ligado el candidato actual"""
    try:
        user_id = get_jwt_identity()
        user = User.query.get_or_404(user_id)
        
        partners_data = []
        for partner in user.partners.all():
            partner_dict = {
                'id': partner.id,
                'name': partner.name,
                'logo_url': partner.logo_url,
                'email': partner.email,
                'phone': partner.phone,
                'website': partner.website,
            }
            
            # Obtener los estados donde el partner tiene presencia
            partner_dict['states'] = [p.state_name for p in partner.state_presences.filter_by(is_active=True).all()]
            
            # Obtener grupos del usuario en este partner
            user_groups = []
            for campus in partner.campuses.filter_by(is_active=True).all():
                for group in campus.groups.filter_by(is_active=True).all():
                    membership = GroupMember.query.filter_by(
                        group_id=group.id,
                        user_id=user_id,
                        status='active'
                    ).first()
                    if membership:
                        user_groups.append({
                            'group_id': group.id,
                            'group_name': group.name,
                            'campus_id': campus.id,
                            'campus_name': campus.name,
                            'campus_city': campus.city,
                            'state_name': campus.state_name,
                            'joined_at': membership.joined_at.isoformat() if membership.joined_at else None
                        })
            partner_dict['my_groups'] = user_groups
            partners_data.append(partner_dict)
        
        return jsonify({
            'partners': partners_data,
            'total': len(partners_data)
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@bp.route('/available', methods=['GET'])
@jwt_required()
def get_available_partners():
    """Obtener lista de partners disponibles para que un candidato se ligue"""
    try:
        user_id = get_jwt_identity()
        user = User.query.get_or_404(user_id)
        
        # Obtener IDs de partners a los que ya está ligado
        current_partner_ids = [p.id for p in user.partners.all()]
        
        # Obtener todos los partners activos
        partners = Partner.query.filter_by(is_active=True).order_by(Partner.name).all()
        
        partners_data = []
        for partner in partners:
            partners_data.append({
                'id': partner.id,
                'name': partner.name,
                'logo_url': partner.logo_url,
                'is_linked': partner.id in current_partner_ids,
                'states': [p.state_name for p in partner.state_presences.filter_by(is_active=True).all()]
            })
        
        return jsonify({
            'partners': partners_data,
            'total': len(partners_data)
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@bp.route('/my-partners/<int:partner_id>', methods=['POST'])
@jwt_required()
def link_to_partner(partner_id):
    """Ligarse a un partner como candidato"""
    try:
        user_id = get_jwt_identity()
        user = User.query.get_or_404(user_id)
        partner = Partner.query.get_or_404(partner_id)
        
        if not partner.is_active:
            return jsonify({'error': 'Este partner no está activo'}), 400
        
        # Verificar si ya está ligado
        if user.partners.filter_by(id=partner_id).first():
            return jsonify({'error': 'Ya estás ligado a este partner'}), 400
        
        user.partners.append(partner)
        db.session.commit()
        
        return jsonify({
            'message': f'Te has ligado exitosamente a {partner.name}',
            'partner': {
                'id': partner.id,
                'name': partner.name,
                'logo_url': partner.logo_url
            }
        }), 201
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500


@bp.route('/my-partners/<int:partner_id>', methods=['DELETE'])
@jwt_required()
def unlink_from_partner(partner_id):
    """Desligarse de un partner como candidato"""
    try:
        user_id = get_jwt_identity()
        user = User.query.get_or_404(user_id)
        partner = Partner.query.get_or_404(partner_id)
        
        # Verificar si está ligado
        if not user.partners.filter_by(id=partner_id).first():
            return jsonify({'error': 'No estás ligado a este partner'}), 400
        
        # Verificar si tiene grupos activos con este partner
        has_active_groups = False
        for campus in partner.campuses.all():
            for group in campus.groups.all():
                membership = GroupMember.query.filter_by(
                    group_id=group.id,
                    user_id=user_id,
                    status='active'
                ).first()
                if membership:
                    has_active_groups = True
                    break
            if has_active_groups:
                break
        
        if has_active_groups:
            return jsonify({
                'error': 'No puedes desligarte de este partner mientras tengas grupos activos. Contacta al coordinador.'
            }), 400
        
        user.partners.remove(partner)
        db.session.commit()
        
        return jsonify({
            'message': f'Te has desligado de {partner.name}'
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500


# ============== EXÁMENES ASIGNADOS A GRUPOS ==============

@bp.route('/groups/<int:group_id>/exams', methods=['GET'])
@jwt_required()
@coordinator_required
def get_group_exams(group_id):
    """Listar exámenes asignados a un grupo"""
    try:
        from app.models import GroupExam, Exam
        from app.models.study_content import StudyMaterial
        
        group = CandidateGroup.query.get_or_404(group_id)
        
        group_exams = GroupExam.query.filter_by(group_id=group_id, is_active=True).all()
        
        exams_data = []
        for ge in group_exams:
            exam_data = ge.to_dict(include_exam=True, include_materials=True)
            exams_data.append(exam_data)
        
        return jsonify({
            'group_id': group_id,
            'group_name': group.name,
            'assigned_exams': exams_data,
            'total': len(exams_data)
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@bp.route('/groups/<int:group_id>/exams/<int:exam_id>/detail', methods=['GET'])
@jwt_required()
@coordinator_required
def get_group_exam_detail(group_id, exam_id):
    """Detalle completo de una asignación de examen a grupo.
    Incluye: config del plantel, config del grupo, config del examen,
    ECM asignado, materiales de estudio, resumen de miembros y resultados.
    """
    try:
        from app.models import GroupExam, Exam, Result
        from app.models.partner import EcmCandidateAssignment, EcmRetake, GroupExamMember, GroupExamMaterial
        from app.models.study_content import StudyMaterial
        from sqlalchemy import func
        
        group = CandidateGroup.query.get_or_404(group_id)
        group_exam = GroupExam.query.filter_by(
            group_id=group_id, exam_id=exam_id, is_active=True
        ).first_or_404()
        
        campus = group.campus
        exam = group_exam.exam
        
        # --- Datos base de la asignación ---
        assignment_data = group_exam.to_dict(include_exam=True, include_materials=True, include_members=True)
        
        # --- Quién asignó ---
        assigned_by_info = None
        if group_exam.assigned_by_id:
            from app.models import User
            assigner = User.query.get(group_exam.assigned_by_id)
            if assigner:
                assigned_by_info = {
                    'id': assigner.id,
                    'name': assigner.name,
                    'first_surname': assigner.first_surname or '',
                    'full_name': assigner.full_name,
                    'email': assigner.email,
                }
        
        # --- ECM info ---
        ecm_info = None
        if exam and exam.competency_standard_id:
            cs = exam.competency_standard
            if cs:
                ecm_info = {
                    'id': cs.id,
                    'code': cs.code,
                    'name': cs.name,
                    'description': getattr(cs, 'description', None),
                    'logo_url': cs.logo_url if hasattr(cs, 'logo_url') else None,
                    'brand_id': cs.brand_id if hasattr(cs, 'brand_id') else None,
                    'brand_name': cs.brand.name if hasattr(cs, 'brand') and cs.brand else None,
                    'brand_logo_url': cs.brand.logo_url if hasattr(cs, 'brand') and cs.brand and hasattr(cs.brand, 'logo_url') else None,
                }
        
        # --- Examen info completa ---
        exam_info = None
        if exam:
            exam_info = {
                'id': exam.id,
                'name': exam.name,
                'version': exam.version,
                'standard': getattr(exam, 'standard', None),
                'description': exam.description,
                'duration_minutes': exam.duration_minutes,
                'passing_score': exam.passing_score,
                'is_published': exam.is_published,
                'total_questions': getattr(exam, 'total_questions', None),
                'total_exercises': getattr(exam, 'total_exercises', None),
            }
        
        # --- Miembros y estadísticas ---
        # Total de miembros asignados
        if group_exam.assignment_type == 'selected':
            total_members = GroupExamMember.query.filter_by(group_exam_id=group_exam.id).count()
        else:
            total_members = group.members.count()
        
        # Asignaciones ECM vinculadas
        ecm_assignments_count = 0
        ecm_stats = {}
        if exam and exam.competency_standard_id:
            ecm_query = EcmCandidateAssignment.query.filter_by(
                group_exam_id=group_exam.id
            )
            ecm_assignments_count = ecm_query.count()
            
            # Stats de vigencia
            expired_count = sum(1 for a in ecm_query.all() if a.is_expired)
            
            # Retomas
            retakes_query = db.session.query(func.count(EcmRetake.id)).filter(
                EcmRetake.assignment_id.in_(
                    db.session.query(EcmCandidateAssignment.id).filter_by(group_exam_id=group_exam.id)
                )
            ).scalar() or 0
            
            ecm_stats = {
                'total_assignments': ecm_assignments_count,
                'expired_count': expired_count,
                'active_count': ecm_assignments_count - expired_count,
                'total_retakes': retakes_query,
            }
        
        # Resultados de examen
        results_count = 0
        passed_count = 0
        avg_score = None
        try:
            results = db.session.query(
                func.count(Result.id),
                func.sum(db.case((Result.score >= (group_exam.passing_score or exam.passing_score or 70), 1), else_=0)),
                func.avg(Result.score)
            ).filter(
                Result.exam_id == exam_id,
                Result.is_active == True
            ).first()
            if results:
                results_count = results[0] or 0
                passed_count = results[1] or 0
                avg_score = round(float(results[2]), 1) if results[2] else None
        except Exception:
            pass
        
        # --- Config del plantel ---
        campus_config = {
            'id': campus.id,
            'name': campus.name,
            'code': campus.code,
            'office_version': campus.office_version or 'office_365',
            'enable_tier_basic': campus.enable_tier_basic if campus.enable_tier_basic is not None else True,
            'enable_tier_standard': campus.enable_tier_standard if campus.enable_tier_standard is not None else True,
            'enable_tier_advanced': campus.enable_tier_advanced if campus.enable_tier_advanced is not None else False,
            'enable_digital_badge': campus.enable_digital_badge or False,
            'enable_partial_evaluations': campus.enable_partial_evaluations or False,
            'enable_unscheduled_partials': campus.enable_unscheduled_partials or False,
            'enable_virtual_machines': campus.enable_virtual_machines or False,
            'enable_online_payments': campus.enable_online_payments or False,
            'enable_candidate_certificates': campus.enable_candidate_certificates or False,
            'require_exam_pin': campus.require_exam_pin or False,
            'enable_session_calendar': campus.enable_session_calendar or False,
            'session_scheduling_mode': campus.session_scheduling_mode or 'leader_only',
            'certification_cost': float(campus.certification_cost) if campus.certification_cost else 0,
            'retake_cost': float(campus.retake_cost) if campus.retake_cost else 0,
            'max_retakes': campus.max_retakes if campus.max_retakes is not None else 0,
            'assignment_validity_months': campus.assignment_validity_months or 12,
        }
        
        # --- Config del grupo (overrides + effective) ---
        group_config = group.to_dict(include_config=True)
        
        # --- Materiales personalizados ---
        has_custom_materials = GroupExamMaterial.query.filter_by(group_exam_id=group_exam.id).count() > 0
        
        # --- Enrich assigned_members with assignment_number and ecm_assignment_id ---
        if assignment_data.get('assigned_members'):
            member_user_ids = [m['user_id'] for m in assignment_data['assigned_members']]
            if member_user_ids and exam and exam.competency_standard_id:
                from sqlalchemy import text as sa_text
                uid_str = ','.join(f"'{uid}'" for uid in member_user_ids)
                eca_rows = db.session.execute(sa_text(
                    f"SELECT id, user_id, assignment_number "
                    f"FROM ecm_candidate_assignments "
                    f"WHERE user_id IN ({uid_str}) AND competency_standard_id = :cs_id"
                ), {'cs_id': exam.competency_standard_id}).fetchall()
                eca_map = {r.user_id: {'assignment_number': r.assignment_number, 'ecm_assignment_id': r.id} for r in eca_rows}
                for m in assignment_data['assigned_members']:
                    eca_info = eca_map.get(m['user_id'], {})
                    m['assignment_number'] = eca_info.get('assignment_number')
                    m['ecm_assignment_id'] = eca_info.get('ecm_assignment_id')
        
        return jsonify({
            'assignment': assignment_data,
            'assigned_by': assigned_by_info,
            'ecm': ecm_info,
            'exam': exam_info,
            'campus_config': campus_config,
            'group': {
                'id': group.id,
                'name': group.name,
                'code': group.code,
                'member_count': total_members,
                'use_custom_config': group_config.get('use_custom_config', False),
                'config': group_config.get('config', {}),
                'effective_config': group_config.get('effective_config', {}),
            },
            'members_summary': {
                'total': total_members,
                'assignment_type': group_exam.assignment_type or 'all',
            },
            'ecm_stats': ecm_stats,
            'results_summary': {
                'total_results': results_count,
                'passed': passed_count,
                'failed': results_count - passed_count,
                'avg_score': avg_score,
                'pass_rate': round((passed_count / results_count * 100), 1) if results_count > 0 else None,
            },
            'has_custom_materials': has_custom_materials,
        })
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@bp.route('/groups/<int:group_id>/assignment-cost-preview', methods=['POST'])
@jwt_required()
@coordinator_required
def assignment_cost_preview(group_id):
    """Calcular el desglose de costo de una asignación antes de confirmar
    
    Body:
    - exam_id: ID del examen (para verificar ECMs ya asignados)
    - assignment_type: 'all' | 'selected'
    - member_ids: lista de user_ids (solo si selected)
    
    Retorna el costo unitario, cantidad de unidades, total y saldo actual
    """
    try:
        group = CandidateGroup.query.get_or_404(group_id)
        data = request.get_json()
        
        assignment_type = data.get('assignment_type', 'all')
        member_ids = data.get('member_ids', [])
        exam_id = data.get('exam_id')
        
        # Calcular lista de usuarios objetivo
        if assignment_type == 'selected':
            target_user_ids = member_ids
        else:
            target_user_ids = [m.user_id for m in group.members.all()]
        
        units = len(target_user_ids)
        if units == 0:
            return jsonify({'error': 'No hay candidatos para asignar'}), 400
        
        # Obtener costo unitario efectivo (grupo override > campus)
        campus = Campus.query.get(group.campus_id)
        if not campus:
            return jsonify({'error': 'Campus no encontrado'}), 404
        
        if group.certification_cost_override is not None:
            unit_cost = float(group.certification_cost_override)
        elif campus.certification_cost is not None:
            unit_cost = float(campus.certification_cost)
        else:
            unit_cost = 0.0
        
        # Verificar cuántos ya tienen el ECM asignado (no se cobran)
        already_assigned_count = 0
        if exam_id:
            from app.models import Exam
            exam = Exam.query.get(exam_id)
            ecm_id = exam.competency_standard_id if exam else None
            if ecm_id:
                from app.models.partner import EcmCandidateAssignment
                already_assigned_count = EcmCandidateAssignment.query.filter(
                    EcmCandidateAssignment.user_id.in_(target_user_ids),
                    EcmCandidateAssignment.competency_standard_id == ecm_id
                ).count()
        
        billable_count = units - already_assigned_count
        total_cost = unit_cost * billable_count
        
        # Obtener saldo actual del coordinador para este grupo
        coordinator_id = g.current_user.id
        balance = CoordinatorBalance.query.filter_by(
            coordinator_id=coordinator_id,
            group_id=group_id
        ).first()
        current_balance = float(balance.current_balance) if balance else 0.0
        
        remaining_balance = current_balance - total_cost
        has_sufficient_balance = remaining_balance >= 0
        
        # Info del coordinador
        is_admin = g.current_user.role in ['admin', 'developer']
        
        return jsonify({
            'unit_cost': unit_cost,
            'units': units,
            'billable_count': billable_count,
            'already_assigned_count': already_assigned_count,
            'total_cost': total_cost,
            'current_balance': current_balance,
            'remaining_balance': remaining_balance,
            'has_sufficient_balance': has_sufficient_balance,
            'is_admin': is_admin,
            'campus_name': campus.name,
            'group_name': group.name,
            'cost_source': 'grupo (override)' if group.certification_cost_override is not None else 'campus',
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@bp.route('/groups/<int:group_id>/exams', methods=['POST'])
@jwt_required()
@coordinator_required
def assign_exam_to_group(group_id):
    """Asignar un examen a un grupo (automáticamente incluye materiales de estudio)
    
    Parámetros:
    - exam_id: ID del examen
    - assignment_type: 'all' (todo el grupo) o 'selected' (candidatos específicos)
    - member_ids: Lista de IDs de usuarios (solo si assignment_type='selected')
    - available_from, available_until: Período de disponibilidad (opcional)
    """
    try:
        from app.models import GroupExam, GroupExamMaterial, Exam
        from app.models.partner import GroupExamMember
        from app.models.study_content import StudyMaterial
        
        group = CandidateGroup.query.get_or_404(group_id)
        data = request.get_json()
        
        exam_id = data.get('exam_id')
        if not exam_id:
            return jsonify({'error': 'El ID del examen es requerido'}), 400
        
        assignment_type = data.get('assignment_type', 'all')
        member_ids = data.get('member_ids', [])
        
        if assignment_type == 'selected' and not member_ids:
            return jsonify({'error': 'Debes seleccionar al menos un candidato'}), 400
        
        # Verificar que el examen existe
        exam = Exam.query.get(exam_id)
        if not exam:
            return jsonify({'error': 'Examen no encontrado'}), 404
        
        # Verificar que no esté ya asignado
        existing = GroupExam.query.filter_by(group_id=group_id, exam_id=exam_id).first()
        if existing:
            if existing.is_active:
                return jsonify({'error': 'Este examen ya está asignado al grupo'}), 400
            else:
                # Reactivar la asignación
                existing.is_active = True
                existing.assigned_at = db.func.now()
                existing.assigned_by_id = g.current_user.id
                existing.assignment_type = assignment_type
                existing.available_from = data.get('available_from')
                existing.available_until = data.get('available_until')
                existing.time_limit_minutes = data.get('time_limit_minutes')
                existing.passing_score = data.get('passing_score')
                existing.max_attempts = data.get('max_attempts', 2)
                existing.max_disconnections = data.get('max_disconnections', 3)
                existing.exam_content_type = data.get('exam_content_type', 'questions_only')
                existing.exam_questions_count = data.get('exam_questions_count')
                existing.exam_exercises_count = data.get('exam_exercises_count')
                existing.simulator_questions_count = data.get('simulator_questions_count')
                existing.simulator_exercises_count = data.get('simulator_exercises_count')
                existing.security_pin = data.get('security_pin')
                existing.require_security_pin = data.get('require_security_pin', False)
                
                # Calcular vigencia para la reactivación
                r_campus = Campus.query.get(group.campus_id)
                if group.assignment_validity_months_override is not None:
                    r_validity_months = group.assignment_validity_months_override
                elif r_campus and r_campus.assignment_validity_months:
                    r_validity_months = r_campus.assignment_validity_months
                else:
                    r_validity_months = 12
                
                from dateutil.relativedelta import relativedelta
                r_assigned_at_now = datetime.utcnow()
                existing.assigned_at = r_assigned_at_now
                existing.validity_months = r_validity_months
                existing.expires_at = r_assigned_at_now + relativedelta(months=r_validity_months)
                existing.extended_months = 0  # Reset extensions on reactivation
                
                # Si es tipo 'selected', agregar miembros
                if assignment_type == 'selected':
                    # Limpiar miembros anteriores
                    GroupExamMember.query.filter_by(group_exam_id=existing.id).delete()
                    
                    for user_id in member_ids:
                        member = GroupExamMember(
                            group_exam_id=existing.id,
                            user_id=user_id
                        )
                        db.session.add(member)
                
                # Si se enviaron materiales personalizados, guardarlos
                material_ids = data.get('material_ids')
                if material_ids is not None:
                    # Limpiar materiales anteriores
                    GroupExamMaterial.query.filter_by(group_exam_id=existing.id).delete()
                    
                    for material_id in material_ids:
                        material = GroupExamMaterial(
                            group_exam_id=existing.id,
                            study_material_id=material_id,
                            is_included=True
                        )
                        db.session.add(material)
                
                # ========== ASIGNACIONES ECM PERMANENTES (Reactivación - verificar primero para cobrar correctamente) ==========
                r_already_assigned = []
                r_new_ecm_user_ids = []
                r_ecm_id = exam.competency_standard_id
                
                if r_ecm_id:
                    from app.models.partner import EcmCandidateAssignment
                    
                    if assignment_type == 'selected':
                        r_target_ids = member_ids
                    else:
                        r_target_ids = [m.user_id for m in group.members.all()]
                    
                    for uid in r_target_ids:
                        existing_ecm = EcmCandidateAssignment.query.filter_by(
                            user_id=uid,
                            competency_standard_id=r_ecm_id
                        ).first()
                        
                        if existing_ecm:
                            user_obj = User.query.get(uid)
                            r_already_assigned.append({
                                'user_id': uid,
                                'user_name': user_obj.full_name if user_obj else uid,
                                'user_email': user_obj.email if user_obj else '',
                                'user_curp': user_obj.curp if user_obj else '',
                                'assignment_number': existing_ecm.assignment_number,
                                'assigned_at': existing_ecm.assigned_at.isoformat() if existing_ecm.assigned_at else None,
                                'original_group': existing_ecm.group_name,
                            })
                        else:
                            r_new_ecm_user_ids.append(uid)
                else:
                    # Sin ECM, todos son "nuevos" para el cobro
                    if assignment_type == 'selected':
                        r_new_ecm_user_ids = member_ids
                    else:
                        r_new_ecm_user_ids = [m.user_id for m in group.members.all()]
                
                # ========== VERIFICACIÓN Y DEDUCCIÓN DE SALDO (Reactivación) ==========
                # Solo cobrar por candidatos que NO tienen ya el ECM asignado
                # Admins y developers no requieren verificación de saldo
                user_role = g.current_user.role if hasattr(g.current_user, 'role') else ''
                is_admin_or_dev = user_role in ('admin', 'developer')
                
                campus = Campus.query.get(group.campus_id)
                if group.certification_cost_override is not None:
                    r_unit_cost = float(group.certification_cost_override)
                elif campus and campus.certification_cost is not None:
                    r_unit_cost = float(campus.certification_cost)
                else:
                    r_unit_cost = 0.0
                
                r_billable_count = len(r_new_ecm_user_ids)
                r_total_cost = r_unit_cost * r_billable_count
                
                if r_total_cost > 0 and not is_admin_or_dev:
                    coordinator_id = g.current_user.id
                    balance = CoordinatorBalance.query.filter_by(
                        coordinator_id=coordinator_id,
                        group_id=group_id
                    ).first()
                    current_bal = float(balance.current_balance) if balance else 0.0
                    
                    if current_bal < r_total_cost:
                        db.session.rollback()
                        return jsonify({
                            'error': f'Saldo insuficiente para este grupo. Necesitas ${r_total_cost:,.2f} pero tu saldo es ${current_bal:,.2f}',
                            'error_type': 'insufficient_balance',
                            'required': r_total_cost,
                            'available': current_bal,
                            'deficit': r_total_cost - current_bal,
                            'already_assigned_count': len(r_already_assigned),
                            'billable_count': r_billable_count
                        }), 400
                    
                    exam_name = exam.name if exam else f'Examen #{exam_id}'
                    skipped_note = f' ({len(r_already_assigned)} omitido(s) por ya tener ECM)' if r_already_assigned else ''
                    notes = f'Reasignación de "{exam_name}" a grupo "{group.name}" - {r_billable_count} unidad(es) x ${r_unit_cost:,.2f}{skipped_note}'
                    
                    create_balance_transaction(
                        coordinator_id=coordinator_id,
                        group_id=group_id,
                        transaction_type='debit',
                        concept='asignacion_certificacion',
                        amount=r_total_cost,
                        reference_type='group_exam',
                        reference_id=existing.id,
                        notes=notes,
                        created_by_id=coordinator_id
                    )
                
                # Crear EcmCandidateAssignment para los nuevos
                if r_ecm_id:
                    for uid in r_new_ecm_user_ids:
                        new_ecm = EcmCandidateAssignment(
                            assignment_number=EcmCandidateAssignment.generate_assignment_number(),
                            user_id=uid,
                            competency_standard_id=r_ecm_id,
                            exam_id=exam_id,
                            campus_id=group.campus_id,
                            group_id=group_id,
                            group_name=group.name,
                            group_exam_id=existing.id,
                            assigned_by_id=g.current_user.id,
                            assignment_source='selected' if assignment_type == 'selected' else 'bulk',
                            validity_months=r_validity_months,
                            assigned_at=r_assigned_at_now,
                            expires_at=r_assigned_at_now + relativedelta(months=r_validity_months),
                        )
                        db.session.add(new_ecm)
                
                db.session.commit()
                
                # Obtener materiales asociados
                materials = StudyMaterial.query.filter_by(exam_id=exam_id, is_published=True).all()
                
                r_response = {
                    'message': 'Examen reactivado exitosamente',
                    'assignment': existing.to_dict(include_exam=True, include_materials=True, include_members=True),
                    'study_materials_count': len(material_ids) if material_ids else len(materials),
                    'assigned_members_count': len(member_ids) if assignment_type == 'selected' else group.members.count()
                }
                
                if r_already_assigned:
                    r_response['already_assigned'] = r_already_assigned
                    r_response['already_assigned_count'] = len(r_already_assigned)
                
                return jsonify(r_response)
        
        # Configuración del examen
        time_limit_minutes = data.get('time_limit_minutes')
        passing_score = data.get('passing_score')
        max_attempts = data.get('max_attempts', 2)
        max_disconnections = data.get('max_disconnections', 3)
        exam_content_type = data.get('exam_content_type', 'questions_only')
        exam_questions_count = data.get('exam_questions_count')
        exam_exercises_count = data.get('exam_exercises_count')
        simulator_questions_count = data.get('simulator_questions_count')
        simulator_exercises_count = data.get('simulator_exercises_count')
        security_pin = data.get('security_pin')
        require_security_pin = data.get('require_security_pin', False)
        material_ids = data.get('material_ids')
        
        # Calcular vigencia de la asignación
        campus = Campus.query.get(group.campus_id)
        if group.assignment_validity_months_override is not None:
            validity_months = group.assignment_validity_months_override
        elif campus and campus.assignment_validity_months:
            validity_months = campus.assignment_validity_months
        else:
            validity_months = 12
        
        from dateutil.relativedelta import relativedelta
        assigned_at_now = datetime.utcnow()
        expires_at = assigned_at_now + relativedelta(months=validity_months)
        
        # Crear nueva asignación
        group_exam = GroupExam(
            group_id=group_id,
            exam_id=exam_id,
            assigned_by_id=g.current_user.id,
            assignment_type=assignment_type,
            available_from=data.get('available_from'),
            available_until=data.get('available_until'),
            time_limit_minutes=time_limit_minutes,
            passing_score=passing_score,
            max_attempts=max_attempts,
            max_disconnections=max_disconnections,
            exam_content_type=exam_content_type,
            exam_questions_count=exam_questions_count,
            exam_exercises_count=exam_exercises_count,
            simulator_questions_count=simulator_questions_count,
            simulator_exercises_count=simulator_exercises_count,
            security_pin=security_pin,
            require_security_pin=require_security_pin,
            validity_months=validity_months,
            assigned_at=assigned_at_now,
            expires_at=expires_at,
        )
        
        db.session.add(group_exam)
        db.session.flush()  # Para obtener el ID
        
        # Si es tipo 'selected', agregar miembros específicos
        if assignment_type == 'selected':
            for user_id in member_ids:
                member = GroupExamMember(
                    group_exam_id=group_exam.id,
                    user_id=user_id
                )
                db.session.add(member)
        
        # Si se enviaron materiales personalizados, guardarlos
        if material_ids is not None:
            for material_id in material_ids:
                material = GroupExamMaterial(
                    group_exam_id=group_exam.id,
                    study_material_id=material_id,
                    is_included=True
                )
                db.session.add(material)
        
        # ========== ASIGNACIONES ECM PERMANENTES (verificar primero para cobrar correctamente) ==========
        already_assigned = []
        new_assignments = []
        ecm_id = exam.competency_standard_id
        new_ecm_user_ids = []
        
        # Determinar lista de user_ids afectados
        if assignment_type == 'selected':
            target_user_ids = member_ids
        else:
            target_user_ids = [m.user_id for m in group.members.all()]
        
        if ecm_id:
            from app.models.partner import EcmCandidateAssignment
            
            for uid in target_user_ids:
                existing_assignment = EcmCandidateAssignment.query.filter_by(
                    user_id=uid,
                    competency_standard_id=ecm_id
                ).first()
                
                if existing_assignment:
                    user_obj = User.query.get(uid)
                    already_assigned.append({
                        'user_id': uid,
                        'user_name': user_obj.full_name if user_obj else uid,
                        'user_email': user_obj.email if user_obj else '',
                        'user_curp': user_obj.curp if user_obj else '',
                        'assignment_number': existing_assignment.assignment_number,
                        'assigned_at': existing_assignment.assigned_at.isoformat() if existing_assignment.assigned_at else None,
                        'original_group': existing_assignment.group_name,
                    })
                else:
                    new_ecm_user_ids.append(uid)
        else:
            # Sin ECM, todos son "nuevos" para el cobro
            new_ecm_user_ids = target_user_ids
        
        # ========== VERIFICACIÓN Y DEDUCCIÓN DE SALDO ==========
        # Solo cobrar por candidatos que NO tienen ya el ECM asignado
        # Admins y developers no requieren verificación de saldo
        user_role = g.current_user.role if hasattr(g.current_user, 'role') else ''
        is_admin_or_dev = user_role in ('admin', 'developer')
        
        campus = Campus.query.get(group.campus_id)
        if group.certification_cost_override is not None:
            unit_cost = float(group.certification_cost_override)
        elif campus and campus.certification_cost is not None:
            unit_cost = float(campus.certification_cost)
        else:
            unit_cost = 0.0
        
        billable_count = len(new_ecm_user_ids)
        total_cost = unit_cost * billable_count
        
        # Solo verificar y deducir si hay costo > 0 y el usuario es coordinador (no admin/dev)
        if total_cost > 0 and not is_admin_or_dev:
            coordinator_id = g.current_user.id
            balance = CoordinatorBalance.query.filter_by(
                coordinator_id=coordinator_id,
                group_id=group_id
            ).first()
            current_balance = float(balance.current_balance) if balance else 0.0
            
            if current_balance < total_cost:
                db.session.rollback()
                return jsonify({
                    'error': f'Saldo insuficiente para este grupo. Necesitas ${total_cost:,.2f} pero tu saldo es ${current_balance:,.2f}',
                    'error_type': 'insufficient_balance',
                    'required': total_cost,
                    'available': current_balance,
                    'deficit': total_cost - current_balance,
                    'already_assigned_count': len(already_assigned),
                    'billable_count': billable_count
                }), 400
            
            # Deducir saldo y crear transacción
            exam_name = exam.name if exam else f'Examen #{exam_id}'
            skipped_note = f' ({len(already_assigned)} omitido(s) por ya tener ECM)' if already_assigned else ''
            notes = f'Asignación de "{exam_name}" a grupo "{group.name}" - {billable_count} unidad(es) x ${unit_cost:,.2f}{skipped_note}'
            
            create_balance_transaction(
                coordinator_id=coordinator_id,
                group_id=group_id,
                transaction_type='debit',
                concept='asignacion_certificacion',
                amount=total_cost,
                reference_type='group_exam',
                reference_id=group_exam.id,
                notes=notes,
                created_by_id=coordinator_id
            )
        
        # Crear EcmCandidateAssignment para los nuevos
        if ecm_id:
            for uid in new_ecm_user_ids:
                new_ecm = EcmCandidateAssignment(
                    assignment_number=EcmCandidateAssignment.generate_assignment_number(),
                    user_id=uid,
                    competency_standard_id=ecm_id,
                    exam_id=exam_id,
                    campus_id=group.campus_id,
                    group_id=group_id,
                    group_name=group.name,
                    group_exam_id=group_exam.id,
                    assigned_by_id=g.current_user.id,
                    assignment_source='selected' if assignment_type == 'selected' else 'bulk',
                    validity_months=validity_months,
                    assigned_at=assigned_at_now,
                    expires_at=expires_at,
                )
                db.session.add(new_ecm)
                new_assignments.append(new_ecm)
        
        db.session.commit()
        
        # Build detailed list of new ECM assignments
        new_assignments_detail = []
        for ecm_assign in new_assignments:
            user_obj = User.query.get(ecm_assign.user_id)
            new_assignments_detail.append({
                'user_id': str(ecm_assign.user_id),
                'user_name': user_obj.full_name if user_obj else str(ecm_assign.user_id),
                'user_email': user_obj.email if user_obj else '',
                'user_curp': user_obj.curp if user_obj else '',
                'assignment_number': ecm_assign.assignment_number,
                'assigned_at': ecm_assign.assigned_at.isoformat() if ecm_assign.assigned_at else None,
                'exam_name': exam.name if exam else '',
                'group_name': group.name,
            })
        
        # Obtener materiales asociados al examen
        materials = StudyMaterial.query.filter_by(exam_id=exam_id, is_published=True).all()
        materials_count = len(material_ids) if material_ids else len(materials)
        
        message = 'Examen asignado exitosamente.'
        if assignment_type == 'all':
            message += f' Disponible para los {group.members.count()} miembros del grupo.'
        else:
            message += f' Disponible para {len(member_ids)} candidato(s) seleccionado(s).'
        
        if materials_count > 0:
            message += f' Con {materials_count} material(es) de estudio.'
        
        if already_assigned:
            message += f' {len(already_assigned)} candidato(s) ya tenían este ECM asignado previamente (sin cobro).'
        
        response_data = {
            'message': message,
            'assignment': group_exam.to_dict(include_exam=True, include_materials=True, include_members=True),
            'study_materials_count': materials_count,
            'assigned_members_count': len(member_ids) if assignment_type == 'selected' else group.members.count(),
            'new_ecm_assignments_count': len(new_assignments),
            'new_assignments': new_assignments_detail,
            'billing': {
                'unit_cost': unit_cost,
                'billable_count': billable_count,
                'total_cost': total_cost,
                'skipped_count': len(already_assigned),
            },
        }
        
        if already_assigned:
            response_data['already_assigned'] = already_assigned
            response_data['already_assigned_count'] = len(already_assigned)
        
        return jsonify(response_data), 201
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500


@bp.route('/groups/<int:group_id>/exams/<int:exam_id>', methods=['DELETE'])
@jwt_required()
@coordinator_required
def unassign_exam_from_group(group_id, exam_id):
    """Desasignar un examen de un grupo"""
    try:
        from app.models import GroupExam
        
        group_exam = GroupExam.query.filter_by(
            group_id=group_id, 
            exam_id=exam_id
        ).first_or_404()
        
        group_exam.is_active = False
        db.session.commit()
        
        return jsonify({
            'message': 'Examen desasignado del grupo'
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500


@bp.route('/groups/<int:group_id>/exams/<int:exam_id>/members', methods=['GET'])
@jwt_required()
@coordinator_required
def get_group_exam_members(group_id, exam_id):
    """Obtener los miembros asignados a un examen específico del grupo"""
    try:
        from app.models import GroupExam
        from app.models.partner import GroupExamMember
        
        group_exam = GroupExam.query.filter_by(
            group_id=group_id, 
            exam_id=exam_id,
            is_active=True
        ).first_or_404()
        
        if group_exam.assignment_type == 'all':
            # Todos los miembros del grupo
            group_members = GroupMember.query.filter_by(group_id=group_id).all()
            assigned_user_ids = [m.user_id for m in group_members]
            members = []  # No hay miembros específicos asignados
        else:
            members = [m.to_dict(include_user=True) for m in group_exam.assigned_members]
            assigned_user_ids = [m.user_id for m in group_exam.assigned_members]
        
        return jsonify({
            'assignment_id': group_exam.id,
            'exam_id': exam_id,
            'assignment_type': group_exam.assignment_type,
            'members': members,
            'assigned_user_ids': assigned_user_ids,
            'total_members': len(assigned_user_ids)
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@bp.route('/groups/<int:group_id>/exams/<int:exam_id>/members', methods=['PUT'])
@jwt_required()
@coordinator_required
def update_group_exam_members(group_id, exam_id):
    """Actualizar los miembros asignados a un examen del grupo"""
    try:
        from app.models import GroupExam
        from app.models.partner import GroupExamMember
        
        group = CandidateGroup.query.get_or_404(group_id)
        
        group_exam = GroupExam.query.filter_by(
            group_id=group_id, 
            exam_id=exam_id,
            is_active=True
        ).first_or_404()
        
        data = request.get_json()
        assignment_type = data.get('assignment_type', group_exam.assignment_type)
        member_ids = data.get('member_ids', [])
        
        if assignment_type == 'selected' and not member_ids:
            return jsonify({'error': 'Debes seleccionar al menos un candidato'}), 400
        
        # Actualizar tipo de asignación
        group_exam.assignment_type = assignment_type
        
        # Limpiar miembros anteriores
        GroupExamMember.query.filter_by(group_exam_id=group_exam.id).delete()
        
        # Si es tipo 'selected', agregar nuevos miembros
        if assignment_type == 'selected':
            for user_id in member_ids:
                member = GroupExamMember(
                    group_exam_id=group_exam.id,
                    user_id=user_id
                )
                db.session.add(member)
        
        db.session.commit()
        
        message = 'Asignación actualizada exitosamente.'
        if assignment_type == 'all':
            message += f' El examen está disponible para los {group.members.count()} miembros del grupo.'
        else:
            message += f' El examen está disponible para {len(member_ids)} candidato(s) seleccionado(s).'
        
        return jsonify({
            'message': message,
            'assignment': group_exam.to_dict(include_exam=True, include_members=True)
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500


@bp.route('/groups/<int:group_id>/exams/<int:exam_id>/members/add', methods=['POST'])
@jwt_required()
@coordinator_required
def add_members_to_exam(group_id, exam_id):
    """Agregar miembros a un examen existente sin afectar los actuales
    
    Parámetros:
    - user_ids: Lista de user_ids a agregar
    """
    try:
        from app.models import GroupExam
        from app.models.partner import GroupExamMember
        
        group_exam = GroupExam.query.filter_by(
            group_id=group_id,
            exam_id=exam_id,
            is_active=True
        ).first_or_404()
        
        data = request.get_json()
        user_ids_to_add = data.get('user_ids', [])
        
        if not user_ids_to_add:
            return jsonify({'error': 'Debes proporcionar user_ids a agregar'}), 400
        
        # Cambiar a 'selected' si era 'all'
        if group_exam.assignment_type == 'all':
            # Si era 'all', primero agregar todos los miembros actuales
            group_members = GroupMember.query.filter_by(group_id=group_id).all()
            for gm in group_members:
                existing = GroupExamMember.query.filter_by(
                    group_exam_id=group_exam.id,
                    user_id=gm.user_id
                ).first()
                if not existing:
                    member = GroupExamMember(
                        group_exam_id=group_exam.id,
                        user_id=gm.user_id
                    )
                    db.session.add(member)
            group_exam.assignment_type = 'selected'
        
        added = []
        for user_id in user_ids_to_add:
            existing = GroupExamMember.query.filter_by(
                group_exam_id=group_exam.id,
                user_id=user_id
            ).first()
            if not existing:
                member = GroupExamMember(
                    group_exam_id=group_exam.id,
                    user_id=user_id
                )
                db.session.add(member)
                added.append(user_id)
        
        db.session.commit()
        
        return jsonify({
            'message': f'{len(added)} usuario(s) agregado(s) al examen',
            'added': added
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500


@bp.route('/groups/<int:group_id>/exams/<int:exam_id>/members-detail', methods=['GET'])
@jwt_required()
@coordinator_required
def get_group_exam_members_detail(group_id, exam_id):
    """Obtener detalle completo paginado de los miembros asignados a un examen.
    Optimizado con SQL-level pagination para cientos de miles de registros.
    
    Query params:
    - page (int, default 1)
    - per_page (int, default 150, max 1000)
    - search (str, optional) — busca en nombre, email, curp, assignment_number
    - sort_by (str, default 'name') — name, email, curp, assignment_number, progress, status
    - sort_dir (str, default 'asc')
    - filter_status (str, optional) — 'locked', 'swappable', 'all'
    """
    try:
        from app.models import GroupExam
        from app.models.partner import GroupExamMember, EcmCandidateAssignment
        from app.models.result import Result
        from app.models.student_progress import StudentTopicProgress
        from app.models.study_content import StudyMaterial
        from app.models.user import User
        from sqlalchemy import text, func, case, literal_column

        page = request.args.get('page', 1, type=int)
        per_page = min(request.args.get('per_page', 150, type=int), 1000)
        search = request.args.get('search', '').strip()
        sort_by = request.args.get('sort_by', 'name')
        sort_dir_param = request.args.get('sort_dir', 'asc')
        filter_status = request.args.get('filter_status', 'all')

        group_exam = GroupExam.query.filter_by(
            group_id=group_id,
            exam_id=exam_id,
            is_active=True
        ).first_or_404()

        exam = group_exam.exam
        ecm_id = exam.competency_standard_id if exam else None

        # ── Pre-compute topic_ids for material progress ──
        total_topics = 0
        topic_ids = []
        try:
            material_ids_rows = db.session.execute(text(
                'SELECT study_material_id FROM study_material_exams WHERE exam_id = :eid'
            ), {'eid': exam_id}).fetchall()
            material_ids = [r[0] for r in material_ids_rows]
            for m in StudyMaterial.query.filter_by(exam_id=exam_id).all():
                if m.id not in material_ids:
                    material_ids.append(m.id)

            if material_ids:
                topic_rows = db.session.execute(text('''
                    SELECT DISTINCT st.id
                    FROM study_topics st
                    INNER JOIN study_sessions ss ON st.session_id = ss.id
                    WHERE ss.study_material_id IN :mids
                ''').bindparams(mids=tuple(material_ids) if len(material_ids) > 1 else tuple(material_ids + [0]))).fetchall()
                topic_ids = [r[0] for r in topic_rows]
                total_topics = len(topic_ids)
        except Exception as prog_err:
            print(f"⚠️ Error calculando topics: {prog_err}")

        # ── Build base member subquery using SQLAlchemy ──
        # Always use ALL group members so that after a swap, the source user
        # still appears with status "Sin asignación" instead of disappearing.
        member_base = db.session.query(
            GroupMember.user_id
        ).filter(GroupMember.group_id == group_id).subquery('member_base')

        # ── Subquery: EcmCandidateAssignment ──
        eca_sub = None
        if ecm_id:
            # Pick the assignment matching this group_exam_id first, otherwise any for this ecm
            eca_sub = db.session.query(
                EcmCandidateAssignment.user_id,
                EcmCandidateAssignment.id.label('eca_id'),
                EcmCandidateAssignment.assignment_number,
                EcmCandidateAssignment.assigned_at,
                func.row_number().over(
                    partition_by=EcmCandidateAssignment.user_id,
                    order_by=case(
                        (EcmCandidateAssignment.group_exam_id == group_exam.id, 0),
                        else_=1
                    )
                ).label('rn')
            ).filter(
                EcmCandidateAssignment.competency_standard_id == ecm_id,
                EcmCandidateAssignment.user_id.in_(db.session.query(member_base.c.user_id))
            ).subquery('eca_ranked')

            eca_final = db.session.query(
                eca_sub.c.user_id,
                eca_sub.c.eca_id,
                eca_sub.c.assignment_number,
                eca_sub.c.assigned_at,
            ).filter(eca_sub.c.rn == 1).subquery('eca_final')

        # ── Subquery: has opened exam (any result exists) ──
        has_result_sub = db.session.query(
            Result.user_id,
            func.count(Result.id).label('results_count')
        ).filter(
            Result.group_exam_id == group_exam.id,
            Result.user_id.in_(db.session.query(member_base.c.user_id))
        ).group_by(Result.user_id).subquery('has_result')

        # ── Subquery: material progress ──
        progress_sub = None
        if topic_ids and total_topics > 0:
            progress_sub = db.session.query(
                StudentTopicProgress.user_id,
                (func.sum(func.coalesce(StudentTopicProgress.progress_percentage, 0.0)) / total_topics).label('progress_pct')
            ).filter(
                StudentTopicProgress.topic_id.in_(topic_ids),
                StudentTopicProgress.user_id.in_(db.session.query(member_base.c.user_id))
            ).group_by(StudentTopicProgress.user_id).subquery('progress')

        # ── Main query ──
        main_q = db.session.query(
            User.id.label('user_id'),
            User.name.label('first_name'),
            User.first_surname,
            User.second_surname,
            func.concat(
                User.name, ' ', User.first_surname,
                func.coalesce(func.concat(' ', User.second_surname), '')
            ).label('full_name'),
            User.email,
            User.curp,
            User.username,
        )

        # Join member base
        main_q = main_q.join(member_base, User.id == member_base.c.user_id)

        # Add ECA columns
        if ecm_id and eca_sub is not None:
            main_q = main_q.outerjoin(eca_final, User.id == eca_final.c.user_id)
            main_q = main_q.add_columns(
                eca_final.c.eca_id,
                eca_final.c.assignment_number,
                eca_final.c.assigned_at.label('eca_assigned_at'),
            )
        else:
            main_q = main_q.add_columns(
                literal_column('NULL').label('eca_id'),
                literal_column('NULL').label('assignment_number'),
                literal_column('NULL').label('eca_assigned_at'),
            )

        # Add result columns
        main_q = main_q.outerjoin(has_result_sub, User.id == has_result_sub.c.user_id)
        main_q = main_q.add_columns(
            func.coalesce(has_result_sub.c.results_count, 0).label('results_count'),
        )

        # Add progress columns
        if progress_sub is not None:
            main_q = main_q.outerjoin(progress_sub, User.id == progress_sub.c.user_id)
            main_q = main_q.add_columns(
                func.coalesce(progress_sub.c.progress_pct, 0.0).label('material_progress'),
            )
        else:
            main_q = main_q.add_columns(
                literal_column('0.0').label('material_progress'),
            )

        # ── Computed: is_locked (progress >= 15 OR has_opened_exam) ──
        # We need to filter on this, so we wrap in a subquery
        main_sub = main_q.subquery('main_data')

        is_locked_expr = case(
            (db.or_(main_sub.c.material_progress >= 15.0, main_sub.c.results_count > 0), True),
            else_=False
        ).label('is_locked')

        final_q = db.session.query(main_sub, is_locked_expr)

        # ── Search filter ──
        if search:
            search_term = f'%{search.lower()}%'
            final_q = final_q.filter(
                db.or_(
                    func.lower(main_sub.c.full_name).like(search_term),
                    func.lower(main_sub.c.email).like(search_term),
                    func.lower(func.coalesce(main_sub.c.curp, '')).like(search_term),
                    func.lower(func.coalesce(main_sub.c.assignment_number, '')).like(search_term),
                )
            )

        # ── Status filter ──
        if filter_status == 'locked':
            final_q = final_q.filter(
                db.or_(main_sub.c.material_progress >= 15.0, main_sub.c.results_count > 0)
            )
        elif filter_status == 'swappable':
            final_q = final_q.filter(
                main_sub.c.material_progress < 15.0,
                main_sub.c.results_count == 0
            )

        # ── Counts (before pagination, after filters) ──
        count_q = final_q.with_entities(
            func.count().label('total_count'),
            func.sum(case(
                (db.or_(main_sub.c.material_progress >= 15.0, main_sub.c.results_count > 0), 1),
                else_=0
            )).label('locked_count'),
        )
        counts = count_q.one()
        total = counts.total_count or 0
        locked_count = counts.locked_count or 0
        swappable_count = total - locked_count

        total_pages = max(1, (total + per_page - 1) // per_page)
        page = min(page, total_pages)

        # ── Sorting ──
        sort_map = {
            'name': main_sub.c.full_name,
            'email': main_sub.c.email,
            'curp': main_sub.c.curp,
            'assignment_number': main_sub.c.assignment_number,
            'progress': main_sub.c.material_progress,
            'status': is_locked_expr,
        }
        sort_col = sort_map.get(sort_by, main_sub.c.full_name)
        if sort_dir_param == 'desc':
            final_q = final_q.order_by(sort_col.desc())
        else:
            final_q = final_q.order_by(sort_col.asc())

        # ── Pagination ──
        results = final_q.offset((page - 1) * per_page).limit(per_page).all()

        # ── Format results ──
        # Pre-fetch retake counts for all users in this page
        page_user_ids = [r.user_id for r in results]
        retake_map = {}  # user_id -> {count, active_count}
        if ecm_id and page_user_ids:
            from app.models.partner import EcmRetake
            retake_rows = db.session.query(
                EcmRetake.user_id,
                func.count(EcmRetake.id).label('total_retakes'),
                func.sum(case((EcmRetake.status == 'active', 1), else_=0)).label('active_retakes')
            ).filter(
                EcmRetake.group_exam_id == group_exam.id,
                EcmRetake.user_id.in_(page_user_ids)
            ).group_by(EcmRetake.user_id).all()
            for row in retake_rows:
                retake_map[row.user_id] = {
                    'total': row.total_retakes or 0,
                    'active': int(row.active_retakes or 0)
                }
        
        # Config de retomas
        group = CandidateGroup.query.get(group_id)
        campus_for_config = Campus.query.get(group.campus_id) if group else None
        max_retakes_config = group.max_retakes_override if group and group.max_retakes_override is not None else (campus_for_config.max_retakes if campus_for_config and campus_for_config.max_retakes is not None else 0)
        
        page_members = []
        max_attempts = group_exam.max_attempts or 1
        for r in results:
            progress_pct = round(float(r.material_progress or 0), 1)
            has_opened = (r.results_count or 0) > 0
            locked = bool(r.is_locked)
            lock_reasons = []
            if progress_pct >= 15.0:
                lock_reasons.append(f'Avance de material: {progress_pct}%')
            if has_opened:
                lock_reasons.append('Ha abierto examen/simulador')
            
            retake_info = retake_map.get(r.user_id, {'total': 0, 'active': 0})
            total_allowed = max_attempts + retake_info['total']
            results_c = r.results_count or 0
            attempts_remaining = max(0, total_allowed - results_c)
            attempts_exhausted = results_c >= total_allowed
            can_retake = (
                ecm_id is not None
                and r.eca_id is not None
                and attempts_exhausted
                and retake_info['total'] < max_retakes_config
            )

            page_members.append({
                'user_id': r.user_id,
                'user': {
                    'id': r.user_id,
                    'name': r.first_name,
                    'first_surname': r.first_surname or '',
                    'second_surname': r.second_surname or '',
                    'full_name': r.full_name or '',
                    'email': r.email,
                    'curp': r.curp,
                    'username': r.username,
                },
                'assignment_number': r.assignment_number,
                'ecm_assignment_id': r.eca_id,
                'ecm_assignment_date': r.eca_assigned_at.isoformat() if r.eca_assigned_at else None,
                'material_progress': progress_pct,
                'has_opened_exam': has_opened,
                'results_count': results_c,
                'is_locked': locked,
                'lock_reasons': lock_reasons,
                'retakes_count': retake_info['total'],
                'retakes_active': retake_info['active'],
                'max_retakes': max_retakes_config,
                'total_allowed_attempts': total_allowed,
                'attempts_remaining': attempts_remaining,
                'attempts_exhausted': attempts_exhausted,
                'can_retake': can_retake,
            })

        ecm_code = None
        if exam and ecm_id:
            try:
                ecm_code = exam.competency_standard.code if exam.competency_standard else None
            except Exception:
                pass

        return jsonify({
            'assignment_id': group_exam.id,
            'exam_id': exam_id,
            'exam_name': exam.name if exam else None,
            'ecm_id': ecm_id,
            'ecm_code': ecm_code,
            'assignment_type': group_exam.assignment_type,
            'max_attempts': max_attempts,
            'max_retakes': max_retakes_config,
            'members': page_members,
            'total': total,
            'page': page,
            'pages': total_pages,
            'per_page': per_page,
            'locked_count': locked_count,
            'swappable_count': swappable_count,
        })

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@bp.route('/groups/<int:group_id>/exams/<int:exam_id>/members/swap', methods=['POST'])
@jwt_required()
@coordinator_required
def swap_exam_member(group_id, exam_id):
    """Reasignar (swap) una asignación ECM de un candidato a otro.
    
    Body JSON:
    {
        "from_user_id": "uuid-del-candidato-actual",
        "to_user_id": "uuid-del-nuevo-candidato"
    }
    
    Validaciones:
    - El candidato origen DEBE tener un número de asignación (EcmCandidateAssignment)
    - El candidato destino NO debe tener número de asignación
    - El candidato origen NO debe tener ≥15% de progreso en material
    - El candidato origen NO debe haber abierto examen/simulador
    - El candidato destino debe ser miembro del grupo
    """
    try:
        from app.models import GroupExam
        from app.models.partner import GroupExamMember, EcmCandidateAssignment, EcmSwapHistory
        from app.models.result import Result
        from app.models.student_progress import StudentTopicProgress, StudentContentProgress
        from app.models.study_content import StudyMaterial
        from sqlalchemy import text

        data = request.get_json()
        from_user_id = data.get('from_user_id')
        to_user_id = data.get('to_user_id')

        if not from_user_id or not to_user_id:
            return jsonify({'error': 'Se requieren from_user_id y to_user_id'}), 400

        if from_user_id == to_user_id:
            return jsonify({'error': 'El candidato origen y destino no pueden ser el mismo'}), 400

        group_exam = GroupExam.query.filter_by(
            group_id=group_id,
            exam_id=exam_id,
            is_active=True
        ).first_or_404()

        exam = group_exam.exam
        ecm_id = exam.competency_standard_id if exam else None

        # Verificar que from_user está asignado
        if group_exam.assignment_type == 'selected':
            from_member = GroupExamMember.query.filter_by(
                group_exam_id=group_exam.id,
                user_id=from_user_id
            ).first()
            if not from_member:
                return jsonify({'error': 'El candidato origen no está asignado a este examen'}), 400
        else:
            # Si es 'all', verificar que es miembro del grupo
            is_member = GroupMember.query.filter_by(group_id=group_id, user_id=from_user_id).first()
            if not is_member:
                return jsonify({'error': 'El candidato origen no es miembro del grupo'}), 400

        # Verificar que to_user es miembro del grupo
        to_group_member = GroupMember.query.filter_by(group_id=group_id, user_id=to_user_id).first()
        if not to_group_member:
            return jsonify({'error': 'El candidato destino no es miembro del grupo'}), 400

        # === VALIDAR NÚMEROS DE ASIGNACIÓN ===
        # Nota: No se valida GroupExamMember del destino porque un miembro del
        # examen puede no tener número de asignación y ser un destino válido.
        # El origen DEBE tener número de asignación, el destino NO debe tener
        if ecm_id:
            from_eca = EcmCandidateAssignment.query.filter_by(
                user_id=from_user_id,
                competency_standard_id=ecm_id
            ).first()
            if not from_eca:
                return jsonify({
                    'error': 'No se puede reasignar: el candidato origen no tiene número de asignación',
                    'reason': 'no_assignment_number'
                }), 400

            dest_eca = EcmCandidateAssignment.query.filter_by(
                user_id=to_user_id,
                competency_standard_id=ecm_id
            ).first()
            if dest_eca:
                return jsonify({
                    'error': 'No se puede reasignar: el candidato destino ya tiene un número de asignación',
                    'reason': 'dest_has_assignment'
                }), 400

        # === VALIDAR QUE EL CANDIDATO ORIGEN PUEDE SER REASIGNADO ===

        # 1. Verificar si tiene resultados (ha abierto examen/simulador)
        has_results = Result.query.filter(
            Result.user_id == from_user_id,
            Result.group_exam_id == group_exam.id
        ).first()
        if has_results:
            return jsonify({
                'error': 'No se puede reasignar: el candidato ya ha abierto el examen o simulador',
                'reason': 'has_opened_exam'
            }), 400

        # 2. Verificar progreso de material de estudio >= 15%
        material_ids = []
        topic_ids = []
        try:
            material_ids_rows = db.session.execute(text(
                'SELECT study_material_id FROM study_material_exams WHERE exam_id = :eid'
            ), {'eid': exam_id}).fetchall()
            material_ids = [r[0] for r in material_ids_rows]

            legacy = StudyMaterial.query.filter_by(exam_id=exam_id).all()
            for m in legacy:
                if m.id not in material_ids:
                    material_ids.append(m.id)

            if material_ids:
                topic_rows = db.session.execute(text('''
                    SELECT DISTINCT st.id
                    FROM study_topics st
                    INNER JOIN study_sessions ss ON st.session_id = ss.id
                    WHERE ss.study_material_id IN :mids
                ''').bindparams(mids=tuple(material_ids) if len(material_ids) > 1 else tuple(material_ids + [0]))).fetchall()
                topic_ids = [r[0] for r in topic_rows]

                if topic_ids:
                    progress_records = StudentTopicProgress.query.filter(
                        StudentTopicProgress.user_id == from_user_id,
                        StudentTopicProgress.topic_id.in_(topic_ids)
                    ).all()

                    total_topics = len(topic_ids)
                    sum_pct = sum(pr.progress_percentage or 0.0 for pr in progress_records)
                    avg_progress = sum_pct / total_topics if total_topics > 0 else 0.0

                    if avg_progress >= 15.0:
                        return jsonify({
                            'error': f'No se puede reasignar: el candidato tiene {round(avg_progress, 1)}% de avance en material de estudio (mínimo permitido: <15%)',
                            'reason': 'material_progress',
                            'progress': round(avg_progress, 1)
                        }), 400
        except Exception as prog_err:
            print(f"⚠️ Error verificando progreso: {prog_err}")

        # === REALIZAR EL SWAP ===

        # Si es 'selected', asegurar que ambos usuarios son miembros del examen
        if group_exam.assignment_type == 'selected':
            # El destino obtiene GroupExamMember si no lo tiene
            # (el origen se eliminará más adelante en la sección de revocación)
            existing_dest_gem = GroupExamMember.query.filter_by(
                group_exam_id=group_exam.id,
                user_id=to_user_id
            ).first()
            if not existing_dest_gem:
                db.session.add(GroupExamMember(
                    group_exam_id=group_exam.id,
                    user_id=to_user_id
                ))
        else:
            # Si era 'all', convertir a 'selected' con TODOS los miembros
            all_members = GroupMember.query.filter_by(group_id=group_id).all()
            group_exam.assignment_type = 'selected'
            for gm in all_members:
                existing = GroupExamMember.query.filter_by(
                    group_exam_id=group_exam.id,
                    user_id=gm.user_id
                ).first()
                if not existing:
                    db.session.add(GroupExamMember(
                        group_exam_id=group_exam.id,
                        user_id=gm.user_id
                    ))

        # Transferir la EcmCandidateAssignment si existe
        # El assignment_number se mantiene y pasa del candidato origen al destino
        transferred_assignment_number = None
        if ecm_id:
            # Buscar primero por group_exam_id, luego sin él como fallback
            eca = EcmCandidateAssignment.query.filter_by(
                user_id=from_user_id,
                competency_standard_id=ecm_id,
                group_exam_id=group_exam.id
            ).first()
            if not eca:
                # Fallback: buscar por user_id + competency_standard_id sin group_exam_id
                eca = EcmCandidateAssignment.query.filter_by(
                    user_id=from_user_id,
                    competency_standard_id=ecm_id
                ).first()
            
            if eca:
                transferred_assignment_number = eca.assignment_number
                
                # Verificar si el destino ya tiene una ECA para este ECM
                existing_dest_eca = EcmCandidateAssignment.query.filter_by(
                    user_id=to_user_id,
                    competency_standard_id=ecm_id
                ).first()
                
                if existing_dest_eca:
                    # El destino ya tiene una ECA — eliminar la del origen ya que 
                    # el destino conserva la suya. No transferimos assignment_number 
                    # en este caso raro.
                    db.session.delete(eca)
                    transferred_assignment_number = existing_dest_eca.assignment_number
                    print(f"⚠️ Destino ya tenía ECA {existing_dest_eca.assignment_number}, se eliminó ECA {eca.assignment_number} del origen")
                else:
                    # Transferir: cambiar user_id al destino
                    eca.user_id = to_user_id
                    eca.group_exam_id = group_exam.id  # Asegurar referencia correcta
                    from app.models.user import User
                    to_user = User.query.get(to_user_id)
                    if to_user:
                        print(f"✅ ECA {eca.assignment_number} transferida de {from_user_id} a {to_user.full_name}")

        # === REVOCAR ACCESO DEL ORIGEN ===
        # Al perder su número de asignación, el origen pierde acceso al examen
        # y se limpian sus registros de progreso de material de estudio.
        revoked_items = []

        # 1. Eliminar GroupExamMember del origen (revoca acceso al examen)
        origin_gem = GroupExamMember.query.filter_by(
            group_exam_id=group_exam.id,
            user_id=from_user_id
        ).first()
        if origin_gem:
            db.session.delete(origin_gem)
            revoked_items.append('exam_access')

        # 2. Eliminar progreso de material de estudio del origen
        try:
            if topic_ids:
                # Eliminar progreso por tema
                deleted_tp = StudentTopicProgress.query.filter(
                    StudentTopicProgress.user_id == from_user_id,
                    StudentTopicProgress.topic_id.in_(topic_ids)
                ).delete(synchronize_session='fetch')

                # Eliminar progreso por contenido
                deleted_cp = StudentContentProgress.query.filter(
                    StudentContentProgress.user_id == from_user_id,
                    StudentContentProgress.topic_id.in_(topic_ids)
                ).delete(synchronize_session='fetch')

                if deleted_tp or deleted_cp:
                    revoked_items.append(f'study_progress({deleted_tp}t,{deleted_cp}c)')
        except Exception as rev_err:
            print(f"⚠️ Error limpiando progreso del origen: {rev_err}")

        if revoked_items:
            print(f"🔒 Revocación para {from_user_id}: {', '.join(revoked_items)}")

        # === REGISTRAR EN HISTORIAL ===
        if transferred_assignment_number:
            current_user_id = get_jwt_identity()
            group_obj = CandidateGroup.query.get(group_id)
            swap_record = EcmSwapHistory(
                assignment_number=transferred_assignment_number,
                ecm_assignment_id=eca.id if eca else None,
                competency_standard_id=ecm_id,
                group_id=group_id,
                group_name=group_obj.name if group_obj else None,
                exam_id=exam_id,
                group_exam_id=group_exam.id,
                from_user_id=from_user_id,
                to_user_id=to_user_id,
                performed_by_id=current_user_id,
                swap_type='single',
            )
            db.session.add(swap_record)

        db.session.commit()

        from app.models.user import User
        from_user = User.query.get(from_user_id)
        to_user = User.query.get(to_user_id)

        return jsonify({
            'message': f'Asignación reasignada de {from_user.full_name if from_user else from_user_id} a {to_user.full_name if to_user else to_user_id}',
            'from_user_id': from_user_id,
            'to_user_id': to_user_id,
            'assignment_number': transferred_assignment_number,
        })

    except Exception as e:
        db.session.rollback()
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@bp.route('/groups/<int:group_id>/exams/<int:exam_id>/members/bulk-swap', methods=['POST'])
@jwt_required()
@coordinator_required
def bulk_swap_exam_members(group_id, exam_id):
    """Reasignación masiva: recibe un array de pares {from_user_id, to_user_id}.

    Body JSON:
    {
        "swaps": [
            {"from_user_id": "uuid-a", "to_user_id": "uuid-x"},
            {"from_user_id": "uuid-b", "to_user_id": "uuid-y"}
        ]
    }

    Todas las reasignaciones se ejecutan en una sola transacción.
    Si alguna falla la validación, se informa del error pero se procesan las demás.
    """
    try:
        from app.models import GroupExam
        from app.models.partner import GroupExamMember, EcmCandidateAssignment, EcmSwapHistory
        from app.models.result import Result
        from app.models.student_progress import StudentTopicProgress, StudentContentProgress
        from app.models.study_content import StudyMaterial
        from app.models.user import User
        from sqlalchemy import text

        data = request.get_json()
        swaps = data.get('swaps', [])

        if not swaps or not isinstance(swaps, list):
            return jsonify({'error': 'Se requiere un array "swaps" con al menos un par'}), 400

        if len(swaps) > 100:
            return jsonify({'error': 'Máximo 100 reasignaciones por lote'}), 400

        # Validar duplicados en el lote (antes de buscar en DB)
        from_ids = [s.get('from_user_id') for s in swaps if s.get('from_user_id')]
        to_ids = [s.get('to_user_id') for s in swaps if s.get('to_user_id')]

        if len(set(from_ids)) != len(from_ids):
            return jsonify({'error': 'Hay candidatos origen duplicados en el lote'}), 400
        if len(set(to_ids)) != len(to_ids):
            return jsonify({'error': 'Hay candidatos destino duplicados en el lote'}), 400

        overlap = set(from_ids) & set(to_ids)
        if overlap:
            return jsonify({'error': f'Los siguientes IDs aparecen como origen y destino a la vez: {list(overlap)}'}), 400

        group_exam = GroupExam.query.filter_by(
            group_id=group_id,
            exam_id=exam_id,
            is_active=True
        ).first_or_404()

        exam = group_exam.exam
        ecm_id = exam.competency_standard_id if exam else None

        # Pre-cargar datos necesarios para validaciones
        # Material de estudio y temas
        material_ids = []
        topic_ids = []
        try:
            material_ids_rows = db.session.execute(text(
                'SELECT study_material_id FROM study_material_exams WHERE exam_id = :eid'
            ), {'eid': exam_id}).fetchall()
            material_ids = [r[0] for r in material_ids_rows]

            legacy = StudyMaterial.query.filter_by(exam_id=exam_id).all()
            for m in legacy:
                if m.id not in material_ids:
                    material_ids.append(m.id)

            if material_ids:
                topic_rows = db.session.execute(text('''
                    SELECT DISTINCT st.id
                    FROM study_topics st
                    INNER JOIN study_sessions ss ON st.session_id = ss.id
                    WHERE ss.study_material_id IN :mids
                ''').bindparams(mids=tuple(material_ids) if len(material_ids) > 1 else tuple(material_ids + [0]))).fetchall()
                topic_ids = [r[0] for r in topic_rows]
        except Exception as e:
            print(f"⚠️ Error pre-cargando material: {e}")

        results = []
        errors = []
        converted_to_selected = False

        for idx, swap in enumerate(swaps):
            from_user_id = swap.get('from_user_id')
            to_user_id = swap.get('to_user_id')

            if not from_user_id or not to_user_id:
                errors.append({'index': idx, 'error': 'Faltan from_user_id o to_user_id'})
                continue

            if from_user_id == to_user_id:
                errors.append({'index': idx, 'from_user_id': from_user_id, 'error': 'Origen y destino son iguales'})
                continue

            # Verificar que from_user está asignado
            if group_exam.assignment_type == 'selected':
                from_gem = GroupExamMember.query.filter_by(
                    group_exam_id=group_exam.id, user_id=from_user_id
                ).first()
                if not from_gem:
                    from_u = User.query.get(from_user_id)
                    errors.append({'index': idx, 'from_user_id': from_user_id,
                                   'error': f'{from_u.full_name if from_u else from_user_id} no está asignado a este examen'})
                    continue
            else:
                is_member = GroupMember.query.filter_by(group_id=group_id, user_id=from_user_id).first()
                if not is_member:
                    errors.append({'index': idx, 'from_user_id': from_user_id, 'error': 'No es miembro del grupo'})
                    continue

            # Verificar destino es miembro del grupo
            to_gm = GroupMember.query.filter_by(group_id=group_id, user_id=to_user_id).first()
            if not to_gm:
                to_u = User.query.get(to_user_id)
                errors.append({'index': idx, 'to_user_id': to_user_id,
                               'error': f'{to_u.full_name if to_u else to_user_id} no es miembro del grupo'})
                continue

            # Validar números de asignación: origen DEBE tener, destino NO debe tener
            # Nota: No se valida GroupExamMember del destino porque un miembro del
            # examen puede no tener número de asignación y ser destino válido.
            if ecm_id:
                from_eca_check = EcmCandidateAssignment.query.filter_by(
                    user_id=from_user_id, competency_standard_id=ecm_id
                ).first()
                if not from_eca_check:
                    from_u = User.query.get(from_user_id)
                    errors.append({'index': idx, 'from_user_id': from_user_id,
                                   'error': f'{from_u.full_name if from_u else from_user_id} no tiene número de asignación'})
                    continue

                dest_eca_check = EcmCandidateAssignment.query.filter_by(
                    user_id=to_user_id, competency_standard_id=ecm_id
                ).first()
                if dest_eca_check:
                    to_u = User.query.get(to_user_id)
                    errors.append({'index': idx, 'to_user_id': to_user_id,
                                   'error': f'{to_u.full_name if to_u else to_user_id} ya tiene número de asignación'})
                    continue

            # Validar que origen puede ser reasignado
            has_results = Result.query.filter(
                Result.user_id == from_user_id,
                Result.group_exam_id == group_exam.id
            ).first()
            if has_results:
                from_u = User.query.get(from_user_id)
                errors.append({'index': idx, 'from_user_id': from_user_id,
                               'error': f'{from_u.full_name if from_u else from_user_id} ya abrió el examen/simulador'})
                continue

            # Verificar progreso >= 15%
            if topic_ids:
                progress_records = StudentTopicProgress.query.filter(
                    StudentTopicProgress.user_id == from_user_id,
                    StudentTopicProgress.topic_id.in_(topic_ids)
                ).all()
                total_topics = len(topic_ids)
                sum_pct = sum(pr.progress_percentage or 0.0 for pr in progress_records)
                avg_progress = sum_pct / total_topics if total_topics > 0 else 0.0
                if avg_progress >= 15.0:
                    from_u = User.query.get(from_user_id)
                    errors.append({'index': idx, 'from_user_id': from_user_id,
                                   'error': f'{from_u.full_name if from_u else from_user_id} tiene {round(avg_progress, 1)}% de avance'})
                    continue

            # === REALIZAR EL SWAP ===
            if group_exam.assignment_type == 'selected':
                # Asegurar que el destino tenga GroupExamMember
                # (el origen se eliminará en la sección de revocación)
                existing_dest_gem = GroupExamMember.query.filter_by(
                    group_exam_id=group_exam.id, user_id=to_user_id
                ).first()
                if not existing_dest_gem:
                    db.session.add(GroupExamMember(group_exam_id=group_exam.id, user_id=to_user_id))
            else:
                # Convertir a 'selected' solo una vez
                if not converted_to_selected:
                    all_members = GroupMember.query.filter_by(group_id=group_id).all()
                    group_exam.assignment_type = 'selected'
                    for gm in all_members:
                        existing = GroupExamMember.query.filter_by(
                            group_exam_id=group_exam.id, user_id=gm.user_id
                        ).first()
                        if not existing:
                            db.session.add(GroupExamMember(group_exam_id=group_exam.id, user_id=gm.user_id))
                    converted_to_selected = True

                # Asegurar que destino tenga GroupExamMember
                existing_to = GroupExamMember.query.filter_by(
                    group_exam_id=group_exam.id, user_id=to_user_id
                ).first()
                if not existing_to:
                    db.session.add(GroupExamMember(group_exam_id=group_exam.id, user_id=to_user_id))

            # Transferir EcmCandidateAssignment
            transferred_number = None
            if ecm_id:
                eca = EcmCandidateAssignment.query.filter_by(
                    user_id=from_user_id, competency_standard_id=ecm_id, group_exam_id=group_exam.id
                ).first()
                if not eca:
                    eca = EcmCandidateAssignment.query.filter_by(
                        user_id=from_user_id, competency_standard_id=ecm_id
                    ).first()

                if eca:
                    transferred_number = eca.assignment_number
                    existing_dest_eca = EcmCandidateAssignment.query.filter_by(
                        user_id=to_user_id, competency_standard_id=ecm_id
                    ).first()
                    if existing_dest_eca:
                        db.session.delete(eca)
                        transferred_number = existing_dest_eca.assignment_number
                    else:
                        eca.user_id = to_user_id
                        eca.group_exam_id = group_exam.id

            # === REVOCAR ACCESO DEL ORIGEN ===
            revoked_items = []

            # 1. Eliminar GroupExamMember del origen (revoca acceso al examen)
            origin_gem = GroupExamMember.query.filter_by(
                group_exam_id=group_exam.id,
                user_id=from_user_id
            ).first()
            if origin_gem:
                db.session.delete(origin_gem)
                revoked_items.append('exam_access')

            # 2. Eliminar progreso de material de estudio del origen
            try:
                if topic_ids:
                    deleted_tp = StudentTopicProgress.query.filter(
                        StudentTopicProgress.user_id == from_user_id,
                        StudentTopicProgress.topic_id.in_(topic_ids)
                    ).delete(synchronize_session='fetch')

                    deleted_cp = StudentContentProgress.query.filter(
                        StudentContentProgress.user_id == from_user_id,
                        StudentContentProgress.topic_id.in_(topic_ids)
                    ).delete(synchronize_session='fetch')

                    if deleted_tp or deleted_cp:
                        revoked_items.append(f'study_progress({deleted_tp}t,{deleted_cp}c)')
            except Exception as rev_err:
                print(f"⚠️ Error limpiando progreso del origen {from_user_id}: {rev_err}")

            if revoked_items:
                print(f"🔒 Revocación bulk para {from_user_id}: {', '.join(revoked_items)}")

            # === REGISTRAR EN HISTORIAL ===
            if transferred_number:
                current_user_id = get_jwt_identity()
                group_obj = CandidateGroup.query.get(group_id)
                db.session.add(EcmSwapHistory(
                    assignment_number=transferred_number,
                    ecm_assignment_id=eca.id if eca else None,
                    competency_standard_id=ecm_id,
                    group_id=group_id,
                    group_name=group_obj.name if group_obj else None,
                    exam_id=exam_id,
                    group_exam_id=group_exam.id,
                    from_user_id=from_user_id,
                    to_user_id=to_user_id,
                    performed_by_id=current_user_id,
                    swap_type='bulk',
                ))

            from_u = User.query.get(from_user_id)
            to_u = User.query.get(to_user_id)
            results.append({
                'from_user_id': from_user_id,
                'from_name': from_u.full_name if from_u else from_user_id,
                'to_user_id': to_user_id,
                'to_name': to_u.full_name if to_u else to_user_id,
                'assignment_number': transferred_number,
            })

        if results:
            db.session.commit()
        else:
            db.session.rollback()

        return jsonify({
            'message': f'{len(results)} reasignación(es) completada(s)' + (f', {len(errors)} error(es)' if errors else ''),
            'success_count': len(results),
            'error_count': len(errors),
            'results': results,
            'errors': errors,
        })

    except Exception as e:
        db.session.rollback()
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


# ============== HISTORIAL DE REASIGNACIONES ==============

@bp.route('/groups/<int:group_id>/exams/<int:exam_id>/swap-history', methods=['GET'])
@jwt_required()
@coordinator_required
def get_swap_history(group_id, exam_id):
    """Obtener el historial de reasignaciones para un examen de grupo.
    
    Query params:
    - page (int): Página actual (default 1)
    - per_page (int): Registros por página (default 50)
    - assignment_number (str): Filtrar por número de asignación
    - user_id (str): Filtrar por usuario (origen o destino)
    - sort (str): Campo de orden (default 'performed_at')
    - dir (str): Dirección 'asc' o 'desc' (default 'desc')
    """
    try:
        from app.models.partner import EcmSwapHistory

        page = request.args.get('page', 1, type=int)
        per_page = min(request.args.get('per_page', 50, type=int), 200)
        assignment_number = request.args.get('assignment_number', '').strip()
        user_id_filter = request.args.get('user_id', '').strip()
        sort_field = request.args.get('sort', 'performed_at')
        sort_dir = request.args.get('dir', 'desc')

        query = EcmSwapHistory.query.filter_by(
            group_id=group_id,
            exam_id=exam_id,
        )

        if assignment_number:
            query = query.filter(EcmSwapHistory.assignment_number.ilike(f'%{assignment_number}%'))

        if user_id_filter:
            from sqlalchemy import or_
            query = query.filter(or_(
                EcmSwapHistory.from_user_id == user_id_filter,
                EcmSwapHistory.to_user_id == user_id_filter,
            ))

        # Ordenamiento
        sort_map = {
            'performed_at': EcmSwapHistory.performed_at,
            'assignment_number': EcmSwapHistory.assignment_number,
        }
        sort_column = sort_map.get(sort_field, EcmSwapHistory.performed_at)
        if sort_dir == 'asc':
            query = query.order_by(sort_column.asc())
        else:
            query = query.order_by(sort_column.desc())

        pagination = query.paginate(page=page, per_page=per_page, error_out=False)

        # Agrupar por assignment_number para la línea de tiempo
        records = [r.to_dict() for r in pagination.items]

        return jsonify({
            'history': records,
            'total': pagination.total,
            'page': pagination.page,
            'pages': pagination.pages,
            'per_page': per_page,
        })

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@bp.route('/groups/<int:group_id>/exams/<int:exam_id>/swap-history/timeline', methods=['GET'])
@jwt_required()
@coordinator_required
def get_swap_timeline(group_id, exam_id):
    """Obtener la trazabilidad agrupada por número de asignación.
    
    Devuelve cada assignment_number con su cadena completa de movimientos.
    """
    try:
        from app.models.partner import EcmSwapHistory, EcmCandidateAssignment
        from app.models import GroupExam

        group_exam = GroupExam.query.filter_by(
            group_id=group_id,
            exam_id=exam_id,
            is_active=True
        ).first_or_404()

        exam = group_exam.exam
        ecm_id = exam.competency_standard_id if exam else None

        # Todos los registros de historial para este grupo/examen
        all_records = EcmSwapHistory.query.filter_by(
            group_id=group_id,
            exam_id=exam_id,
        ).order_by(EcmSwapHistory.performed_at.asc()).all()

        # Agrupar por assignment_number
        timeline = {}
        for r in all_records:
            an = r.assignment_number
            if an not in timeline:
                timeline[an] = {
                    'assignment_number': an,
                    'moves': [],
                    'current_holder': None,
                }
            timeline[an]['moves'].append(r.to_dict())

        # Enriquecer con propietario actual
        if ecm_id:
            for an, data in timeline.items():
                eca = EcmCandidateAssignment.query.filter_by(
                    assignment_number=an,
                    competency_standard_id=ecm_id,
                ).first()
                if eca and eca.user:
                    data['current_holder'] = {
                        'user_id': eca.user_id,
                        'full_name': eca.user.full_name,
                        'email': eca.user.email,
                    }

        return jsonify({
            'timeline': list(timeline.values()),
            'total_assignments': len(timeline),
            'total_moves': len(all_records),
        })

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


# ============== RETOMAS ECM ==============

@bp.route('/groups/<int:group_id>/exams/<int:exam_id>/members/<string:user_id>/retake', methods=['POST'])
@jwt_required()
@coordinator_required
def apply_ecm_retake(group_id, exam_id, user_id):
    """Aplicar una retoma a un candidato que agotó sus intentos.
    
    La retoma:
    1. Verifica que el candidato agotó todos sus intentos (max_attempts)
    2. Verifica que no se excede el máximo de retomas configurado
    3. Cobra retake_cost del saldo del coordinador
    4. Da 1 intento adicional al candidato
    """
    try:
        from app.models import GroupExam
        from app.models.partner import EcmCandidateAssignment, EcmRetake
        from app.models.result import Result
        from app.models.balance import CoordinatorBalance, create_balance_transaction

        group = CandidateGroup.query.get_or_404(group_id)
        campus = Campus.query.get(group.campus_id)

        group_exam = GroupExam.query.filter_by(
            group_id=group_id,
            exam_id=exam_id,
            is_active=True
        ).first()
        if not group_exam:
            return jsonify({'error': 'Asignación de examen no encontrada'}), 404

        exam = group_exam.exam
        ecm_id = exam.competency_standard_id if exam else None
        if not ecm_id:
            return jsonify({'error': 'Este examen no tiene un ECM asignado. Las retomas solo aplican a certificaciones ECM.'}), 400

        # Buscar la asignación ECM del candidato
        ecm_assignment = EcmCandidateAssignment.query.filter_by(
            user_id=user_id,
            competency_standard_id=ecm_id
        ).first()
        if not ecm_assignment:
            return jsonify({'error': 'El candidato no tiene una asignación ECM para este estándar'}), 404

        # Contar resultados del candidato en este group_exam
        results_count = Result.query.filter_by(
            user_id=user_id,
            group_exam_id=group_exam.id
        ).count()

        # Contar retomas activas existentes para este candidato en este group_exam
        active_retakes = EcmRetake.query.filter_by(
            assignment_id=ecm_assignment.id,
            group_exam_id=group_exam.id
        ).filter(EcmRetake.status.in_(['active', 'used'])).count()

        # Los intentos totales = max_attempts original + retomas activas/usadas
        total_allowed = (group_exam.max_attempts or 1) + active_retakes
        
        if results_count < total_allowed:
            remaining = total_allowed - results_count
            return jsonify({
                'error': f'El candidato aún tiene {remaining} intento(s) disponible(s). La retoma solo se puede aplicar cuando se agotan todos los intentos.',
                'results_count': results_count,
                'total_allowed': total_allowed,
                'remaining': remaining
            }), 400

        # Verificar máximo de retomas (0 = ilimitado)
        max_retakes = 0  # Default: ilimitado
        if group.max_retakes_override is not None:
            max_retakes = group.max_retakes_override
        elif campus and campus.max_retakes is not None:
            max_retakes = campus.max_retakes
        
        if max_retakes > 0 and active_retakes >= max_retakes:
            return jsonify({
                'error': f'Se alcanzó el máximo de retomas permitidas ({max_retakes}) para esta asignación.',
                'max_retakes': max_retakes,
                'current_retakes': active_retakes
            }), 400

        # Verificar que no haya aprobado ya
        approved_result = Result.query.filter_by(
            user_id=user_id,
            group_exam_id=group_exam.id,
            result=1  # aprobado
        ).first()
        if approved_result:
            return jsonify({'error': 'El candidato ya aprobó este examen. No necesita retoma.'}), 400

        # Calcular costo de retoma
        if group.retake_cost_override is not None:
            retake_cost = float(group.retake_cost_override)
        elif campus and campus.retake_cost is not None:
            retake_cost = float(campus.retake_cost)
        else:
            retake_cost = 0.0

        # Verificar saldo del coordinador para este grupo
        # Admins y developers no requieren verificación de saldo
        user_role = g.current_user.role if hasattr(g.current_user, 'role') else ''
        is_admin_or_dev = user_role in ('admin', 'developer')
        
        transaction_id = None
        if retake_cost > 0 and not is_admin_or_dev:
            coordinator_id = g.current_user.id
            balance = CoordinatorBalance.query.filter_by(
                coordinator_id=coordinator_id,
                group_id=group_id
            ).first()
            current_balance = float(balance.current_balance) if balance else 0.0
            
            if current_balance < retake_cost:
                return jsonify({
                    'error': f'Saldo insuficiente para este grupo. La retoma cuesta ${retake_cost:,.2f} pero tu saldo es ${current_balance:,.2f}',
                    'error_type': 'insufficient_balance',
                    'required': retake_cost,
                    'available': current_balance,
                    'deficit': retake_cost - current_balance
                }), 400

            # Deducir saldo
            user_obj = User.query.get(user_id)
            user_name = user_obj.full_name if user_obj else user_id
            exam_name = exam.name if exam else f'Examen #{exam_id}'
            notes = f'Retoma #{active_retakes + 1} de "{exam_name}" para {user_name} (Asignación {ecm_assignment.assignment_number}) en grupo "{group.name}"'
            
            transaction = create_balance_transaction(
                coordinator_id=coordinator_id,
                group_id=group_id,
                transaction_type='debit',
                concept='asignacion_retoma',
                amount=retake_cost,
                reference_type='group_exam',
                reference_id=group_exam.id,
                notes=notes,
                created_by_id=coordinator_id
            )
            if transaction:
                transaction_id = transaction.id

        # Crear la retoma
        retake = EcmRetake(
            assignment_id=ecm_assignment.id,
            group_exam_id=group_exam.id,
            user_id=user_id,
            cost=retake_cost,
            transaction_id=transaction_id,
            status='active',
            applied_by_id=g.current_user.id
        )
        db.session.add(retake)
        db.session.commit()

        user_obj = User.query.get(user_id)
        return jsonify({
            'message': f'Retoma aplicada exitosamente para {user_obj.full_name if user_obj else user_id}',
            'retake': retake.to_dict(),
            'retake_number': active_retakes + 1,
            'max_retakes': max_retakes,
            'cost': retake_cost,
            'new_total_attempts': total_allowed + 1,
        }), 201

    except Exception as e:
        db.session.rollback()
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@bp.route('/groups/<int:group_id>/exams/<int:exam_id>/retake-preview', methods=['POST'])
@jwt_required()
@coordinator_required
def preview_ecm_retake(group_id, exam_id):
    """Preview del costo de una retoma antes de aplicarla.
    
    Body JSON: { "user_id": "uuid" }
    """
    try:
        from app.models import GroupExam
        from app.models.partner import EcmCandidateAssignment, EcmRetake
        from app.models.result import Result

        data = request.get_json()
        user_id = data.get('user_id')
        if not user_id:
            return jsonify({'error': 'Se requiere user_id'}), 400

        group = CandidateGroup.query.get_or_404(group_id)
        campus = Campus.query.get(group.campus_id)
        
        group_exam = GroupExam.query.filter_by(
            group_id=group_id, exam_id=exam_id, is_active=True
        ).first()
        if not group_exam:
            return jsonify({'error': 'Asignación no encontrada'}), 404

        exam = group_exam.exam
        ecm_id = exam.competency_standard_id if exam else None
        
        ecm_assignment = EcmCandidateAssignment.query.filter_by(
            user_id=user_id, competency_standard_id=ecm_id
        ).first() if ecm_id else None

        # Contar retomas existentes
        active_retakes = 0
        if ecm_assignment:
            active_retakes = EcmRetake.query.filter_by(
                assignment_id=ecm_assignment.id,
                group_exam_id=group_exam.id
            ).filter(EcmRetake.status.in_(['active', 'used'])).count()

        # Resultados
        results_count = Result.query.filter_by(
            user_id=user_id, group_exam_id=group_exam.id
        ).count()
        total_allowed = (group_exam.max_attempts or 1) + active_retakes

        # Configuración
        max_retakes = group.max_retakes_override if group.max_retakes_override is not None else (campus.max_retakes if campus and campus.max_retakes is not None else 0)
        retake_cost = float(group.retake_cost_override) if group.retake_cost_override is not None else (float(campus.retake_cost) if campus and campus.retake_cost else 0)

        # Saldo del grupo específico
        from app.models.balance import CoordinatorBalance
        balance = CoordinatorBalance.query.filter_by(
            coordinator_id=g.current_user.id,
            group_id=group_id
        ).first()
        current_balance = float(balance.current_balance) if balance else 0.0

        # Verificar si aprobó
        approved = Result.query.filter_by(
            user_id=user_id, group_exam_id=group_exam.id, result=1
        ).first() is not None

        can_apply = (
            ecm_id is not None
            and ecm_assignment is not None
            and results_count >= total_allowed
            and (max_retakes == 0 or active_retakes < max_retakes)
            and not approved
            and current_balance >= retake_cost
        )

        user_obj = User.query.get(user_id)
        return jsonify({
            'can_apply': can_apply,
            'user_name': user_obj.full_name if user_obj else user_id,
            'assignment_number': ecm_assignment.assignment_number if ecm_assignment else None,
            'retake_cost': retake_cost,
            'current_balance': current_balance,
            'sufficient_balance': current_balance >= retake_cost,
            'results_count': results_count,
            'total_allowed': total_allowed,
            'attempts_remaining': max(0, total_allowed - results_count),
            'current_retakes': active_retakes,
            'max_retakes': max_retakes,
            'retakes_remaining': -1 if max_retakes == 0 else max(0, max_retakes - active_retakes),
            'has_approved': approved,
            'has_ecm': ecm_id is not None,
            'has_assignment': ecm_assignment is not None,
            'reasons': _get_retake_block_reasons(
                ecm_id, ecm_assignment, results_count, total_allowed,
                active_retakes, max_retakes, approved, current_balance, retake_cost
            )
        })

    except Exception as e:
        return jsonify({'error': str(e)}), 500


def _get_retake_block_reasons(ecm_id, ecm_assignment, results_count, total_allowed,
                               active_retakes, max_retakes, approved, current_balance, retake_cost):
    """Devuelve lista de razones por las que NO se puede aplicar retoma."""
    reasons = []
    if not ecm_id:
        reasons.append('El examen no tiene un ECM asignado')
    if not ecm_assignment:
        reasons.append('El candidato no tiene asignación ECM')
    if results_count < total_allowed:
        remaining = total_allowed - results_count
        reasons.append(f'Aún tiene {remaining} intento(s) disponible(s)')
    if active_retakes >= max_retakes and max_retakes > 0:
        reasons.append(f'Alcanzó el máximo de retomas ({max_retakes})')
    if approved:
        reasons.append('Ya aprobó el examen')
    if retake_cost > 0 and current_balance < retake_cost:
        reasons.append(f'Saldo insuficiente (necesita ${retake_cost:,.2f}, tiene ${current_balance:,.2f})')
    return reasons


# ============== MATERIALES DE ESTUDIO SIN EXAMEN ==============

@bp.route('/groups/<int:group_id>/study-materials', methods=['GET'])
@jwt_required()
@coordinator_required
def get_group_study_materials(group_id):
    """Obtener materiales de estudio asignados directamente al grupo (sin examen)"""
    try:
        from app.models import GroupStudyMaterial
        
        group = CandidateGroup.query.get_or_404(group_id)
        
        assignments = GroupStudyMaterial.query.filter_by(
            group_id=group_id,
            is_active=True
        ).order_by(GroupStudyMaterial.assigned_at.desc()).all()
        
        materials_data = [a.to_dict(include_material=True, include_members=True) for a in assignments]
        
        return jsonify({
            'group_id': group_id,
            'group_name': group.name,
            'assigned_materials': materials_data,
            'total': len(materials_data)
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@bp.route('/groups/<int:group_id>/study-materials', methods=['POST'])
@jwt_required()
@coordinator_required
def assign_study_materials_to_group(group_id):
    """Asignar materiales de estudio directamente a un grupo (sin examen)
    
    Parámetros:
    - material_ids: Lista de IDs de materiales a asignar
    - assignment_type: 'all' (todo el grupo) o 'selected' (candidatos específicos)
    - member_ids: Lista de IDs de usuarios (solo si assignment_type='selected')
    - available_from, available_until: Período de disponibilidad (opcional)
    """
    try:
        from app.models import GroupStudyMaterial, GroupStudyMaterialMember
        from app.models.study_content import StudyMaterial
        
        group = CandidateGroup.query.get_or_404(group_id)
        data = request.get_json()
        
        material_ids = data.get('material_ids', [])
        if not material_ids:
            return jsonify({'error': 'Debes seleccionar al menos un material de estudio'}), 400
        
        assignment_type = data.get('assignment_type', 'all')
        member_ids = data.get('member_ids', [])
        
        if assignment_type == 'selected' and not member_ids:
            return jsonify({'error': 'Debes seleccionar al menos un candidato'}), 400
        
        # Verificar que los materiales existen y están publicados
        materials = StudyMaterial.query.filter(
            StudyMaterial.id.in_(material_ids),
            StudyMaterial.is_published == True
        ).all()
        
        if len(materials) != len(material_ids):
            return jsonify({'error': 'Algunos materiales no existen o no están publicados'}), 400
        
        # Fechas de disponibilidad
        available_from = None
        available_until = None
        if data.get('available_from'):
            available_from = datetime.fromisoformat(data['available_from'].replace('Z', '+00:00'))
        if data.get('available_until'):
            available_until = datetime.fromisoformat(data['available_until'].replace('Z', '+00:00'))
        
        user_id = get_jwt_identity()
        assignments_created = []
        
        for material in materials:
            # Verificar si ya existe una asignación
            existing = GroupStudyMaterial.query.filter_by(
                group_id=group_id,
                study_material_id=material.id
            ).first()
            
            if existing:
                # Reactivar si estaba inactivo
                if not existing.is_active:
                    existing.is_active = True
                    existing.assigned_at = datetime.utcnow()
                    existing.assigned_by_id = user_id
                    existing.available_from = available_from
                    existing.available_until = available_until
                    existing.assignment_type = assignment_type
                    
                    # Actualizar miembros si es necesario
                    if assignment_type == 'selected':
                        GroupStudyMaterialMember.query.filter_by(group_study_material_id=existing.id).delete()
                        for mid in member_ids:
                            member = GroupStudyMaterialMember(
                                group_study_material_id=existing.id,
                                user_id=mid
                            )
                            db.session.add(member)
                    
                    assignments_created.append(existing)
            else:
                # Crear nueva asignación
                assignment = GroupStudyMaterial(
                    group_id=group_id,
                    study_material_id=material.id,
                    assigned_by_id=user_id,
                    available_from=available_from,
                    available_until=available_until,
                    assignment_type=assignment_type,
                    is_active=True
                )
                db.session.add(assignment)
                db.session.flush()  # Para obtener el ID
                
                # Agregar miembros si es asignación selectiva
                if assignment_type == 'selected':
                    for mid in member_ids:
                        member = GroupStudyMaterialMember(
                            group_study_material_id=assignment.id,
                            user_id=mid
                        )
                        db.session.add(member)
                
                assignments_created.append(assignment)
        
        db.session.commit()
        
        message = f'{len(assignments_created)} material(es) de estudio asignado(s) exitosamente.'
        if assignment_type == 'all':
            message += f' Disponible(s) para los {group.members.count()} miembros del grupo.'
        else:
            message += f' Disponible(s) para {len(member_ids)} candidato(s) seleccionado(s).'
        
        return jsonify({
            'message': message,
            'assignments': [a.to_dict(include_material=True) for a in assignments_created],
            'materials_count': len(assignments_created)
        }), 201
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500


@bp.route('/groups/<int:group_id>/study-materials/<int:material_id>', methods=['DELETE'])
@jwt_required()
@coordinator_required
def unassign_study_material_from_group(group_id, material_id):
    """Desasignar un material de estudio del grupo"""
    try:
        from app.models import GroupStudyMaterial
        
        assignment = GroupStudyMaterial.query.filter_by(
            group_id=group_id,
            study_material_id=material_id
        ).first_or_404()
        
        assignment.is_active = False
        db.session.commit()
        
        return jsonify({
            'message': 'Material de estudio desasignado del grupo'
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500


@bp.route('/groups/<int:group_id>/study-materials/<int:material_id>/members', methods=['GET'])
@jwt_required()
@coordinator_required
def get_study_material_members(group_id, material_id):
    """Obtener los miembros asignados a un material específico del grupo"""
    try:
        from app.models import GroupStudyMaterial, GroupStudyMaterialMember
        
        assignment = GroupStudyMaterial.query.filter_by(
            group_id=group_id,
            study_material_id=material_id,
            is_active=True
        ).first()
        
        if not assignment:
            return jsonify({
                'error': f'No se encontró asignación activa del material {material_id} en el grupo {group_id}'
            }), 404
        
        if assignment.assignment_type == 'all':
            # Todos los miembros del grupo
            group_members = GroupMember.query.filter_by(group_id=group_id).all()
            assigned_user_ids = [m.user_id for m in group_members]
        else:
            # Solo los seleccionados
            assigned_members = GroupStudyMaterialMember.query.filter_by(
                group_study_material_id=assignment.id
            ).all()
            assigned_user_ids = [m.user_id for m in assigned_members]
        
        return jsonify({
            'assignment_id': assignment.id,
            'material_id': material_id,
            'assignment_type': assignment.assignment_type,
            'assigned_user_ids': assigned_user_ids
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@bp.route('/groups/<int:group_id>/study-materials/<int:material_id>/members', methods=['PUT'])
@jwt_required()
@coordinator_required
def update_study_material_members(group_id, material_id):
    """Actualizar los miembros asignados a un material específico del grupo
    
    Parámetros:
    - user_ids: Lista de user_ids que deben estar asignados
    """
    try:
        from app.models import GroupStudyMaterial, GroupStudyMaterialMember
        
        assignment = GroupStudyMaterial.query.filter_by(
            group_id=group_id,
            study_material_id=material_id,
            is_active=True
        ).first()
        
        if not assignment:
            return jsonify({
                'error': f'No se encontró asignación activa del material {material_id} en el grupo {group_id}'
            }), 404
        
        data = request.get_json()
        new_user_ids = set(data.get('user_ids', []))
        
        if not new_user_ids:
            return jsonify({'error': 'Debes seleccionar al menos un candidato'}), 400
        
        # Cambiar a assignment_type 'selected' si era 'all'
        if assignment.assignment_type == 'all':
            assignment.assignment_type = 'selected'
        
        # Obtener asignaciones actuales
        current_members = GroupStudyMaterialMember.query.filter_by(
            group_study_material_id=assignment.id
        ).all()
        current_user_ids = {m.user_id for m in current_members}
        
        # Usuarios a agregar
        to_add = new_user_ids - current_user_ids
        # Usuarios a eliminar
        to_remove = current_user_ids - new_user_ids
        
        # Agregar nuevos
        for user_id in to_add:
            member = GroupStudyMaterialMember(
                group_study_material_id=assignment.id,
                user_id=user_id
            )
            db.session.add(member)
        
        # Eliminar los que ya no están
        for user_id in to_remove:
            GroupStudyMaterialMember.query.filter_by(
                group_study_material_id=assignment.id,
                user_id=user_id
            ).delete()
        
        db.session.commit()
        
        return jsonify({
            'message': 'Miembros actualizados correctamente',
            'added': list(to_add),
            'removed': list(to_remove),
            'total_members': len(new_user_ids)
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500


@bp.route('/groups/<int:group_id>/study-materials/<int:material_id>/members/add', methods=['POST'])
@jwt_required()
@coordinator_required
def add_members_to_study_material(group_id, material_id):
    """Agregar miembros a un material de estudio existente sin afectar los actuales
    
    Parámetros:
    - user_ids: Lista de user_ids a agregar
    """
    try:
        from app.models import GroupStudyMaterial, GroupStudyMaterialMember
        
        assignment = GroupStudyMaterial.query.filter_by(
            group_id=group_id,
            study_material_id=material_id,
            is_active=True
        ).first()
        
        if not assignment:
            return jsonify({
                'error': f'No se encontró asignación activa del material {material_id} en el grupo {group_id}'
            }), 404
        
        data = request.get_json()
        user_ids_to_add = data.get('user_ids', [])
        
        if not user_ids_to_add:
            return jsonify({'error': 'Debes proporcionar user_ids a agregar'}), 400
        
        # Cambiar a 'selected' si era 'all'
        if assignment.assignment_type == 'all':
            # Si era 'all', primero agregar todos los miembros actuales
            group_members = GroupMember.query.filter_by(group_id=group_id).all()
            for gm in group_members:
                existing = GroupStudyMaterialMember.query.filter_by(
                    group_study_material_id=assignment.id,
                    user_id=gm.user_id
                ).first()
                if not existing:
                    member = GroupStudyMaterialMember(
                        group_study_material_id=assignment.id,
                        user_id=gm.user_id
                    )
                    db.session.add(member)
            assignment.assignment_type = 'selected'
        
        added = []
        for user_id in user_ids_to_add:
            existing = GroupStudyMaterialMember.query.filter_by(
                group_study_material_id=assignment.id,
                user_id=user_id
            ).first()
            if not existing:
                member = GroupStudyMaterialMember(
                    group_study_material_id=assignment.id,
                    user_id=user_id
                )
                db.session.add(member)
                added.append(user_id)
        
        db.session.commit()
        
        return jsonify({
            'message': f'{len(added)} usuario(s) agregado(s) al material',
            'added': added
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500


@bp.route('/study-materials/available', methods=['GET'])
@jwt_required()
@coordinator_required
def get_available_study_materials():
    """Obtener materiales de estudio publicados disponibles para asignar"""
    try:
        from app.models.study_content import StudyMaterial, study_material_exams
        from app.models import GroupStudyMaterial, GroupExam
        
        search = request.args.get('search', '')
        page = request.args.get('page', 1, type=int)
        per_page = request.args.get('per_page', 20, type=int)
        group_id = request.args.get('group_id', type=int)
        
        # Obtener IDs de materiales ya asignados al grupo directamente
        assigned_material_ids = set()
        # Obtener materiales que están en exámenes asignados al grupo
        materials_in_assigned_exams = {}  # material_id -> exam_info
        
        if group_id:
            # Materiales asignados directamente
            assigned_materials = GroupStudyMaterial.query.filter_by(
                group_id=group_id,
                is_active=True
            ).all()
            assigned_material_ids = {gm.study_material_id for gm in assigned_materials}
            
            # Obtener exámenes asignados al grupo
            assigned_exams = GroupExam.query.filter_by(
                group_id=group_id,
                is_active=True
            ).all()
            
            # Para cada examen asignado, obtener sus materiales vinculados
            for group_exam in assigned_exams:
                exam = group_exam.exam
                if exam:
                    # Obtener materiales vinculados a este examen (tabla study_material_exams)
                    try:
                        result = db.session.execute(
                            db.select(study_material_exams.c.study_material_id).where(
                                study_material_exams.c.exam_id == exam.id
                            )
                        )
                        for row in result:
                            material_id = row[0]
                            if material_id not in materials_in_assigned_exams:
                                materials_in_assigned_exams[material_id] = {
                                    'exam_id': exam.id,
                                    'exam_name': exam.name,
                                    'group_exam_id': group_exam.id
                                }
                    except Exception:
                        pass  # Tabla puede no existir
                    
                    # También verificar materiales con relación legacy (exam_id directo)
                    legacy_materials = StudyMaterial.query.filter_by(exam_id=exam.id).all()
                    for mat in legacy_materials:
                        if mat.id not in materials_in_assigned_exams:
                            materials_in_assigned_exams[mat.id] = {
                                'exam_id': exam.id,
                                'exam_name': exam.name,
                                'group_exam_id': group_exam.id
                            }
        
        query = StudyMaterial.query.filter(StudyMaterial.is_published == True)
        
        if search:
            search_term = f'%{search}%'
            query = query.filter(
                db.or_(
                    StudyMaterial.title.ilike(search_term),
                    StudyMaterial.description.ilike(search_term)
                )
            )
        
        query = query.order_by(StudyMaterial.title)
        pagination = query.paginate(page=page, per_page=per_page, max_per_page=1000, error_out=False)
        
        materials_data = []
        for mat in pagination.items:
            sessions_count = mat.sessions.count() if mat.sessions else 0
            topics_count = sum(s.topics.count() for s in mat.sessions.all()) if mat.sessions else 0
            
            # Verificar si está en un examen asignado
            exam_info = materials_in_assigned_exams.get(mat.id)
            
            materials_data.append({
                'id': mat.id,
                'title': mat.title,
                'description': mat.description,
                'image_url': mat.image_url,
                'is_published': mat.is_published,
                'sessions_count': sessions_count,
                'topics_count': topics_count,
                'is_assigned_to_group': mat.id in assigned_material_ids,
                'is_in_assigned_exam': exam_info is not None,
                'assigned_exam_info': exam_info,
            })
        
        return jsonify({
            'materials': materials_data,
            'total': pagination.total,
            'pages': pagination.pages,
            'current_page': page
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@bp.route('/ecms/available', methods=['GET'])
@jwt_required()
@coordinator_required
def get_available_ecms():
    """Obtener ECMs disponibles para el campus de un grupo, con conteo de exámenes publicados.
    
    Parámetros:
    - group_id (requerido): ID del grupo para determinar el campus y sus ECMs
    - search: Buscar por código o nombre de ECM
    """
    try:
        from app.models import Exam
        from app.models.competency_standard import CompetencyStandard
        from app.models.brand import Brand
        from app.models.partner import CampusCompetencyStandard
        from sqlalchemy import func
        from sqlalchemy.orm import joinedload
        
        group_id = request.args.get('group_id', type=int)
        search = request.args.get('search', '')
        
        if not group_id:
            return jsonify({'error': 'group_id es requerido'}), 400
        
        group = CandidateGroup.query.get(group_id)
        if not group:
            return jsonify({'error': 'Grupo no encontrado'}), 404
        
        campus_id = group.campus_id
        
        # Obtener IDs de ECMs asignados al campus
        campus_ecm_relations = CampusCompetencyStandard.query.filter_by(campus_id=campus_id).all()
        ecm_ids = [rel.competency_standard_id for rel in campus_ecm_relations]
        
        if not ecm_ids:
            return jsonify({'ecms': [], 'total': 0, 'campus_id': campus_id})
        
        # Query ECMs con su marca
        query = CompetencyStandard.query.options(
            joinedload(CompetencyStandard.brand)
        ).filter(
            CompetencyStandard.id.in_(ecm_ids),
            CompetencyStandard.is_active == True
        )
        
        if search:
            search_term = f'%{search}%'
            query = query.filter(
                db.or_(
                    CompetencyStandard.code.ilike(search_term),
                    CompetencyStandard.name.ilike(search_term),
                    CompetencyStandard.sector.ilike(search_term)
                )
            )
        
        # Ordenar del más reciente al más viejo
        query = query.order_by(CompetencyStandard.created_at.desc())
        ecms = query.all()
        
        # Contar exámenes publicados por ECM
        published_counts = dict(
            db.session.query(
                Exam.competency_standard_id,
                func.count(Exam.id)
            ).filter(
                Exam.competency_standard_id.in_(ecm_ids),
                Exam.is_published == True
            ).group_by(Exam.competency_standard_id).all()
        )
        
        ecms_data = []
        for ecm in ecms:
            published_exam_count = published_counts.get(ecm.id, 0)
            ecms_data.append({
                'id': ecm.id,
                'code': ecm.code,
                'name': ecm.name,
                'description': ecm.description,
                'sector': ecm.sector,
                'level': ecm.level,
                'validity_years': ecm.validity_years,
                'certifying_body': ecm.certifying_body,
                'logo_url': ecm.logo_url,
                'brand_name': ecm.brand.name if ecm.brand else None,
                'brand_logo_url': ecm.brand.logo_url if ecm.brand else None,
                'published_exam_count': published_exam_count,
                'created_at': ecm.created_at.isoformat() if ecm.created_at else None,
            })
        
        return jsonify({
            'ecms': ecms_data,
            'total': len(ecms_data),
            'campus_id': campus_id
        })
        
    except Exception as e:
        import traceback
        print(f"[DEBUG] Error en get_available_ecms: {traceback.format_exc()}")
        return jsonify({'error': str(e)}), 500


@bp.route('/exams/available', methods=['GET'])
@jwt_required()
@coordinator_required
def get_available_exams():
    """Obtener exámenes disponibles para asignar a grupos
    
    Parámetros opcionales:
    - group_id: Si se proporciona, incluye is_assigned_to_group para indicar si ya está asignado
                Además, filtra exámenes por los ECM asignados al campus del grupo
    - campus_id: Si se proporciona (sin group_id), filtra exámenes por ECM del campus
    - filter_by_campus_ecm: Si es 'false', no filtra por ECM del campus (default: true)
    - ecm_id: Si se proporciona, filtra exámenes solo del ECM especificado
    """
    try:
        from app.models import Exam, GroupExam
        from app.models.study_content import StudyMaterial
        from app.models.competency_standard import CompetencyStandard
        from app.models.brand import Brand
        from app.models.partner import CampusCompetencyStandard
        from sqlalchemy import text
        from sqlalchemy.orm import joinedload
        
        search = request.args.get('search', '')
        page = request.args.get('page', 1, type=int)
        per_page = request.args.get('per_page', 20, type=int)
        group_id = request.args.get('group_id', type=int)
        campus_id = request.args.get('campus_id', type=int)
        ecm_id = request.args.get('ecm_id', type=int)
        filter_by_campus_ecm = request.args.get('filter_by_campus_ecm', 'true').lower() != 'false'
        
        # Si se proporciona group_id, obtener exámenes ya asignados a ese grupo
        assigned_exam_ids = set()
        campus_ecm_ids = None
        
        if group_id:
            # Obtener el grupo para saber su campus
            group = CandidateGroup.query.get(group_id)
            if group:
                campus_id = group.campus_id
                
            assigned_exams = GroupExam.query.filter_by(
                group_id=group_id,
                is_active=True
            ).all()
            assigned_exam_ids = {ge.exam_id for ge in assigned_exams}
        
        # Obtener ECMs del campus para filtrar exámenes
        if campus_id and filter_by_campus_ecm:
            campus_ecm_relations = CampusCompetencyStandard.query.filter_by(campus_id=campus_id).all()
            if campus_ecm_relations:
                campus_ecm_ids = {rel.competency_standard_id for rel in campus_ecm_relations}
        
        # Cargar la relación competency_standard y su brand para tener el código ECM y logo
        query = Exam.query.options(
            joinedload(Exam.competency_standard).joinedload(CompetencyStandard.brand)
        ).filter(Exam.is_published == True)
        
        # Filtrar por ECM específico si se proporciona
        if ecm_id:
            query = query.filter(Exam.competency_standard_id == ecm_id)
        # Filtrar por ECM del campus si corresponde
        elif campus_ecm_ids is not None:
            query = query.filter(Exam.competency_standard_id.in_(campus_ecm_ids))
        
        # Siempre hacer outerjoin con CompetencyStandard para búsqueda y ordenamiento
        query = query.outerjoin(CompetencyStandard, Exam.competency_standard_id == CompetencyStandard.id)
        
        if search:
            search_term = f'%{search}%'
            # Buscar también por código ECM
            query = query.filter(
                db.or_(
                    Exam.name.ilike(search_term),
                    Exam.standard.ilike(search_term),
                    Exam.description.ilike(search_term),
                    CompetencyStandard.code.ilike(search_term),
                    CompetencyStandard.name.ilike(search_term)
                )
            )
        
        # Ordenar por código ECM primero, luego por nombre de examen
        query = query.order_by(
            db.func.coalesce(CompetencyStandard.code, ''),
            Exam.name
        )
        pagination = query.paginate(page=page, per_page=per_page, max_per_page=1000, error_out=False)
        
        exams_data = []
        for exam in pagination.items:
            # Contar materiales de estudio asociados (legacy exam_id)
            materials_count = StudyMaterial.query.filter_by(exam_id=exam.id, is_published=True).count()
            
            # También contar materiales de la relación muchos a muchos
            try:
                result = db.session.execute(text('''
                    SELECT COUNT(DISTINCT sme.study_material_id) 
                    FROM study_material_exams sme
                    JOIN study_contents sc ON sme.study_material_id = sc.id
                    WHERE sme.exam_id = :exam_id AND sc.is_published = 1
                '''), {'exam_id': exam.id})
                linked_count = result.scalar() or 0
                # Tomar el máximo de ambos (evitar duplicados si están en ambos)
                materials_count = max(materials_count, linked_count)
            except Exception:
                pass
            
            # Obtener IDs de materiales vinculados al examen
            linked_material_ids = []
            try:
                lm_result = db.session.execute(text('''
                    SELECT DISTINCT sme.study_material_id 
                    FROM study_material_exams sme
                    JOIN study_contents sc ON sme.study_material_id = sc.id
                    WHERE sme.exam_id = :exam_id AND sc.is_published = 1
                '''), {'exam_id': exam.id})
                linked_material_ids = [row[0] for row in lm_result.fetchall()]
                
                # También incluir materiales con relación legacy
                legacy_ids = [m.id for m in StudyMaterial.query.filter_by(exam_id=exam.id, is_published=True).all()]
                linked_material_ids = list(set(linked_material_ids + legacy_ids))
            except Exception:
                pass
            
            # Contar preguntas por tipo (exam vs simulator)
            exam_questions = 0
            simulator_questions = 0
            try:
                questions_result = db.session.execute(text('''
                    SELECT q.type, COUNT(q.id) as cnt
                    FROM questions q
                    JOIN topics t ON q.topic_id = t.id
                    JOIN categories c ON t.category_id = c.id
                    WHERE c.exam_id = :exam_id
                    GROUP BY q.type
                '''), {'exam_id': exam.id})
                for row in questions_result.fetchall():
                    q_type = row[0] or 'exam'
                    count = row[1] or 0
                    if q_type == 'simulator':
                        simulator_questions = count
                    else:
                        exam_questions = count
            except Exception as e:
                print(f"[DEBUG] Error contando preguntas: {e}")
            
            # Contar ejercicios por tipo (exam vs simulator)
            exam_exercises = 0
            simulator_exercises = 0
            try:
                exercises_result = db.session.execute(text('''
                    SELECT e.type, COUNT(e.id) as cnt
                    FROM exercises e
                    JOIN topics t ON e.topic_id = t.id
                    JOIN categories c ON t.category_id = c.id
                    WHERE c.exam_id = :exam_id
                    GROUP BY e.type
                '''), {'exam_id': exam.id})
                for row in exercises_result.fetchall():
                    e_type = row[0] or 'exam'
                    count = row[1] or 0
                    if e_type == 'simulator':
                        simulator_exercises = count
                    else:
                        exam_exercises = count
            except Exception as e:
                print(f"[DEBUG] Error contando ejercicios: {e}")
            
            total_questions = exam_questions + simulator_questions
            total_exercises = exam_exercises + simulator_exercises
            
            # Obtener datos ECM si existe
            ecm_code = None
            ecm_name = None
            ecm_logo_url = None
            ecm_sector = None
            ecm_level = None
            ecm_brand_name = None
            ecm_brand_logo_url = None
            ecm_certifying_body = None
            ecm_validity_years = None
            cs = exam.competency_standard
            if cs:
                ecm_code = cs.code
                ecm_name = cs.name
                ecm_logo_url = cs.logo_url
                ecm_sector = cs.sector
                ecm_level = cs.level
                ecm_certifying_body = cs.certifying_body
                ecm_validity_years = cs.validity_years
                if cs.brand:
                    ecm_brand_name = cs.brand.name
                    ecm_brand_logo_url = cs.brand.logo_url
            
            exams_data.append({
                'id': exam.id,
                'name': exam.name,
                'version': exam.version,
                'standard': exam.standard,
                'ecm_code': ecm_code,
                'ecm_name': ecm_name,
                'ecm_logo_url': ecm_logo_url,
                'ecm_sector': ecm_sector,
                'ecm_level': ecm_level,
                'ecm_brand_name': ecm_brand_name,
                'ecm_brand_logo_url': ecm_brand_logo_url,
                'ecm_certifying_body': ecm_certifying_body,
                'ecm_validity_years': ecm_validity_years,
                'description': exam.description,
                'duration_minutes': exam.duration_minutes,
                'passing_score': exam.passing_score,
                'is_published': exam.is_published,
                'study_materials_count': materials_count,
                'total_questions': total_questions,
                'total_exercises': total_exercises,
                'exam_questions_count': exam_questions,
                'simulator_questions_count': simulator_questions,
                'exam_exercises_count': exam_exercises,
                'simulator_exercises_count': simulator_exercises,
                'is_assigned_to_group': exam.id in assigned_exam_ids,
                # Configuración de asignación por defecto del editor
                'default_max_attempts': exam.default_max_attempts if exam.default_max_attempts is not None else 2,
                'default_max_disconnections': exam.default_max_disconnections if exam.default_max_disconnections is not None else 3,
                'default_exam_content_type': exam.default_exam_content_type or 'mixed',
                'default_exam_questions_count': exam.default_exam_questions_count,
                'default_exam_exercises_count': exam.default_exam_exercises_count,
                'default_simulator_questions_count': exam.default_simulator_questions_count,
                'default_simulator_exercises_count': exam.default_simulator_exercises_count,
                # IDs de materiales vinculados para aceptación rápida
                'linked_material_ids': linked_material_ids,
            })
        
        response_data = {
            'exams': exams_data,
            'total': pagination.total,
            'pages': pagination.pages,
            'current_page': page
        }
        
        # Agregar info de filtrado por ECM si corresponde
        if campus_ecm_ids is not None:
            response_data['filtered_by_campus_ecm'] = True
            response_data['campus_ecm_count'] = len(campus_ecm_ids)
        else:
            response_data['filtered_by_campus_ecm'] = False
        
        return jsonify(response_data)
        
    except Exception as e:
        import traceback
        print(f"[DEBUG] Error en get_available_exams: {traceback.format_exc()}")
        return jsonify({'error': str(e)}), 500


@bp.route('/exams/<int:exam_id>/materials', methods=['GET'])
@jwt_required()
@coordinator_required
def get_exam_materials_for_assignment(exam_id):
    """Obtener materiales disponibles para asignar con un examen.
    Retorna primero los materiales ligados al examen (solo los publicados),
    luego los demás materiales publicados.
    """
    try:
        from app.models.study_content import StudyMaterial
        from app.models import Exam
        from sqlalchemy import text
        
        print(f"=== get_exam_materials_for_assignment: exam_id={exam_id} ===")
        
        exam = Exam.query.get_or_404(exam_id)
        print(f"Exam found: {exam.name}")
        
        # Obtener IDs de materiales vinculados al examen (desde study_material_exams)
        linked_material_ids = []
        try:
            result = db.session.execute(text('''
                SELECT study_material_id FROM study_material_exams 
                WHERE exam_id = :exam_id
            '''), {'exam_id': exam_id})
            linked_material_ids = [r[0] for r in result.fetchall()]
            print(f"Linked materials from study_material_exams: {linked_material_ids}")
        except Exception as e:
            print(f"Error getting linked materials from study_material_exams: {e}")
        
        # También verificar materiales con el campo legacy exam_id
        try:
            legacy_materials = StudyMaterial.query.filter_by(exam_id=exam_id).all()
            print(f"Legacy materials with exam_id={exam_id}: {[m.id for m in legacy_materials]}")
            for mat in legacy_materials:
                if mat.id not in linked_material_ids:
                    linked_material_ids.append(mat.id)
        except Exception as e:
            print(f"Error getting legacy materials: {e}")
        
        print(f"Total linked material IDs: {linked_material_ids}")
        
        materials_data = []
        linked_set = set(linked_material_ids)
        
        # Primero los materiales vinculados al examen (solo mostrar publicados)
        if linked_material_ids:
            linked_materials = StudyMaterial.query.filter(
                StudyMaterial.id.in_(linked_material_ids),
                StudyMaterial.is_published == True
            ).order_by(StudyMaterial.title).all()
            
            print(f"Linked published materials found: {len(linked_materials)}")
            
            for mat in linked_materials:
                # Calcular conteos
                sessions_count = mat.sessions.count() if mat.sessions else 0
                topics_count = sum(s.topics.count() for s in mat.sessions.all()) if mat.sessions else 0
                
                materials_data.append({
                    'id': mat.id,
                    'title': mat.title,
                    'description': mat.description,
                    'cover_image_url': mat.image_url,
                    'is_published': True,
                    'is_linked': True,  # Vinculado directamente al examen
                    'is_selected': True,  # Por defecto seleccionado si está vinculado
                    'sessions_count': sessions_count,
                    'topics_count': topics_count,
                })
        
        # Luego los materiales publicados que no están vinculados
        if linked_material_ids:
            other_published = StudyMaterial.query.filter(
                StudyMaterial.is_published == True,
                ~StudyMaterial.id.in_(linked_material_ids)
            ).order_by(StudyMaterial.title).all()
        else:
            # Si no hay materiales ligados, obtener todos los publicados
            other_published = StudyMaterial.query.filter(
                StudyMaterial.is_published == True
            ).order_by(StudyMaterial.title).all()
        
        print(f"Other published materials found: {len(other_published)}")
        
        for mat in other_published:
            # Calcular conteos
            sessions_count = mat.sessions.count() if mat.sessions else 0
            topics_count = sum(s.topics.count() for s in mat.sessions.all()) if mat.sessions else 0
            
            materials_data.append({
                'id': mat.id,
                'title': mat.title,
                'description': mat.description,
                'cover_image_url': mat.image_url,
                'is_published': True,
                'is_linked': False,  # No vinculado directamente
                'is_selected': False,  # Por defecto no seleccionado
                'sessions_count': sessions_count,
                'topics_count': topics_count,
            })
        
        print(f"Total materials to return: {len(materials_data)}")
        
        return jsonify({
            'exam_id': exam_id,
            'exam_name': exam.name,
            'materials': materials_data,
            'linked_count': len([m for m in materials_data if m['is_linked']]),
            'total_count': len(materials_data)
        })
        
    except Exception as e:
        import traceback
        print(f"Error in get_exam_materials_for_assignment: {e}")
        print(traceback.format_exc())
        return jsonify({'error': str(e)}), 500


# ============== MATERIALES PERSONALIZADOS POR GRUPO-EXAMEN ==============

@bp.route('/group-exams/<int:group_exam_id>/materials', methods=['GET'])
@jwt_required()
@coordinator_required
def get_group_exam_materials(group_exam_id):
    """Obtener materiales disponibles y seleccionados para un grupo-examen"""
    try:
        from app.models import GroupExam, GroupExamMaterial
        from app.models.study_content import StudyMaterial
        from sqlalchemy import text
        
        group_exam = GroupExam.query.get_or_404(group_exam_id)
        
        # Obtener materiales vinculados al examen (desde study_material_exams)
        linked_material_ids = []
        try:
            result = db.session.execute(text('''
                SELECT study_material_id FROM study_material_exams 
                WHERE exam_id = :exam_id
            '''), {'exam_id': group_exam.exam_id})
            linked_material_ids = [r[0] for r in result.fetchall()]
        except Exception:
            pass
        
        # Obtener materiales personalizados para este group_exam
        custom_materials = GroupExamMaterial.query.filter_by(group_exam_id=group_exam_id).all()
        custom_dict = {cm.study_material_id: cm.is_included for cm in custom_materials}
        
        # Obtener todos los materiales publicados para selección adicional
        all_published = StudyMaterial.query.filter_by(is_published=True).all()
        
        # Construir respuesta
        materials_data = []
        
        # Primero los materiales vinculados al examen (solo mostrar publicados en esta sección)
        for mat in StudyMaterial.query.filter(StudyMaterial.id.in_(linked_material_ids), StudyMaterial.is_published == True).all():
            # Por defecto incluido SOLO si está vinculado Y publicado
            is_included = custom_dict.get(mat.id, True)
            materials_data.append({
                'id': mat.id,
                'title': mat.title,
                'description': mat.description,
                'cover_image_url': mat.image_url,
                'is_published': mat.is_published,
                'is_linked': True,  # Vinculado directamente al examen
                'is_included': is_included,
            })
        
        # Luego los materiales publicados que no están vinculados
        linked_set = set(linked_material_ids)
        for mat in all_published:
            if mat.id not in linked_set:
                is_included = custom_dict.get(mat.id, False)  # Por defecto no incluido
                materials_data.append({
                    'id': mat.id,
                    'title': mat.title,
                    'description': mat.description,
                    'cover_image_url': mat.image_url,
                    'is_published': mat.is_published,
                    'is_linked': False,  # No vinculado directamente
                    'is_included': is_included,
                })
        
        return jsonify({
            'group_exam_id': group_exam_id,
            'exam_id': group_exam.exam_id,
            'exam_name': group_exam.exam.name if group_exam.exam else None,
            'materials': materials_data,
            'has_customizations': len(custom_materials) > 0
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@bp.route('/group-exams/<int:group_exam_id>/materials', methods=['PUT'])
@jwt_required()
@coordinator_required
def update_group_exam_materials(group_exam_id):
    """Actualizar materiales seleccionados para un grupo-examen"""
    try:
        from app.models import GroupExam, GroupExamMaterial
        
        group_exam = GroupExam.query.get_or_404(group_exam_id)
        data = request.get_json()
        
        # data.materials = [{id: number, is_included: boolean}, ...]
        materials_config = data.get('materials', [])
        
        # Eliminar configuraciones anteriores
        GroupExamMaterial.query.filter_by(group_exam_id=group_exam_id).delete()
        
        # Crear nuevas configuraciones
        for mat_config in materials_config:
            mat_id = mat_config.get('id')
            is_included = mat_config.get('is_included', False)
            
            if mat_id:
                gem = GroupExamMaterial(
                    group_exam_id=group_exam_id,
                    study_material_id=mat_id,
                    is_included=is_included
                )
                db.session.add(gem)
        
        db.session.commit()
        
        return jsonify({
            'message': 'Materiales actualizados exitosamente',
            'group_exam_id': group_exam_id
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500


@bp.route('/group-exams/<int:group_exam_id>/materials/reset', methods=['POST'])
@jwt_required()
@coordinator_required
def reset_group_exam_materials(group_exam_id):
    """Resetear materiales a los vinculados por defecto del examen"""
    try:
        from app.models import GroupExam, GroupExamMaterial
        
        group_exam = GroupExam.query.get_or_404(group_exam_id)
        
        # Eliminar todas las personalizaciones
        GroupExamMaterial.query.filter_by(group_exam_id=group_exam_id).delete()
        db.session.commit()
        
        return jsonify({
            'message': 'Materiales reseteados a valores por defecto',
            'group_exam_id': group_exam_id
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500


# ============== MOVER CANDIDATOS ENTRE GRUPOS ==============

@bp.route('/groups/<int:source_group_id>/members/move', methods=['POST'])
@jwt_required()
@coordinator_required
def move_members_to_group(source_group_id):
    """Mover candidatos de un grupo a otro - optimizado para miles de candidatos"""
    try:
        data = request.get_json()
        
        target_group_id = data.get('target_group_id')
        user_ids = data.get('user_ids', [])
        
        if not target_group_id:
            return jsonify({'error': 'El grupo destino es requerido'}), 400
        
        if not user_ids or not isinstance(user_ids, list):
            return jsonify({'error': 'Debe especificar al menos un candidato para mover'}), 400
        
        if source_group_id == target_group_id:
            return jsonify({'error': 'El grupo origen y destino no pueden ser el mismo'}), 400
        
        # Verificar grupos existen
        source_group = CandidateGroup.query.get_or_404(source_group_id)
        target_group = CandidateGroup.query.get_or_404(target_group_id)
        
        moved = []
        errors = []
        
        # Procesar en chunks
        CHUNK_SIZE = 500
        
        for i in range(0, len(user_ids), CHUNK_SIZE):
            chunk_ids = user_ids[i:i + CHUNK_SIZE]
            
            # Batch: obtener miembros activos del grupo origen
            source_members = GroupMember.query.filter(
                GroupMember.group_id == source_group_id,
                GroupMember.user_id.in_(chunk_ids),
                GroupMember.status == 'active'
            ).all()
            source_member_map = {m.user_id: m for m in source_members}
            
            # Batch: verificar existencia en grupo destino
            existing_targets = GroupMember.query.filter(
                GroupMember.group_id == target_group_id,
                GroupMember.user_id.in_(chunk_ids)
            ).all()
            existing_target_ids = {m.user_id for m in existing_targets}
            
            # Batch: obtener datos de usuarios para respuesta
            users = User.query.filter(User.id.in_(chunk_ids)).all()
            user_map = {u.id: u for u in users}
            
            for user_id in chunk_ids:
                user = user_map.get(user_id)
                
                if user_id not in source_member_map:
                    errors.append({
                        'user_id': user_id,
                        'name': user.name if user else 'Desconocido',
                        'error': 'No es miembro activo del grupo origen'
                    })
                    continue
                
                if user_id in existing_target_ids:
                    errors.append({
                        'user_id': user_id,
                        'name': user.name if user else 'Desconocido',
                        'error': 'Ya existe en el grupo destino'
                    })
                    continue
                
                # Mover: eliminar del origen y crear en destino
                db.session.delete(source_member_map[user_id])
                
                new_member = GroupMember(
                    group_id=target_group_id,
                    user_id=user_id,
                    status='active',
                    notes=f'Movido desde {source_group.name}'
                )
                db.session.add(new_member)
                
                moved.append({
                    'user_id': user_id,
                    'name': f"{user.name} {user.first_surname}" if user else 'Desconocido',
                    'email': user.email if user else ''
                })
            
            db.session.flush()
        
        db.session.commit()
        
        return jsonify({
            'message': f'{len(moved)} candidato(s) movido(s) exitosamente',
            'moved': moved,
            'errors': errors,
            'source_group': source_group.name,
            'target_group': target_group.name
        }), 200
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500


@bp.route('/groups/<int:group_id>/members/upload/preview', methods=['POST'])
@jwt_required()
@coordinator_required
def preview_group_members_upload(group_id):
    """Preview del archivo Excel antes de procesar la asignación"""
    from openpyxl import load_workbook
    import io

    MAX_ROWS = 50000
    MAX_FILE_MB = 20
    
    try:
        group = CandidateGroup.query.get_or_404(group_id)
        
        if 'file' not in request.files:
            return jsonify({'error': 'No se envió ningún archivo'}), 400
        
        file = request.files['file']
        if not file.filename.endswith(('.xlsx', '.xls')):
            return jsonify({'error': 'El archivo debe ser Excel (.xlsx o .xls)'}), 400

        # File size limit
        file.seek(0, 2)
        file_size = file.tell()
        file.seek(0)
        if file_size > MAX_FILE_MB * 1024 * 1024:
            return jsonify({'error': f'Archivo excede {MAX_FILE_MB}MB'}), 400
        
        # Cargar el archivo Excel
        wb = load_workbook(io.BytesIO(file.read()))
        ws = wb.active
        
        preview = []
        current_count = GroupMember.query.filter_by(group_id=group_id, status='active').count()
        
        # Recolectar todos los identificadores primero
        all_identifiers = []
        row_data = []
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if len(row_data) >= MAX_ROWS:
                return jsonify({'error': f'Límite excedido: máximo {MAX_ROWS:,} filas'}), 400
            identifier = row[0] if len(row) > 0 else None
            notes = row[1] if len(row) > 1 else None
            if not identifier:
                continue
            identifier = str(identifier).strip()
            all_identifiers.append(identifier)
            row_data.append((row_num, identifier, notes))
        
        # Resolver identificadores: email → CURP → username → nombre completo
        user_by_identifier, ambiguous = _resolve_identifiers_to_users(all_identifiers)
        
        # Batch: obtener membresías existentes para este grupo
        found_user_ids = list(set(u.id for u in user_by_identifier.values()))
        existing_member_ids = set()
        CHUNK_SIZE = 500
        if found_user_ids:
            for i in range(0, len(found_user_ids), CHUNK_SIZE):
                chunk = found_user_ids[i:i + CHUNK_SIZE]
                existing = GroupMember.query.filter(
                    GroupMember.group_id == group_id,
                    GroupMember.user_id.in_(chunk)
                ).all()
                existing_member_ids.update(m.user_id for m in existing)
        
        # Construir preview
        for row_num, identifier, notes in row_data:
            status = 'ready'
            user_info = None
            error_message = None
            ambiguous_matches = None
            
            # Verificar si es ambiguo
            if identifier in ambiguous:
                status = 'ambiguous'
                matches = ambiguous[identifier]
                ambiguous_matches = [{
                    'id': u.id,
                    'email': u.email,
                    'username': u.username,
                    'full_name': f"{u.name} {u.first_surname} {u.second_surname or ''}".strip(),
                    'curp': u.curp,
                } for u in matches[:10]]
                error_message = f'{len(matches)} candidatos coinciden con "{identifier}"'
            else:
                user = user_by_identifier.get(identifier)
                
                if not user:
                    status = 'not_found'
                    error_message = 'Candidato no encontrado. Verifique email, CURP, usuario o nombre completo'
                else:
                    if user.id in existing_member_ids:
                        status = 'already_member'
                        error_message = 'Ya es miembro del grupo'
                    
                    user_info = {
                        'id': user.id,
                        'email': user.email,
                        'username': user.username,
                        'name': user.name,
                        'first_surname': user.first_surname,
                        'second_surname': user.second_surname,
                        'full_name': f"{user.name} {user.first_surname} {user.second_surname or ''}".strip(),
                        'curp': user.curp,
                        'gender': user.gender,
                        'created_at': user.created_at.isoformat() if user.created_at else None,
                    }
            
            entry = {
                'row': row_num,
                'identifier': identifier,
                'notes': str(notes) if notes else None,
                'status': status,
                'error': error_message,
                'user': user_info
            }
            if ambiguous_matches:
                entry['ambiguous_matches'] = ambiguous_matches
            preview.append(entry)
        
        # Contar por status
        ready_count = len([p for p in preview if p['status'] == 'ready'])
        already_member_count = len([p for p in preview if p['status'] == 'already_member'])
        not_found_count = len([p for p in preview if p['status'] == 'not_found'])
        ambiguous_count = len([p for p in preview if p['status'] == 'ambiguous'])
        
        return jsonify({
            'group_name': group.name,
            'current_members': current_count,
            'preview': preview,
            'summary': {
                'total': len(preview),
                'ready': ready_count,
                'already_member': already_member_count,
                'not_found': not_found_count,
                'ambiguous': ambiguous_count,
            },
            'can_proceed': ready_count > 0
        }), 200
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@bp.route('/candidates/search/advanced', methods=['GET'])
@jwt_required()
@coordinator_required
def search_candidates_advanced():
    """Búsqueda avanzada de candidatos con filtros múltiples - optimizado para 100K+ registros"""
    try:
        # Parámetros de búsqueda básica
        search = request.args.get('search', '')
        search_field = request.args.get('search_field', '')
        page = request.args.get('page', 1, type=int)
        per_page = min(request.args.get('per_page', 25, type=int), 1000)  # Cap a 1000 max
        
        # Filtros avanzados
        has_group = request.args.get('has_group')
        group_id = request.args.get('group_id', type=int)
        exclude_group_id = request.args.get('exclude_group_id', type=int)
        partner_id = request.args.get('partner_id', type=int)
        campus_id = request.args.get('campus_id', type=int)
        state = request.args.get('state')
        gender = request.args.get('gender')
        
        # Modo liviano: solo devolver IDs+nombre para pageSizes muy grandes
        lightweight = per_page > 1000
        
        query = User.query.filter(
            User.role == 'candidato',
            User.is_active == True
        )
        
        # Multi-tenant: coordinadores solo ven candidatos que les pertenecen
        coord_id = _get_coordinator_filter(g.current_user)
        if coord_id:
            query = query.filter(User.coordinator_id == coord_id)
        
        # Filtro de búsqueda textual
        if search:
            search_term = f'%{search}%'
            if search_field and search_field in ['name', 'first_surname', 'second_surname', 'email', 'curp']:
                field_map = {
                    'name': User.name,
                    'first_surname': User.first_surname,
                    'second_surname': User.second_surname,
                    'email': User.email,
                    'curp': User.curp,
                }
                query = query.filter(field_map[search_field].ilike(search_term))
            else:
                query = query.filter(
                    db.or_(
                        User.name.ilike(search_term),
                        User.first_surname.ilike(search_term),
                        User.second_surname.ilike(search_term),
                        User.email.ilike(search_term),
                        User.curp.ilike(search_term)
                    )
                )
        
        # Filtro por género
        if gender and gender in ['M', 'F', 'O']:
            query = query.filter(User.gender == gender)
        
        # Filtro: tiene email / sin email
        has_email = request.args.get('has_email')
        if has_email == 'yes':
            query = query.filter(User.email.isnot(None), User.email != '')
        elif has_email == 'no':
            query = query.filter(db.or_(User.email.is_(None), User.email == ''))
        
        # Filtro: tiene CURP / sin CURP
        has_curp = request.args.get('has_curp')
        if has_curp == 'yes':
            query = query.filter(User.curp.isnot(None), User.curp != '')
        elif has_curp == 'no':
            query = query.filter(db.or_(User.curp.is_(None), User.curp == ''))
        
        # Filtro: tiene grupo / sin grupo
        if has_group == 'yes':
            # Candidatos que tienen al menos un grupo activo
            candidates_with_group = db.session.query(GroupMember.user_id).filter(
                GroupMember.status == 'active'
            ).subquery()
            query = query.filter(User.id.in_(candidates_with_group))
        elif has_group == 'no':
            # Candidatos sin grupo activo
            candidates_with_group = db.session.query(GroupMember.user_id).filter(
                GroupMember.status == 'active'
            ).subquery()
            query = query.filter(~User.id.in_(candidates_with_group))
        
        # Filtro: candidatos de un grupo específico
        if group_id:
            members_of_group = db.session.query(GroupMember.user_id).filter(
                GroupMember.group_id == group_id,
                GroupMember.status == 'active'
            ).subquery()
            query = query.filter(User.id.in_(members_of_group))
        
        # Filtro: excluir candidatos de un grupo
        if exclude_group_id:
            members_of_group = db.session.query(GroupMember.user_id).filter(
                GroupMember.group_id == exclude_group_id
            ).subquery()
            query = query.filter(~User.id.in_(members_of_group))
        
        # Filtro por partner
        if partner_id:
            # Candidatos que están en grupos de campuses del partner
            groups_of_partner = db.session.query(CandidateGroup.id).join(Campus).filter(
                Campus.partner_id == partner_id
            ).subquery()
            members_of_partner = db.session.query(GroupMember.user_id).filter(
                GroupMember.group_id.in_(groups_of_partner),
                GroupMember.status == 'active'
            ).subquery()
            query = query.filter(User.id.in_(members_of_partner))
        
        # Filtro por campus
        if campus_id:
            groups_of_campus = db.session.query(CandidateGroup.id).filter(
                CandidateGroup.campus_id == campus_id
            ).subquery()
            members_of_campus = db.session.query(GroupMember.user_id).filter(
                GroupMember.group_id.in_(groups_of_campus),
                GroupMember.status == 'active'
            ).subquery()
            query = query.filter(User.id.in_(members_of_campus))
        
        # Filtro por estado
        if state:
            groups_of_state = db.session.query(CandidateGroup.id).join(Campus).filter(
                Campus.state_name == state
            ).subquery()
            members_of_state = db.session.query(GroupMember.user_id).filter(
                GroupMember.group_id.in_(groups_of_state),
                GroupMember.status == 'active'
            ).subquery()
            query = query.filter(User.id.in_(members_of_state))
        
        # Ordenamiento: 'recent' ordena por fecha de creación descendente
        sort_by = request.args.get('sort_by', 'name')
        if sort_by == 'recent':
            query = query.order_by(User.created_at.desc())
        else:
            query = query.order_by(User.first_surname, User.name)
        pagination = query.paginate(page=page, per_page=per_page, max_per_page=1000, error_out=False)
        
        # Batch: obtener membresías activas para todos los usuarios de la página
        # Para pages grandes (>100 items), skip membresía para acelerar respuesta
        user_ids_in_page = [user.id for user in pagination.items]
        
        membership_map = {}
        if user_ids_in_page and not lightweight:
            from sqlalchemy import text
            # Procesar en chunks de 500 para evitar limits de IN clause
            for i in range(0, len(user_ids_in_page), 500):
                chunk_ids = user_ids_in_page[i:i + 500]
                # Obtener el grupo más reciente por joined_at para cada usuario
                membership_rows = db.session.execute(text("""
                    SELECT t.user_id, t.group_id, t.group_name,
                           t.campus_id, t.campus_name, t.state_name, t.city,
                           t.cycle_id, t.cycle_name,
                           t.partner_id, t.partner_name,
                           t.joined_at
                    FROM (
                        SELECT gm.user_id, g.id as group_id, g.name as group_name,
                               c.id as campus_id, c.name as campus_name, c.state_name, c.city,
                               sc.id as cycle_id, sc.name as cycle_name,
                               p.id as partner_id, p.name as partner_name,
                               gm.joined_at,
                               ROW_NUMBER() OVER (PARTITION BY gm.user_id ORDER BY gm.joined_at DESC) as rn
                        FROM group_members gm
                        JOIN candidate_groups g ON gm.group_id = g.id
                        LEFT JOIN campuses c ON g.campus_id = c.id
                        LEFT JOIN school_cycles sc ON g.school_cycle_id = sc.id
                        LEFT JOIN partners p ON c.partner_id = p.id
                        WHERE gm.user_id IN :user_ids AND gm.status = 'active'
                    ) t
                    WHERE t.rn = 1
                """), {'user_ids': tuple(chunk_ids)})
                
                for row in membership_rows:
                    membership_map[row[0]] = {
                        'group_id': row[1],
                        'group_name': row[2],
                        'campus_id': row[3],
                        'campus_name': row[4],
                        'state_name': row[5],
                        'city': row[6],
                        'school_cycle_id': row[7],
                        'school_cycle_name': row[8],
                        'partner_id': row[9],
                        'partner_name': row[10],
                        'joined_at': row[11].isoformat() if row[11] else None,
                    }
        
        candidates = []
        for user in pagination.items:
            group_info = membership_map.get(user.id)
            
            candidate = {
                'id': user.id,
                'email': user.email,
                'username': user.username,
                'name': user.name,
                'first_surname': user.first_surname,
                'second_surname': user.second_surname,
                'full_name': f"{user.name} {user.first_surname} {user.second_surname or ''}".strip(),
                'current_group': group_info,
            }
            
            # Campos extra solo en modo normal (<=100 per_page)
            if not lightweight:
                candidate['curp'] = user.curp
                candidate['gender'] = user.gender
                candidate['created_at'] = user.created_at.isoformat() if user.created_at else None
            
            candidates.append(candidate)
        
        return jsonify({
            'candidates': candidates,
            'total': pagination.total,
            'pages': pagination.pages,
            'current_page': page,
            'filters_applied': {
                'search': search,
                'search_field': search_field,
                'has_group': has_group,
                'group_id': group_id,
                'exclude_group_id': exclude_group_id,
                'partner_id': partner_id,
                'campus_id': campus_id,
                'state': state,
                'gender': gender
            }
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@bp.route('/groups/<int:group_id>/export-members', methods=['GET'])
@jwt_required()
@coordinator_required
def export_group_members(group_id):
    """
    Exportar miembros del grupo a Excel
    Incluye: Grupo, Usuario, Contraseña, Nombre Completo, Email, CURP, Estado, Estatus Certificación
    """
    try:
        from sqlalchemy import text
        from app.models.result import Result
        
        group = CandidateGroup.query.get_or_404(group_id)
        
        # Obtener miembros con sus usuarios
        members = GroupMember.query.filter_by(group_id=group_id).all()
        
        # Obtener exámenes del grupo para buscar resultados
        group_exams = GroupExam.query.filter_by(group_id=group_id, is_active=True).all()
        real_exam_ids = [ge.exam_id for ge in group_exams]
        
        # Obtener resultados de certificación para todos los usuarios
        user_ids = [m.user_id for m in members if m.user]
        certification_status_map = {}
        
        if user_ids and real_exam_ids:
            results = Result.query.filter(
                Result.user_id.in_(user_ids),
                Result.exam_id.in_(real_exam_ids)
            ).all()
            
            for r in results:
                if r.user_id not in certification_status_map:
                    certification_status_map[r.user_id] = {'status': -1, 'result': -1}
                
                current = certification_status_map[r.user_id]
                # Priorizar: aprobado > en proceso > reprobado
                if r.result == 1 and r.status == 1:
                    current['status'] = 1
                    current['result'] = 1
                elif r.status == 0 and current['result'] != 1:
                    current['status'] = 0
                    current['result'] = 0
                elif r.status == 1 and r.result == 0 and current['status'] != 0 and current['result'] != 1:
                    current['status'] = 1
                    current['result'] = 0
        
        # Crear workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Miembros del Grupo"
        
        # Estilos
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Headers
        headers = ["Grupo", "Usuario", "Contraseña", "Nombre Completo", "Email", "CURP", "Correo del Responsable", "Estatus Certificación"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Obtener correo del responsable desde el plantel
        campus = Campus.query.get(group.campus_id) if group.campus_id else None
        responsable_email = (campus.responsable.email if campus and campus.responsable and campus.responsable.email else 'ND')
        
        # Datos
        row_num = 2
        for member in members:
            user = member.user
            if user:
                # Desencriptar contraseña
                password = "No disponible"
                if user.encrypted_password:
                    try:
                        decrypted = decrypt_password(user.encrypted_password)
                        if decrypted:
                            password = decrypted
                    except Exception:
                        password = "Error al desencriptar"
                
                # Determinar estatus de certificación
                cert_info = certification_status_map.get(user.id, None)
                if cert_info:
                    if cert_info['result'] == 1 and cert_info['status'] == 1:
                        cert_status_text = "Certificado"
                    elif cert_info['status'] == 0:
                        cert_status_text = "En proceso"
                    else:
                        cert_status_text = "No aprobado"
                else:
                    cert_status_text = "Pendiente"
                
                ws.cell(row=row_num, column=1, value=group.name or 'ND').border = thin_border
                ws.cell(row=row_num, column=2, value=user.username or 'ND').border = thin_border
                ws.cell(row=row_num, column=3, value=password).border = thin_border
                ws.cell(row=row_num, column=4, value=user.full_name or f"{user.name or ''} {user.last_name or ''}".strip() or 'ND').border = thin_border
                ws.cell(row=row_num, column=5, value=user.email or 'ND').border = thin_border
                ws.cell(row=row_num, column=6, value=user.curp or 'ND').border = thin_border
                ws.cell(row=row_num, column=7, value=responsable_email).border = thin_border
                ws.cell(row=row_num, column=8, value=cert_status_text).border = thin_border
                row_num += 1
        
        # Ajustar anchos de columna
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 35
        ws.column_dimensions['E'].width = 30
        ws.column_dimensions['F'].width = 22
        ws.column_dimensions['G'].width = 30
        ws.column_dimensions['H'].width = 22
        
        # Guardar a BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Nombre del archivo
        safe_name = group.name.replace(' ', '_').replace('/', '-')[:30]
        filename = f"Miembros_{safe_name}_{datetime.now().strftime('%Y%m%d')}.xlsx"
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@bp.route('/groups/<int:group_id>/export-certifications', methods=['GET'])
@jwt_required()
@coordinator_required
def export_group_certifications(group_id):
    """
    Exportar reporte de certificaciones del grupo a Excel.
    Cada fila = un resultado de examen por candidato (puede haber varias filas por persona).
    Columnas: Grupo, Usuario, Nombre Completo, Email, CURP, Tipo, Correo del Responsable,
              Examen, Puntaje, Resultado, Estatus, Fecha
    """
    try:
        from app.models.result import Result
        from app.models.exam import Exam

        group = CandidateGroup.query.get_or_404(group_id)
        members = GroupMember.query.filter_by(group_id=group_id).all()
        group_exams = GroupExam.query.filter_by(group_id=group_id, is_active=True).all()
        real_exam_ids = [ge.exam_id for ge in group_exams]

        # Mapa de exámenes para obtener nombre
        exam_map = {}
        if real_exam_ids:
            exams = Exam.query.filter(Exam.id.in_(real_exam_ids)).all()
            exam_map = {e.id: e for e in exams}

        user_ids = [m.user_id for m in members if m.user]

        # Obtener todos los resultados de estos usuarios en estos exámenes
        results = []
        if user_ids and real_exam_ids:
            results = Result.query.filter(
                Result.user_id.in_(user_ids),
                Result.exam_id.in_(real_exam_ids)
            ).order_by(Result.user_id, Result.exam_id, Result.created_at).all()

        # Mapa de usuarios
        user_map = {m.user_id: m.user for m in members if m.user}
        member_map = {m.user_id: m for m in members}

        # Obtener correo del responsable desde el plantel
        campus = Campus.query.get(group.campus_id) if group.campus_id else None
        responsable_email = (campus.responsable.email if campus and campus.responsable and campus.responsable.email else 'ND')

        # Crear workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Certificaciones"

        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        headers = [
            "Grupo", "Usuario", "Nombre Completo", "Email", "CURP", "Tipo",
            "Correo del Responsable", "Examen", "Puntaje", "Resultado", "Fecha"
        ]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        row_num = 2
        for r in results:
            user = user_map.get(r.user_id)
            if not user:
                continue

            exam = exam_map.get(r.exam_id)
            exam_name = exam.name if exam else f"Examen {r.exam_id}"

            if r.result == 1:
                resultado_text = "Aprobado"
            else:
                resultado_text = "No aprobado"

            role = (user.role if user.role else 'candidato')
            fecha = r.end_date or r.start_date
            fecha_str = fecha.strftime('%Y-%m-%d %H:%M') if fecha else 'ND'

            ws.cell(row=row_num, column=1, value=group.name or 'ND').border = thin_border
            ws.cell(row=row_num, column=2, value=user.username or 'ND').border = thin_border
            ws.cell(row=row_num, column=3, value=user.full_name or f"{user.name or ''} {user.last_name or ''}".strip() or 'ND').border = thin_border
            ws.cell(row=row_num, column=4, value=user.email or 'ND').border = thin_border
            ws.cell(row=row_num, column=5, value=user.curp or 'ND').border = thin_border
            ws.cell(row=row_num, column=6, value=role).border = thin_border
            ws.cell(row=row_num, column=7, value=responsable_email).border = thin_border
            ws.cell(row=row_num, column=8, value=exam_name).border = thin_border
            ws.cell(row=row_num, column=9, value=(r.score * 10) if r.score is not None else 'ND').border = thin_border
            ws.cell(row=row_num, column=10, value=resultado_text).border = thin_border
            ws.cell(row=row_num, column=11, value=fecha_str).border = thin_border
            row_num += 1

        # Ajustar anchos
        widths = [25, 20, 35, 30, 22, 15, 30, 30, 12, 16, 20]
        from openpyxl.utils import get_column_letter
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        safe_name = group.name.replace(' ', '_').replace('/', '-')[:30]
        filename = f"Certificaciones_{safe_name}_{datetime.now().strftime('%Y%m%d')}.xlsx"

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ---------- Helpers para generar filas de reporte ----------
def _get_responsable_email_map(group_ids):
    """Retorna dict: group_id -> email del responsable del plantel del grupo."""
    resp_map = {}
    for gid in set(group_ids):
        group = CandidateGroup.query.get(gid)
        if group and group.campus_id:
            campus = Campus.query.get(group.campus_id)
            if campus and campus.responsable and campus.responsable.email:
                resp_map[gid] = campus.responsable.email
            else:
                resp_map[gid] = 'ND'
        else:
            resp_map[gid] = 'ND'
    return resp_map


def _build_member_row(user, group, campus, partner, cert_status_map, responsable_email='',
                      include_partner=False, include_campus=False, cycle_name=None,
                      include_password=True, include_cert_status=True):
    """Genera una lista de valores para una fila del reporte Excel."""
    password = "No disponible"
    if include_password and user.encrypted_password:
        try:
            decrypted = decrypt_password(user.encrypted_password)
            if decrypted:
                password = decrypted
        except Exception:
            password = "Error al desencriptar"

    cert_info = cert_status_map.get(user.id)
    if cert_info:
        if cert_info['result'] == 1 and cert_info['status'] == 1:
            cert_text = "Certificado"
        elif cert_info['status'] == 0:
            cert_text = "En proceso"
        else:
            cert_text = "No aprobado"
    else:
        cert_text = "Pendiente"

    row = []
    if include_partner:
        row.append(partner.name if partner else 'ND')
        row.append(partner.country if partner else 'ND')
        row.append(cycle_name or 'Sin ciclo')
    if include_campus:
        row.append(campus.name if campus else 'ND')
        row.append((campus.state_name or campus.country) if campus else 'ND')
    row.append(group.name or 'ND')
    row.append(user.username or 'ND')
    if include_password:
        row.append(password)
    row.append(user.full_name or f"{user.name or ''} {user.last_name or ''}".strip() or 'ND')
    row.append(user.email or 'ND')
    row.append(user.curp or 'ND')
    row.append(user.role or 'candidato')
    row.append(responsable_email or 'ND')
    if include_cert_status:
        row.append(cert_text)
    return row


def _build_report_headers(include_partner=False, include_campus=False,
                          include_password=True, include_cert_status=True):
    """Headers del reporte según el nivel."""
    headers = []
    if include_partner:
        headers += ['Partner', 'País Partner', 'Ciclo Escolar']
    if include_campus:
        headers += ['Plantel', 'Estado / País']
    headers.append('Grupo')
    headers.append('Usuario')
    if include_password:
        headers.append('Contraseña')
    headers += ['Nombre Completo', 'Email', 'CURP', 'Tipo', 'Correo del Responsable']
    if include_cert_status:
        headers.append('Estatus Certificación')
    return headers


def _style_report_ws(ws, headers, header_color="4472C4"):
    """Aplica estilos al worksheet."""
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    return thin_border


def _get_cert_status_map(user_ids, exam_ids):
    """Calcula mapa de certificación: user_id -> {status, result}"""
    from app.models.result import Result
    cert_map = {}
    if user_ids and exam_ids:
        results = Result.query.filter(Result.user_id.in_(user_ids), Result.exam_id.in_(exam_ids)).all()
        for r in results:
            if r.user_id not in cert_map:
                cert_map[r.user_id] = {'status': -1, 'result': -1}
            c = cert_map[r.user_id]
            if r.result == 1 and r.status == 1:
                c['status'] = 1; c['result'] = 1
            elif r.status == 0 and c['result'] != 1:
                c['status'] = 0; c['result'] = 0
            elif r.status == 1 and r.result == 0 and c['status'] != 0 and c['result'] != 1:
                c['status'] = 1; c['result'] = 0
    return cert_map


@bp.route('/campuses/<int:campus_id>/export-report', methods=['GET'])
@jwt_required()
@coordinator_required
def export_campus_report(campus_id):
    """Exportar reporte del plantel a Excel — todos los grupos y miembros"""
    try:
        campus = Campus.query.get_or_404(campus_id)
        partner = Partner.query.get(campus.partner_id)
        groups = CandidateGroup.query.filter_by(campus_id=campus_id).all()

        # Recolectar todos los miembros y exam_ids
        all_members = []  # [(member, group)]
        all_exam_ids = set()
        for grp in groups:
            members = GroupMember.query.filter_by(group_id=grp.id).all()
            for m in members:
                all_members.append((m, grp))
            exams = GroupExam.query.filter_by(group_id=grp.id, is_active=True).all()
            all_exam_ids.update(ge.exam_id for ge in exams)

        user_ids = list({m.user_id for m, _ in all_members if m.user})
        cert_map = _get_cert_status_map(user_ids, list(all_exam_ids))
        group_ids = list({grp.id for _, grp in all_members})
        resp_map = _get_responsable_email_map(group_ids)

        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte del Plantel"
        headers = _build_report_headers(include_campus=True, include_password=False, include_cert_status=False)
        thin_border = _style_report_ws(ws, headers)

        row_num = 2
        for member, grp in all_members:
            user = member.user
            if not user:
                continue
            vals = _build_member_row(user, grp, campus, partner, cert_map,
                                     responsable_email=resp_map.get(grp.id, ''),
                                     include_campus=True, include_password=False, include_cert_status=False)
            for col, val in enumerate(vals, 1):
                ws.cell(row=row_num, column=col, value=val).border = thin_border
            row_num += 1

        # Ajustar anchos
        for i, _ in enumerate(headers, 1):
            ws.column_dimensions[chr(64 + i) if i <= 26 else 'A'].width = 22
        from openpyxl.utils import get_column_letter
        for i in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(i)].width = 22

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        safe_name = campus.name.replace(' ', '_').replace('/', '-')[:30]
        filename = f"Reporte_Plantel_{safe_name}_{datetime.now().strftime('%Y%m%d')}.xlsx"
        return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name=filename)
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@bp.route('/partners/<int:partner_id>/export-report', methods=['GET'])
@jwt_required()
@coordinator_required
def export_partner_report(partner_id):
    """Exportar reporte del partner a Excel — todos los planteles, todos los grupos, todos los miembros"""
    try:
        partner = Partner.query.get_or_404(partner_id)
        campuses = Campus.query.filter_by(partner_id=partner_id).all()

        all_members = []  # [(member, group, campus, cycle_name)]
        all_exam_ids = set()
        for campus in campuses:
            groups = CandidateGroup.query.filter_by(campus_id=campus.id).all()
            for grp in groups:
                cycle_name = grp.school_cycle.name if grp.school_cycle else 'Sin ciclo'
                members = GroupMember.query.filter_by(group_id=grp.id).all()
                for m in members:
                    all_members.append((m, grp, campus, cycle_name))
                exams = GroupExam.query.filter_by(group_id=grp.id, is_active=True).all()
                all_exam_ids.update(ge.exam_id for ge in exams)

        user_ids = list({m.user_id for m, _, _, _ in all_members if m.user})
        cert_map = _get_cert_status_map(user_ids, list(all_exam_ids))
        group_ids = list({grp.id for _, grp, _, _ in all_members})
        resp_map = _get_responsable_email_map(group_ids)

        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte del Partner"
        headers = _build_report_headers(include_partner=True, include_campus=True, include_password=False, include_cert_status=False)
        thin_border = _style_report_ws(ws, headers)

        row_num = 2
        for member, grp, campus, cycle_name in all_members:
            user = member.user
            if not user:
                continue
            vals = _build_member_row(user, grp, campus, partner, cert_map,
                                     responsable_email=resp_map.get(grp.id, ''),
                                     include_partner=True, include_campus=True, cycle_name=cycle_name,
                                     include_password=False, include_cert_status=False)
            for col, val in enumerate(vals, 1):
                ws.cell(row=row_num, column=col, value=val).border = thin_border
            row_num += 1

        from openpyxl.utils import get_column_letter
        for i in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(i)].width = 22

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        safe_name = partner.name.replace(' ', '_').replace('/', '-')[:30]
        filename = f"Reporte_Partner_{safe_name}_{datetime.now().strftime('%Y%m%d')}.xlsx"
        return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name=filename)
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ============== ENDPOINTS PARA RESPONSABLE DE PLANTEL ==============

def responsable_required(f):
    """Decorador que requiere rol de responsable"""
    @wraps(f)
    def decorated(*args, **kwargs):
        user_id = get_jwt_identity()
        user = User.query.get(user_id)
        if not user:
            return jsonify({'error': 'No autorizado'}), 401
        if user.role != 'responsable':
            return jsonify({'error': 'Acceso denegado. Se requiere rol de responsable'}), 403
        if not user.campus_id:
            return jsonify({'error': 'No tienes un plantel asignado'}), 403
        g.current_user = user
        return f(*args, **kwargs)
    return decorated


@bp.route('/mi-plantel', methods=['GET'])
@jwt_required()
@responsable_required
def get_mi_plantel():
    """Obtener información del plantel del responsable"""
    try:
        user = g.current_user
        campus = Campus.query.get(user.campus_id)
        
        if not campus:
            return jsonify({'error': 'No se encontró el plantel asignado'}), 404
        
        # Verificar que el usuario es el responsable asignado
        if campus.responsable_id != user.id:
            return jsonify({'error': 'No eres el responsable de este plantel'}), 403
        
        return jsonify({
            'campus': campus.to_dict(include_partner=True, include_responsable=True, include_config=True)
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@bp.route('/mi-plantel/stats', methods=['GET'])
@jwt_required()
@responsable_required
def get_mi_plantel_stats():
    """Obtener estadísticas del plantel del responsable"""
    from app.models import Result, Exam, StudyMaterial
    from app.models.student_progress import StudentContentProgress
    from sqlalchemy import func
    
    try:
        user = g.current_user
        campus = Campus.query.get(user.campus_id)
        
        if not campus:
            return jsonify({'error': 'No se encontró el plantel asignado'}), 404
        
        # Obtener todos los grupos del plantel
        groups = CandidateGroup.query.filter_by(campus_id=campus.id, is_active=True).all()
        group_ids = [g.id for g in groups]
        
        # Obtener todos los candidatos del plantel (miembros de grupos)
        candidate_ids = db.session.query(GroupMember.user_id).filter(
            GroupMember.group_id.in_(group_ids),
            GroupMember.status == 'active'
        ).distinct().all()
        candidate_ids = [c[0] for c in candidate_ids]
        
        total_candidates = len(candidate_ids)
        total_groups = len(groups)
        
        # Estadísticas de exámenes
        if candidate_ids:
            # Total de evaluaciones completadas
            total_evaluations = Result.query.filter(
                Result.user_id.in_(candidate_ids),
                Result.status == 1  # Completado
            ).count()
            
            # Evaluaciones aprobadas
            passed_evaluations = Result.query.filter(
                Result.user_id.in_(candidate_ids),
                Result.status == 1,
                Result.result == 1  # Aprobado
            ).count()
            
            # Evaluaciones reprobadas
            failed_evaluations = Result.query.filter(
                Result.user_id.in_(candidate_ids),
                Result.status == 1,
                Result.result == 0  # Reprobado
            ).count()
            
            # Promedio de calificaciones
            avg_score = db.session.query(func.avg(Result.score)).filter(
                Result.user_id.in_(candidate_ids),
                Result.status == 1
            ).scalar() or 0
            
            # Progreso en materiales de estudio
            total_progress_records = StudentContentProgress.query.filter(
                StudentContentProgress.user_id.in_(candidate_ids)
            ).count()
            
            completed_progress = StudentContentProgress.query.filter(
                StudentContentProgress.user_id.in_(candidate_ids),
                StudentContentProgress.is_completed == True
            ).count()
            
            # Candidatos con al menos una evaluación
            candidates_with_eval = db.session.query(Result.user_id).filter(
                Result.user_id.in_(candidate_ids),
                Result.status == 1
            ).distinct().count()
            
            # Candidatos con certificado (aprobado)
            candidates_certified = db.session.query(Result.user_id).filter(
                Result.user_id.in_(candidate_ids),
                Result.status == 1,
                Result.result == 1
            ).distinct().count()
        else:
            total_evaluations = 0
            passed_evaluations = 0
            failed_evaluations = 0
            avg_score = 0
            total_progress_records = 0
            completed_progress = 0
            candidates_with_eval = 0
            candidates_certified = 0
        
        # Calcular tasa de aprobación
        approval_rate = (passed_evaluations / total_evaluations * 100) if total_evaluations > 0 else 0
        
        # Calcular tasa de progreso en materiales
        material_completion_rate = (completed_progress / total_progress_records * 100) if total_progress_records > 0 else 0
        
        return jsonify({
            'campus': {
                'id': campus.id,
                'name': campus.name,
                'code': campus.code
            },
            'stats': {
                'total_groups': total_groups,
                'total_candidates': total_candidates,
                'candidates_with_evaluations': candidates_with_eval,
                'candidates_certified': candidates_certified,
                'total_evaluations': total_evaluations,
                'passed_evaluations': passed_evaluations,
                'failed_evaluations': failed_evaluations,
                'approval_rate': round(approval_rate, 1),
                'average_score': round(float(avg_score), 1),
                'material_completion_rate': round(material_completion_rate, 1),
                'total_material_progress': total_progress_records,
                'completed_material_progress': completed_progress
            }
        })
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@bp.route('/mi-plantel/evaluations', methods=['GET'])
@jwt_required()
@responsable_required
def get_mi_plantel_evaluations():
    """Obtener lista de evaluaciones del plantel del responsable"""
    from app.models import Result, Exam
    
    try:
        user = g.current_user
        campus = Campus.query.get(user.campus_id)
        
        if not campus:
            return jsonify({'error': 'No se encontró el plantel asignado'}), 404
        
        # Parámetros de paginación y filtrado
        page = request.args.get('page', 1, type=int)
        per_page = request.args.get('per_page', 20, type=int)
        exam_id = request.args.get('exam_id', type=int)
        result_status = request.args.get('result', type=int)  # 0=reprobado, 1=aprobado
        search = request.args.get('search', '')
        
        # Obtener todos los grupos del plantel
        groups = CandidateGroup.query.filter_by(campus_id=campus.id).all()
        group_ids = [g.id for g in groups]
        
        # Obtener todos los candidatos del plantel
        candidate_ids = db.session.query(GroupMember.user_id).filter(
            GroupMember.group_id.in_(group_ids)
        ).distinct().all()
        candidate_ids = [c[0] for c in candidate_ids]
        
        if not candidate_ids:
            return jsonify({
                'evaluations': [],
                'total': 0,
                'page': page,
                'per_page': per_page,
                'pages': 0
            })
        
        # Query base de resultados
        query = Result.query.filter(
            Result.user_id.in_(candidate_ids),
            Result.status == 1  # Solo completados
        )
        
        # Filtros
        if exam_id:
            query = query.filter(Result.exam_id == exam_id)
        
        if result_status is not None:
            query = query.filter(Result.result == result_status)
        
        # Ordenar por fecha más reciente
        query = query.order_by(Result.end_date.desc())
        
        # Paginación
        pagination = query.paginate(page=page, per_page=per_page, max_per_page=1000, error_out=False)
        
        evaluations = []
        for result in pagination.items:
            # Obtener datos del usuario
            candidate = User.query.get(result.user_id)
            
            # Obtener datos del examen
            exam = Exam.query.get(result.exam_id)
            
            # Obtener grupo del candidato
            member = GroupMember.query.filter(
                GroupMember.user_id == result.user_id,
                GroupMember.group_id.in_(group_ids)
            ).first()
            group = CandidateGroup.query.get(member.group_id) if member else None
            
            evaluations.append({
                'id': result.id,
                'candidate': {
                    'id': candidate.id,
                    'full_name': candidate.full_name,
                    'username': candidate.username,
                    'email': candidate.email,
                    'curp': candidate.curp
                } if candidate else None,
                'exam': {
                    'id': exam.id,
                    'name': exam.name,
                    'version': exam.version
                } if exam else None,
                'group': {
                    'id': group.id,
                    'name': group.name
                } if group else None,
                'score': result.score,
                'result': result.result,
                'result_text': 'Aprobado' if result.result == 1 else 'Reprobado',
                'start_date': result.start_date.isoformat() if result.start_date else None,
                'end_date': result.end_date.isoformat() if result.end_date else None,
                'duration_seconds': result.duration_seconds,
                'certificate_url': result.certificate_url,
                'report_url': result.report_url
            })
        
        return jsonify({
            'evaluations': evaluations,
            'total': pagination.total,
            'page': page,
            'per_page': per_page,
            'pages': pagination.pages
        })
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@bp.route('/mi-plantel/evaluations/export', methods=['GET'])
@jwt_required()
@responsable_required
def export_mi_plantel_evaluations():
    """Exportar evaluaciones del plantel a Excel"""
    from app.models import Result, Exam
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    try:
        user = g.current_user
        campus = Campus.query.get(user.campus_id)
        
        if not campus:
            return jsonify({'error': 'No se encontró el plantel asignado'}), 404
        
        # Filtros opcionales
        exam_id = request.args.get('exam_id', type=int)
        result_status = request.args.get('result', type=int)
        
        # Obtener todos los grupos del plantel
        groups = CandidateGroup.query.filter_by(campus_id=campus.id).all()
        group_ids = [g.id for g in groups]
        
        # Obtener todos los candidatos del plantel
        candidate_ids = db.session.query(GroupMember.user_id).filter(
            GroupMember.group_id.in_(group_ids)
        ).distinct().all()
        candidate_ids = [c[0] for c in candidate_ids]
        
        if not candidate_ids:
            return jsonify({'error': 'No hay candidatos en el plantel'}), 404
        
        # Query de resultados
        query = Result.query.filter(
            Result.user_id.in_(candidate_ids),
            Result.status == 1
        )
        
        if exam_id:
            query = query.filter(Result.exam_id == exam_id)
        
        if result_status is not None:
            query = query.filter(Result.result == result_status)
        
        results = query.order_by(Result.end_date.desc()).all()
        
        # Crear workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Evaluaciones"
        
        # Estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_align = Alignment(horizontal='center')
        
        # Título
        ws.merge_cells('A1:I1')
        ws['A1'] = f"Reporte de Evaluaciones - {campus.name}"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Fecha de generación
        ws.merge_cells('A2:I2')
        ws['A2'] = f"Generado el: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        ws['A2'].alignment = Alignment(horizontal='center')
        
        # Encabezados
        headers = ['Grupo', 'Candidato', 'CURP', 'Email', 'Examen', 'Calificación', 'Resultado', 'Fecha', 'Duración']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = center_align
        
        # Datos
        row_num = 5
        for result in results:
            candidate = User.query.get(result.user_id)
            exam = Exam.query.get(result.exam_id)
            
            member = GroupMember.query.filter(
                GroupMember.user_id == result.user_id,
                GroupMember.group_id.in_(group_ids)
            ).first()
            group = CandidateGroup.query.get(member.group_id) if member else None
            
            # Calcular duración en formato legible
            duration_str = ""
            if result.duration_seconds:
                mins = result.duration_seconds // 60
                secs = result.duration_seconds % 60
                duration_str = f"{mins}m {secs}s"
            
            ws.cell(row=row_num, column=1, value=group.name if group else "Sin grupo").border = thin_border
            ws.cell(row=row_num, column=2, value=candidate.full_name if candidate else "").border = thin_border
            ws.cell(row=row_num, column=3, value=candidate.curp if candidate else "").border = thin_border
            ws.cell(row=row_num, column=4, value=candidate.email if candidate else "").border = thin_border
            ws.cell(row=row_num, column=5, value=exam.name if exam else "").border = thin_border
            ws.cell(row=row_num, column=6, value=result.score).border = thin_border
            ws.cell(row=row_num, column=7, value="Aprobado" if result.result == 1 else "Reprobado").border = thin_border
            ws.cell(row=row_num, column=8, value=result.end_date.strftime('%d/%m/%Y %H:%M') if result.end_date else "").border = thin_border
            ws.cell(row=row_num, column=9, value=duration_str).border = thin_border
            
            row_num += 1
        
        # Ajustar anchos
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 22
        ws.column_dimensions['D'].width = 30
        ws.column_dimensions['E'].width = 25
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 12
        ws.column_dimensions['H'].width = 18
        ws.column_dimensions['I'].width = 12
        
        # Guardar a BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"Evaluaciones_{campus.code}_{datetime.now().strftime('%Y%m%d')}.xlsx"
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@bp.route('/mi-plantel/groups', methods=['GET'])
@jwt_required()
@responsable_required
def get_mi_plantel_groups():
    """Obtener grupos del plantel del responsable"""
    try:
        user = g.current_user
        campus = Campus.query.get(user.campus_id)
        
        if not campus:
            return jsonify({'error': 'No se encontró el plantel asignado'}), 404
        
        groups = CandidateGroup.query.filter_by(campus_id=campus.id, is_active=True).all()
        
        return jsonify({
            'groups': [g.to_dict(include_members=False) for g in groups]
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@bp.route('/mi-plantel/exams', methods=['GET'])
@jwt_required()
@responsable_required
def get_mi_plantel_exams():
    """Obtener exámenes asignados al plantel"""
    from app.models import Exam
    from app.models.partner import GroupExam
    
    try:
        user = g.current_user
        campus = Campus.query.get(user.campus_id)
        
        if not campus:
            return jsonify({'error': 'No se encontró el plantel asignado'}), 404
        
        # Obtener grupos del plantel
        groups = CandidateGroup.query.filter_by(campus_id=campus.id, is_active=True).all()
        group_ids = [g.id for g in groups]
        
        # Obtener exámenes asignados a los grupos
        exam_ids = db.session.query(GroupExam.exam_id).filter(
            GroupExam.group_id.in_(group_ids),
            GroupExam.is_active == True
        ).distinct().all()
        exam_ids = [e[0] for e in exam_ids]
        
        exams = Exam.query.filter(Exam.id.in_(exam_ids)).all() if exam_ids else []
        
        return jsonify({
            'exams': [{
                'id': exam.id,
                'name': exam.name,
                'version': exam.version,
                'description': exam.description
            } for exam in exams]
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# =====================================================
# RESPONSABLE - GESTIÓN AVANZADA DEL PLANTEL
# =====================================================

def _validate_responsable_campus(user):
    """Helper: valida que el responsable tenga campus y retorna (campus, error_response)"""
    campus = Campus.query.get(user.campus_id)
    if not campus:
        return None, (jsonify({'error': 'No se encontró el plantel asignado'}), 404)
    if campus.responsable_id != user.id:
        return None, (jsonify({'error': 'No eres el responsable de este plantel'}), 403)
    return campus, None


def _validate_responsable_group(user, group_id):
    """Helper: valida que el grupo pertenece al campus del responsable"""
    campus, err = _validate_responsable_campus(user)
    if err:
        return None, None, err
    group = CandidateGroup.query.get(group_id)
    if not group or group.campus_id != campus.id:
        return None, None, (jsonify({'error': 'Grupo no encontrado en tu plantel'}), 404)
    return campus, group, None


@bp.route('/mi-plantel/dashboard-advanced', methods=['GET'])
@jwt_required()
@responsable_required
def get_mi_plantel_dashboard_advanced():
    """Dashboard avanzado con datos para gráficas"""
    from app.models import Result, Exam
    from app.models.student_progress import StudentContentProgress
    from sqlalchemy import func, extract, case
    
    try:
        user = g.current_user
        campus, err = _validate_responsable_campus(user)
        if err:
            return err
        
        groups = CandidateGroup.query.filter_by(campus_id=campus.id, is_active=True).all()
        group_ids = [gr.id for gr in groups]
        group_map = {gr.id: gr.name for gr in groups}
        
        candidate_ids = []
        if group_ids:
            candidate_ids = [c[0] for c in db.session.query(GroupMember.user_id).filter(
                GroupMember.group_id.in_(group_ids), GroupMember.status == 'active'
            ).distinct().all()]
        
        # ---- Approval by group ----
        approval_by_group = []
        for gr in groups:
            member_ids = [m.user_id for m in GroupMember.query.filter_by(group_id=gr.id, status='active').all()]
            if not member_ids:
                approval_by_group.append({'group_name': gr.name, 'group_id': gr.id, 'approved': 0, 'failed': 0, 'pending': len(member_ids), 'rate': 0})
                continue
            approved = Result.query.filter(Result.user_id.in_(member_ids), Result.status == 1, Result.result == 1).count()
            failed = Result.query.filter(Result.user_id.in_(member_ids), Result.status == 1, Result.result == 0).count()
            total = approved + failed
            approval_by_group.append({
                'group_name': gr.name, 'group_id': gr.id,
                'approved': approved, 'failed': failed,
                'total_members': len(member_ids),
                'rate': round(approved / total * 100, 1) if total > 0 else 0
            })
        
        # ---- Scores by group ----
        scores_by_group = []
        for gr in groups:
            member_ids = [m.user_id for m in GroupMember.query.filter_by(group_id=gr.id, status='active').all()]
            if not member_ids:
                scores_by_group.append({'group_name': gr.name, 'average': 0, 'min': 0, 'max': 0})
                continue
            stats = db.session.query(
                func.avg(Result.score), func.min(Result.score), func.max(Result.score)
            ).filter(Result.user_id.in_(member_ids), Result.status == 1).first()
            scores_by_group.append({
                'group_name': gr.name,
                'average': round(float(stats[0] or 0), 1),
                'min': round(float(stats[1] or 0), 1),
                'max': round(float(stats[2] or 0), 1)
            })
        
        # ---- Score distribution ----
        score_distribution = []
        if candidate_ids:
            for low in range(0, 100, 10):
                high = low + 10
                label = f"{low}-{high}"
                if low == 90:
                    count = Result.query.filter(
                        Result.user_id.in_(candidate_ids), Result.status == 1,
                        Result.score >= low, Result.score <= 100
                    ).count()
                    label = "90-100"
                else:
                    count = Result.query.filter(
                        Result.user_id.in_(candidate_ids), Result.status == 1,
                        Result.score >= low, Result.score < high
                    ).count()
                score_distribution.append({'range': label, 'count': count})
        
        # ---- Evaluations over time (last 6 months) ----
        evaluations_over_time = []
        if candidate_ids:
            from datetime import datetime as dt
            now = dt.utcnow()
            for i in range(5, -1, -1):
                month = now.month - i
                year = now.year
                if month <= 0:
                    month += 12
                    year -= 1
                month_start = dt(year, month, 1)
                if month == 12:
                    month_end = dt(year + 1, 1, 1)
                else:
                    month_end = dt(year, month + 1, 1)
                
                approved = Result.query.filter(
                    Result.user_id.in_(candidate_ids), Result.status == 1, Result.result == 1,
                    Result.end_date >= month_start, Result.end_date < month_end
                ).count()
                failed = Result.query.filter(
                    Result.user_id.in_(candidate_ids), Result.status == 1, Result.result == 0,
                    Result.end_date >= month_start, Result.end_date < month_end
                ).count()
                evaluations_over_time.append({
                    'month': f"{year}-{month:02d}",
                    'approved': approved, 'failed': failed
                })
        
        # ---- Material progress by group ----
        material_progress_by_group = []
        for gr in groups:
            member_ids = [m.user_id for m in GroupMember.query.filter_by(group_id=gr.id, status='active').all()]
            if not member_ids:
                material_progress_by_group.append({'group_name': gr.name, 'completed': 0, 'in_progress': 0, 'not_started': len(member_ids)})
                continue
            total_p = StudentContentProgress.query.filter(StudentContentProgress.user_id.in_(member_ids)).count()
            completed_p = StudentContentProgress.query.filter(
                StudentContentProgress.user_id.in_(member_ids), StudentContentProgress.is_completed == True
            ).count()
            in_progress_p = total_p - completed_p
            material_progress_by_group.append({
                'group_name': gr.name,
                'completed': completed_p, 'in_progress': in_progress_p,
                'total_members': len(member_ids)
            })
        
        # ---- Certification by type ----
        certification_by_type = {
            'constancia_eduit': 0, 'certificado_eduit': 0,
            'certificado_conocer': 0, 'insignia_digital': 0
        }
        if candidate_ids:
            certified_users = db.session.query(Result.user_id).filter(
                Result.user_id.in_(candidate_ids), Result.status == 1, Result.result == 1
            ).distinct().all()
            certified_user_ids = [u[0] for u in certified_users]
            for uid in certified_user_ids:
                u = User.query.get(uid)
                if u:
                    if getattr(u, 'enable_evaluation_report', False):
                        certification_by_type['constancia_eduit'] += 1
                    if getattr(u, 'enable_certificate', False):
                        certification_by_type['certificado_eduit'] += 1
                    if getattr(u, 'enable_conocer_certificate', False):
                        certification_by_type['certificado_conocer'] += 1
                    if getattr(u, 'enable_digital_badge', False):
                        certification_by_type['insignia_digital'] += 1
        
        # ---- Summary stats ----
        total_candidates = len(candidate_ids)
        total_evals = Result.query.filter(Result.user_id.in_(candidate_ids), Result.status == 1).count() if candidate_ids else 0
        passed = Result.query.filter(Result.user_id.in_(candidate_ids), Result.status == 1, Result.result == 1).count() if candidate_ids else 0
        failed = total_evals - passed
        avg_score = 0
        if candidate_ids:
            avg = db.session.query(func.avg(Result.score)).filter(Result.user_id.in_(candidate_ids), Result.status == 1).scalar()
            avg_score = round(float(avg or 0), 1)
        
        return jsonify({
            'campus': {'id': campus.id, 'name': campus.name, 'code': campus.code},
            'stats': {
                'total_candidates': total_candidates,
                'total_groups': len(groups),
                'total_evaluations': total_evals,
                'passed_evaluations': passed,
                'failed_evaluations': failed,
                'approval_rate': round(passed / total_evals * 100, 1) if total_evals > 0 else 0,
                'average_score': avg_score,
                'certification_rate': round(len([u for u in (db.session.query(Result.user_id).filter(Result.user_id.in_(candidate_ids), Result.status == 1, Result.result == 1).distinct().all() if candidate_ids else [])]) / total_candidates * 100, 1) if total_candidates > 0 else 0
            },
            'charts': {
                'approval_by_group': approval_by_group,
                'scores_by_group': scores_by_group,
                'score_distribution': score_distribution,
                'evaluations_over_time': evaluations_over_time,
                'material_progress_by_group': material_progress_by_group,
                'certification_by_type': certification_by_type
            }
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@bp.route('/mi-plantel/certificates-by-group', methods=['GET'])
@jwt_required()
@responsable_required
def get_mi_plantel_certificates_by_group():
    """Certificados emitidos por grupo con tipos habilitados"""
    from app.models import Result, Exam
    
    try:
        user = g.current_user
        campus, err = _validate_responsable_campus(user)
        if err:
            return err
        
        groups = CandidateGroup.query.filter_by(campus_id=campus.id, is_active=True).order_by(CandidateGroup.name).all()
        
        # Tipos habilitados en el campus
        campus_tiers = {
            'constancia_eduit': campus.enable_tier_basic or False,
            'certificado_eduit': campus.enable_tier_standard or False,
            'certificado_conocer': campus.enable_tier_advanced or False,
            'insignia_digital': campus.enable_digital_badge or False
        }
        
        groups_data = []
        for gr in groups:
            members = GroupMember.query.filter_by(group_id=gr.id, status='active').all()
            member_ids = [m.user_id for m in members]
            
            # Obtener tiers del grupo (override o herencia del campus)
            group_tiers = {
                'constancia_eduit': gr.enable_tier_basic_override if gr.enable_tier_basic_override is not None else campus_tiers['constancia_eduit'],
                'certificado_eduit': gr.enable_tier_standard_override if gr.enable_tier_standard_override is not None else campus_tiers['certificado_eduit'],
                'certificado_conocer': gr.enable_tier_advanced_override if gr.enable_tier_advanced_override is not None else campus_tiers['certificado_conocer'],
                'insignia_digital': gr.enable_digital_badge_override if gr.enable_digital_badge_override is not None else campus_tiers['insignia_digital']
            }
            
            certificates = []
            if member_ids:
                results = Result.query.filter(
                    Result.user_id.in_(member_ids), Result.status == 1, Result.result == 1
                ).all()
                
                for r in results:
                    u = User.query.get(r.user_id)
                    exam = Exam.query.get(r.exam_id) if r.exam_id else None
                    if u:
                        certificates.append({
                            'user_id': r.user_id,
                            'user_name': u.full_name,
                            'user_curp': getattr(u, 'curp', ''),
                            'exam_name': exam.name if exam else 'N/A',
                            'score': float(r.score) if r.score else 0,
                            'date': r.end_date.isoformat() if r.end_date else None,
                            'certificate_code': r.certificate_code,
                            'certificate_url': r.certificate_url,
                            'report_url': getattr(r, 'report_url', None),
                            'document_options': {
                                'evaluation_report': getattr(u, 'enable_evaluation_report', False),
                                'certificate': getattr(u, 'enable_certificate', False),
                                'conocer_certificate': getattr(u, 'enable_conocer_certificate', False),
                                'digital_badge': getattr(u, 'enable_digital_badge', False)
                            }
                        })
            
            groups_data.append({
                'group_id': gr.id,
                'group_name': gr.name,
                'group_code': gr.code,
                'total_members': len(member_ids),
                'total_certificates': len(certificates),
                'enabled_tiers': group_tiers,
                'certificates': certificates
            })
        
        return jsonify({
            'campus': {'id': campus.id, 'name': campus.name, 'code': campus.code},
            'campus_tiers': campus_tiers,
            'groups': groups_data
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@bp.route('/mi-plantel/campus', methods=['PUT'])
@jwt_required()
@responsable_required
def update_mi_plantel_campus():
    """Editar datos de contacto y dirección del campus (no configuración de tiers/licencias)"""
    try:
        user = g.current_user
        campus, err = _validate_responsable_campus(user)
        if err:
            return err
        
        data = request.get_json()
        editable_fields = [
            'address', 'city', 'state_name', 'postal_code', 'country',
            'website',
            'director_name', 'director_first_surname', 'director_second_surname',
            'director_email', 'director_phone', 'director_curp', 'director_gender',
            'director_date_of_birth'
        ]
        
        for field in editable_fields:
            if field in data:
                setattr(campus, field, data[field])
        
        # Auto-sync email/phone from director fields
        if campus.director_email:
            campus.email = campus.director_email
        if campus.director_phone:
            campus.phone = campus.director_phone
        
        campus.updated_at = datetime.utcnow()
        db.session.commit()
        
        return jsonify({
            'message': 'Datos del plantel actualizados',
            'campus': campus.to_dict(include_config=True)
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500


@bp.route('/mi-plantel/groups', methods=['POST'])
@jwt_required()
@responsable_required
def create_mi_plantel_group():
    """Crear grupo en el plantel del responsable (requiere can_manage_groups)"""
    try:
        user = g.current_user
        if not user.can_manage_groups:
            return jsonify({'error': 'No tienes permisos para gestionar grupos'}), 403
        
        campus, err = _validate_responsable_campus(user)
        if err:
            return err
        
        data = request.get_json()
        if not data.get('name'):
            return jsonify({'error': 'El nombre es requerido'}), 400
        
        school_cycle_id = data.get('school_cycle_id')
        if school_cycle_id:
            cycle = SchoolCycle.query.get(school_cycle_id)
            if not cycle or cycle.campus_id != campus.id:
                return jsonify({'error': 'Ciclo escolar no válido'}), 400
        
        group = CandidateGroup(
            campus_id=campus.id,
            school_cycle_id=school_cycle_id,
            name=data['name'],
            code=data.get('code'),
            description=data.get('description'),
            start_date=data.get('start_date'),
            end_date=data.get('end_date'),
            is_active=data.get('is_active', True)
        )
        db.session.add(group)
        db.session.commit()
        
        return jsonify({
            'message': 'Grupo creado exitosamente',
            'group': group.to_dict(include_cycle=True)
        }), 201
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500


@bp.route('/mi-plantel/groups/<int:group_id>', methods=['GET'])
@jwt_required()
@responsable_required
def get_mi_plantel_group_detail(group_id):
    """Detalle de un grupo del plantel"""
    try:
        user = g.current_user
        campus, group, err = _validate_responsable_group(user, group_id)
        if err:
            return err
        return jsonify({
            'group': group.to_dict(include_members=True, include_campus=True, include_cycle=True, include_config=True)
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@bp.route('/mi-plantel/groups/<int:group_id>', methods=['PUT'])
@jwt_required()
@responsable_required
def update_mi_plantel_group(group_id):
    """Actualizar grupo del plantel (requiere can_manage_groups)"""
    try:
        user = g.current_user
        if not user.can_manage_groups:
            return jsonify({'error': 'No tienes permisos para gestionar grupos'}), 403
        
        campus, group, err = _validate_responsable_group(user, group_id)
        if err:
            return err
        
        data = request.get_json()
        for field in ['name', 'code', 'description', 'start_date', 'end_date', 'is_active']:
            if field in data:
                setattr(group, field, data[field])
        
        if 'school_cycle_id' in data:
            school_cycle_id = data['school_cycle_id']
            if school_cycle_id:
                cycle = SchoolCycle.query.get(school_cycle_id)
                if not cycle or cycle.campus_id != campus.id:
                    return jsonify({'error': 'Ciclo escolar no válido'}), 400
            group.school_cycle_id = school_cycle_id
        
        db.session.commit()
        return jsonify({
            'message': 'Grupo actualizado',
            'group': group.to_dict(include_cycle=True)
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500


@bp.route('/mi-plantel/groups/<int:group_id>', methods=['DELETE'])
@jwt_required()
@responsable_required
def delete_mi_plantel_group(group_id):
    """Desactivar grupo (requiere can_manage_groups)"""
    try:
        user = g.current_user
        if not user.can_manage_groups:
            return jsonify({'error': 'No tienes permisos para gestionar grupos'}), 403
        
        campus, group, err = _validate_responsable_group(user, group_id)
        if err:
            return err
        
        group.is_active = False
        db.session.commit()
        return jsonify({'message': 'Grupo desactivado exitosamente'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500


@bp.route('/mi-plantel/groups/<int:group_id>/members', methods=['GET'])
@jwt_required()
@responsable_required
def get_mi_plantel_group_members(group_id):
    """Listar miembros de un grupo del plantel"""
    from sqlalchemy import text
    try:
        user = g.current_user
        campus, group, err = _validate_responsable_group(user, group_id)
        if err:
            return err
        
        members = GroupMember.query.filter_by(group_id=group_id).all()
        members_data = []
        for m in members:
            u = User.query.get(m.user_id)
            if u:
                members_data.append({
                    'id': m.id,
                    'user_id': m.user_id,
                    'user_name': u.full_name,
                    'user_email': u.email,
                    'user_curp': getattr(u, 'curp', ''),
                    'status': m.status,
                    'notes': m.notes,
                    'joined_at': m.joined_at.isoformat() if m.joined_at else None
                })
        
        return jsonify({
            'group_id': group_id,
            'group_name': group.name,
            'members': members_data,
            'total': len(members_data)
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@bp.route('/mi-plantel/groups/<int:group_id>/members', methods=['POST'])
@jwt_required()
@responsable_required
def add_mi_plantel_group_member(group_id):
    """Agregar candidato a grupo (requiere can_manage_groups)"""
    try:
        user = g.current_user
        if not user.can_manage_groups:
            return jsonify({'error': 'No tienes permisos para gestionar grupos'}), 403
        
        campus, group, err = _validate_responsable_group(user, group_id)
        if err:
            return err
        
        data = request.get_json()
        user_id = data.get('user_id')
        if not user_id:
            return jsonify({'error': 'El ID del usuario es requerido'}), 400
        
        target_user = User.query.get(user_id)
        if not target_user:
            return jsonify({'error': 'Usuario no encontrado'}), 404
        if target_user.role != 'candidato':
            return jsonify({'error': 'Solo se pueden agregar candidatos'}), 400
        
        existing = GroupMember.query.filter_by(group_id=group_id, user_id=user_id).first()
        if existing:
            return jsonify({'error': 'El usuario ya es miembro de este grupo'}), 400
        
        member = GroupMember(
            group_id=group_id, user_id=user_id,
            status=data.get('status', 'active'), notes=data.get('notes')
        )
        db.session.add(member)
        db.session.commit()
        
        return jsonify({
            'message': 'Miembro agregado exitosamente',
            'member': member.to_dict(include_user=True)
        }), 201
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500


@bp.route('/mi-plantel/groups/<int:group_id>/members/bulk', methods=['POST'])
@jwt_required()
@responsable_required
def add_mi_plantel_group_members_bulk(group_id):
    """Agregar múltiples candidatos (requiere can_bulk_create_candidates)"""
    try:
        user = g.current_user
        if not user.can_bulk_create_candidates:
            return jsonify({'error': 'No tienes permisos para altas masivas'}), 403
        
        campus, group, err = _validate_responsable_group(user, group_id)
        if err:
            return err
        
        data = request.get_json()
        user_ids = data.get('user_ids', [])
        if not user_ids:
            return jsonify({'error': 'Se requiere al menos un ID'}), 400
        
        added = []
        errors = []
        for uid in user_ids:
            target = User.query.get(uid)
            if not target:
                errors.append({'user_id': uid, 'error': 'No encontrado'})
                continue
            if target.role != 'candidato':
                errors.append({'user_id': uid, 'error': 'No es candidato'})
                continue
            existing = GroupMember.query.filter_by(group_id=group_id, user_id=uid).first()
            if existing:
                errors.append({'user_id': uid, 'error': 'Ya es miembro'})
                continue
            m = GroupMember(group_id=group_id, user_id=uid, status='active')
            db.session.add(m)
            added.append(uid)
        
        db.session.commit()
        return jsonify({
            'message': f'{len(added)} miembros agregados',
            'added': len(added), 'errors': errors
        }), 201
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500


@bp.route('/mi-plantel/groups/<int:group_id>/members/<int:member_id>', methods=['DELETE'])
@jwt_required()
@responsable_required
def remove_mi_plantel_group_member(group_id, member_id):
    """Eliminar miembro de grupo (requiere can_manage_groups)"""
    try:
        user = g.current_user
        if not user.can_manage_groups:
            return jsonify({'error': 'No tienes permisos'}), 403
        
        campus, group, err = _validate_responsable_group(user, group_id)
        if err:
            return err
        
        member = GroupMember.query.filter_by(id=member_id, group_id=group_id).first_or_404()
        db.session.delete(member)
        db.session.commit()
        return jsonify({'message': 'Miembro eliminado del grupo'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500


@bp.route('/mi-plantel/groups/<int:group_id>/exams', methods=['GET'])
@jwt_required()
@responsable_required
def get_mi_plantel_group_exams(group_id):
    """Exámenes asignados a un grupo del plantel"""
    try:
        user = g.current_user
        campus, group, err = _validate_responsable_group(user, group_id)
        if err:
            return err
        
        group_exams = GroupExam.query.filter_by(group_id=group_id, is_active=True).all()
        return jsonify({
            'group_id': group_id, 'group_name': group.name,
            'assigned_exams': [ge.to_dict(include_exam=True, include_materials=True) for ge in group_exams],
            'total': len(group_exams)
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@bp.route('/mi-plantel/candidates/search', methods=['GET'])
@jwt_required()
@responsable_required
def search_mi_plantel_candidates():
    """Buscar candidatos para agregar a grupos del plantel"""
    try:
        user = g.current_user
        campus, err = _validate_responsable_campus(user)
        if err:
            return err
        
        search = request.args.get('search', '')
        exclude_group_id = request.args.get('exclude_group_id', type=int)
        page = request.args.get('page', 1, type=int)
        per_page = request.args.get('per_page', 20, type=int)
        
        query = User.query.filter(User.role == 'candidato', User.is_active == True)
        
        # Filtrar solo candidatos que pertenecen al campus del responsable
        campus_group_ids = [g.id for g in CandidateGroup.query.filter_by(campus_id=campus.id).all()]
        if campus_group_ids:
            campus_candidate_ids = [m.user_id for m in GroupMember.query.filter(GroupMember.group_id.in_(campus_group_ids)).all()]
            if campus_candidate_ids:
                query = query.filter(User.id.in_(campus_candidate_ids))
            else:
                query = query.filter(User.id == None)
        else:
            query = query.filter(User.id == None)
        
        if exclude_group_id:
            existing_ids = [m.user_id for m in GroupMember.query.filter_by(group_id=exclude_group_id).all()]
            if existing_ids:
                query = query.filter(~User.id.in_(existing_ids))
        
        if search:
            search_filter = f"%{search}%"
            query = query.filter(
                db.or_(
                    User.name.ilike(search_filter),
                    User.first_surname.ilike(search_filter),
                    User.email.ilike(search_filter),
                    User.curp.ilike(search_filter)
                )
            )
        
        paginated = query.order_by(User.name).paginate(page=page, per_page=per_page, max_per_page=1000, error_out=False)
        
        return jsonify({
            'candidates': [{
                'id': u.id, 'name': u.name, 'first_surname': u.first_surname,
                'second_surname': u.second_surname, 'full_name': u.full_name,
                'email': u.email, 'curp': getattr(u, 'curp', '')
            } for u in paginated.items],
            'total': paginated.total,
            'page': page,
            'pages': paginated.pages
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# =====================================================
# ENDPOINTS PARA CANDIDATOS - Exámenes y Materiales Asignados
# =====================================================

@bp.route('/mis-examenes', methods=['GET'])
@jwt_required()
def get_mis_examenes():
    """Obtener exámenes asignados al candidato basándose en sus grupos"""
    from app.models import Exam
    from app.models.partner import GroupExam, GroupMember, GroupExamMember
    
    try:
        user_id = get_jwt_identity()
        user = User.query.get(user_id)
        
        if not user:
            return jsonify({'error': 'Usuario no encontrado'}), 404
        
        # Solo para candidatos y responsables
        if user.role not in ['candidato', 'responsable']:
            return jsonify({'error': 'Acceso no autorizado'}), 403
        
        # Si es responsable, obtener exámenes del plantel
        if user.role == 'responsable':
            if not user.campus_id:
                return jsonify({'exams': [], 'total': 0, 'pages': 1, 'current_page': 1})
            
            campus = Campus.query.get(user.campus_id)
            if not campus:
                return jsonify({'exams': [], 'total': 0, 'pages': 1, 'current_page': 1})
            
            # Obtener grupos del plantel
            groups = CandidateGroup.query.filter_by(campus_id=campus.id, is_active=True).all()
            group_ids = [g.id for g in groups]
            
            # Obtener exámenes asignados a los grupos
            exam_ids = db.session.query(GroupExam.exam_id).filter(
                GroupExam.group_id.in_(group_ids),
                GroupExam.is_active == True
            ).distinct().all()
            exam_ids = [e[0] for e in exam_ids]
        else:
            # Para candidatos: obtener grupos donde es miembro activo
            memberships = GroupMember.query.filter_by(
                user_id=user_id,
                status='active'
            ).all()
            
            if not memberships:
                return jsonify({'exams': [], 'total': 0, 'pages': 1, 'current_page': 1})
            
            # Filtrar solo grupos que existen y están activos
            group_ids = []
            for m in memberships:
                group = CandidateGroup.query.filter_by(id=m.group_id, is_active=True).first()
                if group:
                    group_ids.append(m.group_id)
            
            if not group_ids:
                return jsonify({'exams': [], 'total': 0, 'pages': 1, 'current_page': 1})
            
            # Obtener exámenes asignados a esos grupos
            group_exams = GroupExam.query.filter(
                GroupExam.group_id.in_(group_ids),
                GroupExam.is_active == True
            ).all()
            
            # Filtrar por asignación específica si es individual
            exam_ids = set()
            exam_group_context = {}  # exam_id -> {group_id, group_exam_id}
            for ge in group_exams:
                if ge.assignment_type == 'all' or ge.assignment_type is None:
                    # Asignado a todos los miembros del grupo
                    exam_ids.add(ge.exam_id)
                    if ge.exam_id not in exam_group_context:
                        exam_group_context[ge.exam_id] = {'group_id': ge.group_id, 'group_exam_id': ge.id}
                elif ge.assignment_type == 'selected':
                    # Verificar si el candidato está asignado específicamente
                    member_assignment = GroupExamMember.query.filter_by(
                        group_exam_id=ge.id,
                        user_id=user_id
                    ).first()
                    if member_assignment:
                        exam_ids.add(ge.exam_id)
                        if ge.exam_id not in exam_group_context:
                            exam_group_context[ge.exam_id] = {'group_id': ge.group_id, 'group_exam_id': ge.id}
            
            exam_ids = list(exam_ids)
        
        if not exam_ids:
            return jsonify({'exams': [], 'total': 0, 'pages': 1, 'current_page': 1})
        
        # Obtener los exámenes publicados
        exams = Exam.query.filter(
            Exam.id.in_(exam_ids),
            Exam.is_published == True
        ).order_by(Exam.updated_at.desc()).all()
        
        # Para candidatos: obtener mejor resultado aprobado por exam_id
        approved_map = {}  # exam_id -> best_score
        if user.role == 'candidato':
            from app.models.result import Result
            from sqlalchemy import func
            best_results = db.session.query(
                Result.exam_id,
                func.max(Result.score).label('best_score'),
                func.max(Result.result).label('best_result')
            ).filter(
                Result.user_id == str(user_id),
                Result.exam_id.in_(exam_ids),
                Result.status == 1  # solo completados
            ).group_by(Result.exam_id).all()
            for r in best_results:
                approved_map[r.exam_id] = {
                    'best_score': r.best_score or 0,
                    'is_approved': (r.best_result or 0) == 1
                }
        
        # Enriquecer con contexto de grupo y estado de aprobación
        exams_data = []
        for exam in exams:
            d = exam.to_dict()
            ctx = exam_group_context.get(exam.id) if user.role == 'candidato' else None
            if ctx:
                d['group_id'] = ctx['group_id']
                d['group_exam_id'] = ctx['group_exam_id']
                # Agregar info de vigencia
                ge_obj = GroupExam.query.get(ctx['group_exam_id'])
                if ge_obj:
                    d['validity_months'] = ge_obj.validity_months
                    d['expires_at'] = ge_obj.effective_expires_at.isoformat() if ge_obj.effective_expires_at else None
                    d['extended_months'] = ge_obj.extended_months or 0
                    d['is_expired'] = ge_obj.is_expired
            # Agregar estado de aprobación
            approval = approved_map.get(exam.id)
            if approval:
                d['is_approved'] = approval['is_approved']
                d['best_score'] = approval['best_score']
            else:
                d['is_approved'] = False
                d['best_score'] = None
            exams_data.append(d)
        
        return jsonify({
            'exams': exams_data,
            'total': len(exams),
            'pages': 1,
            'current_page': 1
        })
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@bp.route('/mis-materiales', methods=['GET'])
@jwt_required()
def get_mis_materiales():
    """Obtener materiales de estudio asignados al candidato basándose en sus grupos"""
    from app.models.study_content import StudyMaterial
    from app.models.partner import GroupExam, GroupMember, GroupExamMaterial
    
    try:
        user_id = get_jwt_identity()
        user = User.query.get(user_id)
        
        if not user:
            return jsonify({'error': 'Usuario no encontrado'}), 404
        
        # Solo para candidatos y responsables
        if user.role not in ['candidato', 'responsable']:
            return jsonify({'error': 'Acceso no autorizado'}), 403
        
        # Si es responsable, obtener materiales del plantel
        if user.role == 'responsable':
            if not user.campus_id:
                return jsonify({'materials': [], 'total': 0, 'pages': 1, 'current_page': 1})
            
            campus = Campus.query.get(user.campus_id)
            if not campus:
                return jsonify({'materials': [], 'total': 0, 'pages': 1, 'current_page': 1})
            
            # Obtener grupos del plantel
            groups = CandidateGroup.query.filter_by(campus_id=campus.id, is_active=True).all()
            group_ids = [g.id for g in groups]
            
            # Obtener exámenes asignados a los grupos
            group_exams = GroupExam.query.filter(
                GroupExam.group_id.in_(group_ids),
                GroupExam.is_active == True
            ).all()
            
            # Obtener materiales de esos exámenes
            material_ids = set()
            for ge in group_exams:
                materials = GroupExamMaterial.query.filter_by(
                    group_exam_id=ge.id,
                    is_included=True
                ).all()
                for m in materials:
                    material_ids.add(m.study_material_id)
        else:
            # Para candidatos: obtener grupos donde es miembro activo
            memberships = GroupMember.query.filter_by(
                user_id=user_id,
                status='active'
            ).all()
            
            if not memberships:
                return jsonify({'materials': [], 'total': 0, 'pages': 1, 'current_page': 1})
            
            # Filtrar solo grupos que existen y están activos
            group_ids = []
            for m in memberships:
                group = CandidateGroup.query.filter_by(id=m.group_id, is_active=True).first()
                if group:
                    group_ids.append(m.group_id)
            
            if not group_ids:
                return jsonify({'materials': [], 'total': 0, 'pages': 1, 'current_page': 1})
            
            # Obtener exámenes asignados a esos grupos
            group_exams = GroupExam.query.filter(
                GroupExam.group_id.in_(group_ids),
                GroupExam.is_active == True
            ).all()
            
            # Obtener materiales de esos exámenes
            material_ids = set()
            for ge in group_exams:
                # Verificar si el candidato tiene acceso al examen
                has_access = False
                if ge.assignment_type == 'all' or ge.assignment_type is None:
                    has_access = True
                elif ge.assignment_type == 'selected':
                    from app.models.partner import GroupExamMember
                    # La existencia del registro indica asignación
                    member_assignment = GroupExamMember.query.filter_by(
                        group_exam_id=ge.id,
                        user_id=user_id
                    ).first()
                    has_access = member_assignment is not None
                
                if has_access:
                    materials = GroupExamMaterial.query.filter_by(
                        group_exam_id=ge.id,
                        is_included=True
                    ).all()
                    for m in materials:
                        material_ids.add(m.study_material_id)
        
        material_ids = list(material_ids)
        
        if not material_ids:
            return jsonify({'materials': [], 'total': 0, 'pages': 1, 'current_page': 1})
        
        # Obtener los materiales publicados
        materials = StudyMaterial.query.filter(
            StudyMaterial.id.in_(material_ids),
            StudyMaterial.is_published == True
        ).order_by(StudyMaterial.updated_at.desc()).all()
        
        return jsonify({
            'materials': [m.to_dict() for m in materials],
            'total': len(materials),
            'pages': 1,
            'current_page': 1
        })
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


# ============== ELIMINACIÓN COMPLETA DE GRUPOS (SOLO ADMIN) ==============

@bp.route('/groups/<int:group_id>/hard-delete', methods=['DELETE'])
@jwt_required()
def hard_delete_group(group_id):
    """
    Eliminar completamente un grupo y todas sus relaciones.
    SOLO PARA USUARIOS ADMIN.
    Elimina:
    - GroupMember (membresías)
    - GroupExamMember (asignaciones individuales de exámenes)
    - GroupExamMaterial (materiales de exámenes)
    - GroupExam (exámenes asignados al grupo)
    - GroupStudyMaterialMember (asignaciones individuales de materiales)
    - GroupStudyMaterial (materiales asignados directamente)
    - CandidateGroup (el grupo)
    """
    from app.models.partner import (
        GroupMember, GroupExam, GroupExamMember, GroupExamMaterial,
        GroupStudyMaterial, GroupStudyMaterialMember
    )
    
    try:
        user_id = get_jwt_identity()
        user = User.query.get(user_id)
        
        if not user or user.role not in ['admin', 'developer']:
            return jsonify({'error': 'Solo los administradores pueden eliminar grupos permanentemente'}), 403
        
        group = CandidateGroup.query.get(group_id)
        
        if not group:
            return jsonify({'error': 'Grupo no encontrado'}), 404
        
        group_name = group.name
        
        # Estadísticas de lo que se va a eliminar
        stats = {
            'members': 0,
            'exam_assignments': 0,
            'exam_member_assignments': 0,
            'exam_materials': 0,
            'study_materials': 0,
            'study_material_members': 0
        }
        
        # 1. Eliminar asignaciones de miembros a exámenes del grupo (GroupExamMember)
        group_exams = GroupExam.query.filter_by(group_id=group_id).all()
        for ge in group_exams:
            # Eliminar materiales del examen
            materials_deleted = GroupExamMaterial.query.filter_by(group_exam_id=ge.id).delete()
            stats['exam_materials'] += materials_deleted
            
            # Eliminar asignaciones individuales
            members_deleted = GroupExamMember.query.filter_by(group_exam_id=ge.id).delete()
            stats['exam_member_assignments'] += members_deleted
        
        # 2. Eliminar exámenes asignados al grupo (GroupExam)
        stats['exam_assignments'] = GroupExam.query.filter_by(group_id=group_id).delete()
        
        # 3. Eliminar asignaciones de materiales de estudio directos
        group_materials = GroupStudyMaterial.query.filter_by(group_id=group_id).all()
        for gm in group_materials:
            # Eliminar asignaciones individuales de materiales
            deleted = GroupStudyMaterialMember.query.filter_by(group_study_material_id=gm.id).delete()
            stats['study_material_members'] += deleted
        
        stats['study_materials'] = GroupStudyMaterial.query.filter_by(group_id=group_id).delete()
        
        # 4. Eliminar membresías (GroupMember)
        stats['members'] = GroupMember.query.filter_by(group_id=group_id).delete()
        
        # 5. Eliminar el grupo
        db.session.delete(group)
        db.session.commit()
        
        return jsonify({
            'message': f'Grupo "{group_name}" eliminado permanentemente',
            'deleted': stats
        })
        
    except Exception as e:
        db.session.rollback()
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@bp.route('/admin/cleanup-orphan-memberships', methods=['POST'])
@jwt_required()
def cleanup_orphan_memberships():
    """
    Limpia membresías huérfanas (de grupos que ya no existen).
    SOLO PARA USUARIOS ADMIN.
    """
    from app.models.partner import GroupMember
    
    try:
        user_id = get_jwt_identity()
        user = User.query.get(user_id)
        
        if not user or user.role not in ['admin', 'developer']:
            return jsonify({'error': 'Solo los administradores pueden ejecutar esta operación'}), 403
        
        # Encontrar membresías huérfanas
        orphan_memberships = db.session.query(GroupMember).outerjoin(
            CandidateGroup, GroupMember.group_id == CandidateGroup.id
        ).filter(CandidateGroup.id == None).all()
        
        count = len(orphan_memberships)
        
        for m in orphan_memberships:
            db.session.delete(m)
        
        db.session.commit()
        
        return jsonify({
            'message': f'Se eliminaron {count} membresías huérfanas',
            'deleted_count': count
        })
        
    except Exception as e:
        db.session.rollback()
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


# ============== ASIGNACIÓN MASIVA DE EXÁMENES POR ECM ==============

@bp.route('/groups/<int:group_id>/exams/bulk-assign-template', methods=['GET'])
@jwt_required()
@coordinator_required
def download_bulk_exam_assign_template(group_id):
    """
    Descargar plantilla Excel para asignación masiva de exámenes.
    La plantilla incluye los miembros del grupo para indicar a quiénes asignar.
    El ECM ya fue seleccionado en el paso 1, aquí solo se indica qué usuarios incluir.
    """
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from flask import send_file
    
    try:
        group = CandidateGroup.query.get_or_404(group_id)
        
        # Obtener miembros del grupo (status='active')
        members = group.members.filter_by(status='active').all()
        if not members:
            return jsonify({'error': 'El grupo no tiene miembros activos'}), 400
        
        # Crear workbook
        wb = Workbook()
        
        # === Hoja 1: Lista de Usuarios ===
        ws = wb.active
        ws.title = "Usuarios a Asignar"
        
        # Estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
        required_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Headers - Solo Nombre de Usuario
        headers = ['Nombre de Usuario', 'Nombre Completo (opcional)']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')
        
        # Datos de miembros
        for row, member in enumerate(members, 2):
            user = member.user
            if user:
                ws.cell(row=row, column=1, value=user.username).border = thin_border
                ws.cell(row=row, column=2, value=user.full_name or '').border = thin_border
        
        # Ajustar anchos de columna
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 45
        
        # === Hoja 2: Instrucciones ===
        ws2 = wb.create_sheet(title="Instrucciones")
        
        instructions = [
            ("INSTRUCCIONES DE USO", None),
            ("", None),
            ("1. Este archivo contiene la lista de usuarios del grupo.", None),
            ("2. El examen a asignar ya fue seleccionado en el paso 1.", None),
            ("3. ELIMINE las filas de los usuarios a quienes NO desea asignar el examen.", None),
            ("4. Mantenga solo los usuarios que SÍ deben recibir la asignación.", None),
            ("5. Suba este archivo para procesar las asignaciones.", None),
            ("", None),
            ("IDENTIFICACIÓN DE USUARIOS:", None),
            ("- Se identifica a los usuarios por su 'Nombre de Usuario'.", None),
            ("- El campo 'Nombre Completo' es solo informativo.", None),
            ("", None),
            ("NOTAS IMPORTANTES:", None),
            ("- Si un usuario ya tiene el examen asignado, se omitirá.", None),
            ("- Solo se procesarán usuarios que pertenezcan al grupo.", None),
        ]
        
        for row, (text, _) in enumerate(instructions, 1):
            cell = ws2.cell(row=row, column=1, value=text)
            if row == 1 or row == 9 or row == 14:
                cell.font = Font(bold=True, size=12)
            else:
                cell.font = Font(size=11)
        
        ws2.column_dimensions['A'].width = 70
        
        # Guardar en memoria
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"plantilla_asignacion_{group.name.replace(' ', '_')}_{group_id}.xlsx"
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@bp.route('/groups/<int:group_id>/exams/bulk-assign', methods=['POST'])
@jwt_required()
@coordinator_required
def bulk_assign_exams_by_ecm(group_id):
    """
    Procesar archivo Excel y asignar exámenes seleccionados masivamente.
    
    El ECM ya fue seleccionado en paso 1, aquí solo se indica qué usuarios.
    El archivo debe tener al menos UNA de estas columnas para identificar usuarios:
    - Nombre de Usuario (username)
    - Email
    - CURP
    
    Parámetros form-data:
    - file: Archivo Excel
    - ecm_code: Código ECM del examen a asignar (requerido)
    - time_limit_minutes: Límite de tiempo (opcional)
    - passing_score: Puntaje mínimo (opcional)
    - max_attempts: Máximo de intentos (default: 2)
    - max_disconnections: Máximo de desconexiones (default: 3)
    - exam_content_type: 'questions_only', 'exercises_only', 'mixed' (default: 'questions_only')
    - dry_run: 'true' para solo previsualizar sin crear asignaciones
    """
    import io
    from openpyxl import load_workbook
    from app.models import GroupExam, Exam, User
    from app.models.partner import GroupExamMember, EcmCandidateAssignment
    from app.models.competency_standard import CompetencyStandard
    
    try:
        group = CandidateGroup.query.get_or_404(group_id)
        
        # Verificar que se envió un archivo
        if 'file' not in request.files:
            return jsonify({'error': 'No se envió ningún archivo'}), 400
        
        file = request.files['file']
        if not file.filename.endswith(('.xlsx', '.xls')):
            return jsonify({'error': 'El archivo debe ser Excel (.xlsx o .xls)'}), 400
        
        # Obtener código ECM (requerido)
        ecm_code = request.form.get('ecm_code', '').strip().upper()
        if not ecm_code:
            return jsonify({'error': 'Debe especificar el código ECM del examen a asignar'}), 400
        
        # Buscar el ECM y su examen publicado
        ecm = CompetencyStandard.query.filter(
            db.func.upper(CompetencyStandard.code) == ecm_code,
            CompetencyStandard.is_active == True
        ).first()
        
        if not ecm:
            return jsonify({'error': f'Código ECM "{ecm_code}" no encontrado'}), 404
        
        exam = ecm.exams.filter_by(is_active=True, is_published=True).order_by(
            Exam.created_at.desc()
        ).first()
        
        if not exam:
            return jsonify({'error': f'No hay examen publicado para el ECM "{ecm_code}"'}), 400
        
        # Modo dry_run: solo previsualizar sin crear asignaciones
        dry_run = request.form.get('dry_run', 'false').lower() == 'true'
        
        # Obtener configuración adicional
        config = {
            'time_limit_minutes': request.form.get('time_limit_minutes', type=int),
            'passing_score': request.form.get('passing_score', type=int),
            'max_attempts': request.form.get('max_attempts', 2, type=int),
            'max_disconnections': request.form.get('max_disconnections', 3, type=int),
            'exam_content_type': request.form.get('exam_content_type', 'questions_only'),
        }
        
        # Leer archivo Excel
        wb = load_workbook(io.BytesIO(file.read()))
        ws = wb.active
        
        # Obtener headers
        headers = [cell.value for cell in ws[1]]
        
        # Encontrar columna de username
        username_col = None
        
        for idx, header in enumerate(headers):
            if header:
                header_lower = str(header).lower().strip()
                if 'nombre de usuario' in header_lower or 'username' in header_lower:
                    username_col = idx
        
        if username_col is None:
            return jsonify({
                'error': 'El archivo debe tener la columna "Nombre de Usuario"'
            }), 400
        
        # Procesar filas
        results = {
            'processed': 0,
            'assigned': [],
            'skipped': [],
            'errors': []
        }
        
        # Obtener miembros del grupo con sus usuarios
        group_members = {m.user_id: m.user for m in group.members.filter_by(status='active').all() if m.user}
        
        # Crear índice para búsqueda rápida por username
        users_by_username = {u.username.lower(): uid for uid, u in group_members.items() if u.username}
        
        # Verificar/crear GroupExam (solo si no es dry_run)
        group_exam = GroupExam.query.filter_by(
            group_id=group_id, 
            exam_id=exam.id,
            is_active=True
        ).first()
        
        if not group_exam and not dry_run:
            group_exam = GroupExam(
                group_id=group_id,
                exam_id=exam.id,
                assigned_by_id=g.current_user.id,
                assignment_type='selected',
                time_limit_minutes=config['time_limit_minutes'],
                passing_score=config['passing_score'],
                max_attempts=config['max_attempts'],
                max_disconnections=config['max_disconnections'],
                exam_content_type=config['exam_content_type']
            )
            db.session.add(group_exam)
            db.session.flush()
        
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            # Obtener username
            username = str(row[username_col]).strip() if username_col is not None and row[username_col] else None
            
            # Si no hay username, saltar fila
            if not username:
                continue
            
            results['processed'] += 1
            
            # Buscar usuario por username
            user_id = None
            if username.lower() in users_by_username:
                user_id = users_by_username[username.lower()]
            
            if not user_id:
                results['errors'].append({
                    'row': row_num,
                    'identifier': username,
                    'user_name': username,
                    'error': 'Usuario no encontrado en el grupo'
                })
                continue
            
            # Verificar si el usuario ya tiene este examen asignado
            if group_exam:
                existing_member = GroupExamMember.query.filter_by(
                    group_exam_id=group_exam.id,
                    user_id=user_id
                ).first()
                
                if existing_member:
                    user_info = group_members.get(user_id)
                    results['skipped'].append({
                        'row': row_num,
                        'user_id': str(user_id),
                        'username': username,
                        'user_name': user_info.full_name if user_info else username,
                        'email': user_info.email if user_info else '',
                        'curp': user_info.curp if user_info else '',
                        'reason': 'Usuario ya tiene este examen asignado'
                    })
                    continue
                
                # Si el GroupExam es tipo 'all', verificar
                if group_exam.assignment_type == 'all':
                    user_info = group_members.get(user_id)
                    results['skipped'].append({
                        'row': row_num,
                        'user_id': str(user_id),
                        'username': username,
                        'user_name': user_info.full_name if user_info else username,
                        'email': user_info.email if user_info else '',
                        'curp': user_info.curp if user_info else '',
                        'reason': 'El examen está asignado a todo el grupo'
                    })
                    continue
            
            # Crear asignación de miembro (solo si no es dry_run)
            if not dry_run:
                member = GroupExamMember(
                    group_exam_id=group_exam.id,
                    user_id=user_id
                )
                db.session.add(member)
                
                # Crear registro ECA con número de asignación
                ecm_id = exam.competency_standard_id
                if ecm_id:
                    existing_ecm = EcmCandidateAssignment.query.filter_by(
                        user_id=user_id,
                        competency_standard_id=ecm_id
                    ).first()
                    if not existing_ecm:
                        new_ecm = EcmCandidateAssignment(
                            assignment_number=EcmCandidateAssignment.generate_assignment_number(),
                            user_id=user_id,
                            competency_standard_id=ecm_id,
                            exam_id=exam.id,
                            campus_id=group.campus_id,
                            group_id=group_id,
                            group_name=group.name,
                            group_exam_id=group_exam.id,
                            assigned_by_id=g.current_user.id,
                            assignment_source='bulk_upload'
                        )
                        db.session.add(new_ecm)
            
            user = group_members[user_id]
            results['assigned'].append({
                'row': row_num,
                'user_id': str(user_id),
                'username': username,
                'user_name': user.full_name if user else username,
                'email': user.email if user else '',
                'curp': user.curp if user else '',
                'exam_name': exam.name
            })
        
        if not dry_run:
            db.session.commit()
        
        action_word = 'previsualizadas' if dry_run else 'realizadas'
        return jsonify({
            'message': f'Procesamiento completado. {len(results["assigned"])} asignaciones {action_word}.',
            'dry_run': dry_run,
            'results': results,
            'summary': {
                'total_processed': results['processed'],
                'assigned': len(results['assigned']),
                'skipped': len(results['skipped']),
                'errors': len(results['errors'])
            }
        })
        
    except Exception as e:
        db.session.rollback()
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


# ============== CERTIFICADOS DEL GRUPO ==============

@bp.route('/groups/<int:group_id>/certificates/stats', methods=['GET'])
@jwt_required()
@coordinator_required
def get_group_certificates_stats(group_id):
    """
    Obtener estadísticas de certificados del grupo con paginación server-side.
    
    Query params:
    - cert_type: 'tier_basic' | 'tier_standard' | 'tier_advanced' | 'digital_badge' (opcional)
    - page (int, default 1)
    - per_page (int, default 150, max 1000)
    - search (str, opcional) — busca en full_name, email, curp
    - sort_by (str, default 'full_name') — full_name, email, curp, status, ready_count
    - sort_dir (str, default 'asc')
    - filter_status (str, opcional) — 'ready', 'pending', 'all'
    
    Si NO se envía page, devuelve solo summary sin candidates (para el hub).
    """
    try:
        from app.models.result import Result
        from app.models.conocer_certificate import ConocerCertificate
        from app.models.exam import Exam
        from sqlalchemy import and_, or_
        
        group = CandidateGroup.query.get_or_404(group_id)
        
        cert_type = request.args.get('cert_type', None)
        include_candidates = request.args.get('page') is not None
        page = request.args.get('page', 1, type=int)
        per_page = min(request.args.get('per_page', 150, type=int), 1000)
        search = request.args.get('search', '').strip()
        sort_by = request.args.get('sort_by', 'full_name')
        sort_dir_param = request.args.get('sort_dir', 'asc')
        filter_status = request.args.get('filter_status', 'all')
        
        # Obtener configuración efectiva del grupo
        if group.campus:
            enable_tier_basic = group.enable_tier_basic_override if group.enable_tier_basic_override is not None else group.campus.enable_tier_basic
            enable_tier_standard = group.enable_tier_standard_override if group.enable_tier_standard_override is not None else group.campus.enable_tier_standard
            enable_tier_advanced = group.enable_tier_advanced_override if group.enable_tier_advanced_override is not None else group.campus.enable_tier_advanced
            enable_digital_badge = group.enable_digital_badge_override if group.enable_digital_badge_override is not None else group.campus.enable_digital_badge
        else:
            enable_tier_basic = group.enable_tier_basic_override or False
            enable_tier_standard = group.enable_tier_standard_override or False
            enable_tier_advanced = group.enable_tier_advanced_override or False
            enable_digital_badge = group.enable_digital_badge_override or False
        
        # Obtener miembros activos del grupo
        members = GroupMember.query.filter_by(group_id=group_id, status='active').all()
        member_user_ids = [m.user_id for m in members]
        
        # Obtener resultados de los miembros (aprobados)
        results = Result.query.filter(
            and_(
                Result.user_id.in_(member_user_ids),
                Result.status == 1,  # completado
                Result.result == 1   # aprobado
            )
        ).all() if member_user_ids else []
        
        # Obtener certificados CONOCER de los miembros
        conocer_certs = ConocerCertificate.query.filter(
            and_(
                ConocerCertificate.user_id.in_(member_user_ids),
                ConocerCertificate.status == 'active'
            )
        ).all() if member_user_ids else []
        
        # Calcular estadísticas por tipo
        tier_basic_count = 0
        tier_standard_count = 0
        tier_basic_pending = 0
        tier_standard_pending = 0
        
        for r in results:
            if r.report_url or r.certificate_code:
                tier_basic_count += 1
            else:
                tier_basic_pending += 1
            
            if r.certificate_url or r.eduit_certificate_code:
                tier_standard_count += 1
            else:
                tier_standard_pending += 1
        
        tier_advanced_count = len(conocer_certs)
        digital_badge_count = len(results)
        
        # Contar usuarios únicos con resultados aprobados (certificados)
        certified_user_ids = set(r.user_id for r in results)
        
        response = {
            'group': {
                'id': group.id,
                'name': group.name,
                'member_count': len(members)
            },
            'config': {
                'enable_tier_basic': enable_tier_basic,
                'enable_tier_standard': enable_tier_standard,
                'enable_tier_advanced': enable_tier_advanced,
                'enable_digital_badge': enable_digital_badge
            },
            'summary': {
                'total_exams_approved': len(results),
                'total_certified': len(certified_user_ids),
                'tier_basic': {
                    'enabled': enable_tier_basic,
                    'ready': tier_basic_count,
                    'pending': tier_basic_pending,
                    'total': tier_basic_count + tier_basic_pending
                },
                'tier_standard': {
                    'enabled': enable_tier_standard,
                    'ready': tier_standard_count,
                    'pending': tier_standard_pending,
                    'total': tier_standard_count + tier_standard_pending
                },
                'tier_advanced': {
                    'enabled': enable_tier_advanced,
                    'count': tier_advanced_count
                },
                'digital_badge': {
                    'enabled': enable_digital_badge,
                    'count': digital_badge_count
                }
            }
        }
        
        # Si no se solicita paginación, devolver solo summary (hub de documentos)
        if not include_candidates:
            return jsonify(response)
        
        # ----- Construir candidates con paginación -----
        user_results_map = {}
        for r in results:
            user_results_map.setdefault(r.user_id, []).append(r)
        
        user_conocer_map = {}
        for c in conocer_certs:
            user_conocer_map.setdefault(c.user_id, []).append(c)
        
        candidates_stats = []
        for member in members:
            user = member.user
            if not user:
                continue
            
            user_results = user_results_map.get(member.user_id, [])
            user_conocer = user_conocer_map.get(member.user_id, [])
            
            exams_approved = len(user_results)
            tb_ready = sum(1 for r in user_results if r.report_url or r.certificate_code)
            tb_pending = sum(1 for r in user_results if not r.report_url and not r.certificate_code)
            ts_ready = sum(1 for r in user_results if r.certificate_url or r.eduit_certificate_code)
            ts_pending = sum(1 for r in user_results if not r.certificate_url and not r.eduit_certificate_code)
            ta_count = len(user_conocer)
            db_count = len(user_results)
            
            # Determinar ready/pending para el tipo solicitado
            if cert_type == 'tier_basic':
                ready_count = tb_ready
                pending_count = tb_pending
            elif cert_type == 'tier_standard':
                ready_count = ts_ready
                pending_count = ts_pending
            elif cert_type == 'tier_advanced':
                ready_count = ta_count
                pending_count = 0
            elif cert_type == 'digital_badge':
                ready_count = db_count
                pending_count = 0
            else:
                ready_count = tb_ready + ts_ready + ta_count + db_count
                pending_count = tb_pending + ts_pending
            
            total_for_type = ready_count + pending_count
            
            # Filtrar candidatos sin resultados aprobados si no aplica
            if exams_approved == 0:
                continue
            # Para tier_advanced y digital_badge, solo mostrar si tienen al menos 1
            if cert_type in ('tier_advanced', 'digital_badge') and total_for_type == 0:
                continue
            
            stat_status = 'pending' if pending_count > 0 else 'ready'
            
            candidates_stats.append({
                'user_id': member.user_id,
                'full_name': user.full_name,
                'email': user.email,
                'curp': getattr(user, 'curp', None),
                'exams_approved': exams_approved,
                'tier_basic_ready': tb_ready,
                'tier_basic_pending': tb_pending,
                'tier_standard_ready': ts_ready,
                'tier_standard_pending': ts_pending,
                'tier_advanced_count': ta_count,
                'digital_badge_count': db_count,
                'ready_count': ready_count,
                'pending_count': pending_count,
                'status': stat_status,
            })
        
        # Búsqueda
        if search:
            q = search.lower()
            candidates_stats = [c for c in candidates_stats if
                q in c['full_name'].lower() or
                q in (c['email'] or '').lower() or
                q in (c['curp'] or '').lower()
            ]
        
        # Filtro de estado
        if filter_status == 'ready':
            candidates_stats = [c for c in candidates_stats if c['status'] == 'ready']
        elif filter_status == 'pending':
            candidates_stats = [c for c in candidates_stats if c['status'] == 'pending']
        
        total_candidates = len(candidates_stats)
        
        # Ordenamiento
        def sort_key(c):
            if sort_by == 'full_name':
                return c['full_name'].lower()
            elif sort_by == 'email':
                return (c['email'] or '').lower()
            elif sort_by == 'curp':
                return (c['curp'] or '').lower()
            elif sort_by == 'status':
                return c['status']
            elif sort_by == 'ready_count':
                return c['ready_count']
            return c['full_name'].lower()
        
        candidates_stats.sort(key=sort_key, reverse=(sort_dir_param == 'desc'))
        
        # Paginación
        total_pages = max(1, (total_candidates + per_page - 1) // per_page)
        page = min(page, total_pages)
        start = (page - 1) * per_page
        page_candidates = candidates_stats[start:start + per_page]
        
        response['candidates'] = page_candidates
        response['total'] = total_candidates
        response['page'] = page
        response['pages'] = total_pages
        response['per_page'] = per_page
        
        return jsonify(response)
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@bp.route('/groups/<int:group_id>/certificates/download', methods=['POST'])
@jwt_required()
@coordinator_required
def download_group_certificates_zip(group_id):
    """
    Descargar certificados del grupo en formato ZIP.
    Genera los PDFs que no existan antes de crear el ZIP.
    
    Body:
    - certificate_types: Array de tipos a incluir ['tier_basic', 'tier_standard', 'tier_advanced']
    - user_ids: (opcional) Lista de user_ids específicos, o todo el grupo si no se especifica
    """
    try:
        from app.models.result import Result
        from app.models.conocer_certificate import ConocerCertificate
        from app.models.exam import Exam
        from flask import current_app
        from io import BytesIO
        from zipfile import ZipFile
        import requests
        from datetime import datetime
        from sqlalchemy import and_
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import letter
        from reportlab.pdfgen import canvas
        from reportlab.lib.colors import HexColor
        from reportlab.pdfbase.pdfmetrics import stringWidth
        import re
        import os
        
        group = CandidateGroup.query.get_or_404(group_id)
        
        data = request.get_json() or {}
        certificate_types = data.get('certificate_types', ['tier_basic', 'tier_standard'])
        user_ids = data.get('user_ids', None)
        
        # --- FUNCIONES DE GENERACIÓN DE PDF ---
        def strip_html(text):
            if not text:
                return ''
            return re.sub(r'<[^>]+>', '', str(text))
        
        def generate_evaluation_report_pdf(result, exam, user):
            """Genera el PDF del reporte de evaluación"""
            buffer = BytesIO()
            page_width, page_height = letter
            margin = 50
            
            c = canvas.Canvas(buffer, pagesize=letter)
            
            primary_color = HexColor('#1e40af')
            success_color = HexColor('#16a34a')
            error_color = HexColor('#dc2626')
            gray_color = HexColor('#6b7280')
            
            y = page_height - margin
            
            # ENCABEZADO
            c.setFillColor(colors.black)
            c.setFont('Helvetica-Bold', 16)
            c.drawString(margin, y, 'Evaluaasi')
            
            c.setFillColor(gray_color)
            c.setFont('Helvetica', 7)
            c.drawRightString(page_width - margin, y, 'Sistema de Evaluación y Certificación')
            c.drawRightString(page_width - margin, y - 12, datetime.now().strftime('%d/%m/%Y %H:%M'))
            
            y -= 45
            c.setStrokeColor(primary_color)
            c.setLineWidth(2)
            c.line(margin, y, page_width - margin, y)
            
            y -= 30
            
            # TÍTULO
            c.setFillColor(colors.black)
            c.setFont('Helvetica-Bold', 14)
            c.drawCentredString(page_width / 2, y, 'REPORTE DE EVALUACIÓN')
            y -= 30
            
            # DATOS DEL ESTUDIANTE
            c.setFillColor(primary_color)
            c.setFont('Helvetica-Bold', 10)
            c.drawString(margin, y, 'DATOS DEL ESTUDIANTE')
            y -= 15
            
            c.setFillColor(colors.black)
            name_parts = [user.name or '']
            if hasattr(user, 'first_surname') and user.first_surname:
                name_parts.append(user.first_surname)
            if hasattr(user, 'second_surname') and user.second_surname:
                name_parts.append(user.second_surname)
            student_name = ' '.join(name_parts).strip() or user.email
            
            c.setFont('Helvetica-Bold', 9)
            c.drawString(margin + 5, y, 'Nombre:')
            c.setFont('Helvetica', 9)
            c.drawString(margin + 50, y, student_name)
            y -= 12
            c.setFont('Helvetica-Bold', 9)
            c.drawString(margin + 5, y, 'Correo:')
            c.setFont('Helvetica', 9)
            c.drawString(margin + 50, y, user.email)
            y -= 20
            
            # DATOS DEL EXAMEN
            c.setFillColor(primary_color)
            c.setFont('Helvetica-Bold', 10)
            c.drawString(margin, y, 'DATOS DEL EXAMEN')
            y -= 15
            
            c.setFillColor(colors.black)
            exam_name = strip_html(exam.name)[:60] if exam.name else 'Sin nombre'
            c.setFont('Helvetica-Bold', 9)
            c.drawString(margin + 5, y, 'Examen:')
            c.setFont('Helvetica', 9)
            c.drawString(margin + 55, y, exam_name)
            y -= 12
            
            ecm_code = exam.version or 'N/A'
            c.setFont('Helvetica-Bold', 9)
            c.drawString(margin + 5, y, 'Código ECM:')
            c.setFont('Helvetica', 9)
            c.drawString(margin + 70, y, ecm_code)
            y -= 12
            
            start_date = result.start_date.strftime('%d/%m/%Y %H:%M') if result.start_date else 'N/A'
            c.setFont('Helvetica-Bold', 9)
            c.drawString(margin + 5, y, 'Fecha:')
            c.setFont('Helvetica', 9)
            c.drawString(margin + 40, y, start_date)
            y -= 25
            
            # RESULTADO
            c.setFillColor(primary_color)
            c.setFont('Helvetica-Bold', 10)
            c.drawString(margin, y, 'RESULTADO DE LA EVALUACIÓN')
            y -= 10
            
            box_height = 40
            c.setStrokeColor(colors.black)
            c.setLineWidth(0.5)
            c.rect(margin, y - box_height, page_width - 2 * margin, box_height)
            
            passing_score = exam.passing_score or 70
            is_passed = result.result == 1
            percentage = result.score or 0
            
            c.setFillColor(colors.black)
            c.setFont('Helvetica-Bold', 9)
            c.drawString(margin + 10, y - 15, 'Calificación:')
            c.setFont('Helvetica-Bold', 18)
            c.drawString(margin + 70, y - 18, f'{percentage}%')
            
            c.setFont('Helvetica-Bold', 9)
            c.drawString(margin + 10, y - 35, 'Resultado:')
            
            if is_passed:
                c.setFillColor(success_color)
                c.setFont('Helvetica-Bold', 12)
                c.drawString(margin + 60, y - 37, 'APROBADO')
            else:
                c.setFillColor(error_color)
                c.setFont('Helvetica-Bold', 12)
                c.drawString(margin + 60, y - 37, 'NO APROBADO')
            
            c.setFillColor(colors.black)
            c.setFont('Helvetica', 9)
            c.drawRightString(page_width - margin - 10, y - 35, f'Puntaje mínimo requerido: {passing_score}%')
            
            y -= 60
            
            if result.certificate_code:
                c.setFillColor(gray_color)
                c.setFont('Helvetica', 8)
                c.drawCentredString(page_width / 2, y, f'Código de verificación: {result.certificate_code}')
            
            c.save()
            buffer.seek(0)
            return buffer
        
        def generate_certificate_pdf_with_template(result, exam, user):
            """Genera el certificado PDF usando plantilla si existe"""
            from pypdf import PdfReader, PdfWriter
            import qrcode
            from reportlab.lib.utils import ImageReader
            
            template_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'static', 'plantilla.pdf')
            
            if not os.path.exists(template_path):
                # Solo generar reporte si no hay plantilla
                return generate_evaluation_report_pdf(result, exam, user)
            
            try:
                reader = PdfReader(template_path)
                page = reader.pages[0]
                width = float(page.mediabox.width)
                height = float(page.mediabox.height)
                
                buffer_overlay = BytesIO()
                c = canvas.Canvas(buffer_overlay, pagesize=(width, height))
                c.setFillColor(HexColor('#1a365d'))
                
                # Configuración del área compartida para nombre y certificado
                x_min = 85
                x_max = 540
                max_width = x_max - x_min
                center_x = (x_min + x_max) / 2
                
                # Función para ajustar tamaño de fuente
                def draw_fitted_text(canv, text, cx, y_pos, mw, font_name, max_font_size, min_font_size=8):
                    font_sz = max_font_size
                    while font_sz >= min_font_size:
                        text_width = stringWidth(text, font_name, font_sz)
                        if text_width <= mw:
                            canv.setFont(font_name, font_sz)
                            canv.drawCentredString(cx, y_pos, text)
                            return font_sz
                        font_sz -= 1
                    canv.setFont(font_name, min_font_size)
                    canv.drawCentredString(cx, y_pos, text)
                    return min_font_size
                
                # Nombre del usuario (Title Case)
                name_parts = [user.name or '']
                if hasattr(user, 'first_surname') and user.first_surname:
                    name_parts.append(user.first_surname)
                if hasattr(user, 'second_surname') and user.second_surname:
                    name_parts.append(user.second_surname)
                student_name = ' '.join(name_parts).strip() or user.email
                student_name = student_name.title()
                
                # NOMBRE en (center_x, 375), max 36pt - POSICIÓN CORRECTA
                draw_fitted_text(c, student_name, center_x, 375, max_width, 'Helvetica-Bold', 36)
                
                # Nombre del certificado (MAYÚSCULAS)
                cert_name = exam.name.upper() if exam.name else "CERTIFICADO DE COMPETENCIA"
                cert_name = strip_html(cert_name)[:80]
                
                # CERTIFICADO en (center_x, 300), max 18pt - POSICIÓN CORRECTA
                draw_fitted_text(c, cert_name, center_x, 300, max_width, 'Helvetica-Bold', 18)
                
                # === QR DE VERIFICACIÓN ===
                verification_code = result.eduit_certificate_code
                if verification_code:
                    # URL de verificación
                    verify_url = f"https://app.evaluaasi.com/verify/{verification_code}"
                    
                    # Crear QR
                    qr = qrcode.QRCode(
                        version=1,
                        error_correction=qrcode.constants.ERROR_CORRECT_M,
                        box_size=3,
                        border=1,
                    )
                    qr.add_data(verify_url)
                    qr.make(fit=True)
                    
                    # Crear imagen QR con fondo transparente
                    qr_img = qr.make_image(fill_color="black", back_color="white").convert('RGBA')
                    
                    # Hacer el fondo transparente
                    data = qr_img.getdata()
                    new_data = []
                    for item in data:
                        # Si es blanco, hacerlo transparente
                        if item[0] > 240 and item[1] > 240 and item[2] > 240:
                            new_data.append((255, 255, 255, 0))
                        else:
                            new_data.append(item)
                    qr_img.putdata(new_data)
                    
                    # Guardar QR en buffer
                    qr_buffer = BytesIO()
                    qr_img.save(qr_buffer, format='PNG')
                    qr_buffer.seek(0)
                    
                    # Dibujar QR en el PDF (esquina inferior izquierda)
                    qr_image = ImageReader(qr_buffer)
                    qr_size = 55
                    qr_x = 25  # Lado izquierdo
                    qr_y = 25
                    c.drawImage(qr_image, qr_x, qr_y, width=qr_size, height=qr_size, mask='auto')
                    
                    # Código de verificación debajo del QR
                    c.setFillColor(HexColor('#444444'))
                    c.setFont('Helvetica-Bold', 7)
                    c.drawCentredString(qr_x + qr_size/2, qr_y - 8, verification_code)
                    c.setFont('Helvetica', 5)
                    c.drawCentredString(qr_x + qr_size/2, qr_y - 15, 'Escanea para verificar')
                
                c.save()
                buffer_overlay.seek(0)
                
                # Combinar con plantilla
                reader2 = PdfReader(template_path)
                page2 = reader2.pages[0]
                overlay_reader = PdfReader(buffer_overlay)
                page2.merge_page(overlay_reader.pages[0])
                
                writer = PdfWriter()
                writer.add_page(page2)
                
                output_buffer = BytesIO()
                writer.write(output_buffer)
                output_buffer.seek(0)
                
                return output_buffer
                
            except Exception as e:
                current_app.logger.error(f"Error generando certificado con plantilla: {e}")
                import traceback
                current_app.logger.error(traceback.format_exc())
                return generate_evaluation_report_pdf(result, exam, user)
        
        def upload_pdf_to_blob(buffer, filename):
            """Sube el PDF a Azure Blob Storage"""
            from azure.storage.blob import BlobServiceClient, ContentSettings
            
            connection_string = os.environ.get('AZURE_STORAGE_CONNECTION_STRING')
            container_name = os.environ.get('AZURE_STORAGE_CONTAINER', 'evaluaasi-files')
            
            if not connection_string:
                return None
            
            try:
                blob_service_client = BlobServiceClient.from_connection_string(connection_string)
                container_client = blob_service_client.get_container_client(container_name)
                
                try:
                    container_client.create_container()
                except:
                    pass
                
                blob_name = f"pdfs/{filename}"
                blob_client = container_client.get_blob_client(blob_name)
                
                buffer.seek(0)
                blob_client.upload_blob(
                    buffer,
                    overwrite=True,
                    content_settings=ContentSettings(content_type='application/pdf')
                )
                
                return blob_client.url
            except Exception as e:
                current_app.logger.error(f"Error subiendo a blob: {e}")
                return None
        
        # --- FIN FUNCIONES DE GENERACIÓN ---
        
        # Obtener miembros
        if user_ids:
            members = GroupMember.query.filter(
                and_(
                    GroupMember.group_id == group_id,
                    GroupMember.user_id.in_(user_ids),
                    GroupMember.status == 'active'
                )
            ).all()
        else:
            members = GroupMember.query.filter_by(group_id=group_id, status='active').all()
        
        member_user_ids = [m.user_id for m in members]
        
        if not member_user_ids:
            return jsonify({'error': 'No hay miembros en el grupo'}), 400
        
        # Obtener resultados aprobados
        results = Result.query.filter(
            and_(
                Result.user_id.in_(member_user_ids),
                Result.status == 1,
                Result.result == 1
            )
        ).all()
        
        # Pre-cargar exámenes
        exam_ids = list(set(r.exam_id for r in results))
        exams_dict = {e.id: e for e in Exam.query.filter(Exam.id.in_(exam_ids)).all()}
        
        # Obtener certificados CONOCER
        conocer_certs = []
        if 'tier_advanced' in certificate_types:
            conocer_certs = ConocerCertificate.query.filter(
                and_(
                    ConocerCertificate.user_id.in_(member_user_ids),
                    ConocerCertificate.status == 'active'
                )
            ).all()
        
        # Crear ZIP en memoria
        zip_buffer = BytesIO()
        files_added = 0
        generated = 0
        errors = []
        
        # Mapa de usuarios
        users_map = {}
        for m in members:
            if m.user:
                users_map[m.user_id] = m.user
        
        current_app.logger.info(f"Procesando {len(results)} resultados para {len(certificate_types)} tipos de certificado")
        
        with ZipFile(zip_buffer, 'w') as zip_file:
            # Procesar tier_basic (Reportes de evaluación)
            if 'tier_basic' in certificate_types:
                for r in results:
                    user = users_map.get(r.user_id)
                    if not user:
                        continue
                    
                    exam = exams_dict.get(r.exam_id)
                    if not exam:
                        continue
                    
                    safe_name = ''.join(c for c in user.full_name if c.isalnum() or c in ' -_').strip().replace(' ', '_')
                    exam_name = exam.name if exam else f'Examen_{r.exam_id}'
                    safe_exam = ''.join(c for c in exam_name if c.isalnum() or c in ' -_').strip().replace(' ', '_')[:30]
                    
                    folder = f"Constancias/{safe_name}"
                    filename = f"{folder}/{safe_exam}_Reporte_{r.certificate_code or r.id[:8]}.pdf"
                    
                    pdf_data = None
                    
                    if r.report_url:
                        # Intentar descargar PDF existente
                        try:
                            response = requests.get(r.report_url, timeout=30)
                            if response.status_code == 200:
                                pdf_data = response.content
                        except:
                            pass
                    
                    if not pdf_data:
                        # Generar PDF nuevo
                        try:
                            pdf_buffer = generate_evaluation_report_pdf(r, exam, user)
                            pdf_data = pdf_buffer.getvalue()
                            
                            # Subir a blob y guardar URL
                            blob_filename = f"report_{r.id}.pdf"
                            blob_url = upload_pdf_to_blob(pdf_buffer, blob_filename)
                            if blob_url:
                                r.report_url = blob_url
                                db.session.commit()
                            
                            generated += 1
                        except Exception as e:
                            errors.append({'user': user.full_name, 'type': 'tier_basic', 'error': str(e)})
                            continue
                    
                    if pdf_data:
                        zip_file.writestr(filename, pdf_data)
                        files_added += 1
            
            # Procesar tier_standard (Certificados Eduit)
            if 'tier_standard' in certificate_types:
                for r in results:
                    user = users_map.get(r.user_id)
                    if not user:
                        continue
                    
                    exam = exams_dict.get(r.exam_id)
                    if not exam:
                        continue
                    
                    safe_name = ''.join(c for c in user.full_name if c.isalnum() or c in ' -_').strip().replace(' ', '_')
                    exam_name = exam.name if exam else f'Examen_{r.exam_id}'
                    safe_exam = ''.join(c for c in exam_name if c.isalnum() or c in ' -_').strip().replace(' ', '_')[:30]
                    
                    folder = f"Certificados_Eduit/{safe_name}"
                    filename = f"{folder}/{safe_exam}_Certificado_{r.eduit_certificate_code or r.id[:8]}.pdf"
                    
                    pdf_data = None
                    
                    if r.certificate_url:
                        try:
                            response = requests.get(r.certificate_url, timeout=30)
                            if response.status_code == 200:
                                pdf_data = response.content
                        except:
                            pass
                    
                    if not pdf_data:
                        # Generar certificado nuevo
                        try:
                            # Asegurar que tenga código de certificado
                            if not r.eduit_certificate_code:
                                import uuid
                                r.eduit_certificate_code = f"EC{uuid.uuid4().hex[:10].upper()}"
                            
                            pdf_buffer = generate_certificate_pdf_with_template(r, exam, user)
                            pdf_data = pdf_buffer.getvalue()
                            
                            # Subir a blob
                            blob_filename = f"certificate_{r.id}.pdf"
                            blob_url = upload_pdf_to_blob(pdf_buffer, blob_filename)
                            if blob_url:
                                r.certificate_url = blob_url
                                db.session.commit()
                            
                            generated += 1
                        except Exception as e:
                            errors.append({'user': user.full_name, 'type': 'tier_standard', 'error': str(e)})
                            continue
                    
                    if pdf_data:
                        zip_file.writestr(filename, pdf_data)
                        files_added += 1
            
            # Procesar tier_advanced (CONOCER)
            if 'tier_advanced' in certificate_types:
                try:
                    from app.services.conocer_blob_service import ConocerBlobService
                    blob_service = ConocerBlobService()
                    
                    for cert in conocer_certs:
                        user = users_map.get(cert.user_id)
                        if not user:
                            continue
                        
                        safe_name = ''.join(c for c in user.full_name if c.isalnum() or c in ' -_').strip().replace(' ', '_')
                        folder = f"Certificados_CONOCER/{safe_name}"
                        filename = f"{folder}/{cert.standard_code}_{cert.certificate_number}.pdf"
                        
                        try:
                            blob_data = blob_service.download_certificate(cert.blob_name)
                            if blob_data:
                                zip_file.writestr(filename, blob_data)
                                files_added += 1
                        except Exception as e:
                            errors.append({'user': user.full_name, 'type': 'tier_advanced', 'error': str(e)})
                except ImportError:
                    pass
        
        current_app.logger.info(f"ZIP creado: {files_added} archivos, {generated} generados nuevos")
        
        if files_added == 0:
            return jsonify({
                'error': 'No hay certificados disponibles para descargar',
                'details': errors
            }), 400
        
        # Preparar respuesta
        zip_buffer.seek(0)
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        zip_filename = f"Certificados_{group.name}_{timestamp}.zip"
        
        from flask import Response
        return Response(
            zip_buffer.getvalue(),
            mimetype='application/zip',
            headers={
                'Content-Disposition': f'attachment; filename="{zip_filename}"',
                'Content-Length': len(zip_buffer.getvalue())
            }
        )
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@bp.route('/groups/<int:group_id>/certificates/generate', methods=['POST'])
@jwt_required()
@coordinator_required
def generate_group_certificates(group_id):
    """
    Generar certificados pendientes para el grupo.
    Este endpoint inicia la generación de PDFs para los certificados que no los tienen.
    Solo admin/developer pueden ejecutar esta acción.
    
    Body:
    - certificate_type: 'tier_basic' o 'tier_standard'
    - user_ids: (opcional) Lista de user_ids específicos
    """
    try:
        # Solo admin/developer pueden generar certificados
        user_id = get_jwt_identity()
        user = User.query.get(user_id)
        if not user or user.role not in ['admin', 'developer']:
            return jsonify({'error': 'Acceso denegado. Solo administradores pueden generar certificados.'}), 403

        from app.models.result import Result
        from app.models.exam import Exam
        from app.utils.queue_utils import enqueue_pdf_generation
        from sqlalchemy import and_, or_
        from flask import current_app
        
        group = CandidateGroup.query.get_or_404(group_id)
        
        data = request.get_json() or {}
        certificate_type = data.get('certificate_type', 'tier_basic')
        user_ids = data.get('user_ids', None)
        
        if certificate_type not in ['tier_basic', 'tier_standard']:
            return jsonify({'error': 'Tipo de certificado inválido. Use tier_basic o tier_standard'}), 400
        
        # Obtener miembros
        if user_ids:
            members = GroupMember.query.filter(
                and_(
                    GroupMember.group_id == group_id,
                    GroupMember.user_id.in_(user_ids),
                    GroupMember.status == 'active'
                )
            ).all()
        else:
            members = GroupMember.query.filter_by(group_id=group_id, status='active').all()
        
        member_user_ids = [m.user_id for m in members]
        
        # Obtener resultados aprobados sin el PDF correspondiente
        if certificate_type == 'tier_basic':
            # Reportes de evaluación - sin report_url
            results = Result.query.filter(
                and_(
                    Result.user_id.in_(member_user_ids),
                    Result.status == 1,
                    Result.result == 1,
                    or_(Result.report_url == None, Result.report_url == '')
                )
            ).all()
        else:
            # Certificados Eduit - sin certificate_url
            results = Result.query.filter(
                and_(
                    Result.user_id.in_(member_user_ids),
                    Result.status == 1,
                    Result.result == 1,
                    or_(Result.certificate_url == None, Result.certificate_url == '')
                )
            ).all()
        
        queued = []
        for r in results:
            try:
                # Encolar generación de PDF
                pdf_type = 'evaluation_report' if certificate_type == 'tier_basic' else 'certificate'
                enqueue_pdf_generation(r.id, pdf_type)
                queued.append({
                    'result_id': r.id,
                    'user_id': r.user_id,
                    'exam_id': r.exam_id
                })
            except Exception as e:
                current_app.logger.error(f"Error encolando PDF para result {r.id}: {e}")
        
        return jsonify({
            'message': f'{len(queued)} certificados encolados para generación',
            'queued_count': len(queued),
            'certificate_type': certificate_type,
            'details': queued
        })
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@bp.route('/groups/<int:group_id>/certificates/clear', methods=['POST'])
@jwt_required()
@coordinator_required
def clear_group_certificates_urls(group_id):
    """
    Limpiar URLs de certificados del grupo para forzar regeneración.
    Esto permite regenerar los PDFs con las posiciones correctas.
    Solo admin/developer pueden ejecutar esta acción.
    """
    try:
        # Solo admin/developer pueden limpiar/regenerar certificados
        user_id = get_jwt_identity()
        user = User.query.get(user_id)
        if not user or user.role not in ['admin', 'developer']:
            return jsonify({'error': 'Acceso denegado. Solo administradores pueden regenerar certificados.'}), 403

        from app.models.result import Result
        from sqlalchemy import and_
        
        group = CandidateGroup.query.get_or_404(group_id)
        
        data = request.get_json() or {}
        clear_reports = data.get('clear_reports', True)  # tier_basic
        clear_certificates = data.get('clear_certificates', True)  # tier_standard
        
        # Obtener miembros del grupo
        members = GroupMember.query.filter_by(group_id=group_id, status='active').all()
        member_user_ids = [m.user_id for m in members]
        
        if not member_user_ids:
            return jsonify({'error': 'No hay miembros en el grupo'}), 400
        
        # Obtener resultados aprobados
        results = Result.query.filter(
            and_(
                Result.user_id.in_(member_user_ids),
                Result.status == 1,
                Result.result == 1
            )
        ).all()
        
        cleared_count = 0
        for r in results:
            changed = False
            if clear_reports and r.report_url:
                r.report_url = None
                changed = True
            if clear_certificates and r.certificate_url:
                r.certificate_url = None
                changed = True
            if changed:
                cleared_count += 1
        
        db.session.commit()
        
        return jsonify({
            'message': f'{cleared_count} certificados limpiados para regeneración',
            'cleared_count': cleared_count,
            'group_id': group_id,
            'group_name': group.name
        })
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@bp.route('/groups/<int:group_id>/analytics', methods=['GET'])
@jwt_required()
@coordinator_required
def get_group_analytics(group_id):
    """
    Dashboard analítico completo del grupo.
    Agrega datos de miembros, exámenes, resultados, certificados, materiales y ECAs.
    
    Query params:
    - exam_id (int, optional): Filtrar resultados por examen específico
    - date_from (str, optional): Filtrar desde fecha (YYYY-MM-DD)
    - date_to (str, optional): Filtrar hasta fecha (YYYY-MM-DD)
    - certification_status (str, optional): certified|in_progress|failed|pending
    """
    try:
        from app.models.result import Result
        from app.models.conocer_certificate import ConocerCertificate
        from app.models.exam import Exam
        from app.models.partner import EcmCandidateAssignment, GroupStudyMaterial, GroupStudyMaterialMember
        from sqlalchemy import and_, func, case
        from datetime import datetime, timedelta
        from collections import defaultdict

        group = CandidateGroup.query.get_or_404(group_id)

        # Parámetros de filtro
        exam_id_filter = request.args.get('exam_id', type=int)
        date_from = request.args.get('date_from')
        date_to = request.args.get('date_to')

        # ── 1. MIEMBROS ──
        members = GroupMember.query.filter_by(group_id=group_id, status='active').all()
        member_user_ids = [m.user_id for m in members]
        total_members = len(members)

        if total_members == 0:
            return jsonify({
                'group': {'id': group.id, 'name': group.name},
                'members': {'total': 0, 'certified': 0, 'in_progress': 0, 'failed': 0, 'pending': 0, 'with_email': 0, 'with_curp': 0},
                'exams': {'assigned': 0, 'details': []},
                'results': {'total': 0, 'completed': 0, 'approved': 0, 'failed': 0, 'in_progress': 0, 'pass_rate': 0, 'avg_score': 0, 'avg_duration_minutes': 0, 'score_distribution': [], 'by_exam': [], 'by_date': []},
                'certificates': {'tier_basic': {'ready': 0, 'pending': 0}, 'tier_standard': {'ready': 0, 'pending': 0}, 'tier_advanced': 0, 'digital_badge': 0},
                'materials': {'assigned': 0, 'details': []},
                'ecm': {'total_assignments': 0, 'unique_ecms': 0, 'details': []},
            })

        # Conteos de miembros por elegibilidad
        from app.models.user import User
        users = User.query.filter(User.id.in_(member_user_ids)).all()
        users_map = {u.id: u for u in users}
        with_email = sum(1 for u in users if u.email)
        with_curp = sum(1 for u in users if u.curp)

        # ── 2. EXÁMENES ASIGNADOS ──
        group_exams = GroupExam.query.filter_by(group_id=group_id, is_active=True).all()
        exam_ids = [ge.exam_id for ge in group_exams]
        exams_map = {}
        if exam_ids:
            exams = Exam.query.filter(Exam.id.in_(exam_ids)).all()
            exams_map = {e.id: e for e in exams}

        exams_detail = []
        for ge in group_exams:
            exam = exams_map.get(ge.exam_id)
            exams_detail.append({
                'exam_id': ge.exam_id,
                'exam_name': exam.name if exam else f'Examen #{ge.exam_id}',
                'assigned_at': ge.assigned_at.isoformat() if ge.assigned_at else None,
                'passing_score': ge.passing_score or 70,
                'max_attempts': ge.max_attempts or 1,
                'time_limit_minutes': ge.time_limit_minutes,
                'assignment_type': ge.assignment_type,
            })

        # ── 3. RESULTADOS ──
        results_query = Result.query.filter(
            Result.user_id.in_(member_user_ids)
        )
        # Filtrar solo exámenes de este grupo si hay group_id en Result
        if exam_id_filter:
            results_query = results_query.filter(Result.exam_id == exam_id_filter)
        if date_from:
            try:
                dt_from = datetime.strptime(date_from, '%Y-%m-%d')
                results_query = results_query.filter(Result.end_date >= dt_from)
            except ValueError:
                pass
        if date_to:
            try:
                dt_to = datetime.strptime(date_to, '%Y-%m-%d') + timedelta(days=1)
                results_query = results_query.filter(Result.end_date < dt_to)
            except ValueError:
                pass

        all_results = results_query.all()

        # Filtrar por exámenes del grupo
        group_exam_ids_set = set(exam_ids)
        results = [r for r in all_results if r.exam_id in group_exam_ids_set]

        completed_results = [r for r in results if r.status == 1]
        approved = [r for r in completed_results if r.result == 1]
        failed_results = [r for r in completed_results if r.result == 0]
        in_progress = [r for r in results if r.status == 0]

        scores = [r.score for r in completed_results if r.score is not None]
        avg_score = round(sum(scores) / len(scores), 1) if scores else 0
        pass_rate = round((len(approved) / len(completed_results)) * 100, 1) if completed_results else 0

        durations = [r.duration_seconds for r in completed_results if r.duration_seconds and r.duration_seconds > 0]
        avg_duration_min = round((sum(durations) / len(durations)) / 60, 1) if durations else 0

        # Score distribution (bins: 0-10, 11-20, ..., 91-100)
        score_bins = [0] * 10
        for s in scores:
            idx = min(s // 10, 9)  # 100 → bin 9
            score_bins[idx] += 1
        score_distribution = [
            {'range': f'{i*10}-{i*10+9}' if i < 9 else '90-100', 'count': score_bins[i]}
            for i in range(10)
        ]

        # Results by exam
        by_exam = defaultdict(lambda: {'approved': 0, 'failed': 0, 'in_progress': 0, 'total_score': 0, 'count': 0})
        for r in results:
            key = r.exam_id
            if r.status == 1 and r.result == 1:
                by_exam[key]['approved'] += 1
            elif r.status == 1 and r.result == 0:
                by_exam[key]['failed'] += 1
            elif r.status == 0:
                by_exam[key]['in_progress'] += 1
            if r.status == 1 and r.score is not None:
                by_exam[key]['total_score'] += r.score
                by_exam[key]['count'] += 1

        results_by_exam = []
        for eid, data in by_exam.items():
            exam = exams_map.get(eid)
            results_by_exam.append({
                'exam_id': eid,
                'exam_name': exam.name if exam else f'Examen #{eid}',
                'approved': data['approved'],
                'failed': data['failed'],
                'in_progress': data['in_progress'],
                'avg_score': round(data['total_score'] / data['count'], 1) if data['count'] > 0 else 0,
                'pass_rate': round((data['approved'] / (data['approved'] + data['failed'])) * 100, 1) if (data['approved'] + data['failed']) > 0 else 0,
            })

        # Results by date (últimos 30 días o rango)
        by_date = defaultdict(lambda: {'approved': 0, 'failed': 0, 'total': 0})
        for r in completed_results:
            if r.end_date:
                day_key = r.end_date.strftime('%Y-%m-%d')
                by_date[day_key]['total'] += 1
                if r.result == 1:
                    by_date[day_key]['approved'] += 1
                else:
                    by_date[day_key]['failed'] += 1

        results_by_date = sorted([
            {'date': k, 'approved': v['approved'], 'failed': v['failed'], 'total': v['total']}
            for k, v in by_date.items()
        ], key=lambda x: x['date'])

        # Certification status per member
        cert_status = {'certified': 0, 'in_progress': 0, 'failed': 0, 'pending': 0}
        for uid in member_user_ids:
            user_results = [r for r in results if r.user_id == uid and r.status == 1]
            user_approved = [r for r in user_results if r.result == 1]
            if user_approved:
                cert_status['certified'] += 1
            elif user_results:
                # Has completed but none approved
                cert_status['failed'] += 1
            elif any(r.user_id == uid and r.status == 0 for r in results):
                cert_status['in_progress'] += 1
            else:
                cert_status['pending'] += 1

        # ── 4. CERTIFICADOS ──
        tier_basic_ready = sum(1 for r in approved if r.report_url or r.certificate_code)
        tier_basic_pending = sum(1 for r in approved if not r.report_url and not r.certificate_code)
        tier_standard_ready = sum(1 for r in approved if r.certificate_url or r.eduit_certificate_code)
        tier_standard_pending = sum(1 for r in approved if not r.certificate_url and not r.eduit_certificate_code)

        conocer_certs = ConocerCertificate.query.filter(
            and_(
                ConocerCertificate.user_id.in_(member_user_ids),
                ConocerCertificate.status == 'active'
            )
        ).all()

        # ── 5. MATERIALES DE ESTUDIO ──
        # Materiales directos del grupo
        group_materials = GroupStudyMaterial.query.filter_by(group_id=group_id).all()
        materials_detail = []
        for gm in group_materials:
            material = gm.study_material
            assigned_members_count = GroupStudyMaterialMember.query.filter_by(
                group_study_material_id=gm.id
            ).count() if gm.assignment_type == 'selected' else total_members
            materials_detail.append({
                'id': gm.id,
                'material_name': material.title if material else f'Material #{gm.study_material_id}',
                'assigned_at': gm.assigned_at.isoformat() if gm.assigned_at else None,
                'assigned_members': assigned_members_count,
                'source': 'direct',
            })
        
        # Materiales vinculados a exámenes del grupo (via GroupExamMaterial)
        from app.models.partner import GroupExamMaterial
        from app.models.study_content import StudyMaterial
        seen_material_ids = set(gm.study_material_id for gm in group_materials)
        for ge in group_exams:
            # Custom materials del group_exam
            custom_mats = GroupExamMaterial.query.filter_by(group_exam_id=ge.id, is_included=True).all()
            if custom_mats:
                mat_ids = [cm.study_material_id for cm in custom_mats if cm.study_material_id not in seen_material_ids]
                if mat_ids:
                    study_mats = StudyMaterial.query.filter(StudyMaterial.id.in_(mat_ids)).all()
                    for sm in study_mats:
                        seen_material_ids.add(sm.id)
                        materials_detail.append({
                            'id': f'exam-{ge.exam_id}-{sm.id}',
                            'material_name': sm.title,
                            'assigned_at': ge.assigned_at.isoformat() if ge.assigned_at else None,
                            'assigned_members': total_members,
                            'source': 'exam',
                            'exam_name': exams_map.get(ge.exam_id, None) and exams_map[ge.exam_id].name or f'Examen #{ge.exam_id}',
                        })
            else:
                # Sin personalizaciones: buscar materiales vinculados al examen
                try:
                    from sqlalchemy import text as sa_text
                    linked = db.session.execute(sa_text('''
                        SELECT sc.id, sc.title
                        FROM study_contents sc
                        INNER JOIN study_material_exams sme ON sc.id = sme.study_material_id
                        WHERE sme.exam_id = :exam_id AND sc.is_published = 1
                    '''), {'exam_id': ge.exam_id}).fetchall()
                    for row in linked:
                        if row[0] not in seen_material_ids:
                            seen_material_ids.add(row[0])
                            materials_detail.append({
                                'id': f'exam-{ge.exam_id}-{row[0]}',
                                'material_name': row[1],
                                'assigned_at': ge.assigned_at.isoformat() if ge.assigned_at else None,
                                'assigned_members': total_members,
                                'source': 'exam',
                                'exam_name': exams_map.get(ge.exam_id, None) and exams_map[ge.exam_id].name or f'Examen #{ge.exam_id}',
                            })
                except Exception:
                    pass

        # ── 6. ECAs ──
        # Buscar ECAs por group_id O por combinación user_id + exam_ids del grupo
        from sqlalchemy import or_
        eca_filters = [EcmCandidateAssignment.user_id.in_(member_user_ids)]
        if exam_ids:
            ecas = EcmCandidateAssignment.query.filter(
                and_(
                    EcmCandidateAssignment.user_id.in_(member_user_ids),
                    or_(
                        EcmCandidateAssignment.group_id == group_id,
                        EcmCandidateAssignment.exam_id.in_(exam_ids)
                    )
                )
            ).all()
        else:
            ecas = EcmCandidateAssignment.query.filter(
                and_(
                    EcmCandidateAssignment.user_id.in_(member_user_ids),
                    EcmCandidateAssignment.group_id == group_id
                )
            ).all()

        ecm_summary = defaultdict(lambda: {'count': 0, 'ecm_name': '', 'ecm_code': '', 'logo_url': None})
        for eca in ecas:
            key = eca.competency_standard_id
            ecm_summary[key]['count'] += 1
            if eca.competency_standard:
                ecm_summary[key]['ecm_name'] = eca.competency_standard.name
                ecm_summary[key]['ecm_code'] = eca.competency_standard.code
                ecm_summary[key]['logo_url'] = eca.competency_standard.logo_url

        ecm_details = [
            {'ecm_id': k, 'ecm_name': v['ecm_name'], 'ecm_code': v['ecm_code'], 'assignments': v['count'], 'logo_url': v['logo_url']}
            for k, v in ecm_summary.items()
        ]

        # ── 7. TOP PERFORMERS ──
        user_best_scores = defaultdict(lambda: {'score': 0, 'count': 0, 'total_score': 0})
        for r in completed_results:
            user_best_scores[r.user_id]['count'] += 1
            user_best_scores[r.user_id]['total_score'] += (r.score or 0)
            if (r.score or 0) > user_best_scores[r.user_id]['score']:
                user_best_scores[r.user_id]['score'] = r.score or 0

        top_performers = sorted([
            {
                'user_id': uid,
                'full_name': users_map[uid].full_name if uid in users_map else 'Desconocido',
                'best_score': data['score'],
                'avg_score': round(data['total_score'] / data['count'], 1) if data['count'] > 0 else 0,
                'exams_completed': data['count'],
            }
            for uid, data in user_best_scores.items()
        ], key=lambda x: x['avg_score'], reverse=True)[:10]

        return jsonify({
            'group': {
                'id': group.id,
                'name': group.name,
                'campus_name': group.campus.name if group.campus else None,
                'partner_name': group.campus.partner.name if group.campus and group.campus.partner else None,
            },
            'members': {
                'total': total_members,
                'certified': cert_status['certified'],
                'in_progress': cert_status['in_progress'],
                'failed': cert_status['failed'],
                'pending': cert_status['pending'],
                'with_email': with_email,
                'with_curp': with_curp,
            },
            'exams': {
                'assigned': len(group_exams),
                'details': exams_detail,
            },
            'results': {
                'total': len(results),
                'completed': len(completed_results),
                'approved': len(approved),
                'failed': len(failed_results),
                'in_progress': len(in_progress),
                'pass_rate': pass_rate,
                'avg_score': avg_score,
                'avg_duration_minutes': avg_duration_min,
                'score_distribution': score_distribution,
                'by_exam': results_by_exam,
                'by_date': results_by_date,
            },
            'certificates': {
                'tier_basic': {'ready': tier_basic_ready, 'pending': tier_basic_pending},
                'tier_standard': {'ready': tier_standard_ready, 'pending': tier_standard_pending},
                'tier_advanced': len(conocer_certs),
                'digital_badge': len(approved),
            },
            'materials': {
                'assigned': len(materials_detail),
                'details': materials_detail,
            },
            'ecm': {
                'total_assignments': len(ecas),
                'unique_ecms': len(ecm_summary),
                'details': ecm_details,
            },
            'top_performers': top_performers,
        })

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@bp.route('/groups/<int:group_id>/candidates/<string:user_id>/certification-detail', methods=['GET'])
@jwt_required()
@coordinator_required
def get_candidate_certification_detail(group_id, user_id):
    """
    Detalle de certificación de un candidato dentro de un grupo.
    Devuelve todos los exámenes aprobados con nombre, ECAs asociadas, y certificados CONOCER.
    """
    try:
        from app.models.result import Result
        from app.models.conocer_certificate import ConocerCertificate
        from app.models.exam import Exam
        from app.models.partner import EcmCandidateAssignment
        from sqlalchemy import and_

        group = CandidateGroup.query.get_or_404(group_id)

        # Verificar que el usuario es miembro activo del grupo
        member = GroupMember.query.filter_by(group_id=group_id, user_id=user_id, status='active').first()
        if not member:
            return jsonify({'error': 'El candidato no es miembro activo del grupo'}), 404

        user = member.user
        if not user:
            return jsonify({'error': 'Usuario no encontrado'}), 404

        # Resultados aprobados del candidato
        results = Result.query.filter(
            and_(
                Result.user_id == user_id,
                Result.status == 1,
                Result.result == 1
            )
        ).order_by(Result.end_date.desc()).all()

        # Obtener IDs de exámenes para buscar nombres
        exam_ids = list(set(r.exam_id for r in results))
        exams_map = {}
        if exam_ids:
            exams = Exam.query.filter(Exam.id.in_(exam_ids)).all()
            exams_map = {e.id: e for e in exams}

        # ECAs del candidato en este grupo
        ecas = EcmCandidateAssignment.query.filter_by(user_id=user_id, group_id=group_id).all()
        # También obtener ECAs globales (sin filtro de grupo) para cross-reference
        all_ecas = EcmCandidateAssignment.query.filter_by(user_id=user_id).all()
        ecas_by_exam = {}
        for eca in all_ecas:
            ecas_by_exam.setdefault(eca.exam_id, []).append(eca)

        # Certificados CONOCER del candidato
        conocer_certs = ConocerCertificate.query.filter(
            and_(
                ConocerCertificate.user_id == user_id,
                ConocerCertificate.status == 'active'
            )
        ).all()

        # Construir detalle de resultados con examen y ECA info
        results_detail = []
        for r in results:
            exam = exams_map.get(r.exam_id)
            related_ecas = ecas_by_exam.get(r.exam_id, [])

            result_item = {
                'result_id': r.id,
                'exam_id': r.exam_id,
                'exam_name': exam.name if exam else f'Examen #{r.exam_id}',
                'exam_code': getattr(exam, 'code', None) if exam else None,
                'score': r.score,
                'end_date': r.end_date.isoformat() if r.end_date else None,
                'has_report': bool(r.report_url),
                'has_certificate': bool(r.certificate_url),
                'certificate_code': r.certificate_code,
                'eduit_certificate_code': r.eduit_certificate_code,
                'ecm_assignments': [{
                    'id': eca.id,
                    'assignment_number': eca.assignment_number,
                    'ecm_code': eca.competency_standard.code if eca.competency_standard else None,
                    'ecm_name': eca.competency_standard.name if eca.competency_standard else None,
                    'group_name': eca.group_name,
                    'assigned_at': eca.assigned_at.isoformat() if eca.assigned_at else None,
                } for eca in related_ecas]
            }
            results_detail.append(result_item)

        # Certificados CONOCER
        conocer_detail = [{
            'id': c.id,
            'certificate_number': c.certificate_number,
            'standard_code': c.standard_code,
            'standard_name': c.standard_name,
            'issue_date': c.issue_date.isoformat() if c.issue_date else None,
        } for c in conocer_certs]

        return jsonify({
            'candidate': {
                'user_id': user.id,
                'full_name': user.full_name,
                'email': user.email,
                'curp': user.curp,
                'username': user.username,
            },
            'group': {
                'id': group.id,
                'name': group.name,
            },
            'summary': {
                'exams_approved': len(results),
                'ecm_count': len(set(eca.competency_standard_id for eca in all_ecas)),
                'conocer_count': len(conocer_certs),
                'reports_ready': sum(1 for r in results if r.report_url or r.certificate_code),
                'certificates_ready': sum(1 for r in results if r.certificate_url or r.eduit_certificate_code),
            },
            'results': results_detail,
            'conocer_certificates': conocer_detail,
        })

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


# ============== VIGENCIA DE ASIGNACIONES ==============

@bp.route('/assignments/<int:assignment_id>/extend-validity', methods=['POST'])
@jwt_required()
def extend_assignment_validity(assignment_id):
    """Extender la vigencia de una asignación (GroupExam) y sus ECM asociados.
    
    Solo admin y coordinator pueden extender.
    Coordinator solo puede extender asignaciones de sus grupos.
    
    Request body:
    {
        "months": 3  // Meses adicionales a agregar
    }
    """
    try:
        user_id = get_jwt_identity()
        user = User.query.get(user_id)
        if not user or user.role not in ['admin', 'developer', 'coordinator']:
            return jsonify({'error': 'Acceso denegado'}), 403
        
        data = request.get_json()
        months = data.get('months')
        
        if not months or not isinstance(months, (int, float)) or months <= 0:
            return jsonify({'error': 'Debe especificar una cantidad de meses válida (mayor a 0)'}), 400
        
        months = int(months)
        
        group_exam = GroupExam.query.get(assignment_id)
        if not group_exam:
            return jsonify({'error': 'Asignación no encontrada'}), 404
        
        # Verificar permisos de coordinador
        if user.role == 'coordinator':
            group = CandidateGroup.query.get(group_exam.group_id)
            if not group:
                return jsonify({'error': 'Grupo no encontrado'}), 404
            # Verificar acceso: coordinador debe tener relación con el partner del campus
            has_access = False
            if group.campus_id:
                campus = Campus.query.get(group.campus_id)
                if campus and campus.partner_id:
                    from app.models.partner import user_partners
                    up = db.session.query(user_partners).filter_by(
                        user_id=user_id, partner_id=campus.partner_id
                    ).first()
                    if up:
                        has_access = True
            # Fallback: verificar si el coordinador gestiona candidatos de este grupo
            if not has_access:
                from app.models.partner import GroupMember
                coordinator_members = db.session.query(GroupMember).join(
                    User, User.id == GroupMember.user_id
                ).filter(
                    GroupMember.group_id == group_exam.group_id,
                    User.coordinator_id == user_id
                ).first()
                if coordinator_members:
                    has_access = True
            if not has_access:
                return jsonify({'error': 'No tienes permiso para extender esta asignación'}), 403
        
        # Extender GroupExam
        old_extended = group_exam.extended_months or 0
        group_exam.extended_months = old_extended + months
        
        # Extender las EcmCandidateAssignments asociadas
        from app.models.partner import EcmCandidateAssignment
        ecm_assignments = EcmCandidateAssignment.query.filter_by(
            group_exam_id=assignment_id
        ).all()
        
        for ecm_a in ecm_assignments:
            ecm_old = ecm_a.extended_months or 0
            ecm_a.extended_months = ecm_old + months
        
        db.session.commit()
        
        new_expires = group_exam.effective_expires_at
        
        return jsonify({
            'message': f'Vigencia extendida {months} mes(es) exitosamente',
            'assignment_id': assignment_id,
            'extended_months_total': group_exam.extended_months,
            'validity_months': group_exam.validity_months,
            'expires_at': new_expires.isoformat() if new_expires else None,
            'is_expired': group_exam.is_expired,
            'ecm_assignments_updated': len(ecm_assignments),
        })
    
    except Exception as e:
        db.session.rollback()
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@bp.route('/ecm-assignments/<int:ecm_assignment_id>/extend-validity', methods=['POST'])
@jwt_required()
def extend_ecm_assignment_validity(ecm_assignment_id):
    """Extender la vigencia de una asignación ECM individual.
    
    Request body:
    {
        "months": 3  // Meses adicionales a agregar
    }
    """
    try:
        user_id = get_jwt_identity()
        user = User.query.get(user_id)
        if not user or user.role not in ['admin', 'developer', 'coordinator']:
            return jsonify({'error': 'Acceso denegado'}), 403
        
        data = request.get_json()
        months = data.get('months')
        
        if not months or not isinstance(months, (int, float)) or months <= 0:
            return jsonify({'error': 'Debe especificar una cantidad de meses válida (mayor a 0)'}), 400
        
        months = int(months)
        
        from app.models.partner import EcmCandidateAssignment
        ecm_assignment = EcmCandidateAssignment.query.get(ecm_assignment_id)
        if not ecm_assignment:
            return jsonify({'error': 'Asignación ECM no encontrada'}), 404
        
        # Verificar permisos de coordinador
        if user.role == 'coordinator':
            has_access = False
            # Verificar si el candidato está bajo este coordinador
            target_user = User.query.get(ecm_assignment.user_id)
            if target_user and target_user.coordinator_id == user_id:
                has_access = True
            # Verificar acceso via user_partners y grupo/campus
            if not has_access and ecm_assignment.group_exam_id:
                ge = GroupExam.query.get(ecm_assignment.group_exam_id)
                if ge:
                    grp = CandidateGroup.query.get(ge.group_id)
                    if grp and grp.campus_id:
                        campus = Campus.query.get(grp.campus_id)
                        if campus and campus.partner_id:
                            from app.models.partner import user_partners
                            up = db.session.query(user_partners).filter_by(
                                user_id=user_id, partner_id=campus.partner_id
                            ).first()
                            if up:
                                has_access = True
            if not has_access:
                return jsonify({'error': 'No tienes permiso para extender esta asignación'}), 403
        
        old_extended = ecm_assignment.extended_months or 0
        ecm_assignment.extended_months = old_extended + months
        
        db.session.commit()
        
        new_expires = ecm_assignment.effective_expires_at
        
        return jsonify({
            'message': f'Vigencia extendida {months} mes(es) exitosamente',
            'ecm_assignment_id': ecm_assignment_id,
            'assignment_number': ecm_assignment.assignment_number,
            'extended_months_total': ecm_assignment.extended_months,
            'validity_months': ecm_assignment.validity_months,
            'expires_at': new_expires.isoformat() if new_expires else None,
            'is_expired': ecm_assignment.is_expired,
        })
    
    except Exception as e:
        db.session.rollback()
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


# ============== MÓDULO DE ASIGNACIONES POR ECM ==============

@bp.route('/ecm-assignments', methods=['GET'])
@jwt_required()
def get_ecm_assignments():
    """Listado de ECMs con resumen de asignaciones (optimizado).
    
    Usa queries SQL agregadas con subqueries en vez de loops N+1.
    Solo accesible por admin y coordinator.
    
    Query params:
    - search (str): Buscar en código o nombre del ECM
    - active_only (bool): Solo ECMs activos (default true)
    """
    try:
        from sqlalchemy import text

        user_id = get_jwt_identity()
        user = User.query.get(user_id)
        if not user or user.role not in ['admin', 'developer', 'coordinator']:
            return jsonify({'error': 'Acceso denegado'}), 403

        search = request.args.get('search', '').strip()
        active_only = request.args.get('active_only', 'true').lower() == 'true'

        # Build filters
        params = {}
        where_parts = []
        if active_only:
            where_parts.append("AND cs.is_active = 1")
        if search:
            where_parts.append("AND (LOWER(cs.code) LIKE :search_term OR LOWER(cs.name) LIKE :search_term)")
            params['search_term'] = f'%{search.lower()}%'

        where_sql = '\n            '.join(where_parts)

        sql = text(f"""
            SELECT
                cs.id, cs.code, cs.name, cs.sector, cs.[level],
                cs.certifying_body, cs.is_active, cs.logo_url,
                COALESCE(ge_stats.total_assignments, 0) AS total_assignments,
                COALESCE(ge_stats.exams_count, 0) AS exams_count,
                COALESCE(user_stats.total_candidates, 0) AS total_candidates,
                CAST(COALESCE(cost_stats.total_cost, 0) AS FLOAT) AS total_cost,
                result_stats.avg_score,
                result_stats.pass_rate
            FROM competency_standards cs
            LEFT JOIN (
                SELECT e.competency_standard_id AS ecm_id,
                       COUNT(DISTINCT ge.id) AS total_assignments,
                       COUNT(DISTINCT e.id) AS exams_count
                FROM exams e
                JOIN group_exams ge ON ge.exam_id = e.id
                GROUP BY e.competency_standard_id
            ) ge_stats ON ge_stats.ecm_id = cs.id
            LEFT JOIN (
                SELECT ecm_id, COUNT(DISTINCT user_id) AS total_candidates
                FROM (
                    SELECT e.competency_standard_id AS ecm_id, gm.user_id
                    FROM group_exams ge
                    JOIN exams e ON e.id = ge.exam_id
                    JOIN group_members gm ON gm.group_id = ge.group_id
                    WHERE ge.assignment_type = 'all'
                    UNION
                    SELECT e.competency_standard_id, gem.user_id
                    FROM group_exams ge
                    JOIN exams e ON e.id = ge.exam_id
                    JOIN group_exam_members gem ON gem.group_exam_id = ge.id
                    WHERE ge.assignment_type != 'all'
                ) AS all_users
                GROUP BY ecm_id
            ) user_stats ON user_stats.ecm_id = cs.id
            LEFT JOIN (
                SELECT e.competency_standard_id AS ecm_id,
                       SUM(bt.amount) AS total_cost
                FROM balance_transactions bt
                JOIN group_exams ge ON ge.id = bt.reference_id
                JOIN exams e ON e.id = ge.exam_id
                WHERE bt.reference_type = 'group_exam'
                  AND bt.concept IN ('asignacion_certificacion', 'asignacion_retoma')
                GROUP BY e.competency_standard_id
            ) cost_stats ON cost_stats.ecm_id = cs.id
            LEFT JOIN (
                SELECT r.competency_standard_id AS ecm_id,
                       ROUND(AVG(CAST(r.score AS FLOAT)), 1) AS avg_score,
                       ROUND(100.0 * SUM(CASE WHEN r.result = 1 THEN 1 ELSE 0 END)
                             / NULLIF(COUNT(*), 0), 1) AS pass_rate
                FROM results r
                WHERE r.competency_standard_id IS NOT NULL AND r.status = 1
                GROUP BY r.competency_standard_id
            ) result_stats ON result_stats.ecm_id = cs.id
            WHERE 1=1
            {where_sql}
            ORDER BY cs.code
        """)

        rows = db.session.execute(sql, params).fetchall()

        result = []
        for row in rows:
            result.append({
                'id': row.id,
                'code': row.code,
                'name': row.name,
                'sector': row.sector,
                'level': row.level,
                'certifying_body': row.certifying_body,
                'is_active': bool(row.is_active),
                'logo_url': row.logo_url,
                'total_assignments': row.total_assignments,
                'total_candidates': row.total_candidates,
                'total_cost': float(row.total_cost),
                'avg_score': float(row.avg_score) if row.avg_score is not None else None,
                'pass_rate': float(row.pass_rate) if row.pass_rate is not None else None,
                'exams_count': row.exams_count,
            })

        return jsonify({
            'ecms': result,
            'total': len(result),
        })

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@bp.route('/ecm-assignments/<int:ecm_id>', methods=['GET'])
@jwt_required()
def get_ecm_assignment_detail(ecm_id):
    """Detalle de asignaciones para un ECM (optimizado para cientos de miles de registros).
    
    Usa CTEs con JOINs, ROW_NUMBER para últimos resultados, paginación SQL
    OFFSET/FETCH NEXT. Progreso de material se calcula en batch solo para la
    página actual (~30 filas).
    
    Query params:
    - page (int): Página (default 1)
    - per_page (int): Por página (default 30)
    - search (str): Buscar por nombre o email
    - user_type (str): candidato, responsable, all
    - status (str): all, completed, pending, passed, failed
    - date_from (str): YYYY-MM-DD
    - date_to (str): YYYY-MM-DD
    - group_id (int): Filtrar por grupo
    - exam_id (int): Filtrar por examen
    - sort_by (str): date, name, score, cost
    - sort_dir (str): asc, desc
    """
    try:
        from app.models.exam import Exam
        from sqlalchemy import text

        user_id = get_jwt_identity()
        user = User.query.get(user_id)
        if not user or user.role not in ['admin', 'developer', 'coordinator']:
            return jsonify({'error': 'Acceso denegado'}), 403

        ecm = CompetencyStandard.query.get(ecm_id)
        if not ecm:
            return jsonify({'error': 'ECM no encontrado'}), 404

        page = request.args.get('page', 1, type=int)
        per_page = min(request.args.get('per_page', 30, type=int), 100)
        search = request.args.get('search', '').strip()
        user_type_filter = request.args.get('user_type', 'all')
        status_filter = request.args.get('status', 'all')
        date_from = request.args.get('date_from')
        date_to = request.args.get('date_to')
        group_id_filter = request.args.get('group_id', type=int)
        exam_id_filter = request.args.get('exam_id', type=int)
        sort_by = request.args.get('sort_by', 'date')
        sort_dir = request.args.get('sort_dir', 'desc')
        offset = (page - 1) * per_page

        # Check if ECM has exams
        exam_count = Exam.query.filter_by(competency_standard_id=ecm_id).count()
        if not exam_count:
            return jsonify({
                'ecm': {
                    'id': ecm.id, 'code': ecm.code, 'name': ecm.name,
                    'sector': ecm.sector, 'level': ecm.level,
                    'logo_url': ecm.logo_url, 'certifying_body': ecm.certifying_body,
                },
                'assignments': [], 'total': 0, 'pages': 0, 'current_page': 1,
                'per_page': per_page,
                'summary': {
                    'total_assignments': 0, 'total_candidates': 0, 'total_cost': 0,
                    'avg_score': None, 'pass_rate': None,
                    'completed_count': 0, 'pending_count': 0, 'passed_count': 0,
                },
                'filters': {'exams': [], 'groups': []}
            })

        # ── Dynamic WHERE clauses ──
        params = {'ecm_id': ecm_id, 'offset_val': offset, 'limit_val': per_page}
        where_parts = []

        if exam_id_filter:
            where_parts.append("AND a.exam_id = :exam_id_f")
            params['exam_id_f'] = exam_id_filter
        if group_id_filter:
            where_parts.append("AND a.group_id = :group_id_f")
            params['group_id_f'] = group_id_filter
        if user_type_filter != 'all':
            where_parts.append("AND a.user_role = :user_type_f")
            params['user_type_f'] = user_type_filter
        if search:
            where_parts.append(
                "AND (LOWER(a.user_name) LIKE :search_f OR LOWER(a.user_email) LIKE :search_f)"
            )
            params['search_f'] = f'%{search.lower()}%'
        if date_from:
            try:
                from_dt = datetime.strptime(date_from, '%Y-%m-%d')
                where_parts.append("AND a.assignment_date >= :date_from_f")
                params['date_from_f'] = from_dt
            except ValueError:
                pass
        if date_to:
            try:
                to_dt = datetime.strptime(date_to, '%Y-%m-%d').replace(
                    hour=23, minute=59, second=59
                )
                where_parts.append("AND a.assignment_date <= :date_to_f")
                params['date_to_f'] = to_dt
            except ValueError:
                pass

        # Status filters (on joined result)
        if status_filter == 'completed':
            where_parts.append("AND lr.result_status_raw = 1")
        elif status_filter == 'pending':
            where_parts.append("AND (lr.result_status_raw IS NULL OR lr.result_status_raw != 1)")
        elif status_filter == 'passed':
            where_parts.append("AND lr.result_status_raw = 1 AND lr.result_raw = 1")
        elif status_filter == 'failed':
            where_parts.append("AND lr.result_status_raw = 1 AND lr.result_raw = 0")

        where_sql = '\n                '.join(where_parts)

        # ── Sort mapping ──
        # Note: MSSQL doesn't allow ORDER BY on column aliases,
        # so computed columns must use the full expression.
        sort_map = {
            'date': 'a.assignment_date',
            'name': 'a.user_name',
            'score': 'COALESCE(lr.score, -1)',
            'cost': 'COALESCE(CAST(COALESCE(co.total_cost, 0) AS FLOAT) / NULLIF(mc.member_count, 0), 0)',
            'role': 'a.user_role',
            'group': 'a.group_name',
            'exam': 'a.exam_name',
            'status': 'COALESCE(lr.result_status_raw, -1)',
            'material': 'COALESCE(lr.result_status_raw, -1)',
            'duration': 'COALESCE(lr.duration_seconds, -1)',
        }
        sort_col = sort_map.get(sort_by, 'a.assignment_date')
        sort_direction = 'DESC' if sort_dir == 'desc' else 'ASC'

        # ── CTE definitions (shared between data + summary queries) ──
        cte_sql = """
            WITH base_assignments AS (
                SELECT
                    u.id AS user_id,
                    CONCAT(u.name, ' ', u.first_surname,
                           COALESCE(' ' + u.second_surname, '')) AS user_name,
                    u.email AS user_email,
                    u.role AS user_role,
                    u.curp AS user_curp,
                    ge.id AS group_exam_id,
                    ge.exam_id,
                    ge.assigned_at AS assignment_date,
                    ge.assignment_type,
                    ge.max_attempts,
                    ge.time_limit_minutes,
                    ge.passing_score,
                    cg.id AS group_id,
                    cg.name AS group_name,
                    cg.code AS group_code,
                    c.id AS campus_id,
                    c.name AS campus_name,
                    COALESCE(c.enable_tier_basic, 0) AS enable_tier_basic,
                    COALESCE(c.enable_tier_standard, 0) AS enable_tier_standard,
                    COALESCE(c.enable_tier_advanced, 0) AS enable_tier_advanced,
                    COALESCE(c.enable_digital_badge, 0) AS enable_digital_badge,
                    p.id AS partner_id,
                    p.name AS partner_name,
                    e.name AS exam_name
                FROM group_exams ge
                JOIN exams e ON e.id = ge.exam_id
                JOIN candidate_groups cg ON cg.id = ge.group_id
                LEFT JOIN campuses c ON c.id = cg.campus_id
                LEFT JOIN partners p ON p.id = c.partner_id
                JOIN group_members gm ON gm.group_id = ge.group_id
                JOIN users u ON u.id = gm.user_id
                WHERE e.competency_standard_id = :ecm_id
                  AND ge.assignment_type = 'all'

                UNION ALL

                SELECT
                    u.id,
                    CONCAT(u.name, ' ', u.first_surname,
                           COALESCE(' ' + u.second_surname, '')),
                    u.email, u.role, u.curp,
                    ge.id, ge.exam_id, ge.assigned_at,
                    ge.assignment_type,
                    ge.max_attempts, ge.time_limit_minutes, ge.passing_score,
                    cg.id, cg.name, cg.code,
                    c.id, c.name,
                    COALESCE(c.enable_tier_basic, 0),
                    COALESCE(c.enable_tier_standard, 0),
                    COALESCE(c.enable_tier_advanced, 0),
                    COALESCE(c.enable_digital_badge, 0),
                    p.id, p.name,
                    e.name
                FROM group_exams ge
                JOIN exams e ON e.id = ge.exam_id
                JOIN candidate_groups cg ON cg.id = ge.group_id
                LEFT JOIN campuses c ON c.id = cg.campus_id
                LEFT JOIN partners p ON p.id = c.partner_id
                JOIN group_exam_members gem ON gem.group_exam_id = ge.id
                JOIN users u ON u.id = gem.user_id
                WHERE e.competency_standard_id = :ecm_id
                  AND ge.assignment_type != 'all'
            ),
            latest_results AS (
                SELECT
                    r.user_id, r.exam_id, r.score,
                    r.status AS result_status_raw,
                    r.result AS result_raw,
                    r.end_date, r.duration_seconds, r.certificate_code,
                    ROW_NUMBER() OVER (
                        PARTITION BY r.user_id, r.exam_id
                        ORDER BY r.created_at DESC
                    ) AS rn
                FROM results r
                WHERE r.exam_id IN (
                    SELECT id FROM exams
                    WHERE competency_standard_id = :ecm_id
                )
            ),
            costs AS (
                SELECT
                    bt.reference_id AS group_exam_id,
                    SUM(bt.amount) AS total_cost
                FROM balance_transactions bt
                WHERE bt.reference_type = 'group_exam'
                  AND bt.concept IN ('asignacion_certificacion', 'asignacion_retoma')
                  AND bt.reference_id IN (
                      SELECT ge.id FROM group_exams ge
                      JOIN exams e ON e.id = ge.exam_id
                      WHERE e.competency_standard_id = :ecm_id
                  )
                GROUP BY bt.reference_id
            ),
            member_counts AS (
                SELECT ge.id AS group_exam_id, COUNT(gm.id) AS member_count
                FROM group_exams ge
                JOIN exams e ON e.id = ge.exam_id
                JOIN group_members gm ON gm.group_id = ge.group_id
                WHERE e.competency_standard_id = :ecm_id
                  AND ge.assignment_type = 'all'
                GROUP BY ge.id

                UNION ALL

                SELECT ge.id, COUNT(gem.id)
                FROM group_exams ge
                JOIN exams e ON e.id = ge.exam_id
                JOIN group_exam_members gem ON gem.group_exam_id = ge.id
                WHERE e.competency_standard_id = :ecm_id
                  AND ge.assignment_type != 'all'
                GROUP BY ge.id
            )
        """

        # ── Data query (paginated) ──
        data_sql = text(cte_sql + f"""
            SELECT
                a.user_id, a.user_name, a.user_email, a.user_role, a.user_curp,
                a.group_exam_id, a.exam_id, a.assignment_date, a.assignment_type,
                a.max_attempts, a.time_limit_minutes, a.passing_score,
                a.group_id, a.group_name, a.group_code,
                a.campus_id, a.campus_name,
                a.enable_tier_basic, a.enable_tier_standard,
                a.enable_tier_advanced, a.enable_digital_badge,
                a.partner_id, a.partner_name,
                a.exam_name,
                lr.score, lr.result_status_raw, lr.result_raw,
                lr.end_date AS result_date, lr.duration_seconds, lr.certificate_code,
                CAST(COALESCE(co.total_cost, 0) AS FLOAT)
                    / NULLIF(mc.member_count, 0) AS unit_cost
            FROM base_assignments a
            LEFT JOIN latest_results lr
                ON lr.user_id = a.user_id
               AND lr.exam_id = a.exam_id
               AND lr.rn = 1
            LEFT JOIN costs co
                ON co.group_exam_id = a.group_exam_id
            LEFT JOIN member_counts mc
                ON mc.group_exam_id = a.group_exam_id
            WHERE 1=1
                {where_sql}
            ORDER BY {sort_col} {sort_direction}
            OFFSET :offset_val ROWS FETCH NEXT :limit_val ROWS ONLY
        """)

        # ── Summary query (same CTEs, aggregated, no pagination) ──
        summary_sql = text(cte_sql + f"""
            SELECT
                COUNT(*) AS total,
                COUNT(DISTINCT a.user_id) AS unique_users,
                SUM(CAST(COALESCE(co.total_cost, 0) AS FLOAT)
                    / NULLIF(mc.member_count, 0)) AS total_cost,
                ROUND(AVG(CASE WHEN lr.result_status_raw = 1
                          THEN CAST(lr.score AS FLOAT) END), 1) AS avg_score,
                SUM(CASE WHEN lr.result_status_raw = 1
                    THEN 1 ELSE 0 END) AS completed_count,
                SUM(CASE WHEN lr.result_status_raw IS NULL
                          OR lr.result_status_raw != 1
                    THEN 1 ELSE 0 END) AS pending_count,
                SUM(CASE WHEN lr.result_status_raw = 1 AND lr.result_raw = 1
                    THEN 1 ELSE 0 END) AS passed_count
            FROM base_assignments a
            LEFT JOIN latest_results lr
                ON lr.user_id = a.user_id
               AND lr.exam_id = a.exam_id
               AND lr.rn = 1
            LEFT JOIN costs co
                ON co.group_exam_id = a.group_exam_id
            LEFT JOIN member_counts mc
                ON mc.group_exam_id = a.group_exam_id
            WHERE 1=1
                {where_sql}
        """)

        # ── Execute ──
        data_rows = db.session.execute(data_sql, params).fetchall()
        summary_row = db.session.execute(summary_sql, params).fetchone()

        total = summary_row.total or 0
        total_pages = (total + per_page - 1) // per_page if total > 0 else 0
        completed_count = summary_row.completed_count or 0
        passed_count = summary_row.passed_count or 0
        pass_rate = (
            round(passed_count / completed_count * 100, 1)
            if completed_count > 0 else None
        )

        # ── Batch material progress (only for page items) ──
        material_progress_map = {}
        if data_rows:
            page_ge_ids = list(set(row.group_exam_id for row in data_rows))
            page_exam_ids = list(set(row.exam_id for row in data_rows))
            page_user_ids = list(set(row.user_id for row in data_rows))

            # 1) Custom materials per group_exam
            ge_ids_str = ','.join(str(int(x)) for x in page_ge_ids)
            custom_mats = db.session.execute(text(
                f"SELECT group_exam_id, study_material_id "
                f"FROM group_exam_materials "
                f"WHERE group_exam_id IN ({ge_ids_str}) AND is_included = 1"
            )).fetchall()

            ge_material_map = {}
            for r in custom_mats:
                ge_material_map.setdefault(r.group_exam_id, []).append(
                    r.study_material_id
                )

            # 2) Fallback: linked materials from exams (M2M)
            exam_ids_str = ','.join(str(int(x)) for x in page_exam_ids)
            linked_mats = db.session.execute(text(
                f"SELECT exam_id, study_material_id "
                f"FROM study_material_exams "
                f"WHERE exam_id IN ({exam_ids_str})"
            )).fetchall()

            exam_material_map = {}
            for r in linked_mats:
                exam_material_map.setdefault(r.exam_id, []).append(
                    r.study_material_id
                )

            # Resolve: custom if available, else linked
            all_material_ids = set()
            ge_final_materials = {}
            for row in data_rows:
                mats = ge_material_map.get(row.group_exam_id) or \
                       exam_material_map.get(row.exam_id, [])
                ge_final_materials[row.group_exam_id] = mats
                all_material_ids.update(mats)

            if all_material_ids:
                # 3) All topics for these materials
                mat_ids_str = ','.join(str(int(x)) for x in all_material_ids)
                topics = db.session.execute(text(
                    f"SELECT st.id AS topic_id, ss.material_id "
                    f"FROM study_topics st "
                    f"JOIN study_sessions ss ON ss.id = st.session_id "
                    f"WHERE ss.material_id IN ({mat_ids_str})"
                )).fetchall()

                material_topics = {}
                all_topic_ids = set()
                for t in topics:
                    material_topics.setdefault(t.material_id, []).append(t.topic_id)
                    all_topic_ids.add(t.topic_id)

                # 4) Completion status
                completed_set = set()
                if all_topic_ids:
                    user_ids_str = ','.join(
                        f"'{str(x)}'" for x in page_user_ids
                    )
                    topic_ids_str = ','.join(str(int(x)) for x in all_topic_ids)
                    completions = db.session.execute(text(
                        f"SELECT user_id, topic_id "
                        f"FROM student_topic_progress "
                        f"WHERE user_id IN ({user_ids_str}) "
                        f"  AND topic_id IN ({topic_ids_str}) "
                        f"  AND is_completed = 1"
                    )).fetchall()
                    completed_set = {(c.user_id, c.topic_id) for c in completions}

                # 5) Assemble per (user_id, group_exam_id)
                for row in data_rows:
                    mats = ge_final_materials.get(row.group_exam_id, [])
                    total_t = 0
                    completed_t = 0
                    for mid in mats:
                        for tid in material_topics.get(mid, []):
                            total_t += 1
                            if (row.user_id, tid) in completed_set:
                                completed_t += 1
                    if total_t > 0:
                        material_progress_map[(row.user_id, row.group_exam_id)] = {
                            'total': total_t,
                            'completed': completed_t,
                            'percentage': round(completed_t / total_t * 100, 1),
                        }

        # ── Build JSON response ──
        # Batch-fetch assignment numbers for this page
        assignment_number_map = {}
        eca_id_map = {}
        vigencia_map = {}  # group_exam_id -> {validity_months, expires_at, extended_months, is_expired}
        ecm_vigencia_map = {}  # (user_id, ecm_id) -> {validity_months, expires_at, extended_months, is_expired}
        if data_rows:
            page_uids = list(set(row.user_id for row in data_rows))
            page_ge_ids_for_vigencia = list(set(row.group_exam_id for row in data_rows))
            if page_uids:
                uid_placeholders = ','.join(f"'{str(u)}'" for u in page_uids)
                ecm_assigns = db.session.execute(text(
                    f"SELECT id, user_id, assignment_number, validity_months, expires_at, extended_months "
                    f"FROM ecm_candidate_assignments "
                    f"WHERE user_id IN ({uid_placeholders}) AND competency_standard_id = :ecm_id"
                ), {'ecm_id': ecm_id}).fetchall()
                for ea in ecm_assigns:
                    assignment_number_map[ea.user_id] = ea.assignment_number
                    eca_id_map[ea.user_id] = ea.id
                    # Calculate effective expires for ECM assignments
                    ecm_expires = ea.expires_at
                    ecm_extended = ea.extended_months or 0
                    effective_ecm_expires = None
                    ecm_is_expired = False
                    if ecm_expires:
                        from dateutil.relativedelta import relativedelta
                        effective_ecm_expires = ecm_expires + relativedelta(months=ecm_extended)
                        ecm_is_expired = datetime.utcnow() > effective_ecm_expires
                    ecm_vigencia_map[ea.user_id] = {
                        'validity_months': ea.validity_months,
                        'expires_at': effective_ecm_expires.isoformat() if effective_ecm_expires else None,
                        'extended_months': ecm_extended,
                        'is_expired': ecm_is_expired,
                    }
            
            # Batch-fetch vigencia for GroupExams
            if page_ge_ids_for_vigencia:
                ge_ids_vig_str = ','.join(str(int(x)) for x in page_ge_ids_for_vigencia)
                ge_vig_rows = db.session.execute(text(
                    f"SELECT id, validity_months, expires_at, extended_months "
                    f"FROM group_exams WHERE id IN ({ge_ids_vig_str})"
                )).fetchall()
                for gv in ge_vig_rows:
                    gv_expires = gv.expires_at
                    gv_extended = gv.extended_months or 0
                    effective_gv_expires = None
                    gv_is_expired = False
                    if gv_expires:
                        from dateutil.relativedelta import relativedelta
                        effective_gv_expires = gv_expires + relativedelta(months=gv_extended)
                        gv_is_expired = datetime.utcnow() > effective_gv_expires
                    vigencia_map[gv.id] = {
                        'validity_months': gv.validity_months,
                        'expires_at': effective_gv_expires.isoformat() if effective_gv_expires else None,
                        'extended_months': gv_extended,
                        'is_expired': gv_is_expired,
                    }
        
        assignments = []
        for row in data_rows:
            # Certificate types from campus config
            cert_types = ['reporte_evaluacion']
            if row.enable_tier_standard:
                cert_types.append('certificado_eduit')
            if row.enable_digital_badge:
                cert_types.append('insignia_digital')
            if row.enable_tier_advanced:
                cert_types.append('certificado_conocer')

            # Result status
            if row.result_status_raw == 1:
                result_status = 'completed'
                passed = row.result_raw == 1
            elif row.result_status_raw == 0:
                result_status = 'in_progress'
                passed = None
            else:
                result_status = 'pending'
                passed = None

            assignments.append({
                'user_id': row.user_id,
                'user_name': row.user_name,
                'user_email': row.user_email,
                'user_role': row.user_role,
                'user_curp': row.user_curp,
                'assignment_number': assignment_number_map.get(row.user_id),
                'group_id': row.group_id,
                'group_name': row.group_name,
                'group_code': row.group_code,
                'campus_name': row.campus_name,
                'campus_id': row.campus_id,
                'partner_name': row.partner_name,
                'partner_id': row.partner_id,
                'exam_id': row.exam_id,
                'exam_name': row.exam_name,
                'exam_ecm_code': ecm.code,
                'assignment_date': (
                    row.assignment_date.isoformat()
                    if row.assignment_date else None
                ),
                'assignment_type': row.assignment_type,
                'unit_cost': (
                    round(float(row.unit_cost), 2)
                    if row.unit_cost is not None else 0
                ),
                'score': row.score,
                'result_status': result_status,
                'passed': passed,
                'result_date': (
                    row.result_date.isoformat()
                    if row.result_date else None
                ),
                'duration_seconds': row.duration_seconds,
                'certificate_code': (
                    row.certificate_code if passed else None
                ),
                'material_progress': material_progress_map.get(
                    (row.user_id, row.group_exam_id)
                ),
                'max_attempts': row.max_attempts,
                'time_limit': row.time_limit_minutes,
                'passing_score': row.passing_score,
                'certificate_types': cert_types,
                'vigencia': ecm_vigencia_map.get(row.user_id) or vigencia_map.get(row.group_exam_id),
                'group_exam_id': row.group_exam_id,
                'ecm_assignment_id': eca_id_map.get(row.user_id),
            })

        # ── Filter options (lightweight) ──
        available_exams = [
            {'id': e.id, 'name': e.name}
            for e in Exam.query.filter_by(
                competency_standard_id=ecm_id
            ).all()
        ]

        groups_data = db.session.execute(text("""
            SELECT DISTINCT cg.id, cg.name
            FROM group_exams ge
            JOIN exams e ON e.id = ge.exam_id
            JOIN candidate_groups cg ON cg.id = ge.group_id
            WHERE e.competency_standard_id = :ecm_id
            ORDER BY cg.name
        """), {'ecm_id': ecm_id}).fetchall()
        available_groups = [{'id': g.id, 'name': g.name} for g in groups_data]

        return jsonify({
            'ecm': {
                'id': ecm.id, 'code': ecm.code, 'name': ecm.name,
                'sector': ecm.sector, 'level': ecm.level,
                'certifying_body': ecm.certifying_body, 'logo_url': ecm.logo_url,
            },
            'assignments': assignments,
            'total': total,
            'pages': total_pages,
            'current_page': page,
            'per_page': per_page,
            'summary': {
                'total_assignments': total,
                'total_candidates': summary_row.unique_users or 0,
                'total_cost': round(float(summary_row.total_cost or 0), 2),
                'avg_score': (
                    float(summary_row.avg_score)
                    if summary_row.avg_score is not None else None
                ),
                'pass_rate': pass_rate,
                'completed_count': completed_count,
                'pending_count': summary_row.pending_count or 0,
                'passed_count': passed_count,
            },
            'filters': {
                'exams': available_exams,
                'groups': available_groups,
            }
        })

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@bp.route('/ecm-assignments/<int:ecm_id>/export', methods=['GET'])
@jwt_required()
def export_ecm_assignments_excel(ecm_id):
    """Exportar todas las asignaciones de un ECM a Excel.
    
    Sin paginación — descarga TODOS los registros con los filtros aplicados.
    Usa las mismas CTEs optimizadas del endpoint de detalle.
    """
    try:
        from app.models.exam import Exam
        from sqlalchemy import text
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from io import BytesIO

        user_id = get_jwt_identity()
        user = User.query.get(user_id)
        if not user or user.role not in ['admin', 'developer', 'coordinator']:
            return jsonify({'error': 'Acceso denegado'}), 403

        ecm = CompetencyStandard.query.get(ecm_id)
        if not ecm:
            return jsonify({'error': 'ECM no encontrado'}), 404

        search = request.args.get('search', '').strip()
        user_type_filter = request.args.get('user_type', 'all')
        status_filter = request.args.get('status', 'all')
        date_from = request.args.get('date_from')
        date_to = request.args.get('date_to')
        group_id_filter = request.args.get('group_id', type=int)
        exam_id_filter = request.args.get('exam_id', type=int)
        sort_by = request.args.get('sort_by', 'date')
        sort_dir = request.args.get('sort_dir', 'desc')

        # ── Dynamic WHERE ──
        params = {'ecm_id': ecm_id}
        where_parts = []

        if exam_id_filter:
            where_parts.append("AND a.exam_id = :exam_id_f")
            params['exam_id_f'] = exam_id_filter
        if group_id_filter:
            where_parts.append("AND a.group_id = :group_id_f")
            params['group_id_f'] = group_id_filter
        if user_type_filter != 'all':
            where_parts.append("AND a.user_role = :user_type_f")
            params['user_type_f'] = user_type_filter
        if search:
            where_parts.append(
                "AND (LOWER(a.user_name) LIKE :search_f OR LOWER(a.user_email) LIKE :search_f)"
            )
            params['search_f'] = f'%{search.lower()}%'
        if date_from:
            try:
                from_dt = datetime.strptime(date_from, '%Y-%m-%d')
                where_parts.append("AND a.assignment_date >= :date_from_f")
                params['date_from_f'] = from_dt
            except ValueError:
                pass
        if date_to:
            try:
                to_dt = datetime.strptime(date_to, '%Y-%m-%d').replace(
                    hour=23, minute=59, second=59
                )
                where_parts.append("AND a.assignment_date <= :date_to_f")
                params['date_to_f'] = to_dt
            except ValueError:
                pass
        if status_filter == 'completed':
            where_parts.append("AND lr.result_status_raw = 1")
        elif status_filter == 'pending':
            where_parts.append("AND (lr.result_status_raw IS NULL OR lr.result_status_raw != 1)")
        elif status_filter == 'passed':
            where_parts.append("AND lr.result_status_raw = 1 AND lr.result_raw = 1")
        elif status_filter == 'failed':
            where_parts.append("AND lr.result_status_raw = 1 AND lr.result_raw = 0")

        where_sql = '\n                '.join(where_parts)

        # MSSQL doesn't allow ORDER BY on column aliases
        sort_map = {
            'date': 'a.assignment_date',
            'name': 'a.user_name',
            'score': 'COALESCE(lr.score, -1)',
            'cost': 'COALESCE(CAST(COALESCE(co.total_cost, 0) AS FLOAT) / NULLIF(mc.member_count, 0), 0)',
            'role': 'a.user_role',
            'group': 'a.group_name',
            'exam': 'a.exam_name',
            'status': 'COALESCE(lr.result_status_raw, -1)',
            'duration': 'COALESCE(lr.duration_seconds, -1)',
        }
        sort_col = sort_map.get(sort_by, 'a.assignment_date')
        sort_direction = 'DESC' if sort_dir == 'desc' else 'ASC'

        # ── CTEs (same as detail endpoint, NO pagination) ──
        cte_sql = """
            WITH base_assignments AS (
                SELECT
                    u.id AS user_id,
                    CONCAT(u.name, ' ', u.first_surname,
                           COALESCE(' ' + u.second_surname, '')) AS user_name,
                    u.email AS user_email,
                    u.role AS user_role,
                    u.curp AS user_curp,
                    ge.id AS group_exam_id,
                    ge.exam_id,
                    ge.assigned_at AS assignment_date,
                    ge.assignment_type,
                    ge.max_attempts,
                    ge.time_limit_minutes,
                    ge.passing_score,
                    cg.id AS group_id,
                    cg.name AS group_name,
                    cg.code AS group_code,
                    c.id AS campus_id,
                    c.name AS campus_name,
                    COALESCE(c.enable_tier_basic, 0) AS enable_tier_basic,
                    COALESCE(c.enable_tier_standard, 0) AS enable_tier_standard,
                    COALESCE(c.enable_tier_advanced, 0) AS enable_tier_advanced,
                    COALESCE(c.enable_digital_badge, 0) AS enable_digital_badge,
                    p.id AS partner_id,
                    p.name AS partner_name,
                    e.name AS exam_name
                FROM group_exams ge
                JOIN exams e ON e.id = ge.exam_id
                JOIN candidate_groups cg ON cg.id = ge.group_id
                LEFT JOIN campuses c ON c.id = cg.campus_id
                LEFT JOIN partners p ON p.id = c.partner_id
                JOIN group_members gm ON gm.group_id = ge.group_id
                JOIN users u ON u.id = gm.user_id
                WHERE e.competency_standard_id = :ecm_id
                  AND ge.assignment_type = 'all'

                UNION ALL

                SELECT
                    u.id,
                    CONCAT(u.name, ' ', u.first_surname,
                           COALESCE(' ' + u.second_surname, '')),
                    u.email, u.role, u.curp,
                    ge.id, ge.exam_id, ge.assigned_at,
                    ge.assignment_type,
                    ge.max_attempts, ge.time_limit_minutes, ge.passing_score,
                    cg.id, cg.name, cg.code,
                    c.id, c.name,
                    COALESCE(c.enable_tier_basic, 0),
                    COALESCE(c.enable_tier_standard, 0),
                    COALESCE(c.enable_tier_advanced, 0),
                    COALESCE(c.enable_digital_badge, 0),
                    p.id, p.name,
                    e.name
                FROM group_exams ge
                JOIN exams e ON e.id = ge.exam_id
                JOIN candidate_groups cg ON cg.id = ge.group_id
                LEFT JOIN campuses c ON c.id = cg.campus_id
                LEFT JOIN partners p ON p.id = c.partner_id
                JOIN group_exam_members gem ON gem.group_exam_id = ge.id
                JOIN users u ON u.id = gem.user_id
                WHERE e.competency_standard_id = :ecm_id
                  AND ge.assignment_type != 'all'
            ),
            latest_results AS (
                SELECT
                    r.user_id, r.exam_id, r.score,
                    r.status AS result_status_raw,
                    r.result AS result_raw,
                    r.end_date, r.duration_seconds, r.certificate_code,
                    ROW_NUMBER() OVER (
                        PARTITION BY r.user_id, r.exam_id
                        ORDER BY r.created_at DESC
                    ) AS rn
                FROM results r
                WHERE r.exam_id IN (
                    SELECT id FROM exams
                    WHERE competency_standard_id = :ecm_id
                )
            ),
            costs AS (
                SELECT
                    bt.reference_id AS group_exam_id,
                    SUM(bt.amount) AS total_cost
                FROM balance_transactions bt
                WHERE bt.reference_type = 'group_exam'
                  AND bt.concept IN ('asignacion_certificacion', 'asignacion_retoma')
                  AND bt.reference_id IN (
                      SELECT ge.id FROM group_exams ge
                      JOIN exams e ON e.id = ge.exam_id
                      WHERE e.competency_standard_id = :ecm_id
                  )
                GROUP BY bt.reference_id
            ),
            member_counts AS (
                SELECT ge.id AS group_exam_id, COUNT(gm.id) AS member_count
                FROM group_exams ge
                JOIN exams e ON e.id = ge.exam_id
                JOIN group_members gm ON gm.group_id = ge.group_id
                WHERE e.competency_standard_id = :ecm_id
                  AND ge.assignment_type = 'all'
                GROUP BY ge.id

                UNION ALL

                SELECT ge.id, COUNT(gem.id)
                FROM group_exams ge
                JOIN exams e ON e.id = ge.exam_id
                JOIN group_exam_members gem ON gem.group_exam_id = ge.id
                WHERE e.competency_standard_id = :ecm_id
                  AND ge.assignment_type != 'all'
                GROUP BY ge.id
            )
        """

        data_sql = text(cte_sql + f"""
            SELECT
                a.user_name, a.user_email, a.user_role, a.user_curp,
                a.group_name, a.group_code,
                a.campus_name, a.partner_name,
                a.exam_name,
                a.assignment_date,
                a.passing_score,
                a.max_attempts,
                a.time_limit_minutes,
                a.enable_tier_basic, a.enable_tier_standard,
                a.enable_tier_advanced, a.enable_digital_badge,
                lr.score, lr.result_status_raw, lr.result_raw,
                lr.end_date AS result_date, lr.duration_seconds, lr.certificate_code,
                CAST(COALESCE(co.total_cost, 0) AS FLOAT)
                    / NULLIF(mc.member_count, 0) AS unit_cost
            FROM base_assignments a
            LEFT JOIN latest_results lr
                ON lr.user_id = a.user_id
               AND lr.exam_id = a.exam_id
               AND lr.rn = 1
            LEFT JOIN costs co
                ON co.group_exam_id = a.group_exam_id
            LEFT JOIN member_counts mc
                ON mc.group_exam_id = a.group_exam_id
            WHERE 1=1
                {where_sql}
            ORDER BY {sort_col} {sort_direction}
        """)

        rows = db.session.execute(data_sql, params).fetchall()

        # Batch-fetch assignment numbers for all rows
        export_assignment_map = {}
        if rows:
            export_uids = list(set(row.user_id for row in rows))
            for i in range(0, len(export_uids), 500):
                batch = export_uids[i:i+500]
                uid_str = ','.join(f"'{u}'" for u in batch)
                ea_rows = db.session.execute(text(
                    f"SELECT user_id, assignment_number FROM ecm_candidate_assignments "
                    f"WHERE user_id IN ({uid_str}) AND competency_standard_id = :ecm_id"
                ), {'ecm_id': ecm_id}).fetchall()
                for ea in ea_rows:
                    export_assignment_map[ea.user_id] = ea.assignment_number

        # ── Build Excel ──
        wb = Workbook()
        ws = wb.active
        ws.title = f"ECM {ecm.code}"[:31]  # Max 31 chars for sheet name

        # Styles
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        thin_border = Border(
            left=Side(style='thin', color='D1D5DB'),
            right=Side(style='thin', color='D1D5DB'),
            top=Side(style='thin', color='D1D5DB'),
            bottom=Side(style='thin', color='D1D5DB'),
        )
        center_align = Alignment(horizontal='center', vertical='center')

        # Title row
        ws.merge_cells('A1:X1')
        title_cell = ws['A1']
        title_cell.value = f"Asignaciones ECM: {ecm.code} - {ecm.name}"
        title_cell.font = Font(bold=True, size=14, color="4F46E5")
        title_cell.alignment = Alignment(horizontal='center')

        ws.merge_cells('A2:X2')
        subtitle = ws['A2']
        subtitle.value = f"Sector: {ecm.sector or 'N/A'} | Nivel: {ecm.level or 'N/A'} | Total registros: {len(rows)} | Exportado: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        subtitle.font = Font(size=10, color="6B7280")
        subtitle.alignment = Alignment(horizontal='center')

        # Headers (row 4)
        headers = [
            'Nº Asignación', 'Nombre', 'Email', 'CURP', 'Rol', 'Grupo', 'Código Grupo',
            'Sede', 'Partner', 'Examen', 'Fecha Asignación',
            'Costo', 'Calificación', 'Estado', 'Aprobado',
            'Fecha Resultado', 'Duración (min)', 'Código Certificado',
            'Rep. Evaluación', 'Cert. Eduit', 'Insignia Digital', 'Cert. CONOCER',
            'Calif. Mínima', 'Intentos Máx', 'Tiempo Límite (min)',
        ]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # Role labels
        role_labels = {
            'candidato': 'Candidato',
            'responsable': 'Responsable',
            'coordinator': 'Coordinador',
            'admin': 'Administrador',
        }

        # Data rows
        for row_idx, row in enumerate(rows, 5):
            if row.result_status_raw == 1:
                status_txt = 'Completado'
                passed_txt = 'Sí' if row.result_raw == 1 else 'No'
            elif row.result_status_raw == 0:
                status_txt = 'En proceso'
                passed_txt = ''
            else:
                status_txt = 'Pendiente'
                passed_txt = ''

            duration_min = round(row.duration_seconds / 60, 1) if row.duration_seconds else None
            assign_date = row.assignment_date.strftime('%Y-%m-%d') if row.assignment_date else ''
            result_date = row.result_date.strftime('%Y-%m-%d %H:%M') if row.result_date else ''

            data = [
                export_assignment_map.get(row.user_id, ''),
                row.user_name,
                row.user_email,
                row.user_curp or '',
                role_labels.get(row.user_role, row.user_role),
                row.group_name,
                row.group_code,
                row.campus_name or '',
                row.partner_name or '',
                row.exam_name,
                assign_date,
                round(float(row.unit_cost), 2) if row.unit_cost else 0,
                row.score if row.score is not None else '',
                status_txt,
                passed_txt,
                result_date,
                duration_min if duration_min else '',
                row.certificate_code if row.result_raw == 1 else '',
                'Sí',  # Rep. Evaluación siempre disponible
                'Sí' if row.enable_tier_standard else 'No',
                'Sí' if row.enable_digital_badge else 'No',
                'Sí' if row.enable_tier_advanced else 'No',
                row.passing_score if row.passing_score else '',
                row.max_attempts if row.max_attempts else '',
                row.time_limit_minutes if row.time_limit_minutes else '',
            ]
            for col, val in enumerate(data, 1):
                cell = ws.cell(row=row_idx, column=col, value=val)
                cell.border = thin_border
                if col >= 10:
                    cell.alignment = center_align

        # Column widths
        col_widths = [16, 30, 30, 20, 14, 20, 14, 20, 20, 30, 14,
                      10, 12, 12, 10, 18, 14, 18, 14, 12, 14, 14, 12, 12, 16]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        # Auto-filter + freeze
        ws.auto_filter.ref = f"A4:{get_column_letter(len(headers))}{4 + len(rows)}"
        ws.freeze_panes = 'A5'

        # Save to buffer
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        safe_code = ecm.code.replace(' ', '_').replace('/', '-')
        filename = f"Asignaciones_{safe_code}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename,
        )

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


# ═══════════════════════════════════════════════════════════════
# DETALLE DE ASIGNACIÓN INDIVIDUAL POR CANDIDATO
# ═══════════════════════════════════════════════════════════════

@bp.route('/ecm-assignments/candidate/<int:eca_id>', methods=['GET'])
@jwt_required()
def get_candidate_assignment_detail(eca_id):
    """Detalle completo de una EcmCandidateAssignment para trazabilidad individual.
    
    Retorna: datos del candidato, ECM, examen, grupo, plantel, partner,
    configuración de la asignación, todos los intentos de resultado,
    historial de retomas, y progreso de material de estudio.
    """
    try:
        from app.models.exam import Exam
        from app.models import Result
        from app.models.partner import EcmCandidateAssignment, EcmRetake
        from sqlalchemy import text
        from dateutil.relativedelta import relativedelta

        user_id = get_jwt_identity()
        user = User.query.get(user_id)
        if not user or user.role not in ['admin', 'developer', 'coordinator']:
            return jsonify({'error': 'Acceso denegado'}), 403

        eca = EcmCandidateAssignment.query.get(eca_id)
        if not eca:
            return jsonify({'error': 'Asignación no encontrada'}), 404

        # ── User ──
        candidate = User.query.get(eca.user_id)
        if not candidate:
            return jsonify({'error': 'Candidato no encontrado'}), 404

        # ── ECM ──
        ecm = CompetencyStandard.query.get(eca.competency_standard_id)

        # ── Exam ──
        exam = Exam.query.get(eca.exam_id) if eca.exam_id else None

        # ── Group Exam (assignment config) ──
        group_exam = GroupExam.query.get(eca.group_exam_id) if eca.group_exam_id else None

        # ── Group ──
        group = CandidateGroup.query.get(eca.group_id) if eca.group_id else None

        # ── Campus ──
        campus = Campus.query.get(eca.campus_id) if eca.campus_id else None

        # ── Partner ──
        partner = None
        if campus and campus.partner_id:
            partner = Partner.query.get(campus.partner_id)

        # ── Assigned by ──
        assigned_by_user = User.query.get(eca.assigned_by_id) if eca.assigned_by_id else None

        # ── Vigencia ──
        ecm_expires = eca.expires_at
        ecm_extended = eca.extended_months or 0
        effective_expires = None
        is_expired = False
        if ecm_expires:
            effective_expires = ecm_expires + relativedelta(months=ecm_extended)
            is_expired = datetime.utcnow() > effective_expires

        # ── All results for this user + exam ──
        results_list = []
        if eca.exam_id:
            results_query = Result.query.filter_by(
                user_id=eca.user_id,
                exam_id=eca.exam_id,
            ).order_by(Result.created_at.desc()).all()
            for r in results_query:
                results_list.append({
                    'id': r.id,
                    'score': r.score,
                    'status': r.status,
                    'result': r.result,
                    'start_date': r.start_date.isoformat() if r.start_date else None,
                    'end_date': r.end_date.isoformat() if r.end_date else None,
                    'duration_seconds': r.duration_seconds,
                    'certificate_code': r.certificate_code,
                    'certificate_url': r.certificate_url,
                    'report_url': r.report_url,
                    'created_at': r.created_at.isoformat() if r.created_at else None,
                })

        # ── Retakes ──
        retakes_list = []
        retakes = EcmRetake.query.filter_by(assignment_id=eca.id).order_by(EcmRetake.applied_at.desc()).all()
        for rt in retakes:
            applied_by = User.query.get(rt.applied_by_id) if rt.applied_by_id else None
            retakes_list.append({
                'id': rt.id,
                'cost': float(rt.cost) if rt.cost else 0,
                'status': rt.status,
                'result_id': rt.result_id,
                'applied_at': rt.applied_at.isoformat() if rt.applied_at else None,
                'used_at': rt.used_at.isoformat() if rt.used_at else None,
                'applied_by': {
                    'id': applied_by.id,
                    'name': applied_by.full_name,
                } if applied_by else None,
            })

        # ── Material progress ──
        material_progress = None
        materials_detail = []
        if group_exam and exam:
            # Custom materials for this group_exam
            custom_mats = db.session.execute(text(
                "SELECT study_material_id FROM group_exam_materials "
                "WHERE group_exam_id = :ge_id AND is_included = 1"
            ), {'ge_id': group_exam.id}).fetchall()
            mat_ids = [r.study_material_id for r in custom_mats]

            # Fallback: linked materials from exam
            if not mat_ids:
                linked = db.session.execute(text(
                    "SELECT study_material_id FROM study_material_exams WHERE exam_id = :eid"
                ), {'eid': exam.id}).fetchall()
                mat_ids = [r.study_material_id for r in linked]

            if mat_ids:
                mat_ids_str = ','.join(str(int(x)) for x in mat_ids)
                # Get materials info
                mat_rows = db.session.execute(text(
                    f"SELECT id, title FROM study_contents WHERE id IN ({mat_ids_str})"
                )).fetchall()
                mat_name_map = {r.id: r.title for r in mat_rows}

                # Get topics per material
                topics_rows = db.session.execute(text(
                    f"SELECT st.id AS topic_id, ss.material_id "
                    f"FROM study_topics st "
                    f"JOIN study_sessions ss ON ss.id = st.session_id "
                    f"WHERE ss.material_id IN ({mat_ids_str})"
                )).fetchall()

                material_topics = {}
                all_topic_ids = set()
                for t in topics_rows:
                    material_topics.setdefault(t.material_id, []).append(t.topic_id)
                    all_topic_ids.add(t.topic_id)

                # Get completions for this user
                completed_topics = set()
                if all_topic_ids:
                    topic_ids_str = ','.join(str(int(x)) for x in all_topic_ids)
                    completions = db.session.execute(text(
                        f"SELECT topic_id FROM student_topic_progress "
                        f"WHERE user_id = :uid AND topic_id IN ({topic_ids_str}) AND is_completed = 1"
                    ), {'uid': eca.user_id}).fetchall()
                    completed_topics = {c.topic_id for c in completions}

                total_topics = 0
                completed_count = 0
                for mid in mat_ids:
                    m_topics = material_topics.get(mid, [])
                    m_completed = sum(1 for tid in m_topics if tid in completed_topics)
                    total_topics += len(m_topics)
                    completed_count += m_completed
                    materials_detail.append({
                        'id': mid,
                        'name': mat_name_map.get(mid, f'Material {mid}'),
                        'topics_total': len(m_topics),
                        'topics_completed': m_completed,
                        'percentage': round(m_completed / len(m_topics) * 100, 1) if m_topics else 0,
                    })

                if total_topics > 0:
                    material_progress = {
                        'total': total_topics,
                        'completed': completed_count,
                        'percentage': round(completed_count / total_topics * 100, 1),
                    }

        # ── Certificate types from campus config ──
        cert_types = ['reporte_evaluacion']
        if campus:
            # Effective config (group override → campus default)
            eff_standard = (group.enable_tier_standard_override if group and group.enable_tier_standard_override is not None
                            else (campus.enable_tier_standard if campus.enable_tier_standard is not None else False))
            eff_advanced = (group.enable_tier_advanced_override if group and group.enable_tier_advanced_override is not None
                           else (campus.enable_tier_advanced if campus.enable_tier_advanced is not None else False))
            eff_badge = (group.enable_digital_badge_override if group and group.enable_digital_badge_override is not None
                         else (campus.enable_digital_badge if campus.enable_digital_badge is not None else False))
            if eff_standard:
                cert_types.append('certificado_eduit')
            if eff_badge:
                cert_types.append('insignia_digital')
            if eff_advanced:
                cert_types.append('certificado_conocer')

        # ── Cost (effective certification_cost from group override → campus default) ──
        certification_cost = 0
        retake_cost = 0
        if group and hasattr(group, 'certification_cost_override') and group.certification_cost_override is not None:
            certification_cost = float(group.certification_cost_override)
        elif campus and campus.certification_cost is not None:
            certification_cost = float(campus.certification_cost)
        if group and hasattr(group, 'retake_cost_override') and group.retake_cost_override is not None:
            retake_cost = float(group.retake_cost_override)
        elif campus and campus.retake_cost is not None:
            retake_cost = float(campus.retake_cost)

        return jsonify({
            'assignment': {
                'id': eca.id,
                'assignment_number': eca.assignment_number,
                'assignment_source': eca.assignment_source,
                'validity_months': eca.validity_months,
                'expires_at': effective_expires.isoformat() if effective_expires else None,
                'extended_months': ecm_extended,
                'is_expired': is_expired,
                'assigned_at': eca.assigned_at.isoformat() if eca.assigned_at else None,
                'certification_cost': certification_cost,
                'retake_cost': retake_cost,
            },
            'user': {
                'id': candidate.id,
                'name': candidate.name,
                'full_name': candidate.full_name,
                'email': candidate.email,
                'curp': candidate.curp,
                'role': candidate.role,
                'username': candidate.username,
            },
            'ecm': {
                'id': ecm.id,
                'code': ecm.code,
                'name': ecm.name,
                'sector': ecm.sector,
                'level': ecm.level,
                'certifying_body': ecm.certifying_body,
                'logo_url': ecm.logo_url,
            } if ecm else None,
            'exam': {
                'id': exam.id,
                'name': exam.name,
                'version': exam.version,
                'duration_minutes': exam.duration_minutes,
                'passing_score': exam.passing_score,
                'description': exam.description,
                'total_questions': exam.get_total_questions() if hasattr(exam, 'get_total_questions') else None,
            } if exam else None,
            'group_exam': {
                'id': group_exam.id,
                'assignment_type': group_exam.assignment_type,
                'max_attempts': group_exam.max_attempts,
                'time_limit_minutes': group_exam.time_limit_minutes,
                'passing_score': group_exam.passing_score,
                'content_type': group_exam.exam_content_type,
                'assigned_at': group_exam.assigned_at.isoformat() if group_exam.assigned_at else None,
                'max_retakes': group.max_retakes_override if group and group.max_retakes_override is not None else (campus.max_retakes if campus and campus.max_retakes is not None else 0),
                'validity_months': group_exam.validity_months,
            } if group_exam else None,
            'group': {
                'id': group.id,
                'name': group.name,
                'code': group.code,
            } if group else {'id': eca.group_id, 'name': eca.group_name, 'code': None},
            'campus': {
                'id': campus.id,
                'name': campus.name,
                'state_name': campus.state_name,
            } if campus else None,
            'partner': {
                'id': partner.id,
                'name': partner.name,
            } if partner else None,
            'assigned_by': {
                'id': assigned_by_user.id,
                'name': assigned_by_user.full_name,
                'email': assigned_by_user.email,
            } if assigned_by_user else None,
            'results': results_list,
            'retakes': retakes_list,
            'material_progress': material_progress,
            'materials_detail': materials_detail,
            'certificate_types': cert_types,
        })

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


# =====================================================
# ENDPOINTS PARA RESPONSABLE DE PARTNER
# =====================================================

def responsable_partner_required(f):
    """Decorador que requiere rol de responsable_partner (o admin/developer) y asocia el partner"""
    @wraps(f)
    def decorated(*args, **kwargs):
        user_id = get_jwt_identity()
        user = User.query.get(user_id)
        if not user:
            return jsonify({'error': 'No autorizado'}), 401
        
        # Admin y developer pueden acceder si pasan partner_id como query param
        if user.role in ['admin', 'developer']:
            partner_id = request.args.get('partner_id', type=int)
            if partner_id:
                partner = Partner.query.get(partner_id)
                if not partner:
                    return jsonify({'error': 'Partner no encontrado'}), 404
                g.current_user = user
                g.partner = partner
                return f(*args, **kwargs)
            # Si es admin sin partner_id, buscar si tiene asociación en user_partners
            from app.models.partner import user_partners
            partner_row = db.session.query(user_partners).filter(
                user_partners.c.user_id == user.id
            ).first()
            if partner_row:
                partner = Partner.query.get(partner_row.partner_id)
                if partner:
                    g.current_user = user
                    g.partner = partner
                    return f(*args, **kwargs)
            # Si admin no tiene partner asociado, mostrar el primer partner
            first_partner = Partner.query.filter_by(is_active=True).first()
            if first_partner:
                g.current_user = user
                g.partner = first_partner
                return f(*args, **kwargs)
            return jsonify({'error': 'No hay partners disponibles'}), 404
        
        if user.role != 'responsable_partner':
            return jsonify({'error': 'Acceso denegado. Se requiere rol de responsable_partner'}), 403
        # Buscar el partner asociado en la tabla user_partners
        from app.models.partner import user_partners
        partner_row = db.session.query(user_partners).filter(
            user_partners.c.user_id == user.id
        ).first()
        if not partner_row:
            return jsonify({'error': 'No tienes un partner asignado'}), 403
        partner = Partner.query.get(partner_row.partner_id)
        if not partner:
            return jsonify({'error': 'Partner no encontrado'}), 404
        g.current_user = user
        g.partner = partner
        return f(*args, **kwargs)
    return decorated


@bp.route('/mi-partner', methods=['GET'])
@jwt_required()
@responsable_partner_required
def get_mi_partner():
    """Obtener información del partner y sus planteles"""
    try:
        partner = g.partner
        campuses = Campus.query.filter_by(partner_id=partner.id).all()
        
        # Obtener estados únicos de los campus
        states = sorted(set([c.state_name for c in campuses if c.state_name]))
        
        return jsonify({
            'partner': partner.to_dict(include_states=True),
            'campuses': [c.to_dict() for c in campuses],
            'states': states
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@bp.route('/mi-partner/dashboard', methods=['GET'])
@jwt_required()
@responsable_partner_required
def get_mi_partner_dashboard():
    """Dashboard con KPIs y gráficos para el responsable del partner"""
    from app.models import Result, Exam
    from app.models.student_progress import StudentContentProgress
    from app.models.conocer_certificate import ConocerCertificate
    from sqlalchemy import func
    
    try:
        partner = g.partner
        state_filter = request.args.get('state', '')
        
        # Obtener campuses del partner (opcionalmente filtrados por estado)
        campus_query = Campus.query.filter_by(partner_id=partner.id)
        if state_filter:
            campus_query = campus_query.filter_by(state_name=state_filter)
        campuses = campus_query.all()
        campus_ids = [c.id for c in campuses]
        
        if not campus_ids:
            return jsonify({
                'partner': {'id': partner.id, 'name': partner.name},
                'filter': {'state': state_filter},
                'stats': {
                    'total_campuses': 0, 'total_groups': 0, 'total_candidates': 0,
                    'total_evaluations': 0, 'passed_evaluations': 0, 'failed_evaluations': 0,
                    'approval_rate': 0, 'average_score': 0, 'certification_rate': 0
                },
                'charts': {
                    'approval_by_campus': [], 'scores_by_campus': [],
                    'score_distribution': [], 'evaluations_over_time': [],
                    'certification_by_type': {}, 'candidates_by_state': []
                }
            })
        
        # Obtener todos los grupos de estos campus
        groups = CandidateGroup.query.filter(
            CandidateGroup.campus_id.in_(campus_ids),
            CandidateGroup.is_active == True
        ).all()
        group_ids = [gr.id for gr in groups]
        
        # Obtener todos los candidatos
        candidate_ids = []
        if group_ids:
            candidate_ids = [c[0] for c in db.session.query(GroupMember.user_id).filter(
                GroupMember.group_id.in_(group_ids), GroupMember.status == 'active'
            ).distinct().all()]
        
        # ---- Stats generales ----
        total_evals = Result.query.filter(Result.user_id.in_(candidate_ids), Result.status == 1).count() if candidate_ids else 0
        passed = Result.query.filter(Result.user_id.in_(candidate_ids), Result.status == 1, Result.result == 1).count() if candidate_ids else 0
        failed = total_evals - passed
        avg_score = 0
        if candidate_ids:
            avg = db.session.query(func.avg(Result.score)).filter(Result.user_id.in_(candidate_ids), Result.status == 1).scalar()
            avg_score = round(float(avg or 0), 1)
        
        certified_count = 0
        if candidate_ids:
            certified_count = db.session.query(Result.user_id).filter(
                Result.user_id.in_(candidate_ids), Result.status == 1, Result.result == 1
            ).distinct().count()
        
        # ---- Aprobación por campus ----
        campus_map = {c.id: c for c in campuses}
        approval_by_campus = []
        scores_by_campus = []
        for c in campuses:
            c_groups = [gr for gr in groups if gr.campus_id == c.id]
            c_group_ids = [gr.id for gr in c_groups]
            if not c_group_ids:
                approval_by_campus.append({'campus_name': c.name, 'campus_id': c.id, 'state': c.state_name or '', 'approved': 0, 'failed': 0, 'rate': 0, 'total_members': 0})
                scores_by_campus.append({'campus_name': c.name, 'average': 0, 'min': 0, 'max': 0})
                continue
            c_member_ids = [m[0] for m in db.session.query(GroupMember.user_id).filter(
                GroupMember.group_id.in_(c_group_ids), GroupMember.status == 'active'
            ).distinct().all()]
            if not c_member_ids:
                approval_by_campus.append({'campus_name': c.name, 'campus_id': c.id, 'state': c.state_name or '', 'approved': 0, 'failed': 0, 'rate': 0, 'total_members': 0})
                scores_by_campus.append({'campus_name': c.name, 'average': 0, 'min': 0, 'max': 0})
                continue
            c_approved = Result.query.filter(Result.user_id.in_(c_member_ids), Result.status == 1, Result.result == 1).count()
            c_failed = Result.query.filter(Result.user_id.in_(c_member_ids), Result.status == 1, Result.result == 0).count()
            c_total = c_approved + c_failed
            approval_by_campus.append({
                'campus_name': c.name, 'campus_id': c.id, 'state': c.state_name or '',
                'approved': c_approved, 'failed': c_failed,
                'total_members': len(c_member_ids),
                'rate': round(c_approved / c_total * 100, 1) if c_total > 0 else 0
            })
            c_stats = db.session.query(
                func.avg(Result.score), func.min(Result.score), func.max(Result.score)
            ).filter(Result.user_id.in_(c_member_ids), Result.status == 1).first()
            scores_by_campus.append({
                'campus_name': c.name,
                'average': round(float(c_stats[0] or 0), 1),
                'min': round(float(c_stats[1] or 0), 1),
                'max': round(float(c_stats[2] or 0), 1)
            })
        
        # ---- Distribución de calificaciones ----
        score_distribution = []
        if candidate_ids:
            for low in range(0, 100, 10):
                high = low + 10
                label = f"{low}-{high}"
                if low == 90:
                    count = Result.query.filter(
                        Result.user_id.in_(candidate_ids), Result.status == 1,
                        Result.score >= low, Result.score <= 100
                    ).count()
                    label = "90-100"
                else:
                    count = Result.query.filter(
                        Result.user_id.in_(candidate_ids), Result.status == 1,
                        Result.score >= low, Result.score < high
                    ).count()
                score_distribution.append({'range': label, 'count': count})
        
        # ---- Evaluaciones en el tiempo (últimos 6 meses) ----
        evaluations_over_time = []
        if candidate_ids:
            from datetime import datetime as dt
            now = dt.utcnow()
            for i in range(5, -1, -1):
                month = now.month - i
                year = now.year
                if month <= 0:
                    month += 12
                    year -= 1
                month_start = dt(year, month, 1)
                if month == 12:
                    month_end = dt(year + 1, 1, 1)
                else:
                    month_end = dt(year, month + 1, 1)
                approved_m = Result.query.filter(
                    Result.user_id.in_(candidate_ids), Result.status == 1, Result.result == 1,
                    Result.end_date >= month_start, Result.end_date < month_end
                ).count()
                failed_m = Result.query.filter(
                    Result.user_id.in_(candidate_ids), Result.status == 1, Result.result == 0,
                    Result.end_date >= month_start, Result.end_date < month_end
                ).count()
                evaluations_over_time.append({
                    'month': f"{year}-{month:02d}",
                    'approved': approved_m, 'failed': failed_m
                })
        
        # ---- Certificación por tipo ----
        certification_by_type = {
            'reporte_evaluacion': 0, 'certificado_eduit': 0,
            'certificado_conocer': 0, 'insignia_digital': 0
        }
        if candidate_ids:
            # Contar resultados con reporte/certificado generado
            with_report = Result.query.filter(
                Result.user_id.in_(candidate_ids), Result.status == 1, Result.result == 1,
                Result.report_url.isnot(None)
            ).count()
            with_cert = Result.query.filter(
                Result.user_id.in_(candidate_ids), Result.status == 1, Result.result == 1,
                Result.certificate_url.isnot(None)
            ).count()
            conocer_count = ConocerCertificate.query.filter(
                ConocerCertificate.user_id.in_(candidate_ids),
                ConocerCertificate.status == 'active'
            ).count()
            # Insignias digitales: candidatos habilitados con resultado aprobado
            badge_count = db.session.query(Result.user_id).join(
                User, Result.user_id == User.id
            ).filter(
                Result.user_id.in_(candidate_ids), Result.status == 1, Result.result == 1,
                User.enable_digital_badge == True
            ).distinct().count()
            
            certification_by_type = {
                'reporte_evaluacion': with_report,
                'certificado_eduit': with_cert,
                'certificado_conocer': conocer_count,
                'insignia_digital': badge_count
            }
        
        # ---- Candidatos por estado ----
        candidates_by_state = []
        state_groups = {}
        for c in campuses:
            st = c.state_name or 'Sin estado'
            if st not in state_groups:
                state_groups[st] = {'groups': [], 'campuses': 0}
            state_groups[st]['campuses'] += 1
            state_groups[st]['groups'].extend([gr for gr in groups if gr.campus_id == c.id])
        
        for st_name, st_info in sorted(state_groups.items()):
            st_group_ids = [gr.id for gr in st_info['groups']]
            st_member_count = 0
            if st_group_ids:
                st_member_count = db.session.query(GroupMember.user_id).filter(
                    GroupMember.group_id.in_(st_group_ids), GroupMember.status == 'active'
                ).distinct().count()
            candidates_by_state.append({
                'state': st_name,
                'campuses': st_info['campuses'],
                'candidates': st_member_count
            })
        
        # Obtener estados disponibles para el filtro
        all_states = sorted(set([c.state_name for c in Campus.query.filter_by(partner_id=partner.id).all() if c.state_name]))
        
        return jsonify({
            'partner': {'id': partner.id, 'name': partner.name, 'logo_url': partner.logo_url},
            'filter': {'state': state_filter, 'available_states': all_states},
            'stats': {
                'total_campuses': len(campuses),
                'total_groups': len(groups),
                'total_candidates': len(candidate_ids),
                'total_evaluations': total_evals,
                'passed_evaluations': passed,
                'failed_evaluations': failed,
                'approval_rate': round(passed / total_evals * 100, 1) if total_evals > 0 else 0,
                'average_score': avg_score,
                'certification_rate': round(certified_count / len(candidate_ids) * 100, 1) if candidate_ids else 0
            },
            'charts': {
                'approval_by_campus': approval_by_campus,
                'scores_by_campus': scores_by_campus,
                'score_distribution': score_distribution,
                'evaluations_over_time': evaluations_over_time,
                'certification_by_type': certification_by_type,
                'candidates_by_state': candidates_by_state
            }
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@bp.route('/mi-partner/certificates', methods=['GET'])
@jwt_required()
@responsable_partner_required
def get_mi_partner_certificates():
    """Tabla plana de certificados con filtros para el responsable del partner"""
    from app.models import Result, Exam
    from app.models.conocer_certificate import ConocerCertificate
    
    try:
        partner = g.partner
        
        # Filtros
        state_filter = request.args.get('state', '')
        campus_filter = request.args.get('campus_id', '', type=str)
        group_filter = request.args.get('group_id', '', type=str)
        cert_type_filter = request.args.get('cert_type', '')  # reporte_evaluacion, certificado_eduit, certificado_conocer, insignia_digital
        search = request.args.get('search', '').strip()
        page = request.args.get('page', 1, type=int)
        per_page = request.args.get('per_page', 25, type=int)
        per_page = min(per_page, 100)
        
        # Obtener campuses del partner filtrados
        campus_query = Campus.query.filter_by(partner_id=partner.id)
        if state_filter:
            campus_query = campus_query.filter_by(state_name=state_filter)
        if campus_filter:
            campus_query = campus_query.filter_by(id=int(campus_filter))
        campuses = campus_query.all()
        campus_ids = [c.id for c in campuses]
        campus_map = {c.id: c for c in campuses}
        
        if not campus_ids:
            return jsonify({
                'certificates': [], 'pagination': {'total': 0, 'page': page, 'per_page': per_page, 'pages': 0},
                'filters': {'states': [], 'campuses': [], 'groups': []}
            })
        
        # Obtener grupos
        group_query = CandidateGroup.query.filter(CandidateGroup.campus_id.in_(campus_ids))
        if group_filter:
            group_query = group_query.filter_by(id=int(group_filter))
        all_groups = group_query.all()
        group_ids = [gr.id for gr in all_groups]
        group_map = {gr.id: gr for gr in all_groups}
        group_campus_map = {gr.id: gr.campus_id for gr in all_groups}
        
        # Obtener miembros y mapeo usuario -> grupo
        member_rows = db.session.query(GroupMember.user_id, GroupMember.group_id).filter(
            GroupMember.group_id.in_(group_ids), GroupMember.status == 'active'
        ).all() if group_ids else []
        
        user_group_map = {}
        for uid, gid in member_rows:
            if uid not in user_group_map:
                user_group_map[uid] = gid  # Tomar el primer grupo
        candidate_ids = list(user_group_map.keys())
        
        if not candidate_ids:
            return jsonify({
                'certificates': [], 'pagination': {'total': 0, 'page': page, 'per_page': per_page, 'pages': 0},
                'filters': _get_partner_cert_filters(partner.id)
            })
        
        # Precomputar configuración de certificados por grupo
        # enable_tier_basic → reporte_evaluacion
        # enable_tier_standard → certificado_eduit
        # enable_tier_advanced → certificado_conocer
        # enable_digital_badge → insignia_digital
        group_config_map = {}
        for grp in all_groups:
            cmp = campus_map.get(grp.campus_id)
            if cmp:
                group_config_map[grp.id] = {
                    'reporte_evaluacion': grp.enable_tier_basic_override if grp.enable_tier_basic_override is not None else (cmp.enable_tier_basic if cmp.enable_tier_basic is not None else True),
                    'certificado_eduit': grp.enable_tier_standard_override if grp.enable_tier_standard_override is not None else (cmp.enable_tier_standard if cmp.enable_tier_standard is not None else False),
                    'certificado_conocer': grp.enable_tier_advanced_override if grp.enable_tier_advanced_override is not None else (cmp.enable_tier_advanced if cmp.enable_tier_advanced is not None else False),
                    'insignia_digital': grp.enable_digital_badge_override if grp.enable_digital_badge_override is not None else (cmp.enable_digital_badge if cmp.enable_digital_badge is not None else False),
                }
            else:
                group_config_map[grp.id] = {
                    'reporte_evaluacion': grp.enable_tier_basic_override if grp.enable_tier_basic_override is not None else True,
                    'certificado_eduit': grp.enable_tier_standard_override if grp.enable_tier_standard_override is not None else False,
                    'certificado_conocer': grp.enable_tier_advanced_override if grp.enable_tier_advanced_override is not None else False,
                    'insignia_digital': grp.enable_digital_badge_override if grp.enable_digital_badge_override is not None else False,
                }
        
        def is_cert_type_enabled(cert_type, user_id, result_group_id=None):
            """Verificar si un tipo de certificado está habilitado para el grupo del usuario"""
            # Usar group_id del resultado si disponible, sino del mapeo de miembros
            gid = result_group_id or user_group_map.get(user_id)
            if not gid or gid not in group_config_map:
                # Sin grupo identificado, mostrar por defecto (compatibilidad)
                return True
            return group_config_map[gid].get(cert_type, False)
        
        # Construir lista de certificados según tipo
        certificates = []
        
        if not cert_type_filter or cert_type_filter == 'reporte_evaluacion':
            results_q = Result.query.filter(
                Result.user_id.in_(candidate_ids), Result.status == 1, Result.result == 1,
                Result.report_url.isnot(None)
            )
            if search:
                # Buscar por nombre de usuario
                matching_users = User.query.filter(
                    User.id.in_(candidate_ids),
                    db.or_(
                        User.name.ilike(f'%{search}%'),
                        User.first_surname.ilike(f'%{search}%'),
                        User.curp.ilike(f'%{search}%')
                    )
                ).with_entities(User.id).all()
                matching_ids = [u[0] for u in matching_users]
                if matching_ids:
                    results_q = results_q.filter(Result.user_id.in_(matching_ids))
                else:
                    results_q = results_q.filter(False)  # No results
            
            for r in results_q.all():
                if not is_cert_type_enabled('reporte_evaluacion', r.user_id, r.group_id):
                    continue
                u = User.query.get(r.user_id)
                gid = r.group_id or user_group_map.get(r.user_id)
                grp = group_map.get(gid) if gid else None
                cmp = campus_map.get(grp.campus_id) if grp else None
                certificates.append({
                    'id': f'report_{r.id}',
                    'cert_type': 'reporte_evaluacion',
                    'cert_type_label': 'Reporte de Evaluación',
                    'user_id': r.user_id,
                    'user_name': u.full_name if u else 'N/A',
                    'user_curp': getattr(u, 'curp', '') if u else '',
                    'score': float(r.score) if r.score else 0,
                    'date': r.end_date.isoformat() if r.end_date else None,
                    'code': r.certificate_code or '',
                    'download_url': r.report_url,
                    'group_name': grp.name if grp else 'N/A',
                    'group_id': gid,
                    'campus_name': cmp.name if cmp else 'N/A',
                    'campus_id': cmp.id if cmp else None,
                    'state': cmp.state_name if cmp else '',
                    'status': 'generado'
                })
        
        if not cert_type_filter or cert_type_filter == 'certificado_eduit':
            results_q = Result.query.filter(
                Result.user_id.in_(candidate_ids), Result.status == 1, Result.result == 1,
                Result.certificate_url.isnot(None)
            )
            if search:
                matching_users = User.query.filter(
                    User.id.in_(candidate_ids),
                    db.or_(
                        User.name.ilike(f'%{search}%'),
                        User.first_surname.ilike(f'%{search}%'),
                        User.curp.ilike(f'%{search}%')
                    )
                ).with_entities(User.id).all()
                matching_ids = [u[0] for u in matching_users]
                if matching_ids:
                    results_q = results_q.filter(Result.user_id.in_(matching_ids))
                else:
                    results_q = results_q.filter(False)
            
            for r in results_q.all():
                if not is_cert_type_enabled('certificado_eduit', r.user_id, r.group_id):
                    continue
                u = User.query.get(r.user_id)
                gid = r.group_id or user_group_map.get(r.user_id)
                grp = group_map.get(gid) if gid else None
                cmp = campus_map.get(grp.campus_id) if grp else None
                certificates.append({
                    'id': f'eduit_{r.id}',
                    'cert_type': 'certificado_eduit',
                    'cert_type_label': 'Certificado Eduit',
                    'user_id': r.user_id,
                    'user_name': u.full_name if u else 'N/A',
                    'user_curp': getattr(u, 'curp', '') if u else '',
                    'score': float(r.score) if r.score else 0,
                    'date': r.end_date.isoformat() if r.end_date else None,
                    'code': r.eduit_certificate_code or '',
                    'download_url': r.certificate_url,
                    'group_name': grp.name if grp else 'N/A',
                    'group_id': gid,
                    'campus_name': cmp.name if cmp else 'N/A',
                    'campus_id': cmp.id if cmp else None,
                    'state': cmp.state_name if cmp else '',
                    'status': 'generado'
                })
        
        if not cert_type_filter or cert_type_filter == 'certificado_conocer':
            conocer_q = ConocerCertificate.query.filter(
                ConocerCertificate.user_id.in_(candidate_ids),
                ConocerCertificate.status == 'active'
            )
            if search:
                matching_users = User.query.filter(
                    User.id.in_(candidate_ids),
                    db.or_(
                        User.name.ilike(f'%{search}%'),
                        User.first_surname.ilike(f'%{search}%'),
                        User.curp.ilike(f'%{search}%')
                    )
                ).with_entities(User.id).all()
                matching_ids = [u[0] for u in matching_users]
                if matching_ids:
                    conocer_q = conocer_q.filter(ConocerCertificate.user_id.in_(matching_ids))
                else:
                    conocer_q = conocer_q.filter(False)
            
            for cc in conocer_q.all():
                if not is_cert_type_enabled('certificado_conocer', cc.user_id):
                    continue
                u = User.query.get(cc.user_id)
                gid = user_group_map.get(cc.user_id)
                grp = group_map.get(gid) if gid else None
                cmp = campus_map.get(grp.campus_id) if grp else None
                certificates.append({
                    'id': f'conocer_{cc.id}',
                    'cert_type': 'certificado_conocer',
                    'cert_type_label': 'Certificado CONOCER',
                    'user_id': cc.user_id,
                    'user_name': u.full_name if u else 'N/A',
                    'user_curp': cc.curp or (getattr(u, 'curp', '') if u else ''),
                    'score': 0,
                    'date': cc.issue_date.isoformat() if cc.issue_date else None,
                    'code': cc.certificate_number or '',
                    'download_url': None,  # CONOCER uses blob download endpoint
                    'conocer_id': cc.id,
                    'standard_code': cc.standard_code,
                    'standard_name': cc.standard_name,
                    'group_name': grp.name if grp else 'N/A',
                    'group_id': gid,
                    'campus_name': cmp.name if cmp else 'N/A',
                    'campus_id': cmp.id if cmp else None,
                    'state': cmp.state_name if cmp else '',
                    'status': cc.status
                })
        
        if not cert_type_filter or cert_type_filter == 'insignia_digital':
            # Insignias digitales: usuarios habilitados con resultado aprobado
            badge_results = db.session.query(Result, User).join(
                User, Result.user_id == User.id
            ).filter(
                Result.user_id.in_(candidate_ids), Result.status == 1, Result.result == 1,
                User.enable_digital_badge == True
            ).all()
            if search:
                badge_results = [
                    (r, u) for r, u in badge_results
                    if search.lower() in (u.full_name or '').lower() or search.lower() in (u.curp or '').lower()
                ]
            for r, u in badge_results:
                if not is_cert_type_enabled('insignia_digital', r.user_id, r.group_id):
                    continue
                gid = r.group_id or user_group_map.get(r.user_id)
                grp = group_map.get(gid) if gid else None
                cmp = campus_map.get(grp.campus_id) if grp else None
                certificates.append({
                    'id': f'badge_{r.id}',
                    'cert_type': 'insignia_digital',
                    'cert_type_label': 'Insignia Digital',
                    'user_id': r.user_id,
                    'user_name': u.full_name if u else 'N/A',
                    'user_curp': getattr(u, 'curp', '') if u else '',
                    'score': float(r.score) if r.score else 0,
                    'date': r.end_date.isoformat() if r.end_date else None,
                    'code': '',
                    'download_url': None,
                    'group_name': grp.name if grp else 'N/A',
                    'group_id': gid,
                    'campus_name': cmp.name if cmp else 'N/A',
                    'campus_id': cmp.id if cmp else None,
                    'state': cmp.state_name if cmp else '',
                    'status': 'habilitado'
                })
        
        # Ordenar por fecha descendente
        certificates.sort(key=lambda x: x.get('date') or '', reverse=True)
        
        # Paginación manual
        total = len(certificates)
        pages = (total + per_page - 1) // per_page
        start = (page - 1) * per_page
        end = start + per_page
        paginated = certificates[start:end]
        
        return jsonify({
            'certificates': paginated,
            'pagination': {'total': total, 'page': page, 'per_page': per_page, 'pages': pages},
            'filters': _get_partner_cert_filters(partner.id)
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


def _get_partner_cert_filters(partner_id):
    """Helper: obtener opciones de filtro disponibles para certificados del partner"""
    campuses = Campus.query.filter_by(partner_id=partner_id).all()
    campus_ids = [c.id for c in campuses]
    states = sorted(set([c.state_name for c in campuses if c.state_name]))
    
    groups = []
    if campus_ids:
        groups = CandidateGroup.query.filter(CandidateGroup.campus_id.in_(campus_ids)).order_by(CandidateGroup.name).all()
    
    return {
        'states': states,
        'campuses': [{'id': c.id, 'name': c.name, 'state': c.state_name or ''} for c in campuses],
        'groups': [{'id': g.id, 'name': g.name, 'campus_id': g.campus_id} for g in groups],
        'cert_types': [
            {'value': 'reporte_evaluacion', 'label': 'Reporte de Evaluación'},
            {'value': 'certificado_eduit', 'label': 'Certificado Eduit'},
            {'value': 'certificado_conocer', 'label': 'Certificado CONOCER'},
            {'value': 'insignia_digital', 'label': 'Insignia Digital'}
        ]
    }


@bp.route('/mi-partner/certificates/export', methods=['GET'])
@jwt_required()
@responsable_partner_required
def export_mi_partner_certificates_excel():
    """Exportar certificados del partner a Excel"""
    from app.models import Result, Exam
    from app.models.conocer_certificate import ConocerCertificate
    from openpyxl.utils import get_column_letter
    
    try:
        partner = g.partner
        state_filter = request.args.get('state', '')
        campus_filter = request.args.get('campus_id', '')
        group_filter = request.args.get('group_id', '')
        cert_type_filter = request.args.get('cert_type', '')
        search = request.args.get('search', '').strip()
        
        # Obtener campuses
        campus_query = Campus.query.filter_by(partner_id=partner.id)
        if state_filter:
            campus_query = campus_query.filter_by(state_name=state_filter)
        if campus_filter:
            campus_query = campus_query.filter_by(id=int(campus_filter))
        campuses = campus_query.all()
        campus_ids = [c.id for c in campuses]
        campus_map = {c.id: c for c in campuses}
        
        # Obtener grupos y miembros
        group_query = CandidateGroup.query.filter(CandidateGroup.campus_id.in_(campus_ids))
        if group_filter:
            group_query = group_query.filter_by(id=int(group_filter))
        all_groups = group_query.all()
        group_ids = [gr.id for gr in all_groups]
        group_map = {gr.id: gr for gr in all_groups}
        
        member_rows = db.session.query(GroupMember.user_id, GroupMember.group_id).filter(
            GroupMember.group_id.in_(group_ids), GroupMember.status == 'active'
        ).all() if group_ids else []
        
        user_group_map = {}
        for uid, gid in member_rows:
            if uid not in user_group_map:
                user_group_map[uid] = gid
        candidate_ids = list(user_group_map.keys())
        
        # Recolectar certificados (misma lógica pero sin paginación)
        rows = []
        
        if not cert_type_filter or cert_type_filter in ['reporte_evaluacion', 'certificado_eduit']:
            search_uids = None
            if search and candidate_ids:
                matching = User.query.filter(
                    User.id.in_(candidate_ids),
                    db.or_(User.name.ilike(f'%{search}%'), User.first_surname.ilike(f'%{search}%'), User.curp.ilike(f'%{search}%'))
                ).with_entities(User.id).all()
                search_uids = [u[0] for u in matching]
            
            base_q = Result.query.filter(Result.user_id.in_(candidate_ids), Result.status == 1, Result.result == 1) if candidate_ids else Result.query.filter(False)
            if search_uids is not None:
                base_q = base_q.filter(Result.user_id.in_(search_uids)) if search_uids else base_q.filter(False)
            
            if not cert_type_filter or cert_type_filter == 'reporte_evaluacion':
                for r in base_q.filter(Result.report_url.isnot(None)).all():
                    u = User.query.get(r.user_id)
                    gid = user_group_map.get(r.user_id)
                    grp = group_map.get(gid) if gid else None
                    cmp = campus_map.get(grp.campus_id) if grp else None
                    rows.append([
                        'Reporte de Evaluación', u.full_name if u else '', getattr(u, 'curp', '') if u else '',
                        float(r.score) if r.score else 0, r.certificate_code or '',
                        r.end_date.strftime('%Y-%m-%d') if r.end_date else '',
                        grp.name if grp else '', cmp.name if cmp else '', cmp.state_name if cmp else '', 'Generado'
                    ])
            
            if not cert_type_filter or cert_type_filter == 'certificado_eduit':
                for r in base_q.filter(Result.certificate_url.isnot(None)).all():
                    u = User.query.get(r.user_id)
                    gid = user_group_map.get(r.user_id)
                    grp = group_map.get(gid) if gid else None
                    cmp = campus_map.get(grp.campus_id) if grp else None
                    rows.append([
                        'Certificado Eduit', u.full_name if u else '', getattr(u, 'curp', '') if u else '',
                        float(r.score) if r.score else 0, r.eduit_certificate_code or '',
                        r.end_date.strftime('%Y-%m-%d') if r.end_date else '',
                        grp.name if grp else '', cmp.name if cmp else '', cmp.state_name if cmp else '', 'Generado'
                    ])
        
        if not cert_type_filter or cert_type_filter == 'certificado_conocer':
            conocer_q = ConocerCertificate.query.filter(
                ConocerCertificate.user_id.in_(candidate_ids), ConocerCertificate.status == 'active'
            ) if candidate_ids else ConocerCertificate.query.filter(False)
            if search and candidate_ids:
                matching = User.query.filter(
                    User.id.in_(candidate_ids),
                    db.or_(User.name.ilike(f'%{search}%'), User.first_surname.ilike(f'%{search}%'), User.curp.ilike(f'%{search}%'))
                ).with_entities(User.id).all()
                matching_ids = [u[0] for u in matching]
                conocer_q = conocer_q.filter(ConocerCertificate.user_id.in_(matching_ids)) if matching_ids else conocer_q.filter(False)
            for cc in conocer_q.all():
                u = User.query.get(cc.user_id)
                gid = user_group_map.get(cc.user_id)
                grp = group_map.get(gid) if gid else None
                cmp = campus_map.get(grp.campus_id) if grp else None
                rows.append([
                    'Certificado CONOCER', u.full_name if u else '', cc.curp or '',
                    0, cc.certificate_number or '',
                    cc.issue_date.strftime('%Y-%m-%d') if cc.issue_date else '',
                    grp.name if grp else '', cmp.name if cmp else '', cmp.state_name if cmp else '', cc.status
                ])
        
        if not cert_type_filter or cert_type_filter == 'insignia_digital':
            if candidate_ids:
                badge_results = db.session.query(Result, User).join(
                    User, Result.user_id == User.id
                ).filter(
                    Result.user_id.in_(candidate_ids), Result.status == 1, Result.result == 1,
                    User.enable_digital_badge == True
                ).all()
                if search:
                    badge_results = [(r, u) for r, u in badge_results if search.lower() in (u.full_name or '').lower() or search.lower() in (u.curp or '').lower()]
                for r, u in badge_results:
                    gid = user_group_map.get(r.user_id)
                    grp = group_map.get(gid) if gid else None
                    cmp = campus_map.get(grp.campus_id) if grp else None
                    rows.append([
                        'Insignia Digital', u.full_name if u else '', getattr(u, 'curp', '') if u else '',
                        float(r.score) if r.score else 0, '',
                        r.end_date.strftime('%Y-%m-%d') if r.end_date else '',
                        grp.name if grp else '', cmp.name if cmp else '', cmp.state_name if cmp else '', 'Habilitado'
                    ])
        
        # Crear Excel
        wb = Workbook()
        ws = wb.active
        ws.title = 'Certificados'
        
        # Título
        ws.merge_cells('A1:J1')
        ws['A1'] = f'Certificados - {partner.name}'
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A2:J2')
        ws['A2'] = f'Generado: {datetime.now().strftime("%Y-%m-%d %H:%M")}'
        ws['A2'].font = Font(size=10, color='666666')
        
        # Headers
        headers = ['Tipo', 'Nombre', 'CURP', 'Calificación', 'Código', 'Fecha', 'Grupo', 'Plantel', 'Estado', 'Estatus']
        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
        
        # Datos
        for row_idx, row_data in enumerate(rows, 5):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Anchos
        col_widths = [22, 30, 20, 14, 18, 14, 25, 25, 20, 14]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        
        ws.auto_filter.ref = f"A4:J{4 + len(rows)}"
        ws.freeze_panes = 'A5'
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"Certificados_{partner.name.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename,
        )
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@bp.route('/mi-partner/certificates/download-zip', methods=['GET'])
@jwt_required()
@responsable_partner_required
def download_mi_partner_certificates_zip():
    """Descargar todos los certificados filtrados como ZIP organizado por candidato"""
    from app.models import Result, Exam
    from app.models.conocer_certificate import ConocerCertificate
    from app.services.conocer_blob_service import get_conocer_blob_service
    import zipfile
    import re
    import urllib.request
    import urllib.error
    
    try:
        partner = g.partner
        state_filter = request.args.get('state', '')
        campus_filter = request.args.get('campus_id', '')
        group_filter = request.args.get('group_id', '')
        cert_type_filter = request.args.get('cert_type', '')
        search = request.args.get('search', '').strip()
        
        # Obtener campuses
        campus_query = Campus.query.filter_by(partner_id=partner.id)
        if state_filter:
            campus_query = campus_query.filter_by(state_name=state_filter)
        if campus_filter:
            campus_query = campus_query.filter_by(id=int(campus_filter))
        campuses = campus_query.all()
        campus_ids = [c.id for c in campuses]
        campus_map = {c.id: c for c in campuses}
        
        # Obtener grupos y miembros
        group_query = CandidateGroup.query.filter(CandidateGroup.campus_id.in_(campus_ids))
        if group_filter:
            group_query = group_query.filter_by(id=int(group_filter))
        all_groups = group_query.all()
        group_ids = [gr.id for gr in all_groups]
        group_map = {gr.id: gr for gr in all_groups}
        
        member_rows = db.session.query(GroupMember.user_id, GroupMember.group_id).filter(
            GroupMember.group_id.in_(group_ids), GroupMember.status == 'active'
        ).all() if group_ids else []
        
        user_group_map = {}
        for uid, gid in member_rows:
            if uid not in user_group_map:
                user_group_map[uid] = gid
        candidate_ids = list(user_group_map.keys())
        
        if not candidate_ids:
            return jsonify({'error': 'No se encontraron candidatos con los filtros seleccionados'}), 404
        
        # Helper: limpiar nombre para filesystem
        def safe_name(name):
            name = re.sub(r'[<>:"/\\|?*]', '_', name or 'Sin_Nombre')
            return name.strip().replace('  ', ' ')[:80]
        
        # Recolectar archivos para el ZIP: lista de (folder, filename, get_content_fn)
        zip_entries = []
        errors = []
        
        # Búsqueda común
        search_uids = None
        if search and candidate_ids:
            matching = User.query.filter(
                User.id.in_(candidate_ids),
                db.or_(User.name.ilike(f'%{search}%'), User.first_surname.ilike(f'%{search}%'), User.curp.ilike(f'%{search}%'))
            ).with_entities(User.id).all()
            search_uids = [u[0] for u in matching]
        
        # Cache de usuarios
        user_cache = {}
        def get_user(uid):
            if uid not in user_cache:
                user_cache[uid] = User.query.get(uid)
            return user_cache[uid]
        
        def make_folder(u):
            name = safe_name(u.full_name if u else 'Sin_Nombre')
            curp = (getattr(u, 'curp', '') or '').strip()
            return f"{name}_{curp}" if curp else name
        
        # --- Reportes de Evaluación ---
        if not cert_type_filter or cert_type_filter == 'reporte_evaluacion':
            base_q = Result.query.filter(
                Result.user_id.in_(candidate_ids), Result.status == 1, Result.result == 1,
                Result.report_url.isnot(None)
            ) if candidate_ids else Result.query.filter(False)
            if search_uids is not None:
                base_q = base_q.filter(Result.user_id.in_(search_uids)) if search_uids else base_q.filter(False)
            
            for r in base_q.all():
                u = get_user(r.user_id)
                folder = make_folder(u)
                code = r.certificate_code or str(r.id)
                filename = f"Reporte_Evaluacion_{code}.pdf"
                url = r.report_url
                zip_entries.append((folder, filename, 'url', url))
        
        # --- Certificados Eduit ---
        if not cert_type_filter or cert_type_filter == 'certificado_eduit':
            base_q = Result.query.filter(
                Result.user_id.in_(candidate_ids), Result.status == 1, Result.result == 1,
                Result.certificate_url.isnot(None)
            ) if candidate_ids else Result.query.filter(False)
            if search_uids is not None:
                base_q = base_q.filter(Result.user_id.in_(search_uids)) if search_uids else base_q.filter(False)
            
            for r in base_q.all():
                u = get_user(r.user_id)
                folder = make_folder(u)
                code = r.eduit_certificate_code or str(r.id)
                filename = f"Certificado_Eduit_{code}.pdf"
                url = r.certificate_url
                zip_entries.append((folder, filename, 'url', url))
        
        # --- Certificados CONOCER ---
        if not cert_type_filter or cert_type_filter == 'certificado_conocer':
            conocer_q = ConocerCertificate.query.filter(
                ConocerCertificate.user_id.in_(candidate_ids), ConocerCertificate.status == 'active'
            ) if candidate_ids else ConocerCertificate.query.filter(False)
            if search_uids is not None:
                conocer_q = conocer_q.filter(ConocerCertificate.user_id.in_(search_uids)) if search_uids else conocer_q.filter(False)
            
            for cc in conocer_q.all():
                if cc.blob_name:
                    u = get_user(cc.user_id)
                    folder = make_folder(u)
                    code = cc.certificate_number or str(cc.id)
                    std = cc.standard_code or ''
                    filename = f"CONOCER_{std}_{code}.pdf"
                    zip_entries.append((folder, filename, 'blob', cc.blob_name))
        
        if not zip_entries:
            return jsonify({'error': 'No se encontraron certificados descargables con los filtros seleccionados'}), 404
        
        # Crear ZIP en memoria
        zip_buffer = BytesIO()
        blob_service = None
        
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
            added_paths = set()
            for folder, filename, source_type, source in zip_entries:
                path = f"{folder}/{filename}"
                # Evitar duplicados
                if path in added_paths:
                    continue
                added_paths.add(path)
                
                try:
                    if source_type == 'url':
                        req = urllib.request.Request(source, headers={'User-Agent': 'Mozilla/5.0'})
                        with urllib.request.urlopen(req, timeout=30) as resp:
                            content = resp.read()
                    elif source_type == 'blob':
                        if blob_service is None:
                            blob_service = get_conocer_blob_service()
                        content, _ = blob_service.download_certificate(source)
                    else:
                        continue
                    
                    zf.writestr(path, content)
                except Exception as e:
                    errors.append(f"{path}: {str(e)[:100]}")
                    continue
        
        zip_buffer.seek(0)
        
        if errors:
            from flask import current_app
            current_app.logger.warning(f"ZIP download: {len(errors)} errores de {len(zip_entries)} archivos: {errors[:5]}")
        
        filename = f"Certificados_{safe_name(partner.name)}_{datetime.now().strftime('%Y%m%d_%H%M')}.zip"
        
        return send_file(
            zip_buffer,
            mimetype='application/zip',
            as_attachment=True,
            download_name=filename,
        )
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


# ═══════════════════════════════════════════════════════════════
# MÓDULO: Trámites CONOCER
# Candidatos aprobados con CONOCER habilitado que aún no tienen
# certificado CONOCER emitido. Permite exportar Excel para trámite.
# ═══════════════════════════════════════════════════════════════

@bp.route('/conocer-tramites', methods=['GET'])
@jwt_required()
@coordinator_required
def get_conocer_tramites():
    """
    Candidatos elegibles para trámite de certificado CONOCER.
    
    Muestra candidatos que:
    1. Están en un grupo con enable_tier_advanced activo
    2. Tienen al menos un examen aprobado (result=1, status=1)
    3. Tienen CURP registrado
    4. NO tienen un ConocerCertificate activo para ese ECM
    
    Query params:
    - page, per_page, search, partner_id, campus_id, group_id, ecm_id
    - sort_by, sort_dir, conocer_status (all|pending|has_certificate)
    """
    try:
        from sqlalchemy import text

        user_id_jwt = get_jwt_identity()
        user = User.query.get(user_id_jwt)
        if not user or user.role not in ['admin', 'developer', 'coordinator']:
            return jsonify({'error': 'Acceso denegado'}), 403

        page = request.args.get('page', 1, type=int)
        per_page = min(request.args.get('per_page', 150, type=int), 1000)
        search = request.args.get('search', '').strip()
        partner_id = request.args.get('partner_id', type=int)
        campus_id = request.args.get('campus_id', type=int)
        group_id = request.args.get('group_id', type=int)
        ecm_id = request.args.get('ecm_id', type=int)
        sort_by = request.args.get('sort_by', 'name')
        sort_dir = request.args.get('sort_dir', 'asc')
        conocer_status = request.args.get('conocer_status', 'pending')

        params = {}
        where_parts = []

        if search:
            where_parts.append(
                "AND (LOWER(u.name + ' ' + COALESCE(u.first_surname, '') + ' ' + COALESCE(u.second_surname, '')) LIKE :search "
                "OR LOWER(u.email) LIKE :search OR LOWER(u.curp) LIKE :search)"
            )
            params['search'] = f'%{search.lower()}%'
        if partner_id:
            where_parts.append("AND p.id = :partner_id")
            params['partner_id'] = partner_id
        if campus_id:
            where_parts.append("AND c.id = :campus_id")
            params['campus_id'] = campus_id
        if group_id:
            where_parts.append("AND cg.id = :group_id")
            params['group_id'] = group_id
        if ecm_id:
            where_parts.append("AND cs.id = :ecm_id")
            params['ecm_id'] = ecm_id

        where_sql = '\n                '.join(where_parts)

        cte_sql = f"""
            WITH conocer_candidates AS (
                SELECT
                    u.id AS user_id,
                    CONCAT(u.name, ' ', COALESCE(u.first_surname, ''), ' ', COALESCE(u.second_surname, '')) AS full_name,
                    u.email,
                    u.curp,
                    u.username,
                    cs.id AS ecm_id,
                    cs.code AS ecm_code,
                    cs.name AS ecm_name,
                    e.id AS exam_id,
                    e.name AS exam_name,
                    r.score,
                    r.end_date AS exam_date,
                    r.id AS result_id,
                    cg.id AS group_id,
                    cg.name AS group_name,
                    cg.code AS group_code,
                    c.id AS campus_id,
                    c.name AS campus_name,
                    p.id AS partner_id,
                    p.name AS partner_name,
                    eca.assignment_number,
                    COALESCE(eca.tramite_status, 'pendiente') AS eca_tramite_status,
                    ROW_NUMBER() OVER (
                        PARTITION BY u.id, cs.id
                        ORDER BY r.score DESC, r.created_at DESC
                    ) AS rn
                FROM results r
                JOIN users u ON u.id = r.user_id
                JOIN exams e ON e.id = r.exam_id
                JOIN competency_standards cs ON cs.id = e.competency_standard_id
                JOIN group_members gm ON gm.user_id = u.id AND gm.status = 'active'
                JOIN candidate_groups cg ON cg.id = gm.group_id AND cg.is_active = 1
                LEFT JOIN campuses c ON c.id = cg.campus_id
                LEFT JOIN partners p ON p.id = c.partner_id
                LEFT JOIN ecm_candidate_assignments eca ON eca.user_id = u.id AND eca.competency_standard_id = cs.id
                WHERE r.status = 1
                  AND r.result = 1
                  AND u.curp IS NOT NULL
                  AND LEN(u.curp) >= 10
                  AND u.enable_conocer_certificate = 1
                  AND (
                      (cg.enable_tier_advanced_override = 1)
                      OR (cg.enable_tier_advanced_override IS NULL AND c.enable_tier_advanced = 1)
                  )
                  AND EXISTS (
                      SELECT 1 FROM group_exams ge
                      WHERE ge.group_id = cg.id AND ge.exam_id = e.id
                  )
                  {where_sql}
            ),
            with_conocer_status AS (
                SELECT
                    cc.*,
                    kc.id AS conocer_cert_id,
                    kc.certificate_number AS conocer_cert_number,
                    kc.status AS conocer_cert_status,
                    kc.issue_date AS conocer_issue_date
                FROM conocer_candidates cc
                LEFT JOIN conocer_certificates kc
                    ON kc.user_id = cc.user_id
                   AND kc.standard_code = cc.ecm_code
                   AND kc.status = 'active'
                WHERE cc.rn = 1
            )
        """

        if conocer_status == 'pending':
            status_where = "WHERE wcs.conocer_cert_id IS NULL AND wcs.eca_tramite_status = 'pendiente'"
        elif conocer_status == 'en_tramite':
            status_where = "WHERE wcs.eca_tramite_status = 'en_tramite'"
        elif conocer_status == 'has_certificate':
            status_where = "WHERE wcs.conocer_cert_id IS NOT NULL"
        elif conocer_status == 'entregado':
            status_where = "WHERE wcs.eca_tramite_status = 'entregado'"
        else:
            status_where = "WHERE 1=1"

        sort_map = {
            'name': 'wcs.full_name', 'curp': 'wcs.curp', 'email': 'wcs.email',
            'group': 'wcs.group_name', 'campus': 'wcs.campus_name',
            'partner': 'wcs.partner_name', 'ecm': 'wcs.ecm_code',
            'score': 'wcs.score', 'exam_date': 'wcs.exam_date',
        }
        sort_col = sort_map.get(sort_by, 'wcs.full_name')
        sort_direction = 'DESC' if sort_dir == 'desc' else 'ASC'

        count_sql = text(cte_sql + f"""
            SELECT COUNT(*) AS total FROM with_conocer_status wcs {status_where}
        """)
        total = db.session.execute(count_sql, params).scalar() or 0
        total_pages = max(1, (total + per_page - 1) // per_page)

        offset = (page - 1) * per_page
        data_sql = text(cte_sql + f"""
            SELECT
                wcs.user_id, wcs.full_name, wcs.email, wcs.curp, wcs.username,
                wcs.ecm_id, wcs.ecm_code, wcs.ecm_name,
                wcs.exam_id, wcs.exam_name, wcs.score, wcs.exam_date, wcs.result_id,
                wcs.group_id, wcs.group_name, wcs.group_code,
                wcs.campus_id, wcs.campus_name,
                wcs.partner_id, wcs.partner_name,
                wcs.assignment_number,
                wcs.eca_tramite_status,
                wcs.conocer_cert_id, wcs.conocer_cert_number,
                wcs.conocer_cert_status, wcs.conocer_issue_date
            FROM with_conocer_status wcs
            {status_where}
            ORDER BY {sort_col} {sort_direction}
            OFFSET :offset ROWS FETCH NEXT :limit ROWS ONLY
        """)
        params['offset'] = offset
        params['limit'] = per_page

        rows = db.session.execute(data_sql, params).fetchall()

        summary_sql = text(cte_sql + """
            SELECT
                COUNT(*) AS total_candidates,
                SUM(CASE WHEN wcs.conocer_cert_id IS NULL AND wcs.eca_tramite_status = 'pendiente' THEN 1 ELSE 0 END) AS pending,
                SUM(CASE WHEN wcs.eca_tramite_status = 'en_tramite' THEN 1 ELSE 0 END) AS en_tramite,
                SUM(CASE WHEN wcs.conocer_cert_id IS NOT NULL THEN 1 ELSE 0 END) AS with_certificate,
                COUNT(DISTINCT wcs.ecm_code) AS total_ecms,
                COUNT(DISTINCT wcs.group_id) AS total_groups,
                COUNT(DISTINCT wcs.campus_id) AS total_campuses,
                COUNT(DISTINCT wcs.partner_id) AS total_partners
            FROM with_conocer_status wcs WHERE 1=1
        """)
        summary_row = db.session.execute(summary_sql, params).fetchone()

        candidates = []
        for row in rows:
            candidates.append({
                'user_id': row.user_id,
                'full_name': row.full_name.strip() if row.full_name else '',
                'email': row.email,
                'curp': row.curp,
                'username': row.username,
                'ecm_id': row.ecm_id,
                'ecm_code': row.ecm_code,
                'ecm_name': row.ecm_name,
                'exam_id': row.exam_id,
                'exam_name': row.exam_name,
                'score': row.score,
                'exam_date': row.exam_date.isoformat() + 'Z' if row.exam_date else None,
                'result_id': row.result_id,
                'group_id': row.group_id,
                'group_name': row.group_name,
                'group_code': row.group_code,
                'campus_id': row.campus_id,
                'campus_name': row.campus_name,
                'partner_id': row.partner_id,
                'partner_name': row.partner_name,
                'assignment_number': row.assignment_number,
                'conocer_cert_id': row.conocer_cert_id,
                'conocer_cert_number': row.conocer_cert_number,
                'conocer_cert_status': row.conocer_cert_status,
                'conocer_issue_date': row.conocer_issue_date.isoformat() if row.conocer_issue_date else None,
                'tramite_status': 'certificado' if row.conocer_cert_id else (row.eca_tramite_status or 'pendiente'),
            })

        return jsonify({
            'candidates': candidates,
            'total': total,
            'pages': total_pages,
            'current_page': page,
            'per_page': per_page,
            'summary': {
                'total_candidates': summary_row.total_candidates if summary_row else 0,
                'pending': summary_row.pending if summary_row else 0,
                'en_tramite': summary_row.en_tramite if summary_row else 0,
                'with_certificate': summary_row.with_certificate if summary_row else 0,
                'total_ecms': summary_row.total_ecms if summary_row else 0,
                'total_groups': summary_row.total_groups if summary_row else 0,
                'total_campuses': summary_row.total_campuses if summary_row else 0,
                'total_partners': summary_row.total_partners if summary_row else 0,
            }
        })

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@bp.route('/conocer-tramites/export', methods=['GET'])
@jwt_required()
@coordinator_required
def export_conocer_tramites_excel():
    """
    Exportar candidatos certificados CONOCER a Excel en formato Válidos RENAPO.
    11 columnas exactas del formato oficial CONOCER.
    Solo incluye candidatos que ya tienen certificado activo.
    """
    try:
        from sqlalchemy import text
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        from io import BytesIO
        import unicodedata

        user_id_jwt = get_jwt_identity()
        user = User.query.get(user_id_jwt)
        if not user or user.role not in ['admin', 'developer', 'coordinator']:
            return jsonify({'error': 'Acceso denegado'}), 403

        # ── Helpers ──
        def strip_accents(s: str) -> str:
            """Remove diacritical marks (accents) from text for RENAPO compatibility."""
            if not s:
                return s
            nfkd = unicodedata.normalize('NFKD', s)
            return ''.join(c for c in nfkd if not unicodedata.category(c).startswith('M'))

        # ── Parámetros de filtro ──
        search = request.args.get('search', '').strip()
        partner_id = request.args.get('partner_id', type=int)
        campus_id = request.args.get('campus_id', type=int)
        ecm_id = request.args.get('ecm_id', type=int)
        user_ids_param = request.args.get('user_ids', '').strip()

        params = {}
        where_parts = []

        if search:
            where_parts.append(
                "AND (LOWER(u.name + ' ' + COALESCE(u.first_surname, '') + ' ' + COALESCE(u.second_surname, '')) LIKE :search "
                "OR LOWER(u.curp) LIKE :search)"
            )
            params['search'] = f'%{search.lower()}%'
        if partner_id:
            where_parts.append("AND ci.partner_id = :partner_id")
            params['partner_id'] = partner_id
        if campus_id:
            where_parts.append("AND COALESCE(ci.campus_id, u.campus_id) = :campus_id")
            params['campus_id'] = campus_id
        if ecm_id:
            where_parts.append("AND cs.id = :ecm_id")
            params['ecm_id'] = ecm_id
        if user_ids_param:
            uid_list = [uid.strip() for uid in user_ids_param.split(',') if uid.strip()]
            if uid_list:
                placeholders = ','.join(f"'{uid}'" for uid in uid_list[:5000])
                where_parts.append(f"AND u.id IN ({placeholders})")

        where_sql = '\n                '.join(where_parts)

        sql = text(f"""
            SELECT
                u.curp,
                u.name AS first_name,
                u.first_surname,
                u.second_surname,
                u.gender,
                COALESCE(cs.code, kc.standard_code) AS ecm_code,
                COALESCE(cs.name, kc.standard_name) AS ecm_name,
                COALESCE(cs.level, TRY_CAST(kc.competency_level AS INT)) AS ecm_level,
                kc.issue_date AS cert_date,
                kc.certificate_number AS folio,
                eca.assignment_number,
                COALESCE(ci.state_name, uc.state_name) AS state_name,
                COALESCE(ci.country, uc.country) AS country
            FROM conocer_certificates kc
            JOIN users u ON u.id = kc.user_id
            LEFT JOIN competency_standards cs ON cs.code = kc.standard_code
            LEFT JOIN campuses uc ON uc.id = u.campus_id
            LEFT JOIN ecm_candidate_assignments eca
                ON eca.user_id = u.id AND eca.competency_standard_id = cs.id
            OUTER APPLY (
                SELECT TOP 1 c.state_name, c.country, c.id AS campus_id, p.id AS partner_id
                FROM group_members gm
                JOIN candidate_groups cg ON cg.id = gm.group_id
                JOIN campuses c ON c.id = cg.campus_id
                LEFT JOIN partners p ON p.id = c.partner_id
                WHERE gm.user_id = u.id AND gm.status = 'active'
                ORDER BY cg.created_at DESC
            ) ci
            WHERE kc.status = 'active'
                {where_sql}
            ORDER BY u.name, u.first_surname, u.second_surname
        """)

        rows = db.session.execute(sql, params).fetchall()

        # ── Valor fijo del Centro Evaluador ──
        CE_NAME = 'ENTRENAMIENTO INFORMATICO AVANZADO S A  DE C V '

        # Países considerados nacionales (para la regla de extranjeros)
        NATIONAL_COUNTRIES = {'México', 'Mexico', 'MEXICO', 'MÉXICO', 'mexico', 'méxico'}

        # ── Generar Excel en formato RENAPO ──
        wb = Workbook()
        ws = wb.active
        ws.title = 'Válidos'

        # Estilos de header (azul RENAPO #4472C4, texto blanco, centrado, wrap)
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

        headers_list = [
            'Núm.', 'CURP/NIP', 'Nombre(s)', 'Apellidos', 'Entidad Federativa',
            'Código del ECM', 'Título del ECM', 'Nivel de Competencia del ECM',
            'Fecha de Certificación', 'Folio Certificado Marca',
            'Nombre o Razón Social del CE/EI',
        ]
        for col, h in enumerate(headers_list, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align

        # ── Rellenar datos ──
        for idx, row in enumerate(rows, 2):
            num = idx - 1

            # Determinar si es candidato extranjero
            country = (row.country or '').strip()
            is_foreign = bool(country) and country not in NATIONAL_COUNTRIES

            # Col 2: CURP/NIP — extranjeros usan placeholder temporal
            if is_foreign:
                curp = 'A' if (row.gender or '').upper() == 'M' else 'B'
            else:
                curp = (row.curp or '').upper()

            # Col 3-4: Nombres y apellidos en MAYÚSCULAS sin acentos
            first_name = strip_accents((row.first_name or '').upper().strip())
            apellidos = strip_accents(
                f"{(row.first_surname or '').upper()} {(row.second_surname or '').upper()}".strip()
            )

            # Col 5: Entidad Federativa — extranjeros → Ciudad de Mexico
            if is_foreign:
                state = 'Ciudad de Mexico'
            else:
                state = strip_accents((row.state_name or 'Ciudad de Mexico').strip())

            # Col 6-7: ECM código y título sin acentos
            ecm_code = (row.ecm_code or '')
            ecm_name = strip_accents(row.ecm_name or '')

            # Col 8: Nivel de competencia (entero)
            ecm_level = row.ecm_level if row.ecm_level is not None else ''

            # Col 9: Fecha de certificación (formato YYYY-MM-DD como texto)
            if row.cert_date:
                cert_date = row.cert_date.strftime('%Y-%m-%d') if hasattr(row.cert_date, 'strftime') else str(row.cert_date)
            else:
                cert_date = ''

            # Col 10: Folio Certificado Marca → el número de asignación
            folio = row.assignment_number or row.folio or ''

            # ── Escribir fila ──
            # Núm. (int, format 0)
            c = ws.cell(row=idx, column=1, value=num)
            c.number_format = '0'

            # CURP/NIP .. Título del ECM (text, format @)
            for col_i, val in enumerate([curp, first_name, apellidos, state, ecm_code, ecm_name], 2):
                c = ws.cell(row=idx, column=col_i, value=val)
                c.number_format = '@'

            # Nivel de Competencia (int, format @)
            c = ws.cell(row=idx, column=8, value=ecm_level)
            c.number_format = '@'

            # Fecha de Certificación (texto con formato visual dd/mm/yyyy)
            c = ws.cell(row=idx, column=9, value=cert_date)
            c.number_format = 'dd/mm/yyyy'

            # Folio Certificado Marca (text @)
            c = ws.cell(row=idx, column=10, value=folio)
            c.number_format = '@'

            # CE/EI (text @, valor fijo)
            c = ws.cell(row=idx, column=11, value=CE_NAME)
            c.number_format = '@'

        # ── Anchos de columna (idénticos al template) ──
        widths = [6, 20, 27, 27, 21, 16, 50, 30, 24, 25, 49]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        # Congelar primera fila (headers siempre visibles)
        ws.freeze_panes = 'A2'

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"validos_renapo_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename,
        )

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

# ============== CONOCER EMAIL CONTACTS ==============

@bp.route('/conocer-contacts', methods=['GET'])
@jwt_required()
@coordinator_required
def get_conocer_contacts():
    """Obtener contactos de correo para solicitudes CONOCER."""
    from app.models.partner import ConocerEmailContact
    contacts = ConocerEmailContact.query.order_by(ConocerEmailContact.name).all()
    return jsonify({'contacts': [c.to_dict() for c in contacts]})


@bp.route('/conocer-contacts', methods=['POST'])
@jwt_required()
@coordinator_required
def create_conocer_contact():
    """Crear un nuevo contacto de correo CONOCER."""
    from app.models.partner import ConocerEmailContact
    data = request.get_json()
    if not data or not data.get('name') or not data.get('email'):
        return jsonify({'error': 'Nombre y email son requeridos'}), 400
    
    contact = ConocerEmailContact(
        name=data['name'].strip(),
        email=data['email'].strip().lower(),
        is_active=data.get('is_active', True),
        created_by_id=get_jwt_identity(),
    )
    db.session.add(contact)
    db.session.commit()
    return jsonify({'contact': contact.to_dict()}), 201


@bp.route('/conocer-contacts/<int:contact_id>', methods=['PUT'])
@jwt_required()
@coordinator_required
def update_conocer_contact(contact_id):
    """Actualizar un contacto de correo CONOCER."""
    from app.models.partner import ConocerEmailContact
    contact = ConocerEmailContact.query.get_or_404(contact_id)
    data = request.get_json()
    if not data:
        return jsonify({'error': 'Datos requeridos'}), 400
    
    if 'name' in data:
        contact.name = data['name'].strip()
    if 'email' in data:
        contact.email = data['email'].strip().lower()
    if 'is_active' in data:
        contact.is_active = data['is_active']
    contact.updated_at = datetime.utcnow()
    db.session.commit()
    return jsonify({'contact': contact.to_dict()})


@bp.route('/conocer-contacts/<int:contact_id>', methods=['DELETE'])
@jwt_required()
@coordinator_required
def delete_conocer_contact(contact_id):
    """Eliminar un contacto de correo CONOCER."""
    from app.models.partner import ConocerEmailContact
    contact = ConocerEmailContact.query.get_or_404(contact_id)
    db.session.delete(contact)
    db.session.commit()
    return jsonify({'message': 'Contacto eliminado'})


# ============== CONOCER SOLICITUD (SEND EMAIL) ==============

@bp.route('/conocer-tramites/send-solicitud', methods=['POST'])
@jwt_required()
@coordinator_required
def send_conocer_solicitud():
    """
    Enviar correo de solicitud de línea de captura a los contactos CONOCER.
    
    1. Obtiene candidatos con tramite_status='pendiente' que tienen assignment_number
    2. Genera tabla resumen por ECM
    3. Genera el Excel RENAPO como adjunto
    4. Descarga el PDF COSU del blob como adjunto
    5. Envía correo a todos los contactos activos
    6. Cambia tramite_status de 'pendiente' a 'en_tramite'
    """
    import json
    import base64
    from sqlalchemy import text
    from app.models.partner import ConocerEmailContact, ConocerSolicitudLog, EcmCandidateAssignment
    from app.services.email_service import send_email as send_email_fn
    from app.services.email_service import LOGO_URL

    try:
        # 1. Get active contacts
        contacts = ConocerEmailContact.query.filter_by(is_active=True).all()
        if not contacts:
            return jsonify({'error': 'No hay contactos activos configurados. Agrega contactos en la sección de configuración.'}), 400

        # 2. Get pending assignments with their ECM data
        pending_sql = text("""
            SELECT 
                eca.id AS eca_id,
                eca.assignment_number,
                eca.user_id,
                u.curp,
                u.name AS user_name,
                COALESCE(u.first_surname, '') AS first_surname,
                COALESCE(u.second_surname, '') AS second_surname,
                u.gender,
                COALESCE(uc.country, '') AS country,
                cs.id AS ecm_id,
                cs.code AS ecm_code,
                cs.name AS ecm_name,
                cs.level AS competency_level
            FROM ecm_candidate_assignments eca
            JOIN users u ON u.id = eca.user_id
            LEFT JOIN campuses uc ON uc.id = u.campus_id
            JOIN competency_standards cs ON cs.id = eca.competency_standard_id
            LEFT JOIN candidate_groups cg ON cg.id = eca.group_id
            LEFT JOIN campuses c ON c.id = eca.campus_id
            WHERE eca.tramite_status = 'pendiente'
              AND eca.assignment_number IS NOT NULL
              AND u.enable_conocer_certificate = 1
              AND (
                  (cg.enable_tier_advanced_override = 1)
                  OR (cg.enable_tier_advanced_override IS NULL AND c.enable_tier_advanced = 1)
              )
            ORDER BY cs.code, u.name
        """)
        pending_rows = db.session.execute(pending_sql).fetchall()

        if not pending_rows:
            return jsonify({'error': 'No hay trámites pendientes para enviar.'}), 400

        # 3. Build ECM summary table
        ecm_counts = {}
        eca_ids = []
        for row in pending_rows:
            eca_ids.append(row.eca_id)
            code = row.ecm_code
            if code not in ecm_counts:
                ecm_counts[code] = 0
            ecm_counts[code] += 1

        total_certs = sum(ecm_counts.values())
        ecm_summary = [{'code': code, 'count': count} for code, count in sorted(ecm_counts.items())]

        # 4. Generate RENAPO Excel attachment
        excel_base64, excel_filename = _generate_renapo_excel_for_pending(pending_rows)

        # 5. Download COSU PDF from blob
        cosu_base64 = _download_cosu_pdf_from_blob()

        # 6. Build email HTML
        table_rows_html = ''
        for item in ecm_summary:
            table_rows_html += f'''
                <tr>
                    <td style="padding:10px 16px;border:1px solid #d1d5db;font-size:14px;color:#374151;">{item['code']}</td>
                    <td style="padding:10px 16px;border:1px solid #d1d5db;font-size:14px;color:#374151;text-align:center;">{item['count']}</td>
                </tr>'''
        table_rows_html += f'''
                <tr style="background-color:#f3f4f6;font-weight:bold;">
                    <td style="padding:10px 16px;border:1px solid #d1d5db;font-size:14px;color:#111827;">Total</td>
                    <td style="padding:10px 16px;border:1px solid #d1d5db;font-size:14px;color:#111827;text-align:center;">{total_certs}</td>
                </tr>'''

        email_body = f'''
<div style="font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',Roboto,'Helvetica Neue',Arial,sans-serif;max-width:640px;margin:0 auto;background:#ffffff;border-radius:12px;overflow:hidden;box-shadow:0 4px 6px rgba(0,0,0,0.07);">
    <div style="background:linear-gradient(135deg,#1e40af 0%,#2563eb 100%);padding:28px 32px;text-align:center;">
        <img src="{LOGO_URL}" alt="Evaluaasi" style="height:40px;margin-bottom:8px;" />
        <h1 style="color:#ffffff;font-size:20px;margin:0;">Solicitud de L\u00ednea de Captura</h1>
    </div>
    <div style="padding:32px;">
        <p style="font-size:15px;color:#374151;line-height:1.7;margin:0 0 20px;">Buen d\u00eda estimados,</p>
        <p style="font-size:15px;color:#374151;line-height:1.7;margin:0 0 24px;">
            Por favor solicito la l\u00ednea de captura para <strong>{total_certs}</strong> certificados sin fotograf\u00eda con descuento.
        </p>
        <table style="width:100%;border-collapse:collapse;margin:0 0 24px;border-radius:8px;overflow:hidden;">
            <thead>
                <tr style="background-color:#2563eb;">
                    <th style="padding:12px 16px;color:#ffffff;font-size:14px;font-weight:600;text-align:left;border:1px solid #2563eb;">Est\u00e1ndar</th>
                    <th style="padding:12px 16px;color:#ffffff;font-size:14px;font-weight:600;text-align:center;border:1px solid #2563eb;">Cantidad</th>
                </tr>
            </thead>
            <tbody>
                {table_rows_html}
            </tbody>
        </table>
        <p style="font-size:15px;color:#374151;line-height:1.7;margin:0 0 8px;">Agradezco sus comentarios.</p>
        <p style="font-size:15px;color:#374151;line-height:1.7;margin:0;">Saludos.</p>
    </div>
    <div style="background:#f8fafc;padding:20px 32px;text-align:center;border-top:1px solid #e2e8f0;">
        <p style="font-size:12px;color:#94a3b8;margin:0;">Enviado desde Evaluaasi &bull; {datetime.now().strftime('%d/%m/%Y %H:%M')}</p>
    </div>
</div>'''

        # 7. Prepare attachments
        attachments = []
        if excel_base64:
            attachments.append({
                'name': excel_filename,
                'content_type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                'content_base64': excel_base64,
            })
        if cosu_base64:
            attachments.append({
                'name': 'COSU_229_2026_GRUPO_EDUIT.pdf',
                'content_type': 'application/pdf',
                'content_base64': cosu_base64,
            })

        # 8. Send email to all contacts
        recipient_emails = [c.email for c in contacts]
        primary_to = recipient_emails[0]
        cc_list = recipient_emails[1:] if len(recipient_emails) > 1 else None

        success = send_email_fn(
            to=primary_to,
            subject='[Evaluaasi] Solicitud de línea de captura',
            html=email_body,
            attachments=attachments,
            cc=cc_list,
        )

        if not success:
            log = ConocerSolicitudLog(
                sent_by_id=get_jwt_identity(),
                recipients=json.dumps(recipient_emails),
                total_certificates=total_certs,
                ecm_summary=json.dumps(ecm_summary),
                attachment_names=', '.join([a['name'] for a in attachments]),
                status='failed',
                error_message='Error al enviar el correo',
                assignment_ids=json.dumps(eca_ids),
            )
            db.session.add(log)
            db.session.commit()
            return jsonify({'error': 'Error al enviar el correo. Verifica la configuración de ACS.'}), 500

        # 9. Update tramite_status to 'en_tramite'
        if eca_ids:
            ids_str = ','.join(str(i) for i in eca_ids)
            db.session.execute(text(f"""
                UPDATE ecm_candidate_assignments 
                SET tramite_status = 'en_tramite'
                WHERE id IN ({ids_str}) AND tramite_status = 'pendiente'
            """))
            db.session.commit()

        # 10. Log the solicitud
        log = ConocerSolicitudLog(
            sent_by_id=get_jwt_identity(),
            recipients=json.dumps(recipient_emails),
            total_certificates=total_certs,
            ecm_summary=json.dumps(ecm_summary),
            attachment_names=', '.join([a['name'] for a in attachments]),
            status='sent',
            assignment_ids=json.dumps(eca_ids),
        )
        db.session.add(log)
        db.session.commit()

        return jsonify({
            'message': f'Solicitud enviada exitosamente a {len(recipient_emails)} contacto(s)',
            'total_certificates': total_certs,
            'ecm_summary': ecm_summary,
            'recipients': recipient_emails,
            'solicitud_id': log.id,
        })

    except Exception as e:
        import traceback
        traceback.print_exc()
        db.session.rollback()
        return jsonify({'error': str(e)}), 500


@bp.route('/conocer-solicitud-logs', methods=['GET'])
@jwt_required()
@coordinator_required
def get_conocer_solicitud_logs():
    """Obtener historial de solicitudes enviadas."""
    from app.models.partner import ConocerSolicitudLog
    logs = ConocerSolicitudLog.query.order_by(ConocerSolicitudLog.sent_at.desc()).limit(50).all()
    return jsonify({'logs': [l.to_dict() for l in logs]})


def _generate_renapo_excel_for_pending(pending_rows):
    """
    Genera el Excel en formato RENAPO para los candidatos pendientes.
    Returns (base64_string, filename).
    """
    import base64
    import unicodedata
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    def strip_accents(s):
        if not s:
            return s
        nfkd = unicodedata.normalize('NFKD', s)
        return ''.join(c for c in nfkd if not unicodedata.combining(c))

    wb = Workbook()
    ws = wb.active
    ws.title = 'Válidos'

    headers = [
        'Núm.', 'CURP/NIP de la Persona Certificada', 'Nombre(s)',
        'Apellidos', 'Entidad Federativa de Nacimiento',
        'Código del ECM', 'Título del ECM',
        'Nivel de Competencia del ECM',
        'Fecha de Certificación',
        'Folio Certificado Marca',
        'Nombre o Razón Social del CE / EI y/o del OC / OC-OPCION'
    ]

    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True, size=10)
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    col_widths = [6, 20, 27, 27, 21, 16, 50, 30, 24, 25, 49]

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = thin_border

    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[chr(64 + i) if i <= 26 else 'A' + chr(64 + i - 26)].width = w

    CURP_STATE_MAP = {
        'AS': 'Aguascalientes', 'BC': 'Baja California', 'BS': 'Baja California Sur',
        'CC': 'Campeche', 'CL': 'Coahuila de Zaragoza', 'CM': 'Colima',
        'CS': 'Chiapas', 'CH': 'Chihuahua', 'DF': 'Ciudad de Mexico',
        'DG': 'Durango', 'GT': 'Guanajuato', 'GR': 'Guerrero',
        'HG': 'Hidalgo', 'JC': 'Jalisco', 'MC': 'Mexico',
        'MN': 'Michoacan de Ocampo', 'MS': 'Morelos', 'NT': 'Nayarit',
        'NL': 'Nuevo Leon', 'OC': 'Oaxaca', 'PL': 'Puebla',
        'QT': 'Queretaro', 'QR': 'Quintana Roo', 'SP': 'San Luis Potosi',
        'SL': 'Sinaloa', 'SR': 'Sonora', 'TC': 'Tabasco',
        'TS': 'Tamaulipas', 'TL': 'Tlaxcala', 'VZ': 'Veracruz de Ignacio de la Llave',
        'YN': 'Yucatan', 'ZS': 'Zacatecas', 'NE': 'Nacido en el Extranjero',
    }

    ce_ei = 'ENTRENAMIENTO INFORMATICO AVANZADO S A  DE C V '
    cert_date = datetime.now().strftime('%Y-%m-%d')

    for idx, row in enumerate(pending_rows, 1):
        curp = (row.curp or '').strip().upper()
        is_foreigner = (row.country or '').strip().lower() not in ('', 'méxico', 'mexico', 'mx')
        if is_foreigner:
            curp_val = 'A' if (row.gender or '').upper() == 'M' else 'B'
            state = 'Ciudad de Mexico'
        else:
            curp_val = curp
            state_code = curp[11:13] if len(curp) >= 13 else ''
            state = strip_accents(CURP_STATE_MAP.get(state_code, ''))

        first = strip_accents((row.user_name or '').strip().upper())
        surnames = strip_accents(f"{row.first_surname} {row.second_surname}".strip().upper())
        ecm_title = strip_accents((row.ecm_name or '').upper())

        data_row = [
            idx,
            curp_val,
            first,
            surnames,
            state,
            (row.ecm_code or '').upper(),
            ecm_title,
            str(row.competency_level or ''),
            cert_date,
            row.assignment_number or '',
            ce_ei,
        ]

        for col_idx, val in enumerate(data_row, 1):
            cell = ws.cell(row=idx + 1, column=col_idx, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(vertical='center')

    ws.freeze_panes = 'A2'

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    excel_bytes = output.getvalue()

    filename = f"validos_renapo_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return base64.b64encode(excel_bytes).decode('utf-8'), filename


def _download_cosu_pdf_from_blob():
    """
    Download the COSU PDF from Azure Blob Storage.
    Returns base64-encoded content or None if not found.
    """
    import base64
    try:
        from azure.storage.blob import BlobServiceClient
        from config import Config
        conn_str = Config.AZURE_STORAGE_CONNECTION_STRING
        if not conn_str:
            print("[COSU PDF] No AZURE_STORAGE_CONNECTION_STRING configured")
            return None
        
        blob_service = BlobServiceClient.from_connection_string(conn_str)
        blob_client = blob_service.get_blob_client(
            container='conocer-certificates',
            blob='documentos/COSU_229_2026_GRUPO_EDUIT.pdf'
        )
        download = blob_client.download_blob()
        pdf_bytes = download.readall()
        return base64.b64encode(pdf_bytes).decode('utf-8')
    except Exception as e:
        print(f"[COSU PDF] Error downloading: {e}")
        return None