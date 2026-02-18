"""
Rutas para gestión de Certificados CONOCER
"""
from datetime import datetime
from flask import Blueprint, request, jsonify, send_file, current_app
from flask_jwt_extended import jwt_required, get_jwt_identity
from app import db
from app.models import User
from app.models.conocer_certificate import ConocerCertificate
from app.services.conocer_blob_service import get_conocer_blob_service
import io

# Lazy import para Azure (puede no estar instalado)
ResourceNotFoundError = None

def _get_azure_exceptions():
    global ResourceNotFoundError
    if ResourceNotFoundError is None:
        try:
            from azure.core.exceptions import ResourceNotFoundError as _ResourceNotFoundError
            ResourceNotFoundError = _ResourceNotFoundError
        except ImportError:
            # Crear una excepción dummy si azure no está instalado
            class _ResourceNotFoundError(Exception):
                pass
            ResourceNotFoundError = _ResourceNotFoundError
    return ResourceNotFoundError

conocer_bp = Blueprint('conocer', __name__)


@conocer_bp.route('/certificates', methods=['GET'])
@jwt_required()
def get_my_certificates():
    """
    Obtener todos los certificados CONOCER del usuario autenticado
    
    Query params:
        - status: Filtrar por estado (active, archived, revoked)
        - standard_code: Filtrar por código de estándar
    
    Returns:
        Lista de certificados con metadata
    """
    current_user_id = get_jwt_identity()
    
    # Obtener filtros
    status_filter = request.args.get('status')
    standard_filter = request.args.get('standard_code')
    
    # Query base
    query = ConocerCertificate.query.filter_by(user_id=current_user_id)
    
    if status_filter:
        query = query.filter_by(status=status_filter)
    
    if standard_filter:
        query = query.filter_by(standard_code=standard_filter)
    
    certificates = query.order_by(ConocerCertificate.issue_date.desc()).all()
    
    return jsonify({
        'certificates': [cert.to_dict() for cert in certificates],
        'total': len(certificates)
    })


@conocer_bp.route('/certificates/<int:certificate_id>', methods=['GET'])
@jwt_required()
def get_certificate(certificate_id):
    """
    Obtener detalle de un certificado específico
    
    Returns:
        Información detallada del certificado
    """
    current_user_id = get_jwt_identity()
    
    certificate = ConocerCertificate.query.filter_by(
        id=certificate_id,
        user_id=current_user_id
    ).first()
    
    if not certificate:
        return jsonify({'error': 'Certificado no encontrado'}), 404
    
    # Obtener estado del blob
    blob_status = None
    try:
        blob_service = get_conocer_blob_service()
        blob_status = blob_service.get_blob_status(certificate.blob_name)
    except Exception as e:
        current_app.logger.error(f"Error obteniendo estado del blob: {e}")
    
    response = certificate.to_dict(include_blob_info=True)
    response['blob_status'] = blob_status
    
    return jsonify(response)


@conocer_bp.route('/certificates/<int:certificate_id>/download', methods=['GET'])
@jwt_required()
def download_certificate(certificate_id):
    """
    Descargar el PDF del certificado
    
    Si el certificado está en Archive tier, inicia la rehidratación
    y retorna un mensaje indicando el tiempo estimado.
    
    Returns:
        PDF del certificado o mensaje de estado si está en Archive
    """
    current_user_id = get_jwt_identity()
    
    certificate = ConocerCertificate.query.filter_by(
        id=certificate_id,
        user_id=current_user_id
    ).first()
    
    if not certificate:
        return jsonify({'error': 'Certificado no encontrado'}), 404
    
    try:
        blob_service = get_conocer_blob_service()
        content, properties = blob_service.download_certificate(certificate.blob_name)
        
        # Verificar integridad del archivo
        import hashlib
        file_hash = hashlib.sha256(content).hexdigest()
        if certificate.file_hash and file_hash != certificate.file_hash:
            current_app.logger.warning(
                f"Hash mismatch para certificado {certificate_id}: "
                f"esperado {certificate.file_hash}, obtenido {file_hash}"
            )
        
        # Crear nombre de archivo para descarga
        filename = f"CONOCER_{certificate.standard_code}_{certificate.certificate_number}.pdf"
        
        return send_file(
            io.BytesIO(content),
            mimetype='application/pdf',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        error_msg = str(e)
        
        # Verificar si es un error de rehidratación
        if 'archivo' in error_msg.lower() or 'archive' in error_msg.lower():
            return jsonify({
                'error': 'certificate_in_archive',
                'message': error_msg,
                'status': 'rehydrating',
                'estimated_time': '15 horas (aprox.)'
            }), 202  # 202 Accepted - proceso en curso
        
        current_app.logger.error(f"Error descargando certificado {certificate_id}: {e}")
        return jsonify({'error': 'Error al descargar el certificado'}), 500


@conocer_bp.route('/certificates/<int:certificate_id>/download-url', methods=['GET'])
@jwt_required()
def get_download_url(certificate_id):
    """
    Obtener URL temporal para descarga directa
    
    Query params:
        - expiry_hours: Horas de validez (default: 1, max: 24)
    
    Returns:
        URL con SAS token para descarga directa
    """
    current_user_id = get_jwt_identity()
    
    certificate = ConocerCertificate.query.filter_by(
        id=certificate_id,
        user_id=current_user_id
    ).first()
    
    if not certificate:
        return jsonify({'error': 'Certificado no encontrado'}), 404
    
    expiry_hours = min(int(request.args.get('expiry_hours', 1)), 24)
    
    try:
        blob_service = get_conocer_blob_service()
        filename = f"CONOCER_{certificate.standard_code}_{certificate.certificate_number}.pdf"
        download_url = blob_service.generate_download_url(
            certificate.blob_name,
            expiry_hours=expiry_hours,
            filename=filename
        )
        
        if not download_url:
            # Verificar si está en Archive
            status = blob_service.get_blob_status(certificate.blob_name)
            if status and status.get('blob_tier') == 'StandardBlobTier.ARCHIVE':
                return jsonify({
                    'error': 'certificate_in_archive',
                    'message': 'El certificado está en almacenamiento de archivo',
                    'rehydration_url': f'/api/conocer/certificates/{certificate_id}/rehydrate'
                }), 202
            
            return jsonify({'error': 'No se pudo generar el URL de descarga'}), 500
        
        return jsonify({
            'download_url': download_url,
            'expires_in_hours': expiry_hours,
            'filename': filename
        })
        
    except Exception as e:
        current_app.logger.error(f"Error generando URL para certificado {certificate_id}: {e}")
        return jsonify({'error': 'Error al generar URL de descarga'}), 500


@conocer_bp.route('/certificates/<int:certificate_id>/rehydrate', methods=['POST'])
@jwt_required()
def rehydrate_certificate(certificate_id):
    """
    Iniciar rehidratación de un certificado desde Archive tier
    
    Body JSON (opcional):
        - priority: 'standard' (~15hrs) o 'high' (~1hr, más costoso)
    
    Returns:
        Estado de la rehidratación
    """
    current_user_id = get_jwt_identity()
    
    certificate = ConocerCertificate.query.filter_by(
        id=certificate_id,
        user_id=current_user_id
    ).first()
    
    if not certificate:
        return jsonify({'error': 'Certificado no encontrado'}), 404
    
    data = request.get_json() or {}
    priority = data.get('priority', 'standard')
    
    try:
        blob_service = get_conocer_blob_service()
        result = blob_service.rehydrate_from_archive(certificate.blob_name, priority)
        
        return jsonify(result)
        
    except Exception as e:
        current_app.logger.error(f"Error rehidratando certificado {certificate_id}: {e}")
        return jsonify({'error': 'Error al iniciar rehidratación'}), 500


@conocer_bp.route('/certificates/<int:certificate_id>/status', methods=['GET'])
@jwt_required()
def get_certificate_status(certificate_id):
    """
    Obtener estado del certificado (útil para verificar rehidratación)
    
    Returns:
        Estado actual del certificado en blob storage
    """
    current_user_id = get_jwt_identity()
    
    certificate = ConocerCertificate.query.filter_by(
        id=certificate_id,
        user_id=current_user_id
    ).first()
    
    if not certificate:
        return jsonify({'error': 'Certificado no encontrado'}), 404
    
    try:
        blob_service = get_conocer_blob_service()
        status = blob_service.get_blob_status(certificate.blob_name)
        
        if not status:
            return jsonify({
                'error': 'Archivo no encontrado en almacenamiento',
                'certificate_id': certificate_id
            }), 404
        
        return jsonify({
            'certificate_id': certificate_id,
            'certificate_number': certificate.certificate_number,
            **status
        })
        
    except Exception as e:
        current_app.logger.error(f"Error obteniendo estado de certificado {certificate_id}: {e}")
        return jsonify({'error': 'Error al obtener estado'}), 500


# ===== ENDPOINTS ADMIN (para gestión) =====

@conocer_bp.route('/admin/certificates', methods=['POST'])
@jwt_required()
def upload_certificate():
    """
    Subir un nuevo certificado CONOCER (solo admin/editor)
    
    Form data:
        - file: Archivo PDF del certificado (required)
        - user_id: ID del usuario propietario (required)
        - certificate_number: Folio oficial CONOCER (required)
        - curp: CURP del certificado (required)
        - standard_code: Código del estándar (required, ej: EC0217)
        - standard_name: Nombre del estándar (required)
        - issue_date: Fecha de emisión YYYY-MM-DD (required)
        - expiration_date: Fecha de vencimiento YYYY-MM-DD (optional)
        - evaluation_date: Fecha de evaluación YYYY-MM-DD (optional)
        - competency_level: Nivel de competencia (optional)
        - evaluation_center_name: Nombre del centro evaluador (optional)
        - evaluation_center_code: Código del centro evaluador (optional)
        - evaluator_name: Nombre del evaluador (optional)
    
    Returns:
        Certificado creado con metadata
    """
    current_user_id = get_jwt_identity()
    current_user = User.query.get(current_user_id)
    
    # Verificar permisos
    if current_user.role not in ['admin', 'developer', 'editor']:
        return jsonify({'error': 'No tiene permisos para esta acción'}), 403
    
    # Validar archivo
    if 'file' not in request.files:
        return jsonify({'error': 'No se envió ningún archivo'}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'Nombre de archivo vacío'}), 400
    
    if not file.filename.lower().endswith('.pdf'):
        return jsonify({'error': 'Solo se permiten archivos PDF'}), 400
    
    # Validar campos requeridos
    required_fields = ['user_id', 'certificate_number', 'curp', 'standard_code', 'standard_name', 'issue_date']
    for field in required_fields:
        if not request.form.get(field):
            return jsonify({'error': f'El campo {field} es requerido'}), 400
    
    # Verificar que el usuario destino existe
    target_user = User.query.get(request.form['user_id'])
    if not target_user:
        return jsonify({'error': 'Usuario no encontrado'}), 404
    
    # Verificar que no exista certificado con el mismo número
    existing = ConocerCertificate.query.filter_by(
        certificate_number=request.form['certificate_number']
    ).first()
    if existing:
        return jsonify({'error': 'Ya existe un certificado con ese número de folio'}), 409
    
    try:
        # Subir archivo al blob storage
        blob_service = get_conocer_blob_service()
        file_content = file.read()
        
        blob_name, file_hash, file_size = blob_service.upload_certificate(
            file_content=file_content,
            user_id=request.form['user_id'],
            certificate_number=request.form['certificate_number'],
            standard_code=request.form['standard_code'],
            metadata={
                'curp': request.form['curp'],
                'standard_name': request.form['standard_name'],
                'uploaded_by': current_user_id
            }
        )
        
        # Parsear fechas
        issue_date = datetime.strptime(request.form['issue_date'], '%Y-%m-%d').date()
        expiration_date = None
        evaluation_date = None
        
        if request.form.get('expiration_date'):
            expiration_date = datetime.strptime(request.form['expiration_date'], '%Y-%m-%d').date()
        
        if request.form.get('evaluation_date'):
            evaluation_date = datetime.strptime(request.form['evaluation_date'], '%Y-%m-%d').date()
        
        # Crear registro en BD
        certificate = ConocerCertificate(
            user_id=request.form['user_id'],
            certificate_number=request.form['certificate_number'],
            curp=request.form['curp'],
            standard_code=request.form['standard_code'],
            standard_name=request.form['standard_name'],
            competency_level=request.form.get('competency_level'),
            evaluation_center_name=request.form.get('evaluation_center_name'),
            evaluation_center_code=request.form.get('evaluation_center_code'),
            evaluator_name=request.form.get('evaluator_name'),
            issue_date=issue_date,
            expiration_date=expiration_date,
            evaluation_date=evaluation_date,
            blob_name=blob_name,
            blob_container='conocer-certificates',
            blob_tier='Cool',
            file_size=file_size,
            file_hash=file_hash,
            status='active'
        )
        
        db.session.add(certificate)
        db.session.commit()
        
        return jsonify({
            'message': 'Certificado subido exitosamente',
            'certificate': certificate.to_dict(include_blob_info=True)
        }), 201
        
    except ValueError as e:
        return jsonify({'error': str(e)}), 400
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error subiendo certificado: {e}")
        return jsonify({'error': 'Error al subir el certificado'}), 500


@conocer_bp.route('/admin/certificates/<int:certificate_id>/archive', methods=['POST'])
@jwt_required()
def archive_certificate(certificate_id):
    """
    Mover un certificado al tier Archive manualmente (solo admin)
    
    Útil para certificados antiguos que ya no se consultan frecuentemente.
    
    Returns:
        Estado actualizado del certificado
    """
    current_user_id = get_jwt_identity()
    current_user = User.query.get(current_user_id)
    
    if current_user.role not in ['admin', 'developer']:
        return jsonify({'error': 'Solo administradores pueden archivar certificados'}), 403
    
    certificate = ConocerCertificate.query.get(certificate_id)
    if not certificate:
        return jsonify({'error': 'Certificado no encontrado'}), 404
    
    try:
        blob_service = get_conocer_blob_service()
        moved = blob_service.move_to_archive(certificate.blob_name)
        
        if moved:
            certificate.blob_tier = 'Archive'
            certificate.archived_at = datetime.utcnow()
            certificate.status = 'archived'
            db.session.commit()
            
            return jsonify({
                'message': 'Certificado movido a archivo exitosamente',
                'certificate': certificate.to_dict()
            })
        else:
            return jsonify({
                'message': 'El certificado ya estaba en archivo',
                'certificate': certificate.to_dict()
            })
        
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error archivando certificado {certificate_id}: {e}")
        return jsonify({'error': 'Error al archivar certificado'}), 500


@conocer_bp.route('/admin/certificates/by-user/<user_id>', methods=['GET'])
@jwt_required()
def get_user_certificates_admin(user_id):
    """
    Obtener certificados de un usuario específico (solo admin/editor)
    
    Returns:
        Lista de certificados del usuario
    """
    current_user_id = get_jwt_identity()
    current_user = User.query.get(current_user_id)
    
    if current_user.role not in ['admin', 'developer', 'editor', 'soporte']:
        return jsonify({'error': 'No tiene permisos para esta acción'}), 403
    
    certificates = ConocerCertificate.query.filter_by(user_id=user_id)\
        .order_by(ConocerCertificate.issue_date.desc()).all()
    
    return jsonify({
        'user_id': user_id,
        'certificates': [cert.to_dict(include_blob_info=True) for cert in certificates],
        'total': len(certificates)
    })


@conocer_bp.route('/verify/<certificate_number>', methods=['GET'])
def verify_certificate_public(certificate_number):
    """
    Verificar un certificado públicamente por su número de folio
    
    Este endpoint es público para permitir verificación externa.
    
    Returns:
        Información básica del certificado si existe
    """
    certificate = ConocerCertificate.query.filter_by(
        certificate_number=certificate_number
    ).first()
    
    if not certificate:
        return jsonify({
            'valid': False,
            'message': 'Certificado no encontrado'
        }), 404
    
    # Retornar solo información pública
    return jsonify({
        'valid': True,
        'certificate_number': certificate.certificate_number,
        'standard_code': certificate.standard_code,
        'standard_name': certificate.standard_name,
        'issue_date': certificate.issue_date.isoformat() if certificate.issue_date else None,
        'expiration_date': certificate.expiration_date.isoformat() if certificate.expiration_date else None,
        'is_expired': certificate.is_expired,
        'status': certificate.status,
        'evaluation_center_name': certificate.evaluation_center_name,
        'verification_url': certificate.verification_url
    })


# ===== ENDPOINTS DE CARGA MASIVA DE CERTIFICADOS CONOCER =====

@conocer_bp.route('/admin/upload-batch', methods=['POST'])
@jwt_required()
def upload_batch():
    """
    Subir un archivo ZIP con certificados CONOCER para procesamiento masivo.
    Responde HTTP 202 e inicia procesamiento en segundo plano.
    """
    current_user_id = get_jwt_identity()
    current_user = User.query.get(current_user_id)
    
    if not current_user or current_user.role not in ['admin', 'coordinator']:
        return jsonify({'error': 'No tiene permisos para esta acción'}), 403
    
    if 'file' not in request.files:
        return jsonify({'error': 'No se envió ningún archivo'}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'Nombre de archivo vacío'}), 400
    
    if not file.filename.lower().endswith('.zip'):
        return jsonify({'error': 'Solo se permiten archivos ZIP'}), 400
    
    try:
        import os
        import uuid
        import zipfile
        from azure.storage.blob import BlobServiceClient as _BSC, ContentSettings as _CS
        from app.models.conocer_upload import ConocerUploadBatch
        from app.services.conocer_batch_processor import process_batch_background
        
        file_content = file.read()
        
        if len(file_content) == 0:
            return jsonify({'error': 'El archivo está vacío'}), 400
        
        # Validar que es un ZIP válido
        try:
            zf_test = zipfile.ZipFile(io.BytesIO(file_content), 'r')
            all_files = [
                n for n in zf_test.namelist()
                if not n.endswith('/') and not n.startswith('__MACOSX')
            ]
            zf_test.close()
            if len(all_files) == 0:
                return jsonify({'error': 'El archivo ZIP está vacío'}), 400
        except zipfile.BadZipFile:
            return jsonify({'error': 'El archivo no es un ZIP válido'}), 400
        
        # Subir ZIP a blob temporal
        conn_str = os.getenv('AZURE_STORAGE_CONNECTION_STRING')
        container_name = os.getenv('AZURE_STORAGE_CONTAINER', 'evaluaasi-files')
        blob_name = f"conocer-uploads/{uuid.uuid4().hex}_{file.filename}"
        
        bsc = _BSC.from_connection_string(conn_str)
        blob_client = bsc.get_blob_client(container=container_name, blob=blob_name)
        blob_client.upload_blob(
            file_content,
            overwrite=True,
            content_settings=_CS(content_type='application/zip')
        )
        
        # Crear registro del batch
        batch = ConocerUploadBatch(
            uploaded_by=current_user_id,
            filename=file.filename,
            blob_name=blob_name,
            total_files=len(all_files),
            status='queued'
        )
        db.session.add(batch)
        db.session.commit()
        
        # Iniciar procesamiento en segundo plano
        process_batch_background(current_app._get_current_object(), batch.id)
        
        return jsonify({
            'message': 'Archivo recibido. El procesamiento se iniciará en segundo plano.',
            'batch_id': batch.id,
            'total_files': len(all_files),
            'status': 'queued'
        }), 202
        
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error en upload-batch: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'Error al procesar el archivo: {str(e)[:200]}'}), 500


@conocer_bp.route('/admin/upload-batches', methods=['GET'])
@jwt_required()
def list_upload_batches():
    """Listar batches de carga con paginación."""
    current_user_id = get_jwt_identity()
    current_user = User.query.get(current_user_id)
    
    if not current_user or current_user.role not in ['admin', 'coordinator']:
        return jsonify({'error': 'No tiene permisos para esta acción'}), 403
    
    from app.models.conocer_upload import ConocerUploadBatch
    
    page = request.args.get('page', 1, type=int)
    per_page = min(request.args.get('per_page', 20, type=int), 100)
    status_filter = request.args.get('status')
    
    query = ConocerUploadBatch.query
    if status_filter:
        query = query.filter_by(status=status_filter)
    
    query = query.order_by(ConocerUploadBatch.created_at.desc())
    pagination = query.paginate(page=page, per_page=per_page, error_out=False)
    
    return jsonify({
        'batches': [b.to_dict(include_uploader=True) for b in pagination.items],
        'total': pagination.total,
        'pages': pagination.pages,
        'current_page': page,
        'per_page': per_page
    })


@conocer_bp.route('/admin/upload-batches/<int:batch_id>', methods=['GET'])
@jwt_required()
def get_upload_batch_detail(batch_id):
    """Obtener detalle de un batch."""
    current_user_id = get_jwt_identity()
    current_user = User.query.get(current_user_id)
    
    if not current_user or current_user.role not in ['admin', 'coordinator']:
        return jsonify({'error': 'No tiene permisos para esta acción'}), 403
    
    from app.models.conocer_upload import ConocerUploadBatch
    
    batch = ConocerUploadBatch.query.get(batch_id)
    if not batch:
        return jsonify({'error': 'Batch no encontrado'}), 404
    
    data = batch.to_dict(include_uploader=True)
    data['progress_percentage'] = batch.progress_percentage
    data['success_count'] = batch.success_count
    
    return jsonify(data)


@conocer_bp.route('/admin/upload-batches/<int:batch_id>/logs', methods=['GET'])
@jwt_required()
def get_upload_batch_logs(batch_id):
    """Obtener logs de un batch con paginación y filtros."""
    current_user_id = get_jwt_identity()
    current_user = User.query.get(current_user_id)
    
    if not current_user or current_user.role not in ['admin', 'coordinator']:
        return jsonify({'error': 'No tiene permisos para esta acción'}), 403
    
    from app.models.conocer_upload import ConocerUploadBatch, ConocerUploadLog
    
    batch = ConocerUploadBatch.query.get(batch_id)
    if not batch:
        return jsonify({'error': 'Batch no encontrado'}), 404
    
    page = request.args.get('page', 1, type=int)
    per_page = min(request.args.get('per_page', 50, type=int), 200)
    status_filter = request.args.get('status')
    search = request.args.get('search', '').strip()
    
    query = ConocerUploadLog.query.filter_by(batch_id=batch_id)
    
    if status_filter:
        query = query.filter_by(status=status_filter)
    
    if search:
        search_lower = f'%{search.lower()}%'
        query = query.filter(
            db.or_(
                db.func.lower(ConocerUploadLog.filename).like(search_lower),
                db.func.lower(ConocerUploadLog.extracted_curp).like(search_lower),
                db.func.lower(ConocerUploadLog.extracted_ecm_code).like(search_lower),
                db.func.lower(ConocerUploadLog.extracted_name).like(search_lower),
                db.func.lower(ConocerUploadLog.extracted_folio).like(search_lower),
            )
        )
    
    query = query.order_by(ConocerUploadLog.created_at.asc())
    pagination = query.paginate(page=page, per_page=per_page, error_out=False)
    
    return jsonify({
        'logs': [log.to_dict() for log in pagination.items],
        'total': pagination.total,
        'pages': pagination.pages,
        'current_page': page,
        'per_page': per_page,
        'batch_status': batch.status,
        'batch_progress': batch.progress_percentage
    })


@conocer_bp.route('/admin/upload-batches/<int:batch_id>/export', methods=['GET'])
@jwt_required()
def export_upload_batch_logs(batch_id):
    """Exportar logs de un batch a Excel."""
    current_user_id = get_jwt_identity()
    current_user = User.query.get(current_user_id)
    
    if not current_user or current_user.role not in ['admin', 'coordinator']:
        return jsonify({'error': 'No tiene permisos para esta acción'}), 403
    
    from app.models.conocer_upload import ConocerUploadBatch, ConocerUploadLog
    
    batch = ConocerUploadBatch.query.get(batch_id)
    if not batch:
        return jsonify({'error': 'Batch no encontrado'}), 404
    
    logs = ConocerUploadLog.query.filter_by(batch_id=batch_id)\
        .order_by(ConocerUploadLog.created_at.asc()).all()
    
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Logs de Carga"
        
        headers = [
            'Archivo', 'CURP', 'ECM', 'Nombre', 'Folio',
            'Nombre ECM', 'Fecha Emisión', 'Entidad Certificadora',
            'Estado', 'Razón', 'Detalle', 'Tiempo (ms)'
        ]
        
        header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True, size=11)
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        status_labels = {
            'matched': 'Nuevo', 'replaced': 'Reemplazado',
            'skipped': 'Omitido (duplicado)', 'discarded': 'Descartado', 'error': 'Error'
        }
        status_fills = {
            'matched': PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid'),
            'replaced': PatternFill(start_color='BDD7EE', end_color='BDD7EE', fill_type='solid'),
            'skipped': PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid'),
            'discarded': PatternFill(start_color='F2DCDB', end_color='F2DCDB', fill_type='solid'),
            'error': PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid'),
        }
        
        for row_idx, log in enumerate(logs, 2):
            ws.cell(row=row_idx, column=1, value=log.filename)
            ws.cell(row=row_idx, column=2, value=log.extracted_curp)
            ws.cell(row=row_idx, column=3, value=log.extracted_ecm_code)
            ws.cell(row=row_idx, column=4, value=log.extracted_name)
            ws.cell(row=row_idx, column=5, value=log.extracted_folio)
            ws.cell(row=row_idx, column=6, value=log.extracted_ecm_name)
            ws.cell(row=row_idx, column=7, value=log.extracted_issue_date)
            ws.cell(row=row_idx, column=8, value=log.extracted_certifying_entity)
            status_cell = ws.cell(row=row_idx, column=9, value=status_labels.get(log.status, log.status))
            if log.status in status_fills:
                status_cell.fill = status_fills[log.status]
            ws.cell(row=row_idx, column=10, value=log.discard_reason)
            ws.cell(row=row_idx, column=11, value=log.discard_detail)
            ws.cell(row=row_idx, column=12, value=log.processing_time_ms)
        
        for col in ws.columns:
            max_length = 0
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 50)
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"Logs_CONOCER_Batch_{batch_id}_{datetime.utcnow().strftime('%Y%m%d')}.xlsx"
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    except Exception as e:
        current_app.logger.error(f"Error exportando logs del batch {batch_id}: {e}")
        return jsonify({'error': 'Error al exportar logs'}), 500


@conocer_bp.route('/admin/upload-batches/<int:batch_id>/retry', methods=['POST'])
@jwt_required()
def retry_upload_batch(batch_id):
    """Reintentar un batch fallido o atascado (>30 min en processing)."""
    current_user_id = get_jwt_identity()
    current_user = User.query.get(current_user_id)
    
    if not current_user or current_user.role not in ['admin', 'coordinator']:
        return jsonify({'error': 'No tiene permisos para esta acción'}), 403
    
    from app.models.conocer_upload import ConocerUploadBatch, ConocerUploadLog
    from app.services.conocer_batch_processor import process_batch_background
    
    batch = ConocerUploadBatch.query.get(batch_id)
    if not batch:
        return jsonify({'error': 'Batch no encontrado'}), 404
    
    if batch.status == 'completed':
        return jsonify({'error': 'Este batch ya fue completado exitosamente'}), 400
    
    if batch.status == 'processing':
        if batch.started_at:
            elapsed = (datetime.utcnow() - batch.started_at).total_seconds()
            if elapsed < 1800:
                return jsonify({
                    'error': 'Este batch aún está en procesamiento',
                    'elapsed_seconds': int(elapsed)
                }), 400
    
    # Limpiar logs existentes y resetear
    ConocerUploadLog.query.filter_by(batch_id=batch_id).delete()
    batch.status = 'queued'
    batch.processed_files = 0
    batch.matched_files = 0
    batch.replaced_files = 0
    batch.skipped_files = 0
    batch.discarded_files = 0
    batch.error_files = 0
    batch.started_at = None
    batch.completed_at = None
    batch.error_message = None
    db.session.commit()
    
    process_batch_background(current_app._get_current_object(), batch.id)
    
    return jsonify({
        'message': 'Reprocesamiento iniciado',
        'batch_id': batch.id,
        'status': 'queued'
    }), 202
