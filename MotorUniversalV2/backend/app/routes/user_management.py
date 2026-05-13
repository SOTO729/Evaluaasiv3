"""
Rutas para gestión de usuarios (admin y coordinadores)
"""
from flask import Blueprint, request, jsonify, g
from functools import wraps
from flask_jwt_extended import jwt_required, get_jwt_identity
from sqlalchemy import or_, and_, func, case, text
from werkzeug.exceptions import HTTPException
from datetime import datetime, timedelta
from app import db, cache
from app.models import User
from app.models.user import encrypt_password
from app.routes.partners import (
    FOREIGN_CURP_MALE, FOREIGN_CURP_FEMALE, GENERIC_FOREIGN_CURPS,
    _get_foreign_curp, _is_generic_foreign_curp,
)
from app.utils.rate_limit import rate_limit
import uuid
import re
import logging

logger = logging.getLogger(__name__)


def _safe_cache_get(key):
    """cache.get tolerante a fallos de Redis (UM-stats-fix)."""
    try:
        return cache.get(key)
    except Exception as _e:
        logger.warning('cache.get fallo (%s): %s', key, _e)
        return None


def _safe_cache_set(key, value, timeout=300):
    """cache.set tolerante a fallos de Redis (UM-stats-fix)."""
    try:
        cache.set(key, value, timeout=timeout)
    except Exception as _e:
        logger.warning('cache.set fallo (%s): %s', key, _e)

bp = Blueprint('user_management', __name__, url_prefix='/api/user-management')

# Roles disponibles en el sistema (sin alumno)
AVAILABLE_ROLES = ['admin', 'developer', 'gerente', 'financiero', 'editor', 'editor_invitado', 'soporte', 'coordinator', 'responsable', 'responsable_partner', 'responsable_estatal', 'candidato', 'auxiliar']

# Roles que soporte puede VER (todos menos admin, developer, gerente)
SOPORTE_VISIBLE_ROLES = ['financiero', 'editor', 'editor_invitado', 'soporte', 'coordinator', 'responsable', 'responsable_partner', 'responsable_estatal', 'candidato', 'auxiliar']

# Roles que soporte puede EDITAR, ver contraseñas y activar/desactivar
SOPORTE_EDITABLE_ROLES = ['candidato', 'responsable', 'responsable_partner', 'responsable_estatal']

# Roles que puede crear cada tipo de usuario
ROLE_CREATE_PERMISSIONS = {
    'admin': ['developer', 'gerente', 'financiero', 'editor', 'editor_invitado', 'soporte', 'coordinator', 'responsable', 'responsable_partner', 'responsable_estatal', 'candidato', 'auxiliar'],  # Todo menos admin
    'developer': ['gerente', 'financiero', 'editor', 'editor_invitado', 'soporte', 'coordinator', 'responsable', 'responsable_partner', 'responsable_estatal', 'candidato', 'auxiliar'],  # Todo menos admin y developer
    'soporte': ['responsable', 'responsable_partner', 'responsable_estatal', 'candidato', 'auxiliar'],  # Soporte puede crear responsables, responsables del partner/estatal, candidatos y auxiliares
    'coordinator': ['responsable', 'responsable_partner', 'responsable_estatal', 'candidato', 'auxiliar'],  # Responsables, Responsables del Partner/Estatal, candidatos y auxiliares
    'responsable': ['candidato']  # Solo candidatos
}


_ROLE_LABELS_ES = {
    'admin': 'Administrador',
    'developer': 'Desarrollador',
    'gerente': 'Gerente',
    'financiero': 'Financiero',
    'editor': 'Editor',
    'editor_invitado': 'Editor invitado',
    'soporte': 'Soporte',
    'coordinator': 'Coordinador',
    'responsable': 'Responsable',
    'responsable_partner': 'Responsable de partner',
    'responsable_estatal': 'Responsable estatal',
    'auxiliar': 'Auxiliar',
    'candidato': 'Candidato',
}


def _describe_user(user):
    """Resumen humano-legible de un usuario para mensajes de error."""
    if not user:
        return ''
    full_name = ' '.join(filter(None, [
        getattr(user, 'name', None),
        getattr(user, 'first_surname', None),
        getattr(user, 'second_surname', None),
    ])).strip() or '(sin nombre)'
    role_label = _ROLE_LABELS_ES.get(getattr(user, 'role', '') or '', getattr(user, 'role', '') or 'usuario')
    parts = [f'{full_name} ({role_label})']
    if getattr(user, 'username', None):
        parts.append(f'usuario: {user.username}')
    if getattr(user, 'email', None):
        parts.append(f'correo: {user.email}')
    status = 'activo' if getattr(user, 'is_active', False) else 'inactivo'
    parts.append(status)
    return ' — '.join(parts)


def _duplicate_email_response(email, existing_user):
    """Mensaje detallado de email duplicado."""
    return jsonify({
        'error': (
            f'El correo "{email}" ya está registrado para {_describe_user(existing_user)}. '
            f'Si se trata de la misma persona, edita ese usuario en vez de crear uno nuevo. '
            f'Si es otra persona, usa un correo distinto.'
        ),
        'duplicate_field': 'email',
        'duplicate_value': email,
        'existing_user': {
            'id': existing_user.id,
            'username': existing_user.username,
            'email': existing_user.email,
            'full_name': ' '.join(filter(None, [existing_user.name, existing_user.first_surname, existing_user.second_surname])).strip(),
            'role': existing_user.role,
            'is_active': existing_user.is_active,
        },
    }), 409


def _duplicate_curp_response(curp, existing_user):
    """Mensaje detallado de CURP duplicado."""
    return jsonify({
        'error': (
            f'La CURP "{curp}" ya está registrada para {_describe_user(existing_user)}. '
            f'Cada CURP es única por persona, así que no puede asignarse a un nuevo usuario. '
            f'Si es la misma persona, edita ese usuario; si es un homónimo, verifica que la CURP capturada sea correcta.'
        ),
        'duplicate_field': 'curp',
        'duplicate_value': curp,
        'existing_user': {
            'id': existing_user.id,
            'username': existing_user.username,
            'email': existing_user.email,
            'full_name': ' '.join(filter(None, [existing_user.name, existing_user.first_surname, existing_user.second_surname])).strip(),
            'role': existing_user.role,
            'is_active': existing_user.is_active,
            'curp': existing_user.curp,
        },
    }), 409


def _assign_existing_to_group_response(existing_user, target_group, dup_field):
    """Cuando se intenta crear un candidato cuyo email/CURP ya existe y el alta
    incluye un grupo destino, en lugar de fallar con 409 asignamos al usuario
    existente al grupo (mismo comportamiento que la carga masiva).
    Retorna (Response, status_code).
    """
    from app.models.partner import GroupMember
    try:
        # Verificar membresía existente
        already_member = GroupMember.query.filter_by(
            group_id=target_group.id, user_id=existing_user.id
        ).first()
        if already_member:
            return jsonify({
                'message': (
                    f'El usuario ya existe y ya pertenece al grupo "{target_group.name}". '
                    f'No se realizó ningún cambio.'
                ),
                'user': existing_user.to_dict(include_private=False),
                'group_id': target_group.id,
                'group_name': target_group.name,
                'existing_user_assigned': False,
                'duplicate_field': dup_field,
            }), 200

        db.session.add(GroupMember(
            group_id=target_group.id,
            user_id=existing_user.id,
            status='existing_assigned',
        ))
        db.session.commit()
        return jsonify({
            'message': (
                f'El usuario ya existía como {_describe_user(existing_user)}. '
                f'Se asignó al grupo "{target_group.name}".'
            ),
            'user': existing_user.to_dict(include_private=False),
            'group_id': target_group.id,
            'group_name': target_group.name,
            'existing_user_assigned': True,
            'duplicate_field': dup_field,
        }), 200
    except Exception:
        db.session.rollback()
        logger.exception('user_management error')
        return jsonify({'error': 'Error interno del servidor'}), 500


def _is_coordinator_role(role):
    """True si el rol es coordinator o auxiliar (auxiliar hereda permisos de coordinador)"""
    return role in ('coordinator', 'auxiliar')


def _get_effective_coordinator_id(user):
    """Retorna el ID del coordinador efectivo para filtrar datos multi-tenant.
    Coordinador usa su propio ID. Auxiliar usa el coordinator_id de su creador.
    Admin/developer retorna None (ve todo)."""
    if user.role == 'coordinator':
        return user.id
    if user.role == 'auxiliar':
        return user.coordinator_id
    return None


def _verify_coordinator_user_access(current_user, target_user):
    """
    Verifica que el target_user pertenezca al mismo tenant que current_user.
    Retorna None si OK, o (response, status_code) si error.
    """
    if target_user is None:
        return jsonify({'error': 'Usuario no encontrado'}), 404
    if not _is_coordinator_role(current_user.role):
        return None  # Solo aplica a coordinator/auxiliar
    if target_user.role not in ['candidato', 'responsable', 'responsable_partner', 'responsable_estatal', 'auxiliar']:
        return jsonify({'error': 'Acceso denegado'}), 403
    if target_user.role == 'candidato':
        return None  # candidatos son compartidos entre coordinadores

    coord_id = _get_effective_coordinator_id(current_user)
    if target_user.coordinator_id == coord_id:
        return None

    # UM-M4: también permitir acceso si el responsable está asignado a un
    # campus del coordinador (caso histórico: responsables creados antes de
    # implementar coordinator_id, o asignados vía Campus.responsable_id).
    try:
        from app.models.partner import Campus
        if target_user.campus_id:
            owned = Campus.query.filter(
                Campus.id == target_user.campus_id,
                Campus.coordinator_id == coord_id,
            ).first()
            if owned:
                return None
        # También caso responsable_id directo en Campus.
        if target_user.id:
            owned = Campus.query.filter(
                Campus.responsable_id == target_user.id,
                Campus.coordinator_id == coord_id,
            ).first()
            if owned:
                return None
    except Exception:
        logger.exception('_verify_coordinator_user_access: error consultando Campus')

    return jsonify({'error': 'Acceso denegado'}), 403


def management_required(f):
    """Decorador que requiere rol de admin, developer, coordinator, auxiliar, soporte o responsable"""
    @wraps(f)
    def decorated(*args, **kwargs):
        user_id = get_jwt_identity()
        user = User.query.get(user_id)
        if not user:
            return jsonify({'error': 'No autorizado'}), 401
        if user.role not in ['admin', 'developer', 'coordinator', 'auxiliar', 'soporte', 'responsable']:
            return jsonify({'error': 'Acceso denegado. Se requiere rol de administrador, desarrollador, soporte o coordinador'}), 403
        if user.role == 'responsable' and not user.campus_id:
            return jsonify({'error': 'No tienes un plantel asignado'}), 403
        g.current_user = user
        return f(*args, **kwargs)
    return decorated


def _ensure_bulk_upload_allowed(current_user):
    """Restringir altas masivas para responsables sin permiso explícito."""
    if current_user.role == 'responsable' and not bool(getattr(current_user, 'can_bulk_create_candidates', False)):
        return jsonify({'error': 'No tienes permiso para realizar altas masivas de candidatos'}), 403
    return None


def admin_required(f):
    """Decorador que requiere rol de admin o developer"""
    @wraps(f)
    def decorated(*args, **kwargs):
        user_id = get_jwt_identity()
        user = User.query.get(user_id)
        if not user:
            return jsonify({'error': 'No autorizado'}), 401
        if user.role not in ['admin', 'developer']:
            return jsonify({'error': 'Acceso denegado. Se requiere rol de administrador'}), 403
        g.current_user = user
        return f(*args, **kwargs)
    return decorated


def validate_email(email):
    """Validar formato de email"""
    pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
    return re.match(pattern, email) is not None


def validate_password(password):
    """Validar requisitos de contraseña"""
    if len(password) < 8:
        return False, "La contraseña debe tener al menos 8 caracteres"
    if not re.search(r'[A-Z]', password):
        return False, "La contraseña debe tener al menos una mayúscula"
    if not re.search(r'[a-z]', password):
        return False, "La contraseña debe tener al menos una minúscula"
    if not re.search(r'[0-9]', password):
        return False, "La contraseña debe tener al menos un número"
    return True, None


def generate_secure_password(length=12):
    """Generar contraseña que siempre cumple los requisitos de validación.
    Excluye caracteres confusos: i, I, l, L, o, O, 0
    """
    import secrets, string
    _CONFUSING = set('iIlLoO0')
    upper = ''.join(c for c in string.ascii_uppercase if c not in _CONFUSING)
    lower = ''.join(c for c in string.ascii_lowercase if c not in _CONFUSING)
    digits = ''.join(c for c in string.digits if c not in _CONFUSING)
    special = '-_'
    all_chars = upper + lower + digits + special
    while True:
        pwd = (
            secrets.choice(upper)
            + secrets.choice(lower)
            + secrets.choice(digits)
            + ''.join(secrets.choice(all_chars) for _ in range(length - 3))
        )
        # Mezclar para que no sea predecible
        pwd_list = list(pwd)
        secrets.SystemRandom().shuffle(pwd_list)
        pwd = ''.join(pwd_list)
        is_valid, _ = validate_password(pwd)
        if is_valid:
            return pwd


# ============== LISTAR USUARIOS ==============

@bp.route('/users', methods=['GET'])
@jwt_required()
@management_required
def list_users():
    """
    Listar usuarios según permisos del solicitante.
    Optimizado para escalar a 100K+ usuarios con:
    - Select específico de columnas (evita SELECT *)
    - Cursor-based pagination opcional
    - Índices optimizados
    """
    try:
        current_user = g.current_user
        
        # Parámetros de paginación
        page = request.args.get('page', 1, type=int)
        per_page = min(request.args.get('per_page', 20, type=int), 100)  # Max 100 por página
        
        # Cursor-based pagination (más eficiente para grandes datasets)
        cursor = request.args.get('cursor', '')  # ID del último usuario
        cursor_date = request.args.get('cursor_date', '')  # Fecha del último usuario
        use_cursor = cursor and cursor_date
        
        # Filtros
        search = request.args.get('search', '').strip()
        role_filter = request.args.get('role', '')
        active_filter = request.args.get('is_active', '')
        created_from = request.args.get('created_from', '').strip()
        created_to = request.args.get('created_to', '').strip()
        sort_by = request.args.get('sort_by', 'created_at')
        sort_order = request.args.get('sort_order', 'desc')
        
        # Seleccionar solo las columnas necesarias para la lista (evita cargar todo)
        columns = [
            User.id, User.email, User.username, User.name, User.first_surname,
            User.second_surname, User.gender, User.role, User.is_active,
            User.is_verified, User.created_at, User.last_login, User.curp,
            User.curp_verified,
            User.phone, User.campus_id, User.date_of_birth,
            User.can_bulk_create_candidates, User.can_manage_groups, User.can_view_reports,
            User.enable_evaluation_report, User.enable_certificate,
            User.enable_conocer_certificate, User.enable_digital_badge
        ]
        
        query = db.session.query(*columns)

        # UM-C3: excluir usuarios soft-eliminados de todos los listados.
        query = query.filter(User.is_deleted == False)  # noqa: E712

        # UM-M3: responsable NUNCA puede listar otros responsables; devolver
        # 403 explícito si se pide rol responsable directamente.
        if current_user.role == 'responsable' and role_filter:
            requested_roles = {r.strip() for r in role_filter.split(',') if r.strip()}
            if requested_roles - {'candidato'}:
                return jsonify({
                    'error': 'No tienes permiso para listar usuarios distintos de candidatos.'
                }), 403

        # Coordinadores y auxiliares: ven TODOS los candidatos (compartidos) + sus propios responsables/auxiliares
        # UM-M4: también incluir responsables asignados a campuses del coordinador,
        # aunque User.coordinator_id no esté seteado (datos históricos o asignados vía Campus).
        if _is_coordinator_role(current_user.role):
            from app.models.partner import Campus as _Campus
            coord_id = _get_effective_coordinator_id(current_user)
            # IDs de campuses propiedad del coordinador.
            owned_campus_ids = db.session.query(_Campus.id).filter(
                _Campus.coordinator_id == coord_id
            ).subquery()
            # IDs de usuarios asignados como responsable_id en algún campus propio.
            owned_responsable_ids = db.session.query(_Campus.responsable_id).filter(
                _Campus.coordinator_id == coord_id,
                _Campus.responsable_id.isnot(None),
            ).subquery()
            query = query.filter(
                or_(
                    User.role == 'candidato',
                    and_(
                        User.role.in_(['responsable', 'responsable_partner', 'responsable_estatal', 'auxiliar']),
                        or_(
                            User.coordinator_id == coord_id,
                            User.campus_id.in_(owned_campus_ids),
                            User.id.in_(owned_responsable_ids),
                        ),
                    )
                )
            )

        # Soporte ve todos los usuarios excepto admin, developer y gerente
        if current_user.role == 'soporte':
            query = query.filter(User.role.in_(SOPORTE_VISIBLE_ROLES))

        # Responsables solo ven candidatos de su plantel (por grupo O por campus_id directo)
        if current_user.role == 'responsable':
            from app.models.partner import GroupMember, CandidateGroup
            campus_user_ids = db.session.query(GroupMember.user_id).join(
                CandidateGroup, GroupMember.group_id == CandidateGroup.id
            ).filter(
                CandidateGroup.campus_id == current_user.campus_id
            ).distinct().subquery()
            query = query.filter(
                User.role == 'candidato',
                or_(
                    User.id.in_(campus_user_ids),
                    User.campus_id == current_user.campus_id
                )
            )
        
        # Filtros
        if role_filter:
            roles = [r.strip() for r in role_filter.split(',') if r.strip()]
            if len(roles) == 1:
                query = query.filter(User.role == roles[0])
            elif len(roles) > 1:
                query = query.filter(User.role.in_(roles))
            
        if active_filter != '':
            is_active = active_filter.lower() == 'true'
            query = query.filter(User.is_active == is_active)
        
        # Búsqueda optimizada
        if search:
            search_term = f'%{search}%'
            # Intentar usar búsqueda por prefijo si es corta (más eficiente con índices)
            if len(search) <= 3:
                search_prefix = f'{search}%'
                query = query.filter(
                    or_(
                        User.name.ilike(search_prefix),
                        User.first_surname.ilike(search_prefix),
                        User.email.ilike(search_prefix),
                        User.curp.ilike(search_prefix),
                        User.username.ilike(search_prefix)
                    )
                )
            else:
                query = query.filter(
                    or_(
                        User.name.ilike(search_term),
                        User.first_surname.ilike(search_term),
                        User.email.ilike(search_term),
                        User.curp.ilike(search_term),
                        User.username.ilike(search_term)
                    )
                )
        
        # Filtros de fecha de creación
        if created_from:
            try:
                date_from = datetime.fromisoformat(created_from)
                query = query.filter(User.created_at >= date_from)
            except ValueError:
                pass
        if created_to:
            try:
                date_to = datetime.fromisoformat(created_to + 'T23:59:59') if 'T' not in created_to else datetime.fromisoformat(created_to)
                query = query.filter(User.created_at <= date_to)
            except ValueError:
                pass
        
        # Ordenamiento dinámico
        sort_columns = {
            'name': User.name,
            'full_name': User.first_surname,
            'email': User.email,
            'role': User.role,
            'is_active': User.is_active,
            'created_at': User.created_at,
            'last_login': User.last_login,
            'curp': User.curp
        }
        
        sort_column = sort_columns.get(sort_by, User.created_at)
        
        # Cursor-based pagination (más eficiente para páginas > 100)
        if use_cursor and sort_by == 'created_at':
            try:
                cursor_datetime = datetime.fromisoformat(cursor_date)
                if sort_order == 'desc':
                    query = query.filter(
                        or_(
                            User.created_at < cursor_datetime,
                            and_(User.created_at == cursor_datetime, User.id < cursor)
                        )
                    )
                else:
                    query = query.filter(
                        or_(
                            User.created_at > cursor_datetime,
                            and_(User.created_at == cursor_datetime, User.id > cursor)
                        )
                    )
            except ValueError:
                pass  # Invalid cursor, use offset pagination
        
        # Aplicar ordenamiento
        if sort_order == 'asc':
            query = query.order_by(sort_column.asc(), User.id.asc())
        else:
            query = query.order_by(sort_column.desc(), User.id.desc())
        
        # Contar total solo si no usamos cursor (es costoso)
        if use_cursor:
            # Con cursor, estimamos el total o lo cacheamos
            total = _get_cached_user_count(current_user.role, role_filter, active_filter, coordinator_id=_get_effective_coordinator_id(current_user), campus_id=current_user.campus_id if current_user.role == 'responsable' else None)
            users_data = query.limit(per_page + 1).all()  # +1 para saber si hay más
            has_more = len(users_data) > per_page
            users_data = users_data[:per_page]
            pages = -1  # Indicar que usamos cursor
        else:
            # Paginación tradicional con offset
            total = query.count()
            pages = (total + per_page - 1) // per_page
            offset = (page - 1) * per_page
            users_data = query.offset(offset).limit(per_page).all()
            has_more = page < pages
        
        # Convertir resultados a diccionarios
        users_list = []
        for row in users_data:
            full_name = ' '.join(filter(None, [row.name, row.first_surname, row.second_surname]))
            users_list.append({
                'id': row.id,
                'email': row.email,
                'username': row.username,
                'name': row.name,
                'first_surname': row.first_surname,
                'second_surname': row.second_surname,
                'full_name': full_name,
                'gender': row.gender,
                'role': row.role,
                'is_active': row.is_active,
                'is_verified': row.is_verified,
                'created_at': row.created_at.isoformat() if row.created_at else None,
                'last_login': row.last_login.isoformat() if row.last_login else None,
                'curp': row.curp,
                'curp_verified': row.curp_verified,
                'phone': row.phone,
                'campus_id': row.campus_id,
                'date_of_birth': row.date_of_birth.isoformat() if row.date_of_birth else None,
                'can_bulk_create_candidates': row.can_bulk_create_candidates,
                'can_manage_groups': row.can_manage_groups,
                'can_view_reports': row.can_view_reports,
                'document_options': {
                    'evaluation_report': row.enable_evaluation_report,
                    'certificate': row.enable_certificate,
                    'conocer_certificate': row.enable_conocer_certificate,
                    'digital_badge': row.enable_digital_badge
                }
            })
        
        # Generar cursor para siguiente página
        next_cursor = None
        next_cursor_date = None
        if users_list and has_more:
            last_user = users_list[-1]
            next_cursor = last_user['id']
            next_cursor_date = last_user['created_at']
        
        return jsonify({
            'users': users_list,
            'total': total,
            'pages': pages,
            'current_page': page,
            'has_more': has_more,
            'next_cursor': next_cursor,
            'next_cursor_date': next_cursor_date
        })
        
    except HTTPException:
        
        raise
        
    except Exception:
        logger.exception('user_management error')
        return jsonify({'error': 'Error interno del servidor'}), 500


def _get_cached_user_count(user_role, role_filter='', active_filter='', coordinator_id=None, campus_id=None):
    """
    Obtener conteo de usuarios con caché.
    El conteo exacto es costoso, así que lo cacheamos por 5 minutos.
    """
    cache_key = f'user_count_{user_role}_{role_filter}_{active_filter}_{coordinator_id or "all"}_{campus_id or "all"}'
    cached = _safe_cache_get(cache_key)
    if cached is not None:
        return cached
    
    query = User.query.filter(User.is_deleted == False)  # noqa: E712
    if _is_coordinator_role(user_role):
        # UM-M4: count consistente con list_users (incluye responsables ligados
        # al coordinador por User.coordinator_id, User.campus_id o
        # Campus.responsable_id sobre campuses propios).
        if coordinator_id:
            from app.models.partner import Campus as _Campus
            owned_campus_ids = db.session.query(_Campus.id).filter(
                _Campus.coordinator_id == coordinator_id
            ).subquery()
            owned_responsable_ids = db.session.query(_Campus.responsable_id).filter(
                _Campus.coordinator_id == coordinator_id,
                _Campus.responsable_id.isnot(None),
            ).subquery()
            query = query.filter(
                or_(
                    User.role == 'candidato',
                    and_(
                        User.role.in_(['responsable', 'responsable_partner', 'responsable_estatal', 'auxiliar']),
                        or_(
                            User.coordinator_id == coordinator_id,
                            User.campus_id.in_(owned_campus_ids),
                            User.id.in_(owned_responsable_ids),
                        ),
                    )
                )
            )
        else:
            query = query.filter(User.role.in_(['candidato', 'responsable', 'responsable_partner', 'responsable_estatal', 'auxiliar']))
    if user_role == 'soporte':
        query = query.filter(User.role.in_(SOPORTE_VISIBLE_ROLES))
    if user_role == 'responsable' and campus_id:
        from app.models.partner import GroupMember, CandidateGroup
        campus_user_ids = db.session.query(GroupMember.user_id).join(
            CandidateGroup, GroupMember.group_id == CandidateGroup.id
        ).filter(
            CandidateGroup.campus_id == campus_id
        ).distinct().subquery()
        query = query.filter(
            User.role == 'candidato',
            or_(
                User.id.in_(campus_user_ids),
                User.campus_id == campus_id
            )
        )
    if role_filter:
        roles = [r.strip() for r in role_filter.split(',') if r.strip()]
        query = query.filter(User.role.in_(roles))
    if active_filter:
        is_active = active_filter.lower() == 'true'
        query = query.filter(User.is_active == is_active)
    
    count = query.count()
    _safe_cache_set(cache_key, count, timeout=300)  # 5 minutos
    return count


@bp.route('/users/<string:user_id>', methods=['GET'])
@jwt_required()
@management_required
def get_user_detail(user_id):
    """Obtener detalle de un usuario"""
    try:
        current_user = g.current_user
        user = User.query.get_or_404(user_id)

        # UM-C3: ocultar usuarios soft-eliminados excepto a admin/developer.
        if getattr(user, 'is_deleted', False) and current_user.role not in ('admin', 'developer'):
            return jsonify({'error': 'Usuario no encontrado'}), 404
        
        # Coordinadores y auxiliares pueden ver candidatos, responsables y responsables del partner
        if _is_coordinator_role(current_user.role):
            err = _verify_coordinator_user_access(current_user, user)
            if err:
                return err

        # Soporte puede ver todos los usuarios excepto admin, developer y gerente
        if current_user.role == 'soporte':
            if user.role not in SOPORTE_VISIBLE_ROLES:
                return jsonify({'error': 'No tienes permiso para ver este usuario'}), 403

        # Responsable solo puede ver candidatos de su plantel
        if current_user.role == 'responsable':
            if user.role != 'candidato':
                return jsonify({'error': 'No tienes permiso para ver este usuario'}), 403
            from app.models.partner import GroupMember, CandidateGroup
            is_in_campus = db.session.query(GroupMember.id).join(
                CandidateGroup, GroupMember.group_id == CandidateGroup.id
            ).filter(
                GroupMember.user_id == user.id,
                CandidateGroup.campus_id == current_user.campus_id
            ).first()
            if not is_in_campus and user.campus_id != current_user.campus_id:
                return jsonify({'error': 'No tienes permiso para ver este usuario'}), 403
        
        return jsonify({
            'user': user.to_dict(include_private=True, include_partners=True)
        })
        
    except HTTPException:
        raise
    except Exception as e:
        logger.exception('user_management error')
        return jsonify({'error': 'Error interno del servidor'}), 500


# ============== HISTORIAL DE GRUPOS DE UN CANDIDATO ==============

@bp.route('/users/<string:user_id>/group-history', methods=['GET'])
@jwt_required()
@management_required
@rate_limit(limit=300, window=60, key_prefix='rl_um_gh')  # UM-N3
def get_user_group_history(user_id):
    """Obtener historial completo de grupos de un candidato con asignaciones y resultados"""
    try:
        current_user = g.current_user
        user = User.query.get_or_404(user_id)

        if user.role != 'candidato':
            return jsonify({'error': 'El historial de grupos solo está disponible para candidatos'}), 400

        # Permisos: misma lógica que get_user_detail
        if _is_coordinator_role(current_user.role):
            err = _verify_coordinator_user_access(current_user, user)
            if err:
                return err
        elif current_user.role == 'responsable':
            from app.models.partner import GroupMember as GM2, CandidateGroup as CG2
            is_in_campus = db.session.query(GM2.id).join(
                CG2, GM2.group_id == CG2.id
            ).filter(GM2.user_id == user.id, CG2.campus_id == current_user.campus_id).first()
            if not is_in_campus and user.campus_id != current_user.campus_id:
                return jsonify({'error': 'No tienes permiso para ver este usuario'}), 403

        from app.models.partner import (
            GroupMember, CandidateGroup, GroupExam, GroupExamMember,
            EcmCandidateAssignment, Campus, SchoolCycle
        )
        from app.models.result import Result
        from app.models.exam import Exam
        from app.models.competency_standard import CompetencyStandard

        # 1) Obtener todas las membresías del candidato
        memberships = GroupMember.query.filter_by(user_id=user_id).order_by(GroupMember.joined_at.desc()).all()

        groups_data = []
        for membership in memberships:
            group = membership.group
            if not group:
                continue

            campus = Campus.query.get(group.campus_id) if group.campus_id else None
            cycle = SchoolCycle.query.get(group.school_cycle_id) if group.school_cycle_id else None

            # 2) Exámenes asignados en este grupo para este candidato
            group_exams = GroupExam.query.filter_by(group_id=group.id).all()
            exams_data = []
            for ge in group_exams:
                # Verificar si el candidato tiene acceso a este examen
                if ge.assignment_type == 'selected':
                    is_assigned = GroupExamMember.query.filter_by(
                        group_exam_id=ge.id, user_id=user_id
                    ).first()
                    if not is_assigned:
                        continue

                exam = Exam.query.get(ge.exam_id)
                cs = CompetencyStandard.query.get(exam.competency_standard_id) if exam and exam.competency_standard_id else None

                # 3) Asignación ECM para este candidato en este grupo+examen
                ecm = EcmCandidateAssignment.query.filter_by(
                    user_id=user_id, group_id=group.id, exam_id=ge.exam_id
                ).first()

                # 4) Resultados del candidato en este grupo+examen
                results = Result.query.filter_by(
                    user_id=user_id, group_id=group.id, group_exam_id=ge.id
                ).order_by(Result.start_date.desc()).all()

                exams_data.append({
                    'group_exam_id': ge.id,
                    'exam_id': ge.exam_id,
                    'exam_name': exam.name if exam else None,
                    'exam_version': exam.version if exam else None,
                    'competency_standard': {
                        'id': cs.id, 'code': cs.code, 'name': cs.name
                    } if cs else None,
                    'assignment_type': ge.assignment_type,
                    'assigned_at': ge.assigned_at.isoformat() if ge.assigned_at else None,
                    'max_attempts': ge.max_attempts,
                    'passing_score': ge.passing_score,
                    'is_active': ge.is_active,
                    'expires_at': ge.effective_expires_at.isoformat() if ge.effective_expires_at else None,
                    'is_expired': ge.is_expired,
                    'ecm_assignment': {
                        'id': ecm.id,
                        'assignment_number': ecm.assignment_number,
                        'tramite_status': ecm.tramite_status,
                        'assigned_at': ecm.assigned_at.isoformat() if ecm.assigned_at else None,
                        'expires_at': ecm.effective_expires_at.isoformat() if ecm.effective_expires_at else None,
                        'is_expired': ecm.is_expired,
                    } if ecm else None,
                    'results': [{
                        'id': r.id,
                        'score': r.score,
                        'status': r.status,
                        'result': r.result,
                        'start_date': r.start_date.isoformat() + 'Z' if r.start_date else None,
                        'end_date': r.end_date.isoformat() + 'Z' if r.end_date else None,
                        'duration_seconds': r.duration_seconds,
                        'certificate_code': r.certificate_code,
                        'eduit_certificate_code': r.eduit_certificate_code,
                    } for r in results],
                    'attempts_used': len([r for r in results if r.status == 1]),
                })

            groups_data.append({
                'group_id': group.id,
                'group_name': group.name,
                'group_code': group.code,
                'is_active': group.is_active,
                'start_date': group.start_date.isoformat() if group.start_date else None,
                'end_date': group.end_date.isoformat() if group.end_date else None,
                'campus': {
                    'id': campus.id, 'name': campus.name,
                    'city': campus.city if hasattr(campus, 'city') else None,
                } if campus else None,
                'cycle': {
                    'id': cycle.id, 'name': cycle.name,
                } if cycle else None,
                'membership_status': membership.status,
                'joined_at': membership.joined_at.isoformat() if membership.joined_at else None,
                'exams': exams_data,
            })

        return jsonify({
            'user_id': user_id,
            'user_name': user.full_name,
            'groups': groups_data,
            'total_groups': len(groups_data),
        })

    except HTTPException:
        raise
    except Exception as e:
        logger.exception('user_management error')
        return jsonify({'error': 'Error interno del servidor'}), 500


# ============== VERIFICACIÓN DE SIMILITUD DE NOMBRE ==============

@bp.route('/users/check-name-similarity', methods=['POST'])
@jwt_required()
@management_required
@rate_limit(limit=120, window=60, key_prefix='rl_um_namesim')  # UM-N1
def check_name_similarity():
    """
    Busca usuarios existentes con nombre similar (exacto o parcial).
    Se usa antes de crear un usuario para advertir sobre posibles duplicados.
    Retorna lista de coincidencias con su nivel de similitud.
    """
    try:
        data = request.get_json()
        name = (data.get('name') or '').strip()
        first_surname = (data.get('first_surname') or '').strip()
        second_surname = (data.get('second_surname') or '').strip()

        if not name or not first_surname:
            return jsonify({'similar_users': [], 'has_exact_match': False}), 200

        similar = []
        seen_ids = set()

        # 1. Coincidencia EXACTA: nombre + primer_apellido + segundo_apellido
        # UM-N1: filtrar is_deleted para no exponer PII de usuarios eliminados.
        exact_filters = [
            func.lower(func.ltrim(func.rtrim(User.name))) == name.lower(),
            func.lower(func.ltrim(func.rtrim(User.first_surname))) == first_surname.lower(),
            User.is_active == True,
            User.is_deleted == False,  # noqa: E712
        ]
        if second_surname:
            exact_filters.append(
                func.lower(func.ltrim(func.rtrim(User.second_surname))) == second_surname.lower()
            )

        exact_matches = User.query.filter(and_(*exact_filters)).limit(10).all()
        for u in exact_matches:
            if u.id not in seen_ids:
                seen_ids.add(u.id)
                similar.append({
                    'id': u.id,
                    'full_name': u.full_name,
                    'email': u.email,
                    'curp': u.curp,
                    'username': u.username,
                    'role': u.role,
                    'is_active': u.is_active,
                    'match_level': 'exact',
                    'match_description': 'Nombre completo idéntico',
                })

        # 2. Coincidencia PARCIAL: nombre + primer_apellido (sin segundo apellido)
        # UM-N1: filtrar is_deleted.
        if second_surname:
            partial_matches = User.query.filter(
                func.lower(func.ltrim(func.rtrim(User.name))) == name.lower(),
                func.lower(func.ltrim(func.rtrim(User.first_surname))) == first_surname.lower(),
                User.is_active == True,
                User.is_deleted == False,  # noqa: E712
                or_(
                    User.second_surname == None,
                    func.lower(func.ltrim(func.rtrim(User.second_surname))) != second_surname.lower()
                )
            ).limit(10).all()
            for u in partial_matches:
                if u.id not in seen_ids:
                    seen_ids.add(u.id)
                    similar.append({
                        'id': u.id,
                        'full_name': u.full_name,
                        'email': u.email,
                        'curp': u.curp,
                        'username': u.username,
                        'role': u.role,
                        'is_active': u.is_active,
                        'match_level': 'partial',
                        'match_description': f'Mismo nombre y primer apellido (segundo apellido diferente: {u.second_surname or "sin dato"})',
                    })

        # 3. Coincidencia por primer_apellido + segundo_apellido (mismos apellidos, nombre diferente)
        # UM-N1: filtrar is_deleted.
        if second_surname:
            surname_matches = User.query.filter(
                func.lower(func.ltrim(func.rtrim(User.first_surname))) == first_surname.lower(),
                func.lower(func.ltrim(func.rtrim(User.second_surname))) == second_surname.lower(),
                func.lower(func.ltrim(func.rtrim(User.name))) != name.lower(),
                User.is_active == True,
                User.is_deleted == False,  # noqa: E712
            ).limit(5).all()
            for u in surname_matches:
                if u.id not in seen_ids:
                    seen_ids.add(u.id)
                    similar.append({
                        'id': u.id,
                        'full_name': u.full_name,
                        'email': u.email,
                        'curp': u.curp,
                        'username': u.username,
                        'role': u.role,
                        'is_active': u.is_active,
                        'match_level': 'surname',
                        'match_description': f'Mismos apellidos, nombre diferente ({u.name})',
                    })

        has_exact = any(s['match_level'] == 'exact' for s in similar)

        return jsonify({
            'similar_users': similar[:20],
            'has_exact_match': has_exact,
            'total_found': len(similar),
        }), 200

    except HTTPException:

        raise

    except Exception as e:
        logger.exception('user_management error')
        return jsonify({'error': 'Error interno del servidor'}), 500


# ============== CREAR USUARIOS ==============

@bp.route('/users', methods=['POST'])
@jwt_required()
@management_required
@rate_limit(limit=60, window=60, key_prefix='rl_um_create')
def create_user():
    """Crear un nuevo usuario"""
    try:
        from app.models.partner import Campus
        
        current_user = g.current_user
        data = request.get_json()
        
        # Validar campos requeridos básicos (name, first_surname, role)
        required_fields = ['name', 'first_surname', 'role']
        for field in required_fields:
            if not data.get(field):
                return jsonify({'error': f'El campo {field} es requerido'}), 400
        
        role = data['role']

        # ── Asignación opcional a grupo (solo aplica para candidato) ──
        # Si se manda group_id y el rol es candidato:
        #   - Si email/CURP ya existe → asignar usuario existente al grupo (no 409)
        #   - Si no existe → crear y asignar en una sola operación
        # El acceso al grupo se valida con la misma regla que el bulk upload.
        target_group = None
        target_group_id = data.get('group_id')
        if target_group_id and role == 'candidato':
            try:
                target_group_id = int(target_group_id)
            except (TypeError, ValueError):
                return jsonify({'error': 'group_id inválido'}), 400
            target_group, group_err = _validate_target_group(target_group_id, current_user)
            if group_err:
                return group_err
        
        # Email es requerido para todos EXCEPTO candidatos (opcional para candidatos)
        if role != 'candidato' and not data.get('email'):
            return jsonify({'error': 'El campo email es requerido'}), 400
        
        # Para candidatos: CURP y email son opcionales
        # - Sin email: no puede recibir insignia digital
        # - Sin CURP: no puede recibir certificado CONOCER
        if role == 'candidato':
            # Solo second_surname y gender son obligatorios para candidatos
            candidato_required = ['second_surname', 'gender']
            field_names = {'second_surname': 'segundo apellido', 'gender': 'género'}
            for field in candidato_required:
                if not data.get(field):
                    return jsonify({'error': f'El campo {field_names[field]} es requerido para candidatos'}), 400
        
        # Para responsables, campos adicionales son obligatorios
        if role == 'responsable':
            responsable_required = ['second_surname', 'curp', 'gender', 'date_of_birth', 'campus_id']
            field_names = {
                'second_surname': 'segundo apellido',
                'curp': 'CURP',
                'gender': 'género',
                'date_of_birth': 'fecha de nacimiento',
                'campus_id': 'plantel'
            }
            for field in responsable_required:
                if not data.get(field):
                    return jsonify({'error': f'El campo {field_names[field]} es requerido para responsables'}), 400
            
            # Validar que el campus exista
            campus = Campus.query.get(data['campus_id'])
            if not campus:
                return jsonify({'error': 'El plantel especificado no existe'}), 404

            # UM-C5: validar que el campus pertenezca al tenant del creador.
            # Coordinador/auxiliar: el campus debe pertenecer a un partner del coordinador.
            if _is_coordinator_role(current_user.role):
                _eff_coord_id = _get_effective_coordinator_id(current_user)
                from app.models.partner import user_partners as _up
                _belongs = db.session.query(_up.c.partner_id).filter(
                    _up.c.user_id == _eff_coord_id,
                    _up.c.partner_id == campus.partner_id,
                ).first()
                if not _belongs:
                    return jsonify({'error': 'El plantel seleccionado no pertenece a tus partners'}), 403
            
            # Validar fecha de nacimiento
            from datetime import datetime as dt
            try:
                date_of_birth = dt.strptime(data['date_of_birth'], '%Y-%m-%d').date()
            except ValueError:
                return jsonify({'error': 'Formato de fecha de nacimiento inválido. Use YYYY-MM-DD'}), 400
            
            # Validar género
            if data['gender'] not in ['M', 'F', 'O']:
                return jsonify({'error': 'Género inválido. Use M (masculino), F (femenino) u O (otro)'}), 400
        
        # Para responsable_partner o responsable_estatal, partner_id es obligatorio
        if role in ('responsable_partner', 'responsable_estatal'):
            if not data.get('partner_id'):
                return jsonify({'error': 'Debe seleccionar un partner para el responsable'}), 400
            from app.models.partner import Partner, user_partners as _up_check
            partner = Partner.query.get(data['partner_id'])
            if not partner:
                return jsonify({'error': 'El partner especificado no existe'}), 404
            # UM-C4: coordinador/auxiliar solo puede asignar partners propios
            if _is_coordinator_role(current_user.role):
                _eff = _get_effective_coordinator_id(current_user)
                _ok = db.session.query(_up_check.c.partner_id).filter(
                    _up_check.c.user_id == _eff,
                    _up_check.c.partner_id == data['partner_id'],
                ).first()
                if not _ok:
                    return jsonify({'error': 'El partner seleccionado no está asignado a tu cuenta'}), 403
            # Para responsable_estatal, assigned_state es obligatorio
            if role == 'responsable_estatal' and not data.get('assigned_state'):
                return jsonify({'error': 'Debe seleccionar un estado para el responsable estatal'}), 400
        
        # Email es opcional para candidatos
        email = None
        if data.get('email'):
            email = data['email'].strip().lower()
            
            # Validar formato email
            if not validate_email(email):
                return jsonify({'error': 'Formato de email inválido'}), 400
            
            # Verificar email único (solo si se proporciona)
            # Roles soporte, editor, editor_invitado y coordinator pueden compartir email con otros usuarios
            if role not in ('soporte', 'editor', 'editor_invitado', 'coordinator'):
                existing_email_user = User.query.filter_by(email=email).first()
                if existing_email_user:
                    # Si el rol es candidato y se especificó un grupo destino,
                    # en lugar de bloquear con 409 reasignamos al usuario existente al grupo
                    # (paridad con el flujo bulk: usuarios pueden pertenecer a múltiples grupos).
                    if role == 'candidato' and target_group:
                        return _assign_existing_to_group_response(existing_email_user, target_group, 'email')
                    return _duplicate_email_response(email, existing_email_user)
        
        # Generar contraseña automática para TODOS los usuarios
        password = generate_secure_password(12)
        
        # Verificar permisos de creación según rol
        # Auxiliar hereda los permisos de creación del coordinator
        effective_role = 'coordinator' if current_user.role == 'auxiliar' else current_user.role
        allowed_roles = ROLE_CREATE_PERMISSIONS.get(effective_role, [])
        if role not in allowed_roles:
            if _is_coordinator_role(current_user.role):
                return jsonify({'error': 'Los coordinadores solo pueden crear candidatos, responsables, responsables del partner y auxiliares'}), 403
            else:
                return jsonify({'error': f'No tienes permiso para crear usuarios con rol {role}'}), 403
        
        # Verificar CURP único si se proporciona (solo para roles diferentes a editor/editor_invitado/gerente/financiero)
        # Los CURPs genéricos de extranjero pueden repetirse
        if data.get('curp') and role not in ['editor', 'editor_invitado', 'gerente', 'financiero']:
            curp = data['curp'].upper().strip()
            if not _is_generic_foreign_curp(curp):
                existing_curp_user = User.query.filter_by(curp=curp).first()
                if existing_curp_user:
                    # Si el rol es candidato y hay grupo destino, asignar al existente
                    if role == 'candidato' and target_group:
                        return _assign_existing_to_group_response(existing_curp_user, target_group, 'curp')
                    return _duplicate_curp_response(curp, existing_curp_user)
        
        # Generar username automáticamente (10 caracteres alfanuméricos únicos EN MAYÚSCULAS)
        import random
        import string
        
        def generate_unique_username():
            """Genera un username único de 10 caracteres alfanuméricos en MAYÚSCULAS.
            Excluye caracteres confusos: I, L, O, 0
            """
            _CONFUSING = set('ILO0')
            chars = ''.join(c for c in string.ascii_uppercase + string.digits if c not in _CONFUSING)
            while True:
                username = ''.join(random.choices(chars, k=10))
                if not User.query.filter_by(username=username).first():
                    return username
        
        username = generate_unique_username()
        
        # Para usuarios tipo editor o candidato, no guardar phone
        # Para usuarios tipo editor/gerente/financiero, no guardar CURP ni phone
        user_curp = None
        user_phone = None
        user_date_of_birth = None
        user_campus_id = None
        
        if role not in ['editor', 'editor_invitado', 'candidato', 'gerente', 'financiero']:
            # Solo admin/coordinator/responsable pueden tener teléfono
            user_phone = data.get('phone', '').strip() or None
        if role not in ['editor', 'editor_invitado', 'gerente', 'financiero']:
            # Solo candidatos, responsables y otros roles (no editor/gerente/financiero) pueden tener CURP
            user_curp = data.get('curp', '').upper().strip() or None
        
        # Para responsables, campos adicionales
        if role == 'responsable':
            from datetime import datetime as dt
            user_date_of_birth = dt.strptime(data['date_of_birth'], '%Y-%m-%d').date()
            user_campus_id = data['campus_id']
        
        # Para candidatos creados por responsable, asignar campus_id del responsable
        if role == 'candidato' and current_user.role == 'responsable' and current_user.campus_id:
            user_campus_id = current_user.campus_id
        # Crear usuario
        # Determinar coordinator_id para multi-tenancy
        # Auxiliar hereda el coordinator_id de su creador (mismo tenant que el coordinador)
        effective_coord_id = _get_effective_coordinator_id(current_user)
        user_coordinator_id = None
        if role == 'responsable':
            # ── Regla de negocio: responsable DEBE tener coordinator_id ──
            if current_user.role == 'coordinator':
                # El coordinador que crea → se liga automáticamente
                user_coordinator_id = current_user.id
            elif current_user.role == 'auxiliar' and effective_coord_id:
                # Auxiliar hereda el coordinator_id de su coordinador
                user_coordinator_id = effective_coord_id
            elif data.get('coordinator_id'):
                # Admin/developer/soporte eligen coordinador del dropdown
                coord_check = User.query.get(data['coordinator_id'])
                if not coord_check or coord_check.role != 'coordinator' or not coord_check.is_active:
                    return jsonify({'error': 'El coordinador seleccionado no es válido o no está activo'}), 400
                user_coordinator_id = data['coordinator_id']
            else:
                return jsonify({'error': 'Debe seleccionar un coordinador para el responsable'}), 400
        elif _is_coordinator_role(current_user.role) and role in ['candidato', 'responsable_partner', 'responsable_estatal', 'auxiliar']:
            # Coordinador o auxiliar: ambos quedan ligados al coordinador efectivo
            # Incluye creación de auxiliar por parte del coordinador
            user_coordinator_id = effective_coord_id
        elif current_user.role == 'responsable' and role == 'candidato':
            # UM-C7: candidato creado por responsable hereda el coordinator_id
            # del responsable para preservar multi-tenancy.
            user_coordinator_id = current_user.coordinator_id
        elif current_user.role in ['admin', 'developer', 'soporte'] and role == 'auxiliar' and data.get('coordinator_id'):
            # Admin/developer/soporte creando auxiliar: validar y persistir coordinator_id del body
            coord_check_aux = User.query.get(data['coordinator_id'])
            if not coord_check_aux or coord_check_aux.role != 'coordinator' or not coord_check_aux.is_active:
                return jsonify({'error': 'El coordinador seleccionado para el auxiliar no es válido o no está activo'}), 400
            user_coordinator_id = data['coordinator_id']
        elif current_user.role in ['admin', 'developer', 'soporte'] and role == 'candidato' and user_campus_id:
            # Admin/developer creando responsable/candidato: buscar el coordinador del campus
            from app.models.partner import Campus as CampusLookup, Partner as PartnerLookup, user_partners as up_table
            campus_lookup = CampusLookup.query.get(user_campus_id)
            if campus_lookup and campus_lookup.partner_id:
                # Buscar coordinador asociado al partner del campus
                coord_row = db.session.query(up_table.c.user_id).join(
                    User, User.id == up_table.c.user_id
                ).filter(
                    up_table.c.partner_id == campus_lookup.partner_id,
                    User.role == 'coordinator',
                    User.is_active == True
                ).first()
                if coord_row:
                    user_coordinator_id = coord_row[0]
        elif current_user.role in ['admin', 'developer'] and role in ('responsable_partner', 'responsable_estatal') and data.get('partner_id'):
            # Admin creando responsable_partner/estatal: buscar el coordinador del partner
            from app.models.partner import user_partners as up_table2
            coord_row2 = db.session.query(up_table2.c.user_id).join(
                User, User.id == up_table2.c.user_id
            ).filter(
                up_table2.c.partner_id == data['partner_id'],
                User.role == 'coordinator',
                User.is_active == True
            ).first()
            if coord_row2:
                user_coordinator_id = coord_row2[0]
        
        # ── Validación CURP ──
        # Mientras RENAPO está en standby (feature flag CURP_RENAPO_ENABLED=false),
        # hacemos validación LOCAL síncrona: formato + dígito verificador.
        # Si el flag está ON, conservamos el flujo original (cache + encolado).
        renapo_result = None
        renapo_rejected = False
        renapo_error_msg = None
        renapo_unavailable = False  # True si debemos encolar para retry async
        local_validated = False     # True si pasó validación local (RENAPO OFF)
        skip_renapo = data.get('skip_renapo_validation', False)

        from app.services.curp_local_validator import is_renapo_enabled
        renapo_on = is_renapo_enabled()

        if user_curp and not _is_generic_foreign_curp(user_curp) and not skip_renapo:
            if not renapo_on:
                # ─── Modo LOCAL: validación síncrona de formato ───
                from app.services.curp_local_validator import validate_curp_local
                is_valid_local, local_err, _ = validate_curp_local(user_curp)
                if not is_valid_local:
                    return jsonify({
                        'error': f'CURP inválida: {local_err}',
                        'reason': 'curp_format',
                    }), 400
                local_validated = True
            else:
                # ─── Modo RENAPO: cache-only + encolado async (legacy) ───
                try:
                    from app.services.renapo_service import (
                        validate_curp_renapo, apply_renapo_to_user,
                    )
                    # cache_only=True: nunca llama a RENAPO; sólo lee cache.
                    cached = validate_curp_renapo(user_curp, use_cache=True, cache_only=True)
                    if cached.valid:
                        # Hit positivo en cache — aplicar de inmediato.
                        renapo_result = cached
                    else:
                        # Cache miss o negativo: encolar para validación async.
                        renapo_unavailable = True
                        if cached.error and cached.error != 'cache_miss':
                            renapo_error_msg = cached.error
                except Exception as renapo_err:
                    import logging
                    logging.getLogger(__name__).warning(f'RENAPO no disponible: {renapo_err}')
                    renapo_result = None
                    renapo_unavailable = True

        # Si el alta es de candidato con grupo destino, derivar campus_id y coordinator_id
        # del grupo (paridad con bulk upload B+C — consistencia y filtros multi-tenant).
        if role == 'candidato' and target_group:
            user_campus_id = target_group.campus_id
            user_coordinator_id = target_group.coordinator_id

        new_user = User(
            id=str(uuid.uuid4()),
            email=email,
            username=username,
            name=data['name'].strip().upper(),
            first_surname=data['first_surname'].strip().upper(),
            second_surname=data.get('second_surname', '').strip().upper() or None,
            gender=data.get('gender'),
            curp=user_curp,
            phone=user_phone,
            role=role,
            campus_id=user_campus_id,
            date_of_birth=user_date_of_birth,
            coordinator_id=user_coordinator_id,
            is_active=data.get('is_active', True),
            is_verified=True if role == 'responsable' else data.get('is_verified', False),
            can_bulk_create_candidates=data.get('can_bulk_create_candidates', False) if role == 'responsable' else False,
            can_manage_groups=data.get('can_manage_groups', False) if role == 'responsable' else False,
            can_view_reports=data.get('can_view_reports', True) if role == 'responsable' else True,
            assigned_state=data.get('assigned_state', '').strip() or None if role == 'responsable_estatal' else None
        )
        new_user.set_password(password)  # Usar la variable password (puede ser generada automáticamente)
        new_user.encrypted_password = encrypt_password(password)
        
        # Aplicar datos RENAPO si la validación fue exitosa
        if renapo_result and renapo_result.valid:
            from app.services.renapo_service import apply_renapo_to_user
            apply_renapo_to_user(new_user, renapo_result)
        elif local_validated:
            # Modo local: marcar verified=True (sin datos RENAPO porque no
            # tenemos nombre/apellidos oficiales, solo el formato pasa).
            from app.services.curp_local_validator import apply_local_validation_to_user
            apply_local_validation_to_user(new_user, mark_verified=True)
        elif _is_generic_foreign_curp(user_curp) if user_curp else False:
            # CURP genérica de extranjero — aceptar sin RENAPO ni local
            new_user.curp_verified = True
            if hasattr(new_user, 'curp_verified_at'):
                from datetime import datetime as _dt_gfc
                new_user.curp_verified_at = _dt_gfc.utcnow()
        
        db.session.add(new_user)
        db.session.flush()  # Persistir el user para que exista antes de la FK en user_partners
        
        # Para responsable_partner o responsable_estatal, crear la asociación en user_partners
        if role in ('responsable_partner', 'responsable_estatal') and data.get('partner_id'):
            from app.models.partner import user_partners
            db.session.execute(user_partners.insert().values(
                user_id=new_user.id,
                partner_id=data['partner_id']
            ))
        
        # Para responsable, asociarlo como responsable del plantel
        if role == 'responsable' and user_campus_id:
            from app.models.partner import Campus as CampusModel
            campus_obj = CampusModel.query.get(user_campus_id)
            if campus_obj:
                campus_obj.responsable_id = new_user.id
                # Avanzar estado de activación si está pendiente
                if campus_obj.activation_status == 'pending':
                    campus_obj.activation_status = 'configuring'

        # Asignar al grupo destino si aplica (solo candidatos)
        if role == 'candidato' and target_group:
            from app.models.partner import GroupMember
            db.session.add(GroupMember(
                group_id=target_group.id,
                user_id=new_user.id,
                status='active',
            ))

        db.session.commit()
        
        # Encolar verificación CURP si RENAPO no respondió o rechazó.
        # Cubre W1 (alta individual sin retries) + W2 (circuit abierto).
        # NOTA: si CURP_RENAPO_ENABLED=false (modo local), NUNCA encolamos.
        if renapo_on and renapo_unavailable and user_curp and not _is_generic_foreign_curp(user_curp):
            try:
                from app.services.curp_queue_worker import enqueue_curp_verification
                enqueue_curp_verification(new_user.id, user_curp, source='individual')
                import logging
                logging.getLogger(__name__).info(
                    f'[USERS-CURP] Encolado user={new_user.id} curp={user_curp} para verificación'
                )
            except Exception as enq_err:
                import logging
                logging.getLogger(__name__).error(
                    f'[USERS-CURP] Error encolando verificación: {enq_err}'
                )
        
        # Enviar email de bienvenida si tiene email
        if new_user.email:
            try:
                from app.services.email_service import send_welcome_email
                send_welcome_email(new_user, password)
            except Exception as email_err:
                import logging
                logging.getLogger(__name__).error(f'Error enviando welcome email: {email_err}')
        
        # Incluir la contraseña temporal en la respuesta para TODOS los roles
        response_data = {
            'message': 'Usuario creado exitosamente',
            'user': new_user.to_dict(include_private=True),
            'temporary_password': password
        }

        # Incluir info de asignación a grupo si aplica
        if target_group and role == 'candidato':
            response_data['group_id'] = target_group.id
            response_data['group_name'] = target_group.name
            response_data['assigned_to_group'] = True
        
        # Incluir info RENAPO si se validó
        if renapo_result and renapo_result.valid:
            response_data['renapo_validated'] = True
            response_data['renapo_data'] = renapo_result.to_dict()
        elif local_validated:
            response_data['curp_validated_locally'] = True
        elif renapo_rejected:
            response_data['curp_pending_validation'] = True
            response_data['renapo_warning'] = (
                f'La CURP no pudo validarse contra RENAPO ({renapo_error_msg}). '
                f'El usuario fue creado pero deberá validarla al iniciar sesión.'
            )
        elif renapo_unavailable:
            response_data['curp_pending_validation'] = True
            response_data['renapo_warning'] = (
                'El servicio RENAPO no está disponible en este momento. '
                'El usuario fue creado y la CURP se validará automáticamente en segundo plano.'
            )
        
        return jsonify(response_data), 201
        
    except HTTPException:
        
        raise
        
    except Exception as e:
        db.session.rollback()
        logger.exception('user_management error')
        return jsonify({'error': 'Error interno del servidor'}), 500


# ============== ACTUALIZAR USUARIOS ==============

@bp.route('/users/<string:user_id>', methods=['PUT'])
@jwt_required()
@management_required
def update_user(user_id):
    """Actualizar un usuario"""
    try:
        current_user = g.current_user
        user = User.query.get_or_404(user_id)
        data = request.get_json()
        
        # Coordinadores solo pueden editar candidatos, responsables y responsables del partner
        if _is_coordinator_role(current_user.role):
            err = _verify_coordinator_user_access(current_user, user)
            if err:
                return err

        # Soporte solo puede editar candidatos, responsables y responsables_partner
        if current_user.role == 'soporte':
            if user.role not in SOPORTE_EDITABLE_ROLES:
                return jsonify({'error': 'Solo puedes editar candidatos, responsables y responsables del partner'}), 403

        # Responsable solo puede editar candidatos de su plantel
        if current_user.role == 'responsable':
            if user.role != 'candidato':
                return jsonify({'error': 'Solo puedes editar candidatos'}), 403
            from app.models.partner import GroupMember, CandidateGroup
            is_in_campus = db.session.query(GroupMember.id).join(
                CandidateGroup, GroupMember.group_id == CandidateGroup.id
            ).filter(
                GroupMember.user_id == user_id,
                CandidateGroup.campus_id == current_user.campus_id
            ).first()
            if not is_in_campus and user.campus_id != current_user.campus_id:
                return jsonify({'error': 'Este candidato no pertenece a tu plantel'}), 403
        
        # No se puede editar a uno mismo por esta ruta (usar perfil)
        if user_id == current_user.id:
            return jsonify({'error': 'Usa la opción de perfil para editar tu propia cuenta'}), 400
        
        # Campos actualizables
        basic_fields = ['name', 'first_surname', 'second_surname', 'gender']
        # Solo actualizar phone si el usuario NO es editor/editor_invitado ni candidato
        if user.role not in ['editor', 'editor_invitado', 'candidato']:
            basic_fields.append('phone')
        name_fields = {'name', 'first_surname', 'second_surname'}
        for field in basic_fields:
            if field in data:
                val = data[field].strip() if data[field] else None
                if val and field in name_fields:
                    val = val.upper()
                setattr(user, field, val)
        
        # CURP - verificar unicidad (solo si el usuario no es editor/editor_invitado)
        # Los CURPs genéricos de extranjero pueden repetirse
        if 'curp' in data and user.role not in ['editor', 'editor_invitado']:
            curp = data['curp'].upper().strip() if data['curp'] else None
            # UM-H2: validar formato antes de cualquier commit
            if curp and not _is_generic_foreign_curp(curp):
                try:
                    from app.services.renapo_service import validate_curp_format
                    fmt_valid, fmt_error = validate_curp_format(curp)
                    if not fmt_valid:
                        return jsonify({'error': f'CURP inválida: {fmt_error}'}), 400
                except Exception:
                    pass
                existing = User.query.filter(User.curp == curp, User.id != user_id).first()
                if existing:
                    return _duplicate_curp_response(curp, existing)

            # UM-H1: si la CURP cambia (incluido pasar a None o a genérica),
            # resetear todas las marcas de verificación RENAPO.
            if curp != user.curp:
                user.curp_verified = False
                user.curp_verified_at = None
                user.curp_renapo_name = None
                user.curp_renapo_first_surname = None
                user.curp_renapo_second_surname = None
            user.curp = curp
        
        # Email - verificar unicidad (solo si se proporciona un valor no vacío)
        if 'email' in data and data['email'] and data['email'].strip():
            email = data['email'].strip().lower()
            if not validate_email(email):
                return jsonify({'error': 'Formato de email inválido'}), 400
            # Roles soporte, editor, editor_invitado y coordinator pueden compartir email con otros usuarios
            if user.role not in ('soporte', 'editor', 'editor_invitado', 'coordinator'):
                existing = User.query.filter(User.email == email, User.id != user_id).first()
                if existing:
                    return _duplicate_email_response(email, existing)
            user.email = email
        
        # Campos de responsable (editables por admin, developer, soporte, coordinator y auxiliar)
        if user.role == 'responsable' and current_user.role in ['admin', 'developer', 'soporte', 'coordinator', 'auxiliar']:
            if 'date_of_birth' in data:
                if data['date_of_birth']:
                    user.date_of_birth = datetime.strptime(data['date_of_birth'], '%Y-%m-%d').date()
                else:
                    user.date_of_birth = None
            if 'campus_id' in data and data['campus_id']:
                from app.models.partner import Campus
                campus = Campus.query.get(data['campus_id'])
                if not campus:
                    return jsonify({'error': 'Plantel no encontrado'}), 400
                # UM-C5: validar tenant scope para coordinador/auxiliar
                if _is_coordinator_role(current_user.role):
                    _eff = _get_effective_coordinator_id(current_user)
                    from app.models.partner import user_partners as _up
                    _ok = db.session.query(_up.c.partner_id).filter(
                        _up.c.user_id == _eff,
                        _up.c.partner_id == campus.partner_id,
                    ).first()
                    if not _ok:
                        return jsonify({'error': 'El plantel seleccionado no pertenece a tus partners'}), 403
                user.campus_id = data['campus_id']
            if 'can_bulk_create_candidates' in data:
                user.can_bulk_create_candidates = bool(data['can_bulk_create_candidates'])
            if 'can_manage_groups' in data:
                user.can_manage_groups = bool(data['can_manage_groups'])
            if 'can_view_reports' in data:
                user.can_view_reports = bool(data['can_view_reports'])
        
        # Cambio de coordinador asignado (admin, developer, gerente, soporte)
        if user.role == 'responsable' and 'coordinator_id' in data:
            if current_user.role in ['admin', 'developer', 'gerente', 'soporte']:
                new_coord_id = data['coordinator_id']
                if new_coord_id:
                    coord = User.query.get(new_coord_id)
                    if not coord or coord.role != 'coordinator' or not coord.is_active:
                        return jsonify({'error': 'El coordinador seleccionado no es válido o no está activo'}), 400
                    user.coordinator_id = new_coord_id
                else:
                    user.coordinator_id = None
            else:
                return jsonify({'error': 'No tienes permisos para cambiar el coordinador asignado'}), 403
        
        # Campos de responsable_partner/responsable_estatal (editables por admin, developer, coordinator y auxiliar)
        if user.role in ('responsable_partner', 'responsable_estatal') and current_user.role in ['admin', 'developer', 'coordinator', 'auxiliar']:
            if 'partner_id' in data and data['partner_id']:
                from app.models.partner import Partner, user_partners
                partner = Partner.query.get(data['partner_id'])
                if not partner:
                    return jsonify({'error': 'Partner no encontrado'}), 400
                # UM-C4: validar que el partner sea del tenant del coordinador
                if _is_coordinator_role(current_user.role):
                    _eff2 = _get_effective_coordinator_id(current_user)
                    _ok2 = db.session.query(user_partners.c.partner_id).filter(
                        user_partners.c.user_id == _eff2,
                        user_partners.c.partner_id == data['partner_id'],
                    ).first()
                    if not _ok2:
                        return jsonify({'error': 'El partner seleccionado no está asignado a tu cuenta'}), 403
                # Eliminar relaciones anteriores
                db.session.execute(user_partners.delete().where(user_partners.c.user_id == user.id))
                # Crear nueva relación
                db.session.execute(user_partners.insert().values(user_id=user.id, partner_id=data['partner_id']))
            if 'assigned_state' in data and user.role == 'responsable_estatal':
                user.assigned_state = data['assigned_state'].strip() if data['assigned_state'] else None
        
        # Campos que solo admin puede cambiar
        if current_user.role in ['admin', 'developer']:
            if 'role' in data:
                new_role = data['role']
                # Admin/developer no puede crear otros admins o developers
                if new_role in ['admin', 'developer'] and user.role not in ['admin', 'developer']:
                    return jsonify({'error': 'No se puede asignar rol de administrador o desarrollador'}), 403
                # Si cambia a editor/editor_invitado o candidato, limpiar phone
                if new_role in ['editor', 'editor_invitado', 'candidato'] and user.role not in ['editor', 'editor_invitado', 'candidato']:
                    user.phone = None
                # Si cambia a editor/editor_invitado, también limpiar CURP
                if new_role in ['editor', 'editor_invitado'] and user.role not in ['editor', 'editor_invitado']:
                    user.curp = None
                user.role = new_role
            
            if 'is_active' in data:
                user.is_active = data['is_active']
            
            if 'is_verified' in data:
                user.is_verified = data['is_verified']
        
        # Coordinadores también pueden activar/desactivar candidatos y responsables
        if _is_coordinator_role(current_user.role) and 'is_active' in data:
            user.is_active = data['is_active']
        
        db.session.commit()
        
        # ── Verificación CURP (si se envió CURP y aún no está verificada) ──
        # Si CURP_RENAPO_ENABLED=false (default): validación LOCAL síncrona.
        # Si está ON: flujo legacy con RENAPO síncrono.
        curp_msg = ''
        if 'curp' in data and user.curp and not user.curp_verified and not _is_generic_foreign_curp(user.curp):
            import logging
            logger = logging.getLogger(__name__)
            from app.services.curp_local_validator import is_renapo_enabled as _renapo_on_upd
            if not _renapo_on_upd():
                # ─── Modo LOCAL ───
                from app.services.curp_local_validator import (
                    validate_curp_local, apply_local_validation_to_user,
                )
                is_v, local_err, _gen = validate_curp_local(user.curp)
                if is_v:
                    apply_local_validation_to_user(user, mark_verified=True)
                    db.session.commit()
                    curp_msg = '. CURP verificada (formato)'
                else:
                    logger.warning(f'[CURP-LOCAL] Formato inválido CURP={user.curp} user={user.id}: {local_err}')
                    user.curp = None
                    user.curp_verified = False
                    user.curp_verified_at = None
                    db.session.commit()
                    curp_msg = f'. CURP rechazada: {local_err}'
            else:
                try:
                    from app.services.renapo_service import (
                        validate_curp_renapo, apply_renapo_to_user, validate_curp_format,
                    )

                    fmt_valid, fmt_error = validate_curp_format(user.curp)
                    if not fmt_valid:
                        logger.warning(f'[CURP] Formato inválido para CURP {user.curp} de usuario {user.id}: {fmt_error}')
                        user.curp = None
                        user.curp_verified = False
                        user.curp_verified_at = None
                        user.curp_renapo_name = None
                        user.curp_renapo_first_surname = None
                        user.curp_renapo_second_surname = None
                        db.session.commit()
                        curp_msg = '. CURP rechazada: formato inválido'
                    else:
                        renapo_result = validate_curp_renapo(user.curp)
                        if renapo_result.valid:
                            apply_renapo_to_user(user, renapo_result)
                            db.session.commit()
                            curp_msg = '. CURP verificada contra RENAPO'
                        else:
                            # CURP rechazada por RENAPO: no removemos el dato pero queda como no verificado
                            logger.warning(f'[CURP] CURP {user.curp} rechazada por RENAPO para usuario {user.id}: {renapo_result.error}')
                            curp_msg = f'. CURP no pudo verificarse en RENAPO: {renapo_result.error}'
                except Exception as e:
                    # Si RENAPO está caído, no fallar la actualización; quedará pendiente
                    logger.warning(f'[CURP] RENAPO no disponible al verificar CURP de usuario {user.id}: {e}')
                    curp_msg = '. La CURP quedará pendiente de verificación (RENAPO no disponible)'
        
        return jsonify({
            'message': f'Usuario actualizado exitosamente{curp_msg}',
            'user': user.to_dict(include_private=True)
        })
        
    except HTTPException:
        
        raise
        
    except Exception as e:
        db.session.rollback()
        logger.exception('user_management error')
        return jsonify({'error': 'Error interno del servidor'}), 500


# ============== CAMBIAR CONTRASEÑA ==============

@bp.route('/users/<string:user_id>/password', methods=['PUT'])
@jwt_required()
@management_required
@rate_limit(limit=30, window=60, key_prefix='rl_um_chpwd')
def change_user_password(user_id):
    """Cambiar contraseña de un usuario (sin requerir contraseña actual)"""
    try:
        current_user = g.current_user
        user = User.query.get_or_404(user_id)
        data = request.get_json()
        
        # Coordinadores/auxiliares solo pueden cambiar contraseñas de sus propios candidatos/responsables
        if _is_coordinator_role(current_user.role):
            err = _verify_coordinator_user_access(current_user, user)
            if err:
                return err

        # Soporte solo puede cambiar contraseñas de candidatos, responsables y responsables_partner
        if current_user.role == 'soporte':
            if user.role not in SOPORTE_EDITABLE_ROLES:
                return jsonify({'error': 'Solo puedes cambiar contraseñas de candidatos, responsables y responsables del partner'}), 403

        new_password = data.get('new_password')
        if not new_password:
            return jsonify({'error': 'La nueva contraseña es requerida'}), 400
        
        is_valid, error_msg = validate_password(new_password)
        if not is_valid:
            return jsonify({'error': error_msg}), 400
        
        user.set_password(new_password)
        user.encrypted_password = encrypt_password(new_password)
        db.session.commit()
        
        return jsonify({
            'message': 'Contraseña actualizada exitosamente'
        })
        
    except HTTPException:
        
        raise
        
    except Exception as e:
        db.session.rollback()
        logger.exception('user_management error')
        return jsonify({'error': 'Error interno del servidor'}), 500


# ============== GENERAR CONTRASEÑA TEMPORAL (ADMIN Y COORDINADOR) ==============

@bp.route('/users/<string:user_id>/generate-password', methods=['POST'])
@jwt_required()
@management_required
@rate_limit(limit=30, window=60, key_prefix='rl_um_genpwd')
def generate_temp_password(user_id):
    """Generar una contraseña temporal para un usuario (admin y coordinador)"""
    try:
        import secrets
        
        current_user = g.current_user
        user = User.query.get_or_404(user_id)
        
        # Coordinadores/auxiliares solo pueden generar contraseñas de sus propios candidatos/responsables
        if _is_coordinator_role(current_user.role):
            err = _verify_coordinator_user_access(current_user, user)
            if err:
                return err

        # Soporte solo puede generar contraseñas de candidatos, responsables y responsables_partner
        if current_user.role == 'soporte':
            if user.role not in SOPORTE_EDITABLE_ROLES:
                return jsonify({'error': 'Solo puedes generar contraseñas de candidatos, responsables y responsables del partner'}), 403

        # Generar contraseña segura de 12 caracteres
        temp_password = generate_secure_password(12)
        
        # Establecer la nueva contraseña
        user.set_password(temp_password)
        user.encrypted_password = encrypt_password(temp_password)
        db.session.commit()
        
        return jsonify({
            'message': 'Contraseña temporal generada exitosamente',
            'password': temp_password,
            'user': user.to_dict(include_private=True)
        })
        
    except HTTPException:
        
        raise
        
    except Exception as e:
        db.session.rollback()
        logger.exception('user_management error')
        return jsonify({'error': 'Error interno del servidor'}), 500


# ============== VER CONTRASEÑA DE USUARIO (ADMIN Y COORDINADOR) ==============

@bp.route('/users/<string:user_id>/password', methods=['GET'])
@jwt_required()
@management_required
@rate_limit(limit=60, window=60, key_prefix='rl_um_getpwd')
def get_user_password(user_id):
    """Obtener la contraseña descifrada de un usuario (admin y coordinador)"""
    try:
        current_user = g.current_user
        user = User.query.get_or_404(user_id)
        
        # Coordinadores/auxiliares solo pueden ver contraseñas de sus propios candidatos/responsables
        if _is_coordinator_role(current_user.role):
            err = _verify_coordinator_user_access(current_user, user)
            if err:
                return err

        # Soporte solo puede ver contraseñas de candidatos, responsables y responsables_partner
        if current_user.role == 'soporte':
            if user.role not in SOPORTE_EDITABLE_ROLES:
                return jsonify({'error': 'Solo puedes ver contraseñas de candidatos, responsables y responsables del partner'}), 403

        # Responsable solo puede ver contraseñas de candidatos de su plantel
        if current_user.role == 'responsable':
            if user.role != 'candidato':
                return jsonify({'error': 'Solo puedes ver contraseñas de candidatos'}), 403
            from app.models.partner import GroupMember, CandidateGroup
            in_campus = db.session.query(GroupMember.id).join(
                CandidateGroup, GroupMember.group_id == CandidateGroup.id
            ).filter(
                GroupMember.user_id == user_id,
                CandidateGroup.campus_id == current_user.campus_id
            ).first()
            # UM-M2: aceptar también candidatos cuyo campus_id directo coincide
            # (los que aún no fueron asignados a un grupo del plantel).
            if not in_campus and user.campus_id != current_user.campus_id:
                return jsonify({'error': 'Este candidato no pertenece a tu plantel'}), 403
        
        # Desencriptar la contraseña
        decrypted_password = user.get_decrypted_password()
        
        if not decrypted_password:
            return jsonify({
                'error': 'No se puede recuperar la contraseña. Genera una nueva contraseña temporal.',
                'has_password': False
            }), 404
        
        return jsonify({
            'password': decrypted_password,
            'has_password': True,
            'user': {
                'id': user.id,
                'email': user.email,
                'full_name': user.full_name
            }
        })
        
    except HTTPException:
        
        raise
        
    except Exception as e:
        logger.exception('user_management error')
        return jsonify({'error': 'Error interno del servidor'}), 500


# ============== ACTIVAR/DESACTIVAR USUARIOS ==============

@bp.route('/users/<string:user_id>/toggle-active', methods=['POST'])
@jwt_required()
@management_required
@rate_limit(limit=60, window=60, key_prefix='rl_um_toggle')  # UM-N4
def toggle_user_active(user_id):
    """Activar o desactivar un usuario"""
    try:
        current_user = g.current_user
        user = User.query.get_or_404(user_id)
        
        # Solo admin/coordinator/soporte pueden activar/desactivar usuarios
        # Coordinadores solo pueden activar/desactivar candidatos, responsables y responsables_partner
        if current_user.role == 'responsable':
            return jsonify({'error': 'Solo administradores, coordinadores o soporte pueden activar/desactivar usuarios'}), 403

        if _is_coordinator_role(current_user.role):
            err = _verify_coordinator_user_access(current_user, user)
            if err:
                return err

        # Soporte solo puede activar/desactivar candidatos, responsables y responsables_partner
        if current_user.role == 'soporte':
            if user.role not in SOPORTE_EDITABLE_ROLES:
                return jsonify({'error': 'Solo puedes activar/desactivar candidatos, responsables y responsables del partner'}), 403
        
        # No se puede desactivar a uno mismo
        if user_id == current_user.id:
            return jsonify({'error': 'No puedes desactivarte a ti mismo'}), 400
        
        was_inactive = not user.is_active
        user.is_active = not user.is_active
        db.session.commit()
        _invalidate_user_caches()  # UM-M6
        
        status = 'activado' if user.is_active else 'desactivado'

        # Si se activa y tiene CURP no extranjera sin verificar → validar RENAPO en background
        if user.is_active and was_inactive and user.curp and not user.curp_verified:
            from app.services.renapo_service import is_generic_foreign_curp
            if not is_generic_foreign_curp(user.curp):
                import threading
                from flask import current_app
                app = current_app._get_current_object()
                _user_id = user.id
                _curp = user.curp
                _username = user.username

                def _verify_curp_on_activate(app_obj, uid, curp, username):
                    import time
                    import logging
                    _logger = logging.getLogger(__name__)
                    _logger.info(f'[ACTIVATE-CURP] Verificando CURP {curp} de {username}')
                    time.sleep(1)  # Dar tiempo a que termine la respuesta HTTP
                    with app_obj.app_context():
                        from app.services.renapo_service import validate_curp_renapo, apply_renapo_to_user
                        try:
                            result = validate_curp_renapo(curp)
                            u = User.query.get(uid)
                            if not u:
                                _logger.warning(f'[ACTIVATE-CURP] Usuario {username} ya no existe')
                                return
                            if result.valid:
                                apply_renapo_to_user(u, result)
                                db.session.commit()
                                _logger.info(f'[ACTIVATE-CURP] CURP {curp} válida para {username}')
                            else:
                                _logger.warning(f'[ACTIVATE-CURP] CURP {curp} inválida para {username}: {result.error}')
                        except Exception as e:
                            _logger.error(f'[ACTIVATE-CURP] Error verificando CURP {curp} de {username}: {e}')
                            try:
                                db.session.rollback()
                            except Exception:
                                pass

                threading.Thread(
                    target=_verify_curp_on_activate,
                    args=(app, _user_id, _curp, _username),
                    daemon=True,
                ).start()
        
        return jsonify({
            'message': f'Usuario {status} exitosamente',
            'user': user.to_dict(include_private=True)
        })
        
    except HTTPException:
        
        raise
        
    except Exception as e:
        db.session.rollback()
        logger.exception('user_management error')
        return jsonify({'error': 'Error interno del servidor'}), 500


# ============== OPCIONES DE DOCUMENTOS ==============

@bp.route('/users/<string:user_id>/document-options', methods=['PUT'])
@jwt_required()
@management_required
@rate_limit(limit=60, window=60, key_prefix='rl_um_docopts')  # UM-N5
def update_user_document_options(user_id):
    """Actualizar opciones de documentos de un usuario (solo admin/developer)."""
    try:
        current_user = g.current_user
        user = User.query.get_or_404(user_id)

        # UM-N5: solo admin/developer. Soporte/coordinador/responsable/auxiliar NO.
        if current_user.role not in ('admin', 'developer'):
            return jsonify({'error': 'Solo administradores pueden modificar opciones de documentos'}), 403
        
        data = request.get_json()
        
        if 'evaluation_report' in data:
            user.enable_evaluation_report = data['evaluation_report']
        if 'certificate' in data:
            user.enable_certificate = data['certificate']
        if 'conocer_certificate' in data:
            user.enable_conocer_certificate = data['conocer_certificate']
        if 'digital_badge' in data:
            user.enable_digital_badge = data['digital_badge']
        
        db.session.commit()
        
        return jsonify({
            'message': 'Opciones de documentos actualizadas',
            'user': user.to_dict(include_private=True)
        })
        
    except HTTPException:
        
        raise
        
    except Exception as e:
        db.session.rollback()
        logger.exception('user_management error')
        return jsonify({'error': 'Error interno del servidor'}), 500


# ============== ELIMINAR USUARIO (SOLO ADMIN) ==============
#
# Política de borrado (UM-C3 — auditoría):
#   - El flujo normal (DELETE /users/<id> y bulk-delete) hace SOFT-DELETE:
#       anonimiza PII, marca is_deleted=True/deleted_at=now, limpia relaciones
#       operativas (group_members, group_exam_members, coordinator_balances,
#       chat_message_templates) pero PRESERVA evidencia (payments, results,
#       conocer_certificates, issued_badges, vouchers, activity_logs,
#       bulk_upload_*, study_*_progress) para auditoría y verificación
#       pública de certificados.
#   - El hard-delete (DELETE /users/<id>/hard) está reservado a admin y
#       cascadea todas las tablas dependientes. Solo debe usarse en respuesta
#       a una solicitud ARCO/GDPR genuina del titular.

# Tablas en las que la FK se debe poner NULL al hard-delete
_USER_NULLABLE_REFS = [
    ('answers', 'created_by'),
    ('answers', 'updated_by'),
    ('balance_requests', 'approved_by_id'),
    ('balance_requests', 'financiero_id'),
    ('balance_transactions', 'created_by_id'),
    ('badge_templates', 'created_by_id'),
    ('bulk_upload_batches', 'uploaded_by_id'),
    ('bulk_upload_members', 'user_id'),
    ('campuses', 'coordinator_id'),
    ('candidate_groups', 'coordinator_id'),
    ('categories', 'created_by'),
    ('categories', 'updated_by'),
    ('certificate_templates', 'updated_by'),
    ('conocer_email_contacts', 'created_by_id'),
    ('conocer_solicitud_logs', 'sent_by_id'),
    ('conocer_upload_batches', 'uploaded_by'),
    ('ecm_candidate_assignments', 'assigned_by_id'),
    ('ecm_retakes', 'applied_by_id'),
    ('ecm_swap_history', 'performed_by_id'),
    ('exams', 'created_by'),
    ('exams', 'updated_by'),
    ('exercises', 'created_by'),
    ('exercises', 'updated_by'),
    ('group_exams', 'assigned_by_id'),
    ('group_study_materials', 'assigned_by_id'),
    ('partners', 'coordinator_id'),
    ('questions', 'created_by'),
    ('questions', 'updated_by'),
    ('study_contents', 'created_by'),
    ('study_contents', 'updated_by'),
    ('study_interactive_exercises', 'created_by'),
    ('study_interactive_exercises', 'updated_by'),
    ('study_materials', 'created_by'),
    ('study_materials', 'updated_by'),
    ('support_conversations', 'assigned_support_user_id'),
    ('support_conversations', 'created_by_user_id'),
    ('support_conversation_satisfaction', 'submitted_by_user_id'),
    ('support_messages', 'sender_user_id'),
    ('topics', 'created_by'),
    ('topics', 'updated_by'),
    ('vm_sessions', 'cancelled_by_id'),
    ('vm_sessions', 'created_by_id'),
]

_USER_DELETE_REFS = [
    ('group_exam_members', 'user_id'),
    ('group_members', 'user_id'),
    ('group_study_material_members', 'user_id'),
    ('activity_logs', 'user_id'),
    ('balance_transactions', 'coordinator_id'),
    ('chat_message_templates', 'owner_user_id'),
    ('conocer_certificates', 'user_id'),
    ('coordinator_balances', 'coordinator_id'),
    ('ecm_candidate_assignments', 'user_id'),
    ('ecm_retakes', 'user_id'),
    ('ecm_swap_history', 'from_user_id'),
    ('ecm_swap_history', 'to_user_id'),
    ('issued_badges', 'user_id'),
    ('payments', 'user_id'),
    ('student_content_progress', 'user_id'),
    ('student_topic_progress', 'user_id'),
    ('study_scorm_attempts', 'user_id'),
    ('user_partners', 'user_id'),
    ('vm_sessions', 'user_id'),
    ('vouchers', 'user_id'),
    ('results', 'user_id'),
]


def _purge_user_from_db(user_id):
    """HARD-DELETE: elimina un user_id limpiando todas las FK.

    Reservado para flujo ARCO/GDPR (admin-only). El flujo normal usa
    `_soft_delete_user` que preserva evidencia.
    """
    for table, col in _USER_NULLABLE_REFS:
        try:
            db.session.execute(
                text(f"UPDATE {table} SET {col} = NULL WHERE {col} = :uid"),
                {'uid': user_id}
            )
        except Exception as _e:
            logger.warning('[HARD-DELETE] no se pudo nullificar %s.%s para %s: %s', table, col, user_id, _e)

    try:
        db.session.execute(
            text("UPDATE users SET coordinator_id = NULL WHERE coordinator_id = :uid"),
            {'uid': user_id}
        )
    except Exception as _e:
        logger.warning('[HARD-DELETE] no se pudo limpiar coordinator_id en cascada para %s: %s', user_id, _e)

    # Borrar cadena de support: messages -> participants -> conversations
    try:
        db.session.execute(
            text("""
                DELETE FROM support_messages
                WHERE conversation_id IN (
                    SELECT id FROM support_conversations WHERE candidate_user_id = :uid
                )
            """),
            {'uid': user_id}
        )
        db.session.execute(
            text("""
                DELETE FROM support_conversation_participants
                WHERE conversation_id IN (
                    SELECT id FROM support_conversations WHERE candidate_user_id = :uid
                )
                   OR user_id = :uid
            """),
            {'uid': user_id}
        )
        db.session.execute(
            text("DELETE FROM support_conversations WHERE candidate_user_id = :uid"),
            {'uid': user_id}
        )
    except Exception as _e:
        logger.warning('[HARD-DELETE] no se pudo limpiar support_* para %s: %s', user_id, _e)

    for table, col in _USER_DELETE_REFS:
        try:
            db.session.execute(
                text(f"DELETE FROM {table} WHERE {col} = :uid"),
                {'uid': user_id}
            )
        except Exception as _e:
            logger.warning('[HARD-DELETE] no se pudo borrar %s.%s para %s: %s', table, col, user_id, _e)

    db.session.execute(
        text("DELETE FROM users WHERE id = :uid"),
        {'uid': user_id}
    )


# Tablas con relaciones puramente OPERATIVAS que sí se limpian al soft-delete.
# NO se incluyen tablas con valor de auditoría/financiero (payments, results,
# conocer_certificates, issued_badges, vouchers, activity_logs,
# bulk_upload_*, student_*_progress, study_scorm_attempts, ecm_*).
_USER_SOFT_DELETE_CLEAR_REFS = [
    ('group_exam_members', 'user_id'),
    ('group_members', 'user_id'),
    ('group_study_material_members', 'user_id'),
    ('chat_message_templates', 'owner_user_id'),
    ('coordinator_balances', 'coordinator_id'),
    ('user_partners', 'user_id'),
]


def _soft_delete_user(user, current_user_id):
    """SOFT-DELETE: anonimiza PII, marca is_deleted, limpia relaciones
    operativas y conserva intactas las tablas de auditoría/evidencia.

    No hace commit; el caller decide cuándo persistir.
    """
    uid = user.id

    # 1. Limpiar relaciones operativas (membresías, plantillas, balances vacíos).
    for table, col in _USER_SOFT_DELETE_CLEAR_REFS:
        try:
            db.session.execute(
                text(f"DELETE FROM {table} WHERE {col} = :uid"),
                {'uid': uid}
            )
        except Exception as _e:
            logger.warning('[SOFT-DELETE] no se pudo limpiar %s.%s para %s: %s', table, col, uid, _e)

    # 2. Si era responsable de un campus, liberar el campus.
    try:
        db.session.execute(
            text("UPDATE campuses SET responsable_id = NULL WHERE responsable_id = :uid"),
            {'uid': uid}
        )
    except Exception as _e:
        logger.warning('[SOFT-DELETE] no se pudo liberar responsable en campuses: %s', _e)

    # 3. Cerrar / liberar conversaciones de soporte SIN borrar mensajes
    # (los mensajes deben conservarse para auditoría — UM-C3).
    try:
        db.session.execute(
            text("UPDATE support_conversations SET status = 'closed' WHERE candidate_user_id = :uid AND status != 'closed'"),
            {'uid': uid}
        )
    except Exception:
        pass

    # 4. Si era coordinador de otros usuarios, dejar a esos huérfanos del
    # coordinator_id para que un admin los reasigne (no se cascadea borrar).
    try:
        db.session.execute(
            text("UPDATE users SET coordinator_id = NULL WHERE coordinator_id = :uid"),
            {'uid': uid}
        )
    except Exception as _e:
        logger.warning('[SOFT-DELETE] no se pudo nullificar coordinator_id en cascada: %s', _e)

    # 5. Anonimizar PII y marcar como eliminado. El username se reserva
    # con prefijo deleted_ para liberar el original.
    original_username = user.username or uid
    suffix = uid[:8] if uid else 'x'
    user.is_deleted = True
    user.deleted_at = datetime.utcnow()
    user.is_active = False
    user.is_verified = False
    user.email = None
    user.username = f'deleted_{suffix}_{original_username}'[:100]
    user.name = '(Usuario eliminado)'
    user.first_surname = ''
    user.second_surname = None
    user.curp = None
    user.curp_verified = False
    user.curp_verified_at = None
    user.curp_renapo_name = None
    user.curp_renapo_first_surname = None
    user.curp_renapo_second_surname = None
    user.phone = None
    user.gender = None
    user.date_of_birth = None
    user.linkedin_token = None
    # Las credenciales se invalidan para que el usuario no pueda autenticarse.
    try:
        import secrets
        user.password_hash = secrets.token_urlsafe(64)
        user.encrypted_password = None
    except Exception:
        pass

    return {
        'id': uid,
        'previous_username': original_username,
    }


def _invalidate_user_caches():
    """Limpiar caches que dependen del set de usuarios (stats, counts)."""
    try:
        # Flask-Caching delete_many no soporta wildcards; limpiamos llaves conocidas.
        from app import cache as _cache
        for role in ('admin', 'developer', 'coordinator', 'auxiliar', 'soporte', 'responsable'):
            try:
                _cache.delete(f'user_stats_{role}_all')
            except Exception:
                pass
        # Heurística: limpiar también la versión sin sufijo coord/campus
        try:
            _cache.delete('user_stats_admin')
            _cache.delete('user_stats_coordinator')
        except Exception:
            pass
        # Como cache.delete_many con wildcard no funciona en SimpleCache,
        # confiamos en el TTL de 5 minutos para los counts dispersos.
    except Exception:
        pass


@bp.route('/users/<string:user_id>', methods=['DELETE'])
@jwt_required()
@admin_required
@rate_limit(limit=30, window=60, key_prefix='rl_um_delete')
def delete_user(user_id):
    """SOFT-DELETE: anonimiza al usuario y lo excluye de listados, preservando
    evidencia financiera y de certificación. Solo admin.
    """
    try:
        current_user = g.current_user

        if current_user.role != 'admin':
            return jsonify({'error': 'Solo el administrador puede eliminar usuarios'}), 403

        user = User.query.get_or_404(user_id)

        if getattr(user, 'is_deleted', False):
            return jsonify({'error': 'El usuario ya fue eliminado'}), 400

        if user_id == current_user.id:
            return jsonify({'error': 'No puedes eliminarte a ti mismo'}), 400

        if user.role in ('admin', 'developer'):
            return jsonify({'error': 'No se pueden eliminar usuarios admin/developer'}), 400

        user_email = user.email
        user_name = user.full_name

        _soft_delete_user(user, current_user.id)
        db.session.commit()
        _invalidate_user_caches()

        # UM-N25: no exponer PII (email/nombre) en respuesta.
        logger.info('[SOFT-DELETE] usuario %s eliminado por %s', user_name, current_user.id)
        return jsonify({
            'message': 'Usuario eliminado correctamente',
            'soft_delete': True,
        })

    except HTTPException:
        raise
    except Exception:
        db.session.rollback()
        logger.exception('user_management error')
        return jsonify({'error': 'Error interno del servidor'}), 500


@bp.route('/users/<string:user_id>/hard', methods=['DELETE'])
@jwt_required()
@admin_required
@rate_limit(limit=10, window=300, key_prefix='rl_um_harddel')
def hard_delete_user(user_id):
    """HARD-DELETE: borra permanentemente y cascadea evidencia. Solo admin.

    Reservado para solicitudes ARCO/GDPR — destruye payments, results,
    certificados emitidos e historial. Confirmar con header
    X-Confirm-Hard-Delete: yes para evitar borrados accidentales.
    """
    try:
        current_user = g.current_user

        if current_user.role != 'admin':
            return jsonify({'error': 'Solo el administrador puede eliminar usuarios'}), 403

        if request.headers.get('X-Confirm-Hard-Delete', '').lower() != 'yes':
            return jsonify({
                'error': 'Esta operación destruye evidencia financiera y de certificación. '
                         'Incluye el header X-Confirm-Hard-Delete: yes para confirmar.'
            }), 400

        user = User.query.get_or_404(user_id)

        if user_id == current_user.id:
            return jsonify({'error': 'No puedes eliminarte a ti mismo'}), 400

        if user.role in ('admin', 'developer'):
            return jsonify({'error': 'No se pueden eliminar usuarios admin/developer'}), 400

        user_email = user.email
        user_name = user.full_name

        db.session.expunge(user)
        _purge_user_from_db(user_id)
        db.session.commit()
        _invalidate_user_caches()

        # UM-N25: no exponer PII en respuesta.
        logger.info('[HARD-DELETE] usuario %s eliminado permanentemente por %s', user_name, current_user.id)
        return jsonify({
            'message': 'Usuario eliminado permanentemente',
            'hard_delete': True,
        })

    except HTTPException:
        raise
    except Exception:
        db.session.rollback()
        logger.exception('user_management error')
        return jsonify({'error': 'Error interno del servidor'}), 500


@bp.route('/users/bulk-delete', methods=['POST'])
@jwt_required()
@admin_required
@rate_limit(limit=5, window=60, key_prefix='rl_um_bulkdel')
def bulk_delete_users():
    """SOFT-DELETE en lote (solo admin). Body: { "user_ids": [...] }"""
    try:
        current_user = g.current_user

        if current_user.role != 'admin':
            return jsonify({'error': 'Solo el administrador puede eliminar usuarios'}), 403

        data = request.get_json() or {}
        user_ids = data.get('user_ids') or []

        if not isinstance(user_ids, list) or not user_ids:
            return jsonify({'error': 'Debes proporcionar una lista de user_ids'}), 400

        if len(user_ids) > 500:
            return jsonify({'error': 'No puedes eliminar más de 500 usuarios a la vez'}), 400

        user_ids = [uid for uid in user_ids if uid and uid != current_user.id]
        if not user_ids:
            return jsonify({'error': 'No hay usuarios válidos para eliminar'}), 400

        users = User.query.filter(
            User.id.in_(user_ids),
            User.is_deleted == False,  # noqa: E712
        ).all()

        if not users:
            return jsonify({'error': 'Ninguno de los usuarios existe o ya fueron eliminados'}), 404

        protected = [u for u in users if u.role in ('admin', 'developer')]
        if protected:
            names = ', '.join(u.username or u.id for u in protected)
            return jsonify({
                'error': f'No se pueden eliminar usuarios admin/developer: {names}'
            }), 400

        deleted = []
        failed = []

        for user in users:
            try:
                sp = db.session.begin_nested()
                try:
                    info = _soft_delete_user(user, current_user.id)
                    sp.commit()
                    deleted.append({'id': info['id'], 'name': info['previous_username']})
                except Exception as inner_err:
                    try:
                        sp.rollback()
                    except Exception:
                        pass
                    logger.warning('[BULK-SOFT-DELETE] error eliminando %s: %s', user.id, inner_err)
                    failed.append({'id': user.id, 'error': 'No se pudo eliminar este usuario'})
            except Exception as outer_err:
                logger.warning('[BULK-SOFT-DELETE] savepoint failed for %s: %s', user.id, outer_err)
                failed.append({'id': user.id, 'error': 'No se pudo iniciar la transacción'})

        try:
            db.session.commit()
        except Exception:
            db.session.rollback()
            logger.exception('[BULK-SOFT-DELETE] commit final falló')
            return jsonify({'error': 'Error interno del servidor'}), 500

        _invalidate_user_caches()

        return jsonify({
            'message': f'{len(deleted)} usuarios eliminados',
            'deleted_count': len(deleted),
            'deleted': deleted,
            'failed': failed,
            'requested': len(user_ids),
        })

    except HTTPException:
        raise
    except Exception as e:
        db.session.rollback()
        logger.exception('user_management error')
        return jsonify({'error': 'Error interno del servidor'}), 500


# ============== ESTADÍSTICAS ==============

@bp.route('/stats', methods=['GET'])
@jwt_required()
@management_required
@rate_limit(limit=300, window=60, key_prefix='rl_um_stats')  # UM-N6
def get_user_stats():
    """
    Obtener estadísticas de usuarios.
    Optimizado con:
    - Caché de 5 minutos
    - Una sola query agregada en lugar de múltiples counts
    """
    try:
        current_user = g.current_user
        
        # Intentar obtener del caché primero
        cache_key = f'user_stats_{current_user.role}_{_get_effective_coordinator_id(current_user) or current_user.id if _is_coordinator_role(current_user.role) or current_user.role == "responsable" else "all"}'
        cached_stats = _safe_cache_get(cache_key)
        if cached_stats is not None:
            return jsonify(cached_stats)
        
        # Query optimizada: obtener todos los conteos en una sola consulta
        if _is_coordinator_role(current_user.role):
            allowed_roles = ['candidato', 'responsable', 'responsable_partner', 'responsable_estatal', 'auxiliar']
        elif current_user.role == 'soporte':
            allowed_roles = SOPORTE_VISIBLE_ROLES
        elif current_user.role == 'responsable':
            allowed_roles = ['candidato']
        else:
            allowed_roles = AVAILABLE_ROLES
        
        # Usar una sola query con GROUP BY para obtener stats por rol
        stats_query = db.session.query(
            User.role,
            func.count(User.id).label('total'),
            func.sum(case((User.is_active == True, 1), else_=0)).label('active'),
            func.sum(case((User.is_verified == True, 1), else_=0)).label('verified')
        ).filter(User.is_deleted == False)  # noqa: E712  # UM-C3
        
        if _is_coordinator_role(current_user.role):
            coord_id = _get_effective_coordinator_id(current_user)
            stats_query = stats_query.filter(
                User.role.in_(allowed_roles),
                or_(
                    User.role == 'candidato',
                    and_(
                        User.role.in_(['responsable', 'responsable_partner', 'responsable_estatal', 'auxiliar']),
                        User.coordinator_id == coord_id
                    )
                )
            )
        elif current_user.role == 'soporte':
            stats_query = stats_query.filter(User.role.in_(allowed_roles))
        elif current_user.role == 'responsable':
            from app.models.partner import GroupMember, CandidateGroup
            campus_user_ids = db.session.query(GroupMember.user_id).join(
                CandidateGroup, GroupMember.group_id == CandidateGroup.id
            ).filter(
                CandidateGroup.campus_id == current_user.campus_id
            ).distinct().subquery()
            stats_query = stats_query.filter(
                User.role == 'candidato',
                or_(
                    User.id.in_(campus_user_ids),
                    User.campus_id == current_user.campus_id
                )
            )
        
        stats_query = stats_query.group_by(User.role)
        role_stats = stats_query.all()
        
        # Procesar resultados
        total_users = 0
        active_users = 0
        verified_users = 0
        users_by_role = []
        
        role_counts = {stat.role: stat.total for stat in role_stats}
        
        for stat in role_stats:
            total_users += stat.total
            active_users += stat.active or 0
            verified_users += stat.verified or 0
            users_by_role.append({'role': stat.role, 'count': stat.total})
        
        # Asegurar que todos los roles aparezcan (aunque tengan 0)
        if current_user.role in ['admin', 'developer']:
            for role in AVAILABLE_ROLES:
                if role not in role_counts:
                    users_by_role.append({'role': role, 'count': 0})
        elif current_user.role in ('coordinator', 'responsable'):
            for role in allowed_roles:
                if role not in role_counts:
                    users_by_role.append({'role': role, 'count': 0})
        
        # Ordenar por rol
        users_by_role.sort(key=lambda x: x['role'])
        
        inactive_users = total_users - active_users
        
        result = {
            'total_users': total_users,
            'active_users': active_users,
            'inactive_users': inactive_users,
            'verified_users': verified_users,
            'users_by_role': users_by_role
        }
        
        # Guardar en caché por 5 minutos
        _safe_cache_set(cache_key, result, timeout=300)
        
        return jsonify(result)
        
    except HTTPException:
        
        raise
        
    except Exception:
        logger.exception('user_management error')
        return jsonify({'error': 'Error interno del servidor'}), 500


@bp.route('/stats/invalidate', methods=['POST'])
@jwt_required()
@admin_required
@rate_limit(limit=30, window=60, key_prefix='rl_um_statsinv')  # UM-N16
def invalidate_stats_cache():
    """Invalidar caché de estadísticas (solo admin/developer — UM-H8)."""
    try:
        _invalidate_user_caches()
        return jsonify({'message': 'Caché de estadísticas invalidada'})
    except HTTPException:
        raise
    except Exception as e:
        logger.exception('user_management error')
        return jsonify({'error': 'Error interno del servidor'}), 500


# ============== ROLES DISPONIBLES ==============

@bp.route('/roles', methods=['GET'])
@jwt_required()
@management_required
@rate_limit(limit=300, window=60, key_prefix='rl_um_roles')  # UM-N15
def get_available_roles():
    """Obtener roles que puede crear el usuario actual"""
    try:
        current_user = g.current_user
        allowed_roles = ROLE_CREATE_PERMISSIONS.get(current_user.role, [])
        
        role_descriptions = {
            'admin': 'Administrador - Acceso total al sistema',
            'gerente': 'Gerente - Supervisión general y aprobación de operaciones',
            'financiero': 'Financiero - Gestión de saldos y aprobaciones financieras',
            'editor': 'Editor - Gestión de exámenes y contenidos',
            'editor_invitado': 'Editor Invitado - Gestión de exámenes con contenido aislado',
            'soporte': 'Soporte - Atención a usuarios y vouchers',
            'coordinator': 'Coordinador - Gestión de partners y candidatos',
            'responsable': 'Responsable - Administra un plantel y sus candidatos',
            'responsable_partner': 'Responsable del Partner - Supervisa al partner y sus planteles',
            'candidato': 'Candidato - Usuario que presenta evaluaciones',
            'auxiliar': 'Auxiliar - Acceso de solo lectura'
        }
        
        # Para soporte, devolver los roles visibles como all_roles para que los filtros funcionen
        if current_user.role == 'soporte':
            visible_roles = SOPORTE_VISIBLE_ROLES
        elif current_user.role in ['admin', 'developer']:
            visible_roles = AVAILABLE_ROLES
        else:
            visible_roles = None

        return jsonify({
            'roles': [
                {'value': role, 'label': role.capitalize(), 'description': role_descriptions.get(role, '')}
                for role in allowed_roles
            ],
            'all_roles': [
                {'value': role, 'label': role.capitalize(), 'description': role_descriptions.get(role, '')}
                for role in visible_roles
            ] if visible_roles else None
        })
        
    except HTTPException:
        
        raise
        
    except Exception as e:
        logger.exception('user_management error')
        return jsonify({'error': 'Error interno del servidor'}), 500


# ============== PLANTELES PARA RESPONSABLES ==============

@bp.route('/available-campuses', methods=['GET'])
@jwt_required()
@management_required
@rate_limit(limit=300, window=60, key_prefix='rl_um_camp')  # UM-N7
def get_available_campuses():
    """
    Obtener lista de planteles disponibles.
    Por defecto muestra todos los planteles (para asignar candidatos a grupos).
    Con for_responsable=true, muestra solo los pendientes de activación.
    """
    try:
        from app.models.partner import Campus, Partner

        current_user = g.current_user
        for_responsable = request.args.get('for_responsable', 'false').lower() == 'true'

        # Usar outerjoin para incluir campus aunque no tengan partner asociado
        query = Campus.query.outerjoin(Partner)

        # Coordinadores solo ven sus propios campuses
        if _is_coordinator_role(current_user.role):
            coord_id = _get_effective_coordinator_id(current_user)
            query = query.filter(Campus.coordinator_id == coord_id)

        if for_responsable:
            # Solo planteles pendientes de activación (para asignar responsables)
            query = query.filter(Campus.is_active == False)
        # Si no es for_responsable, mostrar TODOS los planteles (sin filtro is_active)
        
        campuses = query.order_by(Campus.name).all()
        
        return jsonify({
            'campuses': [{
                'id': c.id,
                'name': c.name,
                'code': c.code,
                'partner_id': c.partner_id,
                'partner_name': c.partner.name if c.partner else 'Sin partner',
                'state_name': c.state_name,
                'city': c.city,
                'has_responsable': c.responsable_id is not None,
                'activation_status': c.activation_status,
                'certification_cost': float(c.certification_cost) if c.certification_cost else 500.0
            } for c in campuses],
            'total': len(campuses)
        })
        
    except HTTPException:
        
        raise
        
    except Exception as e:
        logger.exception('user_management error')
        return jsonify({'error': 'Error interno del servidor'}), 500


# ============== PARTNERS DISPONIBLES (para responsable_partner) ==============

@bp.route('/available-partners', methods=['GET'])
@jwt_required()
@management_required
@rate_limit(limit=300, window=60, key_prefix='rl_um_part')  # UM-N8
def get_available_partners():
    """Obtener lista de partners disponibles para asignar a responsable_partner"""
    try:
        from app.models.partner import Partner

        current_user = g.current_user
        query = Partner.query.filter_by(is_active=True)

        # Coordinadores solo ven sus propios partners
        if _is_coordinator_role(current_user.role):
            coord_id = _get_effective_coordinator_id(current_user)
            query = query.filter(Partner.coordinator_id == coord_id)

        partners = query.order_by(Partner.name).all()
        
        return jsonify({
            'partners': [{
                'id': p.id,
                'name': p.name,
                'rfc': p.rfc or '',
                'contact_email': p.email or '',
                'country': p.country or 'México',
                'total_campuses': p.campuses.count() if p.campuses else 0
            } for p in partners],
            'total': len(partners)
        })
        
    except HTTPException:
        
        raise
        
    except Exception as e:
        logger.exception('user_management error')
        return jsonify({'error': 'Error interno del servidor'}), 500


# ============== ESTADOS DISPONIBLES (para responsable_estatal) ==============

@bp.route('/available-states', methods=['GET'])
@jwt_required()
@management_required
@rate_limit(limit=300, window=60, key_prefix='rl_um_states')  # UM-N9
def get_available_states():
    """Obtener lista de estados únicos de los campus de los partners del coordinador.
    Si el usuario es admin/developer/soporte, devuelve todos los estados de todos los campus."""
    try:
        from app.models.partner import Campus

        current_user = g.current_user
        query = db.session.query(Campus.state_name).filter(
            Campus.state_name.isnot(None),
            Campus.state_name != ''
        )

        # Coordinadores/auxiliares solo ven estados de sus campus
        if _is_coordinator_role(current_user.role):
            coord_id = _get_effective_coordinator_id(current_user)
            query = query.filter(Campus.coordinator_id == coord_id)

        rows = query.distinct().all()
        states = sorted([r[0] for r in rows if r[0]])

        return jsonify({
            'states': states,
            'total': len(states)
        })

    except HTTPException:
        raise

    except Exception as e:
        logger.exception('user_management error')
        return jsonify({'error': 'Error interno del servidor'}), 500


# ============== COORDINADORES DISPONIBLES (para asignar a responsables) ==============

@bp.route('/available-coordinators', methods=['GET'])
@jwt_required()
@management_required
@rate_limit(limit=300, window=60, key_prefix='rl_um_coords')  # UM-N10
def get_available_coordinators():
    """Obtener lista de coordinadores activos para asignar a responsables.

    UM-M7: solo admin/developer/soporte/gerente pueden ver el roster completo
    de coordinadores. Otros roles reciben 403.
    """
    try:
        current_user = g.current_user
        if current_user.role not in ('admin', 'developer', 'soporte', 'gerente'):
            return jsonify({'error': 'No tienes permiso para listar coordinadores'}), 403

        coordinators = User.query.filter(
            User.role == 'coordinator',
            User.is_active == True,
            User.is_deleted == False,  # noqa: E712
        ).order_by(User.name).all()

        return jsonify({
            'coordinators': [{
                'id': c.id,
                'full_name': c.full_name,
                'email': c.email,
                'username': c.username
            } for c in coordinators],
            'total': len(coordinators)
        })

    except HTTPException:
        raise

    except Exception as e:
        logger.exception('user_management error')
        return jsonify({'error': 'Error interno del servidor'}), 500


# ============== HELPERS: CARGA MASIVA DE CANDIDATOS ==============

MAX_BULK_ROWS = 50000
MAX_BULK_FILE_MB = 20
BATCH_COMMIT_SIZE = 500

_COLUMN_MAPPING = {
    'email': ['email', 'correo', 'correo electronico', 'correo electrónico', 'e-mail'],
    'nombre': ['nombre', 'name', 'nombres'],
    'primer_apellido': ['primer_apellido', 'primer apellido', 'apellido paterno', 'apellido_paterno', 'first_surname', 'apellido1'],
    'segundo_apellido': ['segundo_apellido', 'segundo apellido', 'apellido materno', 'apellido_materno', 'second_surname', 'apellido2'],
    'genero': ['genero', 'género', 'sexo', 'gender'],
    'curp': ['curp']
}


def _parse_bulk_candidates_excel(file_storage):
    """
    Parse Excel de candidatos. Retorna (parsed_rows, error_string).
    Valida tamaño, formato, columnas requeridas y límite de filas.
    """
    import io
    from openpyxl import load_workbook

    # Validar tamaño
    file_storage.seek(0, 2)
    file_size = file_storage.tell()
    file_storage.seek(0)
    if file_size > MAX_BULK_FILE_MB * 1024 * 1024:
        return None, f'Archivo excede {MAX_BULK_FILE_MB}MB (tiene {file_size / 1024 / 1024:.1f}MB)'

    # UM-H6: validar magic bytes de xlsx (zip header PK\x03\x04) antes de
    # pasarlo a openpyxl, que de otra forma reporta errores oscuros para
    # archivos maliciosos o renombrados.
    head = file_storage.read(4)
    file_storage.seek(0)
    if head[:2] != b'PK':
        return None, 'El archivo no es un .xlsx válido (debe ser un archivo Excel moderno).'

    try:
        workbook = load_workbook(filename=io.BytesIO(file_storage.read()), data_only=True)
        sheet = workbook.active
    except HTTPException:
        raise
    except Exception as e:
        return None, f'Error al leer el archivo Excel: {str(e)}'

    # Encabezados
    headers = []
    for cell in sheet[1]:
        value = cell.value
        headers.append(str(value).strip().lower() if value else '')

    column_indices = {}
    for field, aliases in _COLUMN_MAPPING.items():
        for i, header in enumerate(headers):
            if header in aliases:
                column_indices[field] = i
                break

    required_columns = ['nombre', 'primer_apellido', 'segundo_apellido', 'genero']
    missing = [c for c in required_columns if c not in column_indices]
    if missing:
        return None, f'Faltan columnas requeridas: {", ".join(missing)}. Requeridas: nombre, primer_apellido, segundo_apellido, genero'

    # Parsear filas (fila 1=headers, fila 2=descripciones, datos desde fila 3)
    parsed_rows = []
    for row_idx, row in enumerate(sheet.iter_rows(min_row=3), start=3):
        if len(parsed_rows) >= MAX_BULK_ROWS:
            return None, f'Límite excedido: máximo {MAX_BULK_ROWS:,} filas permitidas'

        def _val(field):
            if field not in column_indices:
                return None
            idx = column_indices[field]
            if idx < len(row):
                v = row[idx].value
                return str(v).strip() if v is not None else None
            return None

        nombre = _val('nombre')
        primer_ap = _val('primer_apellido')
        email_raw = _val('email')
        # Saltar filas completamente vacías
        if not any([email_raw, nombre, primer_ap]):
            continue

        parsed_rows.append({
            'row': row_idx,
            'email': email_raw.lower().strip() if email_raw else None,
            'nombre': nombre,
            'primer_apellido': primer_ap,
            'segundo_apellido': _val('segundo_apellido'),
            'genero_raw': _val('genero'),
            'curp': _val('curp').upper().strip() if _val('curp') else None,
        })

    return parsed_rows, None


def _validate_rows(parsed_rows):
    """Validar campos requeridos y formatos. Retorna (valid, errors).

    En modo LOCAL (CURP_RENAPO_ENABLED=false): valida CURP completa (formato
    + dígito verificador + entidad) síncronamente. Filas con CURP inválida
    se rechazan aquí. En modo RENAPO: solo valida longitud (mantiene
    compatibilidad con el flujo legacy donde el worker valida después).
    """
    from app.services.curp_local_validator import is_renapo_enabled, validate_curp_local
    renapo_on = is_renapo_enabled()

    valid = []
    errors = []
    for r in parsed_rows:
        errs = []
        if not r['nombre']:
            errs.append('nombre vacío')
        if not r['primer_apellido']:
            errs.append('primer_apellido vacío')
        if not r['segundo_apellido']:
            errs.append('segundo_apellido vacío')
        if not r['genero_raw']:
            errs.append('genero vacío')
        if r['email'] and not validate_email(r['email']):
            errs.append('formato de email inválido')
        if r['curp']:
            if not renapo_on and not _is_generic_foreign_curp(r['curp']):
                # ─── Modo LOCAL: validación completa síncrona ───
                is_v_local, local_err, _ = validate_curp_local(r['curp'])
                if not is_v_local:
                    errs.append(f'CURP inválida: {local_err}')
            elif renapo_on and len(r['curp']) != 18 and not _is_generic_foreign_curp(r['curp']):
                # Modo RENAPO legacy: solo longitud, formato lo valida el worker
                errs.append(f'CURP debe tener 18 caracteres (tiene {len(r["curp"])})')
        genero = None
        if r['genero_raw']:
            g = r['genero_raw'].upper()[0]
            if g in ('M', 'F', 'O'):
                genero = g
            else:
                errs.append('género inválido (usar M, F u O)')
        r['genero'] = genero
        if errs:
            errors.append({'row': r['row'], 'email': r['email'] or '(vacío)', 'nombre': r.get('nombre', ''), 'error': '; '.join(errs)})
        else:
            valid.append(r)
    return valid, errors


def _batch_fetch_existing(valid_rows):
    """
    Batch-fetch usuarios existentes por email y CURP.
    Retorna (existing_by_email, existing_by_curp) como dicts {key: User}.
    """
    from sqlalchemy import func as sf
    CHUNK = 500

    all_emails = list({r['email'] for r in valid_rows if r['email']})
    all_curps = list({r['curp'] for r in valid_rows if r['curp']})

    existing_by_email = {}
    for i in range(0, len(all_emails), CHUNK):
        chunk = all_emails[i:i + CHUNK]
        # UM-C3: excluir soft-deleted para permitir reutilizar email/CURP.
        users = User.query.filter(
            sf.lower(User.email).in_(chunk),
            User.is_deleted == False,  # noqa: E712
        ).all()
        for u in users:
            if u.email:
                existing_by_email[u.email.lower()] = u

    existing_by_curp = {}
    for i in range(0, len(all_curps), CHUNK):
        chunk = all_curps[i:i + CHUNK]
        users = User.query.filter(
            sf.upper(User.curp).in_(chunk),
            User.is_deleted == False,  # noqa: E712
        ).all()
        for u in users:
            if u.curp:
                existing_by_curp[u.curp.upper()] = u

    return existing_by_email, existing_by_curp


def _batch_generate_usernames(rows_to_create):
    """
    Pre-generar usernames únicos sin N+1 queries.
    Retorna dict {row_index: username}.

    Regla de negocio: TODO username debe ser exactamente 10 caracteres
    alfanuméricos en MAYUSCULAS (excluyendo I, L, O, 0 por ser confusos).
    """
    import random
    import string
    from sqlalchemy import func as sf
    CHUNK = 500

    _CONFUSING = set('ILO0')
    _ALPHABET = ''.join(c for c in string.ascii_uppercase + string.digits if c not in _CONFUSING)

    # Pre-cargar todos los usernames existentes activos para evitar colisiones.
    # Como queremos exactamente 10 chars, leemos sólo los de esa longitud.
    existing_usernames = set()
    existing_rows = User.query.filter(
        sf.length(User.username) == 10
    ).with_entities(User.username).all()
    existing_usernames.update(u.username.upper() for u, in existing_rows if u)

    used = set(existing_usernames)
    result = {}
    for r in rows_to_create:
        # Intentar hasta 30 veces; suficiente para >34^10 espacio de claves.
        for _ in range(30):
            candidate = ''.join(random.choices(_ALPHABET, k=10))
            if candidate not in used:
                used.add(candidate)
                result[r['row']] = candidate
                break
        else:
            # Fallback extremadamente improbable: usar timestamp.
            import time as _t
            candidate = (str(int(_t.time() * 1000))[-10:]).rjust(10, 'X')
            result[r['row']] = candidate
            used.add(candidate)
    return result


def _classify_valid_rows(valid_rows, existing_by_email, existing_by_curp, target_group_id=None):
    """
    Clasificar filas válidas en: to_create, duplicates (existing_assigned), skipped.
    Batch-fetch membresías de grupo si aplica.
    """
    from app.models.partner import GroupMember
    CHUNK = 500
    to_create = []
    duplicates = []  # Usuarios que ya existen (por email o CURP)
    skipped = []
    seen_emails = set()
    seen_curps = set()

    # Pre-fetch membresías si hay grupo destino
    group_member_ids = set()
    if target_group_id:
        all_existing_ids = list({u.id for u in list(existing_by_email.values()) + list(existing_by_curp.values())})
        for i in range(0, len(all_existing_ids), CHUNK):
            chunk = all_existing_ids[i:i + CHUNK]
            members = GroupMember.query.filter(
                GroupMember.group_id == target_group_id,
                GroupMember.user_id.in_(chunk)
            ).with_entities(GroupMember.user_id).all()
            group_member_ids.update(m.user_id for m in members)

    for r in valid_rows:
        email = r['email']
        curp = r['curp']

        # Duplicado por email
        if email and email in existing_by_email:
            user = existing_by_email[email]
            user_label = _describe_user(user)
            if target_group_id:
                if user.id in group_member_ids:
                    skipped.append({'row': r['row'], 'email': email,
                                    'reason': f'El correo {email} ya pertenece a un usuario existente ({user_label}) y ya es miembro del grupo destino, por lo que no se hace ningún cambio',
                                    'user_id': user.id, 'name': user.full_name, 'username': user.username, 'curp': user.curp, 'is_existing_user': True})
                else:
                    duplicates.append({'row': r['row'], 'email': email, 'name': user.full_name, 'username': user.username, 'user_id': user.id, 'curp': user.curp,
                                       'reason': f'El correo {email} ya está registrado como {user_label}; en vez de crear uno nuevo, este usuario se asignará al grupo'})
            else:
                skipped.append({'row': r['row'], 'email': email,
                                'reason': f'El correo {email} ya está registrado como {user_label}; no se crea otro usuario con el mismo correo',
                                'user_id': user.id, 'name': user.full_name, 'username': user.username, 'curp': user.curp, 'is_existing_user': True})
            continue

        # Duplicado por CURP (excepto CURPs genéricos de extranjero que pueden repetirse)
        if curp and curp in existing_by_curp and not _is_generic_foreign_curp(curp):
            user = existing_by_curp[curp]
            user_label = _describe_user(user)
            already_dup = any(d.get('user_id') == user.id for d in duplicates)
            if target_group_id:
                if user.id in group_member_ids or already_dup:
                    if not already_dup:
                        skipped.append({'row': r['row'], 'email': email or '(sin email)',
                                        'reason': f'La CURP {curp} ya pertenece a un usuario existente ({user_label}) y ya es miembro del grupo destino',
                                        'user_id': user.id, 'name': user.full_name, 'username': user.username, 'curp': user.curp, 'is_existing_user': True})
                else:
                    duplicates.append({'row': r['row'], 'email': user.email or email or '(sin email)', 'name': user.full_name, 'username': user.username, 'user_id': user.id, 'curp': user.curp,
                                       'reason': f'La CURP {curp} ya está registrada como {user_label}; este usuario se asignará al grupo en vez de crear uno nuevo'})
            else:
                skipped.append({'row': r['row'], 'email': email or '(sin email)',
                                'reason': f'La CURP {curp} ya está registrada como {user_label}; no se crea un usuario nuevo. Verifica que la CURP sea correcta',
                                'user_id': user.id, 'name': user.full_name, 'username': user.username, 'curp': user.curp, 'is_existing_user': True})
            continue

        # Duplicado interno (mismo email/curp repetido en el mismo archivo)
        if email and email in seen_emails:
            skipped.append({'row': r['row'], 'email': email, 'reason': f'El correo {email} aparece más de una vez en el mismo archivo; solo se procesa la primera fila'})
            continue
        # CURPs genéricos de extranjero pueden repetirse en el mismo archivo
        if curp and curp in seen_curps and not _is_generic_foreign_curp(curp):
            skipped.append({'row': r['row'], 'email': email or '(sin email)', 'reason': f'La CURP {curp} aparece más de una vez en el mismo archivo; solo se procesa la primera fila'})
            continue

        if email:
            seen_emails.add(email)
        if curp and not _is_generic_foreign_curp(curp):
            seen_curps.add(curp)
        to_create.append(r)

    return to_create, duplicates, skipped


def _validate_target_group(group_id, current_user):
    """Validar grupo destino. Retorna (group, error_tuple_or_None)."""
    if not group_id:
        return None, None
    from app.models.partner import CandidateGroup, Campus, Partner
    group = CandidateGroup.query.get(group_id)
    if not group:
        return None, (jsonify({'error': f'Grupo con ID {group_id} no encontrado'}), 404)
    if not group.is_active:
        return None, (jsonify({'error': 'El grupo seleccionado no está activo'}), 400)
    # Coordinador o auxiliar: el grupo debe pertenecer al coordinador efectivo
    if _is_coordinator_role(current_user.role):
        coord_id = _get_effective_coordinator_id(current_user)
        if not coord_id or group.coordinator_id != coord_id:
            return None, (jsonify({'error': 'No tienes acceso a este grupo'}), 403)
    if current_user.role == 'responsable':
        # Verificar que el grupo pertenece al plantel del responsable
        if group.campus_id != current_user.campus_id:
            return None, (jsonify({'error': 'No tienes acceso a este grupo'}), 403)
    return group, None


# ============== VALIDACIÓN CURP RENAPO ==============

@bp.route('/validate-curp', methods=['POST'])
@jwt_required()
@management_required
@rate_limit(limit=20, window=60, key_prefix='rl_um_curp')
def validate_curp_renapo_endpoint():
    """Valida una CURP contra RENAPO (síncrono ~10s).
    Retorna datos del ciudadano si la CURP es válida.
    """
    try:
        data = request.get_json()
        curp = (data.get('curp') or '').upper().strip()

        if not curp:
            return jsonify({'error': 'El campo CURP es requerido'}), 400

        # Validar formato básico
        curp_pattern = re.compile(r'^[A-Z]{4}[0-9]{6}[HM][A-Z]{5}[A-Z0-9][0-9]$')
        if not curp_pattern.match(curp) and curp not in GENERIC_FOREIGN_CURPS:
            return jsonify({'error': 'Formato de CURP inválido'}), 400

        # CURPs genéricos de extranjero: no requieren validación
        if _is_generic_foreign_curp(curp):
            return jsonify({
                'valid': True,
                'curp': curp,
                'skip_reason': 'CURP genérico de extranjero — no requiere validación RENAPO',
                'data': None
            }), 200

        from app.services.renapo_service import validate_curp_renapo
        result = validate_curp_renapo(curp)

        if result.valid:
            return jsonify({
                'valid': True,
                'curp': result.curp,
                'data': {
                    'name': result.name,
                    'first_surname': result.first_surname,
                    'second_surname': result.second_surname,
                    'gender': result.gender,
                }
            }), 200
        else:
            return jsonify({
                'valid': False,
                'curp': result.curp,
                'error': result.error,
                'data': None
            }), 200

    except HTTPException:
        raise
    except Exception:
        logger.exception('user_management error')
        return jsonify({'error': 'Error interno del servidor'}), 500


@bp.route('/validate-curp/batch', methods=['POST'])
@jwt_required()
@management_required
def validate_curp_batch_endpoint():
    """Valida un lote de CURPs contra RENAPO.
    Ejecuta de forma síncrona (el frontend puede mostrar spinner).
    Máximo 50 CURPs por solicitud.
    """
    try:
        data = request.get_json()
        curps = data.get('curps', [])

        if not curps or not isinstance(curps, list):
            return jsonify({'error': 'Se requiere una lista de CURPs'}), 400

        if len(curps) > 50:
            return jsonify({'error': 'Máximo 50 CURPs por solicitud'}), 400

        from app.services.renapo_service import validate_curps_batch
        results = validate_curps_batch(curps)

        return jsonify({
            'results': [r.to_dict() for r in results],
            'summary': {
                'total': len(results),
                'valid': sum(1 for r in results if r.valid),
                'invalid': sum(1 for r in results if not r.valid),
            }
        }), 200

    except HTTPException:
        raise
    except Exception:
        logger.exception('user_management error')
        return jsonify({'error': 'Error interno del servidor'}), 500


# ============== AUTO-VALIDACIÓN DE CURP POR EL CANDIDATO ==============

def _run_candidate_curp_verification(app_obj, user_id, curp, max_rounds: int = 6):
    """Ejecuta verificación RENAPO en segundo plano para un solo candidato.

    Aplica las mismas reglas que la verificación masiva (`/candidates/bulk-create`):
      - Hasta 6 rondas iterativas (RENAPO es inconsistente).
      - Cada ronda llama a `validate_curp_renapo` (que internamente reintenta 7 veces).
      - Pausa de 5s entre rondas.
      - Si en alguna ronda la CURP es válida: aplica datos al usuario y mueve
        GroupMembers a 'active'.
      - Si fallan las 6 rondas: deja GroupMembers en 'curp_required' para que
        el candidato pueda corregir su CURP.
    """
    import time
    import logging
    _logger = logging.getLogger(__name__)
    _logger.info(f'[ME-CURP] Iniciando verificación RENAPO en segundo plano para user={user_id} curp={curp} (hasta {max_rounds} rondas)')

    with app_obj.app_context():
        from app.services.renapo_service import validate_curp_renapo, apply_renapo_to_user
        from app.models.partner import GroupMember

        last_error = None
        for round_num in range(1, max_rounds + 1):
            try:
                result = validate_curp_renapo(curp)
            except Exception as e:
                last_error = str(e)[:300]
                _logger.error(f'[ME-CURP] R{round_num} excepción validando {curp}: {e}')
                if round_num < max_rounds:
                    time.sleep(5)
                continue

            if result.valid:
                try:
                    user = User.query.get(user_id)
                    if not user:
                        _logger.warning(f'[ME-CURP] Usuario {user_id} ya no existe, abortando')
                        return
                    # Confirmar que no cambió la CURP mientras tanto
                    if (user.curp or '').strip().upper() != curp.strip().upper():
                        _logger.info(f'[ME-CURP] CURP del usuario {user_id} cambió mientras se verificaba; abortando este job')
                        return
                    apply_renapo_to_user(user, result)
                    pending_gms = GroupMember.query.filter_by(user_id=user.id).filter(
                        GroupMember.status.in_(['curp_pending', 'curp_verifying', 'curp_required'])
                    ).all()
                    for gm in pending_gms:
                        gm.status = 'active'
                    db.session.commit()
                    _logger.info(f'[ME-CURP] R{round_num} CURP {curp} válida — usuario {user_id} verificado')
                    return
                except Exception as commit_err:
                    _logger.error(f'[ME-CURP] R{round_num} error aplicando datos RENAPO a {user_id}: {commit_err}')
                    try:
                        db.session.rollback()
                    except Exception:
                        pass
                    last_error = str(commit_err)[:300]
            else:
                last_error = result.error or 'no encontrada'
                _logger.info(f'[ME-CURP] R{round_num} CURP {curp} rechazada para user={user_id} ({last_error}); reintentará')

            if round_num < max_rounds:
                time.sleep(5)

        # ── Tras 6 rondas sin éxito: dejar al usuario en curp_required ──
        try:
            user = User.query.get(user_id)
            if not user:
                return
            if (user.curp or '').strip().upper() != curp.strip().upper():
                return  # cambió la CURP, otro job tomará el relevo
            for gm in GroupMember.query.filter_by(user_id=user.id).all():
                if gm.status in ('curp_pending', 'curp_verifying'):
                    gm.status = 'curp_required'
            db.session.commit()
            _logger.warning(f'[ME-CURP] CURP {curp} no validada tras {max_rounds} rondas para user={user_id} (último error: {last_error})')
        except Exception as final_err:
            _logger.error(f'[ME-CURP] error marcando curp_required final para {user_id}: {final_err}')
            try:
                db.session.rollback()
            except Exception:
                pass


@bp.route('/me/validate-curp', methods=['POST'])
@jwt_required()
@rate_limit(limit=10, window=60, key_prefix='rl_um_mecurp')
def candidate_validate_own_curp():
    """Permite al usuario (candidato o responsable) corregir su CURP contra RENAPO.

    Flujo 100% asíncrono — el usuario NO necesita quedarse en la página:
      1. Se valida formato + unicidad localmente (síncrono, rápido).
      2. Se persiste la nueva CURP en `users.curp` y se ponen los GroupMembers
         del usuario en estado 'curp_verifying'.
      3. Se cancelan filas previas en la cola para este usuario y se inserta
         una nueva fila con `source='self_fix'`.
      4. El worker procesa la cola en background. Cuando termina:
           - Si RENAPO valida: marca `curp_verified=True`, desbloquea GroupMembers
             a 'active' y envía email al usuario "Tu CURP fue validada".
           - Si RENAPO rechaza tras los reintentos: GroupMembers vuelven a
             'curp_required' y envía email al usuario "No pudimos validar tu CURP".
      5. El endpoint responde inmediatamente con 202 `{verifying: true}` —
         el frontend puede mostrar un mensaje "te avisaremos por correo" o
         hacer polling a `/me/curp-status` si quiere actualización en vivo.

    CURPs genéricas de extranjero se aceptan inmediatamente sin RENAPO (200).
    """
    try:
        current_user_id = get_jwt_identity()
        user = User.query.get(current_user_id)
        if not user:
            return jsonify({'error': 'Usuario no encontrado'}), 404

        if user.role not in ('candidato', 'responsable'):
            return jsonify({'error': 'Solo candidatos y responsables pueden usar este endpoint'}), 403

        if user.curp_verified:
            return jsonify({
                'error': 'Tu CURP ya fue validada anteriormente. Si necesitas cambiarla, contacta a tu coordinador.',
                'already_verified': True
            }), 400

        data = request.get_json() or {}
        new_curp = (data.get('curp') or '').upper().strip()

        if not new_curp:
            return jsonify({'error': 'El campo CURP es requerido'}), 400

        # CURP genérica de extranjero — aceptar sin RENAPO
        if _is_generic_foreign_curp(new_curp):
            user.curp = new_curp
            from app.models.partner import GroupMember
            for gm in GroupMember.query.filter_by(user_id=user.id).all():
                if gm.status in ('curp_required', 'curp_pending', 'curp_verifying'):
                    gm.status = 'active'
            db.session.commit()
            return jsonify({
                'valid': True,
                'curp': new_curp,
                'generic_foreign': True,
                'message': 'CURP genérica de extranjero aceptada'
            }), 200

        # Validar formato local
        from app.services.renapo_service import validate_curp_format
        fmt_valid, fmt_error = validate_curp_format(new_curp)
        if not fmt_valid:
            return jsonify({
                'valid': False,
                'curp': new_curp,
                'error': fmt_error,
                'reason': 'format'
            }), 400

        # Verificar unicidad (excepto el propio usuario)
        existing = User.query.filter(
            User.curp == new_curp,
            User.id != user.id
        ).first()
        if existing:
            return jsonify({
                'valid': False,
                'curp': new_curp,
                'error': 'Esta CURP ya está registrada por otro usuario',
                'reason': 'duplicate'
            }), 409

        # ===== Modo LOCAL (CURP_RENAPO_ENABLED=false) =====
        # Validación síncrona: formato + dígito verificador OK ⇒ verified.
        from app.services.curp_local_validator import is_renapo_enabled
        if not is_renapo_enabled():
            from app.services.curp_local_validator import apply_local_validation_to_user
            from app.models.partner import GroupMember
            try:
                user.curp = new_curp
                apply_local_validation_to_user(user, mark_verified=True)
                for gm in GroupMember.query.filter_by(user_id=user.id).all():
                    if gm.status in ('curp_required', 'curp_pending', 'curp_verifying'):
                        gm.status = 'active'
                db.session.commit()
                return jsonify({
                    'valid': True,
                    'curp': new_curp,
                    'local_validated': True,
                    'message': 'CURP validada (formato + dígito verificador)'
                }), 200
            except Exception as local_err:
                db.session.rollback()
                import logging
                logging.getLogger(__name__).error(f'[ME-CURP-LOCAL] error: {local_err}')
                return jsonify({
                    'valid': False,
                    'curp': new_curp,
                    'error': 'No fue posible registrar tu CURP. Intenta de nuevo.',
                    'reason': 'persist',
                }), 500

        # ===== Encolar para validación ASÍNCRONA contra RENAPO =====
        # No bloqueamos al usuario esperando 25s+. Persistimos la CURP, marcamos
        # los GroupMembers como 'curp_verifying' y delegamos al worker. El
        # worker enviará un email al usuario cuando termine (éxito o rechazo).
        from app.models.curp_verification import (
            CurpVerificationQueue, QUEUE_PENDING, QUEUE_PROCESSING, QUEUE_FAILED,
        )
        from app.models.partner import GroupMember

        try:
            user.curp = new_curp
            for gm in GroupMember.query.filter_by(user_id=user.id).all():
                if gm.status in ('curp_required', 'curp_pending', 'active'):
                    gm.status = 'curp_verifying'

            # Cancelar filas previas en la cola para este usuario (CURPs
            # anteriores que ya no aplican o reintentos del mismo CURP).
            CurpVerificationQueue.query.filter(
                CurpVerificationQueue.user_id == user.id,
                CurpVerificationQueue.status.in_([QUEUE_PENDING, QUEUE_PROCESSING]),
            ).update({
                'status': QUEUE_FAILED,
                'finished_at': datetime.utcnow(),
                'last_error': 'Reemplazado por nueva CURP enviada por el usuario',
                'locked_at': None,
                'locked_by': None,
            }, synchronize_session=False)

            # Insertar nueva entrada en la cola
            new_row = CurpVerificationQueue(
                user_id=user.id,
                curp=new_curp,
                status=QUEUE_PENDING,
                source='self_fix',  # marca para que el worker notifique al usuario
                next_retry_at=datetime.utcnow(),
            )
            db.session.add(new_row)
            db.session.commit()
        except Exception as enq_err:
            db.session.rollback()
            import logging
            logging.getLogger(__name__).error(f'[ME-CURP] error encolando: {enq_err}')
            return jsonify({
                'valid': False,
                'curp': new_curp,
                'error': 'No fue posible registrar tu solicitud. Intenta de nuevo.',
                'reason': 'persist',
            }), 500

        return jsonify({
            'verifying': True,
            'curp': new_curp,
            'message': (
                'Tu CURP se está validando contra RENAPO en segundo plano. '
                'Puedes cerrar esta página: te enviaremos un correo cuando termine.'
            ),
        }), 202

    except HTTPException:
        raise
    except Exception:
        db.session.rollback()
        logger.exception('user_management error')
        return jsonify({'error': 'Error interno del servidor'}), 500


@bp.route('/me/curp-status', methods=['GET'])
@jwt_required()
def candidate_curp_status():
    """Retorna el estado actual de validación CURP del candidato logueado."""
    try:
        current_user_id = get_jwt_identity()
        user = User.query.get(current_user_id)
        if not user:
            return jsonify({'error': 'Usuario no encontrado'}), 404

        from app.models.partner import GroupMember
        from app.models.curp_verification import (
            CurpVerificationQueue, QUEUE_PENDING, QUEUE_PROCESSING,
        )
        gm_statuses = [
            gm.status for gm in GroupMember.query.filter_by(user_id=user.id).all()
        ]
        # 'verifying' es verdadero si:
        #   - algún GroupMember está en estado 'curp_verifying', O
        #   - hay una fila pending/processing en la cola para este usuario
        #     (cubre el caso de responsables que no tienen GroupMember).
        queue_active = db.session.query(CurpVerificationQueue.id).filter(
            CurpVerificationQueue.user_id == user.id,
            CurpVerificationQueue.status.in_([QUEUE_PENDING, QUEUE_PROCESSING]),
        ).first()
        verifying = any(s == 'curp_verifying' for s in gm_statuses) or bool(queue_active)
        return jsonify({
            'user_id': user.id,
            'curp': user.curp,
            'curp_verified': user.curp_verified,
            'curp_verified_at': user.curp_verified_at.isoformat() if user.curp_verified_at else None,
            'verifying': verifying,
            'requires_validation': bool(
                user.role in ('candidato', 'responsable')
                and not user.curp_verified
                and (user.curp or '').upper().strip() not in GENERIC_FOREIGN_CURPS
            ),
            'group_member_statuses': gm_statuses,
        }), 200
    except HTTPException:
        raise
    except Exception as e:
        logger.exception('user_management error')
        return jsonify({'error': 'Error interno del servidor'}), 500


# ============== PREVIEW CARGA MASIVA ==============

@bp.route('/candidates/bulk-upload/preview', methods=['POST'])
@jwt_required()
@management_required
@rate_limit(limit=60, window=60, key_prefix='rl_um_bulkprev')  # UM-N11 (generoso para flujos de hasta 5K candidatos)
def preview_bulk_upload_candidates():
    """
    Preview de carga masiva de candidatos — valida SIN crear usuarios.
    Retorna status por fila: ready, duplicate, error, skipped.
    """
    try:
        current_user = g.current_user
        bulk_perm_err = _ensure_bulk_upload_allowed(current_user)
        if bulk_perm_err:
            return bulk_perm_err
        group_id = request.form.get('group_id', type=int)

        # Validar grupo si se especificó
        target_group, group_err = _validate_target_group(group_id, current_user)
        if group_err:
            return group_err

        # Validar archivo
        if 'file' not in request.files:
            return jsonify({'error': 'No se envió ningún archivo'}), 400
        file = request.files['file']
        if not file.filename or not file.filename.lower().endswith(('.xlsx', '.xls')):
            return jsonify({'error': 'El archivo debe ser formato Excel (.xlsx o .xls)'}), 400

        # Parsear
        parsed_rows, parse_error = _parse_bulk_candidates_excel(file)
        if parse_error:
            return jsonify({'error': parse_error}), 400
        if not parsed_rows:
            return jsonify({'error': 'El archivo no contiene datos (filas vacías)'}), 400

        # Auto-asignar CURP genérico para planteles extranjeros
        is_foreign_campus = False
        if target_group:
            from app.models.partner import Campus as CampusModel
            campus = CampusModel.query.get(target_group.campus_id)
            if campus and campus.country and campus.country != 'México':
                is_foreign_campus = True
                for r in parsed_rows:
                    if not r['curp']:
                        # Auto-asignar CURP según género
                        g_raw = (r.get('genero_raw') or '').upper()[:1]
                        if g_raw == 'M':
                            r['curp'] = FOREIGN_CURP_MALE
                        else:
                            r['curp'] = FOREIGN_CURP_FEMALE

        # Validar campos
        valid_rows, validation_errors = _validate_rows(parsed_rows)

        # Batch-fetch existentes
        existing_by_email, existing_by_curp = _batch_fetch_existing(valid_rows)

        # Clasificar
        to_create, duplicates, skipped = _classify_valid_rows(
            valid_rows, existing_by_email, existing_by_curp,
            target_group.id if target_group else None
        )

        # Pre-generar usernames para los que se van a crear
        username_map = _batch_generate_usernames(to_create) if to_create else {}

        # Detectar coincidencias por nombre+género en to_create
        name_match_map = {}  # row_number -> [matching_users]
        if to_create:
            # Recopilar combinaciones únicas (nombre, primer_apellido, genero)
            name_keys = set()
            for r in to_create:
                key = (r['nombre'].strip().lower(), r['primer_apellido'].strip().lower(), (r['genero'] or '').strip().upper())
                name_keys.add(key)

            # Batch-fetch por primer_apellido (reducir queries)
            surnames = list({k[1] for k in name_keys})
            CHUNK = 500
            matching_users = []
            for i in range(0, len(surnames), CHUNK):
                chunk = surnames[i:i + CHUNK]
                users = User.query.filter(
                    func.lower(func.ltrim(func.rtrim(User.first_surname))).in_(chunk),
                    User.is_active == True,
                ).all()
                matching_users.extend(users)

            # Construir lookup: (nombre_lower, primer_apellido_lower, genero_upper) -> [users]
            name_lookup = {}
            for u in matching_users:
                n = (u.name or '').strip().lower()
                fs = (u.first_surname or '').strip().lower()
                gen = (u.gender or '').strip().upper()
                key = (n, fs, gen)
                name_lookup.setdefault(key, []).append(u)

            # Mapear filas a coincidencias
            for r in to_create:
                key = (r['nombre'].strip().lower(), r['primer_apellido'].strip().lower(), (r['genero'] or '').strip().upper())
                matches = name_lookup.get(key, [])
                if matches:
                    name_match_map[r['row']] = matches[:5]  # Máximo 5 coincidencias (objetos User)

            # Batch-fetch grupo y plantel actual de los usuarios coincidentes
            all_match_user_ids = list({u.id for matches in name_match_map.values() for u in matches})
            user_group_campus_map = {}  # user_id -> {group_name, campus_name}
            if all_match_user_ids:
                from app.models.partner import GroupMember as GM_match, CandidateGroup as CG_match, Campus as Campus_match
                gm_rows = db.session.query(
                    GM_match.user_id,
                    CG_match.name.label('group_name'),
                    Campus_match.name.label('campus_name')
                ).join(
                    CG_match, GM_match.group_id == CG_match.id
                ).join(
                    Campus_match, CG_match.campus_id == Campus_match.id
                ).filter(
                    GM_match.user_id.in_(all_match_user_ids),
                    GM_match.status == 'active'
                ).all()
                for row in gm_rows:
                    uid = row[0]
                    if uid not in user_group_campus_map:
                        user_group_campus_map[uid] = []
                    user_group_campus_map[uid].append({
                        'group_name': row[1],
                        'campus_name': row[2],
                    })

            # Convertir a dicts con info de grupo/plantel
            for row_num in name_match_map:
                name_match_map[row_num] = [
                    {
                        'id': u.id,
                        'full_name': u.full_name,
                        'username': u.username,
                        'email': u.email or '',
                        'curp': u.curp or '',
                        'groups': user_group_campus_map.get(u.id, []),
                    }
                    for u in name_match_map[row_num]
                ]

        # Construir preview
        preview = []
        name_match_count = 0
        for r in to_create:
            matches = name_match_map.get(r['row'])
            row_status = 'name_match' if matches else 'ready'
            if matches:
                name_match_count += 1
            entry = {
                'row': r['row'],
                'status': row_status,
                'email': r['email'],
                'nombre': r['nombre'],
                'primer_apellido': r['primer_apellido'],
                'segundo_apellido': r['segundo_apellido'],
                'genero': r['genero'],
                'curp': r['curp'],
                'username_preview': username_map.get(r['row'], ''),
                'eligibility': {
                    'reporte': True, 'eduit': True,
                    'conocer': bool(r['curp']),
                    'insignia': bool(r['email']),
                },
                'error': None,
            }
            if matches:
                entry['name_matches'] = matches
            preview.append(entry)
        for d in duplicates:
            preview.append({
                'row': d['row'],
                'status': 'duplicate',
                'email': d.get('email'),
                'nombre': d.get('name', ''),
                'curp': d.get('curp'),
                'existing_user': {'id': d['user_id'], 'name': d['name'], 'username': d['username'], 'curp': d.get('curp')},
                'error': ('Usuario ya existente — se asignará al grupo destino' if target_group else 'Usuario ya existente — no se crea uno nuevo'),
            })
        for s in skipped:
            entry = {
                'row': s['row'],
                'status': 'skipped',
                'email': s.get('email'),
                'error': s['reason'],
            }
            if s.get('is_existing_user'):
                entry['existing_user'] = {
                    'id': s['user_id'],
                    'name': s['name'],
                    'username': s['username'],
                    'curp': s.get('curp'),
                }
                entry['curp'] = s.get('curp')
            preview.append(entry)
        for e in validation_errors:
            preview.append({
                'row': e['row'],
                'status': 'error',
                'email': e.get('email'),
                'nombre': e.get('nombre', ''),
                'error': e['error'],
            })

        # Ordenar por número de fila
        preview.sort(key=lambda x: x['row'])

        return jsonify({
            'preview': preview,
            'summary': {
                'total_rows': len(parsed_rows),
                'ready': len(to_create) - name_match_count,
                'name_matches': name_match_count,
                'duplicates': len(duplicates),
                'errors': len(validation_errors),
                'skipped': len(skipped),
            },
            'can_proceed': len(to_create) > 0 or len(duplicates) > 0 or len(skipped) > 0,
            'group_info': {'id': target_group.id, 'name': target_group.name} if target_group else None,
        }), 200

    except HTTPException:

        raise

    except Exception as e:
        logger.exception('user_management error')
        return jsonify({'error': 'Error interno del servidor'}), 500


# ============== CARGA MASIVA DE CANDIDATOS (OPTIMIZADO) ==============

@bp.route('/candidates/bulk-upload', methods=['POST'])
@jwt_required()
@management_required
def bulk_upload_candidates():
    """
    Carga masiva de candidatos desde archivo Excel — versión batch-optimizada.
    Soporta hasta 50,000 filas. Commits en lotes de 500.

    Columnas: nombre, primer_apellido, segundo_apellido, genero (requeridas)
              email, curp (opcionales)
    Form fields: group_id (opcional)
    """
    try:
        import secrets
        import string

        CHUNK = 500
        current_user = g.current_user
        bulk_perm_err = _ensure_bulk_upload_allowed(current_user)
        if bulk_perm_err:
            return bulk_perm_err
        group_id = request.form.get('group_id', type=int)

        # Validar grupo
        target_group, group_err = _validate_target_group(group_id, current_user)
        if group_err:
            return group_err

        # Validar archivo
        if 'file' not in request.files:
            return jsonify({'error': 'No se envió ningún archivo'}), 400
        file = request.files['file']
        if not file.filename or not file.filename.lower().endswith(('.xlsx', '.xls')):
            return jsonify({'error': 'El archivo debe ser formato Excel (.xlsx o .xls)'}), 400

        # Parsear
        parsed_rows, parse_error = _parse_bulk_candidates_excel(file)
        if parse_error:
            return jsonify({'error': parse_error}), 400
        if not parsed_rows:
            return jsonify({'error': 'El archivo no contiene datos'}), 400

        # Auto-asignar CURP genérico para planteles extranjeros
        if target_group:
            from app.models.partner import Campus as CampusModel
            campus = CampusModel.query.get(target_group.campus_id)
            if campus and campus.country and campus.country != 'México':
                for r in parsed_rows:
                    if not r['curp']:
                        g_raw = (r.get('genero_raw') or '').upper()[:1]
                        if g_raw == 'M':
                            r['curp'] = FOREIGN_CURP_MALE
                        else:
                            r['curp'] = FOREIGN_CURP_FEMALE

        # Validar
        valid_rows, validation_errors = _validate_rows(parsed_rows)

        # Batch-fetch existentes
        existing_by_email, existing_by_curp = _batch_fetch_existing(valid_rows)

        # Clasificar
        to_create, existing_assigned, skipped = _classify_valid_rows(
            valid_rows, existing_by_email, existing_by_curp,
            target_group.id if target_group else None
        )

        # Responsable: auto-incluir todos los usuarios existentes (asignación silenciosa)
        if current_user.role == 'responsable' and target_group:
            new_skipped = []
            for s in skipped:
                if s.get('is_existing_user') and s.get('user_id'):
                    existing_assigned.append({
                        'row': s['row'],
                        'email': s.get('email', ''),
                        'name': s.get('name', ''),
                        'username': s.get('username', ''),
                        'user_id': s['user_id'],
                    })
                else:
                    new_skipped.append(s)
            skipped = new_skipped

        # Handle include_existing_ids — skipped users the admin opted to add to group
        include_existing_ids_raw = request.form.get('include_existing_ids', '')
        if include_existing_ids_raw and target_group:
            import json as _json
            try:
                include_ids = _json.loads(include_existing_ids_raw)
                if isinstance(include_ids, list):
                    include_set = set(str(i) for i in include_ids)
                    new_skipped = []
                    for s in skipped:
                        if s.get('is_existing_user') and str(s.get('user_id', '')) in include_set:
                            existing_assigned.append({
                                'row': s['row'],
                                'email': s.get('email', ''),
                                'name': s.get('name', ''),
                                'username': s.get('username', ''),
                                'user_id': s['user_id'],
                            })
                        else:
                            new_skipped.append(s)
                    skipped = new_skipped
            except (ValueError, TypeError):
                pass

        # Handle skip_row_numbers — name_match rows the admin chose NOT to create
        skip_rows_raw = request.form.get('skip_row_numbers', '')
        if skip_rows_raw:
            import json as _json2
            try:
                skip_rows = _json2.loads(skip_rows_raw)
                if isinstance(skip_rows, list):
                    skip_set = set(int(r) for r in skip_rows)
                    new_to_create = []
                    for r in to_create:
                        if r['row'] in skip_set:
                            skipped.append({
                                'row': r['row'],
                                'email': r.get('email', ''),
                                'reason': 'Omitido por coincidencia de nombre (decisión del usuario)',
                            })
                        else:
                            new_to_create.append(r)
                    to_create = new_to_create
            except (ValueError, TypeError):
                pass

        # Generar usernames
        username_map = _batch_generate_usernames(to_create) if to_create else {}

        # Función para generar contraseña (excluye caracteres confusos: i, I, l, L, o, O, 0)
        def _gen_pwd(length=10):
            _CONFUSING = set('iIlLoO0')
            upper = ''.join(c for c in string.ascii_uppercase if c not in _CONFUSING)
            lower = ''.join(c for c in string.ascii_lowercase if c not in _CONFUSING)
            digits = ''.join(c for c in string.digits if c not in _CONFUSING)
            alpha = upper + lower + digits
            pwd = [secrets.choice(upper), secrets.choice(lower), secrets.choice(digits)]
            pwd += [secrets.choice(alpha) for _ in range(length - 3)]
            secrets.SystemRandom().shuffle(pwd)
            return ''.join(pwd)

        # Crear usuarios en lotes con tolerancia a fallos.
        # Estrategia (enfoque B):
        #   1) Fast path: agregar BATCH_COMMIT_SIZE usuarios y hacer commit.
        #   2) Si el commit falla por una fila duplicada/FK (que envenena la
        #      transacción en MSSQL), rollback y reintentar cada usuario del
        #      lote uno por uno con SAVEPOINT (begin_nested) para identificar
        #      exactamente cuál(es) fallaron sin perder los válidos.
        created = []
        create_errors = []
        curp_pending_user_ids = set()  # IDs de usuarios que necesitan verificación CURP

        def _build_user_payload(r):
            """Construye el dict de kwargs + password para un User sin agregarlo
            todavía a la sesión. Devuelve (payload, password, username).

            En modo local (CURP_RENAPO_ENABLED=false), si la CURP pasó la
            validación local en _validate_rows, marca curp_verified=True
            de una vez (sin esperar al worker).
            """
            from app.services.curp_local_validator import is_renapo_enabled
            _renapo_on_bulk = is_renapo_enabled()
            _curp_val = r['curp'] or None
            _curp_pre_verified = bool(
                (not _renapo_on_bulk) and _curp_val  # local mode + CURP presente y ya validada en _validate_rows
            )
            username = username_map[r['row']]
            password = _gen_pwd()
            _campus_id = None
            _coord_id = None
            if target_group:
                _campus_id = target_group.campus_id
                _coord_id = target_group.coordinator_id
            else:
                if _is_coordinator_role(current_user.role):
                    _coord_id = _get_effective_coordinator_id(current_user)
                if current_user.role == 'responsable' and current_user.campus_id:
                    _campus_id = current_user.campus_id
            payload = dict(
                id=str(uuid.uuid4()),
                email=r['email'] if r['email'] else None,
                username=username,
                name=r['nombre'].upper() if r['nombre'] else r['nombre'],
                first_surname=r['primer_apellido'].upper() if r['primer_apellido'] else r['primer_apellido'],
                second_surname=r['segundo_apellido'].upper() if r['segundo_apellido'] else None,
                gender=r['genero'],
                curp=r['curp'] or None,
                role='candidato',
                coordinator_id=_coord_id,
                campus_id=_campus_id,
                is_active=True,
                is_verified=False,
            )
            # Modo local: marcar verificado de una vez
            if _curp_pre_verified:
                payload['curp_verified'] = True
                payload['curp_verified_at'] = datetime.utcnow()
            return payload, password, username

        def _instantiate_and_add(payload, password):
            """Crea User, setea passwords y lo agrega a la sesión actual."""
            u = User(**payload)
            u.set_password(password)
            u.encrypted_password = encrypt_password(password)
            db.session.add(u)
            return u

        def _record_created(r, payload, password, username, uid):
            _parts = [r['nombre'], r['primer_apellido']]
            if r.get('segundo_apellido'):
                _parts.append(r['segundo_apellido'])
            created.append({
                'row': r['row'],
                'email': r['email'],
                'name': f"{r['nombre']} {r['primer_apellido']}",
                'full_name': ' '.join(_parts),
                'username': username,
                'password': password,
                'curp': r.get('curp'),
                'gender': r.get('genero'),
                'user_id': uid,
            })
            needs_curp_verify = bool(
                target_group and r.get('curp') and r['curp'] not in GENERIC_FOREIGN_CURPS
            )
            if needs_curp_verify:
                curp_pending_user_ids.add(uid)

        def _commit_batch(batch_rows):
            """Intenta commit del lote (fast path). Si falla, rollback y
            reintenta cada fila con SAVEPOINT (slow path).
            batch_rows: lista [(r, payload, password, username), ...].
            Modifica `created` y `create_errors` in-place.
            """
            if not batch_rows:
                return
            try:
                db.session.commit()
                # Fast path OK: todos los usuarios persistieron.
                for r, payload, password, username in batch_rows:
                    _record_created(r, payload, password, username, payload['id'])
                return
            except HTTPException:
                raise
            except Exception as batch_err:
                import logging as _lg_b
                _lg_b.getLogger(__name__).warning(
                    f'[BULK-UPLOAD] Commit de lote falló ({len(batch_rows)} filas): '
                    f'{str(batch_err)[:200]}. Reintentando individualmente con SAVEPOINT.'
                )
                try:
                    db.session.rollback()
                except Exception:
                    pass

                # Slow path: cada fila aislada con SAVEPOINT. Las válidas
                # persisten; las inválidas van a errors con mensaje específico.
                for r, payload, password, username in batch_rows:
                    try:
                        sp = db.session.begin_nested()
                        try:
                            u = _instantiate_and_add(payload, password)
                            sp.commit()
                            _record_created(r, payload, password, username, u.id)
                        except Exception as ex_inner:
                            try:
                                sp.rollback()
                            except Exception:
                                pass
                            create_errors.append({
                                'row': r['row'],
                                'email': r['email'] or '(vacío)',
                                'error': str(ex_inner)[:300],
                            })
                    except Exception as ex_outer:
                        create_errors.append({
                            'row': r['row'],
                            'email': r['email'] or '(vacío)',
                            'error': f'savepoint failed: {str(ex_outer)[:200]}',
                        })

                try:
                    db.session.commit()
                except Exception as commit_after_sp:
                    try:
                        db.session.rollback()
                    except Exception:
                        pass
                    _lg_b.getLogger(__name__).error(
                        f'[BULK-UPLOAD] Commit post-savepoints falló: {commit_after_sp}'
                    )

        # ── Bucle principal ──
        batch_rows = []  # [(r, payload, password, username), ...]
        for r in to_create:
            try:
                payload, password, username = _build_user_payload(r)
                _instantiate_and_add(payload, password)
                batch_rows.append((r, payload, password, username))
                if len(batch_rows) >= BATCH_COMMIT_SIZE:
                    _commit_batch(batch_rows)
                    batch_rows = []
            except HTTPException:
                raise
            except Exception as e:
                # Error al construir/agregar a la sesión (antes del commit).
                # Procesamos lo acumulado para no perderlo y registramos esta fila.
                if batch_rows:
                    _commit_batch(batch_rows)
                    batch_rows = []
                create_errors.append({
                    'row': r.get('row'),
                    'email': r.get('email') or '(vacío)',
                    'error': str(e)[:300],
                })

        # Commit remanente
        if batch_rows:
            _commit_batch(batch_rows)
            batch_rows = []

        batch_count = 0  # vestigial; evita romper código posterior
        if False:  # bloque legado deshabilitado
            try:
                db.session.commit()
            except HTTPException:
                raise
            except Exception as e:
                db.session.rollback()
                import logging
                logging.getLogger(__name__).error(f'Error en commit final bulk: {e}')

        # ======= RECONCILIACIÓN: filtrar created[] contra DB real =======
        # Si un commit (intermedio o final) fue rolleado por error de unique/FK,
        # los usuarios quedan en `created[]` aunque NO persistieron en DB. Esto
        # provocaba que el historial reportara N creados pero el grupo solo
        # tuviera M < N (issue reportado). Reconciliamos pidiendo a la DB qué
        # usernames sí existen y movemos los faltantes a create_errors.
        if created:
            try:
                _all_usernames = [c['username'] for c in created]
                _persisted = set()
                for _i in range(0, len(_all_usernames), CHUNK):
                    _chunk = _all_usernames[_i:_i + CHUNK]
                    _rows = User.query.filter(
                        User.username.in_(_chunk)
                    ).with_entities(User.username).all()
                    _persisted.update(r.username for r in _rows)
                if len(_persisted) != len(created):
                    _orphans = [c for c in created if c['username'] not in _persisted]
                    _kept = [c for c in created if c['username'] in _persisted]
                    for _o in _orphans:
                        create_errors.append({
                            'row': _o.get('row'),
                            'email': _o.get('email') or '(vacío)',
                            'error': 'No se persistió por error de DB (commit rolleado)',
                        })
                    import logging as _lg_recon
                    _lg_recon.getLogger(__name__).warning(
                        f'[BULK-UPLOAD] Reconciliación: {len(_orphans)} usuarios reportados '
                        f'como creados pero NO persistieron en DB (de {len(created)} esperados)'
                    )
                    created = _kept
                    # También limpiar curp_pending_user_ids — los IDs ya no existen
                    # (Las IDs eran UUID generados; al rollback no entraron a DB,
                    # así que tampoco hay GroupMembers que asignar para ellos)
            except Exception as _recon_err:
                import logging as _lg_recon2
                _lg_recon2.getLogger(__name__).error(
                    f'[BULK-UPLOAD] Error en reconciliación: {_recon_err}'
                )

        # Asignar a grupo si aplica
        group_assignment = None
        if target_group and (created or existing_assigned):
            from app.models.partner import GroupMember
            assigned_new = 0
            assigned_existing = 0
            assignment_errors = []

            # Batch-fetch membresías del grupo destino
            all_assign_ids = []
            if created:
                cnames = [c['username'] for c in created]
                for i in range(0, len(cnames), CHUNK):
                    chunk = cnames[i:i + CHUNK]
                    users = User.query.filter(User.username.in_(chunk)).with_entities(User.id, User.username).all()
                    all_assign_ids.extend([(u.id, u.username) for u in users])
            for ea in existing_assigned:
                all_assign_ids.append((ea['user_id'], ea.get('username', '?')))

            uid_list = [x[0] for x in all_assign_ids]
            existing_members = set()
            for i in range(0, len(uid_list), CHUNK):
                chunk = uid_list[i:i + CHUNK]
                members = GroupMember.query.filter(
                    GroupMember.group_id == target_group.id,
                    GroupMember.user_id.in_(chunk)
                ).with_entities(GroupMember.user_id).all()
                existing_members.update(m.user_id for m in members)

            batch_count = 0
            for uid, uname in all_assign_ids:
                if uid not in existing_members:
                    try:
                        # Política: el candidato siempre se ve en el grupo desde el
                        # inicio. Si su CURP no se valida, el usuario será redirigido
                        # a /mi-curp en su login. La cola de verificación corre en
                        # background y solo cambia el status si RENAPO RECHAZA
                        # definitivamente la CURP (curp_required) tras MAX intentos.
                        db.session.add(GroupMember(group_id=target_group.id, user_id=uid, status='active'))
                        if uid in {ea['user_id'] for ea in existing_assigned}:
                            assigned_existing += 1
                        else:
                            assigned_new += 1
                        batch_count += 1
                        if batch_count >= BATCH_COMMIT_SIZE:
                            try:
                                db.session.commit()
                            except Exception as _gm_commit_err:
                                import logging as _lg_gm
                                _lg_gm.getLogger(__name__).error(
                                    f'[BULK-UPLOAD] Error en commit intermedio GroupMember: {_gm_commit_err}'
                                )
                                try:
                                    db.session.rollback()
                                except Exception:
                                    pass
                                assignment_errors.append({
                                    'username': '(batch)',
                                    'error': f'Commit intermedio falló: {str(_gm_commit_err)[:200]}',
                                })
                            batch_count = 0
                    except HTTPException:
                        raise
                    except Exception as e:
                        assignment_errors.append({'username': uname, 'error': str(e)})

            if batch_count > 0:
                try:
                    db.session.commit()
                except Exception as _gm_final_err:
                    import logging as _lg_gm2
                    _lg_gm2.getLogger(__name__).error(
                        f'[BULK-UPLOAD] Error en commit final GroupMember: {_gm_final_err}'
                    )
                    try:
                        db.session.rollback()
                    except Exception:
                        pass
                    assignment_errors.append({
                        'username': '(commit final)',
                        'error': f'Commit final falló: {str(_gm_final_err)[:200]}',
                    })

            # Reconciliar: contar GroupMembers reales en DB vs los que esperábamos crear
            try:
                _real_count = GroupMember.query.filter(
                    GroupMember.group_id == target_group.id,
                    GroupMember.user_id.in_([x[0] for x in all_assign_ids])
                ).count()
                _expected = len(all_assign_ids)
                if _real_count != _expected:
                    import logging as _lg_recon_gm
                    _lg_recon_gm.getLogger(__name__).warning(
                        f'[BULK-UPLOAD] Reconciliación GroupMember: esperados {_expected}, '
                        f'reales en DB {_real_count} (grupo {target_group.id})'
                    )
                    # Ajustar contadores reportados al usuario
                    _diff = _expected - _real_count
                    if _diff > 0:
                        # Restar prioritariamente de assigned_new (los recién creados)
                        _restar_new = min(_diff, assigned_new)
                        assigned_new -= _restar_new
                        assigned_existing -= (_diff - _restar_new)
                        if assigned_existing < 0:
                            assigned_existing = 0
                        assignment_errors.append({
                            'username': '(reconciliación)',
                            'error': f'{_diff} GroupMember(s) no persistieron en DB',
                        })
            except Exception as _recon_gm_err:
                import logging as _lg_recon_gm2
                _lg_recon_gm2.getLogger(__name__).error(
                    f'[BULK-UPLOAD] Error reconciliando GroupMember: {_recon_gm_err}'
                )

            group_assignment = {
                'group_id': target_group.id,
                'group_name': target_group.name,
                'assigned': assigned_new + assigned_existing,
                'assigned_new': assigned_new,
                'assigned_existing': assigned_existing,
                'errors': assignment_errors,
            }

        # Combinar errores
        all_errors = validation_errors + create_errors

        # ======= GUARDAR REGISTRO DE CARGA MASIVA EN HISTORIAL =======
        # IMPORTANTE: se guarda ANTES del envío de emails para garantizar que siempre
        # quede registro aunque el envío de emails tarde o falle.
        batch_record_id = None
        try:
            from app.models.partner import BulkUploadBatch, BulkUploadMember, Campus as CampusModel

            # Obtener info del campus/partner/grupo para snapshot
            _partner_name = None
            _campus_name = None
            _country = None
            _state_name = None
            _partner_id = None
            _campus_id = None

            if target_group:
                campus_obj = CampusModel.query.get(target_group.campus_id)
                if campus_obj:
                    _campus_id = campus_obj.id
                    _campus_name = campus_obj.name
                    _country = campus_obj.country
                    _state_name = campus_obj.state_name
                    if campus_obj.partner_id:
                        _partner_id = campus_obj.partner_id
                        from app.models.partner import Partner as PartnerModel
                        partner_obj = PartnerModel.query.get(campus_obj.partner_id)
                        if partner_obj:
                            _partner_name = partner_obj.name

            original_filename = file.filename if hasattr(file, 'filename') else None

            batch_rec = BulkUploadBatch(
                uploaded_by_id=current_user.id,
                partner_id=_partner_id,
                campus_id=_campus_id,
                group_id=target_group.id if target_group else None,
                partner_name=_partner_name,
                campus_name=_campus_name,
                group_name=target_group.name if target_group else None,
                country=_country,
                state_name=_state_name,
                total_processed=len(parsed_rows),
                total_created=len(created),
                total_existing_assigned=len(existing_assigned),
                total_errors=len(all_errors),
                total_skipped=len(skipped),
                emails_sent=0,      # Se actualiza en segundo plano tras el envío async
                emails_failed=0,
                original_filename=original_filename,
            )
            db.session.add(batch_rec)
            db.session.flush()  # get batch_rec.id

            # Build username->user_id map from created users
            created_user_ids = {}
            if created:
                cnames = [c['username'] for c in created]
                for i in range(0, len(cnames), CHUNK):
                    chunk = cnames[i:i + CHUNK]
                    users = User.query.filter(User.username.in_(chunk)).with_entities(User.id, User.username).all()
                    for u in users:
                        created_user_ids[u.username] = u.id

            # Save created members
            for c in created:
                db.session.add(BulkUploadMember(
                    batch_id=batch_rec.id,
                    user_id=created_user_ids.get(c['username']),
                    row_number=c.get('row'),
                    email=c.get('email'),
                    full_name=c.get('full_name', c.get('name', '')),
                    username=c.get('username'),
                    curp=c.get('curp'),
                    gender=c.get('gender'),
                    status='created',
                ))

            # Save existing_assigned members
            for ea in existing_assigned:
                db.session.add(BulkUploadMember(
                    batch_id=batch_rec.id,
                    user_id=ea.get('user_id'),
                    row_number=ea.get('row'),
                    email=ea.get('email'),
                    full_name=ea.get('name', ''),
                    username=ea.get('username'),
                    status='existing_assigned',
                ))

            # Save errors
            for err in all_errors:
                db.session.add(BulkUploadMember(
                    batch_id=batch_rec.id,
                    row_number=err.get('row'),
                    email=err.get('email'),
                    full_name=err.get('nombre', err.get('name', '')),
                    status='error',
                    error_message=(err.get('error') or '')[:500],
                ))

            # Save skipped
            for s in skipped:
                db.session.add(BulkUploadMember(
                    batch_id=batch_rec.id,
                    user_id=s.get('user_id'),
                    row_number=s.get('row'),
                    email=s.get('email'),
                    full_name=s.get('name', ''),
                    username=s.get('username'),
                    status='skipped',
                    error_message=(s.get('reason') or '')[:500],
                ))

            db.session.commit()
            batch_record_id = batch_rec.id
        except Exception as batch_err:
            import logging
            logging.getLogger(__name__).error(f'Error saving bulk upload history: {batch_err}')
            try:
                db.session.rollback()
            except Exception:
                pass

        # ======= ENVÍO DE EMAILS DE BIENVENIDA (EN SEGUNDO PLANO) =======
        # Se ejecuta en un thread independiente para evitar timeouts en cargas grandes.
        # Los contadores emails_sent/emails_failed se actualizan en el batch record
        # una vez que el thread termina.
        emails_sent = 0
        emails_failed = 0
        if created:
            import threading
            from flask import current_app as _cur_app
            _email_app = _cur_app._get_current_object()
            _emails_payload = [
                {'username': c['username'], 'email': c.get('email'), 'password': c['password']}
                for c in created if c.get('email')
            ]

            def _send_welcome_emails_bg(app_obj, payload, b_id):
                import logging as _log
                _logger = _log.getLogger(__name__)
                sent = 0
                failed = 0
                with app_obj.app_context():
                    try:
                        from app.services.email_service import send_welcome_email
                        usernames = [p['username'] for p in payload]
                        users_map = {}
                        for i in range(0, len(usernames), 500):
                            chunk_u = User.query.filter(
                                User.username.in_(usernames[i:i + 500])
                            ).all()
                            for u in chunk_u:
                                users_map[u.username] = u
                        for p in payload:
                            if p['email'] and p['username'] in users_map:
                                try:
                                    send_welcome_email(users_map[p['username']], p['password'])
                                    sent += 1
                                except Exception:
                                    failed += 1
                    except Exception as _e:
                        _logger.error(f'Error enviando welcome emails bulk (bg): {_e}')
                    # Actualizar contadores en el batch record
                    if b_id:
                        try:
                            from app.models.partner import BulkUploadBatch as _BUB
                            _batch = _BUB.query.get(b_id)
                            if _batch:
                                _batch.emails_sent = sent
                                _batch.emails_failed = failed
                                db.session.commit()
                        except Exception as _ue:
                            _logger.error(f'Error actualizando email counts en batch {b_id}: {_ue}')
                            try:
                                db.session.rollback()
                            except Exception:
                                pass

            threading.Thread(
                target=_send_welcome_emails_bg,
                args=(_email_app, _emails_payload, batch_record_id),
                daemon=True,
            ).start()

        response_data = {
            'message': f'Proceso completado: {len(created)} creados, {len(existing_assigned)} existentes asignados',
            'summary': {
                'total_processed': len(parsed_rows),
                'created': len(created),
                'existing_assigned': len(existing_assigned),
                'errors': len(all_errors),
                'skipped': len(skipped),
                'emails_sent': emails_sent,
                'emails_failed': emails_failed,
            },
            'details': {
                'created': created,
                'existing_assigned': existing_assigned,
                'errors': all_errors,
                'skipped': skipped,
                'total_processed': len(parsed_rows),
            },
        }
        if group_assignment:
            response_data['group_assignment'] = group_assignment
        if batch_record_id:
            response_data['batch_id'] = batch_record_id

        # ======= ENCOLAR VERIFICACIÓN CURP =======
        # En vez de threads volátiles, encolar en curp_verification_queue.
        # Un worker en background (services/curp_queue_worker.py) procesa
        # con cache + circuit-breaker + reintentos persistentes.
        # NOTA: en modo LOCAL (CURP_RENAPO_ENABLED=false) NO encolamos —
        # ya marcamos curp_verified=True en _build_user_payload tras la
        # validación local en _validate_rows.
        from app.services.curp_local_validator import is_renapo_enabled as _renapo_on_bulk_q
        if _renapo_on_bulk_q():
            curp_users_to_verify = [
                c for c in created
                if c.get('curp') and c['curp'] not in GENERIC_FOREIGN_CURPS
            ]
            if curp_users_to_verify and batch_record_id:
                try:
                    from app.services.curp_queue_worker import enqueue_curp_verification
                    import logging as _lg_curp
                    _logger_curp = _lg_curp.getLogger(__name__)
                    enqueued = 0
                    for c in curp_users_to_verify:
                        uid = c.get('user_id') or c.get('id')
                        curp_val = c.get('curp')
                        if uid and curp_val:
                            if enqueue_curp_verification(uid, curp_val, source='bulk', batch_id=batch_record_id):
                                enqueued += 1
                    _logger_curp.info(f'[BULK-CURP] {enqueued} CURPs encoladas para validación (batch {batch_record_id})')
                except Exception as enq_err:
                    import logging as _lg_curp_err
                    _lg_curp_err.getLogger(__name__).error(f'[BULK-CURP] Error encolando verificaciones: {enq_err}')

        return jsonify(response_data), 200

    except HTTPException:

        raise

    except Exception as e:
        db.session.rollback()
        logger.exception('user_management error')
        return jsonify({'error': 'Error interno del servidor'}), 500


@bp.route('/candidates/bulk-upload/template', methods=['GET'])
@jwt_required()
@management_required
@rate_limit(limit=60, window=60, key_prefix='rl_um_tmpl')  # UM-N12
def download_bulk_upload_template():
    """
    Descargar plantilla Excel para carga masiva de candidatos
    """
    try:
        current_user = g.current_user
        bulk_perm_err = _ensure_bulk_upload_allowed(current_user)
        if bulk_perm_err:
            return bulk_perm_err

        import io
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from flask import send_file
        
        # Crear workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Candidatos"
        
        # Estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        required_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        optional_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Encabezados - obligatorios a la izquierda, opcionales a la derecha
        headers = [
            ('nombre', 'Requerido'),
            ('primer_apellido', 'Requerido'),
            ('segundo_apellido', 'Requerido'),
            ('genero', 'Requerido (M, F, O)'),
            ('email', 'Opcional (sin email no recibe insignia)'),
            ('curp', 'Opcional (sin CURP no recibe cert. CONOCER)')
        ]
        
        for col, (header, _) in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
        
        # Segunda fila con descripción
        for col, (_, description) in enumerate(headers, start=1):
            cell = ws.cell(row=2, column=col, value=description)
            if 'Requerido' in description:
                cell.fill = required_fill
            else:
                cell.fill = optional_fill
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
            cell.border = thin_border
        
        # Ejemplo de datos (obligatorios primero, opcionales al final)
        example_data = [
            ('Juan', 'García', 'López', 'M', 'candidato1@email.com', 'GALJ900101HDFRPR01'),
            ('María', 'Pérez', 'Sánchez', 'F', '', ''),  # Sin email ni CURP
        ]
        
        for row_idx, data in enumerate(example_data, start=3):
            for col_idx, value in enumerate(data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
        
        # Ajustar ancho de columnas
        column_widths = [15, 18, 18, 18, 30, 25]
        for col, width in enumerate(column_widths, start=1):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width
        
        # Hoja de instrucciones
        ws_instructions = wb.create_sheet("Instrucciones")
        instructions = [
            "INSTRUCCIONES PARA CARGA MASIVA DE CANDIDATOS",
            "",
            "1. Use la hoja 'Candidatos' para ingresar los datos",
            "2. No modifique los encabezados de las columnas",
            "3. Elimine las filas de ejemplo antes de agregar sus datos",
            "",
            "CAMPOS REQUERIDOS:",
            "- nombre: Nombre(s) del candidato",
            "- primer_apellido: Apellido paterno",
            "- segundo_apellido: Apellido materno",
            "- genero: M (Masculino), F (Femenino), O (Otro)",
            "",
            "CAMPOS OPCIONALES:",
            "- email: Sin email, el candidato NO podrá recibir insignia digital",
            "- curp: Sin CURP, el candidato NO podrá recibir certificado CONOCER",
            "  (El reporte de evaluación y certificado Eduit siempre están disponibles)",
            "",
            "GENERACIÓN AUTOMÁTICA:",
            "- Las contraseñas se generan automáticamente para cada usuario",
            "- Los usernames se generan a partir del email o nombre+apellido",
            "",
            "NOTAS:",
            "- Los emails duplicados serán omitidos",
            "- Los CURP duplicados serán omitidos",
            "- Se genera un reporte con los resultados de la carga"
        ]
        
        for row_idx, text in enumerate(instructions, start=1):
            cell = ws_instructions.cell(row=row_idx, column=1, value=text)
            if row_idx == 1:
                cell.font = Font(bold=True, size=14)
            elif text.endswith(':'):
                cell.font = Font(bold=True)
        
        ws_instructions.column_dimensions['A'].width = 60
        
        # Guardar en memoria
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='plantilla_candidatos.xlsx'
        )
        
    except HTTPException:
        
        raise
        
    except Exception as e:
        logger.exception('user_management error')
        return jsonify({'error': 'Error interno del servidor'}), 500


# ============== EXPORTAR USUARIOS CON CONTRASEÑAS (ADMIN Y COORDINADOR) ==============

@bp.route('/export-credentials', methods=['POST'])
@jwt_required()
@management_required
@rate_limit(limit=10, window=60, key_prefix='rl_um_export')
def export_user_credentials():
    """
    Exportar usuarios seleccionados con el mismo formato del reporte de altas masivas.
    Un usuario puede pertenecer a varios grupos, por lo que se genera una fila por cada
    combinación usuario-grupo (o una sola fila si no tiene grupo).
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from flask import send_file
        import io
        from app.models.partner import Campus, CandidateGroup, GroupMember, Partner
        
        current_user = g.current_user
        data = request.get_json()
        user_ids = data.get('user_ids', [])
        
        if not user_ids:
            return jsonify({'error': 'Debes seleccionar al menos un usuario'}), 400

        # UM-N30: límite generoso (20K) para no OOMear el worker.
        if len(user_ids) > 20000:
            return jsonify({
                'error': 'No puedes exportar más de 20,000 usuarios en una sola operación. '
                         'Divide la selección en lotes más pequeños.'
            }), 400

        # Obtener usuarios
        CHUNK = 500
        users_map = {}
        for i in range(0, len(user_ids), CHUNK):
            chunk = user_ids[i:i + CHUNK]
            # UM-C3: excluir usuarios soft-eliminados.
            q = User.query.filter(User.id.in_(chunk), User.is_deleted == False)  # noqa: E712
            # UM-C1: respetar el tenant del solicitante para no filtrar credenciales
            # de otros tenants vía IDs adivinados.
            if _is_coordinator_role(current_user.role):
                coord_id = _get_effective_coordinator_id(current_user)
                q = q.filter(
                    or_(
                        User.role == 'candidato',
                        and_(
                            User.role.in_(['responsable', 'responsable_partner', 'responsable_estatal', 'auxiliar']),
                            User.coordinator_id == coord_id
                        )
                    )
                )
            elif current_user.role == 'soporte':
                q = q.filter(User.role.in_(SOPORTE_VISIBLE_ROLES))
            elif current_user.role == 'responsable':
                from app.models.partner import GroupMember as _GM, CandidateGroup as _CG
                _campus_uids = db.session.query(_GM.user_id).join(
                    _CG, _GM.group_id == _CG.id
                ).filter(_CG.campus_id == current_user.campus_id).distinct().subquery()
                q = q.filter(
                    User.role == 'candidato',
                    or_(
                        User.id.in_(_campus_uids),
                        User.campus_id == current_user.campus_id,
                    )
                )
            for u in q.all():
                users_map[u.id] = u
        
        if not users_map:
            return jsonify({'error': 'No se encontraron usuarios'}), 404
        
        # Obtener membresías de grupo para todos los usuarios seleccionados
        all_uids = list(users_map.keys())
        memberships = []
        for i in range(0, len(all_uids), CHUNK):
            chunk = all_uids[i:i + CHUNK]
            memberships.extend(
                db.session.query(GroupMember.user_id, GroupMember.group_id)
                .filter(GroupMember.user_id.in_(chunk), GroupMember.status == 'active')
                .all()
            )
        
        # Mapear user_id → [group_ids]
        user_group_ids = {}
        all_group_ids = set()
        for uid, gid in memberships:
            user_group_ids.setdefault(uid, []).append(gid)
            all_group_ids.add(gid)
        
        # Cargar grupos con sus campus
        groups_map = {}
        if all_group_ids:
            for i in range(0, len(list(all_group_ids)), CHUNK):
                chunk = list(all_group_ids)[i:i + CHUNK]
                for grp in CandidateGroup.query.filter(CandidateGroup.id.in_(chunk)).all():
                    groups_map[grp.id] = grp
        
        # Cargar campus
        all_campus_ids = set(grp.campus_id for grp in groups_map.values())
        # Agregar campus_id de usuarios sin grupo
        for u in users_map.values():
            if u.campus_id:
                all_campus_ids.add(u.campus_id)
        campus_map = {}
        if all_campus_ids:
            for i in range(0, len(list(all_campus_ids)), CHUNK):
                chunk = list(all_campus_ids)[i:i + CHUNK]
                for c in Campus.query.filter(Campus.id.in_(chunk)).all():
                    campus_map[c.id] = c
        
        # Cargar partners
        all_partner_ids = set(c.partner_id for c in campus_map.values())
        partner_map = {}
        if all_partner_ids:
            for p in Partner.query.filter(Partner.id.in_(list(all_partner_ids))).all():
                partner_map[p.id] = p
        
        # Cargar responsables de campus
        resp_ids = set(c.responsable_id for c in campus_map.values() if c.responsable_id)
        resp_map = {}
        if resp_ids:
            for u in User.query.filter(User.id.in_(list(resp_ids))).all():
                resp_map[u.id] = u
        
        # Crear Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Usuarios Seleccionados"
        
        # Estilos
        header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True, size=11)
        cell_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        headers = [
            'Partner', 'País', 'Estado', 'Plantel', 'Grupo',
            'Nombre de Usuario', 'Nombre Completo',
            'CURP', 'Género', 'Email del Usuario',
            'Email del Responsable', 'Contraseña',
        ]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = cell_border
        
        gender_labels = {'M': 'Masculino', 'F': 'Femenino', 'O': 'Otro'}
        
        # Helper para construir fila de datos
        def build_row(user, campus, partner, group_name, resp_email):
            parts = [user.name or '', user.first_surname or '']
            if user.second_surname:
                parts.append(user.second_surname)
            full_name = ' '.join(p for p in parts if p)
            gender_display = gender_labels.get(user.gender, user.gender or '-')
            password = user.get_decrypted_password() or '(no disponible)'
            return [
                partner.name if partner else '-',
                (campus.country if campus else None) or '-',
                (campus.state_name if campus else None) or '-',
                (campus.name if campus else None) or '-',
                group_name or '-',
                user.username or '-',
                full_name or '-',
                user.curp or '-',
                gender_display,
                user.email or '-',
                resp_email or '-',
                password,
            ]
        
        # Generar filas: una por cada combinación usuario-grupo
        row_idx = 2
        for uid, user in users_map.items():
            group_ids = user_group_ids.get(uid, [])
            
            if group_ids:
                for gid in group_ids:
                    grp = groups_map.get(gid)
                    campus = campus_map.get(grp.campus_id) if grp else None
                    partner = partner_map.get(campus.partner_id) if campus else None
                    resp_email = '-'
                    if campus and campus.responsable_id:
                        resp_user = resp_map.get(campus.responsable_id)
                        if resp_user:
                            resp_email = resp_user.email or '-'
                    
                    row_data = build_row(user, campus, partner, grp.name if grp else '-', resp_email)
                    for col, value in enumerate(row_data, start=1):
                        cell = ws.cell(row=row_idx, column=col, value=value)
                        cell.border = cell_border
                        cell.alignment = Alignment(vertical='center')
                    row_idx += 1
            else:
                # Usuario sin grupo — usar su campus_id directo si tiene
                campus = campus_map.get(user.campus_id) if user.campus_id else None
                partner = partner_map.get(campus.partner_id) if campus else None
                resp_email = '-'
                if campus and campus.responsable_id:
                    resp_user = resp_map.get(campus.responsable_id)
                    if resp_user:
                        resp_email = resp_user.email or '-'
                
                row_data = build_row(user, campus, partner, '-', resp_email)
                for col, value in enumerate(row_data, start=1):
                    cell = ws.cell(row=row_idx, column=col, value=value)
                    cell.border = cell_border
                    cell.alignment = Alignment(vertical='center')
                row_idx += 1
        
        # Ajustar anchos
        column_widths = [25, 15, 20, 25, 20, 20, 35, 22, 12, 30, 30, 20]
        for col, width in enumerate(column_widths, start=1):
            letter = chr(64 + col) if col <= 26 else chr(64 + (col // 26)) + chr(64 + (col % 26))
            ws.column_dimensions[letter].width = width
        
        ws.freeze_panes = 'A2'
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'credenciales_usuarios_{len(users_map)}.xlsx'
        )
        
    except HTTPException:
        
        raise
        
    except Exception as e:
        logger.exception('user_management error')
        return jsonify({'error': 'Error interno del servidor'}), 500


# ============== HISTÓRICO DE ALTAS MASIVAS ==============

@bp.route('/bulk-history', methods=['GET'])
@jwt_required()
@management_required
@rate_limit(limit=300, window=60, key_prefix='rl_um_bhist')  # UM-N13
def list_bulk_upload_history():
    """Listar todas las cargas masivas con paginación y filtros"""
    try:
        from app.models.partner import BulkUploadBatch

        current_user = g.current_user
        bulk_perm_err = _ensure_bulk_upload_allowed(current_user)
        if bulk_perm_err:
            return bulk_perm_err
        page = request.args.get('page', 1, type=int)
        per_page = request.args.get('per_page', 20, type=int)
        per_page = min(per_page, 100)

        query = BulkUploadBatch.query

        # UM-C2: respetar tenant scope.
        if _is_coordinator_role(current_user.role):
            eff_id = _get_effective_coordinator_id(current_user)
            query = query.filter(BulkUploadBatch.uploaded_by_id == eff_id)
        elif current_user.role in ('responsable', 'responsable_partner', 'responsable_estatal', 'soporte'):
            query = query.filter(BulkUploadBatch.uploaded_by_id == current_user.id)

        # Filtros opcionales
        partner_id = request.args.get('partner_id', type=int)
        campus_id = request.args.get('campus_id', type=int)
        if partner_id:
            query = query.filter(BulkUploadBatch.partner_id == partner_id)
        if campus_id:
            query = query.filter(BulkUploadBatch.campus_id == campus_id)

        query = query.order_by(BulkUploadBatch.created_at.desc())
        pagination = query.paginate(page=page, per_page=per_page, error_out=False)

        return jsonify({
            'batches': [b.to_dict() for b in pagination.items],
            'total': pagination.total,
            'page': pagination.page,
            'per_page': pagination.per_page,
            'pages': pagination.pages,
        }), 200
    except HTTPException:
        raise
    except Exception as e:
        logger.exception('user_management error')
        return jsonify({'error': 'Error interno del servidor'}), 500


@bp.route('/bulk-history/<int:batch_id>', methods=['GET'])
@jwt_required()
@management_required
@rate_limit(limit=300, window=60, key_prefix='rl_um_bdet')  # UM-N14
def get_bulk_upload_detail(batch_id):
    """Obtener detalle de una carga masiva (resumen + miembros)"""
    try:
        from app.models.partner import BulkUploadBatch
        from werkzeug.exceptions import HTTPException

        current_user = g.current_user
        bulk_perm_err = _ensure_bulk_upload_allowed(current_user)
        if bulk_perm_err:
            return bulk_perm_err
        batch = BulkUploadBatch.query.get_or_404(batch_id)

        # UM-C2: respetar tenant scope.
        if _is_coordinator_role(current_user.role):
            eff_id = _get_effective_coordinator_id(current_user)
            if batch.uploaded_by_id != eff_id:
                return jsonify({'error': 'No tienes acceso a este registro'}), 403
        elif current_user.role in ('responsable', 'responsable_partner', 'responsable_estatal', 'soporte'):
            if batch.uploaded_by_id != current_user.id:
                return jsonify({'error': 'No tienes acceso a este registro'}), 403

        return jsonify(batch.to_dict(include_members=True)), 200
    except HTTPException:
        raise
    except HTTPException:
        raise
    except Exception as e:
        logger.exception('user_management error')
        return jsonify({'error': 'Error interno del servidor'}), 500


@bp.route('/bulk-history/<int:batch_id>/curp-progress', methods=['GET'])
@jwt_required()
@management_required
@rate_limit(limit=600, window=60, key_prefix='rl_um_curp_prog')
def get_bulk_curp_progress(batch_id):
    """Progreso de validación RENAPO para un batch de carga masiva.

    Devuelve conteos por estado de la cola CurpVerificationQueue + estado
    final de los usuarios del batch. Pensado para polling desde la UI
    mientras el worker en background procesa.
    """
    try:
        from app.models.partner import BulkUploadBatch, BulkUploadMember
        from app.models.curp_verification import (
            CurpVerificationQueue,
            QUEUE_PENDING, QUEUE_PROCESSING, QUEUE_DONE,
            QUEUE_FAILED, QUEUE_REJECTED,
        )
        from werkzeug.exceptions import HTTPException

        current_user = g.current_user
        bulk_perm_err = _ensure_bulk_upload_allowed(current_user)
        if bulk_perm_err:
            return bulk_perm_err
        batch = BulkUploadBatch.query.get_or_404(batch_id)

        # Tenant scope idéntico al detail endpoint.
        if _is_coordinator_role(current_user.role):
            eff_id = _get_effective_coordinator_id(current_user)
            if batch.uploaded_by_id != eff_id:
                return jsonify({'error': 'No tienes acceso a este registro'}), 403
        elif current_user.role in ('responsable', 'responsable_partner', 'responsable_estatal', 'soporte'):
            if batch.uploaded_by_id != current_user.id:
                return jsonify({'error': 'No tienes acceso a este registro'}), 403

        # Conteos por status en la cola para este batch.
        counts = {
            'pending': 0,
            'processing': 0,
            'done': 0,
            'failed': 0,
            'rejected': 0,
        }
        rows = db.session.query(
            CurpVerificationQueue.status,
            db.func.count(CurpVerificationQueue.id),
        ).filter(
            CurpVerificationQueue.batch_id == batch_id
        ).group_by(CurpVerificationQueue.status).all()
        for status, n in rows:
            if status in counts:
                counts[status] = int(n)

        total_enqueued = sum(counts.values())
        finished = counts['done'] + counts['failed'] + counts['rejected']
        in_progress = counts['pending'] + counts['processing']

        # Conteo de miembros del batch que ya quedaron como 'curp_verified'
        # (lo escribe el worker tras éxito) — útil cuando filas viejas en
        # la cola ya no existen.
        verified_members = BulkUploadMember.query.filter(
            BulkUploadMember.batch_id == batch_id,
            BulkUploadMember.status == 'curp_verified',
        ).count()

        return jsonify({
            'batch_id': batch_id,
            'total_created': batch.total_created or 0,
            'total_enqueued': total_enqueued,
            'in_progress': in_progress,
            'finished': finished,
            'counts': counts,
            'verified_members': verified_members,
            'is_complete': total_enqueued > 0 and in_progress == 0,
        }), 200
    except HTTPException:
        raise
    except Exception:
        logger.exception('user_management error get_bulk_curp_progress')
        return jsonify({'error': 'Error interno del servidor'}), 500


@bp.route('/bulk-history/<int:batch_id>/export', methods=['GET'])
@jwt_required()
@management_required
def export_bulk_upload_batch(batch_id):
    """Exportar a Excel los candidatos de una carga masiva"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from flask import send_file
        from app.models.partner import BulkUploadBatch, BulkUploadMember, Campus
        from werkzeug.exceptions import HTTPException
        import io

        current_user = g.current_user
        bulk_perm_err = _ensure_bulk_upload_allowed(current_user)
        if bulk_perm_err:
            return bulk_perm_err
        batch = BulkUploadBatch.query.get_or_404(batch_id)

        # UM-C2: respetar tenant scope.
        if _is_coordinator_role(current_user.role):
            eff_id = _get_effective_coordinator_id(current_user)
            if batch.uploaded_by_id != eff_id:
                return jsonify({'error': 'No tienes acceso a este registro'}), 403
        elif current_user.role in ('responsable', 'responsable_partner', 'responsable_estatal', 'soporte'):
            if batch.uploaded_by_id != current_user.id:
                return jsonify({'error': 'No tienes acceso a este registro'}), 403

        # Filtro opcional por status
        status_filter = request.args.get('status', '').strip()
        valid_statuses = {'created', 'existing_assigned', 'error', 'skipped', 'curp_invalid', 'curp_verified'}

        # Obtener miembros con sus usuarios
        members_query = BulkUploadMember.query.filter_by(batch_id=batch_id)
        if status_filter and status_filter in valid_statuses:
            members_query = members_query.filter(BulkUploadMember.status == status_filter)
        members = members_query.order_by(BulkUploadMember.row_number).all()

        # Fetch users for getting passwords
        user_ids = [m.user_id for m in members if m.user_id]
        users_map = {}
        if user_ids:
            CHUNK = 500
            for i in range(0, len(user_ids), CHUNK):
                chunk = user_ids[i:i + CHUNK]
                users = User.query.filter(User.id.in_(chunk)).all()
                for u in users:
                    users_map[u.id] = u

        # Crear Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Altas Masivas"

        # Estilos
        header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True, size=11)
        cell_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        # Obtener email del responsable del plantel
        responsable_email = '-'
        if batch.campus_id:
            campus_obj = Campus.query.get(batch.campus_id)
            if campus_obj and campus_obj.responsable_id:
                resp_user = User.query.get(campus_obj.responsable_id)
                if resp_user and resp_user.email:
                    responsable_email = resp_user.email

        # Encabezados (orden solicitado)
        headers = [
            'Partner', 'País', 'Estado', 'Plantel', 'Grupo',
            'Fecha de Carga', 'Nombre de Usuario', 'Nombre Completo',
            'CURP', 'Género', 'Email del Usuario',
            'Email del Responsable', 'Contraseña', 'Estado de Carga'
        ]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = cell_border

        # Mapeo de género
        gender_labels = {'M': 'Masculino', 'F': 'Femenino', 'O': 'Otro'}

        # Datos
        status_labels = {
            'created': 'Creado',
            'existing_assigned': 'Existente asignado',
            'error': 'Error',
            'skipped': 'Omitido',
        }
        fecha_carga = batch.created_at.strftime('%d/%m/%Y %H:%M') if batch.created_at else '-'

        for row_idx, member in enumerate(members, start=2):
            user = users_map.get(member.user_id) if member.user_id else None
            password = user.get_decrypted_password() if user else '(no disponible)'

            # Nombre completo: preferir datos del User si existe (incluye segundo_apellido)
            full_name = member.full_name or ''
            if user:
                parts = [user.name or '', user.first_surname or '']
                if user.second_surname:
                    parts.append(user.second_surname)
                full_name = ' '.join(p for p in parts if p)

            # Género: preferir del User, fallback al member snapshot
            gender_raw = (user.gender if user and user.gender else member.gender) or '-'
            gender_display = gender_labels.get(gender_raw, gender_raw)

            row_data = [
                batch.partner_name or '-',
                batch.country or '-',
                batch.state_name or '-',
                batch.campus_name or '-',
                batch.group_name or '-',
                fecha_carga,
                member.username or '-',
                full_name or '-',
                (user.curp if user and user.curp else member.curp) or '-',
                gender_display,
                member.email or '-',
                responsable_email,
                password or '(no disponible)',
                status_labels.get(member.status, member.status),
            ]
            for col, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.border = cell_border
                cell.alignment = Alignment(vertical='center')

        # Ajustar anchos
        column_widths = [25, 15, 20, 25, 20, 18, 20, 35, 22, 12, 30, 30, 20, 18]
        for col, width in enumerate(column_widths, start=1):
            letter = chr(64 + col) if col <= 26 else chr(64 + (col // 26)) + chr(64 + (col % 26))
            ws.column_dimensions[letter].width = width

        ws.freeze_panes = 'A2'

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        safe_name = (batch.group_name or 'carga').replace(' ', '_')[:30]
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'altas_masivas_{safe_name}_{batch.id}.xlsx'
        )
    except HTTPException:
        raise
    except HTTPException:
        raise
    except Exception as e:
        logger.exception('user_management error')
        return jsonify({'error': 'Error interno del servidor'}), 500


# ---------------------------------------------------------------------------
# Recovery: re-verify CURPs for users orphaned by a container restart
# ---------------------------------------------------------------------------

def _recover_orphaned_curp_users(app_obj, users_data):
    """Re-verify CURPs for users stuck in curp_pending/curp_verifying after a restart.

    Uses 6-round iterative verification (same as bulk upload) because RENAPO
    is inconsistent — a valid CURP can fail on one query but succeed on another.
    Only after 6 failed rounds is the user deleted.
    """
    import time
    import logging
    MAX_ROUNDS = 6
    _logger = logging.getLogger(__name__)
    _logger.info(
        '[CURP-RECOVERY] Starting RENAPO verification for %d orphaned users — up to %d rounds',
        len(users_data), MAX_ROUNDS,
    )

    verified = 0
    format_invalid = 0

    try:
        with app_obj.app_context():
            from app.services.renapo_service import (
                validate_curp_renapo, apply_renapo_to_user, validate_curp_format,
            )
            from app.models.partner import BulkUploadMember, GroupMember

            # ── Pre-pass: filter format-invalid — NO se eliminan, quedan curp_required ──
            valid_format_users = []
            for u_data in users_data:
                curp = u_data['curp']
                username = u_data['username']
                try:
                    fmt_valid, fmt_error = validate_curp_format(curp)
                    if not fmt_valid:
                        _logger.warning('[CURP-RECOVERY] Invalid format %s (%s): %s — marcando curp_required', curp, username, fmt_error)
                        user = User.query.filter_by(username=username).first()
                        if user:
                            user.is_active = True
                            for gm in GroupMember.query.filter_by(user_id=user.id).all():
                                if gm.status in ('curp_pending', 'curp_verifying', 'active'):
                                    gm.status = 'curp_required'
                            rec = BulkUploadMember.query.filter_by(user_id=user.id).first()
                            if rec:
                                rec.status = 'curp_required'
                                rec.error_message = f'Formato CURP invalido (recovery): {fmt_error}. El candidato deberá corregirla al iniciar sesión.'
                            db.session.commit()
                        format_invalid += 1
                        continue
                    valid_format_users.append(u_data)
                except Exception as e:
                    _logger.error('[CURP-RECOVERY] Error pre-processing %s: %s', username, e)
                    try:
                        db.session.rollback()
                    except Exception:
                        pass
                    valid_format_users.append(u_data)

            # ── Iterative RENAPO rounds ──
            pending_users = valid_format_users
            for round_num in range(1, MAX_ROUNDS + 1):
                if not pending_users:
                    break

                _logger.info('[CURP-RECOVERY] Round %d/%d: verifying %d CURPs', round_num, MAX_ROUNDS, len(pending_users))
                still_failing = []

                for i, u_data in enumerate(pending_users):
                    curp = u_data['curp']
                    username = u_data['username']
                    try:
                        result = validate_curp_renapo(curp)
                        user = User.query.filter_by(username=username).first()
                        if not user:
                            _logger.warning('[CURP-RECOVERY] User %s no longer exists, skipping', username)
                            continue

                        if result.valid:
                            apply_renapo_to_user(user, result)
                            user.is_active = True
                            for gm in GroupMember.query.filter_by(user_id=user.id).filter(
                                GroupMember.status.in_(['curp_pending', 'curp_verifying'])
                            ).all():
                                gm.status = 'active'
                            rec = BulkUploadMember.query.filter_by(user_id=user.id).first()
                            if rec:
                                rec.status = 'curp_verified'
                            db.session.commit()
                            verified += 1
                            _logger.info('[CURP-RECOVERY] R%d (%d/%d) CURP %s valid — %s activated', round_num, i + 1, len(pending_users), curp, username)
                        else:
                            u_data['_last_error'] = result.error or 'no encontrada'
                            still_failing.append(u_data)
                            _logger.info('[CURP-RECOVERY] R%d (%d/%d) CURP %s rejected — %s (will retry)', round_num, i + 1, len(pending_users), curp, username)

                    except Exception as e:
                        _logger.error('[CURP-RECOVERY] R%d Error verifying %s (%s): %s', round_num, curp, username, e)
                        try:
                            db.session.rollback()
                        except Exception:
                            pass
                        u_data['_last_error'] = str(e)[:300]
                        still_failing.append(u_data)

                    if i < len(pending_users) - 1:
                        time.sleep(2)

                pending_users = still_failing
                if pending_users and round_num < MAX_ROUNDS:
                    _logger.info('[CURP-RECOVERY] Round %d done: %d CURPs pending, waiting 5s before round %d', round_num, len(pending_users), round_num + 1)
                    time.sleep(5)

            # ── After all rounds: dejar a los que fallaron en estado 'curp_required'
            # para que el candidato resuelva al iniciar sesión. NO se eliminan.
            final_rejected = 0
            for u_data in pending_users:
                username = u_data['username']
                curp = u_data['curp']
                last_error = u_data.get('_last_error', 'no encontrada')
                try:
                    user = User.query.filter_by(username=username).first()
                    if user:
                        _logger.warning('[CURP-RECOVERY] CURP %s no validada tras %d rondas — %s queda en curp_required', curp, MAX_ROUNDS, username)
                        user.is_active = True
                        for gm in GroupMember.query.filter_by(user_id=user.id).all():
                            if gm.status in ('curp_pending', 'curp_verifying', 'active'):
                                gm.status = 'curp_required'
                        rec = BulkUploadMember.query.filter_by(user_id=user.id).first()
                        if rec:
                            rec.status = 'curp_required'
                            rec.error_message = f'CURP rechazada tras {MAX_ROUNDS} rondas RENAPO (recovery): {last_error[:300]}. El candidato deberá validarla al iniciar sesión.'
                        db.session.commit()
                        final_rejected += 1
                except Exception as del_err:
                    _logger.error('[CURP-RECOVERY] Failed to mark curp_required %s: %s', username, del_err)
                    try:
                        db.session.rollback()
                    except Exception:
                        pass
                    final_rejected += 1

    except Exception as outer:
        _logger.error('[CURP-RECOVERY] Fatal error in recovery thread: %s', outer)

    _logger.info(
        '[CURP-RECOVERY] Finished: %d verified, %d format invalid, %d rejected after %d rounds',
        verified, format_invalid, final_rejected if 'final_rejected' in dir() else 0, MAX_ROUNDS,
    )


# ─────────────────────────────────────────────────────────────────────────────
# DASHBOARD CURP QUEUE (admin/developer) — monitoreo en tiempo real
# ─────────────────────────────────────────────────────────────────────────────

@bp.route('/curp-queue/dashboard', methods=['GET'])
@jwt_required()
@admin_required
@rate_limit(limit=600, window=60, key_prefix='rl_um_curp_dash')
def curp_queue_dashboard():
    """Dashboard real-time del worker de validación CURP.

    Devuelve métricas agregadas + filas activas/recientes para auditoría:
      - counts por status (pending, processing, done, failed, rejected)
      - circuit breaker RENAPO (open/closed, cooldown remaining)
      - filas actualmente en 'processing' (con lock)
      - últimas N filas terminadas (done/rejected/failed)
      - top errores recientes agrupados
      - throughput por hora (últimas 24h)
      - cache RENAPO: tamaño, hits totales, recientes
      - locks zombi (locked_at > 10 min)

    Admin/developer only. Polling-friendly (rate-limit 600 req/min).
    """
    try:
        from app.models.curp_verification import (
            CurpVerificationQueue, CurpRenapoCache,
            QUEUE_PENDING, QUEUE_PROCESSING, QUEUE_DONE,
            QUEUE_FAILED, QUEUE_REJECTED,
        )
        from sqlalchemy import desc

        now = datetime.utcnow()

        # 1) Counts por status
        counts = {QUEUE_PENDING: 0, QUEUE_PROCESSING: 0, QUEUE_DONE: 0,
                  QUEUE_FAILED: 0, QUEUE_REJECTED: 0}
        rows = db.session.query(
            CurpVerificationQueue.status,
            db.func.count(CurpVerificationQueue.id)
        ).group_by(CurpVerificationQueue.status).all()
        for status, n in rows:
            if status in counts:
                counts[status] = int(n)
        total = sum(counts.values())

        # 2) Filas en processing — incluyen locked_at para detectar zombis
        processing_rows = (
            CurpVerificationQueue.query
            .filter(CurpVerificationQueue.status == QUEUE_PROCESSING)
            .order_by(desc(CurpVerificationQueue.locked_at))
            .limit(50)
            .all()
        )
        processing_data = []
        zombie_locks = 0
        for r in processing_rows:
            age_seconds = None
            is_zombie = False
            if r.locked_at:
                age_seconds = int((now - r.locked_at).total_seconds())
                if age_seconds > 600:  # > 10 min
                    is_zombie = True
                    zombie_locks += 1
            processing_data.append({
                'id': r.id,
                'curp': r.curp,
                'user_id': r.user_id,
                'attempts': r.attempts,
                'circuit_open_retries': r.circuit_open_retries,
                'locked_at': r.locked_at.isoformat() if r.locked_at else None,
                'locked_by': r.locked_by,
                'lock_age_seconds': age_seconds,
                'is_zombie': is_zombie,
                'source': r.source,
                'batch_id': r.batch_id,
            })

        # 3) Próximas filas pending (orden por next_retry_at)
        next_pending = (
            CurpVerificationQueue.query
            .filter(CurpVerificationQueue.status == QUEUE_PENDING)
            .order_by(CurpVerificationQueue.next_retry_at.asc())
            .limit(20)
            .all()
        )
        next_pending_data = [{
            'id': r.id,
            'curp': r.curp,
            'attempts': r.attempts,
            'circuit_open_retries': r.circuit_open_retries,
            'next_retry_at': r.next_retry_at.isoformat() if r.next_retry_at else None,
            'last_error': r.last_error,
            'source': r.source,
            'batch_id': r.batch_id,
            'wait_seconds': int((r.next_retry_at - now).total_seconds()) if r.next_retry_at else None,
        } for r in next_pending]

        # 4) Últimas 100 finalizadas (done/rejected/failed) — orden por finished_at
        recent_finished = (
            CurpVerificationQueue.query
            .filter(CurpVerificationQueue.status.in_([QUEUE_DONE, QUEUE_REJECTED, QUEUE_FAILED]))
            .filter(CurpVerificationQueue.finished_at.isnot(None))
            .order_by(desc(CurpVerificationQueue.finished_at))
            .limit(100)
            .all()
        )
        recent_finished_data = []
        # Acumuladores para estadísticas derivadas (procesamiento puro = finished_at - locked_at).
        _durations_proc = []  # seconds processing
        for r in recent_finished:
            # started_at: preferimos locked_at (ícono de "comenzó a procesarse"),
            # fallback a created_at (cuándo se encoló originalmente).
            started_at = r.locked_at or r.created_at
            duration_total = (
                int((r.finished_at - r.created_at).total_seconds())
                if r.finished_at and r.created_at else None
            )
            processing_seconds = (
                int((r.finished_at - r.locked_at).total_seconds())
                if r.finished_at and r.locked_at else None
            )
            if processing_seconds is not None and processing_seconds >= 0:
                _durations_proc.append(processing_seconds)
            recent_finished_data.append({
                'id': r.id,
                'curp': r.curp,
                'user_id': r.user_id,
                'status': r.status,
                'attempts': r.attempts,
                'last_error': r.last_error,
                'source': r.source,
                'batch_id': r.batch_id,
                'created_at': r.created_at.isoformat() if r.created_at else None,
                'started_at': started_at.isoformat() if started_at else None,
                'finished_at': r.finished_at.isoformat() if r.finished_at else None,
                'duration_seconds': duration_total,
                'processing_seconds': processing_seconds,
                'locked_by': r.locked_by,
            })

        # 5) Throughput últimas 24h — done por hora (bucketing en Python para portabilidad)
        since_24h = now - timedelta(hours=24)
        buckets = {}
        for i in range(24):
            ts = now - timedelta(hours=23 - i)
            key = ts.strftime('%Y-%m-%d %H:00')
            buckets[key] = {'done': 0, 'rejected': 0, 'failed': 0}
        last_hour_total = {'done': 0, 'rejected': 0, 'failed': 0}
        finished_24h = db.session.query(
            CurpVerificationQueue.finished_at,
            CurpVerificationQueue.status,
        ).filter(
            CurpVerificationQueue.finished_at >= since_24h,
            CurpVerificationQueue.status.in_([QUEUE_DONE, QUEUE_REJECTED, QUEUE_FAILED]),
        ).all()
        for fin_at, st in finished_24h:
            if not fin_at:
                continue
            key = fin_at.strftime('%Y-%m-%d %H:00')
            if key in buckets and st in buckets[key]:
                buckets[key][st] += 1
        # Conteo de la última hora exacta
        one_hour_ago = now - timedelta(hours=1)
        recent_hour = db.session.query(
            CurpVerificationQueue.status,
            db.func.count(CurpVerificationQueue.id),
        ).filter(
            CurpVerificationQueue.finished_at >= one_hour_ago,
            CurpVerificationQueue.status.in_([QUEUE_DONE, QUEUE_REJECTED, QUEUE_FAILED]),
        ).group_by(CurpVerificationQueue.status).all()
        for st, c in recent_hour:
            if st in last_hour_total:
                last_hour_total[st] = int(c)

        throughput_24h = [{'bucket': k, **v} for k, v in buckets.items()]

        # 6) Top errores recientes (últimas 24h)
        since_err = now - timedelta(hours=24)
        top_errors_rows = db.session.query(
            CurpVerificationQueue.last_error,
            db.func.count(CurpVerificationQueue.id).label('c'),
        ).filter(
            CurpVerificationQueue.last_error.isnot(None),
            CurpVerificationQueue.created_at >= since_err,
        ).group_by(CurpVerificationQueue.last_error).order_by(text('c DESC')).limit(10).all()
        top_errors = [{'error': (e or '')[:200], 'count': int(c)} for e, c in top_errors_rows]

        # 7) Estado RENAPO circuit breaker (acceso vía módulo para leer estado vivo)
        renapo_state = {'circuit_open': False, 'available': True}
        try:
            from app.services import renapo_service as _rs
            import time as _t
            renapo_state['circuit_open'] = bool(_rs.is_renapo_circuit_open())
            renapo_state['available'] = not renapo_state['circuit_open']
            renapo_state['consecutive_failures'] = int(getattr(_rs, '_consecutive_failures', 0))
            renapo_state['threshold'] = int(getattr(_rs, '_CIRCUIT_THRESHOLD', 10))
            renapo_state['cooldown_seconds'] = int(getattr(_rs, '_CIRCUIT_COOLDOWN', 120))
            opened_at = float(getattr(_rs, '_circuit_opened_at', 0) or 0)
            if renapo_state['circuit_open'] and opened_at:
                elapsed = _t.time() - opened_at
                renapo_state['cooldown_remaining'] = max(0, int(renapo_state['cooldown_seconds'] - elapsed))
            else:
                renapo_state['cooldown_remaining'] = 0
        except Exception as _e:
            logger.warning('No se pudo leer circuit breaker RENAPO: %s', _e)

        # 8) Cache RENAPO stats
        cache_stats = {'total': 0, 'positive': 0, 'negative': 0, 'total_hits': 0, 'fresh': 0}
        try:
            cache_stats['total'] = int(CurpRenapoCache.query.count())
            cache_stats['positive'] = int(
                CurpRenapoCache.query.filter(CurpRenapoCache.valid.is_(True)).count()
            )
            cache_stats['negative'] = cache_stats['total'] - cache_stats['positive']
            hits_row = db.session.query(
                db.func.coalesce(db.func.sum(CurpRenapoCache.hits), 0)
            ).scalar()
            cache_stats['total_hits'] = int(hits_row or 0)
            cache_stats['fresh'] = int(
                CurpRenapoCache.query.filter(CurpRenapoCache.expires_at > now).count()
            )
        except Exception as _e:
            logger.warning('No se pudieron leer stats cache RENAPO: %s', _e)

        # 9) Workers vivos: identificadores únicos de locked_by sobre filas activas
        worker_ids = db.session.query(
            CurpVerificationQueue.locked_by
        ).filter(
            CurpVerificationQueue.locked_by.isnot(None),
            CurpVerificationQueue.locked_at >= now - timedelta(minutes=10),
        ).distinct().all()
        active_workers = [w[0] for w in worker_ids if w[0]]

        # 10) Throughput hour-rate (CURPs/hora basado en última hora)
        throughput_rate_per_hour = sum(last_hour_total.values())

        # 11) Stats derivadas: success rate (done / total finalizadas) y latencias.
        def _success_rate(stats):
            done = stats.get('done', 0)
            total_f = done + stats.get('rejected', 0) + stats.get('failed', 0)
            if total_f == 0:
                return None
            return round(done * 100.0 / total_f, 1)

        # 24h aggregate
        stats_24h_agg = {'done': 0, 'rejected': 0, 'failed': 0}
        for _b in throughput_24h:
            stats_24h_agg['done'] += _b['done']
            stats_24h_agg['rejected'] += _b['rejected']
            stats_24h_agg['failed'] += _b['failed']
        stats_24h_total = sum(stats_24h_agg.values())

        # Latencias de procesamiento (segundos entre locked_at y finished_at) sobre
        # las últimas 100 finalizadas. Más representativo del "trabajo real" del
        # worker que (finished_at - created_at), que mete cola.
        avg_proc = None
        median_proc = None
        max_proc = None
        if _durations_proc:
            avg_proc = round(sum(_durations_proc) / len(_durations_proc), 1)
            sorted_d = sorted(_durations_proc)
            median_proc = sorted_d[len(sorted_d) // 2]
            max_proc = max(sorted_d)

        stats_block = {
            'last_hour': {
                **last_hour_total,
                'total': throughput_rate_per_hour,
                'success_rate': _success_rate(last_hour_total),
            },
            'last_24h': {
                **stats_24h_agg,
                'total': stats_24h_total,
                'success_rate': _success_rate(stats_24h_agg),
            },
            'processing_latency_seconds': {
                'avg': avg_proc,
                'median': median_proc,
                'max': max_proc,
                'sample_size': len(_durations_proc),
            },
        }

        return jsonify({
            'server_time': now.isoformat(),
            'total': total,
            'counts': counts,
            'processing': processing_data,
            'processing_count': len(processing_data),
            'zombie_locks': zombie_locks,
            'next_pending': next_pending_data,
            'recent_finished': recent_finished_data,
            'top_errors': top_errors,
            'renapo': renapo_state,
            'cache': cache_stats,
            'active_workers': active_workers,
            'last_hour': last_hour_total,
            'last_hour_total': throughput_rate_per_hour,
            'throughput_24h': throughput_24h,
            'stats': stats_block,
        }), 200
    except Exception:
        logger.exception('curp_queue_dashboard error')
        return jsonify({'error': 'Error interno del servidor'}), 500


@bp.route('/curp-queue/<int:queue_id>/release', methods=['POST'])
@jwt_required()
@admin_required
def curp_queue_release_lock(queue_id):
    """Liberar manualmente un lock zombi de una fila en 'processing'.

    Pone status='pending', locked_at=NULL, locked_by=NULL para que el worker
    la retome. Solo admin/developer.
    """
    try:
        from app.models.curp_verification import (
            CurpVerificationQueue, QUEUE_PENDING, QUEUE_PROCESSING,
        )
        row = CurpVerificationQueue.query.get_or_404(queue_id)
        if row.status != QUEUE_PROCESSING:
            return jsonify({
                'error': f'La fila no está en processing (status actual: {row.status})'
            }), 400
        row.status = QUEUE_PENDING
        row.locked_at = None
        row.locked_by = None
        row.next_retry_at = datetime.utcnow()
        db.session.commit()
        logger.info(
            '[CURP-DASH] admin %s liberó manualmente lock q=%s',
            g.current_user.id, queue_id,
        )
        return jsonify({'success': True, 'queue_id': queue_id, 'new_status': row.status}), 200
    except HTTPException:
        raise
    except Exception:
        logger.exception('curp_queue_release_lock error')
        try:
            db.session.rollback()
        except Exception:
            pass
        return jsonify({'error': 'Error interno del servidor'}), 500

