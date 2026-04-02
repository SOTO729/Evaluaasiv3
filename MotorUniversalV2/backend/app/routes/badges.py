"""
Rutas de Insignias Digitales — Open Badges 3.0

Endpoints:
  CRUD plantillas   → /api/badges/templates
  Emisión           → /api/badges/issue, /api/badges/issue-batch
  Mis insignias     → /api/badges/my-badges
  Público           → /api/badges/<uuid>/credential.json  (JSON-LD)
                    → /api/badges/issuer                   (Issuer Profile)
  Verificación      → /api/badges/verify/<code>
  Analytics         → POST /api/badges/<id>/share, /api/badges/<id>/verify-count
"""
from flask import Blueprint, request, jsonify, abort, g
from werkzeug.exceptions import HTTPException
from flask_jwt_extended import jwt_required, get_jwt_identity
from app import db
from app.models.badge import BadgeTemplate, IssuedBadge
from app.models.user import User

bp = Blueprint('badges', __name__)


# ──────────────────────────────────────────────
# Helper: emisión retroactiva de insignias
# ──────────────────────────────────────────────

def _retroactive_issue_for_template(template):
    """
    Emite insignias retroactivamente para resultados aprobados
    que coinciden con esta plantilla pero aún no tienen insignia.
    Llamar después de crear o activar una plantilla.
    Returns (issued_count, errors_list)
    """
    from app.models.result import Result
    from app.models.exam import Exam
    from app.services.badge_service import issue_badge_for_result

    if not template or not template.is_active:
        return 0, []

    # Encontrar exámenes que coinciden con esta plantilla
    exam_ids = []
    if template.exam_id:
        exam_ids.append(template.exam_id)
    if template.competency_standard_id:
        matching_exams = Exam.query.filter_by(
            competency_standard_id=template.competency_standard_id
        ).with_entities(Exam.id).all()
        exam_ids.extend(eid for (eid,) in matching_exams)
    exam_ids = list(set(exam_ids))

    if not exam_ids:
        return 0, []

    # Resultados aprobados y completados para esos exámenes
    approved = Result.query.filter(
        Result.exam_id.in_(exam_ids),
        Result.result == 1,
        Result.status == 1
    ).all()

    issued_count = 0
    errors_list = []

    for result in approved:
        # ¿Ya tiene insignia para este resultado?
        existing = IssuedBadge.query.filter_by(result_id=result.id).first()
        if existing:
            continue
        # ¿Ya tiene insignia activa con esta plantilla?
        existing2 = IssuedBadge.query.filter_by(
            user_id=result.user_id, badge_template_id=template.id, status='active'
        ).first()
        if existing2:
            continue

        user = User.query.get(str(result.user_id))
        if not user:
            continue
        exam = Exam.query.get(result.exam_id)
        if not exam:
            continue

        try:
            badge = issue_badge_for_result(result, user, exam, force=True)
            if badge:
                issued_count += 1
        except HTTPException:
            raise
        except Exception as e:
            errors_list.append(f"{getattr(user, 'email', None) or user.username}: {str(e)}")

    if issued_count > 0:
        print(f"[BADGE] Retroactive issue: {issued_count} badge(s) for template '{template.name}' (ID {template.id})")

    return issued_count, errors_list


def _lazy_issue_for_user(user_id):
    """
    Intenta emitir insignias pendientes para un usuario.
    Se usa en my_badges para generación lazy cuando el template
    no existía al momento de aprobar el examen.
    Returns count of newly issued badges.
    """
    from app.models.result import Result
    from app.models.exam import Exam
    from app.services.badge_service import issue_badge_for_result

    # Resultados aprobados del usuario sin insignia
    approved = Result.query.filter_by(
        user_id=str(user_id), result=1, status=1
    ).all()

    issued_count = 0
    for result in approved:
        existing = IssuedBadge.query.filter_by(result_id=result.id).first()
        if existing:
            continue

        exam = Exam.query.get(result.exam_id) if result.exam_id else None
        if not exam:
            continue

        # Buscar plantilla activa
        template = BadgeTemplate.query.filter_by(exam_id=exam.id, is_active=True).first()
        if not template and exam.competency_standard_id:
            template = BadgeTemplate.query.filter_by(
                competency_standard_id=exam.competency_standard_id, is_active=True
            ).first()
        if not template:
            continue

        # ¿Ya tiene insignia activa con esta plantilla?
        existing2 = IssuedBadge.query.filter_by(
            user_id=str(user_id), badge_template_id=template.id, status='active'
        ).first()
        if existing2:
            continue

        try:
            user = User.query.get(str(user_id))
            badge = issue_badge_for_result(result, user, exam, force=True)
            if badge:
                issued_count += 1
        except Exception as e:
            print(f"[BADGE] Lazy issue error for user {user_id}: {e}")

    if issued_count > 0:
        print(f"[BADGE] Lazy issue: {issued_count} badge(s) for user {user_id}")

    return issued_count


# ──────────────────────────────────────────────
# Helper: verificar roles permitidos
# ──────────────────────────────────────────────
def _require_roles(*allowed):
    uid = get_jwt_identity()
    user = User.query.get(str(uid))
    # developer siempre tiene acceso (equivalente a admin)
    effective_allowed = set(allowed) | {'developer'}
    if not user or user.role not in effective_allowed:
        abort(403, description='No tienes permisos para esta acción')
    return user


def _verify_badge_group_access(group_id, user):
    """Verifica que el coordinador tenga acceso al grupo para operaciones de badges.
    Admin/editor/developer ven todo. Coordinador solo sus propios grupos (aislados por coordinator_id)."""
    if user.role in ('admin', 'editor', 'developer'):
        return None  # sin restricción
    if user.role in ('coordinator', 'auxiliar'):
        from app.models.partner import CandidateGroup
        group = CandidateGroup.query.get(group_id)
        if not group:
            return jsonify({'error': 'Grupo no encontrado'}), 404
        # Obtener coordinator_id efectivo (auxiliar hereda de su coordinador)
        coord_id = user.id if user.role == 'coordinator' else user.coordinator_id
        if group.coordinator_id == coord_id:
            return None  # tiene acceso
        return jsonify({'error': 'No tienes acceso a este grupo'}), 403
    return jsonify({'error': 'No tienes permisos'}), 403


# ═══════════════════════════════════════════════
# PLANTILLAS (CRUD)
# ═══════════════════════════════════════════════

@bp.route('/templates', methods=['GET'])
@jwt_required()
def list_templates():
    """Listar plantillas de insignias (admin/editor/coordinator)"""
    _require_roles('admin', 'editor', 'coordinator')

    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 20, type=int)
    search = request.args.get('search', '').strip()
    active_only = request.args.get('active_only', 'false').lower() == 'true'

    query = BadgeTemplate.query
    if active_only:
        query = query.filter_by(is_active=True)
    if search:
        query = query.filter(BadgeTemplate.name.ilike(f'%{search}%'))

    query = query.order_by(BadgeTemplate.created_at.desc())
    pagination = query.paginate(page=page, per_page=per_page, error_out=False)

    return jsonify({
        'templates': [t.to_dict() for t in pagination.items],
        'total': pagination.total,
        'page': pagination.page,
        'pages': pagination.pages,
    })


@bp.route('/templates', methods=['POST'])
@jwt_required()
def create_template():
    """Crear nueva plantilla de insignia"""
    user = _require_roles('admin', 'editor', 'coordinator')
    data = request.get_json(silent=True) or {}

    name = data.get('name', '').strip()
    if not name:
        return jsonify({'error': 'El nombre es requerido'}), 400

    ecm_id = data.get('competency_standard_id')
    if not ecm_id:
        return jsonify({'error': 'competency_standard_id es requerido'}), 400

    # Regla: solo puede haber 1 plantilla activa por ECM
    wants_active = data.get('is_active', True)
    if wants_active:
        existing_active = BadgeTemplate.query.filter_by(
            competency_standard_id=ecm_id, is_active=True
        ).first()
        if existing_active:
            return jsonify({
                'error': f'Ya existe una plantilla activa para este ECM (ID {existing_active.id}: {existing_active.name}). Desactívela primero.'
            }), 409

    # Normalize skills: accept string or list
    raw_skills = data.get('skills', '')
    if isinstance(raw_skills, list):
        raw_skills = ', '.join(str(s).strip() for s in raw_skills if str(s).strip())
    skills_str = (raw_skills or '').strip() or None

    # tags se sincroniza desde skills
    tags_str = skills_str

    # Emisor siempre es Grupo Eduit (regla de negocio)
    template = BadgeTemplate(
        name=name,
        description=data.get('description', '').strip() or None,
        criteria_narrative=data.get('criteria_narrative', '').strip() or None,
        criteria_url=data.get('criteria_url', '').strip() or None,
        exam_id=data.get('exam_id'),
        competency_standard_id=ecm_id,
        issuer_name='Grupo Eduit',
        issuer_url='https://www.grupoeduit.com',
        issuer_image_url=None,
        tags=tags_str,
        skills=skills_str,
        expiry_months=data.get('expiry_months'),
        is_active=wants_active,
        created_by_id=user.id,
    )
    db.session.add(template)
    db.session.commit()

    # Emisión retroactiva: si la plantilla está activa, emitir para resultados aprobados existentes
    retro_issued = 0
    if template.is_active:
        retro_issued, _errors = _retroactive_issue_for_template(template)

    resp = {'message': 'Plantilla creada', 'template': template.to_dict()}
    if retro_issued > 0:
        resp['retroactive_issued'] = retro_issued
        resp['message'] += f' — {retro_issued} insignia(s) emitida(s) retroactivamente'

    return jsonify(resp), 201


@bp.route('/templates/<int:template_id>', methods=['GET'])
@jwt_required()
def get_template(template_id):
    """Obtener una plantilla por ID"""
    _require_roles('admin', 'editor', 'coordinator')
    template = BadgeTemplate.query.get_or_404(template_id)
    return jsonify({'template': template.to_dict()})


@bp.route('/templates/<int:template_id>', methods=['PUT'])
@jwt_required()
def update_template(template_id):
    """Actualizar plantilla"""
    _require_roles('admin', 'editor', 'coordinator')
    template = BadgeTemplate.query.get_or_404(template_id)
    data = request.get_json(silent=True) or {}

    for field in ('name', 'description', 'criteria_narrative', 'criteria_url'):
        if field in data:
            val = data[field]
            if isinstance(val, str):
                setattr(template, field, val.strip() or None)
            else:
                setattr(template, field, val or None)

    # Emisor siempre es Grupo Eduit (regla de negocio) — ignorar valores del cliente
    template.issuer_name = 'Grupo Eduit'
    template.issuer_url = 'https://www.grupoeduit.com'

    # skills: accept string or list — tags se sincroniza desde skills
    if 'skills' in data:
        raw = data['skills']
        if isinstance(raw, list):
            raw = ', '.join(str(s).strip() for s in raw if str(s).strip())
        template.skills = (raw or '').strip() or None
        template.tags = template.skills

    for int_field in ('exam_id', 'competency_standard_id', 'expiry_months'):
        if int_field in data:
            setattr(template, int_field, data[int_field])

    # Regla: solo puede haber 1 plantilla activa por ECM
    was_active_before = template.is_active
    if 'is_active' in data:
        wants_active = bool(data['is_active'])
        ecm_id = template.competency_standard_id
        if wants_active and ecm_id:
            existing_active = BadgeTemplate.query.filter(
                BadgeTemplate.competency_standard_id == ecm_id,
                BadgeTemplate.is_active == True,
                BadgeTemplate.id != template.id
            ).first()
            if existing_active:
                return jsonify({
                    'error': f'Ya existe una plantilla activa para este ECM (ID {existing_active.id}: {existing_active.name}). Desactívela primero.'
                }), 409
        template.is_active = wants_active

    db.session.commit()

    # Emisión retroactiva al activar una plantilla
    retro_issued = 0
    if template.is_active and not was_active_before:
        retro_issued, _errors = _retroactive_issue_for_template(template)

    resp = {'message': 'Plantilla actualizada', 'template': template.to_dict()}
    if retro_issued > 0:
        resp['retroactive_issued'] = retro_issued
        resp['message'] += f' — {retro_issued} insignia(s) emitida(s) retroactivamente'

    return jsonify(resp)


@bp.route('/templates/<int:template_id>', methods=['DELETE'])
@jwt_required()
def delete_template(template_id):
    """Eliminar plantilla (soft: desactivar)"""
    _require_roles('admin', 'editor')
    template = BadgeTemplate.query.get_or_404(template_id)
    template.is_active = False
    db.session.commit()
    return jsonify({'message': 'Plantilla desactivada'})


@bp.route('/templates/<int:template_id>/image', methods=['POST'])
@jwt_required()
def upload_template_image(template_id):
    """Subir imagen personalizada para la plantilla (convertida a WebP).
    Sobreescribe el blob anterior si existe."""
    _require_roles('admin', 'editor', 'coordinator')
    template = BadgeTemplate.query.get_or_404(template_id)

    if 'image' not in request.files:
        return jsonify({'error': 'Se requiere un archivo de imagen'}), 400

    file = request.files['image']
    if not file.filename:
        return jsonify({'error': 'Archivo vacío'}), 400

    from PIL import Image as PILImage
    import io
    from app.utils.azure_storage import AzureStorageService
    storage = AzureStorageService()

    # Convertir a WebP
    try:
        img = PILImage.open(file)
        if img.mode in ('RGBA', 'LA', 'P'):
            bg = PILImage.new('RGBA', img.size, (255, 255, 255, 0))
            if img.mode == 'P':
                img = img.convert('RGBA')
            bg.paste(img, mask=img.split()[-1] if img.mode == 'RGBA' else None)
            img = bg
        webp_buf = io.BytesIO()
        img.save(webp_buf, format='WEBP', quality=90, lossless=False)
        webp_bytes = webp_buf.getvalue()
    except HTTPException:
        raise
    except Exception as e:
        return jsonify({'error': f'Error al procesar la imagen: {e}'}), 400

    # Blob name estable por template: siempre sobreescribe el mismo archivo
    blob_name = template.badge_image_blob_name or f"badge-templates/template-{template_id}.webp"
    # Garantizar extensión .webp
    if not blob_name.endswith('.webp'):
        blob_name = blob_name.rsplit('.', 1)[0] + '.webp'

    # Si tenía otro blob distinto, eliminar el anterior
    if template.badge_image_url and template.badge_image_blob_name and template.badge_image_blob_name != blob_name:
        try:
            storage.delete_file(template.badge_image_url)
        except Exception:
            pass

    url = storage.upload_bytes(webp_bytes, blob_name, content_type='image/webp')
    if not url:
        return jsonify({'error': 'Error al subir la imagen'}), 500

    template.badge_image_url = url
    template.badge_image_blob_name = blob_name
    db.session.commit()

    return jsonify({'message': 'Imagen subida (WebP)', 'image_url': url})


@bp.route('/templates/<int:template_id>/image', methods=['DELETE'])
@jwt_required()
def delete_template_image(template_id):
    """Eliminar imagen de la plantilla (borra blob y limpia campos)"""
    _require_roles('admin', 'editor', 'coordinator')
    template = BadgeTemplate.query.get_or_404(template_id)

    if not template.badge_image_url:
        return jsonify({'message': 'La plantilla no tiene imagen'}), 200

    from app.utils.azure_storage import AzureStorageService
    storage = AzureStorageService()
    try:
        storage.delete_file(template.badge_image_url)
    except Exception:
        pass  # Si falla el borrado del blob, igualmente limpiamos la referencia

    template.badge_image_url = None
    template.badge_image_blob_name = None
    db.session.commit()

    return jsonify({'message': 'Imagen eliminada'})


@bp.route('/templates/<int:template_id>/issuer-logo', methods=['POST'])
@jwt_required()
def upload_issuer_logo(template_id):
    """Subir logo del emisor para la plantilla (convertido a WebP)."""
    _require_roles('admin', 'editor', 'coordinator')
    template = BadgeTemplate.query.get_or_404(template_id)

    if 'image' not in request.files:
        return jsonify({'error': 'Se requiere un archivo de imagen'}), 400

    file = request.files['image']
    if not file.filename:
        return jsonify({'error': 'Archivo vacío'}), 400

    from PIL import Image as PILImage
    import io
    from app.utils.azure_storage import AzureStorageService
    storage = AzureStorageService()

    try:
        img = PILImage.open(file)
        if img.mode in ('RGBA', 'LA', 'P'):
            bg = PILImage.new('RGBA', img.size, (255, 255, 255, 0))
            if img.mode == 'P':
                img = img.convert('RGBA')
            bg.paste(img, mask=img.split()[-1] if img.mode == 'RGBA' else None)
            img = bg
        webp_buf = io.BytesIO()
        img.save(webp_buf, format='WEBP', quality=90, lossless=False)
        webp_bytes = webp_buf.getvalue()
    except HTTPException:
        raise
    except Exception as e:
        return jsonify({'error': f'Error al procesar la imagen: {e}'}), 400

    blob_name = template.issuer_logo_blob_name or f"badge-templates/issuer-logo-{template_id}.webp"
    if not blob_name.endswith('.webp'):
        blob_name = blob_name.rsplit('.', 1)[0] + '.webp'

    if template.issuer_logo_url and template.issuer_logo_blob_name and template.issuer_logo_blob_name != blob_name:
        try:
            storage.delete_file(template.issuer_logo_url)
        except Exception:
            pass

    url = storage.upload_bytes(webp_bytes, blob_name, content_type='image/webp')
    if not url:
        return jsonify({'error': 'Error al subir la imagen'}), 500

    template.issuer_logo_url = url
    template.issuer_logo_blob_name = blob_name
    db.session.commit()

    return jsonify({'message': 'Logo del emisor subido (WebP)', 'issuer_logo_url': url})


@bp.route('/templates/<int:template_id>/issuer-logo', methods=['DELETE'])
@jwt_required()
def delete_issuer_logo(template_id):
    """Eliminar logo del emisor de la plantilla."""
    _require_roles('admin', 'editor', 'coordinator')
    template = BadgeTemplate.query.get_or_404(template_id)

    if not template.issuer_logo_url:
        return jsonify({'message': 'La plantilla no tiene logo de emisor'}), 200

    from app.utils.azure_storage import AzureStorageService
    storage = AzureStorageService()
    try:
        storage.delete_file(template.issuer_logo_url)
    except Exception:
        pass

    template.issuer_logo_url = None
    template.issuer_logo_blob_name = None
    db.session.commit()

    return jsonify({'message': 'Logo del emisor eliminado'})


# ═══════════════════════════════════════════════
# EMISIÓN
# ═══════════════════════════════════════════════

@bp.route('/issue', methods=['POST'])
@jwt_required()
def issue_badge_manual():
    """Emitir insignia manualmente para un resultado"""
    _require_roles('admin', 'editor', 'coordinator')
    data = request.get_json(silent=True) or {}
    result_id = data.get('result_id')
    if not result_id:
        return jsonify({'error': 'result_id requerido'}), 400

    from app.models.result import Result
    from app.models.exam import Exam
    result = Result.query.get(str(result_id))
    if not result:
        return jsonify({'error': 'Resultado no encontrado'}), 404

    user = User.query.get(str(result.user_id))
    exam = Exam.query.get(result.exam_id)

    from app.services.badge_service import issue_badge_for_result
    badge = issue_badge_for_result(result, user, exam, force=True)
    if not badge:
        return jsonify({'error': 'No se pudo emitir la insignia (revise plantilla/resultado)'}), 400

    return jsonify({'message': 'Insignia emitida', 'badge': badge.to_dict(include_credential=True)}), 201


@bp.route('/issue-batch', methods=['POST'])
@jwt_required()
def issue_badge_batch():
    """Emitir insignias en lote"""
    _require_roles('admin', 'editor', 'coordinator')
    data = request.get_json(silent=True) or {}
    result_ids = data.get('result_ids', [])
    if not result_ids:
        return jsonify({'error': 'result_ids requerido'}), 400

    from app.services.badge_service import issue_badges_batch
    badges = issue_badges_batch(result_ids)

    return jsonify({
        'message': f'{len(badges)} insignias emitidas',
        'badges': [b.to_dict() for b in badges],
        'total_issued': len(badges),
    })


# ═══════════════════════════════════════════════
# MIS INSIGNIAS (candidato)
# ═══════════════════════════════════════════════

@bp.route('/my-badges', methods=['GET'])
@jwt_required()
def my_badges():
    """Obtener las insignias del usuario autenticado (con emisión lazy de pendientes).
    También devuelve exámenes aprobados que aún no tienen insignia (pending_exams)."""
    uid = get_jwt_identity()

    # Emisión lazy: si hay resultados aprobados sin insignia y ahora existe plantilla, emitir
    try:
        _lazy_issue_for_user(uid)
    except Exception as e:
        print(f"[BADGE] Lazy issue error in my_badges: {e}")

    badges = IssuedBadge.query.filter_by(
        user_id=str(uid), status='active'
    ).order_by(IssuedBadge.issued_at.desc()).all()

    # Identificar exámenes aprobados que aún no tienen insignia
    pending_exams = []
    try:
        from app.models.result import Result
        from app.models.exam import Exam

        # IDs de resultados que ya tienen insignia
        badged_result_ids = {b.result_id for b in badges if b.result_id}

        approved_results = Result.query.filter_by(
            user_id=str(uid), result=1, status=1
        ).all()

        seen_exam_ids = set()
        for result in approved_results:
            if result.id in badged_result_ids:
                continue
            if result.exam_id in seen_exam_ids:
                continue
            seen_exam_ids.add(result.exam_id)

            exam = Exam.query.get(result.exam_id) if result.exam_id else None
            if not exam:
                continue

            # Verificar si existe plantilla activa (para mostrar razón correcta)
            has_template = False
            tmpl = BadgeTemplate.query.filter_by(exam_id=exam.id, is_active=True).first()
            if not tmpl and exam.competency_standard_id:
                tmpl = BadgeTemplate.query.filter_by(
                    competency_standard_id=exam.competency_standard_id, is_active=True
                ).first()
            if tmpl:
                has_template = True

            pending_exams.append({
                'exam_id': exam.id,
                'exam_name': exam.name,
                'has_template': has_template,
                'approved_at': result.created_at.isoformat() if result.created_at else None,
            })
    except Exception as e:
        print(f"[BADGE] Error building pending_exams: {e}")

    return jsonify({
        'badges': [b.to_dict(include_credential=False) for b in badges],
        'total': len(badges),
        'pending_exams': pending_exams,
    })


@bp.route('/user/<user_id>', methods=['GET'])
@jwt_required()
def user_badges(user_id):
    """Obtener insignias de un usuario (admin/editor/coordinator)"""
    _require_roles('admin', 'editor', 'coordinator')
    badges = IssuedBadge.query.filter_by(
        user_id=str(user_id), status='active'
    ).order_by(IssuedBadge.issued_at.desc()).all()

    return jsonify({
        'badges': [b.to_dict() for b in badges],
        'total': len(badges),
    })


# ═══════════════════════════════════════════════
# ENDPOINTS PÚBLICOS (Open Badges 3.0)
# ═══════════════════════════════════════════════

@bp.route('/<badge_uuid>/credential.json', methods=['GET'])
def public_credential(badge_uuid):
    """
    Endpoint público de credencial OB3 JSON-LD.
    Hosted verification: los verificadores obtienen el JSON de aquí.
    """
    import json as _json
    badge = IssuedBadge.query.filter_by(badge_uuid=badge_uuid).first()
    if not badge:
        return jsonify({'error': 'Badge not found'}), 404

    if badge.status == 'revoked':
        return jsonify({'error': 'This credential has been revoked', 'reason': badge.revocation_reason}), 410

    # Incrementar contador de verificación
    badge.verify_count = (badge.verify_count or 0) + 1
    db.session.commit()

    credential = _json.loads(badge.credential_json) if badge.credential_json else {}
    response = jsonify(credential)
    response.headers['Content-Type'] = 'application/ld+json'
    return response


@bp.route('/issuer', methods=['GET'])
def issuer_profile():
    """Perfil del emisor (Issuer) conforme a OB3"""
    import os, base64

    key_id = os.getenv('ED25519_KEY_ID', 'evaluaasi-ed25519-2026')
    pub_pem = os.getenv('ED25519_PUBLIC_KEY_PEM', '')

    profile = {
        'id': request.url_root.rstrip('/') + '/api/badges/issuer',
        'type': ['Profile'],
        'name': 'Grupo Eduit',
        'url': 'https://www.grupoeduit.com',
        'description': 'Plataforma de evaluación y certificación de competencias laborales.',
        'email': 'soporte@evaluaasi.com',
    }

    # Expose Ed25519 public key for remote verification
    if pub_pem:
        try:
            from cryptography.hazmat.primitives.serialization import load_pem_public_key
            from cryptography.hazmat.primitives.serialization import Encoding, PublicFormat
            pub_pem_clean = pub_pem.replace('\\n', '\n')
            pub_key = load_pem_public_key(pub_pem_clean.encode())
            pub_raw = pub_key.public_bytes(Encoding.Raw, PublicFormat.Raw)
            pub_b64url = base64.urlsafe_b64encode(pub_raw).rstrip(b'=').decode()
            profile['verificationMethod'] = [{
                'id': f"{profile['id']}#key-{key_id}",
                'type': 'Ed25519VerificationKey2020',
                'controller': profile['id'],
                'publicKeyJwk': {
                    'kty': 'OKP',
                    'crv': 'Ed25519',
                    'x': pub_b64url,
                },
            }]
        except Exception as e:
            print(f'[BADGE] Error exposing public key: {e}', flush=True)
    else:
        print(f'[BADGE] No ED25519_PUBLIC_KEY_PEM env var found (length={len(pub_pem)})', flush=True)

    return jsonify(profile)


@bp.route('/verify/<code>', methods=['GET'])
def verify_badge(code):
    """Verificar insignia por código BD..."""
    badge = IssuedBadge.query.filter_by(badge_code=code).first()
    if not badge:
        return jsonify({'valid': False, 'error': 'Código no encontrado'}), 404

    badge.verify_count = (badge.verify_count or 0) + 1
    db.session.commit()

    template = BadgeTemplate.query.get(badge.badge_template_id)
    user = User.query.get(str(badge.user_id))

    # Nombre del candidato
    name_parts = [user.name or ''] if user else ['N/A']
    if user and user.first_surname:
        name_parts.append(user.first_surname)
    if user and user.second_surname:
        name_parts.append(user.second_surname)
    full_name = ' '.join(name_parts).strip() or 'N/A'

    # Fecha formateada
    meses = {
        1: 'enero', 2: 'febrero', 3: 'marzo', 4: 'abril',
        5: 'mayo', 6: 'junio', 7: 'julio', 8: 'agosto',
        9: 'septiembre', 10: 'octubre', 11: 'noviembre', 12: 'diciembre'
    }
    issued_date = badge.issued_at
    formatted_date = None
    if issued_date:
        formatted_date = f"{issued_date.day} de {meses[issued_date.month]} de {issued_date.year}"

    expires_date = None
    if badge.expires_at:
        expires_date = f"{badge.expires_at.day} de {meses[badge.expires_at.month]} de {badge.expires_at.year}"

    from datetime import datetime
    is_expired = badge.expires_at and badge.expires_at < datetime.utcnow()

    resp = {
        'valid': True,
        'badge': {
            'name': template.name if template else 'N/A',
            'description': template.description if template else None,
            'template_image_url': (badge.template_image_url or (template.badge_image_url if template else None)),
            'candidate_name': full_name,
            'issuer_name': 'Grupo Eduit',
            'issuer_logo_url': template.issuer_logo_url if template else None,
            'issued_date': formatted_date,
            'expires_date': expires_date,
            'is_expired': bool(is_expired),
            'status': badge.status,
            'badge_code': badge.badge_code,
            'verify_url': badge.verify_url,
            'credential_url': badge.credential_url,
            'share_count': badge.share_count or 0,
            'verify_count': badge.verify_count or 0,
            'skills': template.skills if template else None,
        }
    }

    # Check if credential has cryptographic proof
    if badge.credential_json:
        try:
            import json as _json
            cred = _json.loads(badge.credential_json)
            if 'proof' in cred and cred['proof'].get('type') == 'Ed25519Signature2020':
                resp['badge']['cryptographically_signed'] = True
                resp['badge']['proof_type'] = 'Ed25519Signature2020'
        except Exception:
            pass

    return jsonify(resp)


# ═══════════════════════════════════════════════
# ANALYTICS & SHARE
# ═══════════════════════════════════════════════

@bp.route('/<int:badge_id>/share', methods=['POST'])
@jwt_required()
def track_share(badge_id):
    """Incrementar contador de compartido"""
    badge = IssuedBadge.query.get_or_404(badge_id)
    badge.share_count = (badge.share_count or 0) + 1
    db.session.commit()
    return jsonify({'share_count': badge.share_count})


@bp.route('/<int:badge_id>/claim', methods=['POST'])
@jwt_required()
def claim_badge(badge_id):
    """Marcar insignia como reclamada (claimed)"""
    uid = get_jwt_identity()
    badge = IssuedBadge.query.get_or_404(badge_id)
    if str(badge.user_id) != str(uid):
        abort(403)
    from datetime import datetime
    badge.claimed_at = badge.claimed_at or datetime.utcnow()
    db.session.commit()
    return jsonify({'message': 'Insignia reclamada', 'claimed_at': str(badge.claimed_at)})


# ═══════════════════════════════════════════════
# REVOCACIÓN
# ═══════════════════════════════════════════════

@bp.route('/<int:badge_id>/revoke', methods=['POST'])
@jwt_required()
def revoke_badge(badge_id):
    """Revocar una insignia emitida"""
    _require_roles('admin', 'editor')
    badge = IssuedBadge.query.get_or_404(badge_id)
    data = request.get_json(silent=True) or {}

    badge.status = 'revoked'
    badge.revocation_reason = data.get('reason', 'Revocada por administrador')
    db.session.commit()

    return jsonify({'message': 'Insignia revocada', 'badge': badge.to_dict()})


# ═══════════════════════════════════════════════
# SHARE PREVIEW (OG meta tags para LinkedIn/redes)
# ═══════════════════════════════════════════════

# Blueprint para ruta corta /s/<code> (URL profesional para compartir)
short_share_bp = Blueprint('short_share', __name__)

@short_share_bp.route('/s/<code>', methods=['GET'])
def short_share_redirect(code):
    """Ruta corta para compartir insignias. Redirige a share-preview."""
    return badge_share_preview(code)

@bp.route('/share-preview/<code>', methods=['GET'])
def badge_share_preview(code):
    """
    Página HTML pública con Open Graph meta tags para que LinkedIn
    y otras redes sociales muestren una preview rica de la insignia.
    Redirige al usuario al verify page de la SPA después de cargar.
    """
    import os
    from markupsafe import escape

    badge = IssuedBadge.query.filter_by(badge_code=code).first()
    if not badge:
        return '<html><body>Insignia no encontrada</body></html>', 404

    template = BadgeTemplate.query.get(badge.badge_template_id)
    user = User.query.get(str(badge.user_id))

    # Build candidate name
    name_parts = [user.name or ''] if user else ['N/A']
    if user and user.first_surname:
        name_parts.append(user.first_surname)
    if user and user.second_surname:
        name_parts.append(user.second_surname)
    full_name = ' '.join(name_parts).strip() or 'N/A'

    badge_name = template.name if template else 'Insignia Digital'
    description_parts = [f'Insignia digital verificada otorgada a {full_name}']

    # Add skills to description
    skills_text = ''
    if template and template.skills:
        skills_list = [s.strip() for s in template.skills.split(',') if s.strip()]
        if skills_list:
            skills_text = ', '.join(skills_list)
            description_parts.append(f'Aptitudes: {skills_text}')

    # Add expiry
    if badge.expires_at:
        meses = {1:'ene',2:'feb',3:'mar',4:'abr',5:'may',6:'jun',
                 7:'jul',8:'ago',9:'sep',10:'oct',11:'nov',12:'dic'}
        description_parts.append(
            f'Válida hasta: {badge.expires_at.day}/{meses[badge.expires_at.month]}/{badge.expires_at.year}'
        )
    else:
        description_parts.append('Sin fecha de caducidad')

    description = '. '.join(description_parts) + '.'

    # Image: serve PNG via our conversion endpoint (LinkedIn requires PNG/JPEG, not WebP)
    api_base = os.environ.get('API_BASE_URL', request.host_url.rstrip('/'))
    # Ensure HTTPS (container app is behind TLS-terminating reverse proxy)
    if api_base.startswith('http://'):
        api_base = 'https://' + api_base[7:]
    image_url = f"{api_base}/api/badges/share-image/{code}.png"

    # Determine SPA verify URL
    swa_base = os.environ.get('SWA_BASE_URL', 'https://app.evaluaasi.com')
    verify_url = f"{swa_base}/verify/{escape(code)}"

    html = f'''<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8">
<title>{escape(badge_name)} — Verificación de Insignia | Evaluaasi</title>
<meta property="og:type" content="website" />
<meta property="og:title" content="{escape(badge_name)}" />
<meta property="og:description" content="{escape(description)}" />
<meta property="og:image" content="{escape(image_url)}" />
<meta property="og:image:type" content="image/png" />
<meta property="og:image:width" content="1200" />
<meta property="og:image:height" content="627" />
<meta property="og:url" content="{escape(verify_url)}" />
<meta property="og:site_name" content="Evaluaasi - Grupo Eduit" />
<meta name="twitter:card" content="summary_large_image" />
<meta name="twitter:title" content="{escape(badge_name)}" />
<meta name="twitter:description" content="{escape(description)}" />
<meta name="twitter:image" content="{escape(image_url)}" />
<meta http-equiv="refresh" content="2;url={escape(verify_url)}" />
</head>
<body>
<p>Redirigiendo a la verificación de <strong>{escape(badge_name)}</strong>...</p>
<p><a href="{escape(verify_url)}">Haz clic aquí si no eres redirigido</a></p>
</body>
</html>'''

    from flask import make_response
    resp = make_response(html, 200)
    resp.headers['Content-Type'] = 'text/html; charset=utf-8'
    return resp


@bp.route('/share-image/<code>.png', methods=['GET'])
def badge_share_image(code):
    """
    Sirve la imagen de la insignia convertida a PNG para LinkedIn/redes.
    LinkedIn no soporta WebP en og:image, así que convertimos al vuelo.
    Cacheable por CDN/browser.
    """
    import requests as http_requests
    from io import BytesIO

    badge = IssuedBadge.query.filter_by(badge_code=code).first()
    if not badge:
        abort(404)

    # Prefer template image (higher quality), fallback to baked badge, then ECM logo
    template = BadgeTemplate.query.get(badge.badge_template_id) if badge.badge_template_id else None
    ecm_logo = None
    if template and template.competency_standard_id:
        from app.models.competency_standard import CompetencyStandard
        cs = CompetencyStandard.query.get(template.competency_standard_id)
        ecm_logo = cs.logo_url if cs else None

    image_url = (badge.template_image_url
                 or (template.badge_image_url if template else None)
                 or ecm_logo
                 or badge.badge_image_url)
    if not image_url:
        abort(404)

    try:
        img_resp = http_requests.get(image_url, timeout=10)
        img_resp.raise_for_status()

        from PIL import Image
        img = Image.open(BytesIO(img_resp.content))

        # Convert to RGB if needed (RGBA/palette → RGB for broader compat)
        if img.mode in ('RGBA', 'LA'):
            background = Image.new('RGB', img.size, (255, 255, 255))
            background.paste(img, mask=img.split()[-1])
            img = background
        elif img.mode != 'RGB':
            img = img.convert('RGB')

        # Ensure minimum size for LinkedIn (1200x627 recommended)
        # Scale up if too small, maintaining aspect ratio, on white bg
        min_w, min_h = 1200, 627
        if img.width < min_w or img.height < min_h:
            scale = max(min_w / img.width, min_h / img.height)
            new_w = int(img.width * scale)
            new_h = int(img.height * scale)
            img = img.resize((new_w, new_h), Image.LANCZOS)

        buf = BytesIO()
        img.save(buf, format='PNG', optimize=True)
        buf.seek(0)

        from flask import make_response as mk
        resp = mk(buf.getvalue())
        resp.headers['Content-Type'] = 'image/png'
        resp.headers['Cache-Control'] = 'public, max-age=86400'
        return resp
    except HTTPException:
        raise
    except Exception as e:
        print(f"[BADGE] Error converting image for {code}: {e}")
        abort(404)


@bp.route('/<int:badge_id>/linkedin-url', methods=['GET'])
@jwt_required()
def linkedin_share_url(badge_id):
    """Generar URLs para compartir insignia en LinkedIn"""
    import os
    badge = IssuedBadge.query.get_or_404(badge_id)
    template = BadgeTemplate.query.get(badge.badge_template_id)

    verify_url = badge.verify_url
    issued_year = badge.issued_at.year if badge.issued_at else ''
    issued_month = badge.issued_at.month if badge.issued_at else ''

    from urllib.parse import quote

    # 1. Add to Profile URL (adds certification to LinkedIn profile)
    params = {
        'name': quote(template.name if template else 'Insignia Digital'),
        'organizationName': quote(template.issuer_name or 'Grupo Eduit'),
        'issueYear': str(issued_year),
        'issueMonth': str(issued_month),
        'certUrl': quote(verify_url),
        'certId': badge.badge_code,
    }
    if badge.expires_at:
        params['expirationYear'] = str(badge.expires_at.year)
        params['expirationMonth'] = str(badge.expires_at.month)

    add_profile_url = 'https://www.linkedin.com/profile/add?startTask=CERTIFICATION_NAME'
    for k, v in params.items():
        add_profile_url += f'&{k}={v}'

    # 2. Share as Post URL
    # Uses sharing/share-offsite which crawls OG tags from the share-preview page
    # The share-preview page has og:image pointing to a PNG conversion endpoint,
    # og:description with skills, and og:title with badge name.
    api_base = os.environ.get('API_BASE_URL', request.host_url.rstrip('/'))
    if api_base.startswith('http://'):
        api_base = 'https://' + api_base[7:]
    share_preview_url = f"{api_base}/api/badges/share-preview/{badge.badge_code}"
    image_png_url = f"{api_base}/api/badges/share-image/{badge.badge_code}.png"
    share_post_url = f"https://www.linkedin.com/sharing/share-offsite/?url={quote(share_preview_url)}"

    # Check LinkedIn API availability
    from app.services.linkedin_service import is_configured
    user_id = get_jwt_identity()
    user = User.query.get(str(user_id))
    linkedin_api_configured = is_configured()
    linkedin_connected = bool(getattr(user, 'linkedin_token', None)) if user else False

    return jsonify({
        'linkedin_url': share_post_url,
        'add_profile_url': add_profile_url,
        'share_post_url': share_post_url,
        'image_url': image_png_url,
        'linkedin_api_available': linkedin_api_configured,
        'linkedin_connected': linkedin_connected,
    })


# ═══════════════════════════════════════════════
# GRUPO: insignias de un grupo
# ═══════════════════════════════════════════════

@bp.route('/group/<int:group_id>', methods=['GET'])
@jwt_required()
def group_badges(group_id):
    """Obtener insignias emitidas para miembros de un grupo"""
    user = _require_roles('admin', 'editor', 'coordinator')

    # Multi-tenant: verificar acceso del coordinador al grupo
    access_error = _verify_badge_group_access(group_id, user)
    if access_error:
        return access_error

    from app.models.partner import GroupMember
    member_ids = [m.user_id for m in GroupMember.query.filter_by(group_id=group_id).all()]
    if not member_ids:
        return jsonify({'badges': [], 'total': 0})

    badges = IssuedBadge.query.filter(
        IssuedBadge.user_id.in_(member_ids)
    ).order_by(IssuedBadge.issued_at.desc()).all()

    result_list = []
    for b in badges:
        d = b.to_dict(include_credential=True)
        u = User.query.get(str(b.user_id))
        if u:
            d['candidate_name'] = f"{u.name or ''} {u.first_surname or ''} {u.second_surname or ''}".strip()
            d['candidate_email'] = u.email
        # Include template image (the one uploaded in badge module)
        d['template_image_url'] = b.template.badge_image_url if b.template else None
        result_list.append(d)

    return jsonify({'badges': result_list, 'total': len(result_list)})


@bp.route('/group/<int:group_id>/export-excel', methods=['GET'])
@jwt_required()
def export_group_badges_excel(group_id):
    """Exportar insignias del grupo a Excel — una fila por insignia"""
    user = _require_roles('admin', 'editor', 'coordinator')

    # Multi-tenant: verificar acceso del coordinador al grupo
    access_error = _verify_badge_group_access(group_id, user)
    if access_error:
        return access_error

    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from flask import send_file
    from app.models.partner import GroupMember, CandidateGroup

    group = CandidateGroup.query.get_or_404(group_id)
    member_ids = [m.user_id for m in GroupMember.query.filter_by(group_id=group_id).all()]
    if not member_ids:
        return jsonify({'error': 'El grupo no tiene miembros'}), 404

    badges = IssuedBadge.query.filter(
        IssuedBadge.user_id.in_(member_ids)
    ).order_by(IssuedBadge.issued_at.desc()).all()

    if not badges:
        return jsonify({'error': 'No hay insignias emitidas para este grupo'}), 404

    # Build user cache
    user_map = {}
    for uid in set(b.user_id for b in badges):
        u = User.query.get(str(uid))
        if u:
            user_map[uid] = u

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = 'Insignias'

    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='D97706', end_color='D97706', fill_type='solid')  # amber-600
    header_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    link_font = Font(color='0563C1', underline='single', size=10)

    headers = [
        'Nombre Completo', 'Email', 'CURP',
        'Insignia', 'Código', 'Estado',
        'Fecha Emisión', 'Fecha Expiración',
        'URL Verificación'
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    row_num = 2
    for b in badges:
        u = user_map.get(b.user_id)
        full_name = f"{u.name or ''} {u.first_surname or ''} {u.second_surname or ''}".strip() if u else ''
        email = u.email if u else ''
        curp = (u.curp or '') if u else ''
        tpl_name = b.template.name if b.template else ''
        status_text = 'Activa' if b.status == 'active' else ('Expirada' if b.status == 'expired' else 'Revocada')
        issued = b.issued_at.strftime('%Y-%m-%d') if b.issued_at else ''
        expires = b.expires_at.strftime('%Y-%m-%d') if b.expires_at else 'Sin expiración'
        verify = b.verify_url

        ws.cell(row=row_num, column=1, value=full_name).border = thin_border
        ws.cell(row=row_num, column=2, value=email).border = thin_border
        ws.cell(row=row_num, column=3, value=curp).border = thin_border
        ws.cell(row=row_num, column=4, value=tpl_name).border = thin_border
        ws.cell(row=row_num, column=5, value=b.badge_code).border = thin_border
        c_status = ws.cell(row=row_num, column=6, value=status_text)
        c_status.border = thin_border
        if b.status == 'active':
            c_status.font = Font(color='15803D')  # green-700
        elif b.status == 'revoked':
            c_status.font = Font(color='DC2626')  # red-600
        ws.cell(row=row_num, column=7, value=issued).border = thin_border
        ws.cell(row=row_num, column=8, value=expires).border = thin_border
        c_url = ws.cell(row=row_num, column=9, value=verify)
        c_url.border = thin_border
        c_url.font = link_font
        c_url.hyperlink = verify
        row_num += 1

    # Column widths
    widths = [35, 30, 22, 30, 16, 14, 16, 16, 55]
    for i, w in enumerate(widths):
        ws.column_dimensions[chr(65 + i)].width = w

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    safe_name = group.name.replace(' ', '_').replace('/', '-')[:30]
    from datetime import datetime
    filename = f'Insignias_{safe_name}_{datetime.now().strftime("%Y%m%d")}.xlsx'

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


@bp.route('/group/<int:group_id>/issue-pending', methods=['POST'])
@jwt_required()
def issue_pending_group_badges(group_id):
    """
    Emite retroactivamente insignias para resultados aprobados de un grupo
    que aún no tienen insignia emitida.
    """
    user = _require_roles('admin', 'editor', 'coordinator')

    # Multi-tenant: verificar acceso del coordinador al grupo
    access_error = _verify_badge_group_access(group_id, user)
    if access_error:
        return access_error

    from app.models.partner import GroupMember, CandidateGroup
    from app.models.result import Result
    from app.models.exam import Exam
    from app.services.badge_service import issue_badge_for_result

    group = CandidateGroup.query.get_or_404(group_id)
    members = GroupMember.query.filter_by(group_id=group_id).all()
    member_ids = [m.user_id for m in members]

    if not member_ids:
        return jsonify({'message': 'No hay miembros en el grupo', 'issued': 0}), 200

    # Buscar resultados aprobados de miembros de este grupo
    approved_results = Result.query.filter(
        Result.user_id.in_(member_ids),
        Result.result == 1,
        Result.status == 1
    ).all()

    if not approved_results:
        return jsonify({'message': 'No hay resultados aprobados pendientes', 'issued': 0}), 200

    # Pre-cargar datos para evitar N+1 queries
    result_ids = [r.id for r in approved_results]
    user_ids = list(set(str(r.user_id) for r in approved_results))
    exam_ids = list(set(r.exam_id for r in approved_results if r.exam_id))

    # Cargar todas las insignias existentes de una vez
    existing_badges_by_result = set(
        row[0] for row in db.session.query(IssuedBadge.result_id)
        .filter(IssuedBadge.result_id.in_(result_ids)).all()
    )

    # Cargar todos los usuarios de una vez
    users_map = {u.id: u for u in User.query.filter(User.id.in_(user_ids)).all()}

    # Cargar todos los exámenes de una vez
    exams_map = {e.id: e for e in Exam.query.filter(Exam.id.in_(exam_ids)).all()} if exam_ids else {}

    # Cargar todas las plantillas activas de una vez
    active_templates = BadgeTemplate.query.filter_by(is_active=True).filter(
        db.or_(
            BadgeTemplate.exam_id.in_(exam_ids),
            BadgeTemplate.competency_standard_id.isnot(None)
        )
    ).all() if exam_ids else []
    templates_by_exam = {}
    templates_by_standard = {}
    for t in active_templates:
        if t.exam_id:
            templates_by_exam[t.exam_id] = t
        if t.competency_standard_id:
            templates_by_standard[t.competency_standard_id] = t

    # Cargar insignias activas existentes por (user_id, template_id)
    template_ids = [t.id for t in active_templates]
    existing_user_badges = set()
    if template_ids:
        for row in db.session.query(IssuedBadge.user_id, IssuedBadge.badge_template_id).filter(
            IssuedBadge.user_id.in_(user_ids),
            IssuedBadge.badge_template_id.in_(template_ids),
            IssuedBadge.status == 'active'
        ).all():
            existing_user_badges.add((row[0], row[1]))

    issued_count = 0
    errors_list = []

    for result in approved_results:
        # Verificar si ya tiene insignia para este resultado
        if result.id in existing_badges_by_result:
            continue

        user = users_map.get(str(result.user_id))
        if not user:
            continue

        exam = exams_map.get(result.exam_id) if result.exam_id else None
        if not exam:
            continue

        # Buscar plantilla activa (desde cache precargado)
        template = templates_by_exam.get(exam.id)
        if not template and hasattr(exam, 'competency_standard_id') and exam.competency_standard_id:
            template = templates_by_standard.get(exam.competency_standard_id)
        if not template:
            continue

        # Verificar que no haya duplicado user+template
        if (user.id, template.id) in existing_user_badges:
            continue

        try:
            badge = issue_badge_for_result(result, user, exam, force=True)
            if badge:
                issued_count += 1
        except HTTPException:
            raise
        except Exception as e:
            errors_list.append(f"{getattr(user, 'email', None) or user.username}: {str(e)}")

    return jsonify({
        'message': f'Se emitieron {issued_count} insignia(s)',
        'issued': issued_count,
        'errors': errors_list[:20] if errors_list else [],
    }), 200


# ═══════════════════════════════════════════════
# LinkedIn API Integration (OAuth2 + Share Posts)
# ═══════════════════════════════════════════════

@bp.route('/linkedin/status', methods=['GET'])
@jwt_required()
def linkedin_api_status():
    """Check if LinkedIn API is configured and if user has a stored token."""
    from app.services.linkedin_service import is_configured
    user_id = get_jwt_identity()
    user = User.query.get(str(user_id))
    has_token = bool(getattr(user, 'linkedin_token', None)) if user else False
    return jsonify({
        'configured': is_configured(),
        'connected': has_token,
    })


@bp.route('/linkedin/authorize', methods=['GET'])
@jwt_required()
def linkedin_authorize():
    """
    Start LinkedIn OAuth2 flow.
    Returns the authorization URL. The frontend opens it in a popup/new tab.
    Query param: badge_id — the badge to share after auth completes.
    """
    from app.services.linkedin_service import is_configured, build_authorize_url
    import json, base64

    if not is_configured():
        return jsonify({'error': 'LinkedIn API no está configurada. Contacte al administrador.'}), 501

    badge_id = request.args.get('badge_id', '')
    user_id = get_jwt_identity()

    # state = base64(JSON{user_id, badge_id}) — used in callback to resume flow
    state_data = json.dumps({'user_id': user_id, 'badge_id': badge_id})
    state = base64.urlsafe_b64encode(state_data.encode()).decode()

    auth_url = build_authorize_url(state)
    return jsonify({'authorize_url': auth_url})


@bp.route('/linkedin/callback', methods=['GET'])
def linkedin_callback():
    """
    LinkedIn OAuth2 callback. Exchanges code for token, stores it,
    then redirects to frontend with success/error.
    """
    import json, base64, os
    from app.services.linkedin_service import exchange_code_for_token

    code = request.args.get('code')
    state = request.args.get('state', '')
    error = request.args.get('error')

    swa_base = os.environ.get('SWA_BASE_URL', 'https://app.evaluaasi.com')

    if error:
        return _redirect_html(f"{swa_base}/certificates?linkedin_error={error}")

    if not code:
        return _redirect_html(f"{swa_base}/certificates?linkedin_error=no_code")

    try:
        state_data = json.loads(base64.urlsafe_b64decode(state))
        user_id = state_data.get('user_id')
        badge_id = state_data.get('badge_id', '')
    except Exception:
        return _redirect_html(f"{swa_base}/certificates?linkedin_error=invalid_state")

    try:
        token_data = exchange_code_for_token(code)
        access_token = token_data.get('access_token')
        if not access_token:
            return _redirect_html(f"{swa_base}/certificates?linkedin_error=no_token")

        # Store token on user record
        user = User.query.get(str(user_id))
        if user:
            user.linkedin_token = access_token
            db.session.commit()

        redirect_url = f"{swa_base}/certificates?linkedin_connected=1"
        if badge_id:
            redirect_url += f"&share_badge={badge_id}"
        return _redirect_html(redirect_url)

    except HTTPException:

        raise

    except Exception as e:
        print(f"[LINKEDIN] Token exchange error: {e}")
        return _redirect_html(f"{swa_base}/certificates?linkedin_error=token_exchange_failed")


def _redirect_html(url):
    """Return a simple HTML redirect page."""
    from markupsafe import escape
    from flask import make_response
    html = f'''<!DOCTYPE html><html><head>
<meta http-equiv="refresh" content="0;url={escape(url)}" />
</head><body><p>Redirigiendo...</p></body></html>'''
    resp = make_response(html, 200)
    resp.headers['Content-Type'] = 'text/html; charset=utf-8'
    return resp


@bp.route('/linkedin/share/<int:badge_id>', methods=['POST'])
@jwt_required()
def linkedin_share_badge(badge_id):
    """
    Share a badge on LinkedIn using the API (with image upload).
    Requires the user to have completed OAuth2 flow (linkedin_token stored).
    Falls back to URL-based sharing if no token.
    """
    import os
    from io import BytesIO
    from app.services.linkedin_service import is_configured, share_badge_with_image

    user_id = get_jwt_identity()
    user = User.query.get(str(user_id))
    badge = IssuedBadge.query.get_or_404(badge_id)
    template = BadgeTemplate.query.get(badge.badge_template_id)

    access_token = getattr(user, 'linkedin_token', None) if user else None

    if not is_configured() or not access_token:
        # Fallback: return URL-based sharing
        return _build_url_share_response(badge, template)

    # Get the PNG badge image
    try:
        api_base = os.environ.get('API_BASE_URL', request.host_url.rstrip('/'))
        if api_base.startswith('http://'):
            api_base = 'https://' + api_base[7:]
        png_url = f"{api_base}/api/badges/share-image/{badge.badge_code}.png"

        import requests as http_req
        img_resp = http_req.get(png_url, timeout=15)
        img_resp.raise_for_status()
        image_bytes = img_resp.content

        result = share_badge_with_image(access_token, badge, template, image_bytes)

        # Track the share
        badge.share_count = (badge.share_count or 0) + 1
        db.session.commit()

        return jsonify({
            'success': True,
            'method': 'linkedin_api',
            'message': 'Insignia publicada en LinkedIn exitosamente',
            'post': result,
        })
    except HTTPException:
        raise
    except Exception as e:
        print(f"[LINKEDIN] Share error: {e}")
        # If API share fails (e.g., expired token), clear token and fallback
        if user and hasattr(user, 'linkedin_token'):
            user.linkedin_token = None
            db.session.commit()
        return _build_url_share_response(badge, template, error=str(e))


def _build_url_share_response(badge, template, error=None):
    """Build fallback URL-based sharing response."""
    import os
    from urllib.parse import quote

    api_base = os.environ.get('API_BASE_URL', request.host_url.rstrip('/'))
    if api_base.startswith('http://'):
        api_base = 'https://' + api_base[7:]

    share_preview_url = f"{api_base}/api/badges/share-preview/{badge.badge_code}"
    image_png_url = f"{api_base}/api/badges/share-image/{badge.badge_code}.png"
    share_post_url = f"https://www.linkedin.com/sharing/share-offsite/?url={quote(share_preview_url)}"

    resp_data = {
        'success': False,
        'method': 'url_fallback',
        'share_post_url': share_post_url,
        'image_url': image_png_url,
    }
    if error:
        resp_data['error'] = f'LinkedIn API no disponible: {error}. Use el enlace alternativo.'
    return jsonify(resp_data)
