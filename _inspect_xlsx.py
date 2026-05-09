import openpyxl, sys
path = r'C:\Users\Diego\Downloads\ActiveDirectoryUsers (1)\plantilla_asignacion_examenes_nacional 1.xlsx'
wb = openpyxl.load_workbook(path, data_only=True)
for s in wb.sheetnames:
    ws = wb[s]
    print('---', s, 'DIMS:', ws.dimensions)
    for i, row in enumerate(ws.iter_rows(max_row=12)):
        print(i, [c.value for c in row])
